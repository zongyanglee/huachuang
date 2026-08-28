# -*- coding: utf-8 -*-

from __future__ import annotations

from configparser import ConfigParser
from datetime import datetime
from pathlib import Path
import re
import time as _time

from iFinDPy import *
import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont
from tqdm import tqdm

from copy import copy
import ctypes
from ctypes import wintypes
from email.message import EmailMessage
from email.utils import formatdate, make_msgid
from pathlib import Path
import mimetypes
import os
import queue
import re
import smtplib
import ssl
import tempfile
import threading
from typing import Callable, Iterable, Sequence

from PIL import Image, ImageTk
from openpyxl import load_workbook


SMTP_HOST = "smtp.126.com"
SMTP_PORT = 465
BCC_BATCH_SIZE = 40
SMTP_BATCH_INTERVAL_SECONDS = 5.0
AUTH_CODE_ENV_VAR = "HC_126_SMTP_AUTH_CODE"
_CREDENTIAL_FILE = (
    Path(os.environ.get("APPDATA", Path.home() / "AppData" / "Roaming"))
    / "HuachuangBondTools"
    / "126_smtp_auth.dpapi"
)
_EMAIL_ADDRESS_PATTERN = re.compile(
    r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}",
    re.IGNORECASE,
)
_EMAIL_COLUMN_HEADERS = {"邮箱地址", "邮箱", "邮件地址", "email", "e-mail"}
_DEFAULT_RECIPIENT_SHEET_NAMES = ("赎回", "强赎")


def _sort_recipient_addresses(addresses: Iterable[str]) -> list[str]:
    """Return non-empty email addresses in case-insensitive alphabetical order."""
    cleaned = [address.strip() for address in addresses if address.strip()]
    return sorted(cleaned, key=lambda address: (address.casefold(), address))


def _filter_recipient_addresses(
    addresses: Sequence[str],
    query: str,
) -> list[str]:
    """Filter addresses by a case-insensitive substring without changing order."""
    normalized_query = query.strip().casefold()
    if not normalized_query:
        return list(addresses)
    return [
        address
        for address in addresses
        if normalized_query in address.casefold()
    ]


class _DataBlob(ctypes.Structure):
    _fields_ = [
        ("cbData", wintypes.DWORD),
        ("pbData", ctypes.POINTER(ctypes.c_byte)),
    ]


def _blob_from_bytes(data: bytes) -> tuple[_DataBlob, ctypes.Array]:
    buffer = ctypes.create_string_buffer(data)
    blob = _DataBlob(
        len(data),
        ctypes.cast(buffer, ctypes.POINTER(ctypes.c_byte)),
    )
    return blob, buffer


def _protect_for_current_windows_user(value: str) -> bytes:
    if os.name != "nt":
        raise RuntimeError("DPAPI 凭据保存仅支持 Windows。")

    raw = value.encode("utf-8")
    input_blob, input_buffer = _blob_from_bytes(raw)
    output_blob = _DataBlob()
    crypt_protect_data = ctypes.windll.crypt32.CryptProtectData
    crypt_protect_data.argtypes = [
        ctypes.POINTER(_DataBlob),
        wintypes.LPCWSTR,
        ctypes.POINTER(_DataBlob),
        wintypes.LPVOID,
        wintypes.LPVOID,
        wintypes.DWORD,
        ctypes.POINTER(_DataBlob),
    ]
    crypt_protect_data.restype = wintypes.BOOL

    if not crypt_protect_data(
        ctypes.byref(input_blob),
        "126 SMTP authorization code",
        None,
        None,
        None,
        0,
        ctypes.byref(output_blob),
    ):
        raise ctypes.WinError()

    try:
        return ctypes.string_at(output_blob.pbData, output_blob.cbData)
    finally:
        ctypes.windll.kernel32.LocalFree(output_blob.pbData)
        del input_buffer


def _unprotect_for_current_windows_user(encrypted: bytes) -> str:
    if os.name != "nt":
        raise RuntimeError("DPAPI 凭据读取仅支持 Windows。")

    input_blob, input_buffer = _blob_from_bytes(encrypted)
    output_blob = _DataBlob()
    crypt_unprotect_data = ctypes.windll.crypt32.CryptUnprotectData
    crypt_unprotect_data.argtypes = [
        ctypes.POINTER(_DataBlob),
        ctypes.POINTER(wintypes.LPWSTR),
        ctypes.POINTER(_DataBlob),
        wintypes.LPVOID,
        wintypes.LPVOID,
        wintypes.DWORD,
        ctypes.POINTER(_DataBlob),
    ]
    crypt_unprotect_data.restype = wintypes.BOOL

    if not crypt_unprotect_data(
        ctypes.byref(input_blob),
        None,
        None,
        None,
        None,
        0,
        ctypes.byref(output_blob),
    ):
        raise ctypes.WinError()

    try:
        return ctypes.string_at(output_blob.pbData, output_blob.cbData).decode("utf-8")
    finally:
        ctypes.windll.kernel32.LocalFree(output_blob.pbData)
        del input_buffer


def credential_file_path() -> Path:
    return _CREDENTIAL_FILE


def save_authorization_code(authorization_code: str) -> Path:
    authorization_code = authorization_code.strip()
    if not authorization_code:
        raise ValueError("授权码不能为空。")

    encrypted = _protect_for_current_windows_user(authorization_code)
    _CREDENTIAL_FILE.parent.mkdir(parents=True, exist_ok=True)
    _CREDENTIAL_FILE.write_bytes(encrypted)
    return _CREDENTIAL_FILE


def load_authorization_code() -> str:
    env_value = os.environ.get(AUTH_CODE_ENV_VAR, "").strip()
    if env_value:
        return env_value

    if not _CREDENTIAL_FILE.exists():
        raise RuntimeError(
            "尚未设置 126 邮箱客户端授权码。请先运行：py 设置126邮箱授权码.py"
        )
    return _unprotect_for_current_windows_user(_CREDENTIAL_FILE.read_bytes()).strip()


def test_smtp_login(sender: str, authorization_code: str | None = None) -> None:
    password = authorization_code or load_authorization_code()
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(
        SMTP_HOST,
        SMTP_PORT,
        context=context,
        timeout=30,
    ) as smtp:
        smtp.login(sender, password)


def _validate_files(paths: Iterable[Path]) -> list[Path]:
    checked = [Path(path) for path in paths]
    missing = [str(path) for path in checked if not path.is_file()]
    empty = [str(path) for path in checked if path.is_file() and path.stat().st_size == 0]
    if missing or empty:
        details = []
        if missing:
            details.append(f"缺失文件：{missing}")
        if empty:
            details.append(f"空文件：{empty}")
        raise RuntimeError("邮件发送前附件校验失败；" + "；".join(details))
    return checked


def _is_valid_email_address(address: str) -> bool:
    if not address or address.count("@") != 1 or any(char.isspace() for char in address):
        return False
    local_part, domain = address.rsplit("@", 1)
    return bool(local_part and "." in domain and not domain.startswith(".") and not domain.endswith("."))


def merge_recipient_addresses(*recipient_groups: Sequence[str]) -> list[str]:
    merged = []
    seen = set()
    for group in recipient_groups:
        for raw_address in group:
            address = str(raw_address).strip()
            normalized = address.lower()
            if not address or normalized in seen:
                continue
            if not _is_valid_email_address(address):
                raise ValueError(f"无效的收件人邮箱地址：{address}")
            seen.add(normalized)
            merged.append(address)
    return merged


def _locate_recipient_sheet_and_column(
    workbook,
    workbook_path: Path,
    preferred_sheet_names: Sequence[str],
):
    sheet_name = next(
        (name for name in preferred_sheet_names if name in workbook.sheetnames),
        None,
    )
    if sheet_name is None:
        raise RuntimeError(
            f"{workbook_path.name} 中未找到工作表"
            f"{list(preferred_sheet_names)}；现有工作表：{workbook.sheetnames}"
        )
    worksheet = workbook[sheet_name]
    if worksheet.max_row is None:
        worksheet.calculate_dimension(force=True)
    worksheet_max_row = worksheet.max_row or 1

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(max(worksheet_max_row, 1), 20),
            values_only=True,
        ),
        start=1,
    ):
        for column_number, value in enumerate(row, start=1):
            header = str(value).strip().lower() if value is not None else ""
            if header in _EMAIL_COLUMN_HEADERS:
                return worksheet, sheet_name, row_number, column_number

    raise RuntimeError(
        f"{workbook_path.name} 的“{sheet_name}”工作表中未找到“邮箱地址”列。"
    )


def _recipient_rows(worksheet, header_row: int, email_column: int):
    rows = []
    invalid_rows = []
    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=header_row + 1,
            min_col=email_column,
            max_col=email_column,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        value = row[0]
        if value is None or str(value).strip() == "":
            continue
        text = str(value).strip()
        matches = _EMAIL_ADDRESS_PATTERN.findall(text)
        if len(matches) != 1 or not _is_valid_email_address(matches[0]):
            invalid_rows.append((row_number, text))
            continue
        rows.append((row_number, matches[0]))
    return rows, invalid_rows


def _save_workbook_atomically(workbook, workbook_path: Path) -> None:
    workbook_path = Path(workbook_path).resolve()
    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{workbook_path.stem}_",
        suffix=workbook_path.suffix,
        dir=workbook_path.parent,
    )
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, workbook_path)
    finally:
        temporary_path.unlink(missing_ok=True)


def add_recipient_to_workbook(
    workbook_path: Path,
    address: str,
    *,
    preferred_sheet_names: Sequence[str] = _DEFAULT_RECIPIENT_SHEET_NAMES,
) -> None:
    address = address.strip()
    if not _is_valid_email_address(address):
        raise ValueError(f"无效的收件人邮箱地址：{address}")

    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        worksheet, _, header_row, email_column = _locate_recipient_sheet_and_column(
            workbook,
            workbook_path,
            preferred_sheet_names,
        )
        rows, invalid_rows = _recipient_rows(worksheet, header_row, email_column)
        if invalid_rows:
            raise RuntimeError("邮箱表存在无效地址，请先修复后再新增。")
        if address.lower() in {existing.lower() for _, existing in rows}:
            raise ValueError(f"邮箱地址已存在：{address}")

        new_row = max([row_number for row_number, _ in rows], default=header_row) + 1
        template_row = new_row - 1
        if template_row > header_row:
            for column_number in range(1, worksheet.max_column + 1):
                source_cell = worksheet.cell(template_row, column_number)
                target_cell = worksheet.cell(new_row, column_number)
                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.protection = copy(source_cell.protection)
            worksheet.row_dimensions[new_row].height = worksheet.row_dimensions[
                template_row
            ].height
        worksheet.cell(new_row, email_column).value = address
        _save_workbook_atomically(workbook, workbook_path)
    finally:
        workbook.close()


def update_recipient_in_workbook(
    workbook_path: Path,
    old_address: str,
    new_address: str,
    *,
    preferred_sheet_names: Sequence[str] = _DEFAULT_RECIPIENT_SHEET_NAMES,
) -> None:
    old_address = old_address.strip()
    new_address = new_address.strip()
    if not _is_valid_email_address(new_address):
        raise ValueError(f"无效的收件人邮箱地址：{new_address}")

    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        worksheet, _, header_row, email_column = _locate_recipient_sheet_and_column(
            workbook,
            workbook_path,
            preferred_sheet_names,
        )
        rows, invalid_rows = _recipient_rows(worksheet, header_row, email_column)
        if invalid_rows:
            raise RuntimeError("邮箱表存在无效地址，请先修复后再修改。")
        old_matches = [
            row_number
            for row_number, address in rows
            if address.lower() == old_address.lower()
        ]
        if not old_matches:
            raise RuntimeError(f"邮箱表中未找到待修改地址：{old_address}")
        if any(
            address.lower() == new_address.lower()
            and address.lower() != old_address.lower()
            for _, address in rows
        ):
            raise ValueError(f"新邮箱地址已存在：{new_address}")

        for row_number in old_matches:
            worksheet.cell(row_number, email_column).value = new_address
        _save_workbook_atomically(workbook, workbook_path)
    finally:
        workbook.close()


def delete_recipients_from_workbook(
    workbook_path: Path,
    addresses: Sequence[str],
    *,
    preferred_sheet_names: Sequence[str] = _DEFAULT_RECIPIENT_SHEET_NAMES,
) -> None:
    normalized_targets = {address.strip().lower() for address in addresses if address.strip()}
    if not normalized_targets:
        return

    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    try:
        worksheet, _, header_row, email_column = _locate_recipient_sheet_and_column(
            workbook,
            workbook_path,
            preferred_sheet_names,
        )
        rows, invalid_rows = _recipient_rows(worksheet, header_row, email_column)
        if invalid_rows:
            raise RuntimeError("邮箱表存在无效地址，请先修复后再删除。")
        rows_to_delete = [
            row_number
            for row_number, address in rows
            if address.lower() in normalized_targets
        ]
        found = {
            address.lower()
            for _, address in rows
            if address.lower() in normalized_targets
        }
        missing = normalized_targets.difference(found)
        if missing:
            raise RuntimeError(f"邮箱表中未找到待删除地址：{sorted(missing)}")

        for row_number in sorted(rows_to_delete, reverse=True):
            worksheet.delete_rows(row_number, 1)
        _save_workbook_atomically(workbook, workbook_path)
    finally:
        workbook.close()


def load_recipients_from_workbook(
    workbook_path: Path,
    *,
    preferred_sheet_names: Sequence[str] = _DEFAULT_RECIPIENT_SHEET_NAMES,
) -> list[str]:
    workbook_path = Path(workbook_path)
    if not workbook_path.is_file():
        raise RuntimeError(f"未找到邮件地址表：{workbook_path}")

    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )
    try:
        worksheet, sheet_name, header_row, email_column = (
            _locate_recipient_sheet_and_column(
                workbook,
                workbook_path,
                preferred_sheet_names,
            )
        )
        rows, invalid_rows = _recipient_rows(worksheet, header_row, email_column)
        addresses = [address for _, address in rows]

        if invalid_rows:
            examples = "；".join(
                f"第{row_number}行：{value}"
                for row_number, value in invalid_rows[:10]
            )
            raise RuntimeError(
                f"{workbook_path.name} 的“{sheet_name}”工作表存在"
                f"{len(invalid_rows)} 条无效邮箱地址：{examples}"
            )

        unique_addresses = merge_recipient_addresses(addresses)
        if not unique_addresses:
            raise RuntimeError(
                f"{workbook_path.name} 的“{sheet_name}”工作表没有有效邮箱地址。"
            )
        print(
            f"[邮件收件人] 从 {workbook_path.name} 的“{sheet_name}”工作表"
            f"读取 {len(addresses)} 条，去重后 {len(unique_addresses)} 条。"
        )
        return unique_addresses
    finally:
        workbook.close()


def confirm_recipients_with_preview(
    *,
    sender: str,
    recipients: Sequence[str],
    recipient_workbook_path: Path | None,
    report_date: str,
    excel_paths: Sequence[Path],
    permille_image_path: Path,
    percent_image_path: Path,
    send_action: Callable[
        [Sequence[str], Callable[[int, str, str], None]],
        None,
    ],
) -> bool | None:
    try:
        import tkinter as tk
        from tkinter import messagebox, ttk
    except ImportError as exc:
        raise RuntimeError("当前 Python 环境缺少 tkinter，无法显示邮件发送确认窗口。") from exc

    recipients = _sort_recipient_addresses(recipients)
    recipient_addresses = list(recipients)
    result: dict[str, bool | None] = {"all_succeeded": None}
    sending_state = {"active": False, "completed": False}
    progress_events: queue.Queue = queue.Queue()

    try:
        root = tk.Tk()
    except tk.TclError as exc:
        raise RuntimeError(f"无法显示邮件发送确认窗口：{exc}") from exc

    root.title("可转债赎回进度邮件发送确认")
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    window_width = min(1380, max(1000, screen_width - 120))
    window_height = min(900, max(700, screen_height - 120))
    x = max((screen_width - window_width) // 2, 0)
    y = max((screen_height - window_height) // 2, 0)
    root.geometry(f"{window_width}x{window_height}+{x}+{y}")
    root.minsize(950, 650)
    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)
    def request_close() -> None:
        if sending_state["active"]:
            messagebox.showinfo(
                "邮件正在发送",
                "请等待本轮邮件发送完成后再关闭窗口。",
                parent=root,
            )
            return
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", request_close)

    title = ttk.Label(
        root,
        text="发送前确认：左侧为完整邮件内容，右侧为本轮收件人",
        font=("Microsoft YaHei UI", 12, "bold"),
    )
    title.grid(row=0, column=0, sticky="w", padx=16, pady=(14, 10))

    content = ttk.Panedwindow(root, orient=tk.HORIZONTAL)
    content.grid(row=1, column=0, sticky="nsew", padx=16)

    preview_frame = ttk.Frame(content)
    preview_frame.grid_rowconfigure(1, weight=1)
    preview_frame.grid_columnconfigure(0, weight=1)
    ttk.Label(
        preview_frame,
        text="邮件内容预览（两张长图内嵌正文，两份 Excel 作为附件）",
        font=("Microsoft YaHei UI", 10, "bold"),
    ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

    canvas = tk.Canvas(preview_frame, background="#FFFFFF", highlightthickness=1)
    preview_scrollbar = ttk.Scrollbar(
        preview_frame,
        orient=tk.VERTICAL,
        command=canvas.yview,
    )
    canvas.configure(yscrollcommand=preview_scrollbar.set)
    canvas.grid(row=1, column=0, sticky="nsew")
    preview_scrollbar.grid(row=1, column=1, sticky="ns")

    recipient_frame = ttk.Frame(content, padding=(14, 0, 0, 0))
    recipient_frame.grid_rowconfigure(3, weight=1)
    recipient_frame.grid_columnconfigure(0, weight=1)
    ttk.Label(
        recipient_frame,
        text="邮件地址列表（增删改同步写入 Excel）",
        font=("Microsoft YaHei UI", 10, "bold"),
    ).grid(row=0, column=0, columnspan=2, sticky="w")

    search_frame = ttk.Frame(recipient_frame)
    search_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(8, 4))
    search_frame.grid_columnconfigure(1, weight=1)
    ttk.Label(search_frame, text="搜索：").grid(row=0, column=0, padx=(0, 4))
    search_var = tk.StringVar()
    search_entry = ttk.Entry(search_frame, textvariable=search_var)
    search_entry.grid(row=0, column=1, sticky="ew")

    count_label = ttk.Label(recipient_frame, text="")
    count_label.grid(row=2, column=0, columnspan=2, sticky="w", pady=(0, 8))

    recipient_list = tk.Listbox(
        recipient_frame,
        selectmode=tk.EXTENDED,
        exportselection=False,
        font=("Segoe UI", 10),
    )
    recipient_scrollbar = ttk.Scrollbar(
        recipient_frame,
        orient=tk.VERTICAL,
        command=recipient_list.yview,
    )
    recipient_list.configure(yscrollcommand=recipient_scrollbar.set)
    recipient_list.grid(row=3, column=0, sticky="nsew")
    recipient_scrollbar.grid(row=3, column=1, sticky="ns")

    ttk.Label(recipient_frame, text="输入邮件地址（用于新增或修改）").grid(
        row=4,
        column=0,
        columnspan=2,
        sticky="w",
        pady=(12, 4),
    )
    address_entry = ttk.Entry(recipient_frame)
    address_entry.grid(row=5, column=0, columnspan=2, sticky="ew")

    edit_buttons = ttk.Frame(recipient_frame)
    edit_buttons.grid(row=6, column=0, columnspan=2, sticky="ew", pady=(8, 0))
    edit_buttons.grid_columnconfigure(0, weight=1)
    edit_buttons.grid_columnconfigure(1, weight=1)
    edit_buttons.grid_columnconfigure(2, weight=1)

    def update_count() -> None:
        total = len(recipient_addresses)
        visible = recipient_list.size()
        if search_var.get().strip():
            summary = f"显示 {visible} / 共 {total} 个地址"
        else:
            summary = f"共 {total} 个地址"
        count_label.configure(text=f"{summary}（单选修改，可多选删除）")

    def refresh_recipient_list(
        *,
        select_addresses: Sequence[str] = (),
    ) -> None:
        selected = {address.casefold() for address in select_addresses}
        recipient_list.delete(0, tk.END)
        first_selected_index = None
        for address in _filter_recipient_addresses(
            recipient_addresses,
            search_var.get(),
        ):
            index = recipient_list.size()
            recipient_list.insert(tk.END, address)
            if address.casefold() in selected:
                recipient_list.selection_set(index)
                if first_selected_index is None:
                    first_selected_index = index
        if first_selected_index is not None:
            recipient_list.see(first_selected_index)
        update_count()

    def clear_search() -> None:
        search_var.set("")
        search_entry.focus_set()

    clear_search_button = ttk.Button(
        search_frame,
        text="清空",
        command=clear_search,
    )
    clear_search_button.grid(row=0, column=2, padx=(4, 0))
    search_var.trace_add("write", lambda *_: refresh_recipient_list())
    refresh_recipient_list()

    def add_recipient(event=None) -> None:
        address = address_entry.get().strip()
        if not _is_valid_email_address(address):
            messagebox.showerror("地址格式错误", "请输入有效的邮件地址。", parent=root)
            return
        existing = {item.casefold() for item in recipient_addresses}
        if address.lower() in existing:
            messagebox.showinfo("地址已存在", "该邮件地址已在列表中。", parent=root)
            return
        if recipient_workbook_path is not None:
            try:
                add_recipient_to_workbook(recipient_workbook_path, address)
            except Exception as exc:
                messagebox.showerror(
                    "Excel 写入失败",
                    f"未新增地址：{exc}",
                    parent=root,
                )
                return
        recipient_addresses.append(address)
        recipient_addresses.sort(key=lambda item: (item.casefold(), item))
        address_entry.delete(0, tk.END)
        search_var.set("")
        refresh_recipient_list(select_addresses=(address,))

    def edit_selected() -> None:
        selected = recipient_list.curselection()
        if len(selected) != 1:
            messagebox.showinfo(
                "请选择一个地址",
                "请只选择一个待修改地址，并在输入框填写新地址。",
                parent=root,
            )
            return
        index = selected[0]
        old_address = recipient_list.get(index).strip()
        new_address = address_entry.get().strip()
        if not _is_valid_email_address(new_address):
            messagebox.showerror("地址格式错误", "请输入有效的新邮件地址。", parent=root)
            return
        existing = {
            item.casefold()
            for item in recipient_addresses
            if item.casefold() != old_address.casefold()
        }
        if new_address.lower() in existing:
            messagebox.showinfo("地址已存在", "新邮件地址已在列表中。", parent=root)
            return
        if recipient_workbook_path is not None:
            try:
                update_recipient_in_workbook(
                    recipient_workbook_path,
                    old_address,
                    new_address,
                )
            except Exception as exc:
                messagebox.showerror(
                    "Excel 写入失败",
                    f"未修改地址：{exc}",
                    parent=root,
                )
                return
        old_key = old_address.casefold()
        recipient_addresses[:] = [
            new_address if item.casefold() == old_key else item
            for item in recipient_addresses
        ]
        recipient_addresses.sort(key=lambda item: (item.casefold(), item))
        address_entry.delete(0, tk.END)
        search_var.set("")
        refresh_recipient_list(select_addresses=(new_address,))

    def delete_selected() -> None:
        selected = recipient_list.curselection()
        if not selected:
            messagebox.showinfo("未选择地址", "请先选择需要删除的邮件地址。", parent=root)
            return
        addresses_to_delete = [recipient_list.get(index).strip() for index in selected]
        preview = "\n".join(addresses_to_delete[:10])
        if len(addresses_to_delete) > 10:
            preview += f"\n……共 {len(addresses_to_delete)} 个地址"
        if not messagebox.askyesno(
            "确认永久删除",
            "以下地址将从 Excel 中永久删除：\n\n"
            f"{preview}\n\n是否继续？",
            parent=root,
        ):
            return
        if recipient_workbook_path is not None:
            try:
                delete_recipients_from_workbook(
                    recipient_workbook_path,
                    addresses_to_delete,
                )
            except Exception as exc:
                messagebox.showerror(
                    "Excel 写入失败",
                    f"未删除地址：{exc}",
                    parent=root,
                )
                return
        deleted_keys = {address.casefold() for address in addresses_to_delete}
        recipient_addresses[:] = [
            address
            for address in recipient_addresses
            if address.casefold() not in deleted_keys
        ]
        refresh_recipient_list()

    add_button = ttk.Button(edit_buttons, text="添加", command=add_recipient)
    add_button.grid(
        row=0,
        column=0,
        sticky="ew",
        padx=(0, 3),
    )
    edit_button = ttk.Button(edit_buttons, text="修改选中", command=edit_selected)
    edit_button.grid(
        row=0,
        column=1,
        sticky="ew",
        padx=3,
    )
    delete_button = ttk.Button(edit_buttons, text="删除选中", command=delete_selected)
    delete_button.grid(
        row=0,
        column=2,
        sticky="ew",
        padx=(3, 0),
    )
    address_entry.bind("<Return>", add_recipient)

    def load_selected_address(event=None) -> None:
        selected = recipient_list.curselection()
        if len(selected) != 1:
            return
        address_entry.delete(0, tk.END)
        address_entry.insert(0, recipient_list.get(selected[0]).strip())
        address_entry.focus_set()

    recipient_list.bind("<Double-Button-1>", load_selected_address)

    content.add(preview_frame, weight=4)
    content.add(recipient_frame, weight=1)

    footer = ttk.Frame(root)
    footer.grid(row=2, column=0, sticky="ew", padx=16, pady=14)
    footer.grid_columnconfigure(0, weight=1)
    footer_status = ttk.Label(
        footer,
        text="关闭窗口或选择“取消发送”均不会发送邮件。",
        wraplength=max(window_width - 360, 500),
    )
    footer_status.grid(row=0, column=0, sticky="w")

    def confirm_send() -> None:
        current_recipients = list(recipient_addresses)
        if not current_recipients:
            messagebox.showerror("无法发送", "本轮收件人列表不能为空。", parent=root)
            return
        sending_state["active"] = True
        result["all_succeeded"] = False

        title.configure(text="分批密送正在提交，请在右侧查看服务器接受状态")
        address_entry.configure(state=tk.DISABLED)
        search_entry.configure(state=tk.DISABLED)
        clear_search_button.configure(state=tk.DISABLED)
        add_button.configure(state=tk.DISABLED)
        edit_button.configure(state=tk.DISABLED)
        delete_button.configure(state=tk.DISABLED)
        confirm_button.configure(state=tk.DISABLED, text="发送中…")
        cancel_button.configure(state=tk.DISABLED)
        recipient_list.configure(selectmode=tk.SINGLE)
        footer_status.configure(text=f"正在发送：0/{len(current_recipients)}")

        search_var.set("")
        recipient_list.delete(0, tk.END)
        for index, address in enumerate(current_recipients):
            recipient_list.insert(tk.END, f"○  {address}")
            recipient_list.itemconfig(index, foreground="#555555")
        count_label.configure(text=f"本轮共 {len(current_recipients)} 个地址")

        def report_progress(index: int, status: str, detail: str = "") -> None:
            progress_events.put(("status", index, status, detail))

        def send_worker() -> None:
            try:
                send_action(current_recipients, report_progress)
            except Exception as exc:
                progress_events.put(("fatal", str(exc)))
            finally:
                progress_events.put(("done",))

        threading.Thread(target=send_worker, daemon=True).start()
        root.after(100, poll_progress)

    cancel_button = ttk.Button(footer, text="取消发送", command=request_close)
    cancel_button.grid(
        row=0,
        column=1,
        padx=(8, 8),
    )
    confirm_button = ttk.Button(footer, text="确认发送", command=confirm_send)
    confirm_button.grid(
        row=0,
        column=2,
    )

    status_by_index: dict[int, str] = {}
    detail_by_index: dict[int, str] = {}
    status_display = {
        "waiting": ("○", "#555555"),
        "sending": ("⏳", "#C77800"),
        "success": ("✅", "#188038"),
        "failed": ("❌", "#C5221F"),
    }

    def update_recipient_status(index: int, status: str, detail: str = "") -> None:
        if index < 0 or index >= recipient_list.size():
            return
        current_text = recipient_list.get(index)
        address = current_text.split("  ", 1)[-1]
        marker, color = status_display.get(status, status_display["waiting"])
        recipient_list.delete(index)
        recipient_list.insert(index, f"{marker}  {address}")
        recipient_list.itemconfig(index, foreground=color)
        recipient_list.see(index)
        status_by_index[index] = status
        if detail:
            detail_by_index[index] = detail

        successes = sum(value == "success" for value in status_by_index.values())
        failures = sum(value == "failed" for value in status_by_index.values())
        completed = successes + failures
        footer_status.configure(
            text=(
                f"正在发送：{completed}/{recipient_list.size()}　"
                f"✅ 服务器接受 {successes}　❌ 提交失败 {failures}"
            )
        )

    def poll_progress() -> None:
        done_received = False
        try:
            while True:
                event = progress_events.get_nowait()
                event_type = event[0]
                if event_type == "status":
                    _, index, status, detail = event
                    update_recipient_status(index, status, detail)
                elif event_type == "fatal":
                    error_detail = event[1]
                    for index in range(recipient_list.size()):
                        if status_by_index.get(index) not in {"success", "failed"}:
                            update_recipient_status(index, "failed", error_detail)
                elif event_type == "done":
                    done_received = True
        except queue.Empty:
            pass

        if done_received:
            sending_state["active"] = False
            sending_state["completed"] = True
            successes = sum(value == "success" for value in status_by_index.values())
            failures = sum(value == "failed" for value in status_by_index.values())
            result["all_succeeded"] = (
                successes == recipient_list.size() and failures == 0
            )
            title.configure(text="本轮邮件提交完成，请核对右侧状态")
            footer_status.configure(
                text=(
                    f"提交完成：✅ 服务器接受 {successes}　❌ 提交失败 {failures}。"
                    "接受不等于最终送达，请结合退信核对。"
                )
            )
            confirm_button.configure(
                state=tk.NORMAL,
                text="关闭窗口",
                command=request_close,
            )
            cancel_button.configure(state=tk.DISABLED, text="发送已完成")
            if failures:
                first_failed_index = next(
                    (
                        index
                        for index, status in status_by_index.items()
                        if status == "failed"
                    ),
                    None,
                )
                detail = (
                    detail_by_index.get(first_failed_index, "")
                    if first_failed_index is not None
                    else ""
                )
                if detail:
                    footer_status.configure(
                        text=(
                            f"提交完成：✅ 服务器接受 {successes}　❌ 提交失败 {failures}。"
                            f"失败原因：{detail}"
                        )
                    )
            return

        if sending_state["active"]:
            root.after(100, poll_progress)

    preview_width = max(window_width - 430, 620)
    content_width = preview_width - 32
    margin_x = 16
    y_position = 16
    canvas_images = []

    def add_preview_text(
        text: str,
        *,
        font=("Microsoft YaHei UI", 10),
        fill="#1F1F1F",
        spacing_after=8,
    ) -> None:
        nonlocal y_position
        item = canvas.create_text(
            margin_x,
            y_position,
            text=text,
            anchor="nw",
            width=content_width,
            font=font,
            fill=fill,
        )
        bounds = canvas.bbox(item)
        y_position = (bounds[3] if bounds else y_position + 20) + spacing_after

    add_preview_text(
        f"主题：{_report_subject(report_date)}",
        font=("Microsoft YaHei UI", 11, "bold"),
        spacing_after=6,
    )
    add_preview_text(f"发件人：{sender}", spacing_after=4)
    estimated_batches = (len(recipients) + BCC_BATCH_SIZE - 1) // BCC_BATCH_SIZE
    add_preview_text(
        f"发送方式：每批最多 {BCC_BATCH_SIZE} 人密送，共约 {estimated_batches} 封；"
        f"批次间隔 {SMTP_BATCH_INTERVAL_SECONDS:g} 秒",
        fill="#188038",
        spacing_after=4,
    )
    add_preview_text(
        f"密送收件人：见右侧本轮邮件地址列表（当前 {len(recipients)} 个）",
        spacing_after=12,
    )
    add_preview_text(
        f"附件（{len(excel_paths)}）：",
        font=("Microsoft YaHei UI", 10, "bold"),
        spacing_after=4,
    )
    for attachment_path in excel_paths:
        size_kb = attachment_path.stat().st_size / 1024
        size_text = (
            f"{size_kb / 1024:.1f} MB"
            if size_kb >= 1024
            else f"{size_kb:.1f} KB"
        )
        add_preview_text(
            f"附件：{attachment_path.name}（{size_text}）",
            fill="#245B86",
            spacing_after=4,
        )

    y_position += 8
    canvas.create_line(
        margin_x,
        y_position,
        preview_width - margin_x,
        y_position,
        fill="#D9D9D9",
    )
    y_position += 16

    add_preview_text("领导：", spacing_after=12)
    add_preview_text("晚上好，可转债赎回进度跟踪请查收", spacing_after=16)
    add_preview_text(
        "下图为赎回触发价以0.001计：",
        font=("Microsoft YaHei UI", 10, "bold"),
        spacing_after=8,
    )

    for image_path, section_title in [
        (permille_image_path, "下图为赎回触发价以0.01计："),
        (percent_image_path, None),
    ]:
        with Image.open(image_path) as source_image:
            preview_image = source_image.convert("RGB")
        if preview_image.width > content_width:
            preview_height = round(
                preview_image.height * content_width / preview_image.width
            )
            preview_image = preview_image.resize(
                (content_width, preview_height),
                Image.Resampling.LANCZOS,
            )
        photo = ImageTk.PhotoImage(preview_image)
        canvas_images.append(photo)
        canvas.create_image(
            margin_x,
            y_position,
            image=photo,
            anchor="nw",
        )
        y_position += preview_image.height + 20
        if section_title:
            add_preview_text(
                section_title,
                font=("Microsoft YaHei UI", 10, "bold"),
                spacing_after=8,
            )

    canvas.configure(
        scrollregion=(0, 0, preview_width, y_position + 16),
    )
    canvas.images = canvas_images

    def scroll_preview(event) -> None:
        canvas.yview_scroll(int(-event.delta / 120), "units")

    canvas.bind("<Enter>", lambda event: canvas.bind_all("<MouseWheel>", scroll_preview))
    canvas.bind("<Leave>", lambda event: canvas.unbind_all("<MouseWheel>"))

    update_count()
    root.after(100, lambda: address_entry.focus_set())
    root.attributes("-topmost", True)
    root.after(500, lambda: root.attributes("-topmost", False))
    root.mainloop()
    return result["all_succeeded"]


def _add_excel_attachment(message: EmailMessage, path: Path) -> None:
    mime_type, _ = mimetypes.guess_type(path.name)
    if mime_type:
        maintype, subtype = mime_type.split("/", 1)
    else:
        maintype = "application"
        subtype = "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    message.add_attachment(
        path.read_bytes(),
        maintype=maintype,
        subtype=subtype,
        filename=path.name,
    )


def _report_subject(report_date: str) -> str:
    return f"【华创固收】可转债赎回进度跟踪-{report_date}"


def _build_message(
    sender: str,
    recipient: str,
    report_date: str,
    excel_paths: Sequence[Path],
    permille_image_path: Path,
    percent_image_path: Path,
) -> EmailMessage:
    message = EmailMessage()
    message["From"] = sender
    message["To"] = recipient
    message["Subject"] = _report_subject(report_date)
    message["Date"] = formatdate(localtime=True)
    message["Message-ID"] = make_msgid(domain=sender.rsplit("@", 1)[-1])

    message.set_content(
        "领导：\n\n"
        "晚上好，可转债赎回进度跟踪请查收。\n\n"
        "下图为赎回触发价以0.001计：\n"
        "（千分位长图见邮件正文）\n\n"
        "下图为赎回触发价以0.01计：\n"
        "（百分位长图见邮件正文）\n"
    )

    html = """
    <html>
      <body style="font-family: 'Microsoft YaHei', Arial, sans-serif; font-size: 14px; color: #000;">
        <p>领导：</p>
        <p>晚上好，可转债赎回进度跟踪请查收</p>
        <p><strong>下图为赎回触发价以0.001计：</strong></p>
        <p><img src="cid:permille_chart" alt="千分位赎回进度长图"
                style="display:block; max-width:100%; height:auto;"></p>
        <p><strong>下图为赎回触发价以0.01计：</strong></p>
        <p><img src="cid:percent_chart" alt="百分位赎回进度长图"
                style="display:block; max-width:100%; height:auto;"></p>
      </body>
    </html>
    """
    message.add_alternative(html, subtype="html")
    html_part = message.get_payload()[-1]
    html_part.add_related(
        permille_image_path.read_bytes(),
        maintype="image",
        subtype="png",
        cid="<permille_chart>",
        filename=permille_image_path.name,
        disposition="inline",
    )
    html_part.add_related(
        percent_image_path.read_bytes(),
        maintype="image",
        subtype="png",
        cid="<percent_chart>",
        filename=percent_image_path.name,
        disposition="inline",
    )

    for excel_path in excel_paths:
        _add_excel_attachment(message, excel_path)
    return message


def _recipient_batches(
    recipients: Sequence[str],
    batch_size: int = BCC_BATCH_SIZE,
) -> list[list[str]]:
    if batch_size <= 0:
        raise ValueError("每批收件人数必须大于0。")
    return [
        list(recipients[start:start + batch_size])
        for start in range(0, len(recipients), batch_size)
    ]


def send_redemption_report(
    *,
    sender: str,
    recipients: Sequence[str],
    report_date: str,
    excel_paths: Sequence[Path],
    permille_image_path: Path,
    percent_image_path: Path,
    recipient_workbook_path: Path | None = None,
    interval_seconds: float = SMTP_BATCH_INTERVAL_SECONDS,
) -> bool:
    recipients = [address.strip() for address in recipients if address.strip()]
    if not recipients:
        raise ValueError("收件人列表不能为空。")

    excel_paths = _validate_files(excel_paths)
    permille_image_path, percent_image_path = _validate_files(
        [permille_image_path, percent_image_path]
    )
    def send_action(
        confirmed_recipients: Sequence[str],
        report_progress: Callable[[int, str, str], None],
    ) -> None:
        authorization_code = load_authorization_code()
        context = ssl.create_default_context()

        with smtplib.SMTP_SSL(
            SMTP_HOST,
            SMTP_PORT,
            context=context,
            timeout=30,
        ) as smtp:
            smtp.login(sender, authorization_code)
            success_count = 0
            failure_count = 0
            batches = _recipient_batches(confirmed_recipients)
            recipient_index = {
                address.casefold(): index
                for index, address in enumerate(confirmed_recipients)
            }

            for batch_number, batch in enumerate(batches, start=1):
                for recipient in batch:
                    report_progress(recipient_index[recipient.casefold()], "sending", "")

                message = _build_message(
                    sender,
                    sender,
                    report_date,
                    excel_paths,
                    permille_image_path,
                    percent_image_path,
                )
                try:
                    refused = smtp.send_message(
                        message,
                        from_addr=sender,
                        to_addrs=batch,
                    )
                except smtplib.SMTPRecipientsRefused as exc:
                    refused = exc.recipients

                refused_by_address = {
                    str(address).casefold(): reason
                    for address, reason in refused.items()
                }
                for recipient in batch:
                    index = recipient_index[recipient.casefold()]
                    refusal = refused_by_address.get(recipient.casefold())
                    if refusal is None:
                        report_progress(
                            index,
                            "success",
                            f"第 {batch_number}/{len(batches)} 批：126服务器已接受",
                        )
                        success_count += 1
                        continue

                    if isinstance(refusal, tuple) and len(refusal) >= 2:
                        code, response = refusal[0], refusal[1]
                        if isinstance(response, bytes):
                            response = response.decode("utf-8", errors="replace")
                        error_detail = f"SMTP {code}: {response}"
                    else:
                        error_detail = str(refusal)
                    report_progress(index, "failed", error_detail)
                    failure_count += 1

                print(
                    f"[密送批次 {batch_number}/{len(batches)}] "
                    f"本批 {len(batch)} 个地址，拒绝 {len(refused)} 个。"
                )
                if batch_number < len(batches) and interval_seconds > 0:
                    _time.sleep(interval_seconds)

            print(
                "[分批密送提交完成] "
                f"服务器接受 {success_count} 个地址，"
                f"拒绝 {failure_count} 个地址。"
            )

    outcome = confirm_recipients_with_preview(
        sender=sender,
        recipients=recipients,
        recipient_workbook_path=recipient_workbook_path,
        report_date=report_date,
        excel_paths=excel_paths,
        permille_image_path=permille_image_path,
        percent_image_path=percent_image_path,
        send_action=send_action,
    )
    if outcome is None:
        print("[邮件已取消] 用户未确认发送，本轮未发送任何邮件。")
        return False
    return outcome



DAYS_TODAY = 0
EXCLUDE_CODES = {"128085.SZ"}
EXCEL_REDEMPTION_NOTICE_FORMULA = "=@cb_clause_calloption_indicativedatey(A2)"
EXCEL_CALCULATION_TIMEOUT_SECONDS = 120

EMAIL_ENABLED = True
EMAIL_SENDER = "lizongyang_bond@126.com"
EMAIL_RECIPIENT_WORKBOOK = Path("条款跟踪邮箱列表.xlsx")
SCRIPT_DIR = Path(__file__).resolve().parent
LOCAL_HEADER_IMAGE = SCRIPT_DIR / "条款表头.png"
LOCAL_IMAGE_ASSETS = (
    LOCAL_HEADER_IMAGE,
)
PERMILLE_FOOTER_TEXT = "备注：赎回触发价的单位为0.001"
PERCENT_FOOTER_TEXT = "备注：赎回触发价的单位为0.01"
FOOTER_IMAGE_SIZE = (2989, 43)
INTERMEDIATE_IMAGE_NAMES = (
    "赎回表题头.png",
    "千分位赎回表尾.png",
    "百分位赎回表尾.png",
    "赎回累计触发天数个券.png",
    "已发布过不提前赎回公告.png",
    "赎回公告个券.png",
    "赎回累计触发天数个券（百分位）.png",
    "已发布过不提前赎回公告（百分位）.png",
    "赎回公告个券（百分位）.png",
)

# 不强赎信息：保留现有手工配置，并直接覆盖对应的 iFinD 数据。
MANUAL_NON_REDEMPTION_OVERRIDES = {
    "123158.SZ": {"不强赎公告日": "2026-08-12", "承诺何日之前不行使": "2027-02-12"},
    "111012.SH": {"不强赎公告日": "2026-07-10", "承诺何日之前不行使": "2026-10-10"},
}

# 强赎信息：仅在相应数据源缺失时补充，不覆盖有效的自动数据。
# 赎回公告日的自动数据源为 Excel 函数；赎回登记日、最后交易日的自动数据源仍为 iFinD。
MANUAL_COMPULSORY_REDEMPTION_OVERRIDES = {
    # "113000.SH": {
    #     "赎回公告日": "2026-01-01",
    #     "赎回登记日": "2026-01-20",
    #     "最后交易日": "2026-01-15",
    # },
}

REDEMPTION_INFO_FIELDS = (
    "ths_convertible_debt_short_name_cbond;"
    "ths_debt_rating_primary_rating_agency_bond;"
    "ths_redemp_triggercalc_mti_cbond;"
    "ths_redemp_triggercalc_time_int_cbond;"
    "ths_redemp_audit_ad_cbond;"
    "ths_last_td_date_convertible_cbond;"
    "ths_bond_balance_cbond;"
    "ths_un_conversion_ratio_cbond;"
    "ths_float_shares_dlt_rate_cbond;"
    "ths_bond_close_cbond;"
    "ths_holder_held_ratio_cbond;"
    "ths_transfer_value_cbond;"
    "ths_conversion_premium_rate_cbond;"
    "ths_pure_bond_premium_rate_cbond;"
    "ths_object_the_sw_bond"
)
REDEMPTION_INFO_PARAMS = ";;;;;;{date};{date};{date};{date},2;{date},1;{date};{date};{date};1,{date}"
NON_REDEMPTION_INFO_PARAMS = ";;;;;;;{date};{date};{date};{date},2;{date},1;{date};{date};{date};1,{date}"
REDEMPTION_INFO_COLUMNS = [
    "转债简称", "债项评级", "时间区间", "计算天数", "赎回登记日", "最后交易日", "转债余额",
    "未转股比例", "对流通股本稀释", "转债价格", "大股东持债比例", "平价", "转股溢价率", "纯债溢价率", "所属行业",
]
REDEMPTION_OUTPUT_COLUMNS = REDEMPTION_INFO_COLUMNS.copy()
REDEMPTION_OUTPUT_COLUMNS.insert(6, "赎回公告日")

NON_REDEMPTION_INFO_FIELDS = (
    "ths_convertible_debt_short_name_cbond;"
    "ths_debt_rating_primary_rating_agency_bond;"
    "ths_redemp_triggercalc_mti_cbond;"
    "ths_redemp_triggercalc_time_int_cbond;"
    "ths_not_compulsory_redemp_indicative_date_bond;"
    "ths_not_compulsory_redemp_enddate_cbond_bond;"
    "ths_redemp_trigger_ratio_cbond;"
    "ths_bond_balance_cbond;"
    "ths_un_conversion_ratio_cbond;"
    "ths_float_shares_dlt_rate_cbond;"
    "ths_bond_close_cbond;"
    "ths_holder_held_ratio_cbond;"
    "ths_transfer_value_cbond;"
    "ths_conversion_premium_rate_cbond;"
    "ths_pure_bond_premium_rate_cbond;"
    "ths_object_the_sw_bond"
)
NON_REDEMPTION_INFO_COLUMNS = [
    "转债简称", "债项评级", "时间区间", "计算天数", "不赎回公告日", "承诺何日之前不行使", "赎回累计触发天数", "转债余额",
    "未转股比例", "对流通股本稀释", "转债价格", "大股东持债比例", "平价", "转股溢价率", "纯债溢价率", "最早触发日期",
]

LAST_TRADE_FIELDS = (
    "ths_convertible_debt_short_name_cbond;"
    "ths_maturity_date_cbond;"
    "ths_last_td_date_convertible_cbond;"
    "ths_conversion_ed_cbond;"
    "ths_delist_date_bond;"
    "ths_surplus_term_d_bond;"
    "ths_bond_balance_cbond;"
    "ths_un_conversion_ratio_cbond;"
    "ths_maturity_redemp_price_cbond;"
    "ths_bond_close_cbond;"
    "ths_transfer_value_cbond;"
    "ths_conversion_premium_rate_cbond;"
    "ths_pure_bond_value_cbond;"
    "ths_pure_bond_premium_rate_cbond;"
    "ths_pure_bond_ytm_cbond;"
    "ths_object_the_sw_bond"
)
LAST_TRADE_COLUMNS = [
    "转债简称", "到期日期", "最后交易日", "最后转股日", "摘牌日期", "剩余天数", "转债余额", "未转股比例", "到期赎回价",
    "转债价格", "平价", "转股溢价率", "纯债价值", "纯债溢价率", "YTM", "所属行业",
]

TABLE_COLUMN_WIDTHS = [0.06, 0.06, 0.06, 0.06, 0.06, 0.1, 0.1, 0.1, 0.06, 0.06, 0.08, 0.06, 0.08, 0.06, 0.06, 0.06, 0.06]
IFIND_CREDENTIAL_FILE = Path(__file__).resolve().parent / "ifind账号.txt"


def load_ifind_credentials() -> tuple[str, str]:
    """从项目目录的 ifind账号.txt 读取统一登录账号。"""
    if not IFIND_CREDENTIAL_FILE.is_file():
        raise FileNotFoundError(f"未找到iFinD账号文件：{IFIND_CREDENTIAL_FILE}")
    config = ConfigParser(interpolation=None)
    config.read(IFIND_CREDENTIAL_FILE, encoding="utf-8")
    username = config.get("ifind", "username", fallback="").strip()
    password = config.get("ifind", "password", fallback="").strip()
    if not username or not password:
        raise RuntimeError("ifind账号.txt中的[ifind] username或password为空")
    return username, password


def print_ifind_usage() -> None:
    """显示iFinD各数据项的已用额度比例。"""
    try:
        result = THS_DataStatistics()
        tables = result.get("tables", {}) if isinstance(result, dict) else {}
        if not tables:
            detail = result.get("errmsg", "未返回额度数据") if isinstance(result, dict) else str(result)
            print(f"[警告] iFinD使用额度查询失败：{detail}")
            return
        print("iFinD使用额度：")
        for key, value in tables.items():
            ratio = value.get("ratio", "N/A") if isinstance(value, dict) else value
            print(f"{key} 已用：{ratio}")
    except Exception as exc:
        print(f"[警告] iFinD使用额度查询失败：{exc}")


def ths_login_demo():
    username, password = load_ifind_credentials()
    ths_login = THS_iFinDLogin(username, password)
    print(ths_login)
    print("登录失败" if ths_login not in (0, -201) else "登录成功")
    if ths_login in (0, -201):
        print_ifind_usage()
    return ths_login


def make_paths():
    mmdd_today = _time.strftime("%m%d", _time.localtime())
    yyyymmdd_today = _time.strftime("%Y%m%d", _time.localtime())
    folder = Path(f"{mmdd_today}数据更新") / "赎回数据更新"
    folder.mkdir(parents=True, exist_ok=True)
    return {
        "mmdd": mmdd_today,
        "yyyymmdd": yyyymmdd_today,
        "folder": folder,
        "permille_xlsx": folder / f"【华创固收】转债赎回信息日度跟踪-{yyyymmdd_today}自动更新【千分位版】.xlsx",
        "percent_xlsx": folder / f"【华创固收】转债赎回信息日度跟踪-{yyyymmdd_today}自动更新【百分位版】.xlsx",
        "permille_png": folder / f"【华创固收】转债赎回信息日度跟踪-{yyyymmdd_today}自动更新【千分位版】.png",
        "percent_png": folder / f"【华创固收】转债赎回信息日度跟踪-{yyyymmdd_today}自动更新【百分位版】.png",
    }


def get_last_trade_date():
    offset = 0
    last_date = _time.strftime("%Y-%m-%d", _time.localtime())
    trade_date = THS_Date_Offset(
        "212001",
        f"dateType:0,period:D,offset:{offset},dateFormat:0,output:singledate",
        last_date,
    ).data
    if trade_date is None:
        raise RuntimeError("THS_Date_Offset返回None，请确认iFinD登录成功后再运行。")
    print(f"使用日期： {trade_date}")
    return trade_date


def get_cb_basic_trade(last_date):
    formatted_date = datetime.strptime(last_date, "%Y-%m-%d").strftime("%Y%m%d")
    cb_list = THS_DR(
        "p00570",
        f"jyzt=未到期;sfdb=全部;jysc=全部;edate={formatted_date}",
        "jydm:Y,jydm_mc:Y,p00570_f001:Y,p00570_f019:Y",
        "format:dataframe",
    ).data
    codes = [code for code in cb_list.set_index("jydm").index.to_list() if code not in EXCLUDE_CODES]
    cb_list_str = ", ".join(codes)

    trade_status = THS_DS(
        cb_list_str,
        "ths_turnover_ratio_cbond",
        "",
        "Fill:Blank,mode:thscode",
        str(last_date),
        str(last_date),
    ).data
    trade_status = trade_status.set_index("time").T.iloc[:, [0, -1]]
    trade_status = trade_status[
        ~(
            (trade_status.iloc[:, 0].isna() & trade_status.iloc[:, -1].isna())
            | ((trade_status.iloc[:, 0] == 0) & (trade_status.iloc[:, -1] == 0))
        )
    ]
    cb_list_trade = ", ".join(trade_status.index.astype(str))

    cb_basic_trade = THS_BD(
        cb_list_trade,
        "ths_convertible_debt_short_name_cbond;ths_stock_code_cbond;ths_stock_short_name_cbond;ths_issue_method_cbond;ths_trading_status_bond;ths_bond_balance_cbond;ths_listed_date_cbond",
        f";;;;;{str(last_date)};",
    ).data
    cb_basic_trade = cb_basic_trade.set_index("thscode").rename_axis("转债代码")
    cb_basic_trade.columns = ["转债简称", "正股代码", "正股简称", "发行方式", "交易状态", "转债余额", "上市日期"]
    cb_basic_trade = cb_basic_trade[~cb_basic_trade["发行方式"].str.contains("定向")]
    cb_basic_trade = cb_basic_trade[~cb_basic_trade.index.str.contains("NQ")]
    return cb_basic_trade, ", ".join(cb_basic_trade.index.astype(str))


def get_date_range(last_date):
    return THS_Date_Offset(
        "212001",
        "dateType:0,period:D,offset:-30,dateFormat:0,output:sequencedate",
        last_date,
        "format:dict",
    ).data["time"]


def convert_status(val):
    if pd.isna(val):
        return "数据缺失"
    if val == 0:
        return "停牌"
    return "正常上市"


def fetch_ds_matrix(code_list, indicator, days, params="", options="mode:thscode", mapper=None):
    frames = []
    for day in tqdm(days):
        data = THS_DS(code_list, indicator, params, options, day, day).data
        frames.append(data.set_index("time"))
    result = pd.concat(frames, axis=0)
    if mapper is not None:
        result = result.apply(lambda col: col.map(mapper))
    return result


def build_daily_sheet(matrix, cb_basic_trade):
    rows = []
    for code, row in cb_basic_trade.iterrows():
        rows.append([code, row["转债简称"]] + list(matrix[str(code)]))
    return pd.DataFrame(rows, columns=["代码", "名称"] + list(matrix.index)).set_index("代码")


def is_excel_error_value(value):
    if isinstance(value, str):
        return value.strip().startswith("#")
    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        return value < -2_000_000_000
    return False


def is_excel_pending_value(value):
    if not isinstance(value, str):
        return False
    text = value.strip().lower()
    return any(marker in text for marker in ["fetching", "loading", "requesting"])


def normalize_date_value(value):
    if value is None or is_excel_error_value(value) or is_excel_pending_value(value):
        return pd.NaT
    if isinstance(value, pd.Timestamp):
        return value.normalize()
    if isinstance(value, datetime):
        return pd.Timestamp(value).normalize()

    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        if pd.isna(value) or value <= 0:
            return pd.NaT
        if float(value).is_integer() and 19_000_101 <= int(value) <= 29_991_231:
            return pd.to_datetime(str(int(value)), format="%Y%m%d", errors="coerce")
        if value < 2_958_466:
            return (pd.Timestamp("1899-12-30") + pd.to_timedelta(float(value), unit="D")).normalize()
        return pd.NaT

    text = str(value).strip()
    if text in {"", "0", "——", "--", "None", "nan", "NaT"}:
        return pd.NaT
    if re.fullmatch(r"\d{8}", text):
        return pd.to_datetime(text, format="%Y%m%d", errors="coerce")
    return pd.to_datetime(text, errors="coerce")


def validate_manual_compulsory_redemption_overrides():
    allowed_fields = {"赎回公告日", "赎回登记日", "最后交易日"}
    for bond_code, values in MANUAL_COMPULSORY_REDEMPTION_OVERRIDES.items():
        unexpected_fields = set(values).difference(allowed_fields)
        if unexpected_fields:
            raise ValueError(f"强赎手工配置 {bond_code} 包含未知字段: {sorted(unexpected_fields)}")
        for field, value in values.items():
            if value not in (None, "") and pd.isna(normalize_date_value(value)):
                raise ValueError(f"强赎手工配置 {bond_code} 的{field}不是有效日期: {value}")


def fetch_excel_redemption_notice_dates(bond_codes):
    codes = [str(code) for code in bond_codes]
    result = pd.Series(pd.NaT, index=codes, dtype="datetime64[ns]", name="赎回公告日")
    result.index.name = "代码"
    if not codes:
        return result

    try:
        import pythoncom
        import win32com.client
    except ImportError as exc:
        raise RuntimeError("自动获取赎回公告日需要安装 pywin32。") from exc

    excel = None
    workbook = None
    pythoncom.CoInitialize()
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.AskToUpdateLinks = False

        workbook = excel.Workbooks.Add()
        excel.Calculation = -4105  # xlCalculationAutomatic
        worksheet = workbook.Worksheets(1)
        worksheet.Name = "赎回公告日批量查询"
        worksheet.Range("A1:B1").Value = (("代码", "赎回公告日"),)

        last_row = len(codes) + 1
        worksheet.Range(f"A2:A{last_row}").Value = tuple((code,) for code in codes)
        formula_cell = worksheet.Range("B2")
        try:
            formula_cell.Formula2 = EXCEL_REDEMPTION_NOTICE_FORMULA
        except Exception:
            formula_cell.Formula = EXCEL_REDEMPTION_NOTICE_FORMULA
        if last_row > 2:
            worksheet.Range(f"B2:B{last_row}").FillDown()

        excel.CalculateFullRebuild()
        try:
            excel.CalculateUntilAsyncQueriesDone()
        except Exception:
            pass

        deadline = _time.time() + EXCEL_CALCULATION_TIMEOUT_SECONDS
        while True:
            raw_values = worksheet.Range(f"B2:B{last_row}").Value2
            if len(codes) == 1 and not isinstance(raw_values, tuple):
                values = [raw_values]
            else:
                values = [row[0] if isinstance(row, tuple) else row for row in raw_values]
            pending_codes = [
                code for code, value in zip(codes, values) if is_excel_pending_value(value)
            ]
            if excel.CalculationState == 0 and not pending_codes:  # xlDone
                break
            if _time.time() >= deadline:
                sample = pending_codes[:10]
                raise TimeoutError(
                    f"Excel 计算赎回公告日超过 {EXCEL_CALCULATION_TIMEOUT_SECONDS} 秒，"
                    f"仍有 {len(pending_codes)} 只处于等待状态，示例: {sample}"
                )
            _time.sleep(0.5)

        error_codes = [code for code, value in zip(codes, values) if is_excel_error_value(value)]
        if len(error_codes) == len(codes):
            raise RuntimeError(
                "Excel 赎回公告日函数全部返回错误，请确认数据插件已加载并登录。"
            )
        if error_codes:
            print(f"[Excel函数] {len(error_codes)} 只转债返回错误，将按缺失值处理并尝试手工补充。")

        for code, value in zip(codes, values):
            result.at[code] = normalize_date_value(value)
    except Exception as exc:
        if isinstance(exc, (RuntimeError, TimeoutError)):
            raise
        raise RuntimeError(f"自动控制 Excel 获取赎回公告日失败: {exc}") from exc
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

    return result


def apply_manual_compulsory_notice_overrides(redemption_notice):
    validate_manual_compulsory_redemption_overrides()
    for bond_code, values in MANUAL_COMPULSORY_REDEMPTION_OVERRIDES.items():
        if bond_code not in redemption_notice.index:
            print(f"[手动补充] 忽略不在当日列表中的强赎信息: {bond_code}")
            continue
        manual_date = normalize_date_value(values.get("赎回公告日"))
        if pd.isna(redemption_notice.at[bond_code]) and pd.notna(manual_date):
            redemption_notice.at[bond_code] = manual_date
            print(f"[手动补充] {bond_code} 赎回公告日: {manual_date.strftime('%Y-%m-%d')}")
    return redemption_notice


def apply_manual_non_redemption_overrides(notice_df, commitment_df):
    if not MANUAL_NON_REDEMPTION_OVERRIDES:
        return
    manual_df = pd.DataFrame.from_dict(MANUAL_NON_REDEMPTION_OVERRIDES, orient="index")
    manual_df.index.name = "thscode"

    if "不强赎公告日" in manual_df.columns:
        manual_notice_dates = pd.to_datetime(manual_df["不强赎公告日"], errors="coerce")
        valid_codes = manual_notice_dates.index.intersection(notice_df.index)
        skipped_codes = manual_notice_dates.index.difference(notice_df.index)
        if len(skipped_codes):
            print(f"[手动补充] 忽略不在当日列表中的不强赎公告日: {list(skipped_codes)}")
        if len(valid_codes):
            notice_df.loc[valid_codes, "ths_not_compulsory_redemp_indicative_date_bond"] = manual_notice_dates.loc[valid_codes]

    if "承诺何日之前不行使" in manual_df.columns:
        manual_commitment_dates = pd.to_datetime(manual_df["承诺何日之前不行使"], errors="coerce")
        valid_codes = manual_commitment_dates.index.intersection(commitment_df.index)
        skipped_codes = manual_commitment_dates.index.difference(commitment_df.index)
        if len(skipped_codes):
            print(f"[手动补充] 忽略不在当日列表中的不赎回承诺期: {list(skipped_codes)}")
        if len(valid_codes):
            commitment_df.loc[valid_codes, "ths_not_compulsory_redemp_enddate_cbond_bond"] = manual_commitment_dates.loc[valid_codes]


def build_total_table(code_list, cb_basic_trade):
    redeem_max_span = THS_BD(code_list, "ths_redemp_triggercalc_mti_cbond", "").data.set_index("thscode")
    redeem_span = THS_BD(code_list, "ths_redemp_triggercalc_time_int_cbond", "").data.set_index("thscode")
    notice_date = THS_BD(code_list, "ths_not_compulsory_redemp_indicative_date_bond", "").data.set_index("thscode")
    notice_date["ths_not_compulsory_redemp_indicative_date_bond"] = pd.to_datetime(
        notice_date["ths_not_compulsory_redemp_indicative_date_bond"],
        format="%Y%m%d",
        errors="coerce",
    )
    commitment_date = THS_BD(code_list, "ths_not_compulsory_redemp_enddate_cbond_bond", "").data.set_index("thscode")
    commitment_date["ths_not_compulsory_redemp_enddate_cbond_bond"] = pd.to_datetime(
        commitment_date["ths_not_compulsory_redemp_enddate_cbond_bond"],
        format="%Y%m%d",
        errors="coerce",
    )
    apply_manual_non_redemption_overrides(notice_date, commitment_date)

    redeem_start = THS_BD(code_list, "ths_cndtn_redemp_sd_cbond", "").data.set_index("thscode")
    redeem_start["ths_cndtn_redemp_sd_cbond"] = pd.to_datetime(
        redeem_start["ths_cndtn_redemp_sd_cbond"],
        format="%Y%m%d",
        errors="coerce",
    )
    redemption_notice = fetch_excel_redemption_notice_dates(cb_basic_trade.index)
    redemption_notice = apply_manual_compulsory_notice_overrides(redemption_notice)

    total_table = pd.DataFrame(index=cb_basic_trade.index)
    total_table.index.name = "代码"
    total_table["名称"] = cb_basic_trade["转债简称"]
    total_table["时间区间MAX"] = redeem_max_span["ths_redemp_triggercalc_mti_cbond"]
    total_table["计算天数"] = redeem_span["ths_redemp_triggercalc_time_int_cbond"]
    total_table["不强赎公告日"] = notice_date["ths_not_compulsory_redemp_indicative_date_bond"]
    total_table["承诺何日之前不行使"] = commitment_date["ths_not_compulsory_redemp_enddate_cbond_bond"]
    total_table["赎回起始日"] = redeem_start["ths_cndtn_redemp_sd_cbond"]
    total_table["赎回公告日"] = redemption_notice.reindex(total_table.index)
    return total_table


def fetch_holder_info(code_list, last_date):
    stock_holder_data = THS_BD(
        code_list,
        "ths_convertible_debt_short_name_cbond;ths_stock_code_cbond;ths_stock_short_name_cbond;ths_major_shareholder_name_bond;ths_big_holder_held_ratio_bond",
        f";;;{last_date},1;{last_date},1",
    ).data
    stock_holder_data = stock_holder_data.set_index("thscode")
    stock_holder_data.columns = ["转债简称", "正股代码", "正股简称", "持股第一名大股东", "持股第一名大股东持有比例"]

    holder_names = {}
    holder_ratios = {}
    for order in tqdm(range(1, 11)):
        holder_names[order] = THS_BD(code_list, "ths_holder_name_cbond", f"{last_date},{order}").data.set_index("thscode")
        holder_names[order].columns = [f"第{order}名"]
        holder_names[order].index.name = "代码"

        holder_ratios[order] = THS_BD(code_list, "ths_holder_held_ratio_cbond", f"{last_date},{order}").data.set_index("thscode")
        holder_ratios[order].columns = [f"第{order}名"]
        holder_ratios[order].index.name = "代码"

    holder_name_df = pd.concat(holder_names.values(), axis=1)
    holder_ratio_df = pd.concat(holder_ratios.values(), axis=1)
    holder_info = pd.concat([holder_name_df, holder_ratio_df], axis=1)

    stock_holder_hold_cb_bond = stock_holder_data.drop(["正股代码", "正股简称", "持股第一名大股东持有比例"], axis=1)
    ratios = []
    for index, holder in zip(stock_holder_hold_cb_bond.index, stock_holder_hold_cb_bond["持股第一名大股东"].values):
        try:
            series = holder_name_df.loc[index]
            matched_index = series[series == holder].index[0]
            ratios.append(holder_ratio_df.loc[index][matched_index])
        except Exception:
            ratios.append(np.nan)
    stock_holder_hold_cb_bond["持股第一名大股东持债比例"] = ratios
    return holder_info, stock_holder_hold_cb_bond


def build_base_tables(cb_basic_trade, code_list, days, last_date):
    stock_close = fetch_ds_matrix(code_list, "ths_stock_close_cbond", days, params="100")
    trigger_price = fetch_ds_matrix(code_list, "ths_redemp_trigger_price_cbond", days)
    trade_status = fetch_ds_matrix(
        code_list,
        "ths_turnover_ratio_cbond",
        days,
        options="Fill:Blank,mode:thscode",
        mapper=convert_status,
    )

    stock_close_sheet = build_daily_sheet(stock_close, cb_basic_trade)
    trigger_price_sheet = build_daily_sheet(trigger_price, cb_basic_trade)
    trade_status_sheet = build_daily_sheet(trade_status, cb_basic_trade)
    total_table = build_total_table(code_list, cb_basic_trade)
    holder_info, stock_holder_hold_cb_bond = fetch_holder_info(code_list, last_date)
    return stock_close_sheet, trigger_price_sheet, trade_status_sheet, total_table, holder_info, stock_holder_hold_cb_bond


def clean_suspended_and_commitment(stock_close_sheet, trigger_price_sheet, trade_status_sheet, total_table):
    stock_close_sheet = stock_close_sheet.drop(columns=["名称"]).copy()
    trigger_price_sheet = trigger_price_sheet.drop(columns=["名称"]).copy()
    trade_status_sheet = trade_status_sheet.drop(columns=["名称"]).copy()
    total_table = total_table.drop(columns=["名称"]).copy()

    mask_suspended = trade_status_sheet.eq("停牌")
    for sheet in [stock_close_sheet, trigger_price_sheet, trade_status_sheet]:
        sheet[mask_suspended] = np.nan

    date_columns = pd.to_datetime(trigger_price_sheet.columns)
    for bond_code in tqdm(stock_close_sheet.index):
        commitment_date = total_table.loc[bond_code, "承诺何日之前不行使"]
        redemption_start_date = total_table.loc[bond_code, "赎回起始日"]
        redemption_notice_date = total_table.loc[bond_code, "赎回公告日"]

        if not pd.isnull(redemption_notice_date):
            trigger_price_sheet.loc[bond_code] = None
            stock_close_sheet.loc[bond_code] = None

        if not pd.isnull(redemption_start_date):
            mask = date_columns < redemption_start_date
            trigger_price_sheet.loc[bond_code, mask] = None
            stock_close_sheet.loc[bond_code, mask] = None

        if not pd.isnull(commitment_date):
            mask = date_columns <= commitment_date
            trigger_price_sheet.loc[bond_code, mask] = None
            stock_close_sheet.loc[bond_code, mask] = None

    return stock_close_sheet, trigger_price_sheet, trade_status_sheet, total_table


def write_base_workbook(path, stock_close_sheet, trigger_price_sheet, trade_status_sheet, total_table, holder_info):
    with pd.ExcelWriter(path, mode="w") as writer:
        stock_close_sheet.to_excel(writer, sheet_name="正股收盘价", index=True)
        trigger_price_sheet.to_excel(writer, sheet_name="赎回触发价", index=True)
        trade_status_sheet.to_excel(writer, sheet_name="交易状态", index=True)
        total_table.to_excel(writer, sheet_name="总表", index=True)
        holder_info.to_excel(writer, sheet_name="前十大转债持有人", index=True)


def write_percent_base_workbook(path, stock_close_sheet, trigger_price_sheet, total_table, holder_info):
    with pd.ExcelWriter(path, mode="w") as writer:
        stock_close_sheet.to_excel(writer, sheet_name="正股收盘价", index=True)
        trigger_price_sheet.to_excel(writer, sheet_name="赎回触发价", index=True)
        total_table.to_excel(writer, sheet_name="总表", index=True)
        holder_info.to_excel(writer, sheet_name="前十大转债持有人", index=True)


def calculate_redemption_count(stock_close_sheet, trigger_price_sheet, total_table, round_trigger=False):
    rows = []
    for bond_code in tqdm(stock_close_sheet.index):
        time_interval = total_table.loc[bond_code, "时间区间MAX"]
        if time_interval == 0:
            continue
        time_interval = int(time_interval)
        stock_close_data = stock_close_sheet.loc[bond_code]
        trigger_data = trigger_price_sheet.loc[bond_code]
        if round_trigger:
            trigger_data = trigger_data.round(2)
        condition_mask = (
            (stock_close_data.iloc[-time_interval:] >= trigger_data.iloc[-time_interval:])
            & (stock_close_data.iloc[-time_interval:] > 0)
        )
        count = condition_mask.sum()
        earliest_date = condition_mask[condition_mask].index[0] if count > 0 else np.nan
        rows.append({"代码": bond_code, "赎回累计触发天数": count, "最早触发日期": earliest_date})
    return pd.DataFrame(rows).set_index("代码")


def apply_manual_compulsory_detail_overrides(redemption_bond_info):
    for bond_code, values in MANUAL_COMPULSORY_REDEMPTION_OVERRIDES.items():
        if bond_code not in redemption_bond_info.index:
            continue
        for field in ["赎回登记日", "最后交易日"]:
            manual_date = normalize_date_value(values.get(field))
            source_date = normalize_date_value(redemption_bond_info.at[bond_code, field])
            if pd.isna(source_date) and pd.notna(manual_date):
                redemption_bond_info.at[bond_code, field] = manual_date.strftime("%Y%m%d")
                print(f"[手动补充] {bond_code} {field}: {manual_date.strftime('%Y-%m-%d')}")


def fetch_redemption_bond_info(stock_close_sheet, total_table, last_date, stock_holder_hold_cb_bond):
    redemption_bond = []
    for bond_code in tqdm(stock_close_sheet.index):
        if not pd.isnull(total_table.loc[bond_code, "赎回公告日"]):
            redemption_bond.append(bond_code)

    if not redemption_bond:
        empty = pd.DataFrame(columns=REDEMPTION_OUTPUT_COLUMNS)
        empty.index.name = "代码"
        return empty

    redemption_bond_str = ", ".join(redemption_bond)
    redemption_bond_info = THS_BD(
        redemption_bond_str,
        REDEMPTION_INFO_FIELDS,
        REDEMPTION_INFO_PARAMS.format(date=last_date),
    ).data
    redemption_bond_info = redemption_bond_info.set_index("thscode")

    if len(redemption_bond_info.columns) != 1:
        redemption_bond_info.columns = REDEMPTION_INFO_COLUMNS
        notice_values = pd.Series(index=redemption_bond_info.index, dtype="object")
        redemption_bond_info.insert(6, "赎回公告日", notice_values)
        for bond_code in redemption_bond_info.index.intersection(total_table.index):
            notice_date = normalize_date_value(total_table.at[bond_code, "赎回公告日"])
            if pd.notna(notice_date):
                redemption_bond_info.at[bond_code, "赎回公告日"] = notice_date.strftime("%Y%m%d")
        apply_manual_compulsory_detail_overrides(redemption_bond_info)

    apply_holder_ratio(redemption_bond_info, stock_holder_hold_cb_bond)
    redemption_bond_info.index.name = "代码"

    if redemption_bond_info.columns[0] != "OUTMESSAGE":
        for col in ["赎回登记日", "最后交易日", "赎回公告日"]:
            redemption_bond_info[col] = pd.to_datetime(redemption_bond_info[col], format="%Y%m%d", errors="coerce")
            redemption_bond_info[col] = redemption_bond_info[col].apply(lambda x: x.strftime("%Y-%m-%d") if isinstance(x, pd.Timestamp) else x)
        redemption_bond_info = redemption_bond_info.round({
            "时间区间": 0,
            "计算天数": 0,
            "转债余额": 2,
            "对流通股本稀释": 2,
            "转债价格": 2,
            "未转股比例": 2,
            "平价": 2,
            "转股溢价率": 2,
            "纯债溢价率": 2,
        })
        redemption_bond_info = redemption_bond_info.sort_values(by="赎回公告日", ascending=False)

    for col in ["赎回登记日", "最后交易日"]:
        if col in redemption_bond_info.columns:
            redemption_bond_info[col] = redemption_bond_info[col].fillna("")
    return redemption_bond_info.fillna("——")


def apply_holder_ratio(df, stock_holder_hold_cb_bond):
    if df.columns[0] == "OUTMESSAGE":
        return
    holder_ratios = stock_holder_hold_cb_bond["持股第一名大股东持债比例"]
    for index in df.index:
        if index in stock_holder_hold_cb_bond.index:
            df.at[index, "大股东持债比例"] = pd.to_numeric(holder_ratios.loc[index], errors="coerce")


def fetch_non_redemption_info(cb_basic_trade, redemption_bond_info, redemption_count, total_table, last_date, stock_holder_hold_cb_bond):
    non_redemption_bond = cb_basic_trade.index.difference(redemption_bond_info.index)
    non_redemption_bond = ",".join(non_redemption_bond)
    non_redemption_bond_info = THS_BD(
        non_redemption_bond,
        NON_REDEMPTION_INFO_FIELDS,
        NON_REDEMPTION_INFO_PARAMS.format(date=last_date),
    ).data
    non_redemption_bond_info = non_redemption_bond_info.set_index("thscode")
    non_redemption_bond_info.columns = NON_REDEMPTION_INFO_COLUMNS

    indexes_to_drop = redemption_count.index.isin(redemption_bond_info.index)
    redemption_count_filtered = redemption_count[~indexes_to_drop]
    notice_date_col = "不赎回公告日" if "不赎回公告日" in total_table.columns else "不强赎公告日"
    notice_date_filtered = total_table.loc[redemption_count_filtered.index, [notice_date_col]].rename(columns={notice_date_col: "不赎回公告日"})
    commitment_date_filtered = total_table.loc[:, ["承诺何日之前不行使"]][~indexes_to_drop]

    if redemption_count_filtered.index.equals(non_redemption_bond_info.index):
        non_redemption_bond_info["赎回累计触发天数"] = redemption_count_filtered["赎回累计触发天数"]
        non_redemption_bond_info["不赎回公告日"] = notice_date_filtered["不赎回公告日"]
        non_redemption_bond_info["承诺何日之前不行使"] = commitment_date_filtered["承诺何日之前不行使"]
        non_redemption_bond_info["最早触发日期"] = redemption_count["最早触发日期"]
    else:
        print("索引顺序不匹配，无法直接赋值。")

    apply_holder_ratio(non_redemption_bond_info, stock_holder_hold_cb_bond)
    non_redemption_bond_info["不赎回公告日"] = pd.to_datetime(non_redemption_bond_info["不赎回公告日"]).dt.strftime("%Y-%m-%d")
    non_redemption_bond_info["承诺何日之前不行使"] = pd.to_datetime(non_redemption_bond_info["承诺何日之前不行使"]).dt.strftime("%Y-%m-%d")
    non_redemption_bond_info = non_redemption_bond_info.round({
        "时间区间": 0,
        "计算天数": 0,
        "转债余额": 2,
        "未转股比例": 2,
        "对流通股本稀释": 2,
        "转债价格": 2,
        "平价": 2,
        "转股溢价率": 2,
        "纯债溢价率": 2,
    })
    non_redemption_bond_info.index.name = "代码"

    counting = non_redemption_bond_info[
        non_redemption_bond_info["赎回累计触发天数"] != 0
    ].sort_values(by="赎回累计触发天数", ascending=False)
    commitment = non_redemption_bond_info[
        (non_redemption_bond_info["不赎回公告日"].notna())
        & (non_redemption_bond_info["赎回累计触发天数"] == 0)
    ].sort_values(by="不赎回公告日", ascending=False)
    other = non_redemption_bond_info[
        ~non_redemption_bond_info.index.isin(counting.index)
        & ~non_redemption_bond_info.index.isin(commitment.index)
    ]
    return counting.replace(np.nan, ""), commitment.replace(np.nan, ""), other.replace(np.nan, "")


def fetch_lasttrade_info(cb_list_trade, redemption_bond_info, last_date):
    lasttradedate = THS_BD(cb_list_trade, "ths_last_td_date_convertible_cbond", "").data
    lasttradedate = lasttradedate.set_index("thscode")
    lasttradedate.dropna(inplace=True)
    lasttradedate = lasttradedate.sort_values(by="ths_last_td_date_convertible_cbond", ascending=False)
    lasttradedate = lasttradedate[~lasttradedate.index.isin(redemption_bond_info.index)]
    lasttrade_list = ", ".join(lasttradedate.index.astype(str))

    if lasttrade_list == "":
        return pd.DataFrame([])

    lasttrade_info = THS_BD(
        lasttrade_list,
        LAST_TRADE_FIELDS,
        f";;;;;{last_date};{last_date};{last_date};;{last_date},2;{last_date};{last_date};{last_date};{last_date};{last_date};1,{last_date}",
    ).data
    lasttrade_info = lasttrade_info.set_index("thscode")
    lasttrade_info.columns = LAST_TRADE_COLUMNS
    lasttrade_info = lasttrade_info.round({
        "转债余额": 2,
        "未转股比例": 2,
        "转债价格": 2,
        "平价": 2,
        "转股溢价率": 2,
        "纯债价值": 2,
        "纯债溢价率": 2,
        "YTM": 2,
    })
    for col in ["到期日期", "最后交易日", "最后转股日", "摘牌日期"]:
        lasttrade_info[col] = pd.to_datetime(lasttrade_info[col], format="%Y%m%d", errors="coerce")
    return lasttrade_info


def write_version_tables(path, redemption_bond_info, counting, commitment, other, lasttrade_info):
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        redemption_bond_info.to_excel(writer, sheet_name="公告赎回转债信息", index=True)
        counting.to_excel(writer, sheet_name="赎回累计触发天数", index=True)
        commitment.to_excel(writer, sheet_name="已发布过不提前赎回公告", index=True)
        other.to_excel(writer, sheet_name="从未发布赎回公告", index=True)
        lasttrade_info.to_excel(writer, sheet_name="将到期摘牌转债", index=True)


def contains_chinese(value):
    return any("\u4e00" <= char <= "\u9fff" for char in str(value))


def set_font(cell_value):
    if contains_chinese(cell_value):
        return Font(name="KaiTi_GB2312", size=10, bold=False)
    return Font(name="Times New Roman", size=10, bold=False)


def format_workbook(path):
    workbook = load_workbook(path)
    excel_format_dict = {
        "公告赎回转债信息": ["963634", "0070C0"],
        "赎回累计触发天数": ["963634", "DDD9C4"],
        "已发布过不提前赎回公告": ["963634", "DCE6F1"],
        "从未发布赎回公告": ["963634", "DCE6F1"],
        "将到期摘牌转债": ["963634", "DCE6F1"],
    }
    excel_basicdata_format_dict = {
        "正股收盘价": ["963634", "0070C0"],
        "赎回触发价": ["963634", "DDD9C4"],
        "总表": ["963634", "DCE6F1"],
        "前十大转债持有人": ["963634", "DCE6F1"],
    }

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    alignment = Alignment(horizontal="center", vertical="center")

    for sheet_name, colors in excel_format_dict.items():
        worksheet = workbook[sheet_name]
        font_title = Font(name="KaiTi_GB2312", size=10, bold=False, color="FFFFFF")
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = font_title if cell.row == 1 else set_font(cell.value)
                cell.fill = PatternFill(start_color=colors[0 if cell.row == 1 else 1], end_color=colors[0 if cell.row == 1 else 1], fill_type="solid")
                cell.border = thin_border
                cell.alignment = alignment
        auto_width(worksheet, multiplier=1.5)

    date_time_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}\s00:00:00$")
    for sheet_name in excel_basicdata_format_dict:
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, datetime):
                    cell.value = cell.value.strftime("%Y-%m-%d")
                elif cell.value and date_time_pattern.match(str(cell.value)):
                    cell.value = str(cell.value)[:10]
                cell.font = Font(name="Times New Roman", size=10, bold=False)
        auto_width(worksheet)

        if sheet_name != "总表":
            for stock_row, redeem_row in zip(workbook["正股收盘价"].iter_rows(min_row=2), workbook["赎回触发价"].iter_rows(min_row=2)):
                for stock_cell, redeem_cell in zip(stock_row[1:], redeem_row[1:]):
                    if stock_cell.value and redeem_cell.value and stock_cell.value > redeem_cell.value:
                        fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        stock_cell.fill = fill
                        redeem_cell.fill = fill

    workbook.save(path)


def auto_width(worksheet, multiplier=1):
    for col in worksheet.columns:
        max_length = 0
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        worksheet.column_dimensions[col[0].column_letter].width = (max_length + 2) * multiplier


def plot_table_image(df, output_path, body_color, highlight_commitment=False, last_date=None):
    plt.rcParams["font.sans-serif"] = ["KaiTi_GB2312"]
    fig, ax = plt.subplots(figsize=(8, 0.01))
    ax.axis("tight")
    ax.axis("off")
    table_df = df.reset_index()
    table = ax.table(cellText=table_df.values.tolist(), colLabels=table_df.columns.tolist(), loc="center")
    table.auto_set_font_size(False)
    table.set_fontsize(5)

    if not (len(df.columns) and df.columns[0] == "OUTMESSAGE"):
        for i, width in enumerate(TABLE_COLUMN_WIDTHS):
            if i >= len(table_df.columns):
                break
            for j in range(len(table_df) + 1):
                table._cells[(j, i)].set_width(width)

    today = tomorrow = None
    if highlight_commitment and last_date is not None:
        today = THS_Date_Offset("212001", "dateType:0,period:D,offset:0,dateFormat:0,output:singledate", f"{last_date}").data
        tomorrow = THS_Date_Offset("212001", "dateType:0,period:D,offset:1,dateFormat:0,output:singledate", f"{last_date}").data

    for cell_key, cell in table._cells.items():
        row, col = cell_key
        text = cell.get_text()
        text.set_ha("center")
        if row == 0:
            cell.set_facecolor("#963634")
            text.set_color("white")
        else:
            cell.set_facecolor(body_color)
        cell.set_linewidth(0.1)

    if highlight_commitment and today is not None:
        max_row = max(cell[0] for cell in table._cells.keys())
        max_col = max(cell[1] for cell in table._cells.keys())
        for row in range(1, max_row + 1):
            try:
                cell_date = datetime.strptime(str(table._cells[(row, 5)].get_text().get_text()), "%Y-%m-%d")
            except ValueError:
                continue
            if cell_date in {today, tomorrow}:
                for col in range(max_col + 1):
                    table._cells[(row, col)].set_facecolor("#F2DCDB")

    fig.set_figheight(0.008)
    fig.savefig(output_path, dpi=300, bbox_inches="tight", pad_inches=0, transparent=False)
    plt.close(fig)


def validate_local_image_assets():
    missing = [path.name for path in LOCAL_IMAGE_ASSETS if not path.is_file()]
    if missing:
        raise FileNotFoundError(
            "缺少赎回长图所需的本地图片："
            f"{', '.join(missing)}。请将图片放在脚本同目录。"
        )

    invalid = []
    for path in LOCAL_IMAGE_ASSETS:
        try:
            with Image.open(path) as image:
                image.verify()
        except Exception as exc:
            invalid.append(f"{path.name}（{exc}）")
    if invalid:
        raise RuntimeError(f"本地图片无法读取：{'；'.join(invalid)}")


def create_header_from_local_asset(folder, trade_date):
    with Image.open(LOCAL_HEADER_IMAGE) as source_image:
        img = source_image.copy()

    draw = ImageDraw.Draw(img)
    text = f"      华创固收·周冠南团队\n可转债赎回信息整理（{trade_date}）"
    font = ImageFont.truetype(str(SCRIPT_DIR / "KaiTi_GB2312.ttf"), 60)
    text_bbox = draw.textbbox((0, 0), text, font=font)
    text_width = text_bbox[2] - text_bbox[0]
    text_height = text_bbox[3] - text_bbox[1]
    x = (img.width - text_width) // 2
    y = (img.height - text_height) // 2 - 40
    draw.text((x, y), text, fill="white", font=font)
    img.save(folder / "赎回表题头.png")


def create_footer(folder, version):
    if version == "permille":
        name = "千分位赎回表尾.png"
        footer_text = PERMILLE_FOOTER_TEXT
    elif version == "percent":
        name = "百分位赎回表尾.png"
        footer_text = PERCENT_FOOTER_TEXT
    else:
        raise ValueError(f"未知的表底版本：{version}")

    image = Image.new("RGBA", FOOTER_IMAGE_SIZE, (255, 255, 255, 255))
    draw = ImageDraw.Draw(image)
    draw.rectangle(
        (0, 0, image.width - 1, image.height - 1),
        outline="black",
        width=2,
    )
    font = ImageFont.truetype(str(SCRIPT_DIR / "KaiTi_GB2312.ttf"), 26)
    text_bbox = draw.textbbox((0, 0), footer_text, font=font)
    text_height = text_bbox[3] - text_bbox[1]
    y = (image.height - text_height) // 2 - text_bbox[1]
    draw.text(
        (7, y),
        footer_text,
        fill="black",
        font=font,
        stroke_width=1,
        stroke_fill="black",
    )
    image.save(folder / name)


def combine_version_images(folder, version, redemption_bond_info, output_path):
    suffix = "" if version == "permille" else "（百分位）"
    footer_name = "千分位赎回表尾.png" if version == "permille" else "百分位赎回表尾.png"
    image_paths = [
        folder / f"赎回累计触发天数个券{suffix}.png",
        folder / f"已发布过不提前赎回公告{suffix}.png",
        folder / f"赎回公告个券{suffix}.png",
        folder / "赎回表题头.png",
        folder / footer_name,
    ]
    images = [Image.open(path) for path in image_paths]
    max_width = max(image.width for image in images[:4])
    resized = []
    heights = []
    for image in images:
        height = round(max_width / image.width * image.height)
        heights.append(height)
        resized.append(image.resize((max_width, height), resample=Image.BILINEAR))

    if redemption_bond_info.columns[0] == "OUTMESSAGE":
        order = [3, 1, 0, 4]
    else:
        order = [3, 2, 1, 0, 4]

    total_height = sum(heights[i] for i in order)
    new_image = Image.new("RGBA", (max_width, total_height), (0, 0, 0, 0))
    y = 0
    for idx in order:
        new_image.paste(resized[idx], (0, y))
        y += heights[idx]
    new_image.save(output_path)

    for image in images:
        image.close()


def cleanup_intermediate_images(folder):
    folder = Path(folder).resolve()
    removed = []
    failures = []
    for name in INTERMEDIATE_IMAGE_NAMES:
        image_path = (folder / name).resolve()
        if image_path.parent != folder:
            raise RuntimeError(f"拒绝清理输出文件夹之外的图片：{image_path}")
        if not image_path.is_file():
            continue
        try:
            image_path.unlink()
            removed.append(name)
        except OSError as exc:
            failures.append(f"{name}（{exc}）")

    if failures:
        raise RuntimeError(
            "完整长图已生成，但以下过程图片清理失败："
            f"{'；'.join(failures)}"
        )
    print(f"[图片清理] 已删除 {len(removed)} 张过程图片，仅保留最终完整图。")
    return removed


def run_version(version, output_path, image_output_path, cb_basic_trade, cb_list_trade, stock_close_sheet, trigger_price_sheet, total_table, stock_holder_hold_cb_bond, last_date, folder):
    round_trigger = version == "percent"
    suffix = "" if version == "permille" else "（百分位）"

    redemption_count = calculate_redemption_count(stock_close_sheet, trigger_price_sheet, total_table, round_trigger=round_trigger)
    redemption_bond_info = fetch_redemption_bond_info(stock_close_sheet, total_table, last_date, stock_holder_hold_cb_bond)
    counting, commitment, other = fetch_non_redemption_info(
        cb_basic_trade,
        redemption_bond_info,
        redemption_count,
        total_table,
        last_date,
        stock_holder_hold_cb_bond,
    )
    lasttrade_info = fetch_lasttrade_info(cb_list_trade, redemption_bond_info, last_date)

    write_version_tables(output_path, redemption_bond_info, counting, commitment, other, lasttrade_info)
    format_workbook(output_path)

    plot_table_image(redemption_bond_info, folder / f"赎回公告个券{suffix}.png", "#0070C0")
    plot_table_image(counting, folder / f"赎回累计触发天数个券{suffix}.png", "#DDD9C4")
    plot_table_image(commitment, folder / f"已发布过不提前赎回公告{suffix}.png", "#DCE6F1", highlight_commitment=True, last_date=last_date)
    create_footer(folder, version)
    combine_version_images(folder, version, redemption_bond_info, image_output_path)


def print_runtime(start_time):
    total_time = _time.time() - start_time
    if total_time > 60:
        minutes = int(total_time // 60)
        seconds = int(total_time % 60)
        print(f"程序总运行时长：{int(minutes)} 分 {int(seconds)} 秒")
    else:
        print(f"程序总运行时长：{int(total_time)} 秒")
    print(datetime.now().strftime("%H:%M:%S"), "运行完成")
    print("\U0001F600", "\U0001F600", "\U0001F600")


def main():
    start_time = _time.time()
    validate_local_image_assets()
    login_result = ths_login_demo()
    if login_result != 0:
        raise RuntimeError(f"iFinD登录失败，返回码：{login_result}")
    paths = make_paths()
    last_date = get_last_trade_date()

    cb_basic_trade_full, cb_list_trade = get_cb_basic_trade(last_date)
    cb_basic_trade = cb_basic_trade_full.iloc[:, :3]
    code_list = ",".join(cb_basic_trade.index.astype(str))
    days = get_date_range(last_date)

    stock_close_sheet, trigger_price_sheet, trade_status_sheet, total_table, holder_info, stock_holder_hold_cb_bond = build_base_tables(
        cb_basic_trade,
        code_list,
        days,
        last_date,
    )
    stock_close_sheet, trigger_price_sheet, trade_status_sheet, total_table = clean_suspended_and_commitment(
        stock_close_sheet,
        trigger_price_sheet,
        trade_status_sheet,
        total_table,
    )

    write_base_workbook(paths["permille_xlsx"], stock_close_sheet, trigger_price_sheet, trade_status_sheet, total_table, holder_info)
    write_percent_base_workbook(paths["percent_xlsx"], stock_close_sheet, trigger_price_sheet, total_table, holder_info)
    create_header_from_local_asset(paths["folder"], last_date)

    run_version(
        "permille",
        paths["permille_xlsx"],
        paths["permille_png"],
        cb_basic_trade,
        cb_list_trade,
        stock_close_sheet,
        trigger_price_sheet,
        total_table,
        stock_holder_hold_cb_bond,
        last_date,
        paths["folder"],
    )
    run_version(
        "percent",
        paths["percent_xlsx"],
        paths["percent_png"],
        cb_basic_trade,
        cb_list_trade,
        stock_close_sheet,
        trigger_price_sheet,
        total_table,
        stock_holder_hold_cb_bond,
        last_date,
        paths["folder"],
    )
    cleanup_intermediate_images(paths["folder"])
    if EMAIL_ENABLED:
        recipients = load_recipients_from_workbook(
            EMAIL_RECIPIENT_WORKBOOK,
        )
        print(f"[邮件收件人] 本轮共 {len(recipients)} 个地址。")
        send_redemption_report(
            sender=EMAIL_SENDER,
            recipients=recipients,
            recipient_workbook_path=EMAIL_RECIPIENT_WORKBOOK,
            report_date=paths["yyyymmdd"],
            excel_paths=[paths["permille_xlsx"], paths["percent_xlsx"]],
            permille_image_path=paths["permille_png"],
            percent_image_path=paths["percent_png"],
        )
    print_runtime(start_time)


if __name__ == "__main__":
    main()
