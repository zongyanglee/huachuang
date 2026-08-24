from __future__ import annotations

import argparse
import json
import os
import re
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd


INDEX_SHEET = "指数"
TARGETS = [
    ("沪深300", "000300.SH", 2),
    ("中证500", "000905.SH", 3),
    ("中证1000", "000852.SH", 4),
    ("中证2000", "932000.CSI", 5),
]


def monthly_files(root: Path) -> list[Path]:
    files = []
    for path in root.glob("[0-9][0-9][0-9][0-9]/*.parquet"):
        if re.fullmatch(r"\d{6}", path.stem):
            files.append(path)
    return sorted(files)


def load_source(workbook_path: Path) -> pd.DataFrame:
    formula_book = openpyxl.load_workbook(
        workbook_path, data_only=False, read_only=True, keep_links=False
    )
    formula_sheet = formula_book[formula_book.sheetnames[0]]
    headers = [formula_sheet.cell(1, col).value for col in range(2, 6)]
    expected_headers = [name for name, _, _ in TARGETS]
    if headers != expected_headers:
        raise RuntimeError(f"源表表头不匹配：{headers}，预期：{expected_headers}")
    for name, code, col in TARGETS:
        formula = str(formula_sheet.cell(2, col).value or "")
        if code not in formula:
            raise RuntimeError(f"{name}的公式未包含预期代码{code}：{formula}")
    formula_book.close()

    value_book = openpyxl.load_workbook(
        workbook_path, data_only=True, read_only=True, keep_links=False
    )
    value_sheet = value_book[value_book.sheetnames[0]]
    records = []
    for row in value_sheet.iter_rows(min_row=2, min_col=1, max_col=5, values_only=True):
        raw_date = row[0]
        if raw_date is None:
            continue
        date = pd.Timestamp(raw_date).normalize()
        values = [pd.to_numeric(value, errors="coerce") for value in row[1:]]
        if any(pd.isna(value) for value in values):
            raise RuntimeError(f"源表{date:%Y-%m-%d}存在空值或非数值：{values}")
        records.append([date, *[float(value) for value in values]])
    value_book.close()

    source = pd.DataFrame(records, columns=["date", *expected_headers]).set_index("date")
    if source.index.has_duplicates:
        duplicates = source.index[source.index.duplicated()].strftime("%Y-%m-%d").tolist()
        raise RuntimeError(f"源表日期重复：{duplicates[:10]}")
    return source.sort_index()


def source_value_as_text(value: float) -> str:
    return format(float(value), ".15g")


def audit_file(path: Path, source: pd.DataFrame) -> dict:
    frame = pd.read_parquet(path)
    date_columns = list(frame.columns[2:])
    parsed_dates = [pd.Timestamp(column).normalize() for column in date_columns]
    missing_source_dates = [
        date.strftime("%Y-%m-%d") for date in parsed_dates if date not in source.index
    ]
    if missing_source_dates:
        raise RuntimeError(f"{path.name}有源表未覆盖日期：{missing_source_dates}")

    index_rows = frame.loc[frame["__sheet_name"] == INDEX_SHEET]
    counts = index_rows["__row_id"].value_counts().to_dict()
    duplicate_targets = [name for name, _, _ in TARGETS if counts.get(name, 0) > 1]
    if duplicate_targets:
        raise RuntimeError(f"{path.name}目标指数存在重复行：{duplicate_targets}")

    hs300_mismatches = 0
    hs300_rows = index_rows.loc[index_rows["__row_id"] == "沪深300"]
    if len(hs300_rows) == 1:
        actual = pd.to_numeric(hs300_rows.iloc[0][date_columns], errors="coerce").to_numpy(
            dtype=float
        )
        expected = source.loc[parsed_dates, "沪深300"].to_numpy(dtype=float)
        hs300_mismatches = int((~np.isclose(actual, expected, rtol=0, atol=1e-8)).sum())

    return {
        "file": str(path),
        "date_count": len(date_columns),
        "first_date": parsed_dates[0].strftime("%Y-%m-%d"),
        "last_date": parsed_dates[-1].strftime("%Y-%m-%d"),
        "index_row_count": len(index_rows),
        "existing_targets": {name: int(counts.get(name, 0)) for name, _, _ in TARGETS},
        "hs300_mismatches": hs300_mismatches,
    }


def build_updated_frame(frame: pd.DataFrame, source: pd.DataFrame) -> pd.DataFrame:
    date_columns = list(frame.columns[2:])
    parsed_dates = [pd.Timestamp(column).normalize() for column in date_columns]
    target_names = [name for name, _, _ in TARGETS]

    target_mask = (frame["__sheet_name"] == INDEX_SHEET) & frame["__row_id"].isin(
        target_names
    )
    base = frame.loc[~target_mask].reset_index(drop=True)
    insertion_candidates = base.index[
        (base["__sheet_name"] == INDEX_SHEET) & (base["__row_id"] == "转债指数")
    ].tolist()
    if len(insertion_candidates) != 1:
        raise RuntimeError("无法在指数块中唯一定位转债指数行")
    insertion = insertion_candidates[0]

    rows = []
    for name, _, _ in TARGETS:
        row = {"__sheet_name": INDEX_SHEET, "__row_id": name}
        for column, date in zip(date_columns, parsed_dates):
            row[column] = source_value_as_text(source.at[date, name])
        rows.append(row)
    targets = pd.DataFrame(rows, columns=frame.columns)
    for column in frame.columns:
        targets[column] = targets[column].astype(frame[column].dtype)

    updated = pd.concat(
        [base.iloc[:insertion], targets, base.iloc[insertion:]], ignore_index=True
    )
    return updated


def validate_updated_frame(
    original: pd.DataFrame, updated: pd.DataFrame, source: pd.DataFrame, path: Path
) -> None:
    original_targets = original.loc[
        (original["__sheet_name"] == INDEX_SHEET)
        & original["__row_id"].isin([name for name, _, _ in TARGETS])
    ]
    expected_rows = len(original) + 4 - len(original_targets)
    if len(updated) != expected_rows:
        raise RuntimeError(f"{path.name}更新后总行数异常：{len(updated)} != {expected_rows}")
    if list(updated.columns) != list(original.columns):
        raise RuntimeError(f"{path.name}更新后列结构变化")

    index_rows = updated.loc[updated["__sheet_name"] == INDEX_SHEET]
    expected_order = [
        "万得全A",
        "十年国债",
        "正股等权指数",
        "沪深300",
        "中证500",
        "中证1000",
        "中证2000",
        "转债指数",
        "中证800",
        "中债新综合财富总指数",
    ]
    if index_rows["__row_id"].tolist() != expected_order:
        raise RuntimeError(f"{path.name}指数行顺序异常：{index_rows['__row_id'].tolist()}")

    date_columns = list(updated.columns[2:])
    parsed_dates = [pd.Timestamp(column).normalize() for column in date_columns]
    for name, _, _ in TARGETS:
        row = index_rows.loc[index_rows["__row_id"] == name].iloc[0]
        actual = pd.to_numeric(row[date_columns], errors="coerce").to_numpy(dtype=float)
        expected = source.loc[parsed_dates, name].to_numpy(dtype=float)
        if not np.allclose(actual, expected, rtol=0, atol=1e-8):
            count = int((~np.isclose(actual, expected, rtol=0, atol=1e-8)).sum())
            raise RuntimeError(f"{path.name}的{name}有{count}个值未通过校验")


def write_monthly_file(path: Path, source: pd.DataFrame) -> None:
    original = pd.read_parquet(path)
    updated = build_updated_frame(original, source)
    validate_updated_frame(original, updated, source, path)

    temp_path = path.with_name(path.name + ".index_tmp")
    try:
        updated.to_parquet(temp_path, index=False, engine="pyarrow", compression="snappy")
        reread = pd.read_parquet(temp_path)
        validate_updated_frame(original, reread, source, path)
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def update_manifest(root: Path) -> None:
    path = root / "_meta" / "sheet_manifest.parquet"
    manifest = pd.read_parquet(path)
    mask = manifest["sheet_name"] == INDEX_SHEET
    if int(mask.sum()) != 1:
        raise RuntimeError("元数据清单中无法唯一定位指数行")
    manifest.loc[mask, "rows"] = 10
    temp_path = path.with_name(path.name + ".index_tmp")
    try:
        manifest.to_parquet(temp_path, index=False, engine="pyarrow", compression="snappy")
        reread = pd.read_parquet(temp_path)
        if int(reread.loc[reread["sheet_name"] == INDEX_SHEET, "rows"].iloc[0]) != 10:
            raise RuntimeError("元数据清单更新校验失败")
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def final_audit(files: list[Path], source: pd.DataFrame, root: Path) -> dict:
    audits = [audit_file(path, source) for path in files]
    manifest = pd.read_parquet(root / "_meta" / "sheet_manifest.parquet")
    manifest_rows = int(
        manifest.loc[manifest["sheet_name"] == INDEX_SHEET, "rows"].iloc[0]
    )
    return {
        "monthly_files": len(files),
        "parquet_dates": sum(item["date_count"] for item in audits),
        "first_date": audits[0]["first_date"],
        "last_date": audits[-1]["last_date"],
        "source_first_date": source.index.min().strftime("%Y-%m-%d"),
        "source_last_date": source.index.max().strftime("%Y-%m-%d"),
        "index_rows_per_file": sorted({item["index_row_count"] for item in audits}),
        "existing_target_file_counts": {
            name: sum(item["existing_targets"][name] == 1 for item in audits)
            for name, _, _ in TARGETS
        },
        "hs300_total_mismatches": sum(item["hs300_mismatches"] for item in audits),
        "manifest_index_rows": manifest_rows,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="将补充指数Excel写入月度parquet的指数块")
    parser.add_argument("excel", type=Path)
    parser.add_argument("parquet_root", type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    excel = args.excel.resolve()
    root = args.parquet_root.resolve()
    source = load_source(excel)
    files = monthly_files(root)
    if len(files) != 140:
        raise RuntimeError(f"月度parquet数量异常：{len(files)}，预期140")

    before = final_audit(files, source, root)
    print("更新前审计：")
    print(json.dumps(before, ensure_ascii=False, indent=2))

    if not args.apply:
        print("仅审计，未写入。添加 --apply 后执行更新。")
        return

    for number, path in enumerate(files, start=1):
        write_monthly_file(path, source)
        if number % 20 == 0 or number == len(files):
            print(f"已更新 {number}/{len(files)} 个文件")
    update_manifest(root)

    after = final_audit(files, source, root)
    print("更新后审计：")
    print(json.dumps(after, ensure_ascii=False, indent=2))
    if after["index_rows_per_file"] != [10]:
        raise RuntimeError("并非所有月度parquet都包含10个指数行")
    if any(count != len(files) for count in after["existing_target_file_counts"].values()):
        raise RuntimeError("并非所有目标指数都已覆盖全部月度parquet")
    if after["hs300_total_mismatches"] != 0 or after["manifest_index_rows"] != 10:
        raise RuntimeError("最终审计未通过")


if __name__ == "__main__":
    main()
