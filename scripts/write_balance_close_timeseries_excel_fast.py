from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs" / "balance_close_timeseries"
SUMMARY_PATH = OUTPUT_DIR / "summary.json"
LONG_CSV = OUTPUT_DIR / "余额与收盘价_成交额大于零清洗_长表.csv"
XLSX_PATH = OUTPUT_DIR / "余额与收盘价时间序列_成交额大于零清洗.xlsx"


def main() -> None:
    summary = json.loads(SUMMARY_PATH.read_text(encoding="utf-8"))
    df = pd.read_csv(LONG_CSV)
    df["日期"] = pd.to_datetime(df["日期"])

    note = pd.DataFrame(
        [
            ("数据来源", summary["source"]),
            ("源文件数量", summary["source_file_count"]),
            ("日期范围", f'{summary["date_start"]} 至 {summary["date_end"]}'),
            ("交易日数量", summary["date_count"]),
            ("转债代码数量", summary["bond_count"]),
            ("清洗规则", "仅保留同一代码、同一日期成交额>0的余额和收盘价；其余置为空。"),
            ("Excel数据行数", len(df)),
            ("成交额>0单元格数", summary["positive_amount_cells"]),
            ("候选单元格数", summary["total_candidate_cells"]),
            ("宽表CSV", "同目录保留余额、收盘价两份宽表CSV。"),
        ],
        columns=["项目", "说明"],
    )

    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        df.to_excel(writer, sheet_name="时间序列", index=False)
        note.to_excel(writer, sheet_name="说明", index=False)

    wb = load_workbook(XLSX_PATH)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    label_fill = PatternFill("solid", fgColor="D9EAF7")
    label_font = Font(color="1F4E78", bold=True)

    ws = wb["时间序列"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        row[0].number_format = "yyyy-mm-dd"
    for col in ["C", "D"]:
        for cell in ws[col][1:]:
            cell.number_format = "0.0000"

    note_ws = wb["说明"]
    for cell in note_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for cell in note_ws["A"][1:]:
        cell.fill = label_fill
        cell.font = label_font
    note_ws.column_dimensions["A"].width = 18
    note_ws.column_dimensions["B"].width = 90

    wb.save(XLSX_PATH)
    print(XLSX_PATH)
    print(f"rows={len(df)}")


if __name__ == "__main__":
    main()
