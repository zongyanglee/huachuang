from __future__ import annotations

import json
from datetime import timedelta
from pathlib import Path

import matplotlib.dates as mdates
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.ticker import PercentFormatter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data/转债个券历史序列"
MASTER_PATH = DATA_DIR / "_special" / "总表.parquet"
FONT_PATH = ROOT / "assets/fonts/KaiTi_GB2312.ttf"

START_2015 = pd.Timestamp("2015-01-01")
START_2017 = pd.Timestamp("2017-01-01")
START_2018 = pd.Timestamp("2018-01-01")
BASE_LEVEL = 100.0

PRICE_COLOR = "#E6121B"
BLUE_COLOR = "#0262BA"
COLOR_SEQUENCE = ["#E6121B", "#0262BA", "#A6A6A6", "#E6B9B8", "#B7DEE8", "#F79646"]

GROUPS: list[tuple[str, float | None, float | None]] = [
    ("70以下", None, 70.0),
    ("70-90", 70.0, 90.0),
    ("90-110", 90.0, 110.0),
    ("110-130", 110.0, 130.0),
    ("130-150", 130.0, 150.0),
    ("150以上", 150.0, None),
]
PLOT_GROUPS_AVG = ["70-90", "90-110", "110-130"]
GROUP_COLORS = {name: color for (name, _, _), color in zip(GROUPS, COLOR_SEQUENCE)}
PLOT_GROUP_COLORS = {name: color for name, color in zip(PLOT_GROUPS_AVG, COLOR_SEQUENCE)}


def parse_last_trade(series: pd.Series) -> pd.Series:
    cleaned = series.astype("string").str.strip()
    cleaned = cleaned.mask(cleaned.isin(["", "0", "0.0", "None", "nan", "<NA>"]))
    return pd.to_datetime(cleaned, errors="coerce")


def find_date_columns(df: pd.DataFrame, start_date: pd.Timestamp) -> list[str]:
    cols: list[str] = []
    for col in df.columns:
        if str(col).startswith("__"):
            continue
        dt = pd.to_datetime(col, errors="coerce")
        if pd.notna(dt) and dt >= start_date:
            cols.append(col)
    return cols


def monthly_parquet_files(start_year: int) -> list[Path]:
    return sorted(
        p
        for year_dir in DATA_DIR.iterdir()
        if year_dir.is_dir() and year_dir.name.isdigit() and int(year_dir.name) >= start_year
        for p in year_dir.glob("*.parquet")
    )


def load_sheet(path: Path, sheet_name: str, start_date: pd.Timestamp) -> pd.DataFrame:
    raw = pd.read_parquet(path)
    sheet = raw[raw["__sheet_name"].eq(sheet_name)].copy()
    if sheet.empty:
        raise ValueError(f"{path} 未找到 sheet: {sheet_name}")
    sheet = sheet.set_index("__row_id")
    cols = find_date_columns(sheet, start_date)
    return sheet[cols].apply(pd.to_numeric, errors="coerce")


def concat_sheet(sheet_name: str, start_date: pd.Timestamp) -> pd.DataFrame:
    frames = [load_sheet(path, sheet_name, start_date) for path in monthly_parquet_files(start_date.year)]
    out = pd.concat(frames, axis=1)
    out = out.loc[:, ~out.columns.duplicated()]
    ordered_cols = sorted(out.columns, key=lambda col: pd.to_datetime(col))
    return out[ordered_cols]


def apply_listing_mask(
    values: pd.DataFrame,
    listing: pd.Series,
    last_trade: pd.Series,
) -> pd.DataFrame:
    out = values.copy()
    trade_dates = pd.to_datetime(out.columns)
    row_listing = listing.reindex(out.index)
    row_last = last_trade.reindex(out.index)
    for col, dt in zip(out.columns, trade_dates):
        invalid = dt < row_listing
        has_last = row_last.notna()
        invalid = invalid | (has_last & (dt > row_last))
        out.loc[invalid.fillna(False), col] = pd.NA
    return out


def prepare_core_data() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.Series, pd.Series]:
    master = pd.read_parquet(MASTER_PATH).set_index("__row_id")
    listing = pd.to_datetime(master["上市日期"], errors="coerce")
    last_trade = parse_last_trade(master["最后交易日"])

    close = concat_sheet("收盘价", START_2015)
    premium = concat_sheet("转股溢价率", START_2015) / 100
    parity = concat_sheet("平价", START_2015)
    balance = concat_sheet("余额", START_2015)

    common_index = close.index.intersection(premium.index).intersection(parity.index).intersection(balance.index)
    common_cols = close.columns.intersection(premium.columns).intersection(parity.columns).intersection(balance.columns)
    close = close.loc[common_index, common_cols]
    premium = premium.loc[common_index, common_cols]
    parity = parity.loc[common_index, common_cols]
    balance = balance.loc[common_index, common_cols]

    close = apply_listing_mask(close, listing, last_trade)
    premium = apply_listing_mask(premium, listing, last_trade)
    parity = apply_listing_mask(parity, listing, last_trade)
    balance = apply_listing_mask(balance, listing, last_trade)
    return close, premium, parity, balance, listing, last_trade


def build_market_stats(close: pd.DataFrame, premium: pd.DataFrame) -> pd.DataFrame:
    stats = pd.DataFrame(
        {
            "交易日期": pd.to_datetime(close.columns),
            "平均价格": close.mean(axis=0, skipna=True).to_numpy(),
            "价格中位数": close.median(axis=0, skipna=True).to_numpy(),
            "平均转股溢价率": premium.mean(axis=0, skipna=True).to_numpy(),
            "价格样本数": close.notna().sum(axis=0).to_numpy(),
            "溢价率样本数": premium.notna().sum(axis=0).to_numpy(),
        }
    )
    return stats[stats["交易日期"] >= START_2018].reset_index(drop=True)


def build_group_average(parity: pd.DataFrame, premium: pd.DataFrame, balance: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for col, dt in zip(parity.columns, pd.to_datetime(parity.columns)):
        valid_balance = balance[col] > 3
        row: dict[str, object] = {"交易日期": dt}
        for name, low, high in GROUPS:
            cond = valid_balance & premium[col].notna() & parity[col].notna()
            if low is None:
                cond = cond & (parity[col] <= high)
            elif high is None:
                cond = cond & (parity[col] > low)
            else:
                cond = cond & (parity[col] > low) & (parity[col] <= high)
            row[name] = premium.loc[cond, col].mean()
            row[f"{name}_样本数"] = int(cond.sum())
        rows.append(row)
    out = pd.DataFrame(rows).drop_duplicates("交易日期").sort_values("交易日期").reset_index(drop=True)
    for name, _, _ in GROUPS:
        out[name] = pd.to_numeric(out[name], errors="coerce")
        out[f"{name}_样本数"] = pd.to_numeric(out[f"{name}_样本数"], errors="coerce").fillna(0).astype(int)
    return out


def percentile_rank(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").rank(method="average", pct=True)


def build_group_percentiles(group_avg: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    pct = group_avg[["交易日期"]].copy()
    smooth = group_avg[["交易日期"]].copy()
    for name, _, _ in GROUPS:
        pct[name] = percentile_rank(group_avg[name])
        smooth[name] = pct[name].rolling(window=5, min_periods=1).mean()
    return pct, smooth


def build_balance_weighted_index(
    close: pd.DataFrame,
    balance: pd.DataFrame,
    listing: pd.Series,
) -> pd.DataFrame:
    returns = close / close.shift(1, axis=1) - 1
    level = BASE_LEVEL
    records: list[dict[str, object]] = []
    prev_weights = pd.Series(dtype="float64")
    cols = list(close.columns)
    dates = pd.to_datetime(cols)

    for idx, (col, dt) in enumerate(zip(cols, dates)):
        if idx == 0:
            bal = pd.to_numeric(balance[col], errors="coerce")
            valid = bal.notna() & (bal > 0) & close[col].notna()
            prev_weights = bal.loc[valid] / bal.loc[valid].sum()
            records.append(
                {
                    "交易日期": dt,
                    "日涨跌幅": 0.0,
                    "指数点位": level,
                    "样本数": int(valid.sum()),
                    "权重和": 1.0,
                    "剔除上市首日新券数": 0,
                }
            )
            continue

        prev_col = cols[idx - 1]
        ret = pd.to_numeric(returns[col], errors="coerce")
        prev_bal = pd.to_numeric(balance[prev_col], errors="coerce")
        prev_close = pd.to_numeric(close[prev_col], errors="coerce")
        cur_close = pd.to_numeric(close[col], errors="coerce")
        listed_today = listing.reindex(close.index).eq(dt)
        eligible = (
            prev_bal.notna()
            & (prev_bal > 0)
            & prev_close.notna()
            & cur_close.notna()
            & ret.notna()
            & (~listed_today.fillna(False))
        )

        weight_base = prev_bal.loc[eligible]
        if weight_base.sum() > 0:
            weights = weight_base / weight_base.sum()
            daily_return = float((ret.loc[eligible] * weights).sum())
            level *= 1 + daily_return
            prev_weights = weights
        else:
            weights = prev_weights.iloc[0:0]
            daily_return = 0.0

        records.append(
            {
                "交易日期": dt,
                "日涨跌幅": daily_return,
                "指数点位": level,
                "样本数": int(eligible.sum()),
                "权重和": float(weights.sum()) if len(weights) else 0.0,
                "剔除上市首日新券数": int(listed_today.fillna(False).sum()),
            }
        )

    return pd.DataFrame(records).reset_index(drop=True)


def setup_font() -> fm.FontProperties:
    if not FONT_PATH.exists():
        raise FileNotFoundError(f"未找到字体文件：{FONT_PATH}")
    font_prop = fm.FontProperties(fname=str(FONT_PATH))
    fm.fontManager.addfont(str(FONT_PATH))
    plt.rcParams["font.family"] = font_prop.get_name()
    plt.rcParams["axes.unicode_minus"] = False
    return font_prop


def style_axes(ax, font_prop: fm.FontProperties) -> None:
    ax.tick_params(axis="both", colors="black")
    for spine in ax.spines.values():
        spine.set_color("black")
    for label in ax.get_xticklabels() + ax.get_yticklabels():
        label.set_fontproperties(font_prop)
        label.set_fontsize(14)


def style_time_axis(ax, x: pd.Series) -> None:
    ax.xaxis.set_major_locator(mdates.YearLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    ax.xaxis.set_minor_locator(mdates.MonthLocator(bymonth=(1, 7)))
    ax.grid(True, which="major", axis="both", color="#D9D9D9", linewidth=0.8, alpha=0.75)
    ax.grid(True, which="minor", axis="x", color="#ECECEC", linewidth=0.5, alpha=0.6)
    ax.set_xlim(pd.to_datetime(x).min(), pd.to_datetime(x).max())


def style_legend(legend, font_prop: fm.FontProperties) -> None:
    for text in legend.get_texts():
        text.set_fontproperties(font_prop)
        text.set_fontsize(15)


def save_market_price_premium_chart(market: pd.DataFrame, output_dir: Path, font_prop: fm.FontProperties) -> Path:
    path = output_dir / "转债平均价格与平均转股溢价率.png"
    x = market["交易日期"]
    fig, ax_price = plt.subplots(figsize=(14, 7), dpi=180)
    ax_premium = ax_price.twinx()
    line_price, = ax_price.plot(x, market["平均价格"], color=PRICE_COLOR, linewidth=2.2, label="平均价格")
    line_premium, = ax_premium.plot(x, market["平均转股溢价率"], color=BLUE_COLOR, linewidth=2.2, label="平均转股溢价率")
    ax_price.set_ylabel("平均价格", fontproperties=font_prop, fontsize=16, color="black")
    ax_premium.set_ylabel("平均转股溢价率", fontproperties=font_prop, fontsize=16, color="black")
    ax_premium.yaxis.set_major_formatter(PercentFormatter(xmax=1, decimals=0))
    style_axes(ax_price, font_prop)
    style_axes(ax_premium, font_prop)
    style_time_axis(ax_price, x)
    legend = ax_price.legend([line_price, line_premium], ["平均价格", "平均转股溢价率"], loc="upper left", frameon=False, prop=font_prop, fontsize=15)
    style_legend(legend, font_prop)
    fig.tight_layout()
    fig.savefig(path, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return path


def save_market_price_median_chart(market: pd.DataFrame, output_dir: Path, font_prop: fm.FontProperties) -> Path:
    path = output_dir / "转债平均价格与价格中位数.png"
    x = market["交易日期"]
    fig, ax = plt.subplots(figsize=(14, 7), dpi=180)
    ax.plot(x, market["平均价格"], color=PRICE_COLOR, linewidth=2.2, label="平均价格")
    ax.plot(x, market["价格中位数"], color=BLUE_COLOR, linewidth=2.2, label="价格中位数")
    ax.set_ylabel("价格", fontproperties=font_prop, fontsize=16, color="black")
    style_axes(ax, font_prop)
    style_time_axis(ax, x)
    legend = ax.legend(loc="upper left", frameon=False, prop=font_prop, fontsize=15)
    style_legend(legend, font_prop)
    fig.tight_layout()
    fig.savefig(path, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return path


def save_group_avg_chart(group_avg: pd.DataFrame, output_dir: Path, font_prop: fm.FontProperties) -> Path:
    path = output_dir / "平价分组平均转股溢价率.png"
    plot_df = group_avg[group_avg["交易日期"] >= START_2017].copy()
    x = plot_df["交易日期"]
    fig, ax = plt.subplots(figsize=(14, 7), dpi=180)
    for name in PLOT_GROUPS_AVG:
        ax.plot(x, plot_df[name], linewidth=2.0, label=name, color=PLOT_GROUP_COLORS[name])
    ax.set_ylabel("平均转股溢价率", fontproperties=font_prop, fontsize=16, color="black")
    ax.yaxis.set_major_formatter(PercentFormatter(xmax=1, decimals=0))
    style_axes(ax, font_prop)
    style_time_axis(ax, x)
    legend = ax.legend(loc="upper left", frameon=False, prop=font_prop, fontsize=15, ncol=2)
    style_legend(legend, font_prop)
    fig.tight_layout()
    fig.savefig(path, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return path


def save_group_percentile_chart(smooth_pct: pd.DataFrame, output_dir: Path, font_prop: fm.FontProperties) -> Path:
    path = output_dir / "平价分组平均转股溢价率分位数.png"
    x = smooth_pct["交易日期"]
    fig, ax = plt.subplots(figsize=(14, 7), dpi=180)
    for name, _, _ in GROUPS:
        ax.plot(x, smooth_pct[name], linewidth=2.0, label=name, color=GROUP_COLORS[name])
    ax.set_ylabel("历史分位数", fontproperties=font_prop, fontsize=16, color="black")
    ax.yaxis.set_major_formatter(PercentFormatter(xmax=1, decimals=0))
    ax.set_ylim(0, 1)
    style_axes(ax, font_prop)
    style_time_axis(ax, x)
    legend = ax.legend(loc="upper left", frameon=False, prop=font_prop, fontsize=15, ncol=2)
    style_legend(legend, font_prop)
    fig.tight_layout()
    fig.savefig(path, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return path


def save_index_chart(index_df: pd.DataFrame, output_dir: Path, font_prop: fm.FontProperties) -> Path:
    path = output_dir / "转债余额加权指数.png"
    x = index_df["交易日期"]
    fig, ax = plt.subplots(figsize=(14, 7), dpi=180)
    ax.plot(x, index_df["指数点位"], color=PRICE_COLOR, linewidth=2.2, label="余额加权指数")
    ax.set_ylabel("指数点位", fontproperties=font_prop, fontsize=16, color="black")
    style_axes(ax, font_prop)
    style_time_axis(ax, x)
    legend = ax.legend(loc="upper left", frameon=False, prop=font_prop, fontsize=15)
    style_legend(legend, font_prop)
    fig.tight_layout()
    fig.savefig(path, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return path


def previous_friday(date: pd.Timestamp) -> pd.Timestamp:
    days_since_friday = (date.weekday() - 4) % 7
    if days_since_friday == 0:
        days_since_friday = 7
    return date - timedelta(days=int(days_since_friday))


def choose_base_date(dates: pd.Series, latest_date: pd.Timestamp) -> pd.Timestamp:
    target = previous_friday(latest_date)
    candidates = dates[dates <= target]
    if candidates.empty:
        raise ValueError("没有找到可用的上周五或节前最近交易日")
    return candidates.max()


def fmt_pct(value: float, digits: int = 2) -> str:
    return f"{value * 100:.{digits}f}%"


def fmt_pct_change(value: float) -> str:
    sign = "+" if value >= 0 else ""
    return f"{sign}{value * 100:.1f}"


def build_commentary(group_avg: pd.DataFrame, group_pct: pd.DataFrame) -> tuple[str, pd.DataFrame]:
    latest_date = group_avg["交易日期"].max()
    base_date = choose_base_date(group_avg["交易日期"], latest_date)
    latest_row = group_avg.loc[group_avg["交易日期"].eq(latest_date)].iloc[0]
    base_row = group_avg.loc[group_avg["交易日期"].eq(base_date)].iloc[0]
    pct_row = group_pct.loc[group_pct["交易日期"].eq(latest_date)].iloc[0]

    names = [name for name, _, _ in GROUPS]
    latest_values = [fmt_pct(float(latest_row[name]), 2) for name in names]
    changes = [fmt_pct_change(float(latest_row[name]) - float(base_row[name])) for name in names]
    percentiles = [fmt_pct(float(pct_row[name]), 1) for name in names]
    group_text = "、".join(["70以下", "70-90", "90-110", "110-130", "130-150", "150元以上"])
    period = f"{base_date:%Y%m%d}-{latest_date:%Y%m%d}"
    text = (
        f"{period}，最新平价{group_text}的平均转股溢价率分别为"
        f"{'、'.join(latest_values)}。"
        f"本周不同平价区间的转债溢价率走势不一，环比分别"
        f"{'、'.join(changes)} pct。"
        f"最新平价{group_text}的平均转股溢价率所处2015年以来历史分位数分别"
        f"{'、'.join(percentiles)}。"
    )
    detail = pd.DataFrame(
        {
            "项目": ["对比区间", "文字描述"],
            "内容": [period, text],
        }
    )
    return text, detail


def write_excel(
    output_path: Path,
    market: pd.DataFrame,
    group_avg: pd.DataFrame,
    group_pct: pd.DataFrame,
    group_pct_smooth: pd.DataFrame,
    index_df: pd.DataFrame,
    commentary_df: pd.DataFrame,
    audit_df: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        market.to_excel(writer, sheet_name="市场统计2018以来", index=False)
        group_avg.to_excel(writer, sheet_name="平价分组均值2015以来", index=False)
        group_pct.to_excel(writer, sheet_name="平价分组分位数", index=False)
        group_pct_smooth.to_excel(writer, sheet_name="平价分组分位数5日平滑", index=False)
        index_df.to_excel(writer, sheet_name="余额加权指数", index=False)
        commentary_df.to_excel(writer, sheet_name="文字描述", index=False)
        audit_df.to_excel(writer, sheet_name="口径说明", index=False)

        wb = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            for col_idx, column_cells in enumerate(ws.columns, 1):
                header = ws.cell(row=1, column=col_idx).value
                width = min(max(len(str(header or "")) + 4, 12), 28)
                ws.column_dimensions[get_column_letter(col_idx)].width = width
                if header in {"交易日期"}:
                    for cell in column_cells[1:]:
                        cell.number_format = "yyyy-mm-dd"
                if header and any(key in str(header) for key in ["溢价率", "分位数", "日涨跌幅", "权重和"]):
                    for cell in column_cells[1:]:
                        cell.number_format = "0.00%"
                if header and any(key in str(header) for key in ["平均价格", "价格中位数", "指数点位"]):
                    for cell in column_cells[1:]:
                        cell.number_format = "0.00"
                if header and "样本数" in str(header):
                    for cell in column_cells[1:]:
                        cell.number_format = "#,##0"
            if ws.title in {"文字描述", "口径说明"}:
                ws.column_dimensions["A"].width = 24
                ws.column_dimensions["B"].width = 110
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")


def main() -> None:
    close, premium, parity, balance, listing, _ = prepare_core_data()
    market = build_market_stats(close, premium)
    group_avg = build_group_average(parity, premium, balance)
    group_pct, group_pct_smooth = build_group_percentiles(group_avg)
    index_df = build_balance_weighted_index(close, balance, listing)

    latest_date = pd.to_datetime(close.columns).max()
    date_tag = latest_date.strftime("%Y%m%d")
    output_dir = ROOT / "runs" / "weekly" / f"鹏华周报{date_tag}"
    output_dir.mkdir(parents=True, exist_ok=True)

    font_prop = setup_font()
    chart_paths = [
        save_market_price_premium_chart(market, output_dir, font_prop),
        save_market_price_median_chart(market, output_dir, font_prop),
        save_group_avg_chart(group_avg, output_dir, font_prop),
        save_group_percentile_chart(group_pct_smooth, output_dir, font_prop),
        save_index_chart(index_df, output_dir, font_prop),
    ]

    commentary_text, commentary_df = build_commentary(group_avg, group_pct)
    text_path = output_dir / "平价分组平均转股溢价率文字描述.txt"
    text_path.write_text(commentary_text, encoding="utf-8")

    audit_rows = [
        ("数据源", "转债个券历史序列 parquet；总表提供上市日期与最后交易日"),
        ("输出目录", str(output_dir)),
        ("最新日期", latest_date.strftime("%Y-%m-%d")),
        ("图表字体", str(FONT_PATH.name)),
        ("图表颜色序列", "、".join(COLOR_SEQUENCE)),
        ("市场统计口径", "2018年以来；清洗上市日前和最后交易日后的样本；平均转股溢价率为算术均值"),
        ("平价分组口径", "<=70、(70,90]、(90,110]、(110,130]、(130,150]、>150；仅余额>3转债；转股溢价率算术均值"),
        ("平价分组均值图", "展示70-90、90-110、110-130三组，2017年以来"),
        ("分位数图", "六个平价分组；分位数按2015年以来各组自身历史序列计算；折线使用5日滚动均值平滑"),
        ("余额加权指数", "基日=100；使用收盘价计算个券日收益；按上一交易日余额权重加权；新券上市首日剔除，不计收益和权重"),
    ]
    audit_df = pd.DataFrame(audit_rows, columns=["项目", "说明"])
    excel_path = output_dir / "鹏华周报图表底稿.xlsx"
    write_excel(excel_path, market, group_avg, group_pct, group_pct_smooth, index_df, commentary_df, audit_df)

    summary = {
        "latest_date": latest_date.strftime("%Y-%m-%d"),
        "output_dir": str(output_dir),
        "charts": [str(path) for path in chart_paths],
        "excel": str(excel_path),
        "commentary": str(text_path),
        "market_rows": int(len(market)),
        "group_rows": int(len(group_avg)),
        "index_rows": int(len(index_df)),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
