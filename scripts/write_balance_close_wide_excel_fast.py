from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs" / "balance_close_timeseries"
SUMMARY_PATH = OUTPUT_DIR / "summary.json"
BALANCE_CSV = OUTPUT_DIR / "余额_成交额大于零清洗.csv"
CLOSE_CSV = OUTPUT_DIR / "收盘价_成交额大于零清洗.csv"
XLSX_PATH = OUTPUT_DIR / "余额与收盘价时间序列_宽表_成交额大于零清洗.xlsx"


def add_note_sheet(writer: pd.ExcelWriter, summary: dict) -> None:
    note = pd.DataFrame(
        [
            ("数据来源", r"D:\JupyterFiles\huachuang\转债个券历史序列"),
            ("源文件数量", summary["source_file_count"]),
            ("日期范围", f'{summary["date_start"]} 至 {summary["date_end"]}'),
            ("交易日数量", summary["date_count"]),
            ("转债代码数量", summary["bond_count"]),
            ("清洗规则", "仅保留同一代码、同一日期成交额>0的余额和收盘价；其余置为空。"),
            ("宽表结构", "行=日期，列=转债代码；分别输出“余额”和“收盘价”两张表。"),
            ("余额有效单元格数", summary["balance_non_null_cells"]),
            ("收盘价有效单元格数", summary["close_non_null_cells"]),
        ],
        columns=["项目", "说明"],
    )
    note.to_excel(writer, sheet_name="说明", index=False)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    label_fill = PatternFill("solid", fgColor="D9EAF7")
    label_font = Font(color="1F4E78", bold=True)

    for sheet_name in ["余额", "收盘价"]:
        ws = wb[sheet_name]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws.column_dimensions["A"].width = 13
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            row[0].number_format = "yyyy-mm-dd"
        for col_idx in range(2, min(ws.max_column, 30) + 1):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 12

    note_ws = wb["说明"]
    for cell in note_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for cell in note_ws["A"][1:]:
        cell.fill = label_fill
        cell.font = label_font
    note_ws.column_dimensions["A"].width = 18
    note_ws.column_dimensions["B"].width = 90

    wb.save(path)


def main() -> None:
    summary = json.loads(SUMMARY_PATH.read_text(encoding="utf-8"))
    balance = pd.read_csv(BALANCE_CSV)
    close = pd.read_csv(CLOSE_CSV)
    balance["日期"] = pd.to_datetime(balance["日期"])
    close["日期"] = pd.to_datetime(close["日期"])

    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        balance.to_excel(writer, sheet_name="余额", index=False)
        close.to_excel(writer, sheet_name="收盘价", index=False)
        add_note_sheet(writer, summary)

    style_workbook(XLSX_PATH)
    print(XLSX_PATH)
    print(f"balance_shape={balance.shape} close_shape={close.shape}")


if __name__ == "__main__":
    main()
