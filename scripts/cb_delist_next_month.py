#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从转债 parquet 取完整 Wind 代码，并汇总未来一个月内摘牌转债。"""

from __future__ import annotations

import argparse
import json
import time
from pathlib import Path

import pandas as pd
from WindPy import w
from openpyxl import Workbook


def load_bond_universe(parquet_root: Path) -> pd.DataFrame:
    path = parquet_root / "_special" / "总表.parquet"
    frame = pd.read_parquet(path)
    required = {"__row_id", "转债名称", "赎回公告日", "上市日期", "最后交易日"}
    missing = required.difference(frame.columns)
    if missing:
        raise ValueError(f"总表缺少字段：{sorted(missing)}")
    out = frame[
        ["__row_id", "转债名称", "赎回公告日", "上市日期", "最后交易日"]
    ].copy()
    out.columns = [
        "wind_code",
        "bond_name",
        "redemption_notice_date",
        "listing_date",
        "last_trade_date",
    ]
    out["wind_code"] = out["wind_code"].astype(str).str.strip()
    for col in ["redemption_notice_date", "listing_date", "last_trade_date"]:
        out[col] = pd.to_datetime(out[col], errors="coerce")
    out = out.drop_duplicates("wind_code").reset_index(drop=True)
    return out


def fetch_delist_dates_python(codes: list[str], batch_size: int = 300) -> pd.Series:
    started = w.start(waitTime=60)
    if getattr(started, "ErrorCode", -1) != 0:
        raise RuntimeError(f"Wind 启动失败：{started}")

    parts: list[pd.Series] = []
    for start in range(0, len(codes), batch_size):
        batch = codes[start : start + batch_size]
        result = w.wss(batch, "s_info_delistdate")
        if result.ErrorCode != 0:
            raise RuntimeError(
                f"Wind s_info_delistdate 查询失败，批次 {start}-{start + len(batch) - 1}，"
                f"错误码 {result.ErrorCode}"
            )
        values = result.Data[0] if result.Data else [None] * len(result.Codes)
        parts.append(pd.Series(values, index=[str(code) for code in result.Codes]))
    w.stop()
    raw = pd.concat(parts) if parts else pd.Series(dtype=object)
    raw.index.name = "wind_code"
    return pd.to_datetime(raw, errors="coerce")


def _excel_serial_to_timestamp(value: object) -> pd.Timestamp | pd.NaT:
    if value is None:
        return pd.NaT
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return pd.Timestamp("1899-12-30") + pd.to_timedelta(float(value), unit="D")
    return pd.to_datetime(value, errors="coerce")


def fetch_delist_dates_excel(
    codes: list[str],
    work_dir: Path,
    timeout_seconds: int = 300,
) -> pd.Series:
    import win32com.client

    work_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = (work_dir / "wind_s_info_delistdate.xlsx").resolve()
    book = Workbook()
    sheet = book.active
    sheet.title = "delist"
    sheet.append(["wind_code", "s_info_delistdate"])
    for row_no, code in enumerate(codes, start=2):
        sheet.cell(row=row_no, column=1, value=code)
        sheet.cell(row=row_no, column=2, value=f'=WSS(A{row_no},"s_info_delistdate")')
    book.save(workbook_path)

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    com_book = None
    try:
        com_book = excel.Workbooks.Open(str(workbook_path))
        excel.CalculateFullRebuild()
        try:
            excel.CalculateUntilAsyncQueriesDone()
        except Exception:
            pass

        com_sheet = com_book.Worksheets(1)
        deadline = time.monotonic() + timeout_seconds
        values: tuple[tuple[object], ...] | None = None
        valid_count = 0
        stable_polls = 0
        previous_valid_count = -1
        while time.monotonic() < deadline:
            time.sleep(2)
            values = com_sheet.Range(f"B2:B{len(codes) + 1}").Value2
            parsed = [_excel_serial_to_timestamp(row[0]) for row in values]
            valid_count = int(pd.Series(parsed).notna().sum())
            if valid_count == len(codes) and excel.CalculationState == 0:
                break
            if valid_count == previous_valid_count and excel.CalculationState == 0:
                stable_polls += 1
            else:
                stable_polls = 0
            previous_valid_count = valid_count
            if stable_polls >= 5:
                break
        if values is None:
            raise RuntimeError("Wind Excel 未返回任何摘牌日期")
        parsed = [_excel_serial_to_timestamp(row[0]) for row in values]
        if valid_count < len(codes):
            missing = [code for code, value in zip(codes, parsed) if pd.isna(value)]
            print(
                f"警告：Wind Excel 摘牌日期有效返回 {valid_count}/{len(codes)}；"
                f"无有效值代码：{','.join(missing)}"
            )
        return pd.Series(parsed, index=codes, name="s_info_delistdate")
    finally:
        if com_book is not None:
            com_book.Close(False)
        excel.Quit()


def fetch_delist_dates(codes: list[str], work_dir: Path) -> pd.Series:
    # 当前 Wind 账户无 Python API 权限，但 Excel 插件有 WSS 权限。
    # 固定使用 Excel 插件，避免 WindPy 在无权限时直接终止宿主进程。
    return fetch_delist_dates_excel(codes, work_dir)


def latest_balances(
    parquet_root: Path,
    codes: list[str],
    as_of: pd.Timestamp,
) -> pd.DataFrame:
    year_dir = parquet_root / str(as_of.year)
    files = sorted(year_dir.glob("*.parquet"))
    if not files:
        raise FileNotFoundError(f"未找到 {as_of.year} 年月度 parquet：{year_dir}")

    records: list[pd.DataFrame] = []
    code_set = set(codes)
    for path in files:
        frame = pd.read_parquet(path)
        frame = frame[
            (frame["__sheet_name"] == "余额")
            & frame["__row_id"].astype(str).isin(code_set)
        ]
        if frame.empty:
            continue
        date_cols = []
        for col in frame.columns:
            if col in {"__sheet_name", "__row_id"}:
                continue
            stamp = pd.to_datetime(col, errors="coerce")
            if pd.notna(stamp) and stamp <= as_of:
                date_cols.append((col, stamp))
        for col, stamp in date_cols:
            part = pd.DataFrame(
                {
                    "wind_code": frame["__row_id"].astype(str),
                    "balance_date": stamp,
                    "balance": pd.to_numeric(frame[col], errors="coerce"),
                }
            )
            records.append(part)

    if not records:
        return pd.DataFrame(columns=["wind_code", "balance_date", "balance"])
    long = pd.concat(records, ignore_index=True).dropna(subset=["balance"])
    if long.empty:
        return pd.DataFrame(columns=["wind_code", "balance_date", "balance"])
    return (
        long.sort_values(["wind_code", "balance_date"])
        .groupby("wind_code", as_index=False)
        .tail(1)
        .reset_index(drop=True)
    )


def build_summary(
    parquet_root: Path,
    as_of: pd.Timestamp,
    work_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    universe = load_bond_universe(parquet_root)
    delist = fetch_delist_dates(universe["wind_code"].tolist(), work_dir)
    universe["delist_date"] = universe["wind_code"].map(delist)

    window_end = as_of + pd.DateOffset(months=1)
    selected = universe[
        universe["delist_date"].gt(as_of)
        & universe["delist_date"].le(window_end)
    ].copy()
    selected = selected.sort_values(["delist_date", "wind_code"]).reset_index(drop=True)

    balances = latest_balances(parquet_root, selected["wind_code"].tolist(), as_of)
    selected = selected.merge(balances, on="wind_code", how="left")

    not_delisted = universe["delist_date"].gt(as_of) | (
        universe["delist_date"].isna()
        & universe["listing_date"].notna()
        & universe["last_trade_date"].gt(as_of)
    )
    redeeming = universe[
        universe["redemption_notice_date"].notna()
        & universe["redemption_notice_date"].le(as_of)
        & not_delisted
    ].copy()
    redeeming_balances = latest_balances(
        parquet_root,
        redeeming["wind_code"].tolist(),
        as_of,
    )
    redeeming = redeeming.merge(redeeming_balances, on="wind_code", how="left")
    redeeming = redeeming[
        pd.to_numeric(redeeming["balance"], errors="coerce").fillna(0).gt(0)
    ].sort_values(["redemption_notice_date", "wind_code"], ascending=[False, True])
    return selected, redeeming.reset_index(drop=True)


def format_paragraph(result: pd.DataFrame, as_of: pd.Timestamp) -> str:
    window_end = as_of + pd.DateOffset(months=1)
    total = result["balance"].sum(min_count=1)
    total_text = "暂无余额数据" if pd.isna(total) else f"合计余额{total:.2f}亿元"
    names = "、".join(result["bond_name"].astype(str)) if len(result) else "无"
    return (
        f"退出方面，{as_of:%Y年%m月%d日}至{window_end:%Y年%m月%d日}期间"
        f"将摘牌转债目前有{len(result)}只，{total_text}，包括{names}。"
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--as-of", required=True, help="报告基准日，YYYY-MM-DD")
    parser.add_argument(
        "--parquet-root",
        type=Path,
        default=Path.cwd() / "data/转债个券历史序列",
    )
    parser.add_argument("--out-dir", type=Path, default=Path.cwd() / "tmp")
    args = parser.parse_args()

    as_of = pd.Timestamp(args.as_of).normalize()
    args.out_dir.mkdir(parents=True, exist_ok=True)
    result, redeeming = build_summary(args.parquet_root, as_of, args.out_dir)
    csv_path = args.out_dir / f"cb_delist_{as_of:%Y%m%d}.csv"
    json_path = args.out_dir / f"cb_delist_{as_of:%Y%m%d}.json"
    result.to_csv(csv_path, index=False, encoding="utf-8-sig", date_format="%Y-%m-%d")
    payload = {
        "as_of": as_of.strftime("%Y-%m-%d"),
        "window_end": (as_of + pd.DateOffset(months=1)).strftime("%Y-%m-%d"),
        "count": int(len(result)),
        "total_balance": None
        if pd.isna(result["balance"].sum(min_count=1))
        else float(result["balance"].sum()),
        "paragraph": format_paragraph(result, as_of),
        "records": result.assign(
            redemption_notice_date=result["redemption_notice_date"].dt.strftime("%Y-%m-%d"),
            listing_date=result["listing_date"].dt.strftime("%Y-%m-%d"),
            last_trade_date=result["last_trade_date"].dt.strftime("%Y-%m-%d"),
            delist_date=result["delist_date"].dt.strftime("%Y-%m-%d"),
            balance_date=result["balance_date"].dt.strftime("%Y-%m-%d"),
        ).to_dict(orient="records"),
        "redeeming_count": int(len(redeeming)),
        "redeeming_total_balance": None
        if pd.isna(redeeming["balance"].sum(min_count=1))
        else float(redeeming["balance"].sum()),
        "redeeming_bonds": redeeming["bond_name"].astype(str).tolist(),
        "redeeming_records": redeeming.assign(
            redemption_notice_date=redeeming["redemption_notice_date"].dt.strftime("%Y-%m-%d"),
            listing_date=redeeming["listing_date"].dt.strftime("%Y-%m-%d"),
            last_trade_date=redeeming["last_trade_date"].dt.strftime("%Y-%m-%d"),
            delist_date=redeeming["delist_date"].dt.strftime("%Y-%m-%d"),
            balance_date=redeeming["balance_date"].dt.strftime("%Y-%m-%d"),
        ).to_dict(orient="records"),
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
