from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
PARQUET_ROOT = ROOT / "data/转债个券历史序列"
OUTPUT_DIR = ROOT / "outputs" / "close_return_wide_exclude_first10"
XLSX_PATH = OUTPUT_DIR / "收盘价与涨跌幅时间序列_宽表_成交额大于零且剔除首10个交易日.xlsx"

CLOSE = "收盘价"
RETURN = "涨跌幅"
AMOUNT = "成交额"
META_COLS = {"__sheet_name", "__row_id"}


def monthly_files() -> list[Path]:
    return sorted(
        path
        for path in PARQUET_ROOT.glob("*/*.parquet")
        if path.parent.name.isdigit() and path.stem.isdigit()
    )


def date_columns(df: pd.DataFrame) -> list[str]:
    parsed: list[tuple[pd.Timestamp, str]] = []
    for col in df.columns:
        if col in META_COLS:
            continue
        try:
            parsed.append((pd.to_datetime(col), col))
        except (TypeError, ValueError):
            continue
    return [col for _, col in sorted(parsed)]


def metric_frame(df: pd.DataFrame, metric: str, dates: list[str]) -> pd.DataFrame:
    frame = df.loc[df["__sheet_name"].eq(metric), ["__row_id", *dates]].copy()
    frame["__row_id"] = frame["__row_id"].astype(str)
    frame = frame.drop_duplicates("__row_id", keep="last").set_index("__row_id")
    return frame[dates].apply(pd.to_numeric, errors="coerce")


def transpose_metric(frame: pd.DataFrame, dates: list[str]) -> pd.DataFrame:
    result = frame.T
    result.index = pd.to_datetime(dates)
    result.index.name = "日期"
    return result


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    label_fill = PatternFill("solid", fgColor="D9EAF7")
    label_font = Font(color="1F4E78", bold=True)

    for sheet_name in [CLOSE, RETURN]:
        ws = wb[sheet_name]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        ws.column_dimensions["A"].width = 13
        for cell in ws["A"][1:]:
            cell.number_format = "yyyy-mm-dd"
        for col_idx in range(2, min(ws.max_column, 30) + 1):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 12

    ws = wb["说明"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for cell in ws["A"][1:]:
        cell.fill = label_fill
        cell.font = label_font
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 96
    wb.save(path)


def main() -> None:
    files = monthly_files()
    if not files:
        raise SystemExit(f"未找到月度 parquet：{PARQUET_ROOT}")

    close_parts: list[pd.DataFrame] = []
    return_parts: list[pd.DataFrame] = []
    amount_parts: list[pd.DataFrame] = []

    for path in files:
        df = pd.read_parquet(path)
        dates = date_columns(df)
        if not dates:
            continue
        available = set(df["__sheet_name"].dropna().astype(str))
        missing = {CLOSE, RETURN, AMOUNT} - available
        if missing:
            raise KeyError(f"{path} 缺少指标：{', '.join(sorted(missing))}")

        close_parts.append(transpose_metric(metric_frame(df, CLOSE, dates), dates))
        return_parts.append(transpose_metric(metric_frame(df, RETURN, dates), dates))
        amount_parts.append(transpose_metric(metric_frame(df, AMOUNT, dates), dates))

    close = pd.concat(close_parts).sort_index()
    returns = pd.concat(return_parts).sort_index()
    amount = pd.concat(amount_parts).sort_index()
    codes = sorted(set(close.columns) | set(returns.columns) | set(amount.columns))
    close = close.reindex(columns=codes)
    returns = returns.reindex(index=close.index, columns=codes)
    amount = amount.reindex(index=close.index, columns=codes)

    # A bond's trading-day sequence is based on dates with a valid close price.
    trading_day_number = close.notna().cumsum(axis=0)
    valid_mask = amount.gt(0) & close.notna() & trading_day_number.gt(10)
    close_clean = close.where(valid_mask)
    return_clean = returns.where(valid_mask)

    first_ten_cells = int((close.notna() & trading_day_number.le(10)).sum().sum())
    positive_amount_cells = int(amount.gt(0).sum().sum())
    close_cells = int(close_clean.notna().sum().sum())
    return_cells = int(return_clean.notna().sum().sum())
    notes = pd.DataFrame(
        [
            ("数据来源", str(PARQUET_ROOT)),
            ("源文件数", len(files)),
            ("日期范围", f"{close.index.min():%Y-%m-%d} 至 {close.index.max():%Y-%m-%d}"),
            ("交易日数", len(close.index)),
            ("转债代码数", len(codes)),
            ("宽表结构", "行=日期，列=转债代码；分别输出收盘价和涨跌幅。"),
            ("首10日定义", "每只转债自首次出现有效收盘价起，按日期排序的前10个有效交易日。"),
            ("清洗规则", "仅保留成交额大于0且交易日序号大于10的记录，其余置为空。"),
            ("首10日观测数", first_ten_cells),
            ("成交额大于0观测数", positive_amount_cells),
            ("收盘价有效单元格数", close_cells),
            ("涨跌幅有效单元格数", return_cells),
        ],
        columns=["项目", "说明"],
    )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        close_clean.reset_index().to_excel(writer, sheet_name=CLOSE, index=False)
        return_clean.reset_index().to_excel(writer, sheet_name=RETURN, index=False)
        notes.to_excel(writer, sheet_name="说明", index=False)
    print(XLSX_PATH)
    print(
        f"files={len(files)} dates={len(close.index)} bonds={len(codes)} "
        f"close_cells={close_cells} return_cells={return_cells}"
    )


if __name__ == "__main__":
    main()
