from __future__ import annotations

import argparse
import json
import os
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import pyarrow.parquet as pq


INDEX_SHEET = "指数"
FORMULA_RE = re.compile(r'i_dq_close\(\s*"([^"]+)"\s*,', re.IGNORECASE)
ROW_ID_OVERRIDES = {
    "885005.WI": "债券型基金",
    "885063.WI": "债券指数型基金",
}
LEGACY_ROW_IDS = {
    "债券型指数基金",
    "债券型基金（885005.WI）",
    "债券型基金（885063.WI）",
}


@dataclass(frozen=True)
class Target:
    source_column: int
    header: str
    code: str
    row_id: str


def monthly_files(root: Path) -> list[Path]:
    files: list[Path] = []
    for path in root.glob("[0-9][0-9][0-9][0-9]/*.parquet"):
        if re.fullmatch(r"\d{6}", path.stem):
            files.append(path)
    return sorted(files)


def extract_targets(workbook_path: Path) -> list[Target]:
    workbook = openpyxl.load_workbook(
        workbook_path, data_only=False, read_only=True, keep_links=False
    )
    sheet = workbook[workbook.sheetnames[0]]
    first_two_rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
    headers = first_two_rows[0]
    formulas = first_two_rows[1]
    raw: list[tuple[int, str, str]] = []
    for column in range(2, len(headers) + 1):
        header = str(headers[column - 1] or "").strip()
        if not header:
            continue
        formula = str(formulas[column - 1] or "")
        match = FORMULA_RE.search(formula)
        if not match:
            raise RuntimeError(
                f"{sheet.title}第{column}列无法从公式第一个参数提取指数代码：{formula}"
            )
        raw.append((column, header, match.group(1).strip()))
    workbook.close()

    if not raw:
        raise RuntimeError("补充指数工作簿未识别到指数列")
    codes = [code for _, _, code in raw]
    if len(codes) != len(set(codes)):
        duplicates = [code for code, count in Counter(codes).items() if count > 1]
        raise RuntimeError(f"指数代码重复：{duplicates}")

    header_counts = Counter(header for _, header, _ in raw)
    return [
        Target(
            source_column=column,
            header=header,
            code=code,
            row_id=ROW_ID_OVERRIDES.get(
                code, header if header_counts[header] == 1 else f"{header}（{code}）"
            ),
        )
        for column, header, code in raw
    ]


def load_source(workbook_path: Path, targets: list[Target]) -> pd.DataFrame:
    workbook = openpyxl.load_workbook(
        workbook_path, data_only=True, read_only=True, keep_links=False
    )
    sheet = workbook[workbook.sheetnames[0]]
    records: list[list[object]] = []
    for row_values in sheet.iter_rows(min_row=2, values_only=True):
        raw_date = row_values[0]
        if raw_date is None:
            continue
        date = pd.Timestamp(raw_date).normalize()
        values: list[float] = []
        for target in targets:
            value = pd.to_numeric(row_values[target.source_column - 1], errors="coerce")
            if pd.isna(value):
                raise RuntimeError(
                    f"源表{date:%Y-%m-%d}的{target.header}（{target.code}）为空或非数值"
                )
            values.append(float(value))
        records.append([date, *values])
    workbook.close()

    source = pd.DataFrame(
        records, columns=["date", *[target.row_id for target in targets]]
    ).set_index("date")
    if source.index.has_duplicates:
        duplicates = source.index[source.index.duplicated()].strftime("%Y-%m-%d").tolist()
        raise RuntimeError(f"源表日期重复：{duplicates[:10]}")
    return source.sort_index()


def value_as_text(value: float) -> str:
    return format(float(value), ".15g")


def index_rows(frame: pd.DataFrame) -> pd.DataFrame:
    return frame.loc[frame["__sheet_name"] == INDEX_SHEET]


def audit_file(
    path: Path,
    source: pd.DataFrame,
    targets: list[Target],
    expected_non_target_order: list[str],
) -> dict:
    date_columns = list(pq.ParquetFile(path).schema.names[2:])
    parsed_dates = [pd.Timestamp(column).normalize() for column in date_columns]
    missing_source_dates = [
        date.strftime("%Y-%m-%d") for date in parsed_dates if date not in source.index
    ]
    if missing_source_dates:
        raise RuntimeError(f"{path.name}存在源表未覆盖日期：{missing_source_dates[:10]}")

    frame = pd.read_parquet(path, columns=["__sheet_name", "__row_id"])
    rows = index_rows(frame)
    target_ids = {target.row_id for target in targets}
    replaceable_ids = target_ids | LEGACY_ROW_IDS
    non_target_order = rows.loc[
        ~rows["__row_id"].isin(replaceable_ids), "__row_id"
    ].tolist()
    if non_target_order != expected_non_target_order:
        raise RuntimeError(f"{path.name}原有指数顺序不一致：{non_target_order}")
    counts = rows["__row_id"].value_counts().to_dict()
    duplicates = [target.row_id for target in targets if counts.get(target.row_id, 0) > 1]
    if duplicates:
        raise RuntimeError(f"{path.name}目标指数存在重复行：{duplicates}")

    return {
        "file": str(path),
        "date_count": len(date_columns),
        "first_date": parsed_dates[0].strftime("%Y-%m-%d"),
        "last_date": parsed_dates[-1].strftime("%Y-%m-%d"),
        "index_row_count": len(rows),
        "target_rows": {target.row_id: int(counts.get(target.row_id, 0)) for target in targets},
    }


def build_updated_frame(
    frame: pd.DataFrame, source: pd.DataFrame, targets: list[Target]
) -> pd.DataFrame:
    date_columns = list(frame.columns[2:])
    parsed_dates = [pd.Timestamp(column).normalize() for column in date_columns]
    target_ids = {target.row_id for target in targets}
    replaceable_ids = target_ids | LEGACY_ROW_IDS
    target_mask = (frame["__sheet_name"] == INDEX_SHEET) & frame["__row_id"].isin(
        replaceable_ids
    )
    base = frame.loc[~target_mask].reset_index(drop=True)

    positions = base.index[base["__sheet_name"] == INDEX_SHEET].tolist()
    if not positions:
        raise RuntimeError("parquet中不存在指数分区")
    if positions != list(range(positions[0], positions[-1] + 1)):
        raise RuntimeError("parquet中的指数分区不是连续区块")
    insertion = positions[-1] + 1

    records: list[dict[str, object]] = []
    for target in targets:
        row: dict[str, object] = {"__sheet_name": INDEX_SHEET, "__row_id": target.row_id}
        for column, date in zip(date_columns, parsed_dates):
            row[column] = value_as_text(source.at[date, target.row_id])
        records.append(row)
    additions = pd.DataFrame(records, columns=frame.columns)
    for column in frame.columns:
        additions[column] = additions[column].astype(frame[column].dtype)

    return pd.concat(
        [base.iloc[:insertion], additions, base.iloc[insertion:]], ignore_index=True
    )


def validate_updated_frame(
    original: pd.DataFrame,
    updated: pd.DataFrame,
    source: pd.DataFrame,
    targets: list[Target],
    expected_non_target_order: list[str],
    path: Path,
) -> None:
    target_ids = {target.row_id for target in targets}
    replaceable_ids = target_ids | LEGACY_ROW_IDS
    original_base = original.loc[
        ~(
            (original["__sheet_name"] == INDEX_SHEET)
            & original["__row_id"].isin(replaceable_ids)
        )
    ].reset_index(drop=True)
    updated_base = updated.loc[
        ~(
            (updated["__sheet_name"] == INDEX_SHEET)
            & updated["__row_id"].isin(replaceable_ids)
        )
    ].reset_index(drop=True)
    pd.testing.assert_frame_equal(original_base, updated_base, check_dtype=True)

    if list(updated.columns) != list(original.columns):
        raise RuntimeError(f"{path.name}更新后列结构变化")
    rows = index_rows(updated)
    expected_order = [*expected_non_target_order, *[target.row_id for target in targets]]
    actual_order = rows["__row_id"].tolist()
    if actual_order != expected_order:
        raise RuntimeError(f"{path.name}更新后指数顺序异常：{actual_order}")

    date_columns = list(updated.columns[2:])
    parsed_dates = [pd.Timestamp(column).normalize() for column in date_columns]
    for target in targets:
        selected = rows.loc[rows["__row_id"] == target.row_id]
        if len(selected) != 1:
            raise RuntimeError(f"{path.name}的{target.row_id}行数为{len(selected)}")
        actual = pd.to_numeric(selected.iloc[0][date_columns], errors="coerce").to_numpy(
            dtype=float
        )
        expected = source.loc[parsed_dates, target.row_id].to_numpy(dtype=float)
        if not np.allclose(actual, expected, rtol=0, atol=1e-8):
            mismatches = int((~np.isclose(actual, expected, rtol=0, atol=1e-8)).sum())
            raise RuntimeError(f"{path.name}的{target.row_id}有{mismatches}个值未通过校验")


def write_monthly_file(
    path: Path,
    source: pd.DataFrame,
    targets: list[Target],
    expected_non_target_order: list[str],
) -> None:
    original = pd.read_parquet(path)
    updated = build_updated_frame(original, source, targets)
    validate_updated_frame(
        original, updated, source, targets, expected_non_target_order, path
    )

    temp_path = path.with_name(path.name + ".fund_index_tmp")
    try:
        updated.to_parquet(temp_path, index=False, engine="pyarrow", compression="snappy")
        reread = pd.read_parquet(temp_path)
        validate_updated_frame(
            original, reread, source, targets, expected_non_target_order, path
        )
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def update_manifest(root: Path, expected_rows: int) -> None:
    path = root / "_meta" / "sheet_manifest.parquet"
    manifest = pd.read_parquet(path)
    mask = manifest["sheet_name"] == INDEX_SHEET
    if int(mask.sum()) != 1:
        raise RuntimeError("元数据清单中无法唯一定位指数行")
    manifest.loc[mask, "rows"] = expected_rows
    temp_path = path.with_name(path.name + ".fund_index_tmp")
    try:
        manifest.to_parquet(temp_path, index=False, engine="pyarrow", compression="snappy")
        reread = pd.read_parquet(temp_path)
        actual = int(reread.loc[reread["sheet_name"] == INDEX_SHEET, "rows"].iloc[0])
        if actual != expected_rows:
            raise RuntimeError(f"元数据清单指数行数校验失败：{actual} != {expected_rows}")
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def final_audit(
    files: list[Path],
    source: pd.DataFrame,
    targets: list[Target],
    expected_non_target_order: list[str],
    root: Path,
) -> dict:
    audits = [
        audit_file(path, source, targets, expected_non_target_order) for path in files
    ]
    manifest = pd.read_parquet(root / "_meta" / "sheet_manifest.parquet")
    manifest_rows = int(
        manifest.loc[manifest["sheet_name"] == INDEX_SHEET, "rows"].iloc[0]
    )
    return {
        "monthly_files": len(files),
        "parquet_dates": sum(item["date_count"] for item in audits),
        "first_date": audits[0]["first_date"],
        "last_date": audits[-1]["last_date"],
        "source_first_date": source.index.min().strftime("%Y-%m-%d"),
        "source_last_date": source.index.max().strftime("%Y-%m-%d"),
        "index_rows_per_file": sorted({item["index_row_count"] for item in audits}),
        "target_file_counts": {
            target.row_id: sum(item["target_rows"][target.row_id] == 1 for item in audits)
            for target in targets
        },
        "manifest_index_rows": manifest_rows,
        "targets": [
            {"header": target.header, "code": target.code, "row_id": target.row_id}
            for target in targets
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="从补充指数Excel公式的第一个参数提取代码，并写入月度parquet指数分区"
    )
    parser.add_argument("excel", type=Path)
    parser.add_argument("parquet_root", type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    excel = args.excel.resolve()
    root = args.parquet_root.resolve()
    targets = extract_targets(excel)
    source = load_source(excel, targets)
    files = monthly_files(root)
    if not files:
        raise RuntimeError("未找到月度parquet文件")

    first = pd.read_parquet(files[0], columns=["__sheet_name", "__row_id"])
    target_ids = {target.row_id for target in targets}
    replaceable_ids = target_ids | LEGACY_ROW_IDS
    expected_non_target_order = index_rows(first).loc[
        ~index_rows(first)["__row_id"].isin(replaceable_ids), "__row_id"
    ].tolist()

    before = final_audit(
        files, source, targets, expected_non_target_order, root
    )
    print("更新前审计：")
    print(json.dumps(before, ensure_ascii=False, indent=2))
    if not args.apply:
        print("仅审计，未写入。添加--apply后执行更新。")
        return

    for number, path in enumerate(files, start=1):
        write_monthly_file(path, source, targets, expected_non_target_order)
        if number % 20 == 0 or number == len(files):
            print(f"已更新 {number}/{len(files)} 个文件", flush=True)

    expected_rows = len(expected_non_target_order) + len(targets)
    update_manifest(root, expected_rows)
    after = final_audit(files, source, targets, expected_non_target_order, root)
    print("更新后审计：")
    print(json.dumps(after, ensure_ascii=False, indent=2))
    if after["index_rows_per_file"] != [expected_rows]:
        raise RuntimeError("并非所有月度parquet都包含预期数量的指数行")
    if any(count != len(files) for count in after["target_file_counts"].values()):
        raise RuntimeError("并非所有目标指数都已覆盖全部月度parquet")
    if after["manifest_index_rows"] != expected_rows:
        raise RuntimeError("元数据清单指数行数未通过最终校验")


if __name__ == "__main__":
    main()
