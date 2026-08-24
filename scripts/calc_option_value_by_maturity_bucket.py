from __future__ import annotations

import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
PARQUET_ROOT = ROOT / "data/转债个券历史序列"
OUTPUT_DIR = ROOT / "outputs" / "option_value_by_maturity_bucket"
XLSX_PATH = OUTPUT_DIR / "转债期权价值_个券层面_剩余期限0.1年分档均值_2017年至今.xlsx"
DETAIL_CSV = OUTPUT_DIR / "转债期权价值样本明细_2017年至今.csv"
SUMMARY_JSON = OUTPUT_DIR / "summary.json"

CLOSE = "收盘价"
PURE_BOND = "纯债价值"
MATURITY = "剩余期限"
META_COLS = ["__sheet_name", "__row_id"]
START_YEAR = 2017
BUCKET_STEP = 0.1


def iter_monthly_parquets(root: Path) -> list[Path]:
    return sorted(
        path
        for path in root.glob("*/*.parquet")
        if path.parent.name.isdigit()
        and path.stem.isdigit()
        and int(path.parent.name) >= START_YEAR
    )


def date_columns(df: pd.DataFrame) -> list[str]:
    cols: list[tuple[pd.Timestamp, str]] = []
    for col in df.columns:
        if col in META_COLS:
            continue
        try:
            cols.append((pd.to_datetime(col), col))
        except Exception:
            continue
    return [col for _, col in sorted(cols)]


def metric_frame(df: pd.DataFrame, metric: str, dates: list[str]) -> pd.DataFrame:
    frame = df.loc[df["__sheet_name"].eq(metric), ["__row_id", *dates]].copy()
    frame["__row_id"] = frame["__row_id"].astype(str)
    frame = frame.drop_duplicates("__row_id", keep="last").set_index("__row_id")
    return frame[dates].apply(pd.to_numeric, errors="coerce")


def metric_long(frame: pd.DataFrame, name: str) -> pd.DataFrame:
    out = frame.stack().dropna().rename(name).reset_index()
    out.columns = ["转债代码", "日期", name]
    out["日期"] = pd.to_datetime(out["日期"])
    return out


def build_samples() -> pd.DataFrame:
    files = iter_monthly_parquets(PARQUET_ROOT)
    if not files:
        raise SystemExit(f"未找到 {START_YEAR} 年以来的月度 parquet: {PARQUET_ROOT}")

    parts: list[pd.DataFrame] = []
    for file in files:
        df = pd.read_parquet(file)
        dates = date_columns(df)
        if not dates:
            continue

        missing = {CLOSE, PURE_BOND, MATURITY} - set(df["__sheet_name"].dropna().unique())
        if missing:
            raise KeyError(f"{file} 缺少指标: {', '.join(sorted(missing))}")

        close = metric_frame(df, CLOSE, dates)
        pure = metric_frame(df, PURE_BOND, dates)
        maturity = metric_frame(df, MATURITY, dates)

        long = metric_long(close, CLOSE)
        long = long.merge(metric_long(pure, PURE_BOND), on=["转债代码", "日期"], how="inner")
        long = long.merge(metric_long(maturity, MATURITY), on=["转债代码", "日期"], how="inner")
        parts.append(long)

    samples = pd.concat(parts, ignore_index=True)
    samples = samples.dropna(subset=[CLOSE, PURE_BOND, MATURITY])
    samples = samples[np.isfinite(samples[CLOSE]) & np.isfinite(samples[PURE_BOND]) & np.isfinite(samples[MATURITY])]
    samples = samples[samples[MATURITY] >= 0].copy()
    samples["期权价值"] = samples[CLOSE] - samples[PURE_BOND]

    bucket_start = np.floor(samples[MATURITY] / BUCKET_STEP) * BUCKET_STEP
    samples["剩余期限分档下限"] = np.round(bucket_start, 1)
    samples["剩余期限分档上限"] = np.round(samples["剩余期限分档下限"] + BUCKET_STEP, 1)
    samples["剩余期限分档"] = samples.apply(
        lambda row: f"[{row['剩余期限分档下限']:.1f}, {row['剩余期限分档上限']:.1f})",
        axis=1,
    )
    return samples


def summarize_sample_weighted(samples: pd.DataFrame) -> pd.DataFrame:
    grouped = (
        samples.groupby(["剩余期限分档下限", "剩余期限分档上限", "剩余期限分档"], as_index=False)
        .agg(
            样本数=("期权价值", "size"),
            期权价值均值=("期权价值", "mean"),
            期权价值中位数=("期权价值", "median"),
            收盘价均值=(CLOSE, "mean"),
            纯债价值均值=(PURE_BOND, "mean"),
            剩余期限均值=(MATURITY, "mean"),
        )
        .sort_values("剩余期限分档下限")
    )
    return grouped


def summarize_bond_level(samples: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    bond_bucket = (
        samples.groupby(["转债代码", "剩余期限分档下限", "剩余期限分档上限", "剩余期限分档"], as_index=False)
        .agg(
            个券样本数=("期权价值", "size"),
            个券期权价值均值=("期权价值", "mean"),
            个券期权价值中位数=("期权价值", "median"),
            个券收盘价均值=(CLOSE, "mean"),
            个券纯债价值均值=(PURE_BOND, "mean"),
            个券剩余期限均值=(MATURITY, "mean"),
        )
        .sort_values(["剩余期限分档下限", "转债代码"])
    )
    bucket_summary = (
        bond_bucket.groupby(["剩余期限分档下限", "剩余期限分档上限", "剩余期限分档"], as_index=False)
        .agg(
            个券数=("转债代码", "nunique"),
            原始样本数=("个券样本数", "sum"),
            期权价值均值_个券等权=("个券期权价值均值", "mean"),
            期权价值中位数_个券等权=("个券期权价值均值", "median"),
            收盘价均值_个券等权=("个券收盘价均值", "mean"),
            纯债价值均值_个券等权=("个券纯债价值均值", "mean"),
            剩余期限均值_个券等权=("个券剩余期限均值", "mean"),
        )
        .sort_values("剩余期限分档下限")
    )
    return bucket_summary, bond_bucket


def write_excel(
    bucket_summary: pd.DataFrame,
    bond_bucket: pd.DataFrame,
    sample_weighted_summary: pd.DataFrame,
    samples: pd.DataFrame,
    source_file_count: int,
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    detail_cols = ["日期", "转债代码", CLOSE, PURE_BOND, "期权价值", MATURITY, "剩余期限分档"]
    samples[detail_cols].to_csv(DETAIL_CSV, index=False, encoding="utf-8-sig")

    note = pd.DataFrame(
        [
            ("数据来源", str(PARQUET_ROOT)),
            ("源文件数量", source_file_count),
            ("日期范围", f"{samples['日期'].min().date()} 至 {samples['日期'].max().date()}"),
            ("样本口径", "2017年至今全部日期-转债观察值；收盘价、纯债价值、剩余期限均有效，且剩余期限>=0。"),
            ("计算公式", "期权价值 = 收盘价 - 纯债价值"),
            ("分档方法", "按剩余期限向下取整至0.1年分档，区间为左闭右开。"),
            ("主结果口径", "先在“转债代码 x 剩余期限分档”层面计算个券期权价值均值，再对同一分档内个券均值做等权平均。"),
            ("对照口径", "“分档均值_样本加权”直接按日期-个券观察值求均值，会让样本天数更多的个券权重更高。"),
            ("明细文件", str(DETAIL_CSV)),
        ],
        columns=["项目", "说明"],
    )

    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        bucket_summary.to_excel(writer, sheet_name="分档均值_个券等权", index=False)
        bond_bucket.to_excel(writer, sheet_name="个券分档明细", index=False)
        sample_weighted_summary.to_excel(writer, sheet_name="分档均值_样本加权", index=False)
        note.to_excel(writer, sheet_name="说明", index=False)

    wb = load_workbook(XLSX_PATH)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    label_fill = PatternFill("solid", fgColor="D9EAF7")
    label_font = Font(color="1F4E78", bold=True)

    ws = wb["分档均值_个券等权"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for col, width in {"A": 16, "B": 16, "C": 18, "D": 12, "E": 14, "F": 20, "G": 20, "H": 18, "I": 18, "J": 18}.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
        for cell in row:
            cell.number_format = "0.0"
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=10):
        for cell in row:
            cell.number_format = "0.0000"

    for sheet_name in ["个券分档明细", "分档均值_样本加权"]:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for col_idx in range(1, min(ws.max_column, 12) + 1):
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = 16

    note_ws = wb["说明"]
    for cell in note_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for cell in note_ws["A"][1:]:
        cell.fill = label_fill
        cell.font = label_font
    note_ws.column_dimensions["A"].width = 18
    note_ws.column_dimensions["B"].width = 110

    wb.save(XLSX_PATH)


def main() -> None:
    files = iter_monthly_parquets(PARQUET_ROOT)
    samples = build_samples()
    sample_weighted_summary = summarize_sample_weighted(samples)
    bucket_summary, bond_bucket = summarize_bond_level(samples)
    write_excel(bucket_summary, bond_bucket, sample_weighted_summary, samples, len(files))

    meta = {
        "source": str(PARQUET_ROOT),
        "source_file_count": len(files),
        "sample_count": int(len(samples)),
        "bond_bucket_count": int(len(bond_bucket)),
        "bucket_count": int(len(bucket_summary)),
        "date_start": str(samples["日期"].min().date()),
        "date_end": str(samples["日期"].max().date()),
        "output_xlsx": str(XLSX_PATH),
        "detail_csv": str(DETAIL_CSV),
    }
    SUMMARY_JSON.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    print(XLSX_PATH)
    print(
        f"samples={meta['sample_count']} bond_buckets={meta['bond_bucket_count']} "
        f"buckets={meta['bucket_count']} files={meta['source_file_count']}"
    )


if __name__ == "__main__":
    main()
