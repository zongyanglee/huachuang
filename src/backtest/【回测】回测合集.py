"""性价比回测：残差修正幂衰减定价。

核心口径：
1. 每 20 个交易日调仓，使用前一交易日截面拟合，避免前视偏差。
2. 先拟合幂衰减基准曲线，再对残差做低维线性修正。
3. 理论价 = 平价 * (1 + 理论转股溢价率 / 100)。
4. 低估比例 = 理论价 / 实际收盘价 - 1；只持有低估比例 > 0 的前 N 只。
5. 同期输出双低、低价、低转股溢价率、动量、偏股、偏债、平衡七个对照组合。
6. 调仓日等权买入，持有期内权重随个券涨跌自然漂移；缺失日收益按 0 处理。
"""

from __future__ import annotations

import hashlib
import os
import pickle
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Optional

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pandas.tseries.offsets import BDay
from scipy.optimize import curve_fit
from tqdm import tqdm

import sys

WORKSPACE_ROOT = Path(__file__).resolve().parents[2]
_COMMON_MODULE_DIR = WORKSPACE_ROOT / "src" / "common"
if str(_COMMON_MODULE_DIR) not in sys.path:
    sys.path.insert(0, str(_COMMON_MODULE_DIR))

from 转债Parquet标准读写模块 import (
    INDEX_NAME,
    INDEX_VALUE,
    INDEX_SHEET,
    TRADE_DATE,
    _write_manifest,
    index_schema,
    read_original_data_from_parquet as _read_standard_parquet,
    write_typed_parquet,
)


RESIDUAL_POWER_ANCHOR = 50.0
RESIDUAL_POWER_LOWER = 50.0
RESIDUAL_POWER_UPPER = 200.0
RESIDUAL_FEATURES = [
    "余额_log",
    "剩余期限",
    "正股20日波动率",
    "赎回累计天数",
    "下修累计天数",
    "隐含波动率",
]

PORTFOLIO_NAMES = ("性价比", "双低", "低价", "低转股溢价率", "动量", "偏股", "偏债", "平衡")
DECOMP_COLUMNS = ["日期", "转债收益率", "债券贡献", "正股贡献", "估值贡献"]
DECOMP_GROUP_COLUMNS = ["分类", "转债收益率", "债券贡献", "正股贡献", "估值贡献"]


@dataclass
class BacktestConfig:
    # 唯一底稿数据源：标准 parquet 目录。
    parquet_root: str = str(WORKSPACE_ROOT / "data" / "转债个券历史序列")
    # True 时跳过内存/磁盘缓存，强制重新读取源数据。
    force_refresh: bool = False

    # 完整计算及 parquet 净值回写口径：自 2015 年以来全部重算。
    backtest_start_date: str = "2015-01-01"
    # Excel、图表及其中的绩效评估仅展示该日期以来的数据，并以区间首个有效值归一为 1。
    report_start_date: str = "2022-01-01"
    rebalance_every_n_days: int = 20
    # 单个组合最多持仓数量；候选不足时按实际可选数量持有。
    max_holdings: int = 30
    # 单边换手交易成本率，净收益按组合换手率扣减；默认不计交易费用。
    transaction_cost_rate: float = 0.0
    # 夏普比率使用指数表中的逐日十年国债收益率作为无风险利率。
    risk_free_rate_name: str = "十年国债"

    # 基础筛选：剔除余额过小、剩余期限过短，以及已公告强赎的转债。
    min_balance: float = 3.0
    min_remaining_years: float = 0.5
    exclude_redeem_announced: bool = True
    # 模型拟合样本过滤：限制换手率，并按溢价率分位数去除极端值。
    fit_turnover_upper: float = 50.0
    fit_premium_low_quantile: float = 0.03
    fit_premium_high_quantile: float = 0.97
    # 每期截面拟合的最低样本数；不足时该期不生成性价比调仓组合。
    minimum_fit_samples: int = len(RESIDUAL_FEATURES) + 30

    # 输入指标名称映射，需与标准 parquet 数据表保持一致。
    close_sheet: str = "收盘价"
    return_sheet: str = "涨跌幅"
    plain_sheet: str = "平价"
    premium_sheet: str = "转股溢价率"
    turnover_sheet: str = "换手率"
    balance_sheet: str = "余额"
    pure_bond_value_sheet: str = "纯债价值"
    amount_sheet: str = "成交额"
    remain_year_sheet: str = "剩余期限"
    stock_vol_sheet: str = "正股20日波动率"
    redeem_days_sheet: str = "赎回累计天数"
    downward_days_sheet: str = "下修累计天数"
    implied_vol_sheet: str = "隐含波动率"
    stock_close_sheet: str = "正股收盘价"
    expma5_sheet: str = "EXPMA5"
    expma10_sheet: str = "EXPMA10"
    expma20_sheet: str = "EXPMA20"
    stock_bond_premium_sheet: str = "平价底价溢价率"
    trade_status_sheet: str = "交易状态"
    total_sheet: str = "总表"
    benchmark_sheet: str = "指数"
    # 净值曲线中额外输出的基准指数名称。
    benchmark_names: tuple[str, ...] = ("万得全A", "转债指数")
    # 回测完成并通过审计后，将八个组合净值按组合名写入 parquet 的“指数”表。
    write_strategy_nav_to_parquet: bool = True

    # 输出路径和文件名；output_dir_template 支持 {end_date} 占位符，取回测截止最新交易日。
    output_dir_template: str = str(
        WORKSPACE_ROOT / "runs" / "research" / "策略回测{end_date}"
    )
    output_file_name: str = "回测合集.xlsx"
    # 回报拆解直接使用当前数据源计算，不读取或调用其它周报文件。
    decomposition_history_start_date: str = "2017-01-01"
    decomposition_output_start_date: str = "2018-01-01"
    decomposition_rolling_window: int = 20


_PARQUET_MEM_CACHE: dict[tuple[str, str], dict[str, pd.DataFrame]] = {}


def _parquet_dir_fingerprint(input_root: str) -> str:
    """根据 parquet 目录内文件路径、修改时间和大小生成缓存指纹。"""
    file_info: list[str] = []
    for root, _, files in os.walk(input_root):
        for file_name in sorted(files):
            if not file_name.endswith(".parquet"):
                continue
            file_path = os.path.join(root, file_name)
            stat = os.stat(file_path)
            file_info.append(f"{file_path}|{stat.st_mtime_ns}|{stat.st_size}")
    return hashlib.md5("|".join(file_info).encode("utf-8")).hexdigest()


def _parse_parquet_date_column(col) -> pd.Timestamp:
    if isinstance(col, pd.Timestamp):
        return col
    if isinstance(col, datetime):
        return pd.Timestamp(col)
    if isinstance(col, date):
        return pd.Timestamp(col)
    if isinstance(col, np.datetime64):
        return pd.Timestamp(col)
    if not isinstance(col, str):
        return pd.NaT
    text = col.strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return pd.Timestamp(datetime.strptime(text, fmt))
        except ValueError:
            pass
    if re.match(r"^\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}(\.\d+)?$", text):
        return pd.to_datetime(text, errors="coerce")
    return pd.NaT


def read_original_data_from_parquet(input_root: str = "data/转债个券历史序列") -> dict[str, pd.DataFrame]:
    return _read_standard_parquet(input_root)
    """读取本地 parquet 底稿目录，还原为 {sheet_name: DataFrame}。"""
    meta_cols = {"__sheet_name", "__row_id", "__date"}
    original_data_parts: dict[str, list[pd.DataFrame]] = {}
    year_entries = [
        entry for entry in sorted(os.listdir(input_root))
        if entry.isdigit() and os.path.isdir(os.path.join(input_root, entry))
    ]

    for year in tqdm(year_entries, desc="Reading Parquet", unit="year"):
        year_path = os.path.join(input_root, year)
        for file_name in sorted(os.listdir(year_path)):
            if not file_name.endswith(".parquet"):
                continue
            file_path = os.path.join(year_path, file_name)
            df = pd.read_parquet(file_path)
            if "__sheet_name" not in df.columns or "__row_id" not in df.columns:
                continue
            parsed_map = {col: _parse_parquet_date_column(col) for col in df.columns}
            date_cols = [col for col, ts in parsed_map.items() if pd.notna(ts) and col not in meta_cols]
            if not date_cols:
                continue
            for sheet_name, sub in df.groupby("__sheet_name", sort=False):
                if pd.isna(sheet_name):
                    continue
                wide_part = sub[["__row_id", *date_cols]].copy()
                wide_part = wide_part.groupby("__row_id", as_index=True).first()
                wide_part.columns = [parsed_map[c] for c in wide_part.columns]
                wide_part = wide_part.sort_index(axis=1)
                original_data_parts.setdefault(str(sheet_name), []).append(wide_part)

    original_data: dict[str, pd.DataFrame] = {}
    for sheet_name, parts in original_data_parts.items():
        merged = pd.concat(parts, axis=1)
        merged = merged.loc[:, ~merged.columns.duplicated(keep="first")]
        original_data[sheet_name] = merged.sort_index(axis=1)

    special_dir = os.path.join(input_root, "_special")
    if os.path.isdir(special_dir):
        legacy_file_name = "基础数据.parquet"
        special_files = sorted(
            file_name for file_name in os.listdir(special_dir)
            if file_name.endswith(".parquet") and file_name != legacy_file_name
        )

        def finalize_special(sub: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
            if "__sheet_name" in sub.columns:
                sub = sub.drop(columns=["__sheet_name"])
            sub = sub.set_index("__row_id")
            sub.index.name = None
            has_monthly_date_cols = sheet_name in original_data and any(
                pd.notna(_parse_parquet_date_column(c)) for c in original_data[sheet_name].columns
            )
            if not has_monthly_date_cols:
                date_like_cols = [c for c in sub.columns if pd.notna(_parse_parquet_date_column(c))]
                if date_like_cols:
                    sub = sub.drop(columns=date_like_cols)
            all_nan_cols = [c for c in sub.columns if sub[c].isna().all()]
            if all_nan_cols:
                sub = sub.drop(columns=all_nan_cols)
            return sub

        for file_name in special_files:
            file_path = os.path.join(special_dir, file_name)
            try:
                special_df = pd.read_parquet(file_path)
            except Exception as exc:
                print(f"[read] 无法读取 special 文件 {file_path}: {exc}")
                continue
            if "__row_id" not in special_df.columns:
                continue
            sheet_name_values = (
                special_df["__sheet_name"].dropna().unique().tolist()
                if "__sheet_name" in special_df.columns else []
            )
            if len(sheet_name_values) == 1:
                sheet_name = str(sheet_name_values[0])
                original_data[sheet_name] = finalize_special(special_df, sheet_name)
            elif len(sheet_name_values) == 0:
                sheet_name = os.path.splitext(file_name)[0]
                original_data[sheet_name] = finalize_special(special_df, sheet_name)
            else:
                for sheet_name, group in special_df.groupby("__sheet_name", sort=False):
                    if pd.notna(sheet_name) and str(sheet_name) not in original_data:
                        original_data[str(sheet_name)] = finalize_special(group, str(sheet_name))

        legacy_path = os.path.join(special_dir, legacy_file_name)
        if os.path.exists(legacy_path):
            try:
                legacy_df = pd.read_parquet(legacy_path)
            except Exception as exc:
                print(f"[read] 无法读取旧 special 合并文件 {legacy_path}: {exc}")
                legacy_df = None
            if legacy_df is not None and "__sheet_name" in legacy_df.columns and "__row_id" in legacy_df.columns:
                for sheet_name, group in legacy_df.groupby("__sheet_name", sort=False):
                    sheet_name = str(sheet_name)
                    if sheet_name in original_data and not original_data[sheet_name].empty:
                        continue
                    original_data[sheet_name] = finalize_special(group, sheet_name)

    print(f"Parquet 读取并还原完成，共恢复 {len(original_data)} 个数据表。")
    return original_data


def load_parquet_with_cache(
    input_root: str = "data/转债个券历史序列",
    force_refresh: bool = False,
    cache_dir: str = "tmp/cache/parquet",
) -> dict[str, pd.DataFrame]:
    fp = _parquet_dir_fingerprint(input_root)
    mem_key = (input_root, fp)
    cache_file = os.path.join(cache_dir, f"parquet_{fp}.pkl")
    if not force_refresh:
        cached = _PARQUET_MEM_CACHE.get(mem_key)
        if cached is not None:
            print(f"[cache] parquet memory hit: {input_root}")
            return cached
        if os.path.exists(cache_file):
            with open(cache_file, "rb") as f:
                cached = pickle.load(f)
            _PARQUET_MEM_CACHE[mem_key] = cached
            print(f"[cache] parquet disk hit: {cache_file}")
            return cached
    print(f"[cache] parquet 缓存未命中，直接读取 parquet 目录：{input_root}")
    data = read_original_data_from_parquet(input_root)
    _PARQUET_MEM_CACHE[mem_key] = data
    return data


def load_original_data(
    parquet_root: str = "data/转债个券历史序列",
    force_refresh: bool = False,
) -> dict[str, pd.DataFrame]:
    """读取唯一底稿数据源：标准 parquet 目录。"""
    if not os.path.isdir(parquet_root):
        raise FileNotFoundError(f"未找到 parquet 底稿目录: {parquet_root}")
    print(f"[source] 使用 parquet 数据源: {parquet_root}")
    return load_parquet_with_cache(parquet_root, force_refresh=force_refresh)


@dataclass
class ModelFit:
    amplitude: float
    scale: float
    power: float
    floor: float
    beta: np.ndarray
    feature_stats: Dict[str, Dict[str, float]]
    sample_count: int
    base_r2: float
    corrected_r2: float
    base_rmse: float
    corrected_rmse: float


def _normalize_wide(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    parsed = pd.to_datetime(pd.Index(out.columns), errors="coerce")
    out.columns = [pd.Timestamp(v) if pd.notna(v) else raw for raw, v in zip(out.columns, parsed)]
    out.index = out.index.map(lambda x: "" if pd.isna(x) else str(x).strip())
    return out


def _date_columns(df: pd.DataFrame) -> pd.DatetimeIndex:
    return pd.DatetimeIndex([c for c in df.columns if isinstance(c, pd.Timestamp)]).sort_values()


def power_decay_with_floor(
    x: np.ndarray | float,
    amplitude: float,
    scale: float,
    power: float,
    floor: float,
) -> np.ndarray:
    x_arr = np.asarray(x, dtype=float)
    base = 1.0 + (x_arr - RESIDUAL_POWER_ANCHOR) / scale
    with np.errstate(invalid="ignore", divide="ignore", over="ignore"):
        return floor + amplitude * np.power(base, -power)


def _feature_design(
    work: pd.DataFrame,
    stats: Optional[Dict[str, Dict[str, float]]] = None,
) -> tuple[pd.DataFrame, Dict[str, Dict[str, float]]]:
    values = work[RESIDUAL_FEATURES].apply(pd.to_numeric, errors="coerce").copy()
    if stats is None:
        stats = {}
        for col in RESIDUAL_FEATURES:
            median = values[col].median()
            median = float(median) if pd.notna(median) else 0.0
            filled = values[col].fillna(median)
            mean = float(filled.mean()) if len(filled) else 0.0
            std = float(filled.std(ddof=0)) if len(filled) else 0.0
            stats[col] = {"median": median, "mean": mean, "std": std if std > 0 else 0.0}

    design = pd.DataFrame(index=work.index)
    design["截距项"] = 1.0
    for col in RESIDUAL_FEATURES:
        col_stats = stats[col]
        filled = values[col].fillna(col_stats["median"])
        design[col] = (
            (filled - col_stats["mean"]) / col_stats["std"]
            if col_stats["std"] > 0
            else 0.0
        )
    return design, stats


def _fit_model(cross_section: pd.DataFrame, config: BacktestConfig) -> ModelFit:
    work = cross_section.copy()
    work["余额_log"] = np.log1p(pd.to_numeric(work["余额"], errors="coerce").clip(lower=0))
    numeric_cols = ["平价", "转股溢价率", "换手率", *RESIDUAL_FEATURES]
    for col in numeric_cols:
        work[col] = pd.to_numeric(work[col], errors="coerce")
    # 与现有残差修正幂衰减模型保持同口径：除了平价和溢价率，其它 0 值视作缺失并由中位数填充。
    work = work.replace("", np.nan).replace(0, np.nan)
    work = work.dropna(subset=["平价", "转股溢价率", "换手率"])
    work = work[
        work["平价"].gt(RESIDUAL_POWER_LOWER)
        & work["平价"].lt(RESIDUAL_POWER_UPPER)
        & work["换手率"].lt(config.fit_turnover_upper)
    ].copy()
    if len(work) > 0:
        low = work["转股溢价率"].quantile(config.fit_premium_low_quantile)
        high = work["转股溢价率"].quantile(config.fit_premium_high_quantile)
        work = work[work["转股溢价率"].gt(low) & work["转股溢价率"].lt(high)]
    if len(work) < config.minimum_fit_samples:
        raise ValueError(f"有效拟合样本不足: {len(work)} < {config.minimum_fit_samples}")

    x = work["平价"].to_numpy(dtype=float)
    y = work["转股溢价率"].to_numpy(dtype=float)
    floor0 = float(np.clip(np.nanpercentile(y, 5), 0, 1))
    amplitude0 = float(max(np.nanpercentile(y, 95) - floor0, 1))
    popt, _ = curve_fit(
        lambda x_data, amplitude, scale, power, floor: power_decay_with_floor(
            x_data, amplitude, scale, power, floor
        ),
        x,
        y,
        p0=[amplitude0, 30.0, 2.0, floor0],
        bounds=([0, 1, 0.05, 0], [np.inf, 500, 20, 1]),
        maxfev=30000,
    )
    amplitude, scale, power, floor = (float(v) for v in popt)
    base_prediction = power_decay_with_floor(x, amplitude, scale, power, floor)
    design, stats = _feature_design(work)
    beta = np.linalg.lstsq(design.to_numpy(dtype=float), y - base_prediction, rcond=None)[0]
    corrected_prediction = base_prediction + design.to_numpy(dtype=float) @ beta

    sst = float(np.sum(np.square(y - y.mean())))
    base_sse = float(np.sum(np.square(y - base_prediction)))
    corrected_sse = float(np.sum(np.square(y - corrected_prediction)))
    return ModelFit(
        amplitude=amplitude,
        scale=scale,
        power=power,
        floor=floor,
        beta=beta,
        feature_stats=stats,
        sample_count=int(len(work)),
        base_r2=float(1 - base_sse / sst) if sst > 0 else np.nan,
        corrected_r2=float(1 - corrected_sse / sst) if sst > 0 else np.nan,
        base_rmse=float(np.sqrt(base_sse / len(y))),
        corrected_rmse=float(np.sqrt(corrected_sse / len(y))),
    )


def _predict_premium(cross_section: pd.DataFrame, model: ModelFit) -> pd.Series:
    work = cross_section.copy()
    work["余额_log"] = np.log1p(pd.to_numeric(work["余额"], errors="coerce").clip(lower=0))
    work[RESIDUAL_FEATURES] = work[RESIDUAL_FEATURES].replace(0, np.nan)
    parity = pd.to_numeric(work["平价"], errors="coerce")
    design, _ = _feature_design(work, stats=model.feature_stats)
    base = power_decay_with_floor(
        parity.to_numpy(dtype=float), model.amplitude, model.scale, model.power, model.floor
    )
    prediction = base + design.to_numpy(dtype=float) @ model.beta
    return pd.Series(prediction, index=work.index, name="理论转股溢价率")


def _cross_section(mats: Dict[str, pd.DataFrame], dt: pd.Timestamp) -> pd.DataFrame:
    return pd.DataFrame({name: frame[dt] for name, frame in mats.items()})


def _equal_weights(holdings: list[str]) -> Dict[str, float]:
    """为调仓后的持仓生成等权权重。"""
    if not holdings:
        return {}
    weight = 1.0 / len(holdings)
    return {code: weight for code in holdings}


def _one_way_turnover(old_weights: Dict[str, float], new_weights: Dict[str, float]) -> float:
    """按资产与现金权重变化计算单边换手率。"""
    codes = set(old_weights) | set(new_weights)
    asset_change = sum(abs(new_weights.get(code, 0.0) - old_weights.get(code, 0.0)) for code in codes)
    old_cash = max(1.0 - sum(old_weights.values()), 0.0)
    new_cash = max(1.0 - sum(new_weights.values()), 0.0)
    return float(0.5 * (asset_change + abs(new_cash - old_cash)))


def _portfolio_return_and_drift(
    weights: Dict[str, float],
    day_returns: pd.Series,
) -> tuple[float, Dict[str, float]]:
    """计算组合当日收益，并让权重按收盘后市值自然漂移；缺失收益按 0。"""
    if not weights:
        return 0.0, {}
    codes = list(weights)
    returns = pd.to_numeric(day_returns.reindex(codes), errors="coerce").fillna(0.0).astype("float64")
    old_weights = pd.Series(weights, dtype="float64").reindex(codes).fillna(0.0)
    gross_return = float((old_weights * returns).sum())
    closing_values = old_weights * (1.0 + returns)
    total_closing_value = float(closing_values.sum())
    if not np.isfinite(total_closing_value) or total_closing_value <= 0:
        return gross_return, dict(weights)
    drifted_weights = closing_values / total_closing_value
    return gross_return, {str(code): float(weight) for code, weight in drifted_weights.items()}


def _clean_total_dates(total_df: Optional[pd.DataFrame], codes: pd.Index) -> Dict[str, pd.Series]:
    empty = pd.Series(pd.NaT, index=codes, dtype="datetime64[ns]")
    result = {"上市日期": empty.copy(), "最后交易日": empty.copy(), "赎回公告日": empty.copy()}
    if total_df is None or total_df.empty:
        return result
    total = total_df.copy()
    total.index = total.index.map(lambda x: "" if pd.isna(x) else str(x).strip())
    for col in result:
        if col in total.columns:
            result[col] = pd.to_datetime(total.reindex(codes)[col], errors="coerce")
    return result


def _max_drawdown(nav: pd.Series) -> float:
    clean = pd.to_numeric(nav, errors="coerce").dropna()
    if clean.empty:
        return np.nan
    return float((clean / clean.cummax() - 1.0).min())


def _daily_risk_free_returns(
    risk_free_rates: Optional[pd.Series],
    dates: pd.Index,
) -> pd.Series:
    """将逐日年化无风险收益率换算为日度收益率；输入单位为小数。"""
    if risk_free_rates is None:
        return pd.Series(0.0, index=dates, dtype="float64")
    annual_rates = pd.to_numeric(risk_free_rates, errors="coerce").reindex(dates).ffill()
    annual_rates = annual_rates.where(annual_rates.gt(-1.0))
    return np.power(1.0 + annual_rates, 1.0 / 252.0) - 1.0


def _series_metrics(nav: pd.Series, risk_free_rates: Optional[pd.Series] = None) -> dict:
    clean = pd.to_numeric(nav, errors="coerce").dropna()
    if len(clean) < 2:
        return {"区间收益率": np.nan, "年化收益率": np.nan, "最大回撤": np.nan, "年化波动率": np.nan, "夏普比率": np.nan}
    daily = clean.pct_change().dropna()
    elapsed_days = max((clean.index[-1] - clean.index[0]).days, 1)
    total_return = float(clean.iloc[-1] / clean.iloc[0] - 1.0)
    annual_return = float((clean.iloc[-1] / clean.iloc[0]) ** (365.25 / elapsed_days) - 1.0)
    volatility = float(daily.std(ddof=1) * np.sqrt(252)) if len(daily) > 1 else np.nan
    daily_risk_free = _daily_risk_free_returns(risk_free_rates, daily.index)
    excess_returns = (daily - daily_risk_free).dropna()
    sharpe = (
        float(excess_returns.mean() / daily.loc[excess_returns.index].std(ddof=1) * np.sqrt(252))
        if len(excess_returns) > 1 and daily.loc[excess_returns.index].std(ddof=1) > 0
        else np.nan
    )
    return {
        "区间收益率": total_return,
        "年化收益率": annual_return,
        "最大回撤": _max_drawdown(clean),
        "年化波动率": volatility,
        "夏普比率": float(sharpe),
    }


def _calmar_ratio(annual_return: float, max_drawdown: float) -> float:
    if pd.isna(annual_return) or pd.isna(max_drawdown) or max_drawdown >= 0:
        return np.nan
    return float(annual_return / abs(max_drawdown))


def _period_nav_with_base(nav: pd.Series, start_date: pd.Timestamp, end_date: pd.Timestamp) -> pd.Series:
    clean = pd.to_numeric(nav, errors="coerce").dropna()
    if clean.empty:
        return clean
    clean = clean.loc[clean.index <= end_date]
    if clean.empty:
        return clean
    before_start = clean.loc[clean.index < start_date]
    in_period = clean.loc[clean.index >= start_date]
    if before_start.empty:
        return in_period
    return pd.concat([before_start.iloc[[-1]], in_period])


def _summary_nav_columns(nav_df: pd.DataFrame) -> list[str]:
    """按固定策略顺序取净值列，并追加基准指数列。"""
    portfolio_cols = [
        f"净值_{name}" for name in PORTFOLIO_NAMES
        if f"净值_{name}" in nav_df.columns
    ]
    benchmark_cols = [c for c in nav_df.columns if str(c).startswith("基准净值_")]
    return portfolio_cols + benchmark_cols


def _build_summary_table(
    nav_df: pd.DataFrame,
    config: BacktestConfig,
    risk_free_rates: Optional[pd.Series] = None,
) -> pd.DataFrame:
    """生成最新运行后的本周、本月与年初至今策略/指数绩效摘要。"""
    nav_cols = _summary_nav_columns(nav_df)
    if nav_df.empty or not nav_cols:
        return pd.DataFrame()

    end_date = pd.Timestamp(nav_df.index.max())
    week_start = end_date - pd.Timedelta(days=end_date.weekday())
    month_start = pd.Timestamp(year=end_date.year, month=end_date.month, day=1)
    year_start = pd.Timestamp(year=end_date.year, month=1, day=1)
    rows = []
    for col in nav_cols:
        name = str(col).replace("基准净值_", "").replace("净值_", "")
        week_metrics = _series_metrics(
            _period_nav_with_base(nav_df[col], week_start, end_date),
            risk_free_rates,
        )
        month_metrics = _series_metrics(
            _period_nav_with_base(nav_df[col], month_start, end_date),
            risk_free_rates,
        )
        ytd_metrics = _series_metrics(
            _period_nav_with_base(nav_df[col], year_start, end_date),
            risk_free_rates,
        )
        rows.append({
            "策略": name,
            "本周收益率": week_metrics["区间收益率"],
            "本周年化收益率": week_metrics["年化收益率"],
            "本周最大回撤": week_metrics["最大回撤"],
            "本周卡玛比率": _calmar_ratio(week_metrics["年化收益率"], week_metrics["最大回撤"]),
            "本月收益率": month_metrics["区间收益率"],
            "本月年化收益率": month_metrics["年化收益率"],
            "本月最大回撤": month_metrics["最大回撤"],
            "本月卡玛比率": _calmar_ratio(month_metrics["年化收益率"], month_metrics["最大回撤"]),
            "年初至今收益率": ytd_metrics["区间收益率"],
            "年初至今年化收益率": ytd_metrics["年化收益率"],
            "年初至今最大回撤": ytd_metrics["最大回撤"],
            "年初至今夏普比率": ytd_metrics["夏普比率"],
            "年初至今卡玛比率": _calmar_ratio(ytd_metrics["年化收益率"], ytd_metrics["最大回撤"]),
        })
    return pd.DataFrame(rows)


def _plot_week_performance(summary_table: pd.DataFrame, png_path: Path, end_date: pd.Timestamp) -> None:
    """输出本周各策略及指数的收益率、最大回撤对比图。"""
    if summary_table is None or summary_table.empty:
        return
    required = {"策略", "本周收益率", "本周最大回撤"}
    if not required.issubset(summary_table.columns):
        return

    week_start = end_date - pd.Timedelta(days=end_date.weekday())
    names = summary_table["策略"].astype(str).tolist()
    returns = pd.to_numeric(summary_table["本周收益率"], errors="coerce").fillna(0.0).to_numpy()
    drawdowns = pd.to_numeric(summary_table["本周最大回撤"], errors="coerce").fillna(0.0).to_numpy()

    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei"]
    plt.rcParams["axes.unicode_minus"] = False
    fig, ax = plt.subplots(figsize=(14, 7))
    x = np.arange(len(names))
    width = 0.36
    bars_ret = ax.bar(x - width / 2, returns * 100, width, label="本周收益率", color="#0262BA", alpha=0.9)
    bars_dd = ax.bar(x + width / 2, drawdowns * 100, width, label="本周最大回撤", color="#E6121B", alpha=0.85)

    def _annotate(bars) -> None:
        for bar in bars:
            height = bar.get_height()
            if abs(height) < 1e-9:
                label = "0.00%"
                y = 0.08 if height >= 0 else -0.08
                va = "bottom" if height >= 0 else "top"
            else:
                label = f"{height:.2f}%"
                y = height + (0.12 if height >= 0 else -0.12)
                va = "bottom" if height >= 0 else "top"
            ax.text(bar.get_x() + bar.get_width() / 2, y, label, ha="center", va=va, fontsize=8)

    _annotate(bars_ret)
    _annotate(bars_dd)
    ax.axhline(0, color="#666666", linewidth=0.8)
    ax.set_xticks(x)
    ax.set_xticklabels(names, rotation=25, ha="right")
    ax.set_ylabel("百分比（%）")
    ax.set_title(f"本周策略与指数：收益率 / 最大回撤（{week_start:%Y-%m-%d} ~ {end_date:%Y-%m-%d}）")
    ax.grid(axis="y", alpha=0.3)
    ax.legend(loc="best")
    fig.tight_layout()
    fig.savefig(png_path, dpi=180)
    plt.close(fig)


def _fmt_pct(value: float, digits: int = 2) -> str:
    if pd.isna(value):
        return "NA"
    return f"{float(value) * 100:.{digits}f}%"


def _weighted_decomposition_group(sample: pd.DataFrame, label: object | None = None) -> dict | None:
    """按期初余额汇总转债收益及三类回报贡献。"""
    valid = pd.to_numeric(sample.get("期初余额"), errors="coerce").gt(0)
    if not valid.any():
        return None
    work = sample.loc[valid]
    weights = pd.to_numeric(work["期初余额"], errors="coerce")
    row: dict[str, object] = {}
    if label is not None:
        row["分类"] = label
    for col in DECOMP_COLUMNS[1:]:
        values = pd.to_numeric(work[col], errors="coerce")
        mask = values.notna() & weights.notna() & weights.gt(0)
        row[col] = float((values[mask] * weights[mask]).sum() / weights[mask].sum()) if mask.any() else np.nan
    return row


def _build_return_decomposition_outputs(
    observations: pd.DataFrame,
    config: BacktestConfig,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """生成滚动20日回报拆解及最新一期行业、平价和转债类型拆解。"""
    empty_ts = pd.DataFrame(columns=DECOMP_COLUMNS)
    empty_group = pd.DataFrame(columns=DECOMP_GROUP_COLUMNS)
    if observations.empty:
        return empty_ts, empty_group, empty_group.copy(), empty_group.copy()

    observations = observations.sort_values(["转债代码", "日期"]).reset_index(drop=True)
    market_dates = pd.DatetimeIndex(sorted(pd.to_datetime(observations["日期"]).dropna().unique()))
    factor_cols = ["纯债价值变动", "平价变动", "转股溢价率变动"]
    window = int(config.decomposition_rolling_window)
    output_start = pd.Timestamp(config.decomposition_output_start_date)
    rows: list[dict[str, object]] = []

    for code, group in observations.groupby("转债代码", sort=False):
        frame = group.sort_values("日期").drop_duplicates("日期", keep="last").set_index("日期").reindex(market_dates)
        if len(frame) < window:
            continue
        x = frame[factor_cols].astype(float).to_numpy()
        y = frame["转债日收益率"].astype(float).to_numpy()
        price = frame["收盘价"].astype(float).to_numpy()
        previous_price = frame["前收盘价"].astype(float).to_numpy()
        row_valid = np.isfinite(y) & np.all(np.isfinite(x), axis=1)

        x_filled = np.where(np.isfinite(x), x, 0.0)
        y_filled = np.where(np.isfinite(y), y, 0.0)
        xtx_terms = np.einsum("ni,nj->nij", x_filled, x_filled)
        xty_terms = x_filled * y_filled[:, None]
        xsum_terms = x_filled
        c_xtx = np.concatenate([np.zeros((1, 3, 3)), np.cumsum(xtx_terms, axis=0)], axis=0)
        c_xty = np.concatenate([np.zeros((1, 3)), np.cumsum(xty_terms, axis=0)], axis=0)
        c_xsum = np.concatenate([np.zeros((1, 3)), np.cumsum(xsum_terms, axis=0)], axis=0)
        c_valid = np.concatenate([[0], np.cumsum(row_valid.astype(int))])

        # 同一转债的全部滚动窗口一次性求解，避免逐日 Python 回归造成长时间等待。
        end_positions = np.arange(window - 1, len(frame), dtype=int)
        start_positions = end_positions - window + 1
        valid_windows = (
            (market_dates[end_positions] >= output_start)
            & ((c_valid[end_positions + 1] - c_valid[start_positions]) == window)
        )
        if not valid_windows.any():
            continue
        end_positions = end_positions[valid_windows]
        start_positions = start_positions[valid_windows]
        xtx = c_xtx[end_positions + 1] - c_xtx[start_positions]
        xty = c_xty[end_positions + 1] - c_xty[start_positions]
        xsum = c_xsum[end_positions + 1] - c_xsum[start_positions]
        beta = np.einsum("nij,nj->ni", np.linalg.pinv(xtx), xty)
        contributions = beta * xsum
        start_prices = previous_price[start_positions]
        end_prices = price[end_positions]
        start_balances = pd.to_numeric(frame["期初余额"], errors="coerce").to_numpy()[start_positions]
        usable = (
            np.isfinite(start_prices) & (start_prices > 0)
            & np.isfinite(end_prices)
            & np.isfinite(start_balances) & (start_balances > 0)
        )
        if not usable.any():
            continue
        end_positions = end_positions[usable]
        start_positions = start_positions[usable]
        contributions = contributions[usable]
        start_prices = start_prices[usable]
        end_prices = end_prices[usable]
        start_balances = start_balances[usable]
        code_rows = pd.DataFrame({
            "日期": market_dates[end_positions],
            "转债代码": code,
            "转债收益率": end_prices / start_prices - 1.0,
            "债券贡献": contributions[:, 0],
            "正股贡献": contributions[:, 1],
            "估值贡献": contributions[:, 2],
            "期初余额": start_balances,
            "申万行业": frame["申万行业"].to_numpy()[start_positions],
            "期初平价": pd.to_numeric(frame["期初平价"], errors="coerce").to_numpy()[start_positions],
            "期初平价底价溢价率": pd.to_numeric(frame["期初平价底价溢价率"], errors="coerce").to_numpy()[start_positions],
        })
        rows.extend(code_rows.to_dict("records"))

    decomposition = pd.DataFrame(rows)
    if decomposition.empty:
        return empty_ts, empty_group, empty_group.copy(), empty_group.copy()

    time_rows: list[dict[str, object]] = []
    for day, sample in decomposition.groupby("日期", sort=True):
        row = _weighted_decomposition_group(sample)
        if row is not None:
            row["日期"] = day
            time_rows.append(row)
    timeseries = pd.DataFrame(time_rows)[DECOMP_COLUMNS].sort_values("日期", ascending=False)

    latest = decomposition[decomposition["日期"].eq(decomposition["日期"].max())].copy()
    industry_rows = []
    for industry, sample in latest.dropna(subset=["申万行业"]).groupby("申万行业", sort=False):
        row = _weighted_decomposition_group(sample, industry)
        if row is not None:
            industry_rows.append(row)
    industry = pd.DataFrame(industry_rows, columns=DECOMP_GROUP_COLUMNS)
    if not industry.empty:
        industry = industry.sort_values("转债收益率", ascending=False)

    parity_bins = [
        ("100及以下", latest["期初平价"].le(100)),
        ("100-110", latest["期初平价"].gt(100) & latest["期初平价"].le(110)),
        ("110-120", latest["期初平价"].gt(110) & latest["期初平价"].le(120)),
        ("120-130", latest["期初平价"].gt(120) & latest["期初平价"].le(130)),
        ("130以上", latest["期初平价"].gt(130)),
    ]
    parity_rows = []
    for label, mask in parity_bins:
        row = _weighted_decomposition_group(latest[mask.fillna(False)], label)
        parity_rows.append(row if row is not None else {"分类": label, **{col: np.nan for col in DECOMP_GROUP_COLUMNS[1:]}})
    parity = pd.DataFrame(parity_rows, columns=DECOMP_GROUP_COLUMNS)

    floor_premium = latest["期初平价底价溢价率"]
    type_bins = [
        ("偏股型", floor_premium.gt(20)),
        ("平衡型", floor_premium.gt(-20) & floor_premium.lt(20)),
        ("偏债型", floor_premium.lt(-20)),
    ]
    type_rows = []
    for label, mask in type_bins:
        row = _weighted_decomposition_group(latest[mask.fillna(False)], label)
        type_rows.append(row if row is not None else {"分类": label, **{col: np.nan for col in DECOMP_GROUP_COLUMNS[1:]}})
    bond_type = pd.DataFrame(type_rows, columns=DECOMP_GROUP_COLUMNS)
    return timeseries, industry, parity, bond_type


def _calculate_integrated_return_decomposition(
    data: dict[str, pd.DataFrame],
    config: BacktestConfig,
    end_date: pd.Timestamp,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """仅使用当前回测数据源计算回报拆解，不依赖任何外部成品文件。"""
    sheet_map = {
        "收盘价": config.close_sheet,
        "涨跌幅": config.return_sheet,
        "纯债价值": config.pure_bond_value_sheet,
        "平价": config.plain_sheet,
        "转股溢价率": config.premium_sheet,
        "余额": config.balance_sheet,
        "成交额": config.amount_sheet,
        "剩余期限": config.remain_year_sheet,
        "平价底价溢价率": config.stock_bond_premium_sheet,
    }
    missing = [sheet for sheet in sheet_map.values() if sheet not in data]
    if missing:
        raise KeyError(f"回报拆解缺少必要原始数据表: {missing}")

    normalized = {name: _normalize_wide(data[sheet]) for name, sheet in sheet_map.items()}
    common_codes: Optional[pd.Index] = None
    common_dates: Optional[pd.DatetimeIndex] = None
    for frame in normalized.values():
        common_codes = frame.index if common_codes is None else common_codes.intersection(frame.index)
        frame_dates = _date_columns(frame)
        common_dates = frame_dates if common_dates is None else common_dates.intersection(frame_dates)
    assert common_codes is not None and common_dates is not None
    dates = common_dates[
        (common_dates >= pd.Timestamp(config.decomposition_history_start_date))
        & (common_dates <= pd.Timestamp(end_date))
    ]
    mats = {
        # pandas 3.x 的 to_numeric 可能返回可空 Float64；显式转为 NumPy 兼容的
        # float64，使缺失值统一为 np.nan，避免 np.isfinite 遇到 pd.NA 报错。
        name: frame.reindex(index=common_codes, columns=dates)
        .apply(pd.to_numeric, errors="coerce")
        .astype("float64")
        for name, frame in normalized.items()
    }

    total = data.get(config.total_sheet, pd.DataFrame()).copy()
    total.index = total.index.map(lambda value: "" if pd.isna(value) else str(value).strip())
    total = total.reindex(common_codes)
    listing = pd.to_datetime(total["上市日期"], errors="coerce") if "上市日期" in total else pd.Series(pd.NaT, index=common_codes)
    last_trade = pd.to_datetime(total["最后交易日"], errors="coerce") if "最后交易日" in total else pd.Series(pd.NaT, index=common_codes)
    industry = total["申万行业"] if "申万行业" in total else pd.Series(pd.NA, index=common_codes, dtype="object")

    previous_names = ["收盘价", "纯债价值", "平价", "转股溢价率", "余额", "平价底价溢价率"]
    previous = {name: mats[name].ffill(axis=1).shift(axis=1) for name in previous_names}
    date_array = dates.to_numpy(dtype="datetime64[ns]")
    listing_start = listing.sub(pd.Timedelta(days=4)).to_numpy(dtype="datetime64[ns]")[:, None]
    last_end = last_trade.add(BDay(1)).to_numpy(dtype="datetime64[ns]")[:, None]
    start_ok = (listing_start <= date_array[None, :]) | mats["成交额"].notna().to_numpy()
    end_ok = last_end >= date_array[None, :]
    stale_nontrading = mats["成交额"].isna().to_numpy() & mats["剩余期限"].notna().to_numpy()
    active = (
        start_ok & end_ok & ~stale_nontrading
        & mats["余额"].gt(0).to_numpy()
        & mats["收盘价"].notna().to_numpy()
    )

    bond_return = mats["涨跌幅"] / 100.0
    bond_value_change = mats["纯债价值"] / previous["纯债价值"] - 1.0
    parity_change = mats["平价"] / previous["平价"] - 1.0
    premium_change = (mats["转股溢价率"] - previous["转股溢价率"]) / 100.0
    valid = (
        active
        & bond_return.notna().to_numpy()
        & mats["收盘价"].notna().to_numpy()
        & previous["收盘价"].gt(0).to_numpy()
        & mats["纯债价值"].notna().to_numpy()
        & previous["纯债价值"].gt(0).to_numpy()
        & mats["平价"].notna().to_numpy()
        & previous["平价"].gt(0).to_numpy()
        & mats["转股溢价率"].notna().to_numpy()
        & previous["转股溢价率"].notna().to_numpy()
        & previous["余额"].gt(0).to_numpy()
        & np.isfinite(bond_value_change.to_numpy())
        & np.isfinite(parity_change.to_numpy())
        & np.isfinite(premium_change.to_numpy())
    )
    row_pos, col_pos = np.where(valid)
    code_array = common_codes.to_numpy(dtype=object)
    industry_array = industry.to_numpy(dtype=object)
    observations = pd.DataFrame({
        "日期": date_array[col_pos],
        "转债代码": code_array[row_pos].astype(str),
        "转债日收益率": bond_return.to_numpy()[row_pos, col_pos],
        "纯债价值变动": bond_value_change.to_numpy()[row_pos, col_pos],
        "平价变动": parity_change.to_numpy()[row_pos, col_pos],
        "转股溢价率变动": premium_change.to_numpy()[row_pos, col_pos],
        "收盘价": mats["收盘价"].to_numpy()[row_pos, col_pos],
        "前收盘价": previous["收盘价"].to_numpy()[row_pos, col_pos],
        "期初余额": previous["余额"].to_numpy()[row_pos, col_pos],
        "申万行业": industry_array[row_pos],
        "期初平价": previous["平价"].to_numpy()[row_pos, col_pos],
        "期初平价底价溢价率": previous["平价底价溢价率"].to_numpy()[row_pos, col_pos],
    })
    return _build_return_decomposition_outputs(observations, config)


def _latest_return_decomposition(timeseries: pd.DataFrame, end_date: pd.Timestamp) -> dict:
    """提取不晚于回测截止日的最新内置回报拆解。"""
    if timeseries.empty or "日期" not in timeseries.columns:
        return {"状态": "内置回报拆解为空"}
    work = timeseries.copy()
    work["日期"] = pd.to_datetime(work["日期"], errors="coerce")
    matched = work.dropna(subset=["日期"])
    matched = matched[matched["日期"].le(end_date)].sort_values("日期")
    if matched.empty:
        return {"状态": "内置回报拆解无可匹配日期"}
    row = matched.iloc[-1]
    return {
        "状态": "成功",
        "来源": "集成回测内置计算",
        "日期": pd.Timestamp(row["日期"]),
        **{col: float(pd.to_numeric(row.get(col), errors="coerce")) for col in DECOMP_COLUMNS[1:]},
    }


def _dominant_driver_text(decomp: dict) -> str:
    if decomp.get("状态") != "成功":
        return "回报拆解数据暂缺，策略归因主要依据区间净值表现。"
    bond = float(pd.to_numeric(decomp.get("债券贡献"), errors="coerce"))
    stock = float(pd.to_numeric(decomp.get("正股贡献"), errors="coerce"))
    valuation = float(pd.to_numeric(decomp.get("估值贡献"), errors="coerce"))
    components = {
        name: abs(value)
        for name, value in {"债券贡献": bond, "正股贡献": stock, "估值贡献": valuation}.items()
        if np.isfinite(value)
    }
    if not components:
        return "近20日回报拆解贡献项均不可用，策略差异主要依据区间净值表现判断。"
    dominant = max(components, key=components.get)
    if dominant == "估值贡献":
        if valuation > 0 and stock < 0:
            return "近20日转债收益主要由估值贡献驱动，估值修复对冲了正股端回撤压力。"
        if valuation > 0:
            return "近20日转债收益主要由估值贡献驱动，市场定价中风险偏好修复更为突出。"
        return "近20日转债收益主要受估值压缩拖累，转债相对正股的定价弹性阶段性受抑。"
    if dominant == "正股贡献":
        return "近20日转债收益主要由正股贡献驱动，权益方向和弹性暴露成为策略收益的核心来源。"
    return "近20日转债收益中债券贡献占比较高，票息、防御属性及价格保护对组合表现形成支撑。"


def _build_period_strategy_commentary(
    summary_table: pd.DataFrame,
    end_date: pd.Timestamp,
    decomp: dict,
    period: str,
) -> tuple[pd.DataFrame, str]:
    """生成固定格式的周度/月度转债策略表现点评。"""
    if period not in {"week", "month"}:
        raise ValueError(f"不支持的点评周期: {period}")
    is_week = period == "week"
    period_name = "本周" if is_week else "本月"
    period_text = "一周" if is_week else "当月"
    commentary_item = "本周策略点评" if is_week else "本月策略点评"
    return_col = "本周收益率" if is_week else "本月收益率"
    columns = ["项目", "内容"]
    if summary_table is None or summary_table.empty or return_col not in summary_table.columns:
        commentary = f"{period_name}策略点评暂缺：未获得有效的{period_name}回测收益表。"
        return pd.DataFrame([{"项目": commentary_item, "内容": commentary}], columns=columns), commentary

    work = summary_table.copy()
    work[return_col] = pd.to_numeric(work[return_col], errors="coerce")
    strategy_names = [name for name in PORTFOLIO_NAMES if name in set(work["策略"].astype(str))]
    strategy_returns = work[work["策略"].isin(strategy_names)].dropna(subset=[return_col])
    benchmark_returns = work[work["策略"].isin(["万得全A", "转债指数"])]

    top = strategy_returns.sort_values(return_col, ascending=False).head(3)
    bottom = strategy_returns.sort_values(return_col, ascending=True).head(3)
    top_text = "、".join(f"{row['策略']}{_fmt_pct(row[return_col])}" for _, row in top.iterrows())
    bottom_text = "、".join(f"{row['策略']}{_fmt_pct(row[return_col])}" for _, row in bottom.iterrows())
    cb_index = benchmark_returns.loc[benchmark_returns["策略"].eq("转债指数"), return_col]
    cb_index_value = float(cb_index.iloc[0]) if not cb_index.empty else np.nan

    equity_names = ["动量", "偏股"]
    defensive_names = ["低价", "偏债", "平衡", "双低", "低转股溢价率"]
    equity_avg = strategy_returns[strategy_returns["策略"].isin(equity_names)][return_col].mean()
    defensive_avg = strategy_returns[strategy_returns["策略"].isin(defensive_names)][return_col].mean()
    if pd.notna(equity_avg) and pd.notna(defensive_avg) and equity_avg > defensive_avg:
        style_text = (
            f"权益弹性类策略平均收益{_fmt_pct(equity_avg)}，高于偏防御类策略的"
            f"{_fmt_pct(defensive_avg)}，市场风格偏向权益弹性。"
        )
        dominant_style = "equity"
    elif pd.notna(equity_avg) and pd.notna(defensive_avg) and defensive_avg > equity_avg:
        style_text = (
            f"偏防御类策略平均收益{_fmt_pct(defensive_avg)}，高于权益弹性类策略的"
            f"{_fmt_pct(equity_avg)}，低价与债性保护阶段性占优。"
        )
        dominant_style = "defensive"
    else:
        style_text = "权益弹性与偏防御策略的平均表现接近，市场风格暂未形成明显倾斜。"
        dominant_style = "balanced"

    if top.empty:
        benchmark_text = ""
    else:
        top_name = str(top.iloc[0]["策略"])
        top_return = float(top.iloc[0][return_col])
        if pd.isna(cb_index_value):
            benchmark_text = f"{top_name}位居策略首位。"
        else:
            excess_return = top_return - cb_index_value
            if excess_return > 1e-12:
                benchmark_text = f"其中{top_name}跑赢转债指数{_fmt_pct(excess_return)}。"
            elif excess_return < -1e-12:
                benchmark_text = f"但最高收益策略{top_name}仍跑输转债指数{_fmt_pct(abs(excess_return))}。"
            else:
                benchmark_text = f"其中{top_name}与转债指数表现基本持平。"

    bottom_names = set(bottom["策略"].astype(str))
    bottom_equity_count = len(bottom_names.intersection(equity_names))
    bottom_defensive_count = len(bottom_names.intersection(defensive_names))
    if bottom_equity_count > bottom_defensive_count:
        if decomp.get("状态") == "成功" and decomp.get("正股贡献", 0.0) < 0:
            lag_reason = "靠后组合以权益弹性策略为主，与近期正股贡献偏弱相互印证。"
        else:
            lag_reason = "靠后组合以权益弹性策略为主，显示弹性暴露未能充分转化为组合收益。"
    elif bottom_defensive_count > bottom_equity_count:
        if dominant_style == "equity":
            lag_reason = "靠后组合以偏防御策略为主，权益上行阶段的债性保护相应牺牲了收益弹性。"
        elif decomp.get("状态") == "成功" and decomp.get("估值贡献", 0.0) < 0:
            lag_reason = "靠后组合以偏防御策略为主，估值压缩削弱了低价与债性保护的相对优势。"
        else:
            lag_reason = "靠后组合以偏防御策略为主，反映低价与债性特征并非当期主要收益来源。"
    else:
        lag_reason = "靠后策略的风格分布较为分散，收益差异更多来自个券选择与组合结构。"

    driver_text = _dominant_driver_text(decomp)
    if decomp.get("状态") == "成功":
        decomp_prefix = "从近20日市场背景看" if is_week else "近20日回报拆解显示"
        decomp_text = (
            f"{decomp_prefix}，转债收益率为{_fmt_pct(decomp['转债收益率'])}，"
            f"债券贡献{_fmt_pct(decomp['债券贡献'])}、正股贡献{_fmt_pct(decomp['正股贡献'])}、"
            f"估值贡献{_fmt_pct(decomp['估值贡献'])}。"
        )
    else:
        decomp_text = f"回报拆解暂未纳入（{decomp.get('状态', '未知原因')}）。"

    end_date_text = f"{end_date.year}年{end_date.month}月{end_date.day}日"
    commentary = (
        f"截至{end_date_text}{period_text}，{top_text}阶段性占优，{bottom_text}表现靠后。"
        f"{benchmark_text}{style_text}{lag_reason}{decomp_text}{driver_text}"
    )
    rows = [
        {"项目": "截止日", "内容": end_date.strftime("%Y-%m-%d")},
        {"项目": f"{period_name}占优策略", "内容": top_text},
        {"项目": f"{period_name}靠后策略", "内容": bottom_text},
        {"项目": f"转债指数{period_name}收益", "内容": _fmt_pct(cb_index_value)},
        {"项目": "回报拆解状态", "内容": "已使用集成回测内置计算" if decomp.get("状态") == "成功" else decomp.get("状态", "")},
        {"项目": commentary_item, "内容": commentary},
    ]
    if decomp.get("状态") == "成功":
        rows.extend([
            {"项目": "近20日转债收益率", "内容": _fmt_pct(decomp["转债收益率"])},
            {"项目": "近20日债券贡献", "内容": _fmt_pct(decomp["债券贡献"])},
            {"项目": "近20日正股贡献", "内容": _fmt_pct(decomp["正股贡献"])},
            {"项目": "近20日估值贡献", "内容": _fmt_pct(decomp["估值贡献"])},
        ])
    return pd.DataFrame(rows, columns=columns), commentary


def _build_weekly_strategy_commentary(
    summary_table: pd.DataFrame,
    end_date: pd.Timestamp,
    decomp: dict,
) -> tuple[pd.DataFrame, str]:
    return _build_period_strategy_commentary(summary_table, end_date, decomp, "week")


def _build_monthly_strategy_commentary(
    summary_table: pd.DataFrame,
    end_date: pd.Timestamp,
    decomp: dict,
) -> tuple[pd.DataFrame, str]:
    return _build_period_strategy_commentary(summary_table, end_date, decomp, "month")


def _build_evaluations(
    nav_df: pd.DataFrame,
    config: BacktestConfig,
    risk_free_rates: Optional[pd.Series] = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    nav_cols = [c for c in nav_df.columns if str(c).startswith("净值_") or str(c).startswith("基准净值_")]
    overall_rows = []
    annual_rows = []
    for col in nav_cols:
        name = str(col).replace("基准净值_", "").replace("净值_", "")
        overall_rows.append({"序列": name, **_series_metrics(nav_df[col], risk_free_rates)})
        daily_returns = pd.to_numeric(nav_df[col], errors="coerce").pct_change(fill_method=None)
        for year, group_returns in daily_returns.dropna().groupby(daily_returns.dropna().index.year):
            if len(group_returns) < 2:
                continue
            annual_return = float((1.0 + group_returns).prod() - 1.0)
            period_nav = pd.concat(
                [pd.Series([1.0], index=[group_returns.index[0] - pd.Timedelta(days=1)]), (1.0 + group_returns).cumprod()]
            )
            annual_volatility = float(group_returns.std(ddof=1) * np.sqrt(252))
            daily_risk_free = _daily_risk_free_returns(risk_free_rates, group_returns.index)
            excess_returns = (group_returns - daily_risk_free).dropna()
            annual_sharpe = (
                float(
                    excess_returns.mean()
                    / group_returns.loc[excess_returns.index].std(ddof=1)
                    * np.sqrt(252)
                )
                if len(excess_returns) > 1
                and group_returns.loc[excess_returns.index].std(ddof=1) > 0
                else np.nan
            )
            annual_rows.append({
                "年份": int(year), "序列": name, "年度收益率": annual_return,
                "最大回撤": _max_drawdown(period_nav), "夏普比率": float(annual_sharpe),
            })
    return pd.DataFrame(overall_rows), pd.DataFrame(annual_rows)


def _build_double_low_price_stats(rebalance_df: pd.DataFrame) -> pd.DataFrame:
    """生成双低组合每期价格统计，便于绘制历史均价曲线。"""
    columns = [
        "调仓日", "平均价格", "中位数价格", "最低价格", "最高价格",
        "持仓数量", "定价依据日", "平均转股溢价率", "平均双低值", "关键节点",
    ]
    if rebalance_df.empty or "组合" not in rebalance_df.columns:
        return pd.DataFrame(columns=columns)
    detail = rebalance_df[rebalance_df["组合"] == "双低"].copy()
    if detail.empty:
        return pd.DataFrame(columns=columns)

    stats = (
        detail.groupby(["调仓日", "定价依据日"], as_index=False)
        .agg(
            平均价格=("收盘价", "mean"),
            中位数价格=("收盘价", "median"),
            最低价格=("收盘价", "min"),
            最高价格=("收盘价", "max"),
            持仓数量=("转债代码", "count"),
            平均转股溢价率=("转股溢价率", "mean"),
            平均双低值=("双低值", "mean"),
        )
        .sort_values("调仓日")
        .reset_index(drop=True)
    )
    stats["关键节点"] = ""
    event_date = pd.Timestamp("2024-09-24")
    pre_event = stats[stats["调仓日"] < event_date]
    if not pre_event.empty:
        stats.loc[pre_event.index[-1], "关键节点"] = "2024年9月行情前最近一期"
    return stats[columns]


def run_backtest(config: BacktestConfig) -> Dict[str, pd.DataFrame]:
    data = load_original_data(
        parquet_root=config.parquet_root,
        force_refresh=config.force_refresh,
    )
    required_sheets = [
        config.close_sheet, config.return_sheet, config.plain_sheet, config.premium_sheet,
        config.turnover_sheet, config.balance_sheet, config.remain_year_sheet,
        config.stock_vol_sheet, config.redeem_days_sheet, config.downward_days_sheet,
        config.implied_vol_sheet, config.stock_close_sheet, config.expma5_sheet,
        config.expma10_sheet, config.expma20_sheet, config.stock_bond_premium_sheet,
    ]
    missing = [name for name in required_sheets if name not in data]
    if missing:
        raise KeyError(f"数据源缺少必要 sheet: {missing}")

    raw_mats = {
        "收盘价": data[config.close_sheet],
        "涨跌幅": data[config.return_sheet],
        "平价": data[config.plain_sheet],
        "转股溢价率": data[config.premium_sheet],
        "换手率": data[config.turnover_sheet],
        "余额": data[config.balance_sheet],
        "剩余期限": data[config.remain_year_sheet],
        "正股20日波动率": data[config.stock_vol_sheet],
        "赎回累计天数": data[config.redeem_days_sheet],
        "下修累计天数": data[config.downward_days_sheet],
        "隐含波动率": data[config.implied_vol_sheet],
        "正股收盘价": data[config.stock_close_sheet],
        "EXPMA5": data[config.expma5_sheet],
        "EXPMA10": data[config.expma10_sheet],
        "EXPMA20": data[config.expma20_sheet],
        "平价底价溢价率": data[config.stock_bond_premium_sheet],
    }
    normalized = {name: _normalize_wide(df) for name, df in raw_mats.items()}
    common_codes: Optional[pd.Index] = None
    common_dates: Optional[pd.DatetimeIndex] = None
    for frame in normalized.values():
        common_codes = frame.index if common_codes is None else common_codes.intersection(frame.index)
        dates = _date_columns(frame)
        common_dates = dates if common_dates is None else common_dates.intersection(dates)
    assert common_codes is not None and common_dates is not None
    dates = common_dates[common_dates >= pd.Timestamp(config.backtest_start_date)]
    if len(dates) < 2:
        raise ValueError("回测区间的公共交易日不足2天。")

    mats = {
        name: frame.reindex(index=common_codes, columns=dates)
        for name, frame in normalized.items()
    }
    numeric_mats = {
        name: frame.apply(pd.to_numeric, errors="coerce")
        for name, frame in mats.items()
        if name != "交易状态"
    }
    status_df = data.get(config.trade_status_sheet)
    if status_df is not None:
        status_df = _normalize_wide(status_df).reindex(index=common_codes, columns=dates)
    total_dates = _clean_total_dates(data.get(config.total_sheet), common_codes)

    benchmark_navs: Dict[str, pd.Series] = {}
    risk_free_rates: Optional[pd.Series] = None
    benchmark_df = data.get(config.benchmark_sheet)
    if benchmark_df is not None and not benchmark_df.empty:
        benchmark_df = _normalize_wide(benchmark_df)
        for name in config.benchmark_names:
            if name not in benchmark_df.index:
                print(f"[benchmark] 未找到 {name}，跳过。")
                continue
            levels = pd.to_numeric(benchmark_df.loc[name].reindex(dates), errors="coerce")
            first = levels.dropna()
            if len(first) >= 2 and first.iloc[0] != 0:
                benchmark_navs[name] = levels / float(first.iloc[0])
        if config.risk_free_rate_name not in benchmark_df.index:
            raise KeyError(f"指数表缺少无风险利率序列: {config.risk_free_rate_name}")
        risk_free_rates = (
            pd.to_numeric(
                benchmark_df.loc[config.risk_free_rate_name].reindex(dates), errors="coerce"
            )
            / 100.0
        ).ffill()
        if risk_free_rates.dropna().empty:
            raise ValueError(f"无风险利率序列无有效数据: {config.risk_free_rate_name}")
        risk_free_rates.name = config.risk_free_rate_name
    else:
        raise KeyError(f"数据源缺少指数表，无法读取无风险利率: {config.risk_free_rate_name}")

    rebalance_positions = set(range(1, len(dates), config.rebalance_every_n_days))
    current_holdings: Dict[str, list[str]] = {name: [] for name in PORTFOLIO_NAMES}
    current_weights: Dict[str, Dict[str, float]] = {name: {} for name in PORTFOLIO_NAMES}
    navs: Dict[str, float] = {name: 1.0 for name in PORTFOLIO_NAMES}
    nav_rows: list[dict] = []
    rebalance_rows: list[dict] = []
    holdings_rows: list[dict] = []
    model_rows: list[dict] = []

    pbar = tqdm(enumerate(dates), total=len(dates), desc="性价比回测", unit="day")
    for pos, dt in pbar:
        turnovers = {name: 0.0 for name in PORTFOLIO_NAMES}
        if pos in rebalance_positions:
            pricing_dt = dates[pos - 1]
            cross = _cross_section(numeric_mats, pricing_dt)
            actual_price = pd.to_numeric(cross["收盘价"], errors="coerce")
            actual_premium = pd.to_numeric(cross["转股溢价率"], errors="coerce")
            parity = pd.to_numeric(cross["平价"], errors="coerce")
            stock_close = pd.to_numeric(cross["正股收盘价"], errors="coerce")
            expma5 = pd.to_numeric(cross["EXPMA5"], errors="coerce")
            expma10 = pd.to_numeric(cross["EXPMA10"], errors="coerce")
            expma20 = pd.to_numeric(cross["EXPMA20"], errors="coerce")
            stock_bond_premium = pd.to_numeric(cross["平价底价溢价率"], errors="coerce")
            double_low = actual_price + actual_premium
            momentum_strength = stock_close / expma20 - 1.0

            # 所有组合共用原回测框架的基础候选池限制。
            base_active = actual_price.gt(0)
            base_active &= pd.to_numeric(cross["余额"], errors="coerce").ge(config.min_balance)
            base_active &= pd.to_numeric(cross["剩余期限"], errors="coerce").ge(config.min_remaining_years)
            listing = total_dates["上市日期"]
            last_trade = total_dates["最后交易日"]
            redeem_announce = total_dates["赎回公告日"]
            if listing.notna().any():
                base_active &= listing.isna() | listing.le(pricing_dt)
            if last_trade.notna().any():
                base_active &= last_trade.isna() | last_trade.ge(pricing_dt)
            if config.exclude_redeem_announced:
                base_active &= redeem_announce.isna() | redeem_announce.gt(pricing_dt)
            if status_df is not None and pricing_dt in status_df.columns:
                base_active &= status_df[pricing_dt].astype("string").isin(["交易", "新股上市"])

            selections: Dict[str, pd.Series] = {}
            candidate_counts: Dict[str, int] = {}
            selection_specs = {
                "双低": (double_low, base_active & double_low.notna(), "双低值", True),
                "低价": (actual_price, base_active & actual_price.notna(), "收盘价", True),
                "低转股溢价率": (actual_premium, base_active & actual_premium.notna(), "转股溢价率", True),
                "动量": (
                    momentum_strength,
                    base_active & stock_close.gt(expma5) & expma5.gt(expma10) & expma10.gt(expma20),
                    "正股/EXPMA20-1",
                    False,
                ),
                "偏股": (stock_bond_premium, base_active & stock_bond_premium.gt(20), "平价底价溢价率", False),
                "偏债": (stock_bond_premium, base_active & stock_bond_premium.lt(-20), "平价底价溢价率", True),
                "平衡": (
                    stock_bond_premium.abs(),
                    base_active & stock_bond_premium.gt(-20) & stock_bond_premium.lt(20),
                    "|平价底价溢价率|",
                    True,
                ),
            }
            selection_labels: Dict[str, str] = {}
            for portfolio_name, (ranking_value, mask, label, ascending) in selection_specs.items():
                candidates = ranking_value[mask].dropna()
                candidate_counts[portfolio_name] = int(len(candidates))
                selections[portfolio_name] = (
                    candidates.nsmallest(config.max_holdings)
                    if ascending
                    else candidates.nlargest(config.max_holdings)
                )
                selection_labels[portfolio_name] = label

            theoretical_premium = pd.Series(np.nan, index=common_codes, dtype="float64")
            theoretical_price = pd.Series(np.nan, index=common_codes, dtype="float64")
            undervaluation = pd.Series(np.nan, index=common_codes, dtype="float64")
            try:
                model = _fit_model(cross, config)
                theoretical_premium = _predict_premium(cross, model)
                theoretical_price = parity * (1.0 + theoretical_premium / 100.0)
                undervaluation = theoretical_price / actual_price - 1.0
                value_active = (
                    base_active & parity.gt(RESIDUAL_POWER_LOWER) & parity.lt(RESIDUAL_POWER_UPPER)
                    & theoretical_premium.notna() & theoretical_price.gt(0) & undervaluation.gt(0)
                )
                value_candidates = undervaluation[value_active].dropna()
                selections["性价比"] = value_candidates.nlargest(config.max_holdings)
                candidate_counts["性价比"] = int(len(value_candidates))
                selection_labels["性价比"] = "低估比例"
                coef_names = ["截距项", *RESIDUAL_FEATURES]
                model_row = {
                    "调仓日": dt, "定价依据日": pricing_dt, "拟合状态": "成功", "样本数": model.sample_count,
                    "A": model.amplitude, "scale": model.scale, "p": model.power, "floor": model.floor,
                    "基准R2": model.base_r2, "残差修正R2": model.corrected_r2,
                    "基准RMSE": model.base_rmse, "残差修正RMSE": model.corrected_rmse,
                    "低估券数": int(value_active.sum()), "入选数": len(selections["性价比"]),
                }
                model_row.update({f"系数_{name}": float(value) for name, value in zip(coef_names, model.beta)})
                model_rows.append(model_row)
            except Exception as exc:
                # 单期拟合失败时沿用上期持仓，不因数据断点中断整段回测。
                model_rows.append({"调仓日": dt, "定价依据日": pricing_dt, "拟合状态": f"失败，沿用上期持仓: {exc}"})

            # 更新各组合持仓、换手率和可审计调仓明细。
            for portfolio_name, selected in selections.items():
                new_holdings = selected.index.tolist()
                new_weights = _equal_weights(new_holdings)
                turnovers[portfolio_name] = _one_way_turnover(
                    current_weights[portfolio_name], new_weights
                )
                current_holdings[portfolio_name] = new_holdings
                current_weights[portfolio_name] = new_weights
                for rank, code in enumerate(new_holdings, start=1):
                    rebalance_rows.append({
                        "调仓日": dt, "定价依据日": pricing_dt, "组合": portfolio_name,
                        "排名": rank, "转债代码": code, "排序指标": selection_labels[portfolio_name],
                        "排序值": selected.loc[code], "候选数": candidate_counts[portfolio_name],
                        "收盘价": actual_price.loc[code], "转股溢价率": actual_premium.loc[code],
                        "双低值": double_low.loc[code], "平价": parity.loc[code],
                        "平价底价溢价率": stock_bond_premium.loc[code],
                        "正股收盘价": stock_close.loc[code], "EXPMA5": expma5.loc[code],
                        "EXPMA10": expma10.loc[code], "EXPMA20": expma20.loc[code],
                        "动量强度": momentum_strength.loc[code],
                        "理论转股溢价率": theoretical_premium.loc[code],
                        "理论价格": theoretical_price.loc[code], "低估比例": undervaluation.loc[code],
                    })

        nav_row = {"日期": dt}
        for portfolio_name in PORTFOLIO_NAMES:
            holdings = current_holdings[portfolio_name]
            if holdings:
                day_returns = numeric_mats["涨跌幅"].loc[holdings, dt] / 100.0
                gross_return, drifted_weights = _portfolio_return_and_drift(
                    current_weights[portfolio_name], day_returns
                )
                current_weights[portfolio_name] = drifted_weights
            else:
                gross_return = 0.0
            cost = turnovers[portfolio_name] * config.transaction_cost_rate
            net_return = gross_return - cost
            navs[portfolio_name] *= 1.0 + net_return
            nav_row[f"净值_{portfolio_name}"] = navs[portfolio_name]
            holdings_rows.append({
                "日期": dt, "组合": portfolio_name,
                **{f"持仓{i}": code for i, code in enumerate(holdings, start=1)},
            })
        nav_rows.append(nav_row)
        pbar.set_postfix(日期=dt.strftime("%Y-%m-%d"), 性价比持仓=len(current_holdings["性价比"]), 净值=f"{navs['性价比']:.4f}")

    nav_df = pd.DataFrame(nav_rows).set_index("日期")
    for name, benchmark_nav in benchmark_navs.items():
        nav_df[f"基准净值_{name}"] = benchmark_nav.reindex(nav_df.index)
    overall, annual = _build_evaluations(nav_df, config, risk_free_rates)
    summary_table = _build_summary_table(nav_df, config, risk_free_rates)
    risk_free_df = risk_free_rates.to_frame(name="十年国债收益率")
    risk_free_df.index.name = "日期"
    end_ts = pd.Timestamp(nav_df.index.max())
    decomp_timeseries, decomp_industry, decomp_parity, decomp_type = _calculate_integrated_return_decomposition(
        data, config, end_ts
    )
    latest_decomp = _latest_return_decomposition(decomp_timeseries, end_ts)
    weekly_commentary, _ = _build_weekly_strategy_commentary(summary_table, end_ts, latest_decomp)
    monthly_commentary, _ = _build_monthly_strategy_commentary(summary_table, end_ts, latest_decomp)
    rebalance_df = pd.DataFrame(rebalance_rows)
    double_low_price_stats = _build_double_low_price_stats(rebalance_df)
    return {
        "净值曲线": nav_df,
        "总结表格": summary_table,
        "本周策略点评": weekly_commentary,
        "本月策略点评": monthly_commentary,
        "整体评估": overall,
        "分年评估": annual,
        "无风险利率": risk_free_df,
        "调仓组合": rebalance_df,
        "双低组合均价": double_low_price_stats,
        "日度持仓": pd.DataFrame(holdings_rows),
        "模型诊断": pd.DataFrame(model_rows),
        "回报拆解_历史": decomp_timeseries,
        "回报拆解_行业": decomp_industry,
        "回报拆解_平价": decomp_parity,
        "回报拆解_类型": decomp_type,
    }


def build_report_results(results: Dict[str, pd.DataFrame], config: BacktestConfig) -> Dict[str, pd.DataFrame]:
    """截取并归一化报告区间，随后重新计算 Excel 中的绩效与点评。"""
    report_start = pd.Timestamp(config.report_start_date)
    report_results = {
        name: frame.copy() if isinstance(frame, pd.DataFrame) else frame
        for name, frame in results.items()
    }

    nav_df = report_results.get("净值曲线", pd.DataFrame()).copy()
    if nav_df.empty:
        raise ValueError("净值曲线为空，无法生成报告结果")
    nav_df.index = pd.to_datetime(nav_df.index, errors="coerce")
    nav_df = nav_df.loc[nav_df.index.notna() & (nav_df.index >= report_start)].copy()
    if nav_df.empty:
        raise ValueError(f"{config.report_start_date} 以来无净值数据，无法生成 Excel 和图表")
    for column in nav_df.columns:
        values = pd.to_numeric(nav_df[column], errors="coerce")
        valid = values.dropna()
        if valid.empty or not np.isfinite(valid.iloc[0]) or valid.iloc[0] == 0:
            nav_df[column] = values
            continue
        nav_df[column] = values / float(valid.iloc[0])
    report_results["净值曲线"] = nav_df

    risk_free_frame = report_results.get("无风险利率", pd.DataFrame()).copy()
    if risk_free_frame.empty or "十年国债收益率" not in risk_free_frame.columns:
        raise ValueError("回测结果缺少十年国债无风险利率序列")
    risk_free_frame.index = pd.to_datetime(risk_free_frame.index, errors="coerce")
    risk_free_frame = risk_free_frame.loc[
        risk_free_frame.index.notna() & (risk_free_frame.index >= report_start)
    ].copy()
    if risk_free_frame.empty:
        raise ValueError(f"{config.report_start_date} 以来无十年国债收益率数据")
    report_results["无风险利率"] = risk_free_frame
    report_risk_free_rates = pd.to_numeric(
        risk_free_frame["十年国债收益率"], errors="coerce"
    )

    date_columns = {
        "调仓组合": "调仓日",
        "双低组合均价": "调仓日",
        "日度持仓": "日期",
        "模型诊断": "调仓日",
        "回报拆解_历史": "日期",
    }
    for result_name, date_column in date_columns.items():
        frame = report_results.get(result_name)
        if not isinstance(frame, pd.DataFrame) or frame.empty or date_column not in frame.columns:
            continue
        dates = pd.to_datetime(frame[date_column], errors="coerce")
        report_results[result_name] = frame.loc[dates.ge(report_start)].copy()

    overall, annual = _build_evaluations(nav_df, config, report_risk_free_rates)
    summary_table = _build_summary_table(nav_df, config, report_risk_free_rates)
    end_ts = pd.Timestamp(nav_df.index.max())
    latest_decomp = _latest_return_decomposition(
        report_results.get("回报拆解_历史", pd.DataFrame()), end_ts
    )
    weekly_commentary, _ = _build_weekly_strategy_commentary(summary_table, end_ts, latest_decomp)
    monthly_commentary, _ = _build_monthly_strategy_commentary(summary_table, end_ts, latest_decomp)
    report_results["整体评估"] = overall
    report_results["分年评估"] = annual
    report_results["总结表格"] = summary_table
    report_results["本周策略点评"] = weekly_commentary
    report_results["本月策略点评"] = monthly_commentary
    return report_results


def validate_results(results: Dict[str, pd.DataFrame], config: BacktestConfig) -> dict:
    """审计组合约束和净值页字段；可恢复的模型拟合失败仅记为告警。"""
    rebalance_df = results.get("调仓组合", pd.DataFrame()).copy()
    nav_df = results.get("净值曲线", pd.DataFrame()).copy()
    diagnostics = results.get("模型诊断", pd.DataFrame()).copy()
    violations = {
        "missingPortfolio": [],
        "maxHoldings": 0,
        "value": 0,
        "momentum": 0,
        "equity": 0,
        "debt": 0,
        "balanced": 0,
        "navNonNavColumns": [],
        "sorting": 0,
    }
    warnings = {"failedFits": 0}

    if rebalance_df.empty:
        violations["missingPortfolio"] = list(PORTFOLIO_NAMES)
    else:
        by_portfolio = {name: frame for name, frame in rebalance_df.groupby("组合")}
        violations["missingPortfolio"] = [name for name in PORTFOLIO_NAMES if name not in by_portfolio]
        max_rank = pd.to_numeric(rebalance_df["排名"], errors="coerce").max()
        violations["maxHoldings"] = int(max_rank) if pd.notna(max_rank) and max_rank > config.max_holdings else 0

        value = by_portfolio.get("性价比", pd.DataFrame())
        if not value.empty:
            violations["value"] = int((pd.to_numeric(value["低估比例"], errors="coerce") <= 0).sum())

        momentum = by_portfolio.get("动量", pd.DataFrame())
        if not momentum.empty:
            momentum_ok = (
                pd.to_numeric(momentum["正股收盘价"], errors="coerce").gt(pd.to_numeric(momentum["EXPMA5"], errors="coerce"))
                & pd.to_numeric(momentum["EXPMA5"], errors="coerce").gt(pd.to_numeric(momentum["EXPMA10"], errors="coerce"))
                & pd.to_numeric(momentum["EXPMA10"], errors="coerce").gt(pd.to_numeric(momentum["EXPMA20"], errors="coerce"))
            )
            violations["momentum"] = int((~momentum_ok).sum())

        equity = by_portfolio.get("偏股", pd.DataFrame())
        if not equity.empty:
            violations["equity"] = int((pd.to_numeric(equity["平价底价溢价率"], errors="coerce") <= 20).sum())
        debt = by_portfolio.get("偏债", pd.DataFrame())
        if not debt.empty:
            violations["debt"] = int((pd.to_numeric(debt["平价底价溢价率"], errors="coerce") >= -20).sum())
        balanced = by_portfolio.get("平衡", pd.DataFrame())
        if not balanced.empty:
            premium = pd.to_numeric(balanced["平价底价溢价率"], errors="coerce")
            violations["balanced"] = int((~(premium.gt(-20) & premium.lt(20))).sum())

        direction = {"双低": 1, "低价": 1, "低转股溢价率": 1, "动量": -1, "偏股": -1, "偏债": 1, "平衡": 1, "性价比": -1}
        sorting_violations = 0
        for name, frame in by_portfolio.items():
            if name not in direction:
                continue
            for _, group in frame.groupby("调仓日"):
                ordered = group.sort_values("排名")
                values = pd.to_numeric(ordered["排序值"], errors="coerce").to_numpy()
                diffs = np.diff(values)
                if direction[name] > 0:
                    sorting_violations += int((diffs < -1e-12).sum())
                else:
                    sorting_violations += int((diffs > 1e-12).sum())
        violations["sorting"] = sorting_violations

    if not diagnostics.empty and "拟合状态" in diagnostics.columns:
        # 主回测对单期拟合失败已有“沿用上期持仓”的可恢复处理；
        # 此处保留诊断记录，但不应让已经完成的整段回测无法输出。
        warnings["failedFits"] = int(diagnostics["拟合状态"].astype(str).ne("成功").sum())

    nav_columns = list(nav_df.columns)
    violations["navNonNavColumns"] = [
        str(col) for col in nav_columns
        if not str(col).startswith("净值_") and not str(col).startswith("基准净值_")
    ]
    summary = {
        "navRows": int(len(nav_df)),
        "rebalanceRows": int(len(rebalance_df)),
        "modelPeriods": int(len(diagnostics)),
        "portfolioRows": {
            name: int((rebalance_df["组合"] == name).sum()) if "组合" in rebalance_df.columns else 0
            for name in PORTFOLIO_NAMES
        },
        "maxHoldings": int(pd.to_numeric(rebalance_df["排名"], errors="coerce").max()) if not rebalance_df.empty else 0,
        "violations": violations,
        "warnings": warnings,
    }
    has_violation = any(len(v) > 0 if isinstance(v, list) else v > 0 for v in violations.values())
    if has_violation:
        raise AssertionError(f"回测结果审计未通过: {summary}")
    return summary


def format_workbook(excel_path: Path, config: BacktestConfig) -> None:
    """用 openpyxl 完成最终 Excel 样式与说明页，避免依赖临时 JS 脚本。"""
    workbook = load_workbook(excel_path)
    navy = "17365D"
    light_blue = "D9EAF7"
    border_color = "D9E2F3"
    header_fill = PatternFill("solid", fgColor=navy)
    label_fill = PatternFill("solid", fgColor=light_blue)
    header_font = Font(bold=True, color="FFFFFF")
    label_font = Font(bold=True, color="000000")
    thin_side = Side(style="thin", color=border_color)
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for sheet_name in [
        "净值曲线", "总结表格", "本周策略点评", "本月策略点评", "整体评估", "分年评估",
        "无风险利率",
        "调仓组合", "双低组合均价", "日度持仓", "模型诊断",
        "回报拆解_历史", "回报拆解_行业", "回报拆解_平价", "回报拆解_类型",
    ]:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 1 or max_col < 1:
            continue
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[1].height = 30

        format_row_count = min(max_row, 1200)
        headers = [str(ws.cell(1, col).value or "") for col in range(1, max_col + 1)]
        for col_idx, label in enumerate(headers, start=1):
            values = [str(ws.cell(row, col_idx).value or "") for row in range(1, min(max_row, 200) + 1)]
            max_len = max([len(label) * 2, *[len(value) for value in values]])
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max(max_len + 2, 10), 42 if "状态" in label else 24)
            for row_idx in range(2, format_row_count + 1):
                cell = ws.cell(row_idx, col_idx)
                cell.border = Border(top=thin_side, bottom=thin_side)
                cell.alignment = Alignment(vertical="center")
            if "日" in label or label == "日期":
                number_format = "yyyy-mm-dd"
            elif any(key in label for key in ["收益率", "回撤", "波动率", "低估比例", "贡献"]) or label in {"换手率", "交易成本"}:
                number_format = "0.00%;[Red](0.00%);-"
            elif any(key in label for key in ["净值", "价格", "平价", "溢价率", "RMSE", "R2", "夏普", "卡玛"]):
                number_format = "0.0000;[Red](0.0000);-"
            else:
                number_format = None
            if number_format:
                for row_idx in range(2, max_row + 1):
                    ws.cell(row_idx, col_idx).number_format = number_format

        if sheet_name in {"总结表格", "整体评估", "分年评估"}:
            for row_idx in range(2, max_row + 1):
                ws.cell(row_idx, 1).fill = label_fill
                ws.cell(row_idx, 1).font = label_font
        if sheet_name in {"总结表格", "整体评估"}:
            ws.column_dimensions["A"].width = 22
        if sheet_name in {"本周策略点评", "本月策略点评"}:
            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 110
            for row_idx in range(2, max_row + 1):
                ws.cell(row_idx, 2).alignment = Alignment(vertical="top", wrap_text=True)
                ws.row_dimensions[row_idx].height = 36 if row_idx != 7 else 110
        if sheet_name == "分年评估" and max_col >= 2:
            ws.column_dimensions["B"].width = 22

    if "回测说明" in workbook.sheetnames:
        del workbook["回测说明"]
    notes = workbook.create_sheet("回测说明")
    notes.sheet_view.showGridLines = False
    notes.merge_cells("A1:D1")
    notes["A1"] = "性价比与七类对照组合回测"
    notes["A1"].fill = header_fill
    notes["A1"].font = Font(bold=True, color="FFFFFF", size=15)
    notes["A1"].alignment = Alignment(vertical="center")
    notes.row_dimensions[1].height = 32

    note_rows = [
        ("项目", "口径"),
        ("完整回测区间", f"{config.backtest_start_date} 至最新公共交易日"),
        ("Excel及绘图区间", f"{config.report_start_date} 至最新公共交易日（各序列以区间首个有效值归一为1）"),
        ("调仓频率", f"每{config.rebalance_every_n_days}个交易日"),
        ("定价时点", "调仓日前一交易日收盘截面（避免前视）"),
        ("持仓数", f"所有组合共用 max_holdings，当前最多{config.max_holdings}只，等权"),
        ("低估比例", "理论价/实际价-1，只选正值"),
        ("模型", "幂衰减基准曲线+六因子截面残差修正"),
        ("双低", "收盘价+转股溢价率由低到高取前N只"),
        ("低价", "收盘价由低到高取前N只"),
        ("低转股溢价率", "转股溢价率由低到高取前N只"),
        ("动量", "正股收盘价>EXPMA5>EXPMA10>EXPMA20；超过N只时按正股/EXPMA20-1由高到低截断"),
        ("偏股", "平价底价溢价率>20，由高到低取前N只"),
        ("偏债", "平价底价溢价率<-20，由低到高取前N只"),
        ("平衡", "-20<平价底价溢价率<20，按绝对值由低到高取前N只"),
        ("拟合样本", "50<平价<200，换手率<50，溢价率3%/97%分位去极值"),
        ("基础过滤", "余额≥3亿元，剩余期限≥0.5年，剔除已公告强赎券"),
        ("收益口径", "调仓日等权买入，持有期权重自然漂移；个券日收益缺失按0处理"),
        ("净值回写", "将完整回测区间的八个组合净值按组合名写入 parquet 的指数表，不写入基准净值"),
        ("回报拆解", f"内置滚动{config.decomposition_rolling_window}日回归：纯债价值、平价与转股溢价率变动三因子；按期初余额汇总"),
        ("交易费用", "默认0，可在脚本配置 transaction_cost_rate"),
        ("无风险利率", f"使用指数表中的{config.risk_free_rate_name}逐日年化收益率，并按252个交易日换算为日度收益率"),
        ("夏普比率", "按日度超额收益均值/日收益波动率×√252计算"),
    ]
    for row_idx, (item, desc) in enumerate(note_rows, start=3):
        notes.cell(row_idx, 1, item)
        notes.cell(row_idx, 2, desc)
        for col_idx in (1, 2):
            cell = notes.cell(row_idx, col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        if row_idx == 3:
            notes.cell(row_idx, 1).fill = header_fill
            notes.cell(row_idx, 2).fill = header_fill
            notes.cell(row_idx, 1).font = header_font
            notes.cell(row_idx, 2).font = header_font
        else:
            notes.cell(row_idx, 1).fill = label_fill
            notes.cell(row_idx, 1).font = label_font
    notes.column_dimensions["A"].width = 20
    notes.column_dimensions["B"].width = 78
    notes.freeze_panes = "A4"
    workbook.save(excel_path)


def save_results(results: Dict[str, pd.DataFrame], config: BacktestConfig) -> tuple[str, str, str]:
    nav_df = results["净值曲线"]
    end_ts = pd.Timestamp(nav_df.index.max())
    end_date = end_ts.strftime("%Y%m%d")
    out_dir = Path(config.output_dir_template.format(end_date=end_date))
    out_dir.mkdir(parents=True, exist_ok=True)
    excel_path = out_dir / config.output_file_name
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for sheet_name, frame in results.items():
            frame.to_excel(writer, sheet_name=sheet_name[:31], index=frame.index.name is not None)

    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei"]
    plt.rcParams["axes.unicode_minus"] = False
    fig, ax = plt.subplots(figsize=(14, 7))
    colors = ["#E6121B", "#0262BA", "#F79646", "#70AD47", "#7030A0", "#00B0F0", "#A5A5A5", "#FFC000", "#5B9BD5", "#C00000"]
    portfolio_cols = [c for c in nav_df.columns if str(c).startswith("净值_")]
    for i, col in enumerate(portfolio_cols):
        label = str(col).replace("净值_", "")
        ax.plot(nav_df.index, nav_df[col], label=label, color=colors[i % len(colors)], linewidth=1.6 if label == "性价比" else 1.1)
    benchmark_cols = [c for c in nav_df.columns if str(c).startswith("基准净值_")]
    for i, col in enumerate(benchmark_cols):
        ax.plot(nav_df.index, nav_df[col], label=str(col).replace("基准净值_", ""), color=colors[(i + len(portfolio_cols)) % len(colors)], linewidth=1.3, linestyle="--")
    ax.set_title("回测合集：性价比与七类对照组合")
    ax.set_xlabel("日期")
    ax.set_ylabel("净值")
    ax.grid(alpha=0.3)
    ax.legend()
    fig.tight_layout()
    png_path = excel_path.with_suffix(".png")
    fig.savefig(png_path, dpi=180)
    plt.close(fig)

    week_png_path = excel_path.with_name(f"{excel_path.stem}_本周收益回撤.png")
    _plot_week_performance(results.get("总结表格", pd.DataFrame()), week_png_path, end_ts)

    double_low_stats = results.get("双低组合均价", pd.DataFrame())
    if double_low_stats is not None and not double_low_stats.empty:
        fig, ax = plt.subplots(figsize=(12, 6))
        x = pd.to_datetime(double_low_stats["调仓日"])
        y = pd.to_numeric(double_low_stats["平均价格"], errors="coerce")
        ax.plot(x, y, color="#0262BA", linewidth=1.8, marker="o", markersize=3, label="双低组合平均价格")
        event_date = pd.Timestamp("2024-09-24")
        ax.axvline(event_date, color="#E6121B", linestyle="--", linewidth=1.2, label="2024年9月行情启动")
        key_rows = double_low_stats[double_low_stats["关键节点"].astype(str).str.len() > 0]
        for _, row in key_rows.iterrows():
            key_date = pd.Timestamp(row["调仓日"])
            key_price = float(row["平均价格"])
            ax.scatter([key_date], [key_price], color="#E6121B", s=45, zorder=5)
            ax.annotate(
                f"{key_date:%Y-%m-%d}\n{key_price:.2f}元",
                xy=(key_date, key_price), xytext=(12, 14), textcoords="offset points",
                color="#E6121B", fontsize=10,
                arrowprops={"arrowstyle": "->", "color": "#E6121B"},
            )
        ax.set_title("双低组合历史平均价格")
        ax.set_xlabel("调仓日")
        ax.set_ylabel("平均价格（元）")
        ax.grid(alpha=0.3)
        ax.legend()
        fig.tight_layout()
        double_low_png = excel_path.with_name(f"{excel_path.stem}_双低组合均价.png")
        fig.savefig(double_low_png, dpi=180)
        plt.close(fig)
    format_workbook(excel_path, config)
    return str(excel_path.resolve()), str(png_path.resolve()), str(week_png_path.resolve())


def write_strategy_nav_to_parquet(results: Dict[str, pd.DataFrame], config: BacktestConfig) -> dict:
    """将组合净值以组合名写入标准 parquet 的“指数”表，不写入基准净值。"""
    if not config.write_strategy_nav_to_parquet:
        return {"status": "disabled", "rows": 0, "names": []}

    nav_df = results.get("净值曲线", pd.DataFrame())
    if nav_df.empty:
        raise ValueError("净值曲线为空，无法写入 parquet 指数表")
    missing_columns = [f"净值_{name}" for name in PORTFOLIO_NAMES if f"净值_{name}" not in nav_df.columns]
    if missing_columns:
        raise KeyError(f"净值曲线缺少组合列: {missing_columns}")

    strategy_parts = []
    nav_dates = pd.to_datetime(nav_df.index, errors="coerce")
    for name in PORTFOLIO_NAMES:
        values = pd.to_numeric(nav_df[f"净值_{name}"], errors="coerce")
        part = pd.DataFrame({INDEX_NAME: name, TRADE_DATE: nav_dates, INDEX_VALUE: values.to_numpy(dtype="float64")})
        strategy_parts.append(part.dropna(subset=[TRADE_DATE, INDEX_VALUE]))
    strategy_long = pd.concat(strategy_parts, ignore_index=True)
    strategy_long[INDEX_NAME] = strategy_long[INDEX_NAME].astype("string")
    strategy_long[TRADE_DATE] = pd.to_datetime(strategy_long[TRADE_DATE]).dt.normalize()
    strategy_long[INDEX_VALUE] = pd.to_numeric(strategy_long[INDEX_VALUE], errors="coerce").astype("float64")
    strategy_long = strategy_long.sort_values([TRADE_DATE, INDEX_NAME], kind="stable").reset_index(drop=True)

    root = Path(config.parquet_root).resolve()
    index_path = root / "_special" / "指数.parquet"
    if not index_path.exists():
        raise FileNotFoundError(f"未找到标准 parquet 指数文件: {index_path}")
    existing = pd.read_parquet(index_path)
    required_columns = [INDEX_NAME, TRADE_DATE, INDEX_VALUE]
    if list(existing.columns) != required_columns:
        raise ValueError(f"指数 parquet 字段不符合标准结构: {list(existing.columns)}")
    existing[INDEX_NAME] = existing[INDEX_NAME].astype("string")
    existing[TRADE_DATE] = pd.to_datetime(existing[TRADE_DATE]).dt.normalize()
    existing[INDEX_VALUE] = pd.to_numeric(existing[INDEX_VALUE], errors="coerce").astype("float64")

    strategy_names = set(PORTFOLIO_NAMES)
    old_strategy = existing[existing[INDEX_NAME].isin(strategy_names)].copy()
    old_strategy = old_strategy.sort_values([TRADE_DATE, INDEX_NAME], kind="stable").reset_index(drop=True)
    same = (
        len(old_strategy) == len(strategy_long)
        and np.array_equal(
            old_strategy[INDEX_NAME].astype(str).to_numpy(),
            strategy_long[INDEX_NAME].astype(str).to_numpy(),
        )
        and np.array_equal(
            old_strategy[TRADE_DATE].to_numpy(dtype="datetime64[ns]"),
            strategy_long[TRADE_DATE].to_numpy(dtype="datetime64[ns]"),
        )
        and np.allclose(
            old_strategy[INDEX_VALUE].to_numpy(dtype="float64"),
            strategy_long[INDEX_VALUE].to_numpy(dtype="float64"),
            equal_nan=True,
        )
    )
    if same:
        return {
            "status": "unchanged",
            "path": str(index_path),
            "rows": int(len(strategy_long)),
            "names": list(PORTFOLIO_NAMES),
        }

    preserved = existing[~existing[INDEX_NAME].isin(strategy_names)]
    combined = pd.concat([preserved, strategy_long], ignore_index=True)
    if combined.duplicated([INDEX_NAME, TRADE_DATE]).any():
        duplicates = combined.loc[combined.duplicated([INDEX_NAME, TRADE_DATE], keep=False), [INDEX_NAME, TRADE_DATE]]
        raise ValueError(f"指数名称＋交易日期主键重复: {duplicates.head().to_dict('records')}")
    combined = combined[required_columns].sort_values([TRADE_DATE, INDEX_NAME], kind="stable").reset_index(drop=True)
    write_typed_parquet(combined, index_path, index_schema())

    manifest_path = root / "_meta" / "sheet_manifest.parquet"
    if manifest_path.exists():
        manifest = pd.read_parquet(manifest_path)
        index_mask = manifest["sheet_name"].astype(str).eq(INDEX_SHEET)
        if index_mask.any():
            manifest.loc[index_mask, "route"] = "_special/指数.parquet"
            manifest.loc[index_mask, "rows"] = len(combined)
            manifest.loc[index_mask, "cols"] = len(required_columns)
            _write_manifest(root, manifest.to_dict("records"))

    # 当前进程中的原始数据缓存已经不再代表写盘后的指数表。
    _PARQUET_MEM_CACHE.clear()
    return {
        "status": "written",
        "path": str(index_path),
        "rows": int(len(strategy_long)),
        "names": list(PORTFOLIO_NAMES),
    }


def main() -> None:
    config = BacktestConfig()
    results = run_backtest(config)
    audit_summary = validate_results(results, config)
    report_results = build_report_results(results, config)
    excel_path, png_path, week_png_path = save_results(report_results, config)
    parquet_nav_summary = write_strategy_nav_to_parquet(results, config)
    print(f"回测完成：{excel_path}")
    print(f"净值图：{png_path}")
    print(f"本周收益回撤图：{week_png_path}")
    print(f"审计通过：{audit_summary}")
    print(f"组合净值写入 parquet：{parquet_nav_summary}")
    if not report_results["总结表格"].empty:
        print("总结表格：")
        print(report_results["总结表格"].to_string(index=False))
    weekly_commentary = report_results.get("本周策略点评", pd.DataFrame())
    if not weekly_commentary.empty and "项目" in weekly_commentary.columns and "内容" in weekly_commentary.columns:
        text = weekly_commentary.loc[weekly_commentary["项目"].eq("本周策略点评"), "内容"]
        if not text.empty:
            print("本周策略点评：")
            print(text.iloc[0])
    monthly_commentary = report_results.get("本月策略点评", pd.DataFrame())
    if not monthly_commentary.empty and "项目" in monthly_commentary.columns and "内容" in monthly_commentary.columns:
        text = monthly_commentary.loc[monthly_commentary["项目"].eq("本月策略点评"), "内容"]
        if not text.empty:
            print("本月策略点评：")
            print(text.iloc[0])
    if not report_results["整体评估"].empty:
        print(report_results["整体评估"].to_string(index=False))


if __name__ == "__main__":
    main()
