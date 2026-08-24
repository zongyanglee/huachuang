"""使用多因子修正模型计算分组拟合溢价率时间序列。

脚本内置数据读取、样本清洗、窗口构建、幂衰减基准曲线和六因子 OLS
修正的完整实现，不依赖其他本地业务脚本。除原有分组结果外，同时计算
单日全市场截面的“百元拟合溢价率”（反三次）和“多因子修正百元拟合
溢价率”。默认复用历史结果，仅计算缺失或新增交易日，输出一个汇总各类
拟合结果的 Excel 和各组折线图。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import pickle
import shutil
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from time import perf_counter
from typing import Callable, Dict, List, Optional

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from scipy.optimize import curve_fit
from tqdm import tqdm


WORKSPACE = Path(__file__).resolve().parents[2]
RUN_DATE = datetime.now().strftime("%Y%m%d")
DEFAULT_OUTPUT_DIR = WORKSPACE / "runs" / "research" / f"多因子修正拟合溢价率_{RUN_DATE}"
RESULT_CACHE_DIR = WORKSPACE / "tmp/cache/multifactor_fit"
DEFAULT_RESULT_CACHE_PATH = RESULT_CACHE_DIR / "多因子修正拟合溢价率历史.parquet"
DEFAULT_WEEKLY_MATURITY_CACHE_PATH = RESULT_CACHE_DIR / "剩余期限周分组.parquet"
INDEX_BACKUP_PATH = RESULT_CACHE_DIR / "指数_写入多因子拟合前备份.parquet"
INDEX_MANIFEST_BACKUP_PATH = RESULT_CACHE_DIR / "sheet_manifest_写入多因子拟合前备份.parquet"
LEGACY_RESULT_PATH = (
    WORKSPACE
    / "outputs"
    / "分组多因子修正拟合溢价率_2022年以来"
    / "分组多因子修正拟合明细_2022年以来.csv"
)
MODEL_CACHE_SCHEMA_VERSION = 1
WEEKLY_MATURITY_CACHE_SCHEMA_VERSION = 1
PARQUET_SCHEMA_VERSION = "2.11.0"
WEEKLY_MATURITY_MAX_WEEKS = 310

# 运行开关：不使用命令行时，直接在这里修改。
# False：复用历史拟合结果，仅计算缺失或新增交易日（默认、速度快）。
# True：忽略历史拟合结果缓存，从 2017 年起全量重新计算。非常耗时！！！
FORCE_REBUILD = False

# True：将全部日度拟合溢价率序列并入 parquet 指数长表。
# “剩余期限周分组”是期限截面，不纳入。
# 重复运行时仅写入新增或变化结果；已有且相同则跳过落盘。
WRITE_DAILY_FITS_TO_INDEX = True

BOND_CODE = "转债代码"
TRADE_DATE = "交易日期"
MASTER_SHEET = "总表"
INDEX_NAME = "指数名称"
INDEX_VALUE = "指数值"
RESIDUAL_POWER_ANCHOR = 50.0
RESIDUAL_POWER_LOWER = 50.0
RESIDUAL_POWER_UPPER = 200.0
RESIDUAL_MIN_SAMPLES = 36
INVERSE_CUBIC_MIN_SAMPLES = 8
ALL_MARKET_CATEGORY = "百元拟合"
INVERSE_CUBIC_HUNDRED_GROUP = "百元拟合溢价率"
MULTIFACTOR_HUNDRED_GROUP = "多因子修正百元拟合溢价率"
ALL_MARKET_GROUPS = (
    INVERSE_CUBIC_HUNDRED_GROUP,
    MULTIFACTOR_HUNDRED_GROUP,
)
ALL_MARKET_INDEX_NAMES = {
    INVERSE_CUBIC_HUNDRED_GROUP: INVERSE_CUBIC_HUNDRED_GROUP,
    MULTIFACTOR_HUNDRED_GROUP: MULTIFACTOR_HUNDRED_GROUP,
}
RESIDUAL_FEATURES = [
    "余额_log",
    "剩余期限",
    "正股20日波动率",
    "赎回累计天数",
    "下修累计天数",
    "隐含波动率",
]
REQUIRED_METRICS = [
    "平价",
    "转股溢价率",
    "换手率",
    "剩余期限",
    "余额",
    "正股20日波动率",
    "赎回累计天数",
    "下修累计天数",
    "隐含波动率",
    "平价底价溢价率",
    "债项评级",
    "正股市值",
]
MASTER_DATE_COLUMNS = ["上市日期", "最后交易日", "发行日期", "赎回公告日", "转股期起始日", "回售起始日期"]
_PARQUET_MEM_CACHE: dict[tuple[str, str], dict[str, pd.DataFrame]] = {}
_EXCEL_MEM_CACHE: dict[tuple[str, str], dict[str, pd.DataFrame]] = {}

# 华创图表标准色：各分组按此顺序着色，超过六条曲线时循环使用。
CHART_COLORS = (
    "#E6121B",  # red
    "#0262BA",  # blue
    "#A6A6A6",  # gray
    "#E6B9B8",  # pink
    "#B7DEE8",  # sky
    "#F79646",  # orange
)

EXCEL_DETAIL_EXCLUDED_COLUMNS = [
    "样本数",
    "R2",
    "基准R2",
    "修正R2",
    "多因子修正R2",
    "R2提升",
    "基准RMSE",
    "修正RMSE",
    "多因子修正RMSE",
]

SECTOR_FIT_INDEX_NAMES = {
    "科技": "多因子修正拟合溢价率_科技",
    "金融": "多因子修正拟合溢价率_金融",
    "制造": "多因子修正拟合溢价率_制造",
    "消费": "多因子修正拟合溢价率_消费",
    "周期": "多因子修正拟合溢价率_周期",
}


def _parquet_dir_fingerprint(input_root: str | Path) -> str:
    """根据所有 parquet 文件的路径、修改时间和大小生成缓存指纹。"""
    root = Path(input_root).resolve()
    file_info = []
    for path in sorted(root.rglob("*.parquet")):
        stat = path.stat()
        file_info.append(f"{path}|{stat.st_mtime_ns}|{stat.st_size}")
    return hashlib.md5("|".join(file_info).encode("utf-8")).hexdigest()


def _file_fingerprint(path: str | Path) -> tuple[str, str]:
    resolved = str(Path(path).resolve())
    stat = Path(resolved).stat()
    key = f"{resolved}|{stat.st_mtime_ns}|{stat.st_size}"
    return hashlib.md5(key.encode("utf-8")).hexdigest(), resolved


def _read_required_data_from_parquet(input_root: str | Path) -> dict[str, pd.DataFrame]:
    """读取标准月度 parquet，并还原模型所需宽表和总表。"""
    root = Path(input_root)
    files = sorted(
        path
        for year in root.iterdir()
        if year.is_dir() and year.name.isdigit()
        for path in year.glob("*.parquet")
    )
    if not files:
        raise FileNotFoundError(f"未找到月度 parquet 文件：{root}")

    parts: dict[str, list[pd.DataFrame]] = {metric: [] for metric in REQUIRED_METRICS}
    all_dates: list[pd.Timestamp] = []
    required_columns = [BOND_CODE, TRADE_DATE, *REQUIRED_METRICS]
    for path in tqdm(files, desc="读取 Parquet", unit="file"):
        frame = pd.read_parquet(path, columns=required_columns)
        missing = sorted(set(required_columns) - set(frame.columns))
        if missing:
            raise ValueError(f"{path} 缺少标准字段：{missing}")
        frame[TRADE_DATE] = pd.to_datetime(frame[TRADE_DATE], errors="coerce")
        frame = frame.dropna(subset=[BOND_CODE, TRADE_DATE])
        all_dates.extend(frame[TRADE_DATE].drop_duplicates().tolist())
        for metric in REQUIRED_METRICS:
            parts[metric].append(frame.pivot(index=BOND_CODE, columns=TRADE_DATE, values=metric))

    master_path = root / "_special" / "总表.parquet"
    if not master_path.exists():
        raise FileNotFoundError(f"未找到总表 parquet：{master_path}")
    master = pd.read_parquet(master_path)
    if BOND_CODE not in master.columns:
        raise ValueError(f"总表 parquet 缺少 `{BOND_CODE}`：{master_path}")
    master[BOND_CODE] = master[BOND_CODE].astype(str)
    code_order = master[BOND_CODE].tolist()
    date_order = pd.DatetimeIndex(sorted(pd.Index(all_dates).unique()))

    data: dict[str, pd.DataFrame] = {}
    for metric, metric_parts in parts.items():
        merged = pd.concat(metric_parts, axis=1) if metric_parts else pd.DataFrame()
        merged = merged.loc[:, ~merged.columns.duplicated(keep="last")]
        merged = merged.reindex(index=code_order, columns=date_order)
        merged.index.name = None
        data[metric] = merged

    master = master.set_index(BOND_CODE)
    master.index.name = None
    for column in MASTER_DATE_COLUMNS:
        if column in master.columns:
            master[column] = pd.to_datetime(master[column], errors="coerce")
    data[MASTER_SHEET] = master
    return data


def _read_parquet_date_index(input_root: str | Path) -> pd.DatetimeIndex:
    """仅读取交易日期列，用于在载入完整矩阵前快速判断是否存在新增日期。"""
    root = Path(input_root)
    files = sorted(
        path
        for year in root.iterdir()
        if year.is_dir() and year.name.isdigit()
        for path in year.glob("*.parquet")
    )
    dates: list[pd.Timestamp] = []
    for path in files:
        frame = pd.read_parquet(path, columns=[TRADE_DATE])
        parsed = pd.to_datetime(frame[TRADE_DATE], errors="coerce").dropna().drop_duplicates()
        dates.extend(parsed.tolist())
    return pd.DatetimeIndex(sorted(pd.Index(dates).unique()))


def _quick_source_dates(config: "RunConfig") -> Optional[pd.DatetimeIndex]:
    source = str(config.source_type or "auto").strip().lower()
    parquet_available = Path(config.parquet_root).is_dir()
    if source in {"parquet", "auto"} and parquet_available:
        try:
            dates = _read_parquet_date_index(config.parquet_root)
            if len(dates) > 0:
                return dates
        except Exception as exc:
            print(f"[result_cache] 快速日期检查失败，改为完整读取：{exc}")
    return None


def _load_parquet_with_cache(
    input_root: str | Path,
    force_refresh: bool = False,
    cache_dir: str | Path = "tmp/cache/parquet",
) -> dict[str, pd.DataFrame]:
    root = str(Path(input_root))
    fingerprint = _parquet_dir_fingerprint(root)
    mem_key = (str(Path(root).resolve()), fingerprint)
    cache_path = Path(cache_dir) / f"parquet_{fingerprint}.pkl"
    if not force_refresh:
        if mem_key in _PARQUET_MEM_CACHE:
            print(f"[cache] parquet memory hit: {root}")
            return _PARQUET_MEM_CACHE[mem_key]
        if cache_path.exists():
            print(f"[cache] 命中 parquet 磁盘缓存：{cache_path}")
            with cache_path.open("rb") as handle:
                cached = pickle.load(handle)
            _PARQUET_MEM_CACHE[mem_key] = cached
            return cached
    print(f"[cache] parquet 缓存未命中，直接读取：{root}")
    data = _read_required_data_from_parquet(root)
    _PARQUET_MEM_CACHE[mem_key] = data
    return data


def _read_excel_all_sheets(path: str | Path) -> dict[str, pd.DataFrame]:
    with pd.ExcelFile(path) as workbook:
        return {
            sheet_name: pd.read_excel(workbook, sheet_name=sheet_name, index_col=0)
            for sheet_name in tqdm(workbook.sheet_names, desc="读取 Excel", unit="sheet")
        }


def _load_excel_with_cache(
    file_name: str | Path,
    force_refresh: bool = False,
    cache_dir: str | Path = ".cache_excel",
) -> dict[str, pd.DataFrame]:
    fingerprint, resolved = _file_fingerprint(file_name)
    mem_key = (resolved, fingerprint)
    cache_path = Path(cache_dir) / f"excel_{fingerprint}.pkl"
    if not force_refresh:
        if mem_key in _EXCEL_MEM_CACHE:
            print(f"[cache] excel memory hit: {Path(resolved).name}")
            return _EXCEL_MEM_CACHE[mem_key]
        if cache_path.exists():
            print(f"[cache] 命中 Excel 磁盘缓存：{cache_path}")
            with cache_path.open("rb") as handle:
                cached = pickle.load(handle)
            _EXCEL_MEM_CACHE[mem_key] = cached
            return cached
    print(f"[cache] Excel 缓存未命中，直接读取：{resolved}")
    data = _read_excel_all_sheets(resolved)
    _EXCEL_MEM_CACHE[mem_key] = data
    return data


def load_original_data(
    source_type: str = "auto",
    excel_file_name: Optional[str] = None,
    parquet_root: str = "data/转债个券历史序列",
    force_refresh: bool = False,
) -> dict[str, pd.DataFrame]:
    """独立数据入口：支持 parquet、Excel，以及优先 parquet 的自动模式。"""
    source = str(source_type or "auto").strip().lower()
    if source not in {"excel", "parquet", "auto"}:
        raise ValueError(f"source_type 无效：{source_type}")
    parquet_available = Path(parquet_root).is_dir()
    excel_available = bool(excel_file_name) and Path(excel_file_name).is_file()

    if source == "parquet":
        if not parquet_available:
            raise FileNotFoundError(f"未找到 parquet 目录：{parquet_root}")
        print(f"[source] 使用 parquet 数据源：{parquet_root}")
        return _load_parquet_with_cache(parquet_root, force_refresh=force_refresh)
    if source == "excel":
        if not excel_available:
            raise FileNotFoundError(f"未找到 Excel 文件：{excel_file_name}")
        print(f"[source] 使用 Excel 数据源：{excel_file_name}")
        return _load_excel_with_cache(excel_file_name, force_refresh=force_refresh)
    if parquet_available:
        try:
            print(f"[source] auto 模式优先 parquet：{parquet_root}")
            return _load_parquet_with_cache(parquet_root, force_refresh=force_refresh)
        except Exception as exc:
            print(f"[source] parquet 读取失败，尝试 Excel：{exc}")
    if excel_available:
        return _load_excel_with_cache(excel_file_name, force_refresh=force_refresh)
    raise FileNotFoundError("未找到可用的 parquet 目录或 Excel 文件。")


def _normalize_columns_to_ts(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    parsed = pd.to_datetime(pd.Index(out.columns), errors="coerce")
    out.columns = [pd.Timestamp(ts) if pd.notna(ts) else raw for raw, ts in zip(out.columns, parsed)]
    return out


def _extract_date_columns(frame: pd.DataFrame) -> pd.DatetimeIndex:
    return pd.DatetimeIndex([column for column in frame.columns if isinstance(column, pd.Timestamp)]).sort_values()


def power_decay_with_floor(
    x: np.ndarray | float,
    amplitude: float,
    scale: float,
    power: float,
    floor: float,
    anchor_x: float = 70.0,
) -> np.ndarray | float:
    base = 1 + (np.asarray(x) - anchor_x) / scale
    return floor + amplitude * np.power(base, -power)


def inverse_cubic(
    x: np.ndarray | float,
    a: float,
    b: float,
    c: float,
    d: float,
) -> np.ndarray | float:
    """历史百元拟合使用的反三次函数。"""
    x_array = np.asarray(x)
    return a / np.power(x_array, 3) + b / np.power(x_array, 2) + c / x_array + d


def _fit_inverse_cubic_hundred_detail(df_subset: pd.DataFrame) -> dict:
    """按历史口径拟合单日截面，并且只读取平价100处的溢价率。"""
    empty = {
        "拟合溢价率": float("nan"),
        "样本数": 0,
        "拟合公式": "",
    }
    if df_subset is None or df_subset.empty:
        return empty
    work = df_subset[["平价", "转股溢价率"]].copy()
    work = work.replace("", np.nan).replace(0, np.nan)
    work["平价"] = pd.to_numeric(work["平价"], errors="coerce")
    work["转股溢价率"] = pd.to_numeric(work["转股溢价率"], errors="coerce")
    work = work.dropna(subset=["平价", "转股溢价率"])
    if len(work) < INVERSE_CUBIC_MIN_SAMPLES:
        return empty
    try:
        x = work["平价"].to_numpy(dtype=float)
        y = work["转股溢价率"].to_numpy(dtype=float)
        popt, _ = curve_fit(inverse_cubic, x, y, maxfev=20000)
        a, b, c, d = (float(value) for value in popt)
        return {
            "拟合溢价率": float(inverse_cubic(100.0, a, b, c, d)),
            "样本数": int(len(work)),
            "拟合公式": (
                f"转股溢价率 = {a:.2f}/平价^3 + {b:.2f}/平价^2 "
                f"+ {c:.2f}/平价 + {d:.2f}"
            ),
        }
    except Exception:
        return empty


def _residual_raw_scale_coefficients(coef_map: dict, stats: dict) -> dict:
    raw_coef = {}
    for feature in RESIDUAL_FEATURES:
        std = float(stats[feature]["std"])
        raw_coef[feature] = float(coef_map[feature]) / std if std > 0 else 0.0
    raw_coef["截距项"] = float(coef_map["截距项"]) - sum(
        raw_coef[feature] * float(stats[feature]["mean"])
        for feature in RESIDUAL_FEATURES
    )
    return raw_coef


def _format_signed_formula_term(coef: float, label: str = "") -> str:
    operator = "+" if coef >= 0 else "-"
    suffix = f" * {label}" if label else ""
    return f" {operator} {abs(coef):.2f}{suffix}"


def _format_residual_power_decay_formula(
    amplitude: float,
    scale: float,
    power: float,
    floor: float,
    coef_map: dict,
    stats: dict,
) -> str:
    raw_coef = _residual_raw_scale_coefficients(coef_map, stats)
    formula = (
        f"转股溢价率 = {floor:.2f} + {amplitude:.2f} * "
        f"(1 + (平价 - {RESIDUAL_POWER_ANCHOR:.2f}) / {scale:.2f})^(-{power:.2f})"
    )
    formula += _format_signed_formula_term(raw_coef["截距项"])
    formula += _format_signed_formula_term(raw_coef["余额_log"], "ln(1 + 余额)")
    formula += _format_signed_formula_term(raw_coef["剩余期限"], "剩余期限")
    formula += _format_signed_formula_term(raw_coef["正股20日波动率"], "正股20日波动率")
    formula += _format_signed_formula_term(raw_coef["赎回累计天数"], "赎回累计天数")
    formula += _format_signed_formula_term(raw_coef["下修累计天数"], "下修累计天数")
    formula += _format_signed_formula_term(raw_coef["隐含波动率"], "隐含波动率")
    return formula + "（缺失因子按当日中位数填充）"


def _empty_residual_power_decay_fit_row() -> dict:
    row = {
        "百元残差修正幂衰减溢价率": float("nan"),
        "百元幂衰减基准溢价率": float("nan"),
        "百元残差修正": float("nan"),
        "样本数": 0,
        "基准R2": float("nan"),
        "残差修正R2": float("nan"),
        "R2提升": float("nan"),
        "基准RMSE": float("nan"),
        "残差修正RMSE": float("nan"),
        "拟合公式": "",
    }
    for feature in ["截距项", *RESIDUAL_FEATURES]:
        row[f"系数_{feature}"] = float("nan")
    return row


def _residual_power_feature_design(
    work: pd.DataFrame,
    stats: Optional[dict] = None,
) -> tuple[pd.DataFrame, dict]:
    factors = work[RESIDUAL_FEATURES].apply(pd.to_numeric, errors="coerce").copy()
    if stats is None:
        stats = {}
        for column in RESIDUAL_FEATURES:
            median = factors[column].median()
            mean = factors[column].fillna(median).mean()
            std = factors[column].fillna(median).std(ddof=0)
            stats[column] = {
                "median": float(median) if pd.notna(median) else 0.0,
                "mean": float(mean) if pd.notna(mean) else 0.0,
                "std": float(std) if pd.notna(std) and std > 0 else 0.0,
            }
    design = pd.DataFrame(index=factors.index)
    design.insert(0, "截距项", 1.0)
    for column in RESIDUAL_FEATURES:
        column_stats = stats[column]
        filled = factors[column].fillna(column_stats["median"])
        if column_stats["std"] > 0:
            design[column] = (filled - column_stats["mean"]) / column_stats["std"]
        else:
            design[column] = 0.0
    return design, stats


def _fit_residual_power_decay_detail(
    df_subset: pd.DataFrame,
    target_x: float = 100.0,
    high_price_subset: Optional[pd.DataFrame] = None,
    include_diagnostics: bool = False,
) -> dict:
    if df_subset is None or df_subset.empty:
        return _empty_residual_power_decay_fit_row()
    required_cols = ["平价", "转股溢价率", *RESIDUAL_FEATURES]
    work = df_subset.copy()
    work["余额_log"] = np.log1p(pd.to_numeric(work["余额"], errors="coerce").clip(lower=0))
    work = work.replace("", np.nan).replace(0, np.nan)
    for column in required_cols:
        work[column] = pd.to_numeric(work[column], errors="coerce")
    work = work.dropna(subset=["平价", "转股溢价率"])
    if len(work) < RESIDUAL_MIN_SAMPLES:
        return _empty_residual_power_decay_fit_row()

    x = work["平价"].astype(float).to_numpy()
    y = work["转股溢价率"].astype(float).to_numpy()
    try:
        floor0 = float(np.clip(np.nanpercentile(y, 5), 0, 1))
        amplitude0 = float(max(np.nanpercentile(y, 95) - floor0, 1))
        popt, _ = curve_fit(
            lambda x_data, amplitude, scale, power, floor: power_decay_with_floor(
                x_data,
                amplitude,
                scale,
                power,
                floor,
                anchor_x=RESIDUAL_POWER_ANCHOR,
            ),
            x,
            y,
            p0=[amplitude0, 30.0, 2.0, floor0],
            bounds=([0, 1, 0.05, 0], [np.inf, 500, 20, 1]),
            maxfev=30000,
        )
        amplitude, scale, power, floor = (float(value) for value in popt)
        base_pred = np.asarray(
            power_decay_with_floor(
                x,
                amplitude,
                scale,
                power,
                floor,
                anchor_x=RESIDUAL_POWER_ANCHOR,
            ),
            dtype=float,
        )
        residual = y - base_pred
        design, stats = _residual_power_feature_design(work)
        beta = np.linalg.lstsq(design.to_numpy(dtype=float), residual, rcond=None)[0]
        residual_hat = design.to_numpy(dtype=float) @ beta
        corrected_pred = base_pred + residual_hat

        sst = float(np.sum(np.power(y - np.mean(y), 2)))
        base_sse = float(np.sum(np.power(y - base_pred, 2)))
        model_sse = float(np.sum(np.power(y - corrected_pred, 2)))
        base_rmse = float(np.sqrt(np.mean(np.power(y - base_pred, 2))))
        model_rmse = float(np.sqrt(np.mean(np.power(y - corrected_pred, 2))))
        base_target = float(
            power_decay_with_floor(
                float(target_x),
                amplitude,
                scale,
                power,
                floor,
                anchor_x=RESIDUAL_POWER_ANCHOR,
            )
        )
        coef_map = dict(zip(design.columns, beta))
        correction_target = float(coef_map["截距项"])
        row = {
            "百元残差修正幂衰减溢价率": base_target + correction_target,
            "百元幂衰减基准溢价率": base_target,
            "百元残差修正": correction_target,
            "样本数": int(len(work)),
            "基准R2": float(1 - base_sse / sst) if sst > 0 else float("nan"),
            "残差修正R2": float(1 - model_sse / sst) if sst > 0 else float("nan"),
            "R2提升": float((base_sse - model_sse) / sst) if sst > 0 else float("nan"),
            "基准RMSE": base_rmse,
            "残差修正RMSE": model_rmse,
            "拟合公式": _format_residual_power_decay_formula(
                amplitude,
                scale,
                power,
                floor,
                coef_map,
                stats,
            ),
        }
        for feature, coef in coef_map.items():
            row[f"系数_{feature}"] = float(coef)
        if include_diagnostics:
            row["_diagnostics"] = {
                "work": work,
                "base_pred": base_pred,
                "corrected_pred": corrected_pred,
                "amplitude": amplitude,
                "scale": scale,
                "power": power,
                "floor": floor,
            }
        return row
    except Exception:
        return _empty_residual_power_decay_fit_row()


def _winsorize_premium_only(
    df_subset: pd.DataFrame,
    low_q: float = 0.03,
    high_q: float = 0.97,
) -> pd.DataFrame:
    if df_subset is None or df_subset.empty or "转股溢价率" not in df_subset.columns:
        return df_subset
    low = np.nanquantile(df_subset["转股溢价率"], low_q)
    high = np.nanquantile(df_subset["转股溢价率"], high_q)
    return df_subset[
        df_subset["转股溢价率"].gt(low)
        & df_subset["转股溢价率"].lt(high)
    ].copy()


def _build_window_df(
    plain_df: pd.DataFrame,
    premium_df: pd.DataFrame,
    turnover_df: pd.DataFrame,
    date_range: pd.DatetimeIndex,
    date_pos: int,
    window_size: int,
    extra_cols: Optional[Dict[str, pd.DataFrame]] = None,
) -> pd.DataFrame:
    columns = date_range[date_pos - window_size + 1 : date_pos + 1]
    data = {
        "平价": plain_df[columns].to_numpy().flatten(),
        "转股溢价率": premium_df[columns].to_numpy().flatten(),
        "换手率": turnover_df[columns].to_numpy().flatten(),
    }
    if extra_cols:
        for name, extra_df in extra_cols.items():
            data[name] = extra_df[columns].to_numpy().flatten()
    out = pd.DataFrame(data).replace("", np.nan).dropna(axis=0, how="all")
    for column in ("平价", "转股溢价率", "换手率"):
        out[column] = pd.to_numeric(out[column], errors="coerce")
    return out.dropna(subset=["平价", "转股溢价率", "换手率"])


@dataclass(frozen=True)
class GroupRule:
    label: str
    target_parity: float
    selector: Callable[[pd.DataFrame], pd.Series]


@dataclass(frozen=True)
class CategoryRule:
    name: str
    window_size: int
    groups: tuple[GroupRule, ...]


@dataclass
class RunConfig:
    start_date: str = "2017-01-01"
    end_date: Optional[str] = None
    source_type: str = "parquet"
    excel_file_name: Optional[str] = None
    parquet_root: str = "data/转债个券历史序列"
    force_refresh: bool = False
    force_rebuild: bool = False
    result_cache_path: Path = DEFAULT_RESULT_CACHE_PATH
    output_dir: Path = DEFAULT_OUTPUT_DIR
    make_plots: bool = True


def _between(column: str, lower: float, upper: float) -> Callable[[pd.DataFrame], pd.Series]:
    return lambda frame: frame[column].gt(lower) & frame[column].lt(upper)


def _in_values(column: str, values: tuple[str, ...]) -> Callable[[pd.DataFrame], pd.Series]:
    return lambda frame: frame[column].isin(values)


def build_category_rules() -> tuple[CategoryRule, ...]:
    industry_groups = {
        "科技": ("传媒", "电子", "国防军工", "计算机", "通信"),
        "金融": ("非银金融", "银行"),
        "制造": ("电力设备", "机械设备", "汽车", "轻工制造"),
        "消费": ("农林牧渔", "纺织服饰", "家用电器", "商贸零售", "社会服务", "食品饮料", "医药生物", "美容护理"),
        "周期": ("基础化工", "钢铁", "公用事业", "环保", "建筑材料", "建筑装饰", "交通运输", "煤炭", "石油石化", "有色金属"),
    }
    return (
        CategoryRule(
            "分平价",
            5,
            (
                GroupRule("70-90", 80.0, _between("平价", 70, 90)),
                GroupRule("90-110", 100.0, _between("平价", 90, 110)),
                GroupRule("110-130", 120.0, _between("平价", 110, 130)),
                GroupRule("130-150", 140.0, _between("平价", 130, 150)),
            ),
        ),
        CategoryRule(
            "分板块",
            10,
            tuple(
                GroupRule(label, 100.0, _in_values("申万行业", values))
                for label, values in industry_groups.items()
            ),
        ),
        CategoryRule(
            "股债型",
            5,
            (
                GroupRule("偏债型", 80.0, lambda frame: frame["平价底价溢价率"].lt(-20)),
                GroupRule(
                    "平衡型",
                    100.0,
                    lambda frame: frame["平价底价溢价率"].gt(-20)
                    & frame["平价底价溢价率"].lt(20),
                ),
                GroupRule("偏股型", 120.0, lambda frame: frame["平价底价溢价率"].gt(20)),
            ),
        ),
        CategoryRule(
            "剩余期限",
            5,
            tuple(
                GroupRule(f"{lo}-{hi}", 100.0, _between("剩余期限", lo, hi))
                for lo, hi in ((0, 1), (1, 2), (2, 3), (3, 4), (4, 5), (5, 6))
            ),
        ),
        CategoryRule(
            "老券新券",
            5,
            (
                GroupRule("老券", 100.0, _between("剩余期限", 0, 5.5)),
                GroupRule("新券", 100.0, _between("剩余期限", 5.5, 6)),
            ),
        ),
        CategoryRule(
            "分评级",
            5,
            (
                GroupRule("AAA/AA+", 100.0, _in_values("债项评级", ("AAA", "AA+"))),
                GroupRule("AA/AA-", 100.0, _in_values("债项评级", ("AA", "AA-"))),
                GroupRule("A+/A", 100.0, _in_values("债项评级", ("A+", "A"))),
            ),
        ),
        CategoryRule(
            "正股市值",
            5,
            (
                GroupRule("0-50亿元", 100.0, _between("正股市值", 0, 50)),
                GroupRule("50-300亿元", 100.0, _between("正股市值", 50, 300)),
                GroupRule("300亿元以上", 100.0, lambda frame: frame["正股市值"].gt(300)),
            ),
        ),
        CategoryRule(
            "分余额",
            5,
            (
                GroupRule("0-3亿元", 100.0, _between("余额", 0, 3)),
                GroupRule("3-10亿元", 100.0, _between("余额", 3, 10)),
                GroupRule("10-20亿元", 100.0, _between("余额", 10, 20)),
                GroupRule("20-50亿元", 100.0, _between("余额", 20, 50)),
                GroupRule("50亿元以上", 100.0, lambda frame: frame["余额"].gt(50)),
            ),
        ),
    )


def _selector_signature(selector: Callable[[pd.DataFrame], pd.Series]) -> dict:
    closure_values = []
    if selector.__closure__:
        closure_values = [repr(cell.cell_contents) for cell in selector.__closure__]
    return {
        "bytecode": selector.__code__.co_code.hex(),
        "names": list(selector.__code__.co_names),
        "constants": [repr(value) for value in selector.__code__.co_consts],
        "closure": closure_values,
    }


def _model_signature_payload(include_all_market: bool = True) -> dict:
    """构建模型签名载荷；可生成扩展前的兼容签名。"""
    categories = build_category_rules()
    payload = {
        "schema_version": MODEL_CACHE_SCHEMA_VERSION,
        "model": "幂衰减基准曲线 + 六因子标准化 OLS 修正",
        "residual_features": RESIDUAL_FEATURES,
        "minimum_samples": RESIDUAL_MIN_SAMPLES,
        "parity_range": [RESIDUAL_POWER_LOWER, RESIDUAL_POWER_UPPER],
        "power_anchor": RESIDUAL_POWER_ANCHOR,
        "power_bounds": [[0, 1, 0.05, 0], ["inf", 500, 20, 1]],
        "curve_fit_maxfev": 30000,
        "turnover_upper": 50,
        "premium_trim_quantiles": [0.03, 0.97],
        "categories": [
            {
                "name": category.name,
                "window_size": category.window_size,
                "groups": [
                    {
                        "label": group.label,
                        "target_parity": group.target_parity,
                        "selector": _selector_signature(group.selector),
                    }
                    for group in category.groups
                ],
            }
            for category in categories
        ],
    }
    if include_all_market:
        payload["all_market_hundred_fits"] = {
            "window_size": 1,
            "target_parity": 100.0,
            "inverse_cubic": {
                "formula": "a/x^3+b/x^2+c/x+d",
                "minimum_samples": INVERSE_CUBIC_MIN_SAMPLES,
                "parity_range": [70, 130],
                "curve_fit_maxfev": 20000,
            },
            "multifactor": {
                "formula": "幂衰减基准曲线 + 六因子标准化 OLS 修正",
                "minimum_samples": RESIDUAL_MIN_SAMPLES,
                "parity_range": [RESIDUAL_POWER_LOWER, RESIDUAL_POWER_UPPER],
            },
            "turnover_upper": 50,
            "premium_trim_quantiles": [0.03, 0.97],
        }
    return payload


def _signature_from_payload(payload: dict) -> str:
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _model_signature() -> str:
    """模型、分组及清洗口径签名；口径变化后旧缓存自动失效。"""
    return _signature_from_payload(_model_signature_payload(include_all_market=True))


def _legacy_model_signature() -> str:
    """扩展两条百元序列前的签名，用于无损沿用已有31项历史缓存。"""
    return _signature_from_payload(_model_signature_payload(include_all_market=False))


def _weekly_maturity_model_signature() -> str:
    """剩余期限周分组的独立缓存签名。"""
    payload = {
        "schema_version": WEEKLY_MATURITY_CACHE_SCHEMA_VERSION,
        # 两条全市场百元序列不影响期限周分组，沿用原分组模型签名，避免误使周缓存失效。
        "base_model_signature": _legacy_model_signature(),
        "bucket_unit": "week",
        "max_weeks": WEEKLY_MATURITY_MAX_WEEKS,
        "target_parity": 100.0,
        "sample_period": "RunConfig.start_date~RunConfig.end_date",
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _cache_metadata_path(cache_path: Path) -> Path:
    return cache_path.with_suffix(".json")


def _normalize_result_frame(result: pd.DataFrame) -> pd.DataFrame:
    if result is None or result.empty:
        return pd.DataFrame()
    out = result.copy()
    out = out.rename(
        columns={
            "残差修正": "多因子修正",
            "残差修正R2": "多因子修正R2",
            "残差修正RMSE": "多因子修正RMSE",
        }
    )
    required = {"日期", "分类", "分组", "拟合溢价率"}
    missing = sorted(required - set(out.columns))
    if missing:
        raise ValueError(f"拟合结果缓存缺少字段：{missing}")
    out["日期"] = pd.to_datetime(out["日期"], errors="coerce")
    out = out.dropna(subset=["日期", "分类", "分组"])
    out = out.drop_duplicates(["日期", "分类", "分组"], keep="last")
    return out.sort_values(["分类", "分组", "日期"], kind="stable").reset_index(drop=True)


def _load_result_cache(config: RunConfig) -> pd.DataFrame:
    if config.force_rebuild:
        print("[result_cache] 强制更新已开启，忽略历史拟合缓存。")
        return pd.DataFrame()

    cache_path = config.result_cache_path
    metadata_path = _cache_metadata_path(cache_path)
    cache_artifact_exists = cache_path.exists() or metadata_path.exists()
    if cache_artifact_exists:
        if not cache_path.exists() or not metadata_path.exists():
            print("[result_cache] 缓存文件不完整，执行全量重建。")
            return pd.DataFrame()
        try:
            metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
            cached_signature = metadata.get("model_signature")
            if cached_signature not in {_model_signature(), _legacy_model_signature()}:
                print("[result_cache] 模型或分组口径已变化，旧缓存失效，执行全量重建。")
                return pd.DataFrame()
            cached = _normalize_result_frame(pd.read_parquet(cache_path))
            compatibility_note = (
                "｜将补算新增百元序列"
                if cached_signature == _legacy_model_signature()
                else ""
            )
            print(
                f"[result_cache] 命中历史拟合缓存：{cache_path}｜"
                f"{cached['日期'].min():%Y-%m-%d}~{cached['日期'].max():%Y-%m-%d}｜"
                f"{len(cached)}行{compatibility_note}"
            )
            return cached
        except Exception as exc:
            print(f"[result_cache] 缓存读取失败，执行全量重建：{exc}")
            return pd.DataFrame()

    if LEGACY_RESULT_PATH.exists():
        try:
            legacy = _normalize_result_frame(pd.read_csv(LEGACY_RESULT_PATH))
            print(
                f"[result_cache] 使用现有历史明细初始化：{LEGACY_RESULT_PATH}｜"
                f"{legacy['日期'].min():%Y-%m-%d}~{legacy['日期'].max():%Y-%m-%d}｜{len(legacy)}行"
            )
            return legacy
        except Exception as exc:
            print(f"[result_cache] 历史明细初始化失败，执行全量重建：{exc}")
    else:
        print("[result_cache] 未找到历史拟合缓存，执行全量重建。")
    return pd.DataFrame()


def _write_result_cache(result: pd.DataFrame, config: RunConfig) -> None:
    result = _normalize_result_frame(result)
    if result.empty:
        raise ValueError("拟合结果为空，拒绝写入历史缓存。")
    cache_path = config.result_cache_path
    metadata_path = _cache_metadata_path(cache_path)
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    temp_cache = cache_path.with_name(f".{cache_path.stem}.tmp.parquet")
    temp_metadata = metadata_path.with_name(f".{metadata_path.stem}.tmp.json")
    result.to_parquet(temp_cache, index=False)
    metadata = {
        "model_signature": _model_signature(),
        "schema_version": MODEL_CACHE_SCHEMA_VERSION,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "start_date": str(result["日期"].min().date()),
        "end_date": str(result["日期"].max().date()),
        "rows": int(len(result)),
        "dates": int(result["日期"].nunique()),
    }
    temp_metadata.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temp_cache, cache_path)
    os.replace(temp_metadata, metadata_path)
    print(f"[result_cache] 历史拟合缓存已更新：{cache_path}")


def _index_schema() -> pa.Schema:
    return pa.schema(
        [
            pa.field(INDEX_NAME, pa.string(), nullable=False),
            pa.field(TRADE_DATE, pa.date32(), nullable=False),
            pa.field(INDEX_VALUE, pa.float64()),
        ],
        metadata={
            b"schema_version": PARQUET_SCHEMA_VERSION.encode("utf-8"),
            b"dataset_type": b"market_index_history",
            b"primary_key": f"{INDEX_NAME},{TRADE_DATE}".encode("utf-8"),
        },
    )


def _write_index_parquet_atomic(frame: pd.DataFrame, path: Path) -> None:
    """按标准指数 Schema 原子写入，并在替换前完整回读校验。"""
    schema = _index_schema()
    if list(frame.columns) != schema.names:
        raise ValueError(f"指数字段不符合标准 Schema：{list(frame.columns)}")
    arrays = [
        pa.array(frame[field.name], type=field.type, from_pandas=True, safe=True)
        for field in schema
    ]
    table = pa.Table.from_arrays(arrays, schema=schema)
    fd, temp_name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        pq.write_table(
            table,
            temp_path,
            compression="zstd",
            use_dictionary=True,
            row_group_size=128_000,
        )
        with temp_path.open("rb") as handle:
            reread = pq.read_table(handle)
        if reread.schema != schema or reread.num_rows != len(frame):
            raise RuntimeError("指数 parquet 回读校验失败")
        del reread
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def _backup_before_index_write(parquet_root: Path, index_path: Path) -> None:
    """保留更新前的指数与 manifest 最近版本，便于回退。"""
    RESULT_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(index_path, INDEX_BACKUP_PATH)
    manifest_path = parquet_root / "_meta" / "sheet_manifest.parquet"
    if manifest_path.exists():
        shutil.copy2(manifest_path, INDEX_MANIFEST_BACKUP_PATH)


def _update_index_manifest(parquet_root: Path, expected_rows: int) -> None:
    """同步指数长表在 sheet_manifest 中登记的行数。"""
    manifest_path = parquet_root / "_meta" / "sheet_manifest.parquet"
    if not manifest_path.exists():
        print("[index_write] 未找到 sheet_manifest，跳过行数同步。")
        return
    manifest = pd.read_parquet(manifest_path)
    mask = manifest["sheet_name"].eq("指数")
    if int(mask.sum()) != 1:
        raise RuntimeError("sheet_manifest 中无法唯一定位指数记录")
    if int(manifest.loc[mask, "rows"].iloc[0]) == expected_rows:
        return
    manifest.loc[mask, "rows"] = expected_rows
    manifest["sheet_name"] = manifest["sheet_name"].astype("string")
    manifest["route"] = manifest["route"].astype("string")
    manifest["rows"] = manifest["rows"].astype("int64")
    manifest["cols"] = manifest["cols"].astype("int64")
    manifest_schema = pa.schema(
        [
            pa.field("sheet_name", pa.string()),
            pa.field("route", pa.string()),
            pa.field("rows", pa.int64()),
            pa.field("cols", pa.int64()),
        ],
        metadata={
            b"schema_version": PARQUET_SCHEMA_VERSION.encode("utf-8"),
            b"dataset_type": b"dataset_manifest",
            b"primary_key": b"sheet_name",
        },
    )
    temp_path = manifest_path.with_name(f".{manifest_path.name}.tmp")
    try:
        arrays = [
            pa.array(manifest[field.name], type=field.type, from_pandas=True, safe=True)
            for field in manifest_schema
        ]
        pq.write_table(
            pa.Table.from_arrays(arrays, schema=manifest_schema),
            temp_path,
            compression="zstd",
            use_dictionary=True,
        )
        with temp_path.open("rb") as handle:
            reread_table = pq.read_table(handle)
        if reread_table.schema != manifest_schema:
            raise RuntimeError("sheet_manifest Schema 回读校验失败")
        reread = reread_table.to_pandas()
        actual = int(reread.loc[reread["sheet_name"].eq("指数"), "rows"].iloc[0])
        if actual != expected_rows:
            raise RuntimeError(f"sheet_manifest 指数行数校验失败：{actual} != {expected_rows}")
        os.replace(temp_path, manifest_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()


def _daily_fit_index_name(category: str, group: str) -> str:
    """生成稳定指数名称；保留已经写入的5个分板块名称不变。"""
    if category == ALL_MARKET_CATEGORY:
        return ALL_MARKET_INDEX_NAMES[group]
    if category == "分板块":
        return SECTOR_FIT_INDEX_NAMES[group]
    return f"多因子修正拟合溢价率_{category}_{group}"


def _expected_daily_pairs() -> set[tuple[str, str]]:
    pairs = {
        (category.name, group.label)
        for category in build_category_rules()
        for group in category.groups
    }
    pairs.update((ALL_MARKET_CATEGORY, group) for group in ALL_MARKET_GROUPS)
    return pairs


def _daily_fit_index_name_map() -> dict[tuple[str, str], str]:
    result = {
        (category.name, group.label): _daily_fit_index_name(category.name, group.label)
        for category in build_category_rules()
        for group in category.groups
    }
    result.update(
        {
            (ALL_MARKET_CATEGORY, group): _daily_fit_index_name(ALL_MARKET_CATEGORY, group)
            for group in ALL_MARKET_GROUPS
        }
    )
    return result


def write_daily_fits_to_index(result: pd.DataFrame, config: RunConfig) -> dict:
    """将全部日度拟合溢价率序列并入标准指数长表。"""
    if not WRITE_DAILY_FITS_TO_INDEX:
        print("[index_write] 写入开关已关闭，跳过日度拟合结果并入指数。")
        return {"written": False, "reason": "disabled"}

    index_path = Path(config.parquet_root) / "_special" / "指数.parquet"
    if not index_path.exists():
        raise FileNotFoundError(f"未找到指数 parquet：{index_path}")

    name_map = _daily_fit_index_name_map()
    series_count = len(name_map)
    expected_pairs = set(name_map)
    daily = result.loc[
        result[["分类", "分组"]].apply(tuple, axis=1).isin(expected_pairs),
        ["日期", "分组", "拟合溢价率"],
    ].copy()
    daily.insert(1, "分类", result.loc[daily.index, "分类"])
    daily[TRADE_DATE] = pd.to_datetime(daily.pop("日期"), errors="coerce").dt.normalize()
    daily[INDEX_NAME] = [
        name_map.get((category, group))
        for category, group in zip(daily.pop("分类"), daily.pop("分组"))
    ]
    daily[INDEX_VALUE] = pd.to_numeric(daily.pop("拟合溢价率"), errors="coerce")
    daily = daily.dropna(subset=[INDEX_NAME, TRADE_DATE, INDEX_VALUE])
    daily = daily.loc[daily[INDEX_VALUE].ne(0), [INDEX_NAME, TRADE_DATE, INDEX_VALUE]]
    daily = daily.drop_duplicates([INDEX_NAME, TRADE_DATE], keep="last")
    if daily.empty:
        raise ValueError("日度分组拟合结果为空，拒绝更新指数 parquet。")

    actual_names = set(daily[INDEX_NAME].unique())
    expected_names = set(name_map.values())
    if not actual_names.issubset(expected_names):
        raise ValueError(f"发现未登记的日度拟合指数：{sorted(actual_names - expected_names)}")

    current = pd.read_parquet(index_path)
    if list(current.columns) != [INDEX_NAME, TRADE_DATE, INDEX_VALUE]:
        raise ValueError(f"指数 parquet 字段异常：{list(current.columns)}")
    current[TRADE_DATE] = pd.to_datetime(current[TRADE_DATE], errors="coerce").dt.normalize()
    if current[[INDEX_NAME, TRADE_DATE]].isna().any().any():
        raise ValueError("指数 parquet 主键存在空值")
    if current.duplicated([INDEX_NAME, TRADE_DATE]).any():
        raise ValueError("指数 parquet 存在重复主键")

    target_names = expected_names
    current_targets = current.loc[current[INDEX_NAME].isin(target_names)].copy()
    old_values = current_targets.set_index([INDEX_NAME, TRADE_DATE])[INDEX_VALUE]
    new_values = daily.set_index([INDEX_NAME, TRADE_DATE])[INDEX_VALUE]
    common_keys = old_values.index.intersection(new_values.index)
    changed = int(
        (~np.isclose(
            old_values.loc[common_keys].to_numpy(dtype=float),
            new_values.loc[common_keys].to_numpy(dtype=float),
            rtol=0.0,
            atol=1e-12,
            equal_nan=True,
        )).sum()
    )
    added = int(len(new_values.index.difference(old_values.index)))
    if changed == 0 and added == 0:
        print(
            f"[index_write] {series_count}条日度拟合指数已是最新｜"
            f"{daily[TRADE_DATE].min():%Y-%m-%d}~{daily[TRADE_DATE].max():%Y-%m-%d}｜"
            f"{len(daily)}条"
        )
        return {"written": False, "reason": "unchanged", "rows": int(len(daily))}

    current_indexed = current.set_index([INDEX_NAME, TRADE_DATE])
    daily_indexed = daily.set_index([INDEX_NAME, TRADE_DATE])
    current_indexed.loc[common_keys, INDEX_VALUE] = daily_indexed.loc[
        common_keys, INDEX_VALUE
    ].to_numpy()
    new_keys = daily_indexed.index.difference(current_indexed.index)
    updated = pd.concat(
        [current_indexed, daily_indexed.loc[new_keys]],
        axis=0,
    ).reset_index()
    updated[INDEX_NAME] = updated[INDEX_NAME].astype("string")
    updated[INDEX_VALUE] = pd.to_numeric(updated[INDEX_VALUE], errors="coerce").astype("float64")
    updated = updated.dropna(subset=[INDEX_NAME, TRADE_DATE, INDEX_VALUE])
    updated = updated.loc[updated[INDEX_VALUE].ne(0)]
    updated = updated.sort_values([TRADE_DATE, INDEX_NAME], kind="stable").reset_index(drop=True)
    if updated.duplicated([INDEX_NAME, TRADE_DATE]).any():
        raise RuntimeError("拟合结果并入后指数 parquet 出现重复主键")
    if not set(daily.set_index([INDEX_NAME, TRADE_DATE]).index).issubset(
        set(updated.set_index([INDEX_NAME, TRADE_DATE]).index)
    ):
        raise RuntimeError("拟合结果并入后目标数据不完整")

    parquet_root = Path(config.parquet_root)
    _backup_before_index_write(parquet_root, index_path)
    _write_index_parquet_atomic(updated, index_path)
    _update_index_manifest(parquet_root, len(updated))
    print(
        f"[index_write] 日度拟合结果已并入指数｜{series_count}项｜{len(daily)}条｜"
        f"新增 {added}｜变更 {changed}｜{index_path}"
    )
    return {
        "written": True,
        "rows": int(len(daily)),
        "added": added,
        "changed": changed,
        "path": str(index_path),
    }


def _normalize_weekly_maturity_result(result: pd.DataFrame) -> pd.DataFrame:
    if result is None or result.empty:
        return pd.DataFrame()
    out = result.copy().rename(
        columns={
            "残差修正": "多因子修正",
            "残差修正R2": "多因子修正R2",
            "残差修正RMSE": "多因子修正RMSE",
        }
    )
    required = {"剩余期限（周）", "拟合溢价率"}
    missing = sorted(required - set(out.columns))
    if missing:
        raise ValueError(f"剩余期限周分组缓存缺少字段：{missing}")
    out["周序号"] = pd.to_numeric(out.get("周序号"), errors="coerce")
    if out["周序号"].isna().any():
        extracted = out["剩余期限（周）"].astype(str).str.extract(r"^(\d+)-")[0]
        out["周序号"] = out["周序号"].fillna(pd.to_numeric(extracted, errors="coerce"))
    return out.sort_values("周序号", kind="stable").reset_index(drop=True)


def _load_weekly_maturity_cache(
    config: RunConfig,
    start: pd.Timestamp,
    end: pd.Timestamp,
) -> pd.DataFrame:
    if config.force_rebuild:
        print("[weekly_cache] 强制更新已开启，重新计算剩余期限周分组。")
        return pd.DataFrame()

    cache_path = DEFAULT_WEEKLY_MATURITY_CACHE_PATH
    metadata_path = _cache_metadata_path(cache_path)
    if not cache_path.exists() or not metadata_path.exists():
        print("[weekly_cache] 未找到剩余期限周分组缓存，本次重新计算。")
        return pd.DataFrame()
    try:
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        cache_matches = (
            metadata.get("model_signature") == _weekly_maturity_model_signature()
            and metadata.get("start_date") == str(start.date())
            and metadata.get("end_date") == str(end.date())
        )
        if not cache_matches:
            print("[weekly_cache] 模型口径或样本期间已变化，本次重新计算。")
            return pd.DataFrame()
        cached = _normalize_weekly_maturity_result(pd.read_parquet(cache_path))
        print(
            f"[weekly_cache] 命中剩余期限周分组缓存：{cache_path}｜"
            f"{start:%Y-%m-%d}~{end:%Y-%m-%d}｜{len(cached)}组"
        )
        return cached
    except Exception as exc:
        print(f"[weekly_cache] 缓存读取失败，本次重新计算：{exc}")
        return pd.DataFrame()


def _write_weekly_maturity_cache(
    result: pd.DataFrame,
    start: pd.Timestamp,
    end: pd.Timestamp,
) -> None:
    normalized = _normalize_weekly_maturity_result(result)
    if normalized.empty:
        raise ValueError("剩余期限周分组结果为空，拒绝写入缓存。")
    cache_path = DEFAULT_WEEKLY_MATURITY_CACHE_PATH
    metadata_path = _cache_metadata_path(cache_path)
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    temp_cache = cache_path.with_name(f".{cache_path.stem}.tmp.parquet")
    temp_metadata = metadata_path.with_name(f".{metadata_path.stem}.tmp.json")
    normalized.to_parquet(temp_cache, index=False)
    metadata = {
        "model_signature": _weekly_maturity_model_signature(),
        "schema_version": WEEKLY_MATURITY_CACHE_SCHEMA_VERSION,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "start_date": str(start.date()),
        "end_date": str(end.date()),
        "groups": int(len(normalized)),
    }
    temp_metadata.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temp_cache, cache_path)
    os.replace(temp_metadata, metadata_path)
    print(f"[weekly_cache] 剩余期限周分组缓存已更新：{cache_path}")


def _complete_result_dates(result: pd.DataFrame) -> set[pd.Timestamp]:
    if result is None or result.empty:
        return set()
    expected_pairs = _expected_daily_pairs()
    normalized = _normalize_result_frame(result)
    normalized = normalized[
        normalized[["分类", "分组"]].apply(tuple, axis=1).isin(expected_pairs)
    ]
    expected_count = len(expected_pairs)
    counts = normalized.groupby("日期", sort=False).size()
    return {pd.Timestamp(date) for date, count in counts.items() if int(count) == expected_count}


def _normalize_frame(frame: pd.DataFrame) -> pd.DataFrame:
    out = _normalize_columns_to_ts(frame.copy())
    out.index = out.index.map(str)
    return out


def _prepare_data(config: RunConfig) -> tuple[dict[str, pd.DataFrame], pd.Series, pd.DatetimeIndex]:
    data = load_original_data(
        source_type=config.source_type,
        excel_file_name=config.excel_file_name,
        parquet_root=config.parquet_root,
        force_refresh=config.force_refresh,
    )
    required_sheets = {
        "平价",
        "转股溢价率",
        "换手率",
        "剩余期限",
        "余额",
        "正股20日波动率",
        "赎回累计天数",
        "下修累计天数",
        "隐含波动率",
        "平价底价溢价率",
        "债项评级",
        "正股市值",
        "总表",
    }
    missing = sorted(required_sheets - set(data))
    if missing:
        raise KeyError(f"缺少必要数据表：{missing}")

    frame_names = sorted(required_sheets - {"总表"})
    frames = {name: _normalize_frame(data[name]) for name in frame_names}
    common_index = frames["平价"].index
    for frame in frames.values():
        common_index = common_index.intersection(frame.index)

    all_dates = _extract_date_columns(frames["平价"])
    for name, frame in frames.items():
        frames[name] = frame.reindex(index=common_index, columns=all_dates)

    total = data["总表"].copy()
    total.index = total.index.map(str)
    if "申万行业" not in total.columns:
        raise KeyError("总表缺少 `申万行业` 列")
    industry = total.reindex(common_index)["申万行业"]
    return frames, industry, all_dates


def _build_window_base(
    frames: dict[str, pd.DataFrame],
    industry: pd.Series,
    all_dates: pd.DatetimeIndex,
    date_pos: int,
    window_size: int,
) -> pd.DataFrame:
    window_dates = all_dates[max(0, date_pos - window_size + 1) : date_pos + 1]
    industry_window = pd.DataFrame(
        np.repeat(industry.to_numpy()[:, None], len(window_dates), axis=1),
        index=industry.index,
        columns=window_dates,
    )
    extra_cols = {
        "剩余期限": frames["剩余期限"],
        "余额": frames["余额"],
        "正股20日波动率": frames["正股20日波动率"],
        "赎回累计天数": frames["赎回累计天数"],
        "下修累计天数": frames["下修累计天数"],
        "隐含波动率": frames["隐含波动率"],
        "平价底价溢价率": frames["平价底价溢价率"],
        "债项评级": frames["债项评级"],
        "正股市值": frames["正股市值"],
        "申万行业": industry_window,
    }
    base = _build_window_df(
        frames["平价"],
        frames["转股溢价率"],
        frames["换手率"],
        all_dates,
        date_pos,
        window_size=window_size,
        extra_cols=extra_cols,
    )
    for column in ("剩余期限", "余额", "平价底价溢价率", "正股市值"):
        base[column] = pd.to_numeric(base[column], errors="coerce")
    return base


def _fit_group(
    base: pd.DataFrame,
    category: CategoryRule,
    group: GroupRule,
) -> dict:
    common_mask = (
        base["平价"].gt(RESIDUAL_POWER_LOWER)
        & base["平价"].lt(RESIDUAL_POWER_UPPER)
        & base["换手率"].lt(50)
    )
    selected = base.loc[common_mask & group.selector(base)].copy()
    selected = _winsorize_premium_only(selected)
    detail = _fit_residual_power_decay_detail(
        selected,
        target_x=group.target_parity,
        high_price_subset=None,
    )
    return {
        "分类": category.name,
        "分组": group.label,
        "目标平价": group.target_parity,
        "拟合溢价率": detail["百元残差修正幂衰减溢价率"],
        "幂衰减基准溢价率": detail["百元幂衰减基准溢价率"],
        "多因子修正": detail["百元残差修正"],
        "样本数": detail["样本数"],
        "基准R2": detail["基准R2"],
        "多因子修正R2": detail["残差修正R2"],
        "R2提升": detail["R2提升"],
        "基准RMSE": detail["基准RMSE"],
        "多因子修正RMSE": detail["残差修正RMSE"],
        "拟合公式": detail["拟合公式"],
    }


def _fit_all_market_hundred(base: pd.DataFrame) -> dict[str, dict]:
    """在单日全市场截面上计算反三次和多因子修正两条百元序列。"""
    turnover_mask = base["换手率"].lt(50)

    inverse_sample = base.loc[
        turnover_mask & base["平价"].gt(70) & base["平价"].lt(130)
    ].copy()
    inverse_sample = _winsorize_premium_only(inverse_sample)
    inverse_detail = _fit_inverse_cubic_hundred_detail(inverse_sample)

    multifactor_sample = base.loc[
        turnover_mask
        & base["平价"].gt(RESIDUAL_POWER_LOWER)
        & base["平价"].lt(RESIDUAL_POWER_UPPER)
    ].copy()
    multifactor_sample = _winsorize_premium_only(multifactor_sample)
    multifactor_detail = _fit_residual_power_decay_detail(
        multifactor_sample,
        target_x=100.0,
        high_price_subset=None,
    )

    return {
        INVERSE_CUBIC_HUNDRED_GROUP: {
            "分类": ALL_MARKET_CATEGORY,
            "分组": INVERSE_CUBIC_HUNDRED_GROUP,
            "目标平价": 100.0,
            "拟合溢价率": inverse_detail["拟合溢价率"],
            "幂衰减基准溢价率": float("nan"),
            "多因子修正": float("nan"),
            "样本数": inverse_detail["样本数"],
            "拟合公式": inverse_detail["拟合公式"],
        },
        MULTIFACTOR_HUNDRED_GROUP: {
            "分类": ALL_MARKET_CATEGORY,
            "分组": MULTIFACTOR_HUNDRED_GROUP,
            "目标平价": 100.0,
            "拟合溢价率": multifactor_detail["百元残差修正幂衰减溢价率"],
            "幂衰减基准溢价率": multifactor_detail["百元幂衰减基准溢价率"],
            "多因子修正": multifactor_detail["百元残差修正"],
            "样本数": multifactor_detail["样本数"],
            "基准R2": multifactor_detail["基准R2"],
            "多因子修正R2": multifactor_detail["残差修正R2"],
            "R2提升": multifactor_detail["R2提升"],
            "基准RMSE": multifactor_detail["基准RMSE"],
            "多因子修正RMSE": multifactor_detail["残差修正RMSE"],
            "拟合公式": multifactor_detail["拟合公式"],
        },
    }


def _result_pairs_by_date(result: pd.DataFrame) -> dict[pd.Timestamp, set[tuple[str, str]]]:
    if result is None or result.empty:
        return {}
    normalized = _normalize_result_frame(result)
    expected_pairs = _expected_daily_pairs()
    pair_map: dict[pd.Timestamp, set[tuple[str, str]]] = {}
    for date, part in normalized.groupby("日期", sort=False):
        pairs = set(part[["分类", "分组"]].itertuples(index=False, name=None))
        pair_map[pd.Timestamp(date)] = pairs & expected_pairs
    return pair_map


def calculate_group_series(
    config: RunConfig,
    existing_result: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    existing = _normalize_result_frame(existing_result)
    complete_dates = _complete_result_dates(existing)
    start = pd.Timestamp(config.start_date)
    quick_dates = _quick_source_dates(config)
    if quick_dates is not None:
        quick_end = pd.Timestamp(config.end_date) if config.end_date else quick_dates.max()
        quick_output_dates = quick_dates[(quick_dates >= start) & (quick_dates <= quick_end)]
        if len(quick_output_dates) > 0 and all(
            pd.Timestamp(date) in complete_dates for date in quick_output_dates
        ):
            print(
                f"[result_cache] {quick_output_dates.min():%Y-%m-%d}~{quick_output_dates.max():%Y-%m-%d} "
                "均已有完整结果，本次跳过原始矩阵载入和拟合。"
            )
            return existing

    frames, industry, all_dates = _prepare_data(config)
    end = pd.Timestamp(config.end_date) if config.end_date else all_dates.max()
    output_dates = all_dates[(all_dates >= start) & (all_dates <= end)]
    if output_dates.empty:
        raise ValueError(f"指定期间没有交易日：{start:%Y-%m-%d} 至 {end:%Y-%m-%d}")
    expected_pairs = _expected_daily_pairs()
    existing_pairs_by_date = _result_pairs_by_date(existing)
    missing_pairs_by_date = {
        pd.Timestamp(date): expected_pairs - existing_pairs_by_date.get(pd.Timestamp(date), set())
        for date in output_dates
    }
    dates_to_compute = pd.DatetimeIndex(
        [date for date in output_dates if missing_pairs_by_date[pd.Timestamp(date)]]
    )
    if len(dates_to_compute) == 0:
        print(
            f"[result_cache] {output_dates.min():%Y-%m-%d}~{output_dates.max():%Y-%m-%d} "
            "均已有完整结果，本次跳过拟合。"
        )
        return existing
    print(
        f"[result_cache] 本次需计算 {len(dates_to_compute)}/{len(output_dates)} 个交易日｜"
        f"{dates_to_compute.min():%Y-%m-%d}~{dates_to_compute.max():%Y-%m-%d}"
    )

    date_positions = {date: pos for pos, date in enumerate(all_dates)}
    categories = build_category_rules()
    records: List[dict] = []
    group_timings: dict[tuple[str, str], list[float]] = {}
    total_tasks = sum(len(missing_pairs_by_date[pd.Timestamp(date)]) for date in dates_to_compute)
    with tqdm(
        total=total_tasks,
        desc="日度拟合",
        unit="group",
        ncols=140,
    ) as progress:
        for date in dates_to_compute:
            pos = date_positions[date]
            missing_pairs = missing_pairs_by_date[pd.Timestamp(date)]
            required_windows = {
                category.window_size
                for category in categories
                if any((category.name, group.label) in missing_pairs for group in category.groups)
            }
            if any((ALL_MARKET_CATEGORY, group) in missing_pairs for group in ALL_MARKET_GROUPS):
                required_windows.add(1)
            bases = {
                window: _build_window_base(frames, industry, all_dates, pos, window)
                for window in sorted(required_windows)
            }
            for category in categories:
                category_missing = [
                    group
                    for group in category.groups
                    if (category.name, group.label) in missing_pairs
                ]
                if not category_missing:
                    continue
                base = bases[category.window_size]
                for group in category_missing:
                    started_at = perf_counter()
                    row = {"日期": pd.Timestamp(date)}
                    row.update(_fit_group(base, category, group))
                    elapsed = perf_counter() - started_at
                    records.append(row)

                    timing_key = (category.name, group.label)
                    group_timings.setdefault(timing_key, []).append(elapsed)
                    progress.set_postfix_str(
                        f"{date:%Y-%m-%d}｜{category.name}/{group.label}｜{elapsed:.3f}s",
                        refresh=False,
                    )
                    progress.update(1)

            all_market_missing = [
                group
                for group in ALL_MARKET_GROUPS
                if (ALL_MARKET_CATEGORY, group) in missing_pairs
            ]
            if all_market_missing:
                started_at = perf_counter()
                all_market_rows = _fit_all_market_hundred(bases[1])
                shared_elapsed = perf_counter() - started_at
                for group in all_market_missing:
                    row = {"日期": pd.Timestamp(date)}
                    row.update(all_market_rows[group])
                    records.append(row)
                    timing_key = (ALL_MARKET_CATEGORY, group)
                    group_timings.setdefault(timing_key, []).append(shared_elapsed)
                    progress.set_postfix_str(
                        f"{date:%Y-%m-%d}｜{ALL_MARKET_CATEGORY}/{group}｜{shared_elapsed:.3f}s",
                        refresh=False,
                    )
                    progress.update(1)

    timing_summary = sorted(
        (
            (sum(values) / len(values), sum(values), len(values), category, group)
            for (category, group), values in group_timings.items()
        ),
        reverse=True,
    )
    if timing_summary:
        print("[timing] 平均耗时最慢的10个分组：")
        for average, total, count, category, group in timing_summary[:10]:
            print(
                f"  {category}/{group}｜平均 {average:.3f}s｜"
                f"累计 {total:.2f}s｜计算 {count}次"
            )

    if not records:
        return existing
    new_result = pd.DataFrame(records)
    combined = pd.concat([existing, new_result], ignore_index=True, sort=False)
    return _normalize_result_frame(combined)


def _build_weekly_maturity_cross_section(
    frames: dict[str, pd.DataFrame],
    sample_dates: pd.DatetimeIndex,
) -> pd.DataFrame:
    """将指定期间的债券日样本汇总为剩余期限周分档截面。"""
    source_columns = [
        "平价",
        "转股溢价率",
        "换手率",
        "剩余期限",
        "余额",
        "正股20日波动率",
        "赎回累计天数",
        "下修累计天数",
        "隐含波动率",
    ]
    flattened: dict[str, np.ndarray] = {}
    for column in source_columns:
        values = frames[column].loc[:, sample_dates].to_numpy(copy=False).reshape(-1)
        flattened[column] = pd.to_numeric(pd.Series(values, copy=False), errors="coerce").to_numpy()

    parity = flattened["平价"]
    premium = flattened["转股溢价率"]
    turnover = flattened["换手率"]
    maturity = flattened["剩余期限"]
    common_mask = (
        np.isfinite(parity)
        & np.isfinite(premium)
        & np.isfinite(turnover)
        & np.isfinite(maturity)
        & (parity > RESIDUAL_POWER_LOWER)
        & (parity < RESIDUAL_POWER_UPPER)
        & (turnover < 50)
        & (maturity >= 0)
        & (maturity * 52 < WEEKLY_MATURITY_MAX_WEEKS)
    )
    cross_section = pd.DataFrame(
        {column: values[common_mask] for column, values in flattened.items()}
    )
    cross_section["周序号"] = np.floor(cross_section["剩余期限"] * 52).astype(int)
    return cross_section


def calculate_weekly_maturity_fit(
    config: RunConfig,
) -> tuple[pd.DataFrame, pd.Timestamp, pd.Timestamp]:
    """按剩余期限逐周分档，在完整样本期间上做一次截面多因子拟合。"""
    prepared: Optional[tuple[dict[str, pd.DataFrame], pd.Series, pd.DatetimeIndex]] = None
    quick_dates = _quick_source_dates(config)
    if quick_dates is None or len(quick_dates) == 0:
        prepared = _prepare_data(config)
        available_dates = prepared[2]
    else:
        available_dates = quick_dates

    start = pd.Timestamp(config.start_date)
    end = pd.Timestamp(config.end_date) if config.end_date else available_dates.max()
    sample_dates = available_dates[(available_dates >= start) & (available_dates <= end)]
    if sample_dates.empty:
        raise ValueError(f"剩余期限周分组指定期间没有交易日：{start:%Y-%m-%d} 至 {end:%Y-%m-%d}")
    start = pd.Timestamp(sample_dates.min())
    end = pd.Timestamp(sample_dates.max())

    cached = _load_weekly_maturity_cache(config, start, end)
    if not cached.empty:
        return cached, start, end

    if prepared is None:
        prepared = _prepare_data(config)
    frames, _, all_dates = prepared
    sample_dates = all_dates[(all_dates >= start) & (all_dates <= end)]
    cross_section = _build_weekly_maturity_cross_section(frames, sample_dates)
    grouped = cross_section.groupby("周序号", sort=False)

    rows: List[dict] = []
    for week in tqdm(
        range(WEEKLY_MATURITY_MAX_WEEKS),
        desc="剩余期限周分组多因子修正拟合",
        unit="week",
    ):
        if week in grouped.indices:
            selected = grouped.get_group(week).drop(columns="周序号")
            selected = _winsorize_premium_only(selected)
            detail = _fit_residual_power_decay_detail(
                selected,
                target_x=100.0,
                high_price_subset=None,
            )
        else:
            detail = _empty_residual_power_decay_fit_row()
        rows.append(
            {
                "周序号": week,
                "剩余期限（周）": f"{week}-{week + 1}",
                "拟合溢价率": detail["百元残差修正幂衰减溢价率"],
                "幂衰减基准溢价率": detail["百元幂衰减基准溢价率"],
                "多因子修正": detail["百元残差修正"],
                "样本数": detail["样本数"],
                "基准R2": detail["基准R2"],
                "多因子修正R2": detail["残差修正R2"],
                "R2提升": detail["R2提升"],
                "基准RMSE": detail["基准RMSE"],
                "多因子修正RMSE": detail["残差修正RMSE"],
                "拟合公式": detail["拟合公式"],
            }
        )
    result = _normalize_weekly_maturity_result(pd.DataFrame(rows))
    _write_weekly_maturity_cache(result, start, end)
    return result, start, end


def _set_chinese_font() -> None:
    plt.rcParams["font.sans-serif"] = [
        "Microsoft YaHei",
        "SimHei",
        "KaiTi_GB2312",
        "Arial Unicode MS",
        "DejaVu Sans",
    ]
    plt.rcParams["axes.unicode_minus"] = False


def _safe_filename(name: str) -> str:
    return name.replace("/", "_").replace("\\", "_")


def plot_category_series(category: str, wide: pd.DataFrame, save_path: Path) -> None:
    _set_chinese_font()
    fig, ax = plt.subplots(figsize=(13, 6.8))
    for idx, column in enumerate(wide.columns):
        ax.plot(
            wide.index,
            wide[column],
            linewidth=1.25,
            color=CHART_COLORS[idx % len(CHART_COLORS)],
            label=str(column),
        )
    start_year = int(pd.DatetimeIndex(wide.index).min().year)
    if category == ALL_MARKET_CATEGORY:
        title = f"百元拟合溢价率：反三次与多因子修正（{start_year}年以来）"
        methodology = (
            "口径：均为单日全市场截面并读取平价100处的拟合值；"
            "反三次使用平价70—130且至少8条样本，多因子修正使用平价50—200且至少36条样本；"
            "换手率<50%，剔除溢价率两端各3%。"
        )
    else:
        title = f"{category}：多因子修正拟合溢价率（{start_year}年以来）"
        methodology = (
            "口径：幂衰减基准＋余额、期限、正股波动率、赎回/下修累计天数、隐含波动率多因子修正；"
            "平价50—200，换手率<50%，剔除溢价率两端各3%。"
        )
    ax.set_title(title, loc="left", fontsize=15)
    ax.set_ylabel("拟合转股溢价率（%）")
    ax.set_xlabel("")
    ax.grid(axis="y", alpha=0.25)
    ax.spines[["top", "right"]].set_visible(False)
    ax.xaxis.set_major_locator(mdates.YearLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    ax.legend(ncol=min(3, max(1, len(wide.columns))), frameon=False, loc="upper left")
    fig.text(
        0.01,
        0.012,
        methodology,
        fontsize=9,
        color="#667085",
    )
    fig.tight_layout(rect=(0, 0.045, 1, 1))
    fig.savefig(save_path, dpi=220, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def plot_weekly_maturity_fit(
    result: pd.DataFrame,
    save_path: Path,
    start: pd.Timestamp,
    end: pd.Timestamp,
) -> None:
    """绘制剩余期限周分组的期限截面曲线。"""
    _set_chinese_font()
    x = pd.to_numeric(result["周序号"], errors="coerce") + 0.5
    y = pd.to_numeric(result["拟合溢价率"], errors="coerce")
    fig, ax = plt.subplots(figsize=(13, 6.8))
    ax.plot(x, y, linewidth=1.35, color=CHART_COLORS[0])
    ax.set_title("剩余期限周分组：多因子修正拟合溢价率", loc="left", fontsize=15)
    ax.set_ylabel("拟合转股溢价率（%）")
    ax.set_xlabel("剩余期限（周）")
    ax.grid(axis="y", alpha=0.25)
    ax.spines[["top", "right"]].set_visible(False)
    ax.set_xlim(0, WEEKLY_MATURITY_MAX_WEEKS)
    ax.set_xticks(np.arange(0, WEEKLY_MATURITY_MAX_WEEKS + 1, 26))
    fig.text(
        0.01,
        0.012,
        f"样本期间：{start:%Y-%m-%d}—{end:%Y-%m-%d}；每周为一个剩余期限分档。"
        "口径：幂衰减基准＋六因子修正；平价50—200，换手率<50%，剔除溢价率两端各3%。",
        fontsize=9,
        color="#667085",
    )
    fig.tight_layout(rect=(0, 0.045, 1, 1))
    fig.savefig(save_path, dpi=220, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def plot_latest_multifactor_cross_section(
    config: RunConfig,
    latest_date: pd.Timestamp,
) -> Path:
    """绘制最新单日截面的散点、幂衰减基准与多因子修正结果。"""
    frames, industry, all_dates = _prepare_data(config)
    latest_date = pd.Timestamp(latest_date).normalize()
    date_positions = np.flatnonzero(all_dates.normalize() == latest_date)
    if len(date_positions) == 0:
        raise ValueError(f"原始数据中找不到最新拟合日期：{latest_date:%Y-%m-%d}")

    base = _build_window_base(
        frames,
        industry,
        all_dates,
        int(date_positions[-1]),
        window_size=1,
    )
    all_scatter = base[["平价", "转股溢价率"]].apply(
        pd.to_numeric,
        errors="coerce",
    ).dropna()
    sample_mask = (
        base["平价"].gt(RESIDUAL_POWER_LOWER)
        & base["平价"].lt(RESIDUAL_POWER_UPPER)
        & base["换手率"].lt(50)
    )
    fit_sample = _winsorize_premium_only(base.loc[sample_mask].copy())
    detail = _fit_residual_power_decay_detail(
        fit_sample,
        target_x=100.0,
        high_price_subset=None,
        include_diagnostics=True,
    )
    diagnostics = detail.get("_diagnostics")
    if diagnostics is None:
        raise RuntimeError(
            f"{latest_date:%Y-%m-%d} 有效样本不足，无法绘制多因子修正拟合截面图"
        )

    work = diagnostics["work"]
    x_fit = work["平价"].to_numpy(dtype=float)
    y_fit = work["转股溢价率"].to_numpy(dtype=float)
    corrected_pred = np.asarray(diagnostics["corrected_pred"], dtype=float)
    order = np.argsort(x_fit)
    smooth_window = min(31, max(9, (len(work) // 12) | 1))
    smooth_x = x_fit[order]
    smooth_y = pd.Series(corrected_pred[order]).rolling(
        smooth_window,
        center=True,
        min_periods=1,
    ).mean()

    amplitude = float(diagnostics["amplitude"])
    scale = float(diagnostics["scale"])
    power = float(diagnostics["power"])
    floor = float(diagnostics["floor"])
    curve_lower = max(0.0, RESIDUAL_POWER_ANCHOR - scale + 0.01)
    curve_x = np.linspace(curve_lower, 300.0, 700)
    curve_y = np.asarray(
        power_decay_with_floor(
            curve_x,
            amplitude,
            scale,
            power,
            floor,
            anchor_x=RESIDUAL_POWER_ANCHOR,
        ),
        dtype=float,
    )

    _set_chinese_font()
    fig, ax = plt.subplots(figsize=(14, 7.5))
    ax.axvspan(
        RESIDUAL_POWER_LOWER,
        RESIDUAL_POWER_UPPER,
        color="#EAF2FF",
        alpha=0.55,
        label="拟合区间",
        zorder=0,
    )
    ax.scatter(
        all_scatter["平价"],
        all_scatter["转股溢价率"],
        s=18,
        color="#A6A6A6",
        alpha=0.22,
        label="当日散点",
        zorder=1,
    )
    ax.scatter(
        x_fit,
        y_fit,
        s=21,
        color="#F79646",
        alpha=0.58,
        label="拟合样本",
        zorder=2,
    )
    ax.plot(
        curve_x,
        curve_y,
        linewidth=2.2,
        color="#0262BA",
        label="幂衰减基准曲线",
        zorder=3,
    )
    ax.scatter(
        x_fit,
        corrected_pred,
        s=18,
        color="#6F4EB2",
        alpha=0.48,
        label="残差修正拟合点",
        zorder=4,
    )
    ax.plot(
        smooth_x,
        smooth_y,
        linewidth=2.0,
        linestyle="--",
        color="#6F4EB2",
        label="残差修正平滑线",
        zorder=5,
    )
    ax.axhline(
        floor,
        linewidth=1.1,
        linestyle="--",
        color="#667085",
        label=f"右侧渐近线：{floor:.2f}%",
    )
    ax.axvline(
        100,
        linewidth=1.0,
        linestyle=":",
        color="#344054",
        label="平价100",
    )

    base_r2 = detail["基准R2"]
    corrected_r2 = detail["残差修正R2"]
    ax.set_title(
        f"{latest_date:%Y-%m-%d} 残差修正幂衰减："
        f"基准R2={base_r2:.3f}，修正R2={corrected_r2:.3f}",
        loc="left",
        fontsize=15,
    )
    ax.set_xlabel("平价")
    ax.set_ylabel("转股溢价率（%）")
    ax.set_xlim(0, 300)
    visible_premium = all_scatter.loc[
        all_scatter["平价"].between(0, 300, inclusive="both"),
        "转股溢价率",
    ]
    if visible_premium.notna().any():
        y_min = min(-5.0, float(visible_premium.min()) - 3.0)
        y_max = max(80.0, float(visible_premium.max()) * 1.04)
        ax.set_ylim(y_min, min(y_max, 450.0))
    ax.grid(alpha=0.22)
    ax.spines[["top", "right"]].set_visible(False)
    ax.legend(loc="upper right", frameon=True, framealpha=0.92, ncol=2)
    fig.text(
        0.01,
        0.012,
        f"当日散点{len(all_scatter)}只；拟合样本{len(work)}只；"
        "口径：平价50—200、换手率<50%，转股溢价率按3%—97%分位数剔除。",
        fontsize=9,
        color="#667085",
    )
    fig.tight_layout(rect=(0, 0.04, 1, 1))

    save_path = (
        config.output_dir
        / f"{latest_date:%m%d}【华创固收】残差修正幂衰减散点拟合.jpg"
    )
    fig.savefig(save_path, dpi=220, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"[plot] 最新截面拟合曲线与散点图：{save_path}")
    return save_path


def _format_excel_sheet(worksheet, frame: pd.DataFrame) -> None:
    """设置冻结首行、表头样式、列宽及日期/数值显示格式。"""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="E6121B")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(frame.columns, start=1):
        column_letter = get_column_letter(col_idx)
        series = frame[col_name]
        if col_name == "日期":
            worksheet.column_dimensions[column_letter].width = 12
            number_format = "yyyy-mm-dd"
        elif pd.api.types.is_numeric_dtype(series):
            worksheet.column_dimensions[column_letter].width = min(
                max(sum(2 if ord(char) > 127 else 1 for char in str(col_name)) + 4, 12),
                24,
            )
            number_format = "0.00"
        else:
            sample_values = series.dropna().astype(str).head(500)
            display_width = max(
                [sum(2 if ord(char) > 127 else 1 for char in str(col_name))]
                + [sum(2 if ord(char) > 127 else 1 for char in value) for value in sample_values]
            )
            worksheet.column_dimensions[column_letter].width = min(max(display_width + 2, 12), 80)
            number_format = None

        if number_format is not None:
            for column_cells in worksheet.iter_cols(
                min_col=col_idx,
                max_col=col_idx,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                for cell in column_cells:
                    cell.number_format = number_format


def save_outputs(
    result: pd.DataFrame,
    config: RunConfig,
    weekly_maturity_result: pd.DataFrame,
    weekly_start: pd.Timestamp,
    weekly_end: pd.Timestamp,
) -> dict:
    config.output_dir.mkdir(parents=True, exist_ok=True)
    excel_path = config.output_dir / f"多因子修正拟合溢价率_{RUN_DATE}.xlsx"

    group_order = {
        category.name: [group.label for group in category.groups]
        for category in build_category_rules()
    }
    group_order[ALL_MARKET_CATEGORY] = list(ALL_MARKET_GROUPS)
    category_summaries: Dict[str, dict] = {}
    category_tables: Dict[str, pd.DataFrame] = {}
    for category, part in result.groupby("分类", sort=False):
        wide = part.pivot(index="日期", columns="分组", values="拟合溢价率").sort_index()
        wide = wide.reindex(columns=group_order[category])
        category_tables[category] = wide.reset_index()
        if config.make_plots:
            plot_category_series(
                category,
                wide,
                config.output_dir / f"{_safe_filename(category)}_拟合溢价率.png",
            )
        category_summaries[category] = {
            group: {
                "有效天数": int(series.notna().sum()),
                "总交易日": int(len(series)),
                "有效率": float(series.notna().mean()),
                "最新值": float(series.dropna().iloc[-1]) if series.notna().any() else None,
                "最新有效日期": (
                    str(pd.Timestamp(series.dropna().index[-1]).date())
                    if series.notna().any()
                    else None
                ),
            }
            for group, series in wide.items()
        }

    if config.make_plots:
        plot_weekly_maturity_fit(
            weekly_maturity_result,
            config.output_dir / "剩余期限周分组_拟合溢价率.png",
            weekly_start,
            weekly_end,
        )

    detail_excel = result.drop(
        columns=EXCEL_DETAIL_EXCLUDED_COLUMNS,
        errors="ignore",
    ).copy()
    detail_excel["日期"] = pd.to_datetime(detail_excel["日期"], errors="coerce")
    weekly_excel = weekly_maturity_result.drop(
        columns=["周序号", *EXCEL_DETAIL_EXCLUDED_COLUMNS],
        errors="ignore",
    ).copy()
    with pd.ExcelWriter(
        excel_path,
        engine="openpyxl",
        mode="w",
        datetime_format="yyyy-mm-dd",
    ) as writer:
        detail_excel.to_excel(writer, sheet_name="拟合明细", index=False)
        _format_excel_sheet(writer.sheets["拟合明细"], detail_excel)
        weekly_excel.to_excel(writer, sheet_name="剩余期限周分组", index=False)
        _format_excel_sheet(writer.sheets["剩余期限周分组"], weekly_excel)
        for category, table in category_tables.items():
            table = table.copy()
            table["日期"] = pd.to_datetime(table["日期"], errors="coerce")
            table.to_excel(writer, sheet_name=category, index=False)
            _format_excel_sheet(writer.sheets[category], table)

    summary = {
        "开始日期": str(pd.Timestamp(result["日期"].min()).date()),
        "结束日期": str(pd.Timestamp(result["日期"].max()).date()),
        "交易日数": int(result["日期"].nunique()),
        "分类数": int(result["分类"].nunique()),
        "分组数": int(result[["分类", "分组"]].drop_duplicates().shape[0]),
        "明细行数": int(len(result)),
        "模型": "分组多因子修正拟合溢价率 + 全市场百元拟合溢价率",
        "模型方法": "分组/百元多因子修正：幂衰减基准曲线 + 六因子标准化 OLS 修正；传统百元拟合：反三次函数",
        "分组窗口": "分板块10日，其余分组5日；两条全市场百元序列1日",
        "分类结果": category_summaries,
    }
    return {"excel": str(excel_path), "output_dir": str(config.output_dir), **summary}


def parse_args() -> RunConfig:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--start-date", default="2017-01-01")
    parser.add_argument("--end-date")
    parser.add_argument("--source-type", choices=("parquet", "excel", "auto"), default="parquet")
    parser.add_argument("--excel-file-name")
    parser.add_argument("--parquet-root", default="data/转债个券历史序列")
    parser.add_argument("--force-refresh", action="store_true")
    parser.add_argument("--force-rebuild", action="store_true")
    parser.add_argument("--result-cache-path", type=Path, default=DEFAULT_RESULT_CACHE_PATH)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--no-plots", action="store_true")
    args = parser.parse_args()
    return RunConfig(
        start_date=args.start_date,
        end_date=args.end_date,
        source_type=args.source_type,
        excel_file_name=args.excel_file_name,
        parquet_root=args.parquet_root,
        force_refresh=args.force_refresh,
        force_rebuild=FORCE_REBUILD or args.force_rebuild,
        result_cache_path=args.result_cache_path,
        output_dir=args.output_dir,
        make_plots=not args.no_plots,
    )


def main() -> None:
    config = parse_args()
    existing = _load_result_cache(config)
    cached_result = calculate_group_series(config, existing_result=existing)
    _write_result_cache(cached_result, config)
    weekly_result, weekly_start, weekly_end = calculate_weekly_maturity_fit(config)

    start = pd.Timestamp(config.start_date)
    end = pd.Timestamp(config.end_date) if config.end_date else cached_result["日期"].max()
    output_result = cached_result[
        cached_result["日期"].between(start, end, inclusive="both")
    ].copy()
    summary = save_outputs(
        output_result,
        config,
        weekly_result,
        weekly_start,
        weekly_end,
    )
    if config.make_plots:
        latest_plot_path = plot_latest_multifactor_cross_section(
            config,
            pd.Timestamp(output_result["日期"].max()),
        )
        summary["最新截面拟合图"] = str(latest_plot_path)
    write_daily_fits_to_index(output_result, config)
    terminal_summary = {key: value for key, value in summary.items() if key != "分类结果"}
    print(json.dumps(terminal_summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
