import copy
import json
import os
import re
import shutil
import subprocess
import sys
import textwrap
import tempfile
import time
from collections import Counter, defaultdict
from configparser import ConfigParser
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import urlencode
from zipfile import BadZipFile, ZipFile

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[2]
START_DATE = (datetime.now() - timedelta(days=10)).strftime("%Y-%m-%d")
QUERY_END_DATE = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")  # 巨潮需闭区间；截止日设为当日 T+1。
MAPPING_DATE = datetime.now().strftime("%Y-%m-%d")
DATABASE_XLSX = ROOT / "data/clauses/【华创固收】下修和不下修公告统计.xlsx"
REVISION_BACKUP_DIR = ROOT / "archive/backups" / "条款公告更新"
QUERY_RANGE_LABEL = f"{START_DATE}至{QUERY_END_DATE}"
QUERY_RANGE_SUFFIX = f"{START_DATE.replace('-', '')}_{QUERY_END_DATE.replace('-', '')}"
OUTPUT_XLSX = ROOT / f"可转债修正公告信息收集_{QUERY_RANGE_SUFFIX}.xlsx"
CACHE_DIR = ROOT / "outputs" / f"kzzq_revision_{QUERY_RANGE_SUFFIX}"

IFIND_CREDENTIAL_FILE = ROOT / "private/ifind账号.txt"
THS_LOGIN_OK_CODES = (0, -201)
ENABLE_DATABASE_WRITE = True # 是否将结果写入数据库
ENABLE_AUDIT_WORKBOOK = False  # 默认不输出“公告信息收集”核对表；需要核对时改为 True。
ENABLE_CACHE_SAVE = False  # 默认不保存 outputs 缓存；需要复核 JSON/PDF 时改为 True。
ENABLE_REDEMPTION_UPDATE = True  # 默认同步运行强赎/不赎回公告更新。


def load_ifind_credentials() -> tuple[str, str]:
    """从项目目录的 ifind账号.txt 读取统一登录账号。"""
    if not IFIND_CREDENTIAL_FILE.is_file():
        raise FileNotFoundError(f"未找到iFinD账号文件：{IFIND_CREDENTIAL_FILE}")
    config = ConfigParser(interpolation=None)
    config.read(IFIND_CREDENTIAL_FILE, encoding="utf-8")
    username = config.get("ifind", "username", fallback="").strip()
    password = config.get("ifind", "password", fallback="").strip()
    if not username or not password:
        raise RuntimeError("ifind账号.txt中的[ifind] username或password为空")
    return username, password


def print_ifind_usage() -> None:
    """显示iFinD各数据项的已用额度比例。"""
    try:
        from iFinDPy import THS_DataStatistics

        result = THS_DataStatistics()
        tables = result.get("tables", {}) if isinstance(result, dict) else {}
        if not tables:
            detail = result.get("errmsg", "未返回额度数据") if isinstance(result, dict) else str(result)
            print(f"[警告] iFinD使用额度查询失败：{detail}")
            return
        print("iFinD使用额度：")
        for key, value in tables.items():
            ratio = value.get("ratio", "N/A") if isinstance(value, dict) else value
            print(f"{key} 已用：{ratio}")
    except Exception as exc:
        print(f"[警告] iFinD使用额度查询失败：{exc}")

EXCLUDE_TITLE_WORDS = ("修订", "修正案", "修改", "装修")
CNINFO_QUERY_URL = "https://www.cninfo.com.cn/new/hisAnnouncement/query"
CNINFO_STATIC_PREFIX = "https://static.cninfo.com.cn/"


def run_cninfo_json_request(
    command: list[str],
    context: str,
    timeout: int,
    attempts: int = 3,
) -> dict:
    last_detail = ""
    for attempt in range(1, attempts + 1):
        try:
            result = subprocess.run(
                command,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=timeout,
            )
        except subprocess.TimeoutExpired:
            last_detail = f"请求超时（{timeout}秒）"
        else:
            response_text = result.stdout.lstrip("\ufeff").strip()
            stderr_text = result.stderr.strip()
            if result.returncode != 0:
                last_detail = f"curl退出码={result.returncode}，stderr={stderr_text or '空'}"
            elif not response_text:
                last_detail = f"响应为空，stderr={stderr_text or '空'}"
            else:
                try:
                    payload = json.loads(response_text)
                except json.JSONDecodeError as exc:
                    preview = re.sub(r"\s+", " ", response_text)[:200]
                    last_detail = (
                        f"返回内容不是JSON（{exc.msg}，位置{exc.pos}），"
                        f"响应片段={preview!r}"
                    )
                else:
                    if isinstance(payload, dict):
                        return payload
                    last_detail = f"JSON根节点类型异常：{type(payload).__name__}"
        if attempt < attempts:
            time.sleep(attempt * 2)
    raise RuntimeError(f"{context}；连续{attempts}次请求失败；{last_detail}")
SHAREHOLDER_MEETING_CACHE = {}


def clean_title(title):
    return re.sub(r"<[^>]+>", "", title or "").strip()


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).replace("“", '"').replace("”", '"').replace("＂", '"').replace("：", ":")
    text = re.sub(r"^[^:]{1,24}:", "", text)
    text = re.sub(r"\s+", "", text)
    return text.replace('""', '"')


def parse_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if not value:
        return None
    value = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    match = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", value)
    if match:
        return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return None


def day_key(value):
    dt = parse_datetime(value)
    return dt.date().isoformat() if dt else ""


def cninfo_time(milliseconds):
    dt = datetime.fromtimestamp(
        int(milliseconds) / 1000,
        tz=timezone.utc,
    ).astimezone(timezone(timedelta(hours=8))).replace(tzinfo=None)
    # 巨潮接口若只给公告日期，通常表现为北京时间 00:00:00。
    # 这类记录按前一日 23:59:59 处理；若接口已有具体时点，则保留原时点。
    if dt.time() == datetime.min.time():
        return dt - timedelta(seconds=1)
    return dt


def show_popup(title, message):
    try:
        import tkinter as tk

        root = tk.Tk()
        root.title(title)
        root.geometry("760x520")
        root.attributes("-topmost", True)
        frame = tk.Frame(root, padx=12, pady=12)
        frame.pack(fill="both", expand=True)
        text = tk.Text(frame, wrap="word", font=("KaiTi_GB2312", 10))
        text.insert("1.0", message)
        text.configure(state="disabled")
        text.pack(fill="both", expand=True)
        button = tk.Button(frame, text="关闭", command=root.destroy, width=12)
        button.pack(pady=(10, 0))
        root.after(60_000, root.destroy)
        root.mainloop()
    except Exception:
        print(f"\n{title}\n{message}")


def post_cninfo(page_num, page_size=30):
    data = {
        "pageNum": str(page_num),
        "pageSize": str(page_size),
        "column": "szse",
        "tabName": "fulltext",
        "plate": "",
        "stock": "",
        "searchkey": "修",
        "secid": "",
        "category": "category_kzzq_szsh;",
        "trade": "",
        "seDate": f"{START_DATE}~{QUERY_END_DATE}",
        "sortName": "",
        "sortType": "",
        "isHLtitle": "true",
    }
    body = urlencode(data, safe=";")
    command = [
        "curl.exe",
        "-sS",
        "-L",
        "--compressed",
        "--connect-timeout",
        "30",
        CNINFO_QUERY_URL,
        "-H",
        "User-Agent: Mozilla/5.0",
        "-H",
        "Referer: https://www.cninfo.com.cn/new/commonUrl/pageOfSearch?url=disclosure/list/search&checkedCategory=category_kzzq_szsh",
        "-H",
        "X-Requested-With: XMLHttpRequest",
        "--data",
        body,
    ]
    return run_cninfo_json_request(
        command,
        context=f"巨潮接口请求失败: page={page_num}",
        timeout=360,
    )


def fetch_announcements():
    if ENABLE_CACHE_SAVE:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
    all_rows = []
    page_size = 30
    for page in range(1, 50):
        obj = post_cninfo(page, page_size)
        announcements = obj.get("announcements") or []
        if not announcements:
            break
        all_rows.extend(announcements)
        if len(announcements) < page_size:
            break

    seen = set()
    unique = []
    for item in all_rows:
        announcement_id = str(item.get("announcementId"))
        if announcement_id in seen:
            continue
        seen.add(announcement_id)
        unique.append(item)

    if ENABLE_CACHE_SAVE:
        raw_path = CACHE_DIR / "announcements_raw.json"
        raw_path.write_text(json.dumps({"announcements": unique}, ensure_ascii=False, indent=2), encoding="utf-8")
    return unique


def ths_login_errmsg(code: int) -> str:
    try:
        from iFinDPy import THS_GetErrorInfo

        info = THS_GetErrorInfo(code)
        if isinstance(info, dict):
            return str(info.get("errmsg", info))
        return str(info)
    except Exception:
        return f"未知错误（状态码 {code}）"


def ths_login() -> int:
    from iFinDPy import THS_iFinDLogin

    username, password = load_ifind_credentials()
    code = THS_iFinDLogin(username, password)
    if code not in THS_LOGIN_OK_CODES:
        raise RuntimeError(f"iFinD 登录失败（状态码 {code}）: {ths_login_errmsg(code)}")
    print_ifind_usage()
    return code


def fetch_cb_codes_on_date(edate_yyyymmdd: str) -> str:
    """参考日报脚本：按日期获取未到期可转债成分列表，返回逗号分隔代码。"""
    from iFinDPy import THS_DR

    df = THS_DR(
        "p00570",
        f"jyzt=未到期;sfdb=全部;jysc=全部;edate={edate_yyyymmdd}",
        "jydm:Y,jydm_mc:Y,p00570_f001:Y,p00570_f019:Y",
        "format:dataframe",
    ).data
    if df is None or df.empty:
        raise RuntimeError(f"THS_DR 未返回可转债列表：{edate_yyyymmdd}")
    return ",".join(df.set_index("jydm").index.astype(str))


CB_BASIC_COLUMNS = ["转债简称", "正股代码", "正股简称", "发行方式", "交易状态", "转债余额", "上市日期"]


def filter_cb_basic(df: pd.DataFrame) -> pd.DataFrame:
    """参考日报脚本：统一列名并剔除定向、NQ、终止上市。"""
    out = df.set_index("thscode").rename_axis("转债代码")
    out.columns = CB_BASIC_COLUMNS
    out = out[~out["发行方式"].astype(str).str.contains("定向", na=False)]
    out = out[~out.index.astype(str).str.contains("NQ", na=False)]
    if "交易状态" in out.columns:
        out = out[~out["交易状态"].astype(str).str.contains("终止上市", na=False)]
    return out


def fetch_cb_basic(codes: str, last_date: str) -> pd.DataFrame:
    """参考日报脚本：获取转债简称、正股代码、正股简称等基础信息。"""
    from iFinDPy import THS_BD

    raw = THS_BD(
        codes,
        "ths_convertible_debt_short_name_cbond;ths_stock_code_cbond;ths_stock_short_name_cbond;"
        "ths_issue_method_cbond;ths_trading_status_bond;ths_bond_balance_cbond;ths_listed_date_cbond",
        f";;;;;{last_date};",
    ).data
    if raw is None or raw.empty:
        raise RuntimeError("THS_BD 未返回转债基础信息")
    return filter_cb_basic(raw)


def load_bond_mapping():
    """通过 THS API 动态获取转债-正股映射；不再使用本地 Excel 兜底。"""
    last_date = MAPPING_DATE
    edate_yyyymmdd = datetime.strptime(last_date, "%Y-%m-%d").strftime("%Y%m%d")
    ths_login()
    codes = fetch_cb_codes_on_date(edate_yyyymmdd)
    cb_basic = fetch_cb_basic(codes, last_date)
    cb_basic = cb_basic.reset_index()
    for column in ["转债代码", "转债简称", "正股代码", "正股简称"]:
        cb_basic[column] = cb_basic[column].astype(str).str.strip()
    cb_basic["正股6"] = cb_basic["正股代码"].str.extract(r"(\d{6})", expand=False)
    mapping = {
        stock_code: group[["转债代码", "转债简称"]].drop_duplicates().to_dict("records")
        for stock_code, group in cb_basic.dropna(subset=["正股6"]).groupby("正股6", sort=False)
    }
    if not mapping:
        raise RuntimeError(f"THS API 未获取到有效转债-正股映射：{last_date}")
    return mapping


def normalize_bond_match_text(value):
    """标准化公告标题和转债简称，供同一正股多只转债时做精确简称匹配。"""
    text = str(value or "")
    text = text.replace("“", "").replace("”", "").replace('"', "").replace("＂", "")
    return re.sub(r"\s+", "", text).upper()


def resolve_bond_from_title(stock_code, title, mapping):
    """按正股取得全部转债候选，并优先用公告标题中的转债简称唯一匹配。"""
    candidates = mapping.get(stock_code, [])
    # 兼容旧调用方或测试中仍传入“正股 -> 单只转债”的映射结构。
    if isinstance(candidates, dict):
        candidates = [candidates]
    candidates = [candidate for candidate in candidates if isinstance(candidate, dict)]
    if not candidates:
        return {"转债代码": "未匹配", "转债简称": "未匹配"}

    normalized_title = normalize_bond_match_text(title)
    name_matches = []
    for candidate in candidates:
        bond_name = normalize_bond_match_text(candidate.get("转债简称"))
        if bond_name and bond_name in normalized_title:
            name_matches.append(candidate)
    if len(name_matches) == 1:
        return name_matches[0]

    # 少数公告标题可能只写转债代码；简称未能唯一命中时再尝试代码。
    code_matches = []
    for candidate in candidates:
        bond_code = str(candidate.get("转债代码") or "").strip().upper()
        bond_code_base = bond_code.split(".")[0]
        if bond_code and (bond_code in normalized_title or bond_code_base in normalized_title):
            code_matches.append(candidate)
    if len(code_matches) == 1:
        return code_matches[0]

    if len(candidates) == 1:
        return candidates[0]
    return {"转债代码": "未匹配", "转债简称": "未匹配"}


def classify_title(title):
    if "不向下修正" in title or "不下修" in title:
        return "不下修"
    if "董事会" in title and any(word in title for word in ("提议", "提出")) and "向下修正" in title:
        return "董事会提议下修"
    if ("预计" in title or "可能" in title) and ("修正" in title or "下修" in title):
        return "预计下修"
    if ("向下修正" in title or "下修" in title or "修正" in title) and ("转股价格" in title or "转股价" in title):
        return "实际下修"
    return "其他含修"


def build_classified_rows(announcements, mapping):
    rows = []
    for item in announcements:
        title = clean_title(item.get("announcementTitle"))
        if "修" not in title or any(word in title for word in EXCLUDE_TITLE_WORDS):
            continue
        stock_code = str(item.get("secCode", "")).zfill(6)
        bond = resolve_bond_from_title(stock_code, title, mapping)
        rows.append(
            {
                "公告ID": str(item.get("announcementId")),
                "正股代码": stock_code,
                "正股简称": item.get("secName"),
                "转债代码": bond["转债代码"],
                "转债简称": bond["转债简称"],
                "公告时间": cninfo_time(item.get("announcementTime")),
                "公告类型": classify_title(title),
                "公告标题": title,
                "公告URL": CNINFO_STATIC_PREFIX + item.get("adjunctUrl", ""),
                "PDF文件": str(CACHE_DIR / "pdfs" / f"{item.get('announcementId')}.pdf"),
            }
        )
    rows.sort(key=lambda row: row["公告时间"], reverse=True)
    if ENABLE_CACHE_SAVE:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        (CACHE_DIR / "classified_rows.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return rows


def download_pdf(row):
    pdf_path = Path(row["PDF文件"])
    if ENABLE_CACHE_SAVE:
        pdf_path.parent.mkdir(parents=True, exist_ok=True)
        if pdf_path.exists() and pdf_path.stat().st_size > 0:
            return pdf_path
    else:
        fd, temp_name = tempfile.mkstemp(prefix=f"cninfo_{row['公告ID']}_", suffix=".pdf")
        os.close(fd)
        pdf_path = Path(temp_name)
    result = subprocess.run(
        ["curl.exe", "-s", "-L", row["公告URL"], "-H", "User-Agent: Mozilla/5.0", "-o", str(pdf_path)],
        capture_output=True,
        text=True,
        timeout=60,
    )
    if result.returncode != 0 or not pdf_path.exists() or pdf_path.stat().st_size == 0:
        raise RuntimeError(f"PDF 下载失败: {row['公告URL']}")
    return pdf_path


def read_pdf_text(row):
    pdf_path = download_pdf(row)
    try:
        reader = PdfReader(str(pdf_path))
        return "\n".join(clean_pdf_page_text(page.extract_text() or "", index + 1) for index, page in enumerate(reader.pages))
    finally:
        if not ENABLE_CACHE_SAVE:
            try:
                Path(pdf_path).unlink(missing_ok=True)
            except Exception:
                pass


def post_cninfo_shareholder_meeting(page_num, query_date, searchkey, page_size=30):
    data = {
        "pageNum": str(page_num),
        "pageSize": str(page_size),
        "column": "szse",
        "tabName": "fulltext",
        "plate": "",
        "stock": "",
        "searchkey": searchkey,
        "secid": "",
        "category": "",
        "trade": "",
        "seDate": f"{query_date}~{query_date}",
        "sortName": "",
        "sortType": "",
        "isHLtitle": "true",
    }
    body = urlencode(data, safe=";")
    command = [
        "curl.exe",
        "-sS",
        "-L",
        "--compressed",
        "--connect-timeout",
        "30",
        CNINFO_QUERY_URL,
        "-H",
        "User-Agent: Mozilla/5.0",
        "-H",
        "Referer: https://www.cninfo.com.cn/new/commonUrl/pageOfSearch?url=disclosure/list/search",
        "-H",
        "X-Requested-With: XMLHttpRequest",
        "--data",
        body,
    ]
    return run_cninfo_json_request(
        command,
        context=(
            "巨潮股东会公告查询失败: "
            f"date={query_date}, page={page_num}, searchkey={searchkey}"
        ),
        timeout=120,
    )


def fetch_shareholder_meeting_announcements(query_date):
    if query_date in SHAREHOLDER_MEETING_CACHE:
        return SHAREHOLDER_MEETING_CACHE[query_date]

    unique = {}
    for searchkey in ("股东大会", "股东会"):
        for page in range(1, 30):
            payload = post_cninfo_shareholder_meeting(page, query_date, searchkey)
            announcements = payload.get("announcements") or []
            if not announcements:
                break
            for item in announcements:
                announcement_id = str(item.get("announcementId") or "")
                if announcement_id:
                    unique[announcement_id] = item
            if len(announcements) < 30:
                break

    rows = list(unique.values())
    SHAREHOLDER_MEETING_CACHE[query_date] = rows
    return rows


def extract_shareholder_meeting_date(text):
    compact = compact_text(text)
    cn_date = r"\d{4}年\d{1,2}月\d{1,2}日"
    patterns = [
        rf"(?:现场会议召开时间|现场会议时间|会议召开时间|召开时间|会议时间)[:：]?({cn_date})",
        rf"(?:于|定于|拟于)({cn_date})(?:召开|举行)[^。；;]{{0,40}}股东(?:大)?会",
        rf"({cn_date})[^。；;]{{0,45}}(?:召开|举行)[^。；;]{{0,30}}股东(?:大)?会",
        rf"股东(?:大)?会[^。；;]{{0,70}}(?:于|定于)({cn_date})(?:召开|举行)",
    ]
    for pattern in patterns:
        match = re.search(pattern, compact)
        if match:
            return cn_date_from_match(match.group(1))
    return None


def shareholder_meeting_notice_score(title):
    title = clean_title(title)
    if any(word in title for word in ("决议", "法律意见", "会议资料")):
        return 2
    if "召开" in title and any(word in title for word in ("通知", "提示性公告")):
        return 0
    return 1


def fetch_proposal_shareholder_meeting_date(row):
    board_time = parse_datetime(row.get("公告时间"))
    stock_match = re.search(r"\d{6}", str(row.get("正股代码") or ""))
    if not board_time or not stock_match:
        return None

    stock_code = stock_match.group(0)
    query_dates = [
        board_time.strftime("%Y-%m-%d"),
        (board_time + timedelta(days=1)).strftime("%Y-%m-%d"),
    ]
    candidates = []
    for query_date in dict.fromkeys(query_dates):
        for item in fetch_shareholder_meeting_announcements(query_date):
            title = clean_title(item.get("announcementTitle"))
            if str(item.get("secCode") or "").strip() != stock_code:
                continue
            if "股东大会" not in title and "股东会" not in title:
                continue
            candidates.append((shareholder_meeting_notice_score(title), query_date, title, item))

    candidates.sort(key=lambda item: (item[0], item[1], item[2]))
    for _, query_date, title, item in candidates:
        announcement_id = str(item.get("announcementId") or "")
        adjunct_url = item.get("adjunctUrl") or ""
        notice_url = CNINFO_STATIC_PREFIX + adjunct_url if adjunct_url.startswith("finalpage/") else adjunct_url
        notice_row = {
            "公告ID": announcement_id,
            "公告URL": notice_url,
            "PDF文件": str(CACHE_DIR / "pdfs" / f"shareholder_meeting_{announcement_id}.pdf"),
        }
        try:
            meeting_date = extract_shareholder_meeting_date(read_pdf_text(notice_row))
        except Exception as exc:
            print(f"股东会日期抽取失败：{row.get('转债简称')} {title} ({query_date})：{exc}")
            continue
        if meeting_date:
            row["股东大会公告标题"] = title
            row["股东大会公告URL"] = notice_url
            row["股东大会公告披露日"] = query_date
            return meeting_date
    return None


def clean_pdf_page_text(text, page_number):
    lines = (text or "").splitlines()
    while lines and not lines[0].strip():
        lines.pop(0)
    while lines and not lines[-1].strip():
        lines.pop()
    page_marker = str(page_number)
    if lines and lines[0].strip() == page_marker:
        lines.pop(0)
    if lines and lines[-1].strip() == page_marker:
        lines.pop()
    return "\n".join(lines)


def compact_text(text):
    return re.sub(r"\s+", "", text or "")


def cn_date_from_match(value):
    match = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", value or "")
    if not match:
        return None
    return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))


def extract_no_down_commitment(text):
    compact = compact_text(text)
    cn_date = r"\d{4}年\d{1,2}月\d{1,2}日"
    bond_maturity_name = r'(?:可转债|债券|[“"][^”"]{1,24}(?:转债|债券)[”"]|[^“”"，。；（）()]{1,16}(?:转债|债券))'
    patterns = [
        # 先识别公告明确给出的月份和日期区间，避免被“转股期至债券到期日”等背景信息抢先命中。
        rf"(?:本次)?董事会审议通过(?:后(?:的)?|的)?次一交易日(?:为)?[（(]?(?:即)?{cn_date}[）)]?(?:起)?至({cn_date})[^。；]{{0,220}}?(?:亦)?不(?:提出|向下修正|行使)",
        rf"(?:自)?(?:本次)?(?:本公告(?:披露)?(?:之日|日)?起|公告(?:披露)?(?:之日|日)?起|董事会(?:审议通过)?(?:之日|日)?起|董事会会议审议通过(?:之日|日)?起|董事会审议通过(?:后(?:的)?|的)次一交易日(?:起)?).{{0,100}}?至.{{0,100}}?[（(](?:即)?(?:自)?{cn_date}(?:起)?至({cn_date})(?:期间|止)?[）)](?:内)?.{{0,220}}?(?:亦)?不(?:提出|向下修正|行使)",
        rf"(?:自)?(?:本公告(?:披露)?(?:之日|日)?起|公告(?:披露)?(?:之日|日)?起|董事会(?:审议通过)?(?:之日|日)?起|董事会会议审议通过(?:之日|日)?起)至({cn_date}).{{0,220}}?(?:亦)?不(?:提出|向下修正|行使)",
        rf"未来.{{0,120}}?(?:自)?(?:本公告(?:披露)?(?:之日|日)?起|公告(?:披露)?(?:之日|日)?起|董事会(?:审议通过)?(?:之日|日)?起|董事会会议审议通过(?:之日|日)?起)至({cn_date})",
        rf"(?:[一二三四五六七八九十两0-9]+个?月)(?:内)?.{{0,30}}?[（(](?:即)?(?:自)?{cn_date}(?:起)?至({cn_date})(?:期间|止)?[）)](?:内)?.{{0,220}}?(?:亦)?不(?:提出|向下修正|行使)",
        rf"未来.{{0,80}}?[（(](?:即)?(?:自)?(?:本公告(?:披露)?日|公告披露日|本公告日|{cn_date}).{{0,20}}?至({cn_date})[）)]",
        rf"未来.{{0,80}}?(?:即)?(?:自)?(?:本公告(?:披露)?日|公告披露日|本公告日)至({cn_date}).{{0,220}}?(?:亦)?不(?:提出|向下修正|行使)",
        rf"未来.{{0,180}}?至({cn_date}).{{0,220}}?(?:亦)?不(?:提出|向下修正|行使)",
        rf"未来[^，。；（）()]{{0,30}}[（(](?:即)?(?:自)?{cn_date}至({cn_date})[）)]",
        rf"未来[^，。；（）()]{{0,50}}[（(](?:即)?(?:自)?(?:本次触发修正条件的次一交易日)?{cn_date}至({cn_date})[）)]",
        rf"未来[^，。；]{{0,120}}至({cn_date})[）)]?[^，。；]{{0,100}}(?:亦)?不(?:提出|向下修正|行使)",
        rf"自{cn_date}至({cn_date})[^，。；]{{0,100}}不(?:提出|向下修正|行使)",
        rf"截至({cn_date})[^，。；]{{0,100}}不(?:提出|向下修正|行使)",
        rf"({cn_date})前[^，。；]{{0,100}}不(?:提出|向下修正|行使)",
        # “承诺至债券到期日”放在明确期限之后，并限制在同一句内，禁止跨段抓取后文的不下修表述。
        rf"至{bond_maturity_name}(?:的)?到期日(?:内)?[（(](?:即)?(?:自)?{cn_date}(?:起)?至({cn_date})(?:期间|止)?[）)][^。；]{{0,220}}?(?:亦)?不(?:提出|向下修正|行使)",
        rf"至{bond_maturity_name}(?:的)?到期日(?:内)?[（(](?:即)?({cn_date})(?:，?如遇节假日，?向后顺延)?[）)][^。；]{{0,220}}?(?:亦)?不(?:提出|向下修正|行使)",
        rf"至{bond_maturity_name}(?:的)?到期日(?:为)?({cn_date})(?:止)?[^。；]{{0,220}}?(?:亦)?不(?:提出|向下修正|行使)",
    ]
    for pattern in patterns:
        match = re.search(pattern, compact)
        if match:
            return cn_date_from_match(match.group(1)) or match.group(1)
    return "——"


def extract_first_price(patterns, text):
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return float(match.group(1))
    return None


def actual_down_search_spaces(compact):
    spaces = []

    special_start = compact.find("特别提示")
    if special_start >= 0:
        section_start = compact.find("一、", special_start + 1)
        spaces.append(compact[special_start:section_start] if section_start > special_start else compact[special_start:special_start + 800])

    result_section_patterns = [
        r"本次向下修正(?:“[^”]+”|\"[^\"]+\")转股价格",
        r"本次向下修正(?:可转债|可转换公司债券)?转股价格的(?:结果|具体内容|相关情况)",
    ]
    for pattern in result_section_patterns:
        matches = list(re.finditer(pattern, compact))
        if matches:
            marker_start = matches[-1].start()
            spaces.append(compact[marker_start:marker_start + 1200])
            break

    for marker in [
        "本次向下修正可转债转股价格的结果",
        "本次向下修正可转换公司债券转股价格的结果",
        "本次向下修正转股价格的结果",
        "本次向下修正可转债转股价格",
        "本次向下修正可转换公司债券转股价格",
    ]:
        marker_start = compact.rfind(marker)
        if marker_start >= 0:
            spaces.append(compact[marker_start:marker_start + 1000])
            break

    spaces.append(compact)
    unique_spaces = []
    for space in spaces:
        if space and space not in unique_spaces:
            unique_spaces.append(space)
    return unique_spaces


def extract_actual_down_fields(row, text):
    compact = compact_text(text)
    search_spaces = actual_down_search_spaces(compact)
    old_price = None
    new_price = None

    pair_patterns = [
        r"修正前(?:的)?(?:“[^”]+”|\"[^\"]+\")?转股价格(?:为|[:：])?([0-9]+(?:\.[0-9]+)?)元/股[^。；，]*?修正后(?:的)?(?:“[^”]+”|\"[^\"]+\")?转股价格(?:为|[:：])?([0-9]+(?:\.[0-9]+)?)元/股",
        r"转股价格由([0-9]+(?:\.[0-9]+)?)元/股(?:向下)?修正为([0-9]+(?:\.[0-9]+)?)元/股",
        r"转股价格由([0-9]+(?:\.[0-9]+)?)元/股调整为([0-9]+(?:\.[0-9]+)?)元/股",
    ]
    for search_space in search_spaces:
        for pattern in pair_patterns:
            match = re.search(pattern, search_space)
            if match:
                old_price = float(match.group(1))
                new_price = float(match.group(2))
                break
        if old_price is not None and new_price is not None:
            break

    if old_price is None:
        for search_space in search_spaces:
            old_price = extract_first_price([r"修正前(?:的)?(?:“[^”]+”|\"[^\"]+\")?转股价格(?:为|[:：])?([0-9]+(?:\.[0-9]+)?)元/股"], search_space)
            if old_price is not None:
                break
    if new_price is None:
        for search_space in search_spaces:
            new_price = extract_first_price([r"修正后(?:的)?(?:“[^”]+”|\"[^\"]+\")?转股价格(?:为|[:：])?([0-9]+(?:\.[0-9]+)?)元/股"], search_space)
            if new_price is not None:
                break

    floor_price = extract_first_price(
        [
            r"转股价格应不低于([0-9]+(?:\.[0-9]+)?)元/股",
            r"转股价格不低于([0-9]+(?:\.[0-9]+)?)元/股",
            r"修正后[^。；]{0,30}不低于([0-9]+(?:\.[0-9]+)?)元/股",
        ],
        compact,
    )
    if floor_price is None:
        for search_space in search_spaces:
            twenty_day_match = re.search(
                r"(?:前)?二十个交易日[^。；，]{0,100}?交易均价(?:为|[:：])?[（(]?(?:人民币)?([0-9]+(?:\.[0-9]+)?)元/股",
                search_space,
            )
            previous_day_match = re.search(
                r"前(?:一|一个)交易日[^。；，]{0,100}?交易均价(?:为|[:：])?[（(]?(?:人民币)?([0-9]+(?:\.[0-9]+)?)元/股",
                search_space,
            )
            if twenty_day_match and previous_day_match:
                floor_price = max(float(twenty_day_match.group(1)), float(previous_day_match.group(1)))
                break
            avg_prices = [
                float(value)
                for value in re.findall(
                    r"均价(?:为|[:：])?[（(]?(?:人民币)?([0-9]+(?:\.[0-9]+)?)元/股",
                    search_space,
                )
            ]
            if len(avg_prices) >= 2:
                floor_price = max(avg_prices[-3:] if len(avg_prices) >= 3 and "三十个交易日" in search_space else avg_prices[-2:])
                break

    effective_date = None
    effective_patterns = [
        r"修正后(?:的)?(?:“[^”]+”|\"[^\"]+\")?转股价格生效日期[:：为]*(\d{4}年\d{1,2}月\d{1,2}日)",
        r"(?:本次)?转股价格调整生效日期[:：为]*(\d{4}年\d{1,2}月\d{1,2}日)",
        r"本次转股价格修正实施日期[:：为]*(\d{4}年\d{1,2}月\d{1,2}日)",
        r"转股价格修正实施日期[:：为]*(\d{4}年\d{1,2}月\d{1,2}日)",
        r"修正后的(?:“[^”]+”|\"[^\"]+\")?转股价格自(\d{4}年\d{1,2}月\d{1,2}日)起",
        r"修正后的转股价格自(\d{4}年\d{1,2}月\d{1,2}日)起",
        r"自(\d{4}年\d{1,2}月\d{1,2}日)起恢复转股",
        r"转股价格调整生效日期为(\d{4}年\d{1,2}月\d{1,2}日)",
        r"转股价格(?:将)?于(\d{4}年\d{1,2}月\d{1,2}日)起生效",
    ]
    for search_space in search_spaces:
        for pattern in effective_patterns:
            match = re.search(pattern, search_space)
            if match:
                effective_date = cn_date_from_match(match.group(1))
                break
        if effective_date is not None:
            break

    return {
        "向下修正发布日期": row["公告时间"],
        "转股价变动日期": effective_date,
        "修正前转股价": old_price,
        "修正后转股价": new_price,
        "修正转股价应不低于": floor_price,
    }


def enrich_pdf_fields(rows):
    for row in rows:
        if row["公告类型"] == "不下修":
            row["承诺何日之前不行使"] = extract_no_down_commitment(read_pdf_text(row))
        elif row["公告类型"] == "董事会提议下修":
            row["股东大会日期"] = fetch_proposal_shareholder_meeting_date(row)
        elif row["公告类型"] == "实际下修":
            row.update(extract_actual_down_fields(row, read_pdf_text(row)))
    return rows


def style_sheet(ws, widths):
    header_fill = PatternFill(fill_type="solid", fgColor="244062")
    no_fill = PatternFill(fill_type=None)
    header_font = Font(name="KaiTi_GB2312", size=9, bold=True, color="FFFFFF")
    body_font = Font(name="KaiTi_GB2312", size=9, color="000000")
    link_font = Font(name="KaiTi_GB2312", size=9, color="0563C1", underline="single")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = header_fill if cell.row == 1 else no_fill
            cell.font = header_font if cell.row == 1 else (link_font if cell.hyperlink else body_font)
            cell.border = border
            cell.alignment = Alignment(vertical="top", horizontal="center" if cell.row == 1 else None, wrap_text=True)
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def set_hyperlink(cell, title, url):
    cell.value = title
    cell.hyperlink = url


def write_result_workbook(rows):
    wb = Workbook()
    wb.remove(wb.active)
    sheet_specs = [
        ("不下修", "不下修", ["转债代码", "转债简称", "公告时间", "承诺何日之前不行使", "公告链接"], [13, 14, 18, 22, 75]),
        ("预计下修", "预计下修", ["转债代码", "转债简称", "公告时间", "公告链接"], [13, 14, 18, 75]),
        ("董事会提议下修", "董事会提议下修", ["转债代码", "转债简称", "董事会发布日期", "提议公告"], [13, 14, 18, 75]),
        (
            "实际下修",
            "实际下修",
            ["转债代码", "转债简称", "向下修正发布日期", "转股价变动日期", "修正前转股价", "修正后转股价", "修正转股价应不低于", "下修公告"],
            [13, 14, 18, 18, 14, 14, 18, 75],
        ),
    ]
    for sheet_name, category, headers, widths in sheet_specs:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for row in rows:
            if row["公告类型"] != category:
                continue
            display_title = f"{row['正股简称']}:{row['公告标题']}"
            if category == "不下修":
                ws.append([row["转债代码"], row["转债简称"], row["公告时间"], row.get("承诺何日之前不行使", "——"), display_title])
                ws.cell(ws.max_row, 3).number_format = "yyyy/m/d h:mm"
                if isinstance(ws.cell(ws.max_row, 4).value, datetime):
                    ws.cell(ws.max_row, 4).number_format = "yyyy/m/d"
                set_hyperlink(ws.cell(ws.max_row, 5), display_title, row["公告URL"])
            elif category == "实际下修":
                ws.append(
                    [
                        row["转债代码"],
                        row["转债简称"],
                        row.get("向下修正发布日期"),
                        row.get("转股价变动日期"),
                        row.get("修正前转股价"),
                        row.get("修正后转股价"),
                        row.get("修正转股价应不低于"),
                        display_title,
                    ]
                )
                ws.cell(ws.max_row, 3).number_format = "yyyy/m/d h:mm"
                ws.cell(ws.max_row, 4).number_format = "yyyy/m/d"
                set_hyperlink(ws.cell(ws.max_row, 8), display_title, row["公告URL"])
            else:
                ws.append([row["转债代码"], row["转债简称"], row["公告时间"], display_title])
                ws.cell(ws.max_row, 3).number_format = "yyyy/m/d h:mm"
                set_hyperlink(ws.cell(ws.max_row, 4), display_title, row["公告URL"])
        style_sheet(ws, widths)

    ws = wb.create_sheet("查询结果明细")
    ws.append(["正股代码", "正股简称", "转债代码", "转债简称", "公告时间", "公告类型", "公告链接"])
    for row in rows:
        display_title = f"{row['正股简称']}:{row['公告标题']}"
        ws.append([row["正股代码"], row["正股简称"], row["转债代码"], row["转债简称"], row["公告时间"], row["公告类型"], display_title])
        ws.cell(ws.max_row, 5).number_format = "yyyy/m/d h:mm"
        set_hyperlink(ws.cell(ws.max_row, 7), display_title, row["公告URL"])
    style_sheet(ws, [12, 12, 13, 14, 18, 18, 85])

    ws = wb.create_sheet("说明")
    ws.append(["项目", "说明"])
    ws.append(["查询口径", f"巨潮可转债公告；公告日期 {QUERY_RANGE_LABEL}；标题含“修”；排除“{'、'.join(EXCLUDE_TITLE_WORDS)}”"])
    ws.append(["输出文件", str(OUTPUT_XLSX)])
    style_sheet(ws, [16, 100])
    wb.save(OUTPUT_XLSX)


def copy_style_from_row(ws, source_row, target_row, max_col):
    for column in range(1, max_col + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.number_format = source.number_format


def validate_xlsx_zip(path):
    try:
        with ZipFile(path) as archive:
            names = set(archive.namelist())
            return "[Content_Types].xml" in names and "xl/workbook.xml" in names
    except BadZipFile:
        return False


def backup_existing_workbook(path):
    REVISION_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = REVISION_BACKUP_DIR / f"{path.stem}_{datetime.now():%Y%m%d_%H%M%S}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def save_workbook_safely(wb, path):
    path = Path(path)
    backup_path = backup_existing_workbook(path) if path.exists() else None
    fd, temp_name = tempfile.mkstemp(prefix=f"{path.stem}_", suffix=path.suffix, dir=str(path.parent))
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        wb.save(temp_path)
        if not validate_xlsx_zip(temp_path):
            raise RuntimeError(f"临时保存文件不是有效xlsx：{temp_path}")
        os.replace(temp_path, path)
    finally:
        try:
            temp_path.unlink(missing_ok=True)
        except Exception:
            pass
    return backup_path


def snapshot_hyperlinks(ws):
    """记录当前工作表所有单元格超链接；用于绕开 openpyxl 插行不平移 hyperlink 的问题。"""
    links = []
    for row in ws.iter_rows():
        for cell in row:
            hyperlink = cell.hyperlink
            if hyperlink:
                links.append(
                    {
                        "row": cell.row,
                        "column": cell.column,
                        "target": hyperlink.target,
                        "location": hyperlink.location,
                        "tooltip": hyperlink.tooltip,
                        "display": hyperlink.display,
                    }
                )
    return links


def clear_hyperlinks(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink:
                cell.hyperlink = None


def restore_hyperlinks(ws, links):
    for item in links:
        cell = ws.cell(item["row"], item["column"])
        target = item.get("target")
        location = item.get("location")
        if target:
            cell.hyperlink = target
        elif location:
            cell.hyperlink = f"#{location}"
        else:
            continue
        if item.get("tooltip"):
            cell.hyperlink.tooltip = item["tooltip"]
        if item.get("display"):
            cell.hyperlink.display = item["display"]


def insert_rows_preserve_hyperlinks(ws, idx, amount=1):
    """
    openpyxl 的 insert_rows 会移动单元格文本/值，但不会可靠移动 worksheet rels 中的超链接。
    这里手动将原有超链接按插入行数平移，避免出现“文本是 A 公告、链接指向 B 公告”。
    """
    links = snapshot_hyperlinks(ws)
    ws.insert_rows(idx, amount)
    shifted_links = []
    for item in links:
        shifted = item.copy()
        if shifted["row"] >= idx:
            shifted["row"] += amount
        shifted_links.append(shifted)
    clear_hyperlinks(ws)
    restore_hyperlinks(ws, shifted_links)


def cell_snapshot(cell):
    hyperlink = None
    if cell.hyperlink:
        hyperlink = {
            "target": cell.hyperlink.target,
            "location": cell.hyperlink.location,
            "tooltip": cell.hyperlink.tooltip,
            "display": cell.hyperlink.display,
        }
    return {
        "value": cell.value,
        "style": copy.copy(cell._style) if cell.has_style else None,
        "font": copy.copy(cell.font),
        "fill": copy.copy(cell.fill),
        "border": copy.copy(cell.border),
        "alignment": copy.copy(cell.alignment),
        "number_format": cell.number_format,
        "protection": copy.copy(cell.protection),
        "hyperlink": hyperlink,
        "comment": copy.copy(cell.comment) if cell.comment else None,
    }


def restore_cell(snapshot, target):
    target.value = snapshot["value"]
    if snapshot["style"] is not None:
        target._style = copy.copy(snapshot["style"])
    target.font = copy.copy(snapshot["font"])
    target.fill = copy.copy(snapshot["fill"])
    target.border = copy.copy(snapshot["border"])
    target.alignment = copy.copy(snapshot["alignment"])
    target.number_format = snapshot["number_format"]
    target.protection = copy.copy(snapshot["protection"])
    target.hyperlink = None
    hyperlink = snapshot["hyperlink"]
    if hyperlink:
        if hyperlink.get("target"):
            target.hyperlink = hyperlink["target"]
        elif hyperlink.get("location"):
            target.hyperlink = f"#{hyperlink['location']}"
        if target.hyperlink:
            target.hyperlink.tooltip = hyperlink.get("tooltip")
            target.hyperlink.display = hyperlink.get("display")
    target.comment = copy.copy(snapshot["comment"]) if snapshot["comment"] else None


def sort_sheet_by_date_desc(ws, date_column, start_row=2):
    """
    按日期列降序重排数据行。逐单元格复制值/样式/超链接，确保公告文本和链接不串行。
    """
    max_row = ws.max_row
    max_col = ws.max_column
    rows = []
    for row_index in range(start_row, max_row + 1):
        if all(ws.cell(row_index, col).value in (None, "") for col in range(1, max_col + 1)):
            continue
        key_dt = parse_datetime(ws.cell(row_index, date_column).value) or datetime.min
        row_cells = [cell_snapshot(ws.cell(row_index, col)) for col in range(1, max_col + 1)]
        rows.append((key_dt, row_index, row_cells))
    rows.sort(key=lambda item: (item[0], item[1]), reverse=True)

    for offset, (_, _, row_cells) in enumerate(rows, start=start_row):
        for col, source in enumerate(row_cells, start=1):
            restore_cell(source, ws.cell(offset, col))


def refresh_down_sheet_formulas(ws):
    """重建下修 sheet 中依赖当前行号的 I/J 列公式，避免插行、排序后公式引用串行。"""
    headers = [ws.cell(1, column).value for column in range(1, ws.max_column + 1)]
    col = {name: index + 1 for index, name in enumerate(headers)}
    distance_col = col.get("距离修正底线")
    bottom_col = next(
        (index + 1 for index, name in enumerate(headers) if str(name or "").startswith("是否下修到底")),
        None,
    )
    if not distance_col and not bottom_col:
        return

    for row_index in range(2, ws.max_row + 1):
        row_has_data = any(ws.cell(row_index, column).value not in (None, "") for column in range(1, ws.max_column + 1))
        if not row_has_data:
            continue

        effective_value = ws.cell(row_index, 5).value
        revised_price = ws.cell(row_index, 7).value
        floor_price = ws.cell(row_index, 8).value
        actual_notice = ws.cell(row_index, 13).value
        has_actual_or_failure = any(value not in (None, "") for value in (effective_value, revised_price, floor_price, actual_notice))

        if distance_col:
            distance_cell = ws.cell(row_index, distance_col)
            if revised_price not in (None, "") and floor_price not in (None, ""):
                distance_cell.value = f'=IFERROR((G{row_index}-H{row_index})/H{row_index},"")'
                distance_cell.number_format = "0.00%"
            else:
                distance_cell.value = None

        if bottom_col:
            bottom_cell = ws.cell(row_index, bottom_col)
            if has_actual_or_failure:
                bottom_cell.value = (
                    f'=IFERROR(IF(E{row_index}="下修失败",E{row_index},'
                    f'_xlfn.IFS(OR(G{row_index}<H{row_index}+0.01,I{row_index}=0),"是",'
                    f'(I{row_index}<0.05)*AND(I{row_index}>0),"基本到底",'
                    f'I{row_index}>0.05,"否",'
                    f'OR(E{row_index}="下修失败",E{row_index}="违规取消"),"失败",TRUE,"")),E{row_index})'
                )
            else:
                bottom_cell.value = None


def fill_database(rows):
    if not ENABLE_DATABASE_WRITE:
        print("数据库写入已暂停：当前 openpyxl 写库会导致外链公式/超链接错位。请改用 Excel COM 写入后再开启。")
        return {"数据库写入": {"source": len(rows), "inserted": 0, "updated": 0, "skipped": len(rows)}}

    if not DATABASE_XLSX.exists():
        return {"数据库": f"未找到：{DATABASE_XLSX}"}

    try:
        wb = load_workbook(DATABASE_XLSX)
    except BadZipFile as exc:
        raise RuntimeError(
            f"数据库文件不是完整有效的xlsx，可能上次保存时被中断或文件被占用：{DATABASE_XLSX}"
        ) from exc
    report = defaultdict(lambda: {"source": 0, "inserted": 0, "updated": 0, "skipped": 0, "details": []})

    def mark_summary_status(row, action):
        row["摘要状态"] = action
        return row

    def existing_key(code, time_value, title):
        return (str(code or "").strip(), day_key(time_value), normalize_text(title))

    def days_between(left, right):
        left_dt = parse_datetime(left)
        right_dt = parse_datetime(right)
        if not left_dt or not right_dt:
            return None
        return abs((left_dt.date() - right_dt.date()).days)

    def extract_bond_aliases(text):
        text = str(text or "")
        aliases = set(re.findall(r"[\u4e00-\u9fa5A-Za-z0-9*]+转(?:债|0?2|[A-Za-z0-9]+)", text))
        aliases.update(re.findall(r"[\u4e00-\u9fa5A-Za-z0-9*]+转债", text))
        cleaned = set()
        for alias in aliases:
            alias = alias.replace('"', "").replace("“", "").replace("”", "")
            if len(alias) < 3:
                continue
            if "可转债" in alias:
                continue
            if alias.startswith(("关于", "公司", "本次", "预计", "触发")):
                continue
            cleaned.add(alias)
        return cleaned

    def is_valid_bond_code(code):
        code = str(code or "").strip()
        return bool(code) and code != "未匹配"

    def has_same_bond_within_days(ws, code, time_value, code_col, time_col, max_days=10, title=None, title_col=None):
        incoming_aliases = extract_bond_aliases(title)
        for index in range(2, ws.max_row + 1):
            same_code = str(ws.cell(index, code_col).value or "").strip() == str(code or "").strip() and is_valid_bond_code(code)
            same_alias = False
            if incoming_aliases and title_col:
                existing_aliases = extract_bond_aliases(ws.cell(index, title_col).value)
                same_alias = bool(incoming_aliases & existing_aliases)
            if not (same_code or same_alias):
                continue
            delta = days_between(ws.cell(index, time_col).value, time_value)
            if delta is not None and delta <= max_days:
                return index
        return None

    def is_empty_cell_value(value):
        return value is None or str(value).strip() in ("", "——")

    if "不下修" in wb.sheetnames:
        ws = wb["不下修"]
        insert_at = 2
        for row in [item for item in rows if item["公告类型"] == "不下修"]:
            report["不下修"]["source"] += 1
            display_title = f"{row['正股简称']}:{row['公告标题']}"
            # 规则：同一转债，已有公告时间与抓取公告时间差异在 10 天以内，认定为同一次公告，跳过。
            existing_row = has_same_bond_within_days(ws, row["转债代码"], row["公告时间"], 1, 3, 10, display_title, 5)
            if existing_row:
                incoming_commitment = row.get("承诺何日之前不行使", "——")
                updated = False
                if not is_empty_cell_value(incoming_commitment) and is_empty_cell_value(ws.cell(existing_row, 4).value):
                    ws.cell(existing_row, 4).value = incoming_commitment
                    ws.cell(existing_row, 4).number_format = "yyyy/m/d"
                    updated = True
                if not ws.cell(existing_row, 5).hyperlink:
                    ws.cell(existing_row, 5).hyperlink = row["公告URL"]
                    updated = True
                if updated:
                    report["不下修"]["updated"] += 1
                    report["不下修"]["details"].append({"action": "更新", "row": mark_summary_status(row, "更新")})
                else:
                    report["不下修"]["skipped"] += 1
                    mark_summary_status(row, "已存在")
                continue
            insert_rows_preserve_hyperlinks(ws, insert_at)
            copy_style_from_row(ws, insert_at + 1, insert_at, 5)
            values = [row["转债代码"], row["转债简称"], row["公告时间"], row.get("承诺何日之前不行使", "——"), display_title]
            for column, value in enumerate(values, start=1):
                ws.cell(insert_at, column).value = value
            ws.cell(insert_at, 3).number_format = "yyyy/m/d h:mm"
            ws.cell(insert_at, 5).hyperlink = row["公告URL"]
            report["不下修"]["inserted"] += 1
            report["不下修"]["details"].append({"action": "新增", "row": mark_summary_status(row, "新增")})
            insert_at += 1

    if "预计下修" in wb.sheetnames:
        ws = wb["预计下修"]
        insert_at = 2
        for row in [item for item in rows if item["公告类型"] == "预计下修"]:
            report["预计下修"]["source"] += 1
            display_title = f"{row['正股简称']}:{row['公告标题']}"
            # 规则：同一转债，已有公告时间与抓取公告时间差异在 10 天以内，认定为同一次公告，跳过。
            if has_same_bond_within_days(ws, row["转债代码"], row["公告时间"], 1, 3, 10, display_title, 4):
                report["预计下修"]["skipped"] += 1
                mark_summary_status(row, "已存在")
                continue
            insert_rows_preserve_hyperlinks(ws, insert_at)
            copy_style_from_row(ws, insert_at + 1, insert_at, max(5, ws.max_column))
            for column, value in enumerate([row["转债代码"], row["转债简称"], row["公告时间"], display_title], start=1):
                ws.cell(insert_at, column).value = value
            ws.cell(insert_at, 3).number_format = "yyyy/m/d h:mm"
            ws.cell(insert_at, 4).hyperlink = row["公告URL"]
            report["预计下修"]["inserted"] += 1
            report["预计下修"]["details"].append({"action": "新增", "row": mark_summary_status(row, "新增")})
            insert_at += 1

    if "下修" in wb.sheetnames:
        ws = wb["下修"]
        headers = [ws.cell(1, column).value for column in range(1, ws.max_column + 1)]
        col = {name: index + 1 for index, name in enumerate(headers)}
        bond_name_col = col.get("转债简称") or col.get("转债名称")
        if not bond_name_col:
            raise KeyError("下修 sheet 缺少“转债简称/转债名称”列")

        def same_down_bond(index, row, display_title):
            incoming_code = str(row.get("转债代码") or "").strip()
            existing_code = str(ws.cell(index, col["转债代码"]).value or "").strip()
            if is_valid_bond_code(incoming_code) and incoming_code == existing_code:
                return True
            incoming_aliases = set()
            incoming_aliases.update(extract_bond_aliases(row.get("转债简称")))
            incoming_aliases.update(extract_bond_aliases(display_title))
            existing_aliases = set()
            existing_aliases.update(extract_bond_aliases(ws.cell(index, bond_name_col).value))
            if "提议公告" in col:
                existing_aliases.update(extract_bond_aliases(ws.cell(index, col["提议公告"]).value))
            if "下修公告" in col:
                existing_aliases.update(extract_bond_aliases(ws.cell(index, col["下修公告"]).value))
            return bool(incoming_aliases & existing_aliases)

        def has_down_completion(index):
            completion_fields = [
                "下修公告",
                "转股价变动日期",
                "修正前转股价",
                "修正后转股价",
                "修正转股价应不低于",
                "下修失败/取消公告日期",
            ]
            for field in completion_fields:
                if field in col and ws.cell(index, col[field]).value not in (None, ""):
                    return True
            return False

        def update_provisional_meeting_date(index, row):
            meeting_date = row.get("股东大会日期")
            if not meeting_date or "向下修正发布日期" not in col:
                return False
            if has_down_completion(index):
                return False
            cell = ws.cell(index, col["向下修正发布日期"])
            old_value = parse_datetime(cell.value)
            new_value = parse_datetime(meeting_date)
            if old_value == new_value:
                return False
            cell.value = meeting_date
            cell.number_format = "yyyy/m/d h:mm"
            return True

        insert_at = 2
        for row in [item for item in rows if item["公告类型"] == "董事会提议下修"]:
            report["董事会提议下修"]["source"] += 1
            display_title = f"{row['正股简称']}:{row['公告标题']}"
            # 规则：同一转债当前仍有下修流程（已有提议但尚无实际下修/失败公告），认定为同一次下修，跳过插入。
            # 若抓取到股东会召开日，则临时填入“向下修正发布日期”；实际下修公告到来后再覆盖。
            open_flow_row = None
            for index in range(2, ws.max_row + 1):
                if not same_down_bond(index, row, display_title):
                    continue
                if (
                    ws.cell(index, col["董事会发布日期"]).value
                    and ("提议公告" not in col or ws.cell(index, col["提议公告"]).value)
                    and not has_down_completion(index)
                ):
                    open_flow_row = index
                    break
            if open_flow_row:
                if update_provisional_meeting_date(open_flow_row, row):
                    report["董事会提议下修"]["updated"] += 1
                    report["董事会提议下修"]["details"].append({"action": "更新", "row": mark_summary_status(row, "更新")})
                else:
                    report["董事会提议下修"]["skipped"] += 1
                    mark_summary_status(row, "已存在")
                continue

            # 同日同标题也视为已存在，防止重复插入。
            exists_same = None
            for index in range(2, ws.max_row + 1):
                if not same_down_bond(index, row, display_title):
                    continue
                board_delta = days_between(ws.cell(index, col["董事会发布日期"]).value, row["公告时间"])
                if board_delta is not None and board_delta <= 3:
                    exists_same = index
                    break
                if existing_key(ws.cell(index, col["转债代码"]).value, ws.cell(index, col["董事会发布日期"]).value, ws.cell(index, col["提议公告"]).value) == existing_key(row["转债代码"], row["公告时间"], display_title):
                    exists_same = index
                    break
            if exists_same:
                if update_provisional_meeting_date(exists_same, row):
                    report["董事会提议下修"]["updated"] += 1
                    report["董事会提议下修"]["details"].append({"action": "更新", "row": mark_summary_status(row, "更新")})
                else:
                    report["董事会提议下修"]["skipped"] += 1
                    mark_summary_status(row, "已存在")
                continue

            insert_rows_preserve_hyperlinks(ws, insert_at)
            copy_style_from_row(ws, insert_at + 1, insert_at, ws.max_column)
            for column in range(1, ws.max_column + 1):
                ws.cell(insert_at, column).value = None
                ws.cell(insert_at, column).hyperlink = None
            ws.cell(insert_at, col["转债代码"]).value = row["转债代码"]
            ws.cell(insert_at, bond_name_col).value = row["转债简称"]
            ws.cell(insert_at, col["董事会发布日期"]).value = row["公告时间"]
            ws.cell(insert_at, col["董事会发布日期"]).number_format = "yyyy/m/d h:mm"
            if row.get("股东大会日期") and "向下修正发布日期" in col:
                ws.cell(insert_at, col["向下修正发布日期"]).value = row["股东大会日期"]
                ws.cell(insert_at, col["向下修正发布日期"]).number_format = "yyyy/m/d h:mm"
            ws.cell(insert_at, col["提议公告"]).value = display_title
            ws.cell(insert_at, col["提议公告"]).hyperlink = row["公告URL"]
            report["董事会提议下修"]["inserted"] += 1
            report["董事会提议下修"]["details"].append({"action": "新增", "row": mark_summary_status(row, "新增")})
            insert_at += 1

        for row in [item for item in rows if item["公告类型"] == "实际下修"]:
            report["实际下修"]["source"] += 1
            display_title = f"{row['正股简称']}:{row['公告标题']}"
            target_row = None

            # 规则：实际下修公告优先写入对应的已提议下修行。
            # 匹配顺序：同一转债、有董事会发布日期和提议公告、董事会日期早于实际下修公告；
            # 优先未完成流程，其次选择公告日前最近的一条提议行。这样可覆盖华海、章鼓、文科等已存在提议行。
            candidates = []
            actual_time = parse_datetime(row.get("向下修正发布日期") or row.get("公告时间"))
            for index in range(2, ws.max_row + 1):
                if not same_down_bond(index, row, display_title):
                    continue
                board_time = parse_datetime(ws.cell(index, col["董事会发布日期"]).value)
                proposal_title = ws.cell(index, col["提议公告"]).value
                if not board_time or not proposal_title or not actual_time:
                    continue
                gap = (actual_time.date() - board_time.date()).days
                if gap < 0 or gap > 120:
                    continue
                open_priority = 0 if not has_down_completion(index) else 1
                candidates.append((open_priority, gap, index))
            if candidates:
                target_row = sorted(candidates, key=lambda item: (item[0], item[1]))[0][2]

            # 若没有提议行，则尝试匹配同一转债且实际下修发布日期同日的既有行。
            if target_row is None:
                for index in range(2, ws.max_row + 1):
                    if not same_down_bond(index, row, display_title):
                        continue
                    if day_key(ws.cell(index, col["向下修正发布日期"]).value) == day_key(row.get("向下修正发布日期")):
                        target_row = index
                        break

            # 仍无匹配才新插入一行。
            if target_row is None:
                insert_rows_preserve_hyperlinks(ws, insert_at)
                copy_style_from_row(ws, insert_at + 1, insert_at, ws.max_column)
                for column in range(1, ws.max_column + 1):
                    ws.cell(insert_at, column).value = None
                    ws.cell(insert_at, column).hyperlink = None
                target_row = insert_at
                ws.cell(target_row, col["转债代码"]).value = row["转债代码"]
                ws.cell(target_row, bond_name_col).value = row["转债简称"]
                insert_at += 1
                report["实际下修"]["inserted"] += 1
                action = "新增"
            else:
                action = "更新"

            field_map = {
                "向下修正发布日期": row.get("向下修正发布日期"),
                "转股价变动日期": row.get("转股价变动日期"),
                "修正前转股价": row.get("修正前转股价"),
                "修正后转股价": row.get("修正后转股价"),
                "修正转股价应不低于": row.get("修正转股价应不低于"),
                "下修公告": display_title,
            }
            existing_notice_cell = ws.cell(target_row, col["下修公告"]) if "下修公告" in col else None
            existing_notice_link = existing_notice_cell.hyperlink.target if existing_notice_cell and existing_notice_cell.hyperlink else None
            target_had_actual_completion = has_down_completion(target_row) if action == "更新" else False
            same_notice_for_update = (
                action == "更新"
                and row.get("公告URL")
                and (
                    existing_notice_link == row.get("公告URL")
                    or normalize_text(existing_notice_cell.value if existing_notice_cell else None) == normalize_text(display_title)
                )
            )
            changed = action == "新增"
            for field, value in field_map.items():
                if field in col and value is not None:
                    cell = ws.cell(target_row, col[field])
                    old_value = cell.value
                    old_hyperlink = cell.hyperlink.target if cell.hyperlink else None
                    if field == "向下修正发布日期":
                        cell.number_format = "yyyy/m/d h:mm"
                    if field == "转股价变动日期":
                        cell.number_format = "yyyy/m/d"
                    # 实际下修的历史库可能已有人工校验值；不同公告不覆盖已有非空字段。
                    # 同一公告链接重跑时允许校正抽取规则升级前写入的错误字段。
                    if action == "更新" and field != "下修公告" and old_value not in (None, "") and not same_notice_for_update:
                        if not (field == "向下修正发布日期" and not target_had_actual_completion):
                            continue
                    if action == "更新" and field == "下修公告" and old_value not in (None, ""):
                        if not old_hyperlink and row.get("公告URL"):
                            cell.hyperlink = row["公告URL"]
                            changed = True
                        continue
                    if field == "下修公告":
                        value_changed = normalize_text(old_value) != normalize_text(value) or old_hyperlink != row["公告URL"]
                    elif isinstance(value, datetime) or isinstance(old_value, datetime):
                        value_changed = parse_datetime(old_value) != parse_datetime(value)
                    else:
                        value_changed = old_value != value
                    if value_changed:
                        changed = True
                    cell.value = value
            old_notice_cell = ws.cell(target_row, col["下修公告"])
            old_notice_link = old_notice_cell.hyperlink.target if old_notice_cell.hyperlink else None
            if old_notice_cell.value in (None, ""):
                old_notice_cell.value = display_title
                if row.get("公告URL"):
                    old_notice_cell.hyperlink = row["公告URL"]
                changed = True
            elif not old_notice_link and row.get("公告URL"):
                old_notice_cell.hyperlink = row["公告URL"]
                changed = True
            if action == "更新":
                if changed:
                    report["实际下修"]["updated"] += 1
                    mark_summary_status(row, "更新")
                else:
                    report["实际下修"]["skipped"] += 1
                    mark_summary_status(row, "已存在")
            if changed:
                report["实际下修"]["details"].append({"action": action, "row": mark_summary_status(row, action)})

    if "不下修" in wb.sheetnames:
        sort_sheet_by_date_desc(wb["不下修"], 3)
    if "预计下修" in wb.sheetnames:
        sort_sheet_by_date_desc(wb["预计下修"], 3)
    if "下修" in wb.sheetnames:
        ws = wb["下修"]
        sort_sheet_by_date_desc(ws, 3)
        headers = [ws.cell(1, column).value for column in range(1, ws.max_column + 1)]
        col = {name: index + 1 for index, name in enumerate(headers)}
        date_formats = {
            "董事会发布日期": "yyyy/m/d h:mm",
            "向下修正发布日期": "yyyy/m/d h:mm",
            "转股价变动日期": "yyyy/m/d",
        }
        for field, number_format in date_formats.items():
            if field not in col:
                continue
            for row_index in range(2, ws.max_row + 1):
                ws.cell(row_index, col[field]).number_format = number_format
        refresh_down_sheet_formulas(ws)

    changed = any(item.get("inserted", 0) or item.get("updated", 0) for item in report.values())
    if not changed:
        result = dict(report)
        result["写入提示"] = "没有需要写入或补全的记录，底稿未保存。"
        return result

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    backup_path = save_workbook_safely(wb, DATABASE_XLSX)
    result = dict(report)
    result["写入提示"] = f"底稿已保存：{DATABASE_XLSX}"
    if backup_path:
        result["备份提示"] = f"更新前备份：{backup_path}"
    return result

def format_category_names(rows):
    names = defaultdict(list)
    for row in rows:
        if row["公告类型"] in ("不下修", "预计下修", "董事会提议下修", "实际下修"):
            item = f"{row['转债代码']} {row['转债简称']}"
            if item not in names[row["公告类型"]]:
                names[row["公告类型"]].append(item)
    return names


def format_cn_date(value):
    dt = parse_datetime(value)
    if not dt:
        return "——"
    return f"{dt.year}年{dt.month}月{dt.day}日"


def effective_notice_date(row):
    dt = parse_datetime(row.get("公告时间"))
    if not dt:
        return None
    if dt.hour == 23 and dt.minute == 59 and dt.second == 59:
        return dt + timedelta(days=1)
    return dt


def format_price(value):
    if value is None or value == "":
        return "——"
    try:
        return f"{float(value):.2f}".rstrip("0").rstrip(".")
    except Exception:
        return str(value)


def detail_rows(db_report, category):
    return [item["row"] for item in db_report.get(category, {}).get("details", [])]


def summary_status_prefix(row):
    return ""


def announcement_date_prefix(row):
    return f"{format_cn_date(row.get('公告时间'))} "


def category_rows(rows, category):
    return [
        row
        for row in rows
        if row.get("公告类型") == category and row.get("摘要状态") in ("新增", "更新")
    ]


def build_summary(rows, db_report):
    lines = [f"查询区间：{QUERY_RANGE_LABEL}", f"数据库：{DATABASE_XLSX}", ""]

    proposal_rows = category_rows(rows, "董事会提议下修")
    lines.append("1、提议下修类：")
    if proposal_rows:
        for row in proposal_rows:
            lines.append(f"{announcement_date_prefix(row)}{summary_status_prefix(row)}【{row['转债简称']}】{row['正股简称']}:{row['公告标题']}")
            lines.append(f"本次股东大会日期：{format_cn_date(row.get('股东大会日期'))}")
            lines.append("")
    else:
        lines.append("无新增/更新")
        lines.append("")

    actual_rows = category_rows(rows, "实际下修")
    lines.append("2、实际下修类：")
    if actual_rows:
        for row in actual_rows:
            lines.append(f"{announcement_date_prefix(row)}{summary_status_prefix(row)}【{row['转债简称']}】{row['正股简称']}:{row['公告标题']}")
            lines.append(f"修正前转股价格：{format_price(row.get('修正前转股价'))} 元/股；")
            lines.append(f"修正后转股价格：{format_price(row.get('修正后转股价'))} 元/股；")
            lines.append(f"修正后转股价格生效日期：{format_cn_date(row.get('转股价变动日期'))}")
            lines.append(f"本次下修底价：{format_price(row.get('修正转股价应不低于'))} 元/股")
            lines.append("")
    else:
        lines.append("无新增/更新")
        lines.append("")

    no_down_rows = category_rows(rows, "不下修")

    lines.append("3、不下修：")
    if no_down_rows:
        for row in no_down_rows:
            lines.append(f"{announcement_date_prefix(row)}{summary_status_prefix(row)}【{row['转债简称']}】{row['正股简称']}:{row['公告标题']}")
            if row.get("承诺何日之前不行使") == "——":
                notice_dt = effective_notice_date(row)
                recalc_dt = notice_dt + timedelta(days=1) if notice_dt else None
                lines.append(f"本次不向下修正转股价格，从{format_cn_date(recalc_dt)}起重新计算")
            else:
                start_dt = effective_notice_date(row)
                lines.append(
                    f"本次不向下修正转股价格，且{format_cn_date(start_dt)}至{format_cn_date(row.get('承诺何日之前不行使'))}"
                    "如再次触发转股价格向下修正条款的，亦不提出向下修正方案"
                )
            lines.append("")
    else:
        lines.append("无新增/更新")
        lines.append("")

    expected_rows = category_rows(rows, "预计下修")
    lines.append("4、预计下修公告：")
    if expected_rows:
        for row in expected_rows:
            lines.append(f"{announcement_date_prefix(row)}{summary_status_prefix(row)}【{row['转债简称']}】{row['正股简称']}:{row['公告标题']}")
            lines.append("预计触发下修条件")
            lines.append("")
    else:
        lines.append("无新增/更新")

    write_message = db_report.get("写入提示")
    if write_message:
        lines.extend(["", write_message])
    backup_message = db_report.get("备份提示")
    if backup_message:
        lines.append(backup_message)

    return "\n".join(lines)


def run_redemption_update():
    # Embedded redemption notice updater; no external strong-redemption script is loaded.
    import argparse
    import calendar
    import json
    import os
    import re
    import shutil
    import subprocess
    import sys
    import tempfile
    import time
    from collections import defaultdict
    from dataclasses import asdict, dataclass
    from datetime import date, datetime, timedelta, timezone
    from pathlib import Path
    from urllib.parse import urlencode

    import pandas as pd
    from pypdf import PdfReader


    ROOT = Path(__file__).resolve().parents[2]
    DEFAULT_WORKBOOK = ROOT / "data/clauses/【华创固收】赎回和不赎回公告统计.xlsx"
    BACKUP_DIR = ROOT / "archive/backups" / "转债强赎公告更新"
    AUDIT_DIR = ROOT / "outputs" / "转债强赎公告更新"

    CNINFO_QUERY_URL = "https://www.cninfo.com.cn/new/hisAnnouncement/query"
    CNINFO_STATIC_PREFIX = "https://static.cninfo.com.cn/"
    THS_LOGIN_OK_CODES = (0, -201)
    EXCEL_TIMEOUT_SECONDS = 180

    NUMBERED_FOLLOWUP_RE = re.compile(r"第(?:二|三|四|五|六|七|八|九|十)次")
    NON_REDEMPTION_RE = re.compile(r"不.{0,12}赎回")
    EXCLUDE_TITLE_WORDS = ("停止", "预计")

    XL_CALCULATION_AUTOMATIC = -4105
    XL_UP = -4162
    XL_PASTE_FORMATS = -4122
    XL_SHIFT_DOWN = -4121


    @dataclass
    class Announcement:
        category: str
        bond_code: str
        bond_name: str
        stock_code: str
        stock_name: str
        announcement_time: datetime
        effective_date: date
        title: str
        url: str
        announcement_id: str
        match_note: str
        strong_notice_date: date | None = None
        non_redemption_notice_date: date | None = None
        stop_trading_date: date | None = None
        stop_conversion_date: date | None = None
        redemption_registration_date: date | None = None
        redemption_date: date | None = None
        redemption_schedule_pending: bool = False
        commitment_start: date | None = None
        commitment_deadline: date | None = None
        deadline_source: str = ""
        deadline_evidence: str = ""
        deadline_score: int = 0
        summary_action: str = ""

        @property
        def display_title(self) -> str:
            return f"{self.stock_name}:{self.title}" if self.stock_name else self.title


    def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
        parser = argparse.ArgumentParser(
            description="抓取最近若干日可转债赎回公告，匹配Excel条款日期并更新赎回/不赎回底稿。"
        )
        parser.add_argument(
            "--workbook",
            type=Path,
            default=DEFAULT_WORKBOOK,
            help=f"待更新底稿，默认：{DEFAULT_WORKBOOK}",
        )
        parser.add_argument("--days", type=int, default=10, help="向前抓取自然日数，默认10天。")
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="仅抓取、筛选和解析，不写入Excel底稿。",
        )
        parser.add_argument(
            "--keep-pdfs",
            action="store_true",
            help="保留本次下载的PDF；默认仅保留审计JSON。",
        )
        parser.add_argument(
            "--no-popup",
            action="store_true",
            help="不显示运行结果弹窗。",
        )
        return parser.parse_args(argv)


    def clean_title(value: object) -> str:
        return re.sub(r"<[^>]+>", "", str(value or "")).strip()


    def normalize_code(value: object) -> str:
        match = re.search(r"(\d{6})", str(value or ""))
        return match.group(1) if match else ""


    def normalize_bond_code(value: object) -> str:
        text = str(value or "").strip().upper()
        match = re.search(r"(\d{6})(?:\.(SH|SZ))?", text)
        if not match:
            return ""
        code, market = match.group(1), match.group(2)
        return f"{code}.{market}" if market else code


    def parse_datetime(value: object) -> datetime | None:
        if isinstance(value, datetime):
            return value.replace(tzinfo=None)
        if isinstance(value, date):
            return datetime(value.year, value.month, value.day)
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            if value > 0:
                return datetime(1899, 12, 30) + timedelta(days=float(value))
            return None
        text = str(value).strip()
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%Y/%m/%d %H:%M",
            "%Y/%m/%d",
        ):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                pass
        match = re.search(r"(20\d{2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", text)
        if match:
            return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        return None


    def cninfo_time(milliseconds: object) -> datetime:
        value = datetime.fromtimestamp(
            int(milliseconds) / 1000,
            tz=timezone.utc,
        ).astimezone(timezone(timedelta(hours=8))).replace(tzinfo=None)
        # 巨潮只有日期而无具体时点时返回北京时间00:00:00。
        # 延续下修公告脚本的口径，将其记为前一日23:59:59。
        if value.time() == datetime.min.time():
            return value - timedelta(seconds=1)
        return value


    def effective_notice_date(value: datetime) -> date:
        if value.hour == 23 and value.minute == 59 and value.second == 59:
            return (value + timedelta(days=1)).date()
        return value.date()


    def normalize_excel_date(value: object) -> date | None:
        if value in (None, "", 0):
            return None
        if isinstance(value, (datetime, date, pd.Timestamp)):
            return pd.Timestamp(value).date()
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            if value <= 0 or value < -2_000_000_000:
                return None
            if value < 2_958_466:
                return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        text = str(value).strip()
        if text.startswith("#") or text.lower() in {"fetching", "loading", "requesting"}:
            return None
        if re.fullmatch(r"\d{8}", text):
            parsed = pd.to_datetime(text, format="%Y%m%d", errors="coerce")
        else:
            parsed = pd.to_datetime(text, errors="coerce")
        return None if pd.isna(parsed) else parsed.date()


    def is_pending_excel_value(value: object) -> bool:
        return isinstance(value, str) and any(
            marker in value.strip().lower()
            for marker in ("fetching", "loading", "requesting")
        )


    def show_popup(title: str, message: str) -> None:
        try:
            import tkinter as tk

            root = tk.Tk()
            root.title(title)
            root.geometry("840x560")
            root.attributes("-topmost", True)
            frame = tk.Frame(root, padx=12, pady=12)
            frame.pack(fill="both", expand=True)
            text = tk.Text(frame, wrap="word", font=("KaiTi_GB2312", 10))
            text.insert("1.0", message)
            text.configure(state="disabled")
            text.pack(fill="both", expand=True)
            tk.Button(frame, text="关闭", command=root.destroy, width=12).pack(pady=(10, 0))
            root.after(90_000, root.destroy)
            root.mainloop()
        except Exception:
            print(f"\n{title}\n{message}")


    def ths_login() -> int:
        from iFinDPy import THS_GetErrorInfo, THS_iFinDLogin

        user_id, password = load_ifind_credentials()

        code = THS_iFinDLogin(user_id, password)
        if code not in THS_LOGIN_OK_CODES:
            try:
                detail = THS_GetErrorInfo(code)
            except Exception:
                detail = ""
            raise RuntimeError(f"iFinD登录失败（{code}）：{detail}")
        print_ifind_usage()
        return code


    def get_trade_date_offset(offset: int = 0) -> str:
        from iFinDPy import THS_Date_Offset

        today = datetime.now().strftime("%Y-%m-%d")
        result = THS_Date_Offset(
            "212001",
            f"dateType:0,period:D,offset:{offset},dateFormat:0,output:singledate",
            today,
        ).data
        if result is None:
            raise RuntimeError("THS_Date_Offset未返回最近交易日。")
        if isinstance(result, (list, tuple)):
            result = result[0]
        return str(result)[:10]


    def get_last_trade_date() -> str:
        return get_trade_date_offset(0)


    def fetch_active_cb_universe(last_date: str) -> pd.DataFrame:
        """复用日报/赎回日更口径：未到期列表中，仅保留当日仍有交易的转债。"""
        from iFinDPy import THS_BD, THS_DR, THS_DS

        formatted_date = datetime.strptime(last_date, "%Y-%m-%d").strftime("%Y%m%d")
        cb_list = THS_DR(
            "p00570",
            f"jyzt=未到期;sfdb=全部;jysc=全部;edate={formatted_date}",
            "jydm:Y,jydm_mc:Y,p00570_f001:Y,p00570_f019:Y",
            "format:dataframe",
        ).data
        if cb_list is None or cb_list.empty:
            raise RuntimeError(f"THS_DR未返回未到期可转债列表：{last_date}")

        all_codes = cb_list.set_index("jydm").index.astype(str).tolist()
        code_string = ",".join(all_codes)
        turnover = THS_DS(
            code_string,
            "ths_turnover_ratio_cbond",
            "",
            "Fill:Blank,mode:thscode",
            last_date,
            last_date,
        ).data
        if turnover is None or turnover.empty:
            raise RuntimeError(f"THS_DS未返回可转债交易状态：{last_date}")

        turnover = turnover.set_index("time").T
        first_column = turnover.iloc[:, 0]
        last_column = turnover.iloc[:, -1]
        active_mask = ~(
            (first_column.isna() & last_column.isna())
            | ((first_column == 0) & (last_column == 0))
        )
        active_codes = turnover.index[active_mask].astype(str).tolist()
        if not active_codes:
            raise RuntimeError(f"未筛选出存续交易可转债：{last_date}")

        raw = THS_BD(
            ",".join(active_codes),
            "ths_convertible_debt_short_name_cbond;ths_stock_code_cbond;"
            "ths_stock_short_name_cbond;ths_issue_method_cbond;"
            "ths_trading_status_bond;ths_bond_balance_cbond;ths_listed_date_cbond",
            f";;;;;{last_date};",
        ).data
        if raw is None or raw.empty:
            raise RuntimeError("THS_BD未返回存续交易转债基础信息。")

        basic = raw.set_index("thscode").rename_axis("转债代码")
        basic.columns = [
            "转债简称",
            "正股代码",
            "正股简称",
            "发行方式",
            "交易状态",
            "转债余额",
            "上市日期",
        ]
        basic = basic[~basic["发行方式"].astype(str).str.contains("定向", na=False)]
        basic = basic[~basic.index.astype(str).str.contains("NQ", na=False)]
        basic = basic[~basic["交易状态"].astype(str).str.contains("终止上市", na=False)]
        basic.index = basic.index.astype(str).str.strip()
        return basic


    def fetch_latest_active_cb_universe(max_lookback: int = 5) -> tuple[str, pd.DataFrame]:
        errors = []
        for offset in range(0, -max_lookback - 1, -1):
            trade_date = get_trade_date_offset(offset)
            try:
                cb_basic = fetch_active_cb_universe(trade_date)
            except RuntimeError as exc:
                errors.append(f"{trade_date}: {exc}")
                continue
            if not cb_basic.empty:
                return trade_date, cb_basic
        raise RuntimeError("未找到可用的存续交易转债列表：" + "；".join(errors))


    def fetch_excel_clause_dates(bond_codes: list[str]) -> pd.DataFrame:
        import pythoncom
        import win32com.client

        codes = [str(code) for code in bond_codes]
        result = pd.DataFrame(
            {
                "强赎公告提示日期_函数": [None] * len(codes),
                "最新不强赎公告日期_函数": [None] * len(codes),
            },
            index=codes,
        )
        if not codes:
            return result

        excel = None
        workbook = None
        pythoncom.CoInitialize()
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            excel.AskToUpdateLinks = False
            try:
                excel.Calculation = XL_CALCULATION_AUTOMATIC
            except Exception:
                pass

            workbook = excel.Workbooks.Add()
            sheet = workbook.Worksheets(1)
            sheet.Name = "强赎条款日期批量查询"
            sheet.Range("A1:C1").Value = (("代码", "强赎公告提示日期", "最新不强赎公告日期"),)
            last_row = len(codes) + 1
            sheet.Range(f"A2:A{last_row}").Value = tuple((code,) for code in codes)

            for anchor, formula, fill_range in (
                ("B2", "=@cb_clause_calloption_indicativedatey(A2)", f"B2:B{last_row}"),
                ("C2", "=@cb_clause_calloption_indicativedaten(A2)", f"C2:C{last_row}"),
            ):
                cell = sheet.Range(anchor)
                try:
                    cell.Formula2 = formula
                except Exception:
                    cell.Formula = formula
                if last_row > 2:
                    sheet.Range(fill_range).FillDown()

            excel.CalculateFullRebuild()
            try:
                excel.CalculateUntilAsyncQueriesDone()
            except Exception:
                pass

            deadline = time.time() + EXCEL_TIMEOUT_SECONDS
            rows = ()
            while True:
                raw_values = sheet.Range(f"B2:C{last_row}").Value2
                rows = raw_values if isinstance(raw_values, tuple) else ((raw_values,),)
                pending = [
                    value
                    for row in rows
                    for value in (row if isinstance(row, tuple) else (row,))
                    if is_pending_excel_value(value)
                ]
                if excel.CalculationState == 0 and not pending:
                    break
                if time.time() >= deadline:
                    raise TimeoutError(f"Excel条款日期计算超时，仍有{len(pending)}个待返回值。")
                time.sleep(0.5)

            for code, row in zip(codes, rows):
                values = row if isinstance(row, tuple) else (row,)
                result.at[code, "强赎公告提示日期_函数"] = normalize_excel_date(
                    values[0] if len(values) > 0 else None
                )
                result.at[code, "最新不强赎公告日期_函数"] = normalize_excel_date(
                    values[1] if len(values) > 1 else None
                )
        finally:
            if workbook is not None:
                try:
                    workbook.Close(SaveChanges=False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
        return result


    def post_cninfo(page_num: int, start_date: str, end_date: str, page_size: int = 30) -> dict:
        data = {
            "pageNum": str(page_num),
            "pageSize": str(page_size),
            "column": "szse",
            "tabName": "fulltext",
            "plate": "",
            "stock": "",
            "searchkey": "赎回",
            "secid": "",
            "category": "category_kzzq_szsh;",
            "trade": "",
            "seDate": f"{start_date}~{end_date}",
            "sortName": "",
            "sortType": "",
            "isHLtitle": "true",
        }
        command = [
            "curl.exe",
            "-sS",
            "-L",
            "--compressed",
            "--connect-timeout",
            "30",
            CNINFO_QUERY_URL,
            "-H",
            "User-Agent: Mozilla/5.0",
            "-H",
            "Referer: https://www.cninfo.com.cn/new/commonUrl/pageOfSearch?"
            "url=disclosure/list/search&checkedCategory=category_kzzq_szsh",
            "-H",
            "X-Requested-With: XMLHttpRequest",
            "--data",
            urlencode(data, safe=";"),
        ]
        return run_cninfo_json_request(
            command,
            context=f"巨潮接口请求失败：page={page_num}",
            timeout=360,
        )


    def fetch_cninfo_announcements(start_date: str, end_date: str) -> list[dict]:
        all_items: list[dict] = []
        page_size = 30
        for page_num in range(1, 50):
            payload = post_cninfo(page_num, start_date, end_date, page_size)
            items = payload.get("announcements") or []
            if not items:
                break
            all_items.extend(items)
            if len(items) < page_size:
                break
        unique = {}
        for item in all_items:
            unique[str(item.get("announcementId"))] = item
        return list(unique.values())


    def map_announcement_to_bond(
        item: dict,
        cb_basic: pd.DataFrame,
    ) -> tuple[str | None, str]:
        stock6 = normalize_code(item.get("secCode"))
        title = clean_title(item.get("announcementTitle"))
        candidates = cb_basic[cb_basic["正股代码"].map(normalize_code).eq(stock6)]
        if candidates.empty:
            return None, "正股未匹配当前存续交易转债"

        title_matches = [
            code
            for code, row in candidates.iterrows()
            if str(row["转债简称"]).replace("（退市）", "").replace("(退市)", "") in title
        ]
        if len(title_matches) == 1:
            return title_matches[0], "按转债简称匹配"
        if len(candidates) == 1:
            return candidates.index[0], "按正股唯一匹配"
        return None, f"同一正股存在{len(candidates)}只转债且标题未唯一匹配"


    def build_matched_announcements(
        raw_items: list[dict],
        cb_basic: pd.DataFrame,
        clause_dates: pd.DataFrame,
    ) -> tuple[list[Announcement], list[dict]]:
        captured: list[Announcement] = []
        excluded: list[dict] = []

        for item in raw_items:
            title = clean_title(item.get("announcementTitle"))
            exclusion_reason = ""
            if "赎回" not in title:
                exclusion_reason = "标题不含赎回"
            elif any(word in title for word in EXCLUDE_TITLE_WORDS):
                matched_word = next(word for word in EXCLUDE_TITLE_WORDS if word in title)
                exclusion_reason = f"含排除词：{matched_word}"
            elif NUMBERED_FOLLOWUP_RE.search(title):
                exclusion_reason = "第二至第十次后续公告"

            try:
                announcement_time = cninfo_time(item.get("announcementTime"))
            except Exception:
                announcement_time = datetime.min
                exclusion_reason = exclusion_reason or "公告时间无法解析"

            if exclusion_reason:
                excluded.append(
                    {
                        "公告时间": announcement_time,
                        "正股代码": normalize_code(item.get("secCode")),
                        "公告标题": title,
                        "排除原因": exclusion_reason,
                    }
                )
                continue

            bond_code, match_note = map_announcement_to_bond(item, cb_basic)
            if not bond_code:
                excluded.append(
                    {
                        "公告时间": announcement_time,
                        "正股代码": normalize_code(item.get("secCode")),
                        "公告标题": title,
                        "排除原因": match_note,
                    }
                )
                continue

            category = "不赎回" if NON_REDEMPTION_RE.search(title) else "赎回"
            notice_date = effective_notice_date(announcement_time)
            target_column = (
                "最新不强赎公告日期_函数"
                if category == "不赎回"
                else "强赎公告提示日期_函数"
            )
            target_date = clause_dates.at[bond_code, target_column]
            if target_date != notice_date:
                excluded.append(
                    {
                        "公告时间": announcement_time,
                        "正股代码": normalize_code(item.get("secCode")),
                        "转债代码": bond_code,
                        "公告标题": title,
                        "排除原因": f"公告对应日{notice_date}与函数日期{target_date or '空'}不一致",
                    }
                )
                continue

            captured.append(
                Announcement(
                    category=category,
                    bond_code=str(bond_code),
                    bond_name=str(cb_basic.at[bond_code, "转债简称"]),
                    stock_code=normalize_code(item.get("secCode")),
                    stock_name=str(item.get("secName") or ""),
                    announcement_time=announcement_time,
                    effective_date=notice_date,
                    title=title,
                    url=CNINFO_STATIC_PREFIX + str(item.get("adjunctUrl") or ""),
                    announcement_id=str(item.get("announcementId") or ""),
                    match_note=match_note,
                    strong_notice_date=clause_dates.at[bond_code, "强赎公告提示日期_函数"],
                    non_redemption_notice_date=clause_dates.at[
                        bond_code, "最新不强赎公告日期_函数"
                    ],
                )
            )

        groups: dict[tuple[str, str, date], list[Announcement]] = defaultdict(list)
        for row in captured:
            groups[(row.category, row.bond_code, row.effective_date)].append(row)

        deduplicated: list[Announcement] = []
        for rows in groups.values():
            if len(rows) > 1 and any("提示" not in row.title for row in rows):
                for row in rows:
                    if "提示" in row.title:
                        excluded.append(
                            {
                                "公告时间": row.announcement_time,
                                "正股代码": row.stock_code,
                                "转债代码": row.bond_code,
                                "公告标题": row.title,
                                "排除原因": "同一转债同日存在非提示公告",
                            }
                        )
                rows = [row for row in rows if "提示" not in row.title]
            # 用户规则只要求移除同日同券的“提示”公告；其余不同公告均保留。
            rows.sort(key=lambda row: (row.announcement_time, row.announcement_id), reverse=True)
            seen_ids = set()
            for row in rows:
                unique_key = row.announcement_id or (row.title, row.url)
                if unique_key in seen_ids:
                    continue
                seen_ids.add(unique_key)
                deduplicated.append(row)

        deduplicated.sort(key=lambda row: row.announcement_time, reverse=True)
        excluded.sort(key=lambda row: row.get("公告时间") or datetime.min, reverse=True)
        return deduplicated, excluded


    CN_RANGE_RE = re.compile(
        r"(?P<sy>20\d{2})\s*年\s*(?P<sm>\d{1,2})\s*月\s*(?P<sd>\d{1,2})\s*日"
        r"\s*(?:至|到|—|－|–|~|～|-)\s*"
        r"(?:(?P<ey>20\d{2})\s*年\s*)?"
        r"(?P<em>\d{1,2})\s*月\s*(?P<ed>\d{1,2})\s*日"
    )
    NUMERIC_RANGE_RE = re.compile(
        r"(?P<sy>20\d{2})\s*[./-]\s*(?P<sm>\d{1,2})\s*[./-]\s*(?P<sd>\d{1,2})"
        r"\s*(?:至|到|—|－|–|~|～)\s*"
        r"(?P<ey>20\d{2})\s*[./-]\s*(?P<em>\d{1,2})\s*[./-]\s*(?P<ed>\d{1,2})"
    )
    MONTH_COUNT_RE = re.compile(
        r"未来\s*(?P<count>\d+|一|二|三|四|五|六|七|八|九|十|十一|十二|两)\s*个?月内"
    )
    CHINESE_NUMBERS = {
        "一": 1,
        "二": 2,
        "两": 2,
        "三": 3,
        "四": 4,
        "五": 5,
        "六": 6,
        "七": 7,
        "八": 8,
        "九": 9,
        "十": 10,
        "十一": 11,
        "十二": 12,
    }


    def add_months(value: date, months: int) -> date:
        month_index = value.year * 12 + value.month - 1 + months
        year, zero_based_month = divmod(month_index, 12)
        month = zero_based_month + 1
        day = min(value.day, calendar.monthrange(year, month)[1])
        return date(year, month, day)


    def range_dates(match: re.Match) -> tuple[date, date]:
        start = date(
            int(match.group("sy")),
            int(match.group("sm")),
            int(match.group("sd")),
        )
        end_year_text = match.group("ey")
        end_year = int(end_year_text) if end_year_text else start.year
        end_month = int(match.group("em"))
        if not end_year_text and end_month < start.month:
            end_year += 1
        end = date(end_year, end_month, int(match.group("ed")))
        return start, end


    def commitment_range_score(text: str, match: re.Match) -> int:
        before = text[max(0, match.start() - 160) : match.start()]
        after = text[match.end() : min(len(text), match.end() + 180)]
        context = before + after
        score = 0
        if re.search(r"未来.{0,16}个月内", before):
            score += 130
        if re.search(r"剩余(?:转股|存续)期内|直至.{0,16}(?:到期|转股期结束)", before):
            score += 130
        if re.search(r"均?不行使.{0,20}赎回|不提前赎回", context):
            score += 80
        if "赎回权" in context or "赎回条款" in context:
            score += 30
        if "重新起算" in after or "重新计算" in after:
            score += 30
        if "触发" in before and not re.search(r"未来|剩余|不行使|不提前赎回", before):
            score -= 70
        if re.search(r"前(?:的)?6个月内|前六个月内", before):
            score -= 90
        return score


    def extract_non_redemption_deadline(
        text: str,
        announcement_date: date,
    ) -> dict:
        compact = re.sub(r"\s+", "", text or "")
        candidates = []
        for pattern in (CN_RANGE_RE, NUMERIC_RANGE_RE):
            for match in pattern.finditer(compact):
                try:
                    start, end = range_dates(match)
                except ValueError:
                    continue
                if end < start or (end - start).days > 3650:
                    continue
                candidates.append(
                    {
                        "deadline": end,
                        "start": start,
                        "score": commitment_range_score(compact, match),
                        "source": "公告明确日期区间",
                        "evidence": compact[
                            max(0, match.start() - 60) : min(len(compact), match.end() + 90)
                        ],
                    }
                )
        if candidates:
            best = max(candidates, key=lambda item: (item["score"], item["deadline"]))
            if best["score"] >= 80:
                return best

        date_token = r"(20\d{2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日"
        # “自10月17日之后的首个交易日重新起算”中的10月17日本身就是截止日。
        reset_after = re.search(
            rf"自({date_token})之后(?:的)?首个交易日重新(?:起算|计算)",
            compact,
        )
        if reset_after:
            deadline = date(
                int(reset_after.group(2)),
                int(reset_after.group(3)),
                int(reset_after.group(4)),
            )
            return {
                "deadline": deadline,
                "start": None,
                "score": 65,
                "source": "公告明确的重新起算前截止日",
                "evidence": reset_after.group(0),
            }

        # “以12月15日为首个交易日重新起算”表示承诺期截止12月14日。
        reset_next_day = re.search(
            rf"(?:以|自)({date_token})(?:[（(][^）)]{{0,30}}[）)])?"
            r"[^。；;]{0,50}为首个交易日重新(?:起算|计算)",
            compact,
        )
        if reset_next_day:
            reset = date(
                int(reset_next_day.group(2)),
                int(reset_next_day.group(3)),
                int(reset_next_day.group(4)),
            )
            return {
                "deadline": reset - timedelta(days=1),
                "start": None,
                "score": 65,
                "source": "首个重新起算日减一天",
                "evidence": reset_next_day.group(0),
            }

        month_match = MONTH_COUNT_RE.search(compact)
        if month_match:
            raw_count = month_match.group("count")
            months = int(raw_count) if raw_count.isdigit() else CHINESE_NUMBERS[raw_count]
            return {
                "deadline": add_months(announcement_date, months) - timedelta(days=1),
                "start": announcement_date,
                "score": 30,
                "source": "按公告对应日起算月数（低置信度）",
                "evidence": month_match.group(0),
            }

        return {
            "deadline": None,
            "start": None,
            "score": 0,
            "source": "未识别，需人工核验",
            "evidence": "",
        }


    def download_pdf(url: str, announcement_id: str, pdf_dir: Path) -> Path:
        pdf_dir.mkdir(parents=True, exist_ok=True)
        output_path = pdf_dir / f"{announcement_id or 'unknown'}.pdf"
        result = subprocess.run(
            [
                "curl.exe",
                "-s",
                "-L",
                "--retry",
                "2",
                "--connect-timeout",
                "20",
                "-H",
                "User-Agent: Mozilla/5.0",
                "-o",
                str(output_path),
                url,
            ],
            capture_output=True,
            text=True,
            timeout=120,
        )
        if (
            result.returncode != 0
            or not output_path.exists()
            or output_path.stat().st_size < 100
            or output_path.read_bytes()[:4] != b"%PDF"
        ):
            raise RuntimeError(f"PDF下载失败：{url}")
        return output_path


    def clean_pdf_page_text(text: str, page_number: int) -> str:
        lines = (text or "").splitlines()
        while lines and not lines[0].strip():
            lines.pop(0)
        while lines and not lines[-1].strip():
            lines.pop()
        page_marker = str(page_number)
        if lines and lines[0].strip() == page_marker:
            lines.pop(0)
        if lines and lines[-1].strip() == page_marker:
            lines.pop()
        return "\n".join(lines)


    def read_pdf_text(pdf_path: Path) -> str:
        return "\n".join(
            clean_pdf_page_text(page.extract_text() or "", index + 1)
            for index, page in enumerate(PdfReader(pdf_path).pages)
        )


    def format_cn_date(value: date | datetime | None) -> str:
        if value is None:
            return ""
        return f"{value.year}年{value.month}月{value.day}日"


    def parse_cn_or_iso_date(value: str) -> date | None:
        text = re.sub(r"\s+", "", value or "")
        match = re.search(r"(20\d{2})年(\d{1,2})月(\d{1,2})日", text)
        if not match:
            match = re.search(r"(20\d{2})[-/](\d{1,2})[-/](\d{1,2})", text)
        if not match:
            return None
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None


    def extract_labeled_date(compact: str, labels: tuple[str, ...]) -> date | None:
        date_pattern = r"(20\d{2}年\d{1,2}月\d{1,2}日|20\d{2}[-/]\d{1,2}[-/]\d{1,2})"
        label_pattern = "|".join(re.escape(label) for label in labels)
        for match in re.finditer(rf"(?:{label_pattern})(?:为|[:：])?[^，。；;]{{0,30}}?({date_pattern})", compact):
            parsed = parse_cn_or_iso_date(match.group(1))
            if parsed:
                return parsed
        return None


    def extract_strong_redemption_fields(text: str) -> dict:
        compact = re.sub(r"\s+", "", text or "")
        fields = {
            "stop_trading_date": extract_labeled_date(compact, ("停止交易日", "最后交易日")),
            "stop_conversion_date": extract_labeled_date(compact, ("停止转股日", "最后转股日")),
            "redemption_registration_date": extract_labeled_date(compact, ("赎回登记日", "赎回股权登记日")),
            "redemption_date": extract_labeled_date(compact, ("赎回日", "赎回款发放日", "赎回资金到账日")),
        }
        fields["redemption_schedule_pending"] = not any(fields.values()) or bool(
            re.search(
                r"(?:赎回安排|具体安排|后续安排).{0,40}(?:尚未|另行|后续|届时).{0,40}(?:公告|披露)|"
                r"(?:尚未|暂未).{0,20}(?:确定|明确).{0,40}(?:赎回安排|具体安排)",
                compact,
            )
        )
        return fields


    def enrich_strong_redemption_fields(
        announcements: list[Announcement],
        pdf_dir: Path,
    ) -> list[str]:
        warnings: list[str] = []
        for row in announcements:
            if row.category != "赎回":
                continue
            try:
                pdf_path = download_pdf(row.url, row.announcement_id, pdf_dir)
                extracted = extract_strong_redemption_fields(read_pdf_text(pdf_path))
                row.stop_trading_date = extracted["stop_trading_date"]
                row.stop_conversion_date = extracted["stop_conversion_date"]
                row.redemption_registration_date = extracted["redemption_registration_date"]
                row.redemption_date = extracted["redemption_date"]
                row.redemption_schedule_pending = extracted["redemption_schedule_pending"]
            except Exception as exc:
                row.redemption_schedule_pending = True
                warnings.append(f"{row.bond_code} {row.bond_name} 强赎公告PDF解析失败：{exc}")
        return warnings


    def enrich_non_redemption_deadlines(
        announcements: list[Announcement],
        pdf_dir: Path,
    ) -> list[str]:
        warnings: list[str] = []
        for row in announcements:
            if row.category != "不赎回":
                continue
            try:
                pdf_path = download_pdf(row.url, row.announcement_id, pdf_dir)
                text = read_pdf_text(pdf_path)
                extracted = extract_non_redemption_deadline(text, row.effective_date)
                row.commitment_start = extracted["start"]
                row.commitment_deadline = extracted["deadline"]
                row.deadline_source = extracted["source"]
                row.deadline_evidence = extracted["evidence"]
                row.deadline_score = extracted["score"]
                if row.commitment_deadline is None or row.deadline_score < 60:
                    warnings.append(
                        f"{row.bond_code} {row.bond_name}承诺截止日低置信度："
                        f"{row.deadline_source}，结果={row.commitment_deadline or '未识别'}"
                    )
            except Exception as exc:
                row.deadline_source = f"PDF解析失败：{exc}"
                warnings.append(f"{row.bond_code} {row.bond_name} PDF解析失败：{exc}")
        return warnings


    def excel_last_row(sheet, column: int = 1) -> int:
        return int(sheet.Cells(sheet.Rows.Count, column).End(XL_UP).Row)


    def excel_cell_value(sheet, row: int, column: int):
        return sheet.Cells(row, column).Value


    def excel_hyperlink_address(cell) -> str:
        try:
            if cell.Hyperlinks.Count:
                return str(cell.Hyperlinks(1).Address or "")
        except Exception:
            pass
        return ""


    def to_excel_serial(value: datetime | date) -> float:
        """使用无时区的Excel 1900日期序列，避免pywin32自动做UTC/本地时区转换。"""
        if isinstance(value, date) and not isinstance(value, datetime):
            value = datetime.combine(value, datetime.min.time())
        return (value - datetime(1899, 12, 30)).total_seconds() / 86400


    def set_excel_formula(cell, formula: str) -> None:
        try:
            cell.Formula2 = formula
        except Exception:
            cell.Formula = formula


    def set_excel_hyperlink(sheet, cell, title: str, url: str) -> None:
        try:
            while cell.Hyperlinks.Count:
                cell.Hyperlinks(1).Delete()
        except Exception:
            pass
        cell.Value = title
        sheet.Hyperlinks.Add(
            Anchor=cell,
            Address=url,
            TextToDisplay=title,
        )


    def copy_row_format_for_insert(sheet, target_row: int, max_column: int) -> None:
        source_row = target_row + 1
        sheet.Range(
            sheet.Cells(source_row, 1),
            sheet.Cells(source_row, max_column),
        ).Copy()
        sheet.Range(
            sheet.Cells(target_row, 1),
            sheet.Cells(target_row, max_column),
        ).PasteSpecial(Paste=XL_PASTE_FORMATS)
        sheet.Application.CutCopyMode = False


    def insert_new_row(
        sheet,
        max_column: int,
        managed_columns: set[int],
    ) -> int:
        sheet.Rows(2).Insert(Shift=XL_SHIFT_DOWN)
        copy_row_format_for_insert(sheet, 2, max_column)
        target = sheet.Range(sheet.Cells(2, 1), sheet.Cells(2, max_column))
        target.ClearContents()
        source_row = 3
        for column in range(1, max_column + 1):
            if column in managed_columns:
                continue
            source_cell = sheet.Cells(source_row, column)
            if source_cell.HasFormula:
                sheet.Cells(2, column).FormulaR1C1 = source_cell.FormulaR1C1
        return 2


    def worksheet_header_map(sheet) -> dict[str, int]:
        last_column = max(1, int(sheet.UsedRange.Columns.Count))
        headers: dict[str, int] = {}
        for column in range(1, last_column + 1):
            header = str(sheet.Cells(1, column).Value or "").strip()
            if not header:
                continue
            if header in headers:
                raise ValueError(f"{sheet.Name}表头存在重复列：{header}")
            headers[header] = column
        return headers


    def validate_target_workbook(workbook) -> dict[str, dict[str, int]]:
        sheet_names = {str(workbook.Worksheets(index).Name) for index in range(1, workbook.Worksheets.Count + 1)}
        required = {"赎回", "不赎回", "统计", "所有退市"}
        missing = required - sheet_names
        if missing:
            raise KeyError(f"底稿缺少工作表：{sorted(missing)}")

        required_headers = {
            "赎回": {"转债代码", "转债简称", "公告时间", "公告标题"},
            "不赎回": {"转债代码", "转债简称", "公告时间", "承诺何日之前不行使", "公告链接"},
        }
        header_maps: dict[str, dict[str, int]] = {}
        for sheet_name, expected in required_headers.items():
            sheet = workbook.Worksheets(sheet_name)
            headers = worksheet_header_map(sheet)
            missing_headers = expected - set(headers)
            if missing_headers:
                raise ValueError(
                    f"{sheet_name}缺少必要表头：{sorted(missing_headers)}；"
                    f"当前表头：{list(headers)}"
                )
            header_maps[sheet_name] = headers
        return header_maps


    def find_existing_strong_row(
        sheet,
        bond_code: str,
        code_column: int,
    ) -> int | None:
        incoming = normalize_bond_code(bond_code)
        for row in range(2, excel_last_row(sheet, code_column) + 1):
            existing = normalize_bond_code(excel_cell_value(sheet, row, code_column))
            if existing and (existing == incoming or normalize_code(existing) == normalize_code(incoming)):
                return row
        return None


    def find_existing_non_redemption_row(
        sheet,
        item: Announcement,
        code_column: int,
        time_column: int,
    ) -> int | None:
        incoming_code = normalize_code(item.bond_code)
        for row in range(2, excel_last_row(sheet, code_column) + 1):
            if normalize_code(excel_cell_value(sheet, row, code_column)) != incoming_code:
                continue
            existing_time = parse_datetime(excel_cell_value(sheet, row, time_column))
            if not existing_time:
                continue
            delta_raw = abs((existing_time.date() - item.announcement_time.date()).days)
            delta_effective = abs((existing_time.date() - item.effective_date).days)
            if min(delta_raw, delta_effective) <= 1:
                return row
        return None


    def backup_workbook(path: Path) -> Path:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = BACKUP_DIR / f"{path.stem}_{timestamp}{path.suffix}"
        shutil.copy2(path, backup_path)
        return backup_path


    def ensure_workbook_available(path: Path) -> None:
        if not path.exists():
            raise FileNotFoundError(f"未找到底稿：{path}")
        lock_file = path.with_name(f"~${path.name}")
        if lock_file.exists():
            raise PermissionError(
                f"检测到底稿正在Excel中打开：{lock_file.name}。"
                "请先关闭底稿后再运行，避免写入冲突。"
            )


    def update_workbook(
        path: Path,
        announcements: list[Announcement],
    ) -> tuple[dict, Path | None]:
        import pythoncom
        import win32com.client

        ensure_workbook_available(path)
        report = {
            "赎回": {"新增": 0, "补全": 0, "跳过": 0, "明细": []},
            "不赎回": {"新增": 0, "补全": 0, "跳过": 0, "明细": []},
        }
        changed = False
        backup_path: Path | None = None
        excel = None
        workbook = None

        pythoncom.CoInitialize()
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            excel.AskToUpdateLinks = False
            workbook = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=False)
            header_maps = validate_target_workbook(workbook)

            strong_sheet = workbook.Worksheets("赎回")
            non_sheet = workbook.Worksheets("不赎回")
            strong_columns = header_maps["赎回"]
            non_columns = header_maps["不赎回"]
            strong_max_column = max(strong_columns.values())
            non_max_column = max(non_columns.values())

            # 按旧到新依次插到第2行，最终保持最新公告在最上方。
            for item in sorted(announcements, key=lambda row: row.announcement_time):
                if item.category == "赎回":
                    code_column = strong_columns["转债代码"]
                    name_column = strong_columns["转债简称"]
                    time_column = strong_columns["公告时间"]
                    link_column = strong_columns["公告标题"]
                    existing_row = find_existing_strong_row(
                        strong_sheet,
                        item.bond_code,
                        code_column,
                    )
                    if existing_row:
                        updated_fields = []
                        name_cell = strong_sheet.Cells(existing_row, name_column)
                        time_cell = strong_sheet.Cells(existing_row, time_column)
                        link_cell = strong_sheet.Cells(existing_row, link_column)
                        if name_cell.Value in (None, ""):
                            code_address = (
                                f"{get_column_letter(code_column)}{existing_row}"
                            )
                            set_excel_formula(
                                name_cell,
                                f"=@cb_info_name({code_address})",
                            )
                            updated_fields.append("转债简称")
                        if time_cell.Value in (None, ""):
                            time_cell.Value2 = to_excel_serial(item.announcement_time)
                            time_cell.NumberFormat = "yyyy/m/d h:mm"
                            updated_fields.append("公告时间")
                        if not excel_hyperlink_address(link_cell):
                            set_excel_hyperlink(
                                strong_sheet,
                                link_cell,
                                item.display_title,
                                item.url,
                            )
                            updated_fields.append("公告链接")
                        if updated_fields:
                            item.summary_action = "补全"
                            changed = True
                            report["赎回"]["补全"] += 1
                            report["赎回"]["明细"].append(
                                {"动作": "补全", "字段": updated_fields, "公告": item}
                            )
                        else:
                            item.summary_action = "已存在"
                            report["赎回"]["跳过"] += 1
                        continue

                    managed_columns = {
                        code_column,
                        name_column,
                        time_column,
                        link_column,
                    }
                    row = insert_new_row(
                        strong_sheet,
                        strong_max_column,
                        managed_columns,
                    )
                    strong_sheet.Cells(row, code_column).Value = item.bond_code
                    code_address = f"{get_column_letter(code_column)}{row}"
                    set_excel_formula(
                        strong_sheet.Cells(row, name_column),
                        f"=@cb_info_name({code_address})",
                    )
                    strong_sheet.Cells(row, time_column).Value2 = to_excel_serial(item.announcement_time)
                    strong_sheet.Cells(row, time_column).NumberFormat = "yyyy/m/d h:mm"
                    set_excel_hyperlink(
                        strong_sheet,
                        strong_sheet.Cells(row, link_column),
                        item.display_title,
                        item.url,
                    )
                    changed = True
                    item.summary_action = "新增"
                    report["赎回"]["新增"] += 1
                    report["赎回"]["明细"].append({"动作": "新增", "公告": item})
                    continue

                code_column = non_columns["转债代码"]
                name_column = non_columns["转债简称"]
                time_column = non_columns["公告时间"]
                deadline_column = non_columns["承诺何日之前不行使"]
                link_column = non_columns["公告链接"]
                existing_row = find_existing_non_redemption_row(
                    non_sheet,
                    item,
                    code_column,
                    time_column,
                )
                deadline_value = item.commitment_deadline or "——"
                if existing_row:
                    updated_fields = []
                    name_cell = non_sheet.Cells(existing_row, name_column)
                    deadline_cell = non_sheet.Cells(existing_row, deadline_column)
                    link_cell = non_sheet.Cells(existing_row, link_column)
                    if name_cell.Value in (None, ""):
                        code_address = (
                            f"{get_column_letter(code_column)}{existing_row}"
                        )
                        set_excel_formula(
                            name_cell,
                            f"=@cb_info_name({code_address})",
                        )
                        updated_fields.append("转债简称")
                    if deadline_cell.Value in (None, "", "——") and item.commitment_deadline:
                        deadline_cell.Value2 = to_excel_serial(item.commitment_deadline)
                        deadline_cell.NumberFormat = "yyyy/m/d"
                        updated_fields.append("承诺截止日")
                    if not excel_hyperlink_address(link_cell):
                        set_excel_hyperlink(
                            non_sheet,
                            link_cell,
                            item.display_title,
                            item.url,
                        )
                        updated_fields.append("公告链接")
                    if updated_fields:
                        item.summary_action = "补全"
                        changed = True
                        report["不赎回"]["补全"] += 1
                        report["不赎回"]["明细"].append(
                            {"动作": "补全", "字段": updated_fields, "公告": item}
                        )
                    else:
                        item.summary_action = "已存在"
                        report["不赎回"]["跳过"] += 1
                    continue

                managed_columns = {
                    code_column,
                    name_column,
                    time_column,
                    deadline_column,
                    link_column,
                }
                row = insert_new_row(
                    non_sheet,
                    non_max_column,
                    managed_columns,
                )
                non_sheet.Cells(row, code_column).Value = item.bond_code
                code_address = f"{get_column_letter(code_column)}{row}"
                set_excel_formula(
                    non_sheet.Cells(row, name_column),
                    f"=@cb_info_name({code_address})",
                )
                non_sheet.Cells(row, time_column).Value2 = to_excel_serial(item.announcement_time)
                non_sheet.Cells(row, time_column).NumberFormat = "yyyy/m/d h:mm"
                if isinstance(deadline_value, date):
                    non_sheet.Cells(row, deadline_column).Value2 = to_excel_serial(deadline_value)
                    non_sheet.Cells(row, deadline_column).NumberFormat = "yyyy/m/d"
                else:
                    non_sheet.Cells(row, deadline_column).Value = deadline_value
                set_excel_hyperlink(
                    non_sheet,
                    non_sheet.Cells(row, link_column),
                    item.display_title,
                    item.url,
                )
                changed = True
                item.summary_action = "新增"
                report["不赎回"]["新增"] += 1
                report["不赎回"]["明细"].append({"动作": "新增", "公告": item})

            if changed:
                backup_path = backup_workbook(path)
                try:
                    excel.CalculateFullRebuild()
                except Exception:
                    pass
                workbook.Save()
                report["写入提示"] = f"底稿已保存：{path}"
                report["备份提示"] = f"更新前备份：{backup_path}"
            else:
                report["写入提示"] = "没有需要写入或补全的记录，底稿未保存。"
        finally:
            if workbook is not None:
                try:
                    workbook.Close(SaveChanges=False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
        return report, backup_path


    def json_default(value):
        if isinstance(value, Announcement):
            return asdict(value)
        if isinstance(value, (datetime, date, pd.Timestamp)):
            return value.isoformat()
        if isinstance(value, Path):
            return str(value)
        return str(value)


    def save_audit(
        run_id: str,
        payload: dict,
    ) -> Path:
        AUDIT_DIR.mkdir(parents=True, exist_ok=True)
        path = AUDIT_DIR / f"转债强赎公告更新_{run_id}.json"
        path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2, default=json_default),
            encoding="utf-8",
        )
        return path


    def build_summary(
        query_range: str,
        active_count: int,
        raw_count: int,
        announcements: list[Announcement],
        warnings: list[str],
        write_report: dict | None,
        dry_run: bool,
        backup_path: Path | None,
    ) -> str:
        display_announcements = [
            row for row in announcements if row.summary_action in ("新增", "补全")
        ]
        strong = [row for row in display_announcements if row.category == "赎回"]
        non = [row for row in display_announcements if row.category == "不赎回"]
        lines = [
            f"查询区间：{query_range}",
            f"规则后保留：{len(display_announcements)}条（赎回{len(strong)}条，不赎回{len(non)}条）",
            "",
            "1、强赎公告：",
        ]
        if strong:
            for index, row in enumerate(strong):
                if index:
                    lines.append("")
                lines.extend(format_strong_summary_block(row))
        else:
            lines.append("无")

        lines.extend(["", "2、不强赎公告："])
        if non:
            for index, row in enumerate(non):
                if index:
                    lines.append("")
                lines.extend(format_non_redemption_summary_block(row))
        else:
            lines.append("无")

        if dry_run:
            lines.extend(["", "当前为dry-run，未写入Excel底稿。"])
        elif write_report:
            lines.extend(["", "写入结果："])
            for category in ("赎回", "不赎回"):
                item = write_report[category]
                lines.append(
                    f"- {category}：新增{item['新增']}，补全{item['补全']}，跳过{item['跳过']}"
                )
            if write_report.get("写入提示"):
                lines.append(write_report["写入提示"])
            if write_report.get("备份提示"):
                lines.append(write_report["备份提示"])

        if warnings:
            lines.extend(["", "需关注："])
            lines.extend(f"- {warning}" for warning in warnings)
        return "\n".join(lines)


    def summary_action_prefix(row: Announcement) -> str:
        return ""


    def announcement_time_prefix(row: Announcement) -> str:
        return f"{format_cn_date(row.announcement_time)} "


    def format_strong_summary_block(row: Announcement) -> list[str]:
        lines = [f"{announcement_time_prefix(row)}{summary_action_prefix(row)}【{row.bond_name}】{row.display_title}"]
        date_fields = [
            ("停止交易日", row.stop_trading_date),
            ("停止转股日", row.stop_conversion_date),
            ("赎回登记日", row.redemption_registration_date),
            ("赎回日", row.redemption_date),
        ]
        for label, value in date_fields:
            if value:
                lines.append(f"{label}：{format_cn_date(value)}")
        if len(lines) == 1 and row.redemption_schedule_pending:
            lines.append("赎回安排尚未公告")
        return lines


    def format_non_redemption_summary_block(row: Announcement) -> list[str]:
        lines = [f"{announcement_time_prefix(row)}{summary_action_prefix(row)}【{row.bond_name}】{row.display_title}"]
        if row.commitment_deadline:
            start = row.commitment_start or row.effective_date
            lines.append(
                f"本次不行使“{row.bond_name}”的提前赎回权利，"
                f"不提前赎回“{row.bond_name}”，"
                f"{format_cn_date(start)}至{format_cn_date(row.commitment_deadline)}期间，"
                "如再次触发有条件赎回条款，公司均不行使提前赎回权利"
            )
        else:
            lines.append("承诺期截止日未识别，需人工核验")
        return lines


    def main(args: argparse.Namespace | None = None) -> dict:
        if args is None:
            args = parse_args()
        if args.days < 1:
            raise ValueError("--days必须大于等于1。")

        run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        start_date = (datetime.now() - timedelta(days=args.days)).strftime("%Y-%m-%d")
        end_date = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
        query_range = f"{start_date}~{end_date}"

        ths_login()
        last_trade_date, cb_basic = fetch_latest_active_cb_universe()

        clause_dates = fetch_excel_clause_dates(cb_basic.index.tolist())
        raw_items = fetch_cninfo_announcements(start_date, end_date)
        announcements, excluded = build_matched_announcements(
            raw_items,
            cb_basic,
            clause_dates,
        )

        temp_pdf_dir = Path(tempfile.mkdtemp(prefix="redeem_notice_pdfs_"))
        warnings = enrich_strong_redemption_fields(announcements, temp_pdf_dir)
        warnings.extend(enrich_non_redemption_deadlines(announcements, temp_pdf_dir))
        write_report = None
        backup_path = None
        try:
            if not args.dry_run:
                write_report, backup_path = update_workbook(
                    args.workbook.resolve(),
                    announcements,
                )
            summary = build_summary(
                query_range=query_range,
                active_count=len(cb_basic),
                raw_count=len(raw_items),
                announcements=announcements,
                warnings=warnings,
                write_report=write_report,
                dry_run=args.dry_run,
                backup_path=backup_path,
            )
            payload = {
                "运行时间": datetime.now(),
                "查询区间": query_range,
                "存续交易列表日期": last_trade_date,
                "存续交易转债数量": len(cb_basic),
                "巨潮原始公告数量": len(raw_items),
                "候选公告": announcements,
                "排除公告": excluded,
                "警告": warnings,
                "写入报告": write_report,
                "备份文件": backup_path,
                "dry_run": args.dry_run,
                "摘要": summary,
            }
            audit_path = save_audit(run_id, payload)
            print(summary)
            print(f"\n审计记录：{audit_path}")
            if not args.no_popup:
                show_popup("转债强赎公告更新结果", summary)
            return payload
        finally:
            if args.keep_pdfs:
                kept_dir = AUDIT_DIR / f"PDF_{run_id}"
                kept_dir.parent.mkdir(parents=True, exist_ok=True)
                shutil.move(str(temp_pdf_dir), str(kept_dir))
                print(f"PDF已保留：{kept_dir}")
            else:
                shutil.rmtree(temp_pdf_dir, ignore_errors=True)

    args = parse_args(["--no-popup"])
    print("\n========== 转债强赎/不赎回公告更新结果 ==========")
    payload = main(args)
    summary = payload.get("\u6458\u8981", "") if isinstance(payload, dict) else ""
    if not summary:
        print("强赎公告更新已运行，但未返回摘要。")
    return summary


def main():
    announcements = fetch_announcements()
    mapping = load_bond_mapping()
    rows = build_classified_rows(announcements, mapping)
    rows = [row for row in rows if row["公告类型"] in ("不下修", "预计下修", "董事会提议下修", "实际下修")]
    enrich_pdf_fields(rows)
    if ENABLE_AUDIT_WORKBOOK:
        write_result_workbook(rows)
    db_report = fill_database(rows)
    summary = build_summary(rows, db_report)
    print("\n========== 可转债转股价修正公告更新结果 ==========")
    print(summary)

    redemption_summary = None
    if ENABLE_REDEMPTION_UPDATE:
        redemption_summary = run_redemption_update()

    if redemption_summary:
        popup_message = (
            "========== 可转债转股价修正公告更新结果 ==========\n"
            f"{summary}\n\n"
            "========== 转债强赎/不赎回公告更新结果 ==========\n"
            f"{redemption_summary}"
        )
        show_popup("可转债公告更新结果", popup_message)
    else:
        show_popup("可转债转股价修正公告分类结果", summary)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        message = f"运行失败：{exc}"
        print(message)
        show_popup("可转债转股价修正公告更新失败", message)
        raise
