# -*- coding: utf-8 -*-
"""
强赎进度跟踪本地化脚本。

数据入口只有：
1. 转债个券历史序列下的标准 Parquet；
2. 【华创固收】赎回和不赎回公告统计.xlsx。

报表排版和长图生成均已内置于本文件，
不依赖其他强赎派生脚本，也不登录、不调用 iFinD 数据接口。

默认运行：先全历史重算并回写 Parquet 中的赎回累计天数，
再生成两份 Excel 和两张长图。
安全验证：py 【条款】P强赎进度跟踪.py --validate
"""

from __future__ import annotations

import argparse
from collections import deque
from dataclasses import dataclass
from datetime import datetime
import json
from pathlib import Path
import re
import sys
import time as _time
from types import SimpleNamespace
from typing import Iterable

import numpy as np
import pandas as pd

import sys

_COMMON_MODULE_DIR = Path(__file__).resolve().parents[1] / "common"
if str(_COMMON_MODULE_DIR) not in sys.path:
    sys.path.insert(0, str(_COMMON_MODULE_DIR))

from 转债Parquet标准读写模块 import replace_monthly_metric_from_wide


# ========== 报表与长图功能（内置） ==========

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont
from tqdm import tqdm



DAYS_TODAY = 0
EXCLUDE_CODES = {"128085.SZ"}

SCRIPT_DIR = Path(__file__).resolve().parents[2]
LOCAL_HEADER_IMAGE = SCRIPT_DIR / "assets/images/条款表头.png"
LOCAL_IMAGE_ASSETS = (
    LOCAL_HEADER_IMAGE,
)
PERMILLE_FOOTER_TEXT = "备注：赎回触发价的单位为0.001"
PERCENT_FOOTER_TEXT = "备注：赎回触发价的单位为0.01"
FOOTER_IMAGE_SIZE = (2989, 43)
INTERMEDIATE_IMAGE_NAMES = (
    "赎回表题头.png",
    "千分位赎回表尾.png",
    "百分位赎回表尾.png",
    "赎回累计触发天数个券.png",
    "当前不强赎承诺期内.png",
    "赎回公告个券.png",
    "赎回累计触发天数个券（百分位）.png",
    "当前不强赎承诺期内（百分位）.png",
    "赎回公告个券（百分位）.png",
)

# 不强赎信息：保留现有手工配置，用于补充本地公告 Excel。
MANUAL_NON_REDEMPTION_OVERRIDES = {
    "123158.SZ": {"不强赎公告日": "2026-08-12", "承诺何日之前不行使": "2027-02-12"},
    "111012.SH": {"不强赎公告日": "2026-07-10", "承诺何日之前不行使": "2026-10-10"},
}

# 强赎信息：仅在本地公告 Excel 缺失时补充，不覆盖有效的公告数据。
MANUAL_COMPULSORY_REDEMPTION_OVERRIDES = {
    # "113000.SH": {
    #     "赎回公告日": "2026-01-01",
    #     "赎回登记日": "2026-01-20",
    #     "最后交易日": "2026-01-15",
    # },
}

REDEMPTION_INFO_COLUMNS = [
    "转债简称", "债项评级", "时间区间", "计算天数", "赎回登记日", "最后交易日", "转债余额",
    "未转股比例", "对流通股本稀释", "转债价格", "大股东持债比例", "平价", "转股溢价率", "纯债溢价率", "所属行业",
]
REDEMPTION_OUTPUT_COLUMNS = REDEMPTION_INFO_COLUMNS.copy()
REDEMPTION_OUTPUT_COLUMNS.insert(6, "赎回公告日")

NON_REDEMPTION_INFO_COLUMNS = [
    "转债简称", "债项评级", "时间区间", "计算天数", "不赎回公告日", "承诺何日之前不行使", "赎回累计触发天数", "转债余额",
    "未转股比例", "对流通股本稀释", "转债价格", "大股东持债比例", "平价", "转股溢价率", "纯债溢价率", "最早触发日期",
]

LAST_TRADE_COLUMNS = [
    "转债简称", "到期日期", "最后交易日", "最后转股日", "摘牌日期", "剩余天数", "转债余额", "未转股比例", "到期赎回价",
    "转债价格", "平价", "转股溢价率", "纯债价值", "纯债溢价率", "YTM", "所属行业",
]

TABLE_COLUMN_WIDTHS = [0.06, 0.06, 0.06, 0.06, 0.06, 0.1, 0.1, 0.1, 0.06, 0.06, 0.08, 0.06, 0.08, 0.06, 0.06, 0.06, 0.06]


def make_paths():
    mmdd_today = _time.strftime("%m%d", _time.localtime())
    yyyymmdd_today = _time.strftime("%Y%m%d", _time.localtime())
    folder = Path(f"{mmdd_today}数据更新") / "赎回数据更新"
    folder.mkdir(parents=True, exist_ok=True)
    return {
        "mmdd": mmdd_today,
        "yyyymmdd": yyyymmdd_today,
        "folder": folder,
        "permille_xlsx": folder / f"【华创固收】转债赎回信息日度跟踪-{yyyymmdd_today}【千分位版】.xlsx",
        "percent_xlsx": folder / f"【华创固收】转债赎回信息日度跟踪-{yyyymmdd_today}【百分位版】.xlsx",
        "permille_png": folder / f"【华创固收】转债赎回信息日度跟踪-{yyyymmdd_today}【千分位版】.png",
        "percent_png": folder / f"【华创固收】转债赎回信息日度跟踪-{yyyymmdd_today}【百分位版】.png",
    }


def write_version_tables(
    path,
    redemption_bond_info,
    counting,
    commitment,
    other,
    recent_parity,
    lasttrade_info,
):
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        redemption_bond_info.to_excel(writer, sheet_name="公告赎回转债信息", index=True)
        counting.to_excel(writer, sheet_name="赎回累计触发天数", index=True)
        commitment.to_excel(writer, sheet_name="当前不强赎承诺期内", index=True)
        other.to_excel(writer, sheet_name="非承诺期且无累计天数", index=True)
        recent_parity.to_excel(writer, sheet_name="近30日平价情况", index=False)
        lasttrade_info.to_excel(writer, sheet_name="到期将摘牌转债", index=True)


def contains_chinese(value):
    return any("\u4e00" <= char <= "\u9fff" for char in str(value))


def set_font(cell_value):
    if contains_chinese(cell_value):
        return Font(name="KaiTi_GB2312", size=10, bold=False)
    return Font(name="Times New Roman", size=10, bold=False)


def format_workbook(path, recent_parity: pd.DataFrame | None = None):
    workbook = load_workbook(path)
    recent_style_metadata = (
        recent_parity.attrs if recent_parity is not None else {}
    )
    recent_date_columns = set(recent_style_metadata.get("history_date_columns", []))
    recent_hit_dates = recent_style_metadata.get("hit_dates_by_code", {})
    excel_format_dict = {
        # 与本地化下修表的分组配色保持一致；摘牌表继续沿用原有格式。
        "公告赎回转债信息": ["963634", "F2DCDB"],
        "赎回累计触发天数": ["963634", "E4DFEC"],
        "当前不强赎承诺期内": ["963634", "DCE6F1"],
        "非承诺期且无累计天数": ["963634", "E4DFEC"],
        "近30日平价情况": ["963634", "E4DFEC"],
        "到期将摘牌转债": ["963634", "DCE6F1"],
    }
    excel_basicdata_format_dict = {
        "正股收盘价": ["963634", "0070C0"],
        "赎回触发价": ["963634", "DDD9C4"],
        "总表": ["963634", "DCE6F1"],
        "前十大转债持有人": ["963634", "DCE6F1"],
    }

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    alignment = Alignment(horizontal="center", vertical="center")

    for sheet_name, colors in excel_format_dict.items():
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "E2" if sheet_name == "近30日平价情况" else "C2"
        font_title = Font(name="KaiTi_GB2312", size=10, bold=False, color="FFFFFF")
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = font_title if cell.row == 1 else set_font(cell.value)
                # 下修表的“近30日平价情况”正文不铺底色，保留白底便于查看逐日数据。
                body_color = None if sheet_name == "近30日平价情况" else colors[1]
                cell.fill = (
                    PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")
                    if cell.row == 1
                    else (
                        PatternFill(start_color=body_color, end_color=body_color, fill_type="solid")
                        if body_color
                        else PatternFill(fill_type=None)
                    )
                )
                cell.border = thin_border
                cell.alignment = alignment
                if cell.row >= 1:
                    cell.parent.row_dimensions[cell.row].height = 15
                if sheet_name == "近30日平价情况" and cell.row > 1:
                    if cell.column == 4:
                        cell.number_format = "0"
                        cell.font = Font(
                            name="Times New Roman",
                            size=10,
                            bold=True,
                            color="C00000",
                        )
                    elif cell.column >= 5:
                        cell.number_format = "0.00"
        if sheet_name != "近30日平价情况":
            header_by_col = {
                cell.column: str(cell.value).strip()
                for cell in worksheet[1]
                if cell.value is not None
            }
            cumulative_col = next(
                (
                    column
                    for column, header in header_by_col.items()
                    if header in {"赎回累计触发天数", "目前天数累计"}
                ),
                None,
            )
            if cumulative_col is not None:
                for row_no in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row_no, cumulative_col)
                    cell.number_format = "0"
                    cell.font = Font(
                        name="Times New Roman",
                        size=10,
                        bold=True,
                        color="C00000",
                    )
        else:
            header_by_col = {
                cell.column: str(cell.value).strip()
                for cell in worksheet[1]
                if cell.value is not None
            }
            code_col = next(
                (
                    column
                    for column, header in header_by_col.items()
                    if header == "转债代码"
                ),
                None,
            )
            if code_col is not None and recent_date_columns:
                hit_fill = PatternFill(
                    start_color="FFEB9C",
                    end_color="FFEB9C",
                    fill_type="solid",
                )
                for row_no in range(2, worksheet.max_row + 1):
                    code = str(worksheet.cell(row_no, code_col).value or "").strip()
                    code_hit_dates = recent_hit_dates.get(code, set())
                    for column, date_label in header_by_col.items():
                        if (
                            date_label in recent_date_columns
                            and date_label in code_hit_dates
                        ):
                            worksheet.cell(row_no, column).fill = hit_fill
        auto_width(worksheet, multiplier=1.5)

    date_time_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}\s00:00:00$")
    for sheet_name in excel_basicdata_format_dict:
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "C2"
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, datetime):
                    cell.value = cell.value.strftime("%Y-%m-%d")
                elif cell.value and date_time_pattern.match(str(cell.value)):
                    cell.value = str(cell.value)[:10]
                cell.font = Font(name="Times New Roman", size=10, bold=False)
        auto_width(worksheet)

        if sheet_name != "总表":
            for stock_row, redeem_row in zip(workbook["正股收盘价"].iter_rows(min_row=2), workbook["赎回触发价"].iter_rows(min_row=2)):
                for stock_cell, redeem_cell in zip(stock_row[1:], redeem_row[1:]):
                    if stock_cell.value and redeem_cell.value and stock_cell.value > redeem_cell.value:
                        fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        stock_cell.fill = fill
                        redeem_cell.fill = fill

    workbook.save(path)


def auto_width(worksheet, multiplier=1):
    for col in worksheet.columns:
        max_length = 0
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        worksheet.column_dimensions[col[0].column_letter].width = (max_length + 2) * multiplier


def plot_table_image(df, output_path, body_color, highlight_commitment=False, last_date=None):
    plt.rcParams["font.sans-serif"] = ["KaiTi_GB2312"]
    fig, ax = plt.subplots(figsize=(8, 0.01))
    ax.axis("tight")
    ax.axis("off")
    table_df = df.reset_index()
    table = ax.table(cellText=table_df.values.tolist(), colLabels=table_df.columns.tolist(), loc="center")
    table.auto_set_font_size(False)
    table.set_fontsize(5)

    if not (len(df.columns) and df.columns[0] == "OUTMESSAGE"):
        for i, width in enumerate(TABLE_COLUMN_WIDTHS):
            if i >= len(table_df.columns):
                break
            for j in range(len(table_df) + 1):
                table._cells[(j, i)].set_width(width)

    today = tomorrow = None
    if highlight_commitment and last_date is not None:
        today = _local_date_offset("212001", "dateType:0,period:D,offset:0,dateFormat:0,output:singledate", f"{last_date}").data
        tomorrow = _local_date_offset("212001", "dateType:0,period:D,offset:1,dateFormat:0,output:singledate", f"{last_date}").data

    for cell_key, cell in table._cells.items():
        row, col = cell_key
        text = cell.get_text()
        text.set_ha("center")
        if row == 0:
            cell.set_facecolor("#963634")
            text.set_color("white")
        else:
            cell.set_facecolor(body_color)
        cell.set_linewidth(0.1)

    if highlight_commitment and today is not None:
        max_row = max(cell[0] for cell in table._cells.keys())
        max_col = max(cell[1] for cell in table._cells.keys())
        for row in range(1, max_row + 1):
            try:
                cell_date = datetime.strptime(str(table._cells[(row, 5)].get_text().get_text()), "%Y-%m-%d")
            except ValueError:
                continue
            if cell_date in {today, tomorrow}:
                for col in range(max_col + 1):
                    table._cells[(row, col)].set_facecolor("#F2DCDB")

    fig.set_figheight(0.008)
    fig.savefig(output_path, dpi=300, bbox_inches="tight", pad_inches=0, transparent=False)
    plt.close(fig)


def validate_local_image_assets():
    missing = [path.name for path in LOCAL_IMAGE_ASSETS if not path.is_file()]
    if missing:
        raise FileNotFoundError(
            "缺少赎回长图所需的本地图片："
            f"{', '.join(missing)}。请将图片放在脚本同目录。"
        )

    invalid = []
    for path in LOCAL_IMAGE_ASSETS:
        try:
            with Image.open(path) as image:
                image.verify()
        except Exception as exc:
            invalid.append(f"{path.name}（{exc}）")
    if invalid:
        raise RuntimeError(f"本地图片无法读取：{'；'.join(invalid)}")


def create_header_from_local_asset(folder, trade_date):
    with Image.open(LOCAL_HEADER_IMAGE) as source_image:
        img = source_image.copy()

    draw = ImageDraw.Draw(img)
    text = f"      华创固收·周冠南团队\n可转债赎回信息整理（{trade_date}）"
    font = ImageFont.truetype(str(SCRIPT_DIR / "assets/fonts/KaiTi_GB2312.ttf"), 60)
    text_bbox = draw.textbbox((0, 0), text, font=font)
    text_width = text_bbox[2] - text_bbox[0]
    text_height = text_bbox[3] - text_bbox[1]
    x = (img.width - text_width) // 2
    y = (img.height - text_height) // 2 - 40
    draw.text((x, y), text, fill="white", font=font)
    img.save(folder / "赎回表题头.png")


def create_footer(folder, version):
    if version == "permille":
        name = "千分位赎回表尾.png"
        footer_text = PERMILLE_FOOTER_TEXT
    elif version == "percent":
        name = "百分位赎回表尾.png"
        footer_text = PERCENT_FOOTER_TEXT
    else:
        raise ValueError(f"未知的表底版本：{version}")

    image = Image.new("RGBA", FOOTER_IMAGE_SIZE, (255, 255, 255, 255))
    draw = ImageDraw.Draw(image)
    draw.rectangle(
        (0, 0, image.width - 1, image.height - 1),
        outline="black",
        width=2,
    )
    font = ImageFont.truetype(str(SCRIPT_DIR / "assets/fonts/KaiTi_GB2312.ttf"), 26)
    text_bbox = draw.textbbox((0, 0), footer_text, font=font)
    text_height = text_bbox[3] - text_bbox[1]
    y = (image.height - text_height) // 2 - text_bbox[1]
    draw.text(
        (7, y),
        footer_text,
        fill="black",
        font=font,
        stroke_width=1,
        stroke_fill="black",
    )
    image.save(folder / name)


def combine_version_images(folder, version, redemption_bond_info, output_path):
    suffix = "" if version == "permille" else "（百分位）"
    footer_name = "千分位赎回表尾.png" if version == "permille" else "百分位赎回表尾.png"
    image_paths = [
        folder / f"赎回累计触发天数个券{suffix}.png",
        folder / f"当前不强赎承诺期内{suffix}.png",
        folder / f"赎回公告个券{suffix}.png",
        folder / "赎回表题头.png",
        folder / footer_name,
    ]
    images = [Image.open(path) for path in image_paths]
    max_width = max(image.width for image in images[:4])
    resized = []
    heights = []
    for image in images:
        height = round(max_width / image.width * image.height)
        heights.append(height)
        resized.append(image.resize((max_width, height), resample=Image.BILINEAR))

    if redemption_bond_info.columns[0] == "OUTMESSAGE":
        order = [3, 1, 0, 4]
    else:
        order = [3, 2, 1, 0, 4]

    total_height = sum(heights[i] for i in order)
    new_image = Image.new("RGBA", (max_width, total_height), (0, 0, 0, 0))
    y = 0
    for idx in order:
        new_image.paste(resized[idx], (0, y))
        y += heights[idx]
    new_image.save(output_path)

    for image in images:
        image.close()


def cleanup_intermediate_images(folder):
    folder = Path(folder).resolve()
    removed = []
    failures = []
    for name in INTERMEDIATE_IMAGE_NAMES:
        image_path = (folder / name).resolve()
        if image_path.parent != folder:
            raise RuntimeError(f"拒绝清理输出文件夹之外的图片：{image_path}")
        if not image_path.is_file():
            continue
        try:
            image_path.unlink()
            removed.append(name)
        except OSError as exc:
            failures.append(f"{name}（{exc}）")

    if failures:
        raise RuntimeError(
            "完整长图已生成，但以下过程图片清理失败："
            f"{'；'.join(failures)}"
        )
    return removed


def run_version(version, output_path, image_output_path, cb_basic_trade, cb_list_trade, stock_close_sheet, trigger_price_sheet, total_table, stock_holder_hold_cb_bond, last_date, folder):
    round_trigger = version == "percent"
    suffix = "" if version == "permille" else "（百分位）"

    redemption_count = calculate_redemption_count(stock_close_sheet, trigger_price_sheet, total_table, round_trigger=round_trigger)
    redemption_bond_info = fetch_redemption_bond_info(stock_close_sheet, total_table, last_date, stock_holder_hold_cb_bond)
    counting, commitment, other = fetch_non_redemption_info(
        cb_basic_trade,
        redemption_bond_info,
        redemption_count,
        total_table,
        last_date,
        stock_holder_hold_cb_bond,
    )
    recent_parity = build_recent_parity_sheet(
        stock_close_sheet,
        trigger_price_sheet,
        counting,
        commitment,
        other,
        round_trigger=round_trigger,
    )
    lasttrade_info = fetch_lasttrade_info(cb_list_trade, redemption_bond_info, last_date)

    write_version_tables(
        output_path,
        redemption_bond_info,
        counting,
        commitment,
        other,
        recent_parity,
        lasttrade_info,
    )
    format_workbook(output_path, recent_parity=recent_parity)

    plot_table_image(redemption_bond_info, folder / f"赎回公告个券{suffix}.png", "#0070C0")
    plot_table_image(counting, folder / f"赎回累计触发天数个券{suffix}.png", "#DDD9C4")
    plot_table_image(commitment, folder / f"当前不强赎承诺期内{suffix}.png", "#DCE6F1", highlight_commitment=True, last_date=last_date)
    create_footer(folder, version)
    combine_version_images(folder, version, redemption_bond_info, image_output_path)


def print_runtime(start_time):
    total_time = _time.time() - start_time
    if total_time > 60:
        minutes = int(total_time // 60)
        seconds = int(total_time % 60)
        print(f"程序总运行时长：{int(minutes)} 分 {int(seconds)} 秒")
    else:
        print(f"程序总运行时长：{int(total_time)} 秒")
    print(datetime.now().strftime("%H:%M:%S"), "运行完成")
    print("\U0001F600", "\U0001F600", "\U0001F600")

# 兼容既有调用写法：本文件自身即为报表功能命名空间。
ORIGINAL = sys.modules[__name__]

SCRIPT_DIR = Path(__file__).resolve().parents[2]
PARQUET_ROOT = SCRIPT_DIR / "data/转债个券历史序列"
ANNOUNCEMENT_FILE = SCRIPT_DIR / "data/clauses/【华创固收】赎回和不赎回公告统计.xlsx"
MASTER_FILE = PARQUET_ROOT / "_special" / "总表.parquet"

VALID_STOCK_TRADING_STATUSES = {"交易", "新股上市", "正常上市"}
VALID_BOND_TRADING_STATUSES = {"交易", "新股上市", "正常上市"}
HISTORY_MIN_TRADE_DATES = 70

MASTER_REQUIRED = {
    "转债代码",
    "转债名称",
    "最后交易日",
    "最后转股日",
    "摘牌日期",
    "到期日期",
    "到期赎回价",
    "申万行业",
    "转股期起始日",
    "赎回触发比例",
    "赎回触发计算时间区间",
    "赎回触发计算最大时间区间",
}
PANEL_REQUIRED = {
    "转债代码",
    "交易日期",
    "余额",
    "收盘价",
    "平价",
    "转股价",
    "转股溢价率",
    "纯债价值",
    "纯债溢价率",
    "YTM",
    "债项评级",
    "正股收盘价",
    "正股交易状态",
    "累计转股比例",
    "转股稀释率",
    "交易状态",
}

HISTORY_PANEL_COLUMNS = [
    "转债代码",
    "交易日期",
    "正股收盘价",
    "正股交易状态",
    "转股价",
]
HISTORY_MASTER_REQUIRED = {
    "转债代码",
    "发行日期",
    "上市日期",
    "最后交易日",
    "转股期起始日",
    "赎回触发比例",
    "赎回触发计算时间区间",
    "赎回触发计算最大时间区间",
}


# 本地化版不再列示大股东持债数据：同时从 Excel 表和长图中删除该列。
ORIGINAL.REDEMPTION_OUTPUT_COLUMNS = [
    col for col in ORIGINAL.REDEMPTION_OUTPUT_COLUMNS if col != "大股东持债比例"
]
ORIGINAL.NON_REDEMPTION_INFO_COLUMNS = [
    col for col in ORIGINAL.NON_REDEMPTION_INFO_COLUMNS if col != "大股东持债比例"
]
# TABLE_COLUMN_WIDTHS 含索引列，原第 13 列对应“大股东持债比例”。
ORIGINAL.TABLE_COLUMN_WIDTHS = [
    width for pos, width in enumerate(ORIGINAL.TABLE_COLUMN_WIDTHS) if pos != 12
]


def _normalise_code(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip().upper()
    match = re.search(r"\d{6}(?:\.(?:SH|SZ|BJ|NQ))?", text)
    return match.group(0) if match else text


def _parse_date(value: object) -> pd.Timestamp:
    if value is None or pd.isna(value):
        return pd.NaT
    if isinstance(value, pd.Timestamp):
        return value.normalize()
    if isinstance(value, (int, float, np.integer, np.floating)):
        number = float(value)
        if number == 0:
            return pd.NaT
        if 20_000 <= number <= 80_000:
            return (pd.Timestamp("1899-12-30") + pd.to_timedelta(number, unit="D")).normalize()
    parsed = pd.to_datetime(value, errors="coerce")
    return pd.Timestamp(parsed).normalize() if pd.notna(parsed) else pd.NaT


def _format_date(value: object) -> str:
    parsed = _parse_date(value)
    return parsed.strftime("%Y-%m-%d") if pd.notna(parsed) else ""


def _ensure_columns(frame: pd.DataFrame, required: set[str], label: str) -> None:
    missing = sorted(required.difference(frame.columns))
    if missing:
        raise ValueError(f"{label}缺少必需字段：{missing}")


def _read_recent_panel() -> tuple[pd.DataFrame, list[Path]]:
    files = sorted(PARQUET_ROOT.glob("[0-9][0-9][0-9][0-9]/*.parquet"), reverse=True)
    if not files:
        raise FileNotFoundError(f"未在 {PARQUET_ROOT} 下找到月度 Parquet")

    frames: list[pd.DataFrame] = []
    used_files: list[Path] = []
    unique_dates: set[pd.Timestamp] = set()
    for path in files:
        frame = pd.read_parquet(path)
        _ensure_columns(frame, PANEL_REQUIRED, f"月度 Parquet {path.name}")
        frame = frame.copy()
        frame["转债代码"] = frame["转债代码"].map(_normalise_code)
        frame["交易日期"] = pd.to_datetime(frame["交易日期"], errors="coerce").dt.normalize()
        frame = frame.dropna(subset=["交易日期"])
        frames.append(frame)
        used_files.append(path)
        unique_dates.update(pd.Timestamp(x) for x in frame["交易日期"].unique())
        if len(unique_dates) >= HISTORY_MIN_TRADE_DATES:
            break

    panel = pd.concat(frames, ignore_index=True)
    panel = panel.sort_values(["转债代码", "交易日期"])
    panel = panel.drop_duplicates(["转债代码", "交易日期"], keep="last")
    return panel, list(reversed(used_files))


def _read_announcements() -> tuple[pd.DataFrame, pd.DataFrame]:
    if not ANNOUNCEMENT_FILE.is_file():
        raise FileNotFoundError(f"未找到公告文件：{ANNOUNCEMENT_FILE}")
    redeem = pd.read_excel(ANNOUNCEMENT_FILE, sheet_name="赎回")
    non_redeem = pd.read_excel(ANNOUNCEMENT_FILE, sheet_name="不赎回")
    _ensure_columns(redeem, {"转债代码", "公告时间"}, "公告 Excel/赎回")
    _ensure_columns(
        non_redeem,
        {"转债代码", "公告时间", "承诺何日之前不行使"},
        "公告 Excel/不赎回",
    )
    for frame in (redeem, non_redeem):
        frame["转债代码"] = frame["转债代码"].map(_normalise_code)
        frame["公告时间"] = frame["公告时间"].map(_parse_date)
        frame.dropna(subset=["转债代码", "公告时间"], inplace=True)
    non_redeem["承诺何日之前不行使"] = non_redeem["承诺何日之前不行使"].map(_parse_date)
    for col in ["最后交易日", "最后转股日", "赎回登记日", "摘牌日期"]:
        if col in redeem.columns:
            redeem[col] = redeem[col].map(_parse_date)
    return redeem, non_redeem


def _latest_event_by_code(frame: pd.DataFrame, as_of: pd.Timestamp) -> pd.DataFrame:
    eligible = frame[frame["公告时间"] <= as_of].copy()
    if eligible.empty:
        return eligible.set_index("转债代码")
    eligible = eligible.sort_values(["转债代码", "公告时间"])
    return eligible.groupby("转债代码", as_index=False).tail(1).set_index("转债代码")


def _base_code(code: object) -> str:
    match = re.search(r"\d{6}", _normalise_code(code))
    return match.group(0) if match else _normalise_code(code)


def _event_map(frame: pd.DataFrame, *, include_commitment: bool) -> dict[str, list[object]]:
    result: dict[str, list[object]] = {}
    for row in frame.itertuples(index=False):
        code = _normalise_code(getattr(row, "转债代码"))
        publish = _parse_date(getattr(row, "公告时间"))
        if not code or pd.isna(publish):
            continue
        if include_commitment:
            commitment = _parse_date(getattr(row, "承诺何日之前不行使"))
            value: object = (publish, commitment)
        else:
            value = publish
        for key in {code, _base_code(code)}:
            result.setdefault(key, []).append(value)
    return result


def _events_for_code(event_map: dict[str, list[object]], code: str) -> list[object]:
    values = [*event_map.get(code, []), *event_map.get(_base_code(code), [])]
    unique: list[object] = []
    seen: set[object] = set()
    for value in values:
        key = tuple(value) if isinstance(value, tuple) else value
        if key not in seen:
            seen.add(key)
            unique.append(value)
    return sorted(unique, key=lambda value: value[0] if isinstance(value, tuple) else value)


def _read_historical_trigger_panel() -> tuple[pd.DataFrame, list[Path]]:
    files = sorted(
        path
        for year_dir in PARQUET_ROOT.iterdir()
        if year_dir.is_dir() and year_dir.name.isdigit()
        for path in year_dir.glob("*.parquet")
    )
    if not files:
        raise FileNotFoundError(f"未在 {PARQUET_ROOT} 下找到月度 Parquet")

    frames: list[pd.DataFrame] = []
    for path in files:
        frame = pd.read_parquet(path, columns=HISTORY_PANEL_COLUMNS)
        frame["转债代码"] = frame["转债代码"].map(_normalise_code)
        frame["交易日期"] = pd.to_datetime(frame["交易日期"], errors="coerce").dt.normalize()
        frames.append(frame.dropna(subset=["转债代码", "交易日期"]))
    panel = pd.concat(frames, ignore_index=True)
    panel = panel.sort_values(["转债代码", "交易日期"], kind="stable")
    panel = panel.drop_duplicates(["转债代码", "交易日期"], keep="last")
    return panel, files


def _calculate_one_redemption_history(
    dates: pd.DatetimeIndex,
    observations: pd.DataFrame,
    metadata: pd.Series,
    redeem_events: list[object],
    non_redeem_events: list[object],
) -> pd.Series:
    """按正股收盘价和正股交易状态计算单券全历史赎回累计天数。"""
    result = pd.Series(np.nan, index=dates, dtype="float64")
    issue_date = _parse_date(metadata.get("发行日期"))
    listing_date = _parse_date(metadata.get("上市日期"))
    last_trade_date = _parse_date(metadata.get("最后交易日"))
    conversion_start = _parse_date(metadata.get("转股期起始日"))
    trigger_ratio = pd.to_numeric(metadata.get("赎回触发比例"), errors="coerce")
    trigger_days = pd.to_numeric(metadata.get("赎回触发计算时间区间"), errors="coerce")
    window_days = pd.to_numeric(metadata.get("赎回触发计算最大时间区间"), errors="coerce")
    if pd.isna(conversion_start) or pd.isna(trigger_ratio) or pd.isna(trigger_days) or pd.isna(window_days):
        return result

    trigger_days = max(int(trigger_days), 1)
    window_days = max(int(window_days), 1)
    obs = observations.set_index("交易日期").reindex(dates)
    close = pd.to_numeric(obs["正股收盘价"], errors="coerce")
    conversion_price = pd.to_numeric(obs["转股价"], errors="coerce")
    trigger_price = (conversion_price * float(trigger_ratio) / 100.0).round(3)
    stock_status = obs["正股交易状态"].astype("string").str.strip()

    redeem_dates = sorted({_parse_date(value) for value in redeem_events if pd.notna(_parse_date(value))})
    first_redeem = redeem_dates[0] if redeem_dates else pd.NaT
    non_redeem_intervals: list[tuple[pd.Timestamp, pd.Timestamp]] = []
    non_redeem_publish_days: set[pd.Timestamp] = set()
    for event in non_redeem_events:
        publish, commitment = event
        publish = _parse_date(publish)
        commitment = _parse_date(commitment)
        if pd.isna(publish):
            continue
        end = commitment if pd.notna(commitment) and commitment >= publish else publish
        non_redeem_intervals.append((publish, end))
        non_redeem_publish_days.add(publish)

    rolling_hits: deque[int] = deque()
    rolling_sum = 0
    trigger_reached_previous_day = False
    for date in dates:
        if pd.notna(issue_date) and date < issue_date:
            continue
        if pd.notna(listing_date) and date < listing_date:
            continue
        if date < conversion_start:
            continue
        if pd.notna(last_trade_date) and date > last_trade_date:
            continue
        if pd.notna(first_redeem) and date > first_redeem:
            continue

        in_commitment = any(start < date <= end for start, end in non_redeem_intervals)
        if in_commitment:
            rolling_hits.clear()
            rolling_sum = 0
            trigger_reached_previous_day = False
            result.at[date] = 0.0
            continue

        is_valid_trade = stock_status.at[date] in VALID_STOCK_TRADING_STATUSES
        if not is_valid_trade:
            result.at[date] = float(rolling_sum)
            continue

        hit_today = int(
            pd.notna(close.at[date])
            and float(close.at[date]) > 0
            and pd.notna(trigger_price.at[date])
            and float(close.at[date]) >= float(trigger_price.at[date])
        )
        rolling_hits.append(hit_today)
        rolling_sum += hit_today
        if len(rolling_hits) > window_days:
            rolling_sum -= rolling_hits.popleft()

        result.at[date] = float(rolling_sum)
        triggered_today = rolling_sum >= trigger_days
        # 达到条款即完成一轮计数：强赎后续为空，不强赎或默认不行使后下一交易日重新起算。
        should_reset = date in non_redeem_publish_days or (
            triggered_today
            and not trigger_reached_previous_day
            and (pd.isna(first_redeem) or date != first_redeem)
        )
        if should_reset:
            rolling_hits.clear()
            rolling_sum = 0
        trigger_reached_previous_day = triggered_today

    return result


def calculate_historical_redemption_days() -> tuple[pd.DataFrame, dict[str, object]]:
    """全历史计算入口；不读取旧的「赎回累计天数」结果列。"""
    master = pd.read_parquet(MASTER_FILE).copy()
    _ensure_columns(master, HISTORY_MASTER_REQUIRED, "总表 Parquet")
    master["转债代码"] = master["转债代码"].map(_normalise_code)
    master = master.drop_duplicates("转债代码", keep="last").set_index("转债代码")

    panel, files = _read_historical_trigger_panel()
    dates = pd.DatetimeIndex(panel["交易日期"].drop_duplicates().sort_values())
    redeem, non_redeem = _read_announcements()
    redeem_map = _event_map(redeem, include_commitment=False)
    non_redeem_map = _event_map(non_redeem, include_commitment=True)
    groups = {code: frame for code, frame in panel.groupby("转债代码", sort=False)}

    result = pd.DataFrame(np.nan, index=master.index, columns=dates, dtype="float64")
    calculated = 0
    with tqdm(
        total=len(master.index),
        desc="重算赎回累计天数",
        unit="只",
        dynamic_ncols=True,
        mininterval=0.2,
    ) as progress:
        for code in master.index:
            observations = groups.get(code)
            if observations is None:
                observations = groups.get(_base_code(code))
            if observations is None:
                progress.update(1)
                continue
            result.loc[code] = _calculate_one_redemption_history(
                dates,
                observations,
                master.loc[code],
                _events_for_code(redeem_map, code),
                _events_for_code(non_redeem_map, code),
            ).to_numpy()
            calculated += 1
            progress.update(1)

    # 延续现有底稿口径：0 表示当日尚未触发，落盘时按空值存储，正数才是有效触发结果。
    stored_result = result.mask(result.eq(0))
    stats = {
        "parquet_files": len(files),
        "trade_dates": len(dates),
        "master_bonds": len(master),
        "calculated_bonds": calculated,
        "positive_observations": int(stored_result.notna().sum().sum()),
        "last_date": dates[-1].strftime("%Y-%m-%d") if len(dates) else None,
    }
    return stored_result, stats


def refresh_historical_redemption_days() -> dict[str, object]:
    """由本地化副本独立重算并回写 Parquet 的赎回累计天数。"""
    print("[赎回累计天数] 正在基于本地 Parquet 和公告 Excel 全历史重算…")
    wide, stats = calculate_historical_redemption_days()
    with tqdm(
        desc="写回赎回累计天数",
        unit="月",
        dynamic_ncols=True,
        mininterval=0.2,
    ) as progress:
        observations = replace_monthly_metric_from_wide(
            PARQUET_ROOT,
            "赎回累计天数",
            wide,
            progress=progress,
        )
    stats["written_observations"] = observations
    return stats


@dataclass
class LocalContext:
    master: pd.DataFrame
    panel: pd.DataFrame
    latest: pd.DataFrame
    active_codes: pd.Index
    last_date: pd.Timestamp
    as_of: pd.Timestamp
    redeem_events: pd.DataFrame
    non_redeem_events: pd.DataFrame
    latest_redeem: pd.DataFrame
    latest_non_redeem: pd.DataFrame
    parquet_files: list[Path]
    cb_basic_trade: pd.DataFrame
    stock_close_sheet: pd.DataFrame
    trigger_price_sheet: pd.DataFrame
    trade_status_sheet: pd.DataFrame
    total_table: pd.DataFrame
    holder_info: pd.DataFrame
    stock_holder: pd.DataFrame


CTX: LocalContext | None = None


def _pivot(panel: pd.DataFrame, metric: str, codes: Iterable[str]) -> pd.DataFrame:
    wide = panel.pivot(index="转债代码", columns="交易日期", values=metric)
    wide = wide.reindex(index=pd.Index(codes), columns=sorted(wide.columns))
    wide.index.name = "代码"
    return wide


def load_local_context() -> LocalContext:
    global CTX
    if not MASTER_FILE.is_file():
        raise FileNotFoundError(f"未找到总表 Parquet：{MASTER_FILE}")

    master = pd.read_parquet(MASTER_FILE).copy()
    _ensure_columns(master, MASTER_REQUIRED, "总表 Parquet")
    master["转债代码"] = master["转债代码"].map(_normalise_code)
    master = master.drop_duplicates("转债代码", keep="last").set_index("转债代码")
    for col in ["最后交易日", "最后转股日", "摘牌日期", "到期日期", "转股期起始日"]:
        master[col] = master[col].map(_parse_date)

    panel, parquet_files = _read_recent_panel()
    last_date = pd.Timestamp(panel["交易日期"].max()).normalize()
    latest = panel[panel["交易日期"] == last_date].copy().set_index("转债代码")
    active_mask = latest["交易状态"].astype("string").str.strip().isin(VALID_BOND_TRADING_STATUSES)
    active_codes = latest.index[active_mask]
    active_codes = active_codes.intersection(master.index)
    active_codes = active_codes.difference(pd.Index(getattr(ORIGINAL, "EXCLUDE_CODES", [])))
    active_codes = pd.Index(sorted(active_codes), name="代码")

    redeem_events, non_redeem_events = _read_announcements()
    as_of = pd.Timestamp.now().normalize()
    latest_redeem = _latest_event_by_code(redeem_events, as_of)
    latest_non_redeem = _latest_event_by_code(non_redeem_events, as_of)

    cb_basic_trade = pd.DataFrame(index=active_codes)
    cb_basic_trade["转债简称"] = master.reindex(active_codes)["转债名称"]
    cb_basic_trade["正股代码"] = ""
    cb_basic_trade["正股简称"] = ""
    cb_basic_trade.index.name = "转债代码"

    active_panel = panel[panel["转债代码"].isin(active_codes)].copy()
    ratio_map = pd.to_numeric(master["赎回触发比例"], errors="coerce") / 100.0
    active_panel["本地赎回触发价"] = (
        pd.to_numeric(active_panel["转股价"], errors="coerce")
        * active_panel["转债代码"].map(ratio_map)
    ).round(3)
    stock_close = _pivot(active_panel, "正股收盘价", active_codes)
    trigger_price = _pivot(active_panel, "本地赎回触发价", active_codes)
    stock_status_raw = _pivot(active_panel, "正股交易状态", active_codes)
    trade_status = stock_status_raw.apply(
        lambda column: column.map(
            lambda value: "正常上市"
            if str(value).strip() in VALID_STOCK_TRADING_STATUSES
            else ("数据缺失" if pd.isna(value) else "停牌")
        )
    )

    names = cb_basic_trade["转债简称"]
    for frame in (stock_close, trigger_price, trade_status):
        frame.insert(0, "名称", names.reindex(frame.index))

    total = pd.DataFrame(index=active_codes)
    total.index.name = "代码"
    total["名称"] = names.reindex(active_codes)
    total["时间区间MAX"] = pd.to_numeric(
        master.reindex(active_codes)["赎回触发计算最大时间区间"], errors="coerce"
    )
    total["计算天数"] = pd.to_numeric(
        master.reindex(active_codes)["赎回触发计算时间区间"], errors="coerce"
    )
    total["不强赎公告日"] = pd.NaT
    total["承诺何日之前不行使"] = pd.NaT
    common_non = active_codes.intersection(latest_non_redeem.index)
    if len(common_non):
        total.loc[common_non, "不强赎公告日"] = latest_non_redeem.loc[common_non, "公告时间"]
        total.loc[common_non, "承诺何日之前不行使"] = latest_non_redeem.loc[
            common_non, "承诺何日之前不行使"
        ]
    total["赎回起始日"] = master.reindex(active_codes)["转股期起始日"]
    total["赎回公告日"] = pd.NaT
    common_redeem = active_codes.intersection(latest_redeem.index)
    if len(common_redeem):
        total.loc[common_redeem, "赎回公告日"] = latest_redeem.loc[common_redeem, "公告时间"]

    holder_info = pd.DataFrame(index=active_codes)
    holder_info.index.name = "代码"
    stock_holder = pd.DataFrame(index=active_codes)
    stock_holder["转债简称"] = names.reindex(active_codes)
    stock_holder["持股第一名大股东"] = np.nan
    stock_holder["持股第一名大股东持债比例"] = np.nan

    CTX = LocalContext(
        master=master,
        panel=active_panel,
        latest=latest.reindex(active_codes),
        active_codes=active_codes,
        last_date=last_date,
        as_of=as_of,
        redeem_events=redeem_events,
        non_redeem_events=non_redeem_events,
        latest_redeem=latest_redeem,
        latest_non_redeem=latest_non_redeem,
        parquet_files=parquet_files,
        cb_basic_trade=cb_basic_trade,
        stock_close_sheet=stock_close,
        trigger_price_sheet=trigger_price,
        trade_status_sheet=trade_status,
        total_table=total,
        holder_info=holder_info,
        stock_holder=stock_holder,
    )
    return CTX


def clean_suspended_and_commitment(
    stock_close_sheet: pd.DataFrame,
    trigger_price_sheet: pd.DataFrame,
    trade_status_sheet: pd.DataFrame,
    total_table: pd.DataFrame,
):
    stock_close = stock_close_sheet.drop(columns=["名称"], errors="ignore").copy()
    trigger_price = trigger_price_sheet.drop(columns=["名称"], errors="ignore").copy()
    stock_status = trade_status_sheet.drop(columns=["名称"], errors="ignore").copy()
    total = total_table.drop(columns=["名称"], errors="ignore").copy()

    invalid_stock_day = stock_status.ne("正常上市")
    stock_close = stock_close.mask(invalid_stock_day)
    trigger_price = trigger_price.mask(invalid_stock_day)
    stock_status = stock_status.mask(invalid_stock_day)

    date_columns = pd.DatetimeIndex(pd.to_datetime(stock_close.columns))
    for code in stock_close.index:
        redemption_notice = _parse_date(total.at[code, "赎回公告日"])
        redemption_start = _parse_date(total.at[code, "赎回起始日"])
        non_redeem_notice = _parse_date(total.at[code, "不强赎公告日"])
        commitment_end = _parse_date(total.at[code, "承诺何日之前不行使"])

        if pd.notna(redemption_notice):
            stock_close.loc[code, :] = np.nan
            trigger_price.loc[code, :] = np.nan
            continue
        if pd.notna(redemption_start):
            before_start = date_columns < redemption_start
            stock_close.loc[code, before_start] = np.nan
            trigger_price.loc[code, before_start] = np.nan
        reset_boundary = commitment_end if pd.notna(commitment_end) else non_redeem_notice
        if pd.notna(reset_boundary):
            before_reset = date_columns <= reset_boundary
            stock_close.loc[code, before_reset] = np.nan
            trigger_price.loc[code, before_reset] = np.nan

    return stock_close, trigger_price, stock_status, total


def calculate_redemption_count(
    stock_close_sheet: pd.DataFrame,
    trigger_price_sheet: pd.DataFrame,
    total_table: pd.DataFrame,
    round_trigger: bool = False,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for code in stock_close_sheet.index:
        window = int(pd.to_numeric(total_table.at[code, "时间区间MAX"], errors="coerce") or 0)
        close = pd.to_numeric(stock_close_sheet.loc[code], errors="coerce")
        trigger = pd.to_numeric(trigger_price_sheet.loc[code], errors="coerce")
        trigger = trigger.round(2 if round_trigger else 3)
        observations = pd.DataFrame({"close": close, "trigger": trigger}).dropna()
        observations = observations[observations["close"] > 0].tail(max(window, 0))
        hit = observations["close"].ge(observations["trigger"])
        rows.append(
            {
                "代码": code,
                "赎回累计触发天数": int(hit.sum()),
                "最早触发日期": hit.index[hit][0] if hit.any() else np.nan,
                "有效观察天数": int(len(observations)),
            }
        )
    return pd.DataFrame(rows).set_index("代码")


def _snapshot_output(codes: Iterable[str]) -> pd.DataFrame:
    if CTX is None:
        raise RuntimeError("本地数据上下文未初始化")
    codes = pd.Index(codes)
    latest = CTX.latest.reindex(codes)
    master = CTX.master.reindex(codes)
    cumulative = pd.to_numeric(latest["累计转股比例"], errors="coerce")
    result = pd.DataFrame(index=codes)
    result["转债简称"] = master["转债名称"]
    result["债项评级"] = latest["债项评级"]
    result["时间区间"] = pd.to_numeric(master["赎回触发计算最大时间区间"], errors="coerce")
    result["计算天数"] = pd.to_numeric(master["赎回触发计算时间区间"], errors="coerce")
    result["转债余额"] = pd.to_numeric(latest["余额"], errors="coerce")
    result["未转股比例"] = 100.0 - cumulative
    result["对流通股本稀释"] = pd.to_numeric(latest["转股稀释率"], errors="coerce")
    result["转债价格"] = pd.to_numeric(latest["收盘价"], errors="coerce")
    result["平价"] = pd.to_numeric(latest["平价"], errors="coerce")
    result["转股溢价率"] = pd.to_numeric(latest["转股溢价率"], errors="coerce")
    result["纯债溢价率"] = pd.to_numeric(latest["纯债溢价率"], errors="coerce")
    result["所属行业"] = master["申万行业"]
    return result


def fetch_redemption_bond_info(
    stock_close_sheet: pd.DataFrame,
    total_table: pd.DataFrame,
    last_date: str,
    stock_holder_hold_cb_bond: pd.DataFrame,
) -> pd.DataFrame:
    del last_date, stock_holder_hold_cb_bond
    if CTX is None:
        raise RuntimeError("本地数据上下文未初始化")
    codes = total_table.index[total_table["赎回公告日"].notna()]
    if len(codes) == 0:
        empty = pd.DataFrame(columns=ORIGINAL.REDEMPTION_OUTPUT_COLUMNS)
        empty.index.name = "代码"
        return empty

    result = _snapshot_output(codes)
    details = CTX.latest_redeem.reindex(codes)
    result.insert(4, "赎回登记日", details.get("赎回登记日", pd.Series(index=codes, dtype="object")).map(_format_date))
    # 强赎后的最后交易日必须来自赎回公告；若公告尚未给出，就留空。
    # 不能回退到总表中的到期最后交易日，否则会生成误导性日期。
    last_trade = details.get("最后交易日", pd.Series(index=codes, dtype="object")).copy()
    result.insert(5, "最后交易日", last_trade.map(_format_date))
    result.insert(6, "赎回公告日", details["公告时间"].map(_format_date))
    result = result.reindex(columns=ORIGINAL.REDEMPTION_OUTPUT_COLUMNS)
    result.index.name = "代码"
    result = result.round(
        {
            "时间区间": 0,
            "计算天数": 0,
            "转债余额": 2,
            "未转股比例": 2,
            "对流通股本稀释": 2,
            "转债价格": 2,
            "平价": 2,
            "转股溢价率": 2,
            "纯债溢价率": 2,
        }
    ).sort_values("赎回公告日", ascending=False)
    return result.fillna("——")


def fetch_non_redemption_info(
    cb_basic_trade: pd.DataFrame,
    redemption_bond_info: pd.DataFrame,
    redemption_count: pd.DataFrame,
    total_table: pd.DataFrame,
    last_date: str,
    stock_holder_hold_cb_bond: pd.DataFrame,
):
    del stock_holder_hold_cb_bond
    codes = cb_basic_trade.index.difference(redemption_bond_info.index)
    result = _snapshot_output(codes)
    result.insert(4, "不赎回公告日", total_table.reindex(codes)["不强赎公告日"].map(_format_date))
    result.insert(
        5,
        "承诺何日之前不行使",
        total_table.reindex(codes)["承诺何日之前不行使"].map(_format_date),
    )
    result.insert(
        6,
        "赎回累计触发天数",
        pd.to_numeric(redemption_count.reindex(codes)["赎回累计触发天数"], errors="coerce").fillna(0).astype(int),
    )
    earliest = redemption_count.reindex(codes)["最早触发日期"].map(_format_date)
    result["最早触发日期"] = earliest
    result = result.reindex(columns=ORIGINAL.NON_REDEMPTION_INFO_COLUMNS)
    result.index.name = "代码"
    result = result.round(
        {
            "时间区间": 0,
            "计算天数": 0,
            "转债余额": 2,
            "未转股比例": 2,
            "对流通股本稀释": 2,
            "转债价格": 2,
            "平价": 2,
            "转股溢价率": 2,
            "纯债溢价率": 2,
        }
    )
    counting = result[result["赎回累计触发天数"] > 0].sort_values("赎回累计触发天数", ascending=False)
    commitment_end = pd.to_datetime(
        result["承诺何日之前不行使"], errors="coerce"
    )
    as_of = _parse_date(last_date)
    current_commitment_mask = (
        result["不赎回公告日"].ne("")
        & result["赎回累计触发天数"].eq(0)
        & commitment_end.notna()
        & commitment_end.ge(as_of)
    )
    commitment = result[current_commitment_mask].sort_values(
        ["不赎回公告日", "承诺何日之前不行使", "转债简称"],
        ascending=[False, True, True],
        kind="mergesort",
    )
    other = result[
        result["赎回累计触发天数"].eq(0)
        & ~result.index.isin(commitment.index)
    ].copy()
    other["__承诺截止日排序"] = pd.to_datetime(
        other["承诺何日之前不行使"], errors="coerce"
    )
    other = other.sort_values(
        ["__承诺截止日排序", "转债简称"],
        ascending=[True, True],
        na_position="last",
    ).drop(columns="__承诺截止日排序")
    return counting.fillna(""), commitment.fillna(""), other.fillna("")


def build_recent_parity_sheet(
    stock_close_sheet: pd.DataFrame,
    trigger_price_sheet: pd.DataFrame,
    counting: pd.DataFrame,
    commitment: pd.DataFrame,
    other: pd.DataFrame,
    *,
    round_trigger: bool = False,
) -> pd.DataFrame:
    """按下修进度表的布局构造近 30 日平价 sheet。

    每行包含分组、转债代码、转债名称、当前累计天数和近 30 个交易日的平价。
    黄色标注仍使用正股收盘价与赎回触发价比较，不改变强赎判断口径。
    强赎公告券不再放入该跟踪区域，与下修表不展示已终止跟踪标的口径一致。
    """
    if CTX is None:
        raise RuntimeError("本地数据上下文未初始化")
    raw_dates = pd.to_datetime(pd.Index(stock_close_sheet.columns), errors="coerce")
    valid_dates = pd.DatetimeIndex(raw_dates[raw_dates.notna()]).sort_values().unique()
    recent_dates = valid_dates[-30:]
    date_labels = [date.strftime("%Y-%m-%d") for date in recent_dates]
    close = stock_close_sheet.copy()
    close.columns = raw_dates
    close = close.loc[:, ~close.columns.isna()]
    close = close.loc[:, ~close.columns.duplicated(keep="last")]
    parity = _pivot(CTX.panel, "平价", close.index)
    parity = parity.reindex(index=close.index, columns=close.columns)
    # 沿用正股交易状态、转股期起始日和不赎回承诺期的屏蔽范围。
    parity = parity.where(close.notna())
    trigger = trigger_price_sheet.copy()
    trigger.columns = pd.to_datetime(pd.Index(trigger.columns), errors="coerce")
    trigger = trigger.loc[:, ~trigger.columns.isna()]
    trigger = trigger.loc[:, ~trigger.columns.duplicated(keep="last")]
    trigger = trigger.apply(pd.to_numeric, errors="coerce").round(
        2 if round_trigger else 3
    )

    groups = [
        ("赎回累计触发天数", counting),
        ("当前不强赎承诺期内", commitment),
        ("非承诺期且无累计天数", other),
    ]
    rows: list[dict[str, object]] = []
    hit_dates_by_code: dict[str, set[str]] = {}
    for group_name, frame in groups:
        if frame is None or frame.empty:
            continue
        for code, source_row in frame.iterrows():
            code_text = str(code).strip()
            row: dict[str, object] = {
                "分组": group_name,
                "转债代码": code_text,
                "转债名称": source_row.get("转债简称", ""),
                "目前天数累计": pd.to_numeric(
                    source_row.get("赎回累计触发天数", 0), errors="coerce"
                ),
            }
            if code_text in close.index:
                close_row = pd.to_numeric(close.loc[code_text], errors="coerce")
            else:
                close_row = pd.Series(dtype="float64")
            if code_text in parity.index:
                parity_row = pd.to_numeric(parity.loc[code_text], errors="coerce")
            else:
                parity_row = pd.Series(dtype="float64")
            if code_text in trigger.index:
                trigger_row = pd.to_numeric(trigger.loc[code_text], errors="coerce")
            else:
                trigger_row = pd.Series(dtype="float64")
            hit_dates: set[str] = set()
            for date, label in zip(recent_dates, date_labels):
                close_value = close_row.get(date, np.nan)
                parity_value = parity_row.get(date, np.nan)
                trigger_value = trigger_row.get(date, np.nan)
                row[label] = parity_value
                if (
                    pd.notna(parity_value)
                    and
                    pd.notna(close_value)
                    and pd.notna(trigger_value)
                    and float(close_value) >= float(trigger_value)
                ):
                    hit_dates.add(label)
            hit_dates_by_code[code_text] = hit_dates
            rows.append(row)

    columns = ["分组", "转债代码", "转债名称", "目前天数累计", *date_labels]
    result = pd.DataFrame(rows, columns=columns)
    if not result.empty:
        group_order = {name: pos for pos, (name, _) in enumerate(groups)}
        result["__group_order"] = result["分组"].map(group_order)
        result = result.sort_values(
            ["__group_order", "目前天数累计", "转债名称", "转债代码"],
            ascending=[True, False, True, True],
            na_position="last",
            kind="mergesort",
        ).drop(columns="__group_order")
    result.attrs["history_date_columns"] = date_labels
    result.attrs["hit_dates_by_code"] = hit_dates_by_code
    return result


def fetch_lasttrade_info(cb_list_trade: str, redemption_bond_info: pd.DataFrame, last_date: str) -> pd.DataFrame:
    del cb_list_trade
    if CTX is None:
        raise RuntimeError("本地数据上下文未初始化")
    codes = CTX.active_codes.difference(redemption_bond_info.index)
    latest = CTX.latest.reindex(codes)
    master = CTX.master.reindex(codes)
    as_of = _parse_date(last_date)
    maturity = pd.to_datetime(master["到期日期"], errors="coerce").dt.normalize()
    last_trade = pd.to_datetime(master["最后交易日"], errors="coerce").dt.normalize()
    result = pd.DataFrame(index=codes)
    result["转债简称"] = master["转债名称"]
    result["到期日期"] = maturity
    result["最后交易日"] = last_trade
    result["最后转股日"] = pd.to_datetime(master["最后转股日"], errors="coerce").dt.normalize()
    result["摘牌日期"] = pd.to_datetime(master["摘牌日期"], errors="coerce").dt.normalize()
    result["剩余天数"] = (maturity - as_of).dt.days
    result["转债余额"] = pd.to_numeric(latest["余额"], errors="coerce")
    result["未转股比例"] = 100.0 - pd.to_numeric(latest["累计转股比例"], errors="coerce")
    result["到期赎回价"] = pd.to_numeric(master["到期赎回价"], errors="coerce")
    result["转债价格"] = pd.to_numeric(latest["收盘价"], errors="coerce")
    result["平价"] = pd.to_numeric(latest["平价"], errors="coerce")
    result["转股溢价率"] = pd.to_numeric(latest["转股溢价率"], errors="coerce")
    result["纯债价值"] = pd.to_numeric(latest["纯债价值"], errors="coerce")
    result["纯债溢价率"] = pd.to_numeric(latest["纯债溢价率"], errors="coerce")
    result["YTM"] = pd.to_numeric(latest["YTM"], errors="coerce")
    result["所属行业"] = master["申万行业"]
    result = result.reindex(columns=ORIGINAL.LAST_TRADE_COLUMNS)
    result.index.name = "代码"
    one_month_later = as_of + pd.DateOffset(months=1)
    result = result[
        result["最后交易日"].notna()
        & result["到期日期"].ge(as_of)
        & result["到期日期"].le(one_month_later)
    ].sort_values("最后交易日")
    for date_column in ("到期日期", "最后交易日", "最后转股日", "摘牌日期"):
        result[date_column] = result[date_column].map(_format_date)
    return result.round(
        {
            "转债余额": 2,
            "未转股比例": 2,
            "转债价格": 2,
            "平价": 2,
            "转股溢价率": 2,
            "纯债价值": 2,
            "纯债溢价率": 2,
            "YTM": 2,
        }
    )


def _parquet_trading_dates() -> pd.DatetimeIndex:
    """返回本轮已读取月度Parquet中的实际交易日期。"""
    if CTX is None:
        raise RuntimeError("本地数据上下文未初始化，无法读取Parquet交易日期")
    dates = pd.DatetimeIndex(
        pd.to_datetime(CTX.panel["交易日期"], errors="coerce").dropna().dt.normalize().unique()
    ).sort_values()
    if dates.empty:
        raise RuntimeError("月度Parquet中没有可用的交易日期")
    return dates


def _local_date_offset(_calendar: str, params: str, date_value: str, *_args, **_kwargs):
    """严格按本轮月度Parquet的实际交易日期进行偏移。"""
    offset_match = re.search(r"offset:([+-]?\d+)", params)
    offset = int(offset_match.group(1)) if offset_match else 0
    start = _parse_date(date_value)
    if pd.isna(start):
        return SimpleNamespace(data=None)

    dates = _parquet_trading_dates()
    # offset:0 定位到不晚于输入日期的最近一个 Parquet 交易日。
    base_position = int(dates.searchsorted(start, side="right")) - 1
    if base_position < 0:
        return SimpleNamespace(data=None)

    target_position = base_position + offset
    if "output:sequencedate" in params.lower():
        if target_position < 0 or target_position >= len(dates):
            return SimpleNamespace(data={"time": []})
        left, right = sorted((base_position, target_position))
        values = [value.to_pydatetime() for value in dates[left : right + 1]]
        return SimpleNamespace(data={"time": values})

    if target_position < 0 or target_position >= len(dates):
        # Parquet没有未来日期时不使用工作日规则猜测。
        return SimpleNamespace(data=None)
    return SimpleNamespace(data=dates[target_position].to_pydatetime())


def install_local_hooks() -> None:
    ORIGINAL.calculate_redemption_count = calculate_redemption_count
    ORIGINAL.fetch_redemption_bond_info = fetch_redemption_bond_info
    ORIGINAL.fetch_non_redemption_info = fetch_non_redemption_info
    ORIGINAL.fetch_lasttrade_info = fetch_lasttrade_info


def write_local_base_workbook(
    path: Path,
    stock_close_sheet: pd.DataFrame,
    trigger_price_sheet: pd.DataFrame,
    trade_status_sheet: pd.DataFrame,
    total_table: pd.DataFrame,
) -> None:
    """本地化底表不再创建空的“前十大转债持有人”sheet。"""
    with pd.ExcelWriter(path, mode="w") as writer:
        stock_close_sheet.to_excel(writer, sheet_name="正股收盘价", index=True)
        trigger_price_sheet.to_excel(writer, sheet_name="赎回触发价", index=True)
        trade_status_sheet.to_excel(writer, sheet_name="交易状态", index=True)
        total_table.to_excel(writer, sheet_name="总表", index=True)
        # 原排版函数依赖该 sheet；排版完成后会从最终文件中删除。
        pd.DataFrame(index=stock_close_sheet.index).to_excel(
            writer, sheet_name="前十大转债持有人", index=True
        )


def write_local_percent_base_workbook(
    path: Path,
    stock_close_sheet: pd.DataFrame,
    trigger_price_sheet: pd.DataFrame,
    total_table: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(path, mode="w") as writer:
        stock_close_sheet.to_excel(writer, sheet_name="正股收盘价", index=True)
        trigger_price_sheet.to_excel(writer, sheet_name="赎回触发价", index=True)
        total_table.to_excel(writer, sheet_name="总表", index=True)
        pd.DataFrame(index=stock_close_sheet.index).to_excel(
            writer, sheet_name="前十大转债持有人", index=True
        )


def remove_internal_workbook_sheets(path: Path) -> None:
    """删除只为运行排版使用的内部底表，只保留最终展示结果。

    这些 sheet 在生成过程中仍会临时创建，以供原有排版函数读取；不会出现在交付 Excel 中。
    """
    workbook = ORIGINAL.load_workbook(path)
    internal_sheet_names = {
        "正股收盘价",
        "前十大转债持有人",
        "赎回触发价",
        "交易状态",
        "总表",
    }
    for sheet_name in list(workbook.sheetnames):
        if sheet_name in internal_sheet_names:
            workbook.remove(workbook[sheet_name])
    workbook.save(path)


def _audit_payload(
    context: LocalContext,
    permille_count: pd.DataFrame,
    percent_count: pd.DataFrame,
) -> dict[str, object]:
    latest = context.latest
    required_latest = ["余额", "收盘价", "平价", "转股价", "正股收盘价", "正股交易状态"]
    return {
        "data_sources": {
            "master": str(MASTER_FILE),
            "monthly_parquet": [str(path) for path in context.parquet_files],
            "announcements": str(ANNOUNCEMENT_FILE),
            "external_api_calls": 0,
        },
        "redemption_count_source": "local_runtime_from_raw_parquet_and_announcement_excel",
        "stored_redemption_count_is_input": False,
        "last_parquet_date": context.last_date.strftime("%Y-%m-%d"),
        "announcement_as_of": context.as_of.strftime("%Y-%m-%d"),
        "active_bonds": int(len(context.active_codes)),
        "active_strong_redemption": int(context.total_table["赎回公告日"].notna().sum()),
        "active_non_redemption_history": int(context.total_table["不强赎公告日"].notna().sum()),
        "latest_missing": {col: int(latest[col].isna().sum()) for col in required_latest},
        "allowed_nullable": {
            "累计转股比例": int(latest["累计转股比例"].isna().sum())
        },
        "trigger_count_positive": {
            "permille": int(permille_count["赎回累计触发天数"].gt(0).sum()),
            "percent": int(percent_count["赎回累计触发天数"].gt(0).sum()),
        },
        "stock_status_basis": sorted(VALID_STOCK_TRADING_STATUSES),
        "window_rule": "只使用正股交易状态有效的最近N个观察日，停牌日不占用窗口",
        "known_local_gaps": [],
    }


def validate_only() -> dict[str, object]:
    context = load_local_context()
    stock_close, trigger_price, stock_status, total = clean_suspended_and_commitment(
        context.stock_close_sheet,
        context.trigger_price_sheet,
        context.trade_status_sheet,
        context.total_table,
    )
    del stock_status
    permille_count = calculate_redemption_count(stock_close, trigger_price, total, round_trigger=False)
    percent_count = calculate_redemption_count(stock_close, trigger_price, total, round_trigger=True)
    payload = _audit_payload(context, permille_count, percent_count)
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return payload


def run_report() -> dict[str, Path]:
    start_time = _time.time()
    ORIGINAL.validate_local_image_assets()
    refresh_historical_redemption_days()
    context = load_local_context()
    install_local_hooks()
    paths = ORIGINAL.make_paths()
    last_date_text = context.last_date.strftime("%Y-%m-%d")
    cb_list_trade = ",".join(context.active_codes)

    stock_close, trigger_price, stock_status, total = clean_suspended_and_commitment(
        context.stock_close_sheet,
        context.trigger_price_sheet,
        context.trade_status_sheet,
        context.total_table,
    )
    write_local_base_workbook(
        paths["permille_xlsx"], stock_close, trigger_price, stock_status, total
    )
    write_local_percent_base_workbook(
        paths["percent_xlsx"], stock_close, trigger_price, total
    )
    ORIGINAL.create_header_from_local_asset(paths["folder"], last_date_text)

    for version, xlsx_key, png_key in (
        ("permille", "permille_xlsx", "permille_png"),
        ("percent", "percent_xlsx", "percent_png"),
    ):
        ORIGINAL.run_version(
            version,
            paths[xlsx_key],
            paths[png_key],
            context.cb_basic_trade,
            cb_list_trade,
            stock_close,
            trigger_price,
            total,
            context.stock_holder,
            last_date_text,
            paths["folder"],
        )
        remove_internal_workbook_sheets(paths[xlsx_key])
    ORIGINAL.cleanup_intermediate_images(paths["folder"])

    ORIGINAL.print_runtime(start_time)
    return {key: Path(value) for key, value in paths.items() if key.endswith("xlsx") or key.endswith("png")}


def main() -> None:
    parser = argparse.ArgumentParser(description="【条款】P强赎进度跟踪")
    parser.add_argument("--validate", action="store_true", help="只验证本地数据和触发计算，不生成文件")
    args = parser.parse_args()
    if args.validate:
        validate_only()
        return
    run_report()


if __name__ == "__main__":
    main()

