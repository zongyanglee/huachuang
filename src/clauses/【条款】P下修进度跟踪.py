# -*- coding: utf-8 -*-
"""仅使用本地 Parquet 与公告数据库生成可转债下修进度跟踪表。

本文件是 ``下修进度跟踪.py`` 的独立本地化副本，不登录或调用 iFinD。
正股近 1 日/20 日均价来自 Parquet；每股净资产和股票发行面值缺失时
不参与理论下修底价计算。
"""

from __future__ import annotations

import argparse
import re
import shutil
import time as _time
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any, Iterable

import matplotlib
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont

matplotlib.use("Agg")

import matplotlib.pyplot as plt

import sys

_COMMON_MODULE_DIR = Path(__file__).resolve().parents[1] / "common"
if str(_COMMON_MODULE_DIR) not in sys.path:
    sys.path.insert(0, str(_COMMON_MODULE_DIR))

from 转债Parquet标准读写模块 import (
    BOND_CODE,
    read_metric_wide,
    replace_monthly_metric_from_wide,
)


ROOT = Path(__file__).resolve().parents[2]
OUTPUT_DATE = datetime.now().strftime("%Y%m%d")
OUTPUT_MMDD = datetime.now().strftime("%m%d")
DEFAULT_OUTPUT_DIR = ROOT / "runs" / "daily" / f"{OUTPUT_MMDD}数据更新" / "下修数据更新"
DEFAULT_OUTPUT = DEFAULT_OUTPUT_DIR / f"【华创固收】可转债转股价修正信息跟踪数据库-{OUTPUT_DATE}.xlsx"
DEFAULT_IMAGE_OUTPUT = DEFAULT_OUTPUT.with_suffix(".png")
DEFAULT_DATABASE = ROOT / "data/clauses/【华创固收】下修和不下修公告统计.xlsx"
DEFAULT_PARQUET_ROOT = ROOT / "data/转债个券历史序列"
DEFAULT_HEADER_IMAGE = ROOT / "assets/images/条款表头.png"
DOWNWARD_DAYS_SHEET = "下修累计天数"
PARITY_SHEET = "平价"
RECENT_PARITY_SHEET = "近30日平价情况"
PARQUET_META_COLS = frozenset({"__sheet_name", "__row_id", "__date"})
RECENT_PARITY_DAYS = 30
EXCLUDE_BONDS = {"128085.SZ"}
LOCAL_AS_OF_DATE: pd.Timestamp | None = None
INTEGER_OUTPUT_COLUMNS = {"时间区间", "计算天数", "目前天数累计", "距离回售天数"}
GROUP_SHEET_BODY_COLORS = {
    "实施下修中": "F2DCDB",
    "当前具有累计天数": "E4DFEC",
    "当前不下修承诺期内": "DCE6F1",
    "非承诺期且无累计天数": "E4DFEC",
    RECENT_PARITY_SHEET: "E4DFEC",
}
GROUP_SHEET_HEADER_COLOR = "963634"
COUNTED_DAY_COLOR = "FFEB9C"
DOWNWARD_ANNOUNCEMENT_DAY_COLOR = "963634"
NON_DOWNWARD_ANNOUNCEMENT_DAY_COLOR = "0070C0"
CUMULATIVE_DAYS_FONT_COLOR = "C00000"
IMAGE_GROUP_ORDER = ["实施下修中", "当前具有累计天数"]
IMAGE_TITLE_TEMPLATE = "      华创固收·周冠南团队\n可转债下修信息整理（{date}）"
IMAGE_TABLE_DPI = 300
IMAGE_TABLE_WIDTH_INCH = 9.85
EXCEL_ROW_HEIGHT_POINTS = 15
IMAGE_ROW_HEIGHT_POINTS = 10


def _load_local_parquet_snapshot(
    parquet_root: Path,
    requested_date: str | None = None,
) -> tuple[pd.DataFrame, pd.Timestamp]:
    """读取不晚于指定日期的最新月度快照；未指定时使用 Parquet 最新日期。"""
    monthly_paths = sorted(
        path
        for year_dir in parquet_root.iterdir()
        if year_dir.is_dir() and year_dir.name.isdigit()
        for path in year_dir.glob("*.parquet")
    )
    if not monthly_paths:
        raise FileNotFoundError(f"未找到月度 Parquet: {parquet_root}")

    requested = pd.Timestamp(requested_date).normalize() if requested_date else None
    selected_path: Path | None = None
    selected_date = pd.NaT
    for path in reversed(monthly_paths):
        dates = pd.to_datetime(
            pd.read_parquet(path, columns=["交易日期"])["交易日期"],
            errors="coerce",
        ).dropna()
        if dates.empty:
            continue
        eligible = dates if requested is None else dates[dates.dt.normalize().le(requested)]
        if not eligible.empty:
            selected_path = path
            selected_date = pd.Timestamp(eligible.max()).normalize()
            break
    if selected_path is None or pd.isna(selected_date):
        raise RuntimeError(f"Parquet 中没有不晚于 {requested_date} 的交易日期")

    panel = pd.read_parquet(selected_path)
    panel_dates = pd.to_datetime(panel["交易日期"], errors="coerce").dt.normalize()
    snapshot = panel.loc[panel_dates.eq(selected_date)].copy()
    snapshot["转债代码"] = snapshot["转债代码"].astype(str).str.strip().str.upper()
    snapshot = snapshot.drop_duplicates("转债代码", keep="last").set_index("转债代码")
    return snapshot, selected_date


def fetch_downward_base_data_local(
    last_date: str | None = None,
    parquet_root: Path = DEFAULT_PARQUET_ROOT,
) -> pd.DataFrame:
    """仅从 Parquet 生成下修跟踪所需基础数据，不访问任何外部接口。"""
    global LOCAL_AS_OF_DATE

    total_path = parquet_root / "_special" / "总表.parquet"
    if not total_path.is_file():
        raise FileNotFoundError(f"未找到总表 Parquet: {total_path}")
    total = pd.read_parquet(total_path)
    snapshot, cutoff = _load_local_parquet_snapshot(parquet_root, last_date)
    LOCAL_AS_OF_DATE = cutoff

    required_total = {
        "转债代码",
        "转债名称",
        "发行日期",
        "最后交易日",
        "发行规模",
        "回售起始日期",
        "下修触发比例",
        "重设触发计算时间区间",
        "重设触发计算最大时间区间",
        "股票发行面值",
        "下修条款全文",
    }
    required_snapshot = {
        "余额",
        "收盘价",
        "平价",
        "转股价",
        "正股收盘价",
        "正股近1日均价",
        "正股近20日均价",
        "转股稀释率",
        "每股净资产",
    }
    missing_total = sorted(required_total - set(total.columns))
    missing_snapshot = sorted(required_snapshot - set(snapshot.columns))
    if missing_total or missing_snapshot:
        raise KeyError(
            f"本地化下修测算缺少字段；总表={missing_total}，月度序列={missing_snapshot}"
        )

    total["转债代码"] = total["转债代码"].astype(str).str.strip().str.upper()
    total = total.drop_duplicates("转债代码", keep="last").set_index("转债代码")
    issue_dates = pd.to_datetime(total["发行日期"], errors="coerce").dt.normalize()
    last_trade_dates = pd.to_datetime(total["最后交易日"], errors="coerce").dt.normalize()
    universe_mask = (
        (issue_dates.isna() | issue_dates.le(cutoff))
        & (last_trade_dates.isna() | last_trade_dates.ge(cutoff))
        & ~total.index.to_series().str.contains("NQ", case=False, na=False)
        & ~total.index.to_series().isin(EXCLUDE_BONDS)
    )
    total = total.loc[universe_mask].copy()
    snapshot = snapshot.reindex(total.index)

    base = pd.DataFrame(index=total.index)
    base.index.name = "转债代码"
    base["转债简称"] = total["转债名称"]
    base["转股价"] = pd.to_numeric(snapshot["转股价"], errors="coerce")
    base["正股收盘价"] = pd.to_numeric(snapshot["正股收盘价"], errors="coerce")
    base["转债收盘价"] = pd.to_numeric(snapshot["收盘价"], errors="coerce")
    base["平价"] = pd.to_numeric(snapshot["平价"], errors="coerce")
    base["转债余额"] = pd.to_numeric(snapshot["余额"], errors="coerce").combine_first(
        pd.to_numeric(total["发行规模"], errors="coerce")
    )
    base["条件回售起始日期"] = total["回售起始日期"]
    base["时间区间"] = pd.to_numeric(
        total["重设触发计算最大时间区间"], errors="coerce"
    )
    base["计算天数"] = pd.to_numeric(
        total["重设触发计算时间区间"], errors="coerce"
    )
    base["过去20日成交均价"] = pd.to_numeric(
        snapshot["正股近20日均价"], errors="coerce"
    )
    base["上一交易日成交均价"] = pd.to_numeric(
        snapshot["正股近1日均价"], errors="coerce"
    )
    base["转股稀释率"] = pd.to_numeric(snapshot["转股稀释率"], errors="coerce")
    base["每股净资产"] = pd.to_numeric(snapshot["每股净资产"], errors="coerce")
    base["股票发行面值"] = pd.to_numeric(total["股票发行面值"], errors="coerce")
    base["下修触发比例"] = pd.to_numeric(total["下修触发比例"], errors="coerce")
    base["下修条款全文"] = total["下修条款全文"]
    base.attrs["data_date"] = cutoff
    return base


def _non_empty_mask(series: pd.Series) -> pd.Series:
    return series.notna() & series.astype(str).str.strip().ne("") & series.astype(str).str.strip().ne("——")


def _empty_mask(series: pd.Series) -> pd.Series:
    return ~_non_empty_mask(series)


def _first_existing_column(df: pd.DataFrame, candidates: list[str]) -> str:
    for col in candidates:
        if col in df.columns:
            return col
    raise KeyError(f"缺少必要字段，候选字段: {candidates}；当前字段: {list(df.columns)}")


def _optional_existing_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for col in candidates:
        if col in df.columns:
            return col
    return None


def _merge_with_base(base_data: pd.DataFrame, detail_df: pd.DataFrame, code_col: str = "转债代码") -> pd.DataFrame:
    if detail_df.empty:
        return detail_df
    base = base_data.reset_index()
    merged = detail_df.merge(base, on=code_col, how="left", suffixes=("", "_基础数据"))
    return merged


def _numeric_max(df: pd.DataFrame, columns: list[str]) -> pd.Series:
    values = pd.DataFrame(index=df.index)
    for col in columns:
        values[col] = pd.to_numeric(df[col], errors="coerce") if col in df.columns else pd.NA
    return values.max(axis=1, skipna=True)


def _calculate_theoretical_downward_floor(group: pd.DataFrame) -> pd.Series:
    """按真实下修条款决定净资产与股票面值是否参与底价测算。"""
    values = pd.DataFrame(index=group.index)
    for column in ["过去20日成交均价", "上一交易日成交均价"]:
        values[column] = (
            pd.to_numeric(group[column], errors="coerce")
            if column in group.columns
            else pd.NA
        )
    clause_text = (
        group["下修条款全文"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
        if "下修条款全文" in group.columns
        else pd.Series("", index=group.index, dtype="string")
    )
    net_asset = (
        pd.to_numeric(group["每股净资产"], errors="coerce")
        if "每股净资产" in group.columns
        else pd.Series(pd.NA, index=group.index, dtype="Float64")
    )
    stock_par_value = (
        pd.to_numeric(group["股票发行面值"], errors="coerce")
        if "股票发行面值" in group.columns
        else pd.Series(pd.NA, index=group.index, dtype="Float64")
    )
    values["每股净资产"] = net_asset.where(
        clause_text.str.contains("每股净资产", regex=False, na=False)
    )
    values["股票发行面值"] = stock_par_value.where(
        clause_text.str.contains("股票面值", regex=False, na=False)
    )
    return values.max(axis=1, skipna=True)


def _rename_column_if_present(df: pd.DataFrame, source_col: str | None, target_col: str) -> None:
    if source_col is None or source_col == target_col:
        return
    if target_col in df.columns:
        df[target_col] = df[target_col].combine_first(df[source_col])
    else:
        df.rename(columns={source_col: target_col}, inplace=True)


def _select_existing_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    return df[[col for col in columns if col in df.columns]]


def _is_date_output_column(column: str) -> bool:
    if column == "时间区间":
        return False
    return any(keyword in column for keyword in ["日期", "公告时间", "公告日", "决议时间", "起始日", "何日", "报告期"])


def _format_date_column(series: pd.Series) -> pd.Series:
    raw = series.astype("string").str.strip()
    parsed = pd.to_datetime(raw, errors="coerce")
    yyyymmdd_mask = raw.str.fullmatch(r"\d{8}", na=False)
    if yyyymmdd_mask.any():
        parsed.loc[yyyymmdd_mask] = pd.to_datetime(raw.loc[yyyymmdd_mask], format="%Y%m%d", errors="coerce")

    formatted = raw.astype("object")
    mask = parsed.notna()
    formatted.loc[mask] = parsed.loc[mask].dt.strftime("%Y-%m-%d")
    formatted.loc[raw.isna()] = pd.NA
    return formatted


def _prepare_excel_output(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    formatted = df.copy()
    numeric_columns: list[str] = []

    for col in formatted.columns:
        col_name = str(col)
        if _is_date_output_column(col_name):
            formatted[col] = _format_date_column(formatted[col])
            continue
        if col_name in INTEGER_OUTPUT_COLUMNS:
            formatted[col] = pd.to_numeric(formatted[col], errors="coerce").round(0)
            continue

        non_empty = formatted[col].notna() & formatted[col].astype(str).str.strip().ne("")
        numeric_values = pd.to_numeric(formatted[col], errors="coerce")
        if non_empty.any() and numeric_values.loc[non_empty].notna().all():
            formatted[col] = numeric_values.round(2)
            numeric_columns.append(col_name)

    return formatted, numeric_columns


def _contains_chinese(value: Any) -> bool:
    return any("\u4e00" <= char <= "\u9fff" for char in str(value))


def _group_sheet_body_font(value: Any) -> Font:
    font_name = "KaiTi_GB2312" if _contains_chinese(value) else "Times New Roman"
    return Font(name=font_name, size=10, bold=False)


def _auto_width(worksheet, multiplier: float = 1.5) -> None:
    for column in worksheet.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value is not None), default=0)
        worksheet.column_dimensions[column[0].column_letter].width = (max_length + 2) * multiplier


def _format_group_sheet(
    worksheet,
    body_color: str | None,
    style_metadata: dict[str, Any] | None = None,
) -> None:
    style_metadata = style_metadata or {}
    history_date_columns = set(style_metadata.get("history_date_columns", []))
    hit_dates_by_code = style_metadata.get("hit_dates_by_code", {})
    downward_dates_by_code = style_metadata.get("downward_dates_by_code", {})
    non_downward_dates_by_code = style_metadata.get("non_downward_dates_by_code", {})
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    alignment = Alignment(horizontal="center", vertical="center")
    title_font = Font(name="KaiTi_GB2312", size=10, bold=False, color="FFFFFF")
    title_fill = PatternFill(
        start_color=GROUP_SHEET_HEADER_COLOR,
        end_color=GROUP_SHEET_HEADER_COLOR,
        fill_type="solid",
    )
    body_fill = (
        PatternFill(start_color=body_color, end_color=body_color, fill_type="solid")
        if body_color
        else PatternFill(fill_type=None)
    )
    counted_fill = PatternFill(
        start_color=COUNTED_DAY_COLOR,
        end_color=COUNTED_DAY_COLOR,
        fill_type="solid",
    )
    downward_fill = PatternFill(
        start_color=DOWNWARD_ANNOUNCEMENT_DAY_COLOR,
        end_color=DOWNWARD_ANNOUNCEMENT_DAY_COLOR,
        fill_type="solid",
    )
    non_downward_fill = PatternFill(
        start_color=NON_DOWNWARD_ANNOUNCEMENT_DAY_COLOR,
        end_color=NON_DOWNWARD_ANNOUNCEMENT_DAY_COLOR,
        fill_type="solid",
    )

    header_by_col = {
        cell.column: str(cell.value).strip()
        for cell in worksheet[1]
        if cell.value is not None
    }
    code_col = next(
        (col for col, header in header_by_col.items() if header == "转债代码"),
        None,
    )
    cumulative_col = next(
        (col for col, header in header_by_col.items() if header == "目前天数累计"),
        None,
    )

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.row == 1:
                cell.font = title_font
                cell.fill = title_fill
            else:
                cell.font = _group_sheet_body_font(cell.value)
                cell.fill = body_fill
            cell.border = thin_border
            cell.alignment = alignment

    for row_no in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_no].height = EXCEL_ROW_HEIGHT_POINTS

    if cumulative_col is not None:
        for row_no in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row_no, cumulative_col)
            font_name = "KaiTi_GB2312" if _contains_chinese(cell.value) else "Times New Roman"
            cell.font = Font(
                name=font_name,
                size=10,
                bold=True,
                color=CUMULATIVE_DAYS_FONT_COLOR,
            )

    if code_col is not None and history_date_columns:
        for row_no in range(2, worksheet.max_row + 1):
            code = str(worksheet.cell(row_no, code_col).value or "").strip()
            hit_dates = hit_dates_by_code.get(code, set())
            downward_dates = downward_dates_by_code.get(code, set())
            non_downward_dates = non_downward_dates_by_code.get(code, set())
            for col_no, date_label in header_by_col.items():
                if date_label not in history_date_columns:
                    continue
                cell = worksheet.cell(row_no, col_no)
                if date_label in downward_dates:
                    cell.fill = downward_fill
                    cell.font = Font(name="Times New Roman", size=10, bold=False, color="FFFFFF")
                elif date_label in non_downward_dates:
                    cell.fill = non_downward_fill
                    cell.font = Font(name="Times New Roman", size=10, bold=False, color="FFFFFF")
                elif date_label in hit_dates:
                    cell.fill = counted_fill

    _auto_width(worksheet)


def write_excel_groups(out_path: Path, groups: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, df in groups.items():
            output_df, numeric_columns = _prepare_excel_output(df)
            write_index = sheet_name == "基础数据"
            output_df.to_excel(writer, sheet_name=sheet_name, index=write_index)

            worksheet = writer.sheets[sheet_name]
            start_col = 2 if write_index else 1
            for idx, col in enumerate(output_df.columns, start=start_col):
                if str(col) in INTEGER_OUTPUT_COLUMNS:
                    number_format = "0"
                elif str(col) in numeric_columns:
                    number_format = "0.00"
                else:
                    continue
                for row in worksheet.iter_rows(min_row=2, min_col=idx, max_col=idx):
                    row[0].number_format = number_format

            body_color = GROUP_SHEET_BODY_COLORS.get(sheet_name)
            if body_color or sheet_name == RECENT_PARITY_SHEET:
                parity_body_color = None if sheet_name == RECENT_PARITY_SHEET else body_color
                _format_group_sheet(worksheet, parity_body_color, style_metadata=df.attrs)


def _load_image_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    font_candidates = [
        str(ROOT / "assets/fonts/KaiTi_GB2312.ttf"),
        r"C:\Windows\Fonts\simkai.ttf",
        r"C:\Windows\Fonts\simfang.ttf",
        r"C:\Windows\Fonts\msyh.ttc",
    ]
    for font_path in font_candidates:
        try:
            return ImageFont.truetype(font_path, size)
        except OSError:
            continue
    return ImageFont.load_default()


def _format_image_value(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, (pd.Timestamp, datetime)):
        return pd.Timestamp(value).strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.2f}"
    return str(value)


def _table_column_width_weights(columns: list[str]) -> list[float]:
    widths: list[float] = []
    for column in columns:
        col = str(column)
        if col == "分组":
            widths.append(0.55)
        elif col in {"转债代码", "正股代码"}:
            widths.append(0.85)
        elif col in {"转债名称", "转债简称", "正股名称"}:
            widths.append(0.85)
        elif re.fullmatch(r"\d{4}-\d{2}-\d{2}", col):
            widths.append(0.72)
        elif "日期" in col or "公告" in col or "决议" in col or "时间" in col:
            widths.append(0.92)
        elif col in INTEGER_OUTPUT_COLUMNS:
            widths.append(0.65)
        else:
            widths.append(0.72)
    return widths


def _normalize_column_widths(widths: list[float]) -> list[float]:
    total = sum(widths) or 1
    return [width / total for width in widths]


def _table_column_widths(columns: list[str]) -> list[float]:
    return _normalize_column_widths(_table_column_width_weights(columns))


def _shared_table_column_widths(column_groups: list[list[str]]) -> list[float]:
    """按列位置合并多张表的宽度，确保上下表格的竖向边界完全一致。"""
    if not column_groups:
        return []
    column_counts = {len(columns) for columns in column_groups}
    if len(column_counts) != 1:
        counts = ", ".join(str(count) for count in sorted(column_counts))
        raise ValueError(f"上下表格列数不一致（{counts}），无法生成严格对齐的长图")

    weight_groups = [_table_column_width_weights(columns) for columns in column_groups]
    shared_weights = [
        max(weights[column_index] for weights in weight_groups)
        for column_index in range(len(column_groups[0]))
    ]
    return _normalize_column_widths(shared_weights)


def _style_table_image_cells(table, df: pd.DataFrame, body_color: str) -> None:
    header_color = f"#{GROUP_SHEET_HEADER_COLOR}"
    body_color = f"#{body_color}"
    counted_color = f"#{COUNTED_DAY_COLOR}"
    downward_color = f"#{DOWNWARD_ANNOUNCEMENT_DAY_COLOR}"
    non_downward_color = f"#{NON_DOWNWARD_ANNOUNCEMENT_DAY_COLOR}"

    columns = [str(col) for col in df.columns]
    date_columns = set(df.attrs.get("history_date_columns", []))
    hit_dates_by_code = df.attrs.get("hit_dates_by_code", {})
    downward_dates_by_code = df.attrs.get("downward_dates_by_code", {})
    non_downward_dates_by_code = df.attrs.get("non_downward_dates_by_code", {})
    code_col_idx = columns.index("转债代码") if "转债代码" in columns else None
    cumulative_col_idx = columns.index("目前天数累计") if "目前天数累计" in columns else None

    for (row, col), cell in table._cells.items():
        text = cell.get_text()
        text.set_ha("center")
        text.set_va("center")
        text.set_fontfamily("KaiTi_GB2312" if _contains_chinese(text.get_text()) else "Times New Roman")
        cell.set_linewidth(0.15)
        if row == 0:
            cell.set_facecolor(header_color)
            text.set_color("white")
            continue

        cell.set_facecolor(body_color)
        if col == cumulative_col_idx:
            text.set_color(f"#{CUMULATIVE_DAYS_FONT_COLOR}")
            text.set_fontweight("bold")
        if code_col_idx is None or col >= len(columns):
            continue
        column_label = columns[col]
        if column_label not in date_columns:
            continue
        code = str(df.iloc[row - 1, code_col_idx] or "").strip().upper()
        if column_label in downward_dates_by_code.get(code, set()):
            cell.set_facecolor(downward_color)
            text.set_color("white")
        elif column_label in non_downward_dates_by_code.get(code, set()):
            cell.set_facecolor(non_downward_color)
            text.set_color("white")
        elif column_label in hit_dates_by_code.get(code, set()):
            cell.set_facecolor(counted_color)


def plot_downward_table_image(
    df: pd.DataFrame,
    output_path: Path,
    sheet_name: str,
    *,
    column_widths: list[float] | None = None,
) -> None:
    plt.rcParams["font.sans-serif"] = ["KaiTi_GB2312", "SimHei", "Microsoft YaHei"]
    output_df, _ = _prepare_excel_output(df)
    output_df.attrs = df.attrs.copy()
    table_df = output_df.fillna("").map(_format_image_value)
    figure_width = min(18, max(IMAGE_TABLE_WIDTH_INCH, len(table_df.columns) * 0.32))
    figure_height = max(
        0.6,
        (len(table_df) + 1) * IMAGE_ROW_HEIGHT_POINTS / 72,
    )

    widths = column_widths or _table_column_widths(table_df.columns.tolist())
    if len(widths) != len(table_df.columns):
        raise ValueError(
            f"{sheet_name} 的列数与共享列宽数量不一致："
            f"{len(table_df.columns)} != {len(widths)}"
        )

    fig, ax = plt.subplots(figsize=(figure_width, figure_height))
    fig.subplots_adjust(left=0, right=1, top=1, bottom=0)
    ax.axis("off")
    table = ax.table(
        cellText=table_df.values.tolist(),
        colLabels=table_df.columns.tolist(),
        colWidths=widths,
        bbox=[0, 0, 1, 1],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(5)
    table.scale(1, 1.0)
    _style_table_image_cells(table, output_df, GROUP_SHEET_BODY_COLORS.get(sheet_name, "FFFFFF"))
    fig.savefig(output_path, dpi=IMAGE_TABLE_DPI, bbox_inches=None, pad_inches=0, transparent=False)
    plt.close(fig)


def render_downward_header_image(header_template: Path, output_path: Path, title_date: str) -> None:
    img = Image.open(header_template).convert("RGBA")
    draw = ImageDraw.Draw(img)
    text = IMAGE_TITLE_TEMPLATE.format(date=title_date)
    font = _load_image_font(60)
    bbox = draw.multiline_textbbox((0, 0), text, font=font, spacing=10)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]
    # 蓝色标题区大致为模板中部横条；按模板比例定位，避免依赖固定像素。
    x = int(img.width * 0.555 - text_width / 2)
    y = int(img.height * 0.43 - text_height / 2)
    draw.multiline_text((x, y), text, fill="white", font=font, spacing=10, align="center")
    img.save(output_path)
    img.close()


def combine_downward_images(image_paths: list[Path], output_path: Path) -> None:
    images = [Image.open(path).convert("RGBA") for path in image_paths if path.exists()]
    if not images:
        raise FileNotFoundError("没有可拼接的下修表格图片")
    max_width = max(image.width for image in images)
    resized: list[Image.Image] = []
    for image in images:
        if image.width == max_width:
            resized.append(image)
            continue
        height = round(max_width / image.width * image.height)
        resized.append(image.resize((max_width, height), resample=Image.BILINEAR))

    total_height = sum(image.height for image in resized)
    combined = Image.new("RGBA", (max_width, total_height), (255, 255, 255, 255))
    y = 0
    for image in resized:
        combined.paste(image, (0, y))
        y += image.height
    combined.save(output_path)

    for image in images:
        image.close()
    for image in resized:
        if image not in images:
            image.close()
    combined.close()


def write_downward_tracking_image(
    image_out_path: Path,
    groups: dict[str, pd.DataFrame],
    *,
    header_template: Path = DEFAULT_HEADER_IMAGE,
    title_date: str,
) -> None:
    if not header_template.exists():
        raise FileNotFoundError(f"未找到图片表头模板: {header_template}")
    image_out_path.parent.mkdir(parents=True, exist_ok=True)
    temp_dir = image_out_path.parent / f"_{image_out_path.stem}_parts"
    temp_dir.mkdir(parents=True, exist_ok=True)
    try:
        header_path = temp_dir / "下修表题头.png"
        render_downward_header_image(header_template, header_path, title_date)
        image_paths = [header_path]
        image_column_groups: list[list[str]] = []
        for sheet_name in IMAGE_GROUP_ORDER:
            output_df, _ = _prepare_excel_output(groups[sheet_name])
            image_column_groups.append(output_df.columns.tolist())
        shared_column_widths = _shared_table_column_widths(image_column_groups)

        for sheet_name in IMAGE_GROUP_ORDER:
            part_path = temp_dir / f"{sheet_name}.png"
            plot_downward_table_image(
                groups[sheet_name],
                part_path,
                sheet_name,
                column_widths=shared_column_widths,
            )
            image_paths.append(part_path)

        combine_downward_images(image_paths, image_out_path)
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def _add_downward_valuation_columns(group: pd.DataFrame) -> pd.DataFrame:
    group["理论下修底价"] = _calculate_theoretical_downward_floor(group)
    group["下修到底对应平价"] = (
        pd.to_numeric(group.get("正股收盘价"), errors="coerce")
        / pd.to_numeric(group["理论下修底价"], errors="coerce")
        * 100
    )
    reference_date = LOCAL_AS_OF_DATE or pd.Timestamp.today().normalize()
    group["距离回售天数"] = (
        pd.to_datetime(group.get("条件回售起始日期"), errors="coerce") - reference_date
    ).dt.days
    group.rename(columns={"条件回售起始日期": "回售起始日期"}, inplace=True)
    return group


def _apply_announced_downward_floor(
    group: pd.DataFrame,
    announced_floor_col: str | None,
) -> pd.DataFrame:
    """实施下修已披露底价时优先采用公告值；公告值缺失则保留原理论底价。"""
    if announced_floor_col is None or announced_floor_col not in group.columns:
        return group

    announced_floor = pd.to_numeric(group[announced_floor_col], errors="coerce")
    valid_announced_floor = announced_floor.notna() & announced_floor.gt(0)
    group.loc[valid_announced_floor, "理论下修底价"] = announced_floor.loc[valid_announced_floor]
    group["下修到底对应平价"] = (
        pd.to_numeric(group.get("正股收盘价"), errors="coerce")
        / pd.to_numeric(group["理论下修底价"], errors="coerce")
        * 100
    )
    return group


def build_implementing_downward_group(
    base_data: pd.DataFrame,
    database_path: Path = DEFAULT_DATABASE,
    as_of_date: pd.Timestamp | None = None,
) -> pd.DataFrame:
    """按每只券最新提议判断：转股价尚未实际生效且未失败/取消的均属实施下修中。"""
    downward_df = pd.read_excel(database_path, sheet_name="下修")
    board_col = _first_existing_column(downward_df, ["董事会发布日期", "董事会提议时间", "董事会提议日期"])
    resolution_col = _first_existing_column(
        downward_df,
        ["股东大会决议时间", "股东大会决议日期", "向下修正发布日期"],
    )
    code_col = _first_existing_column(downward_df, ["转债代码"])
    name_col = _optional_existing_column(downward_df, ["转债名称", "转债简称"])
    revised_price_col = _optional_existing_column(downward_df, ["修正后转股价", "向下修正后转股价"])
    announced_floor_col = _optional_existing_column(
        downward_df,
        ["修正转股价应不低于", "修正后转股价应不低于", "下修底价"],
    )
    change_date_col = _optional_existing_column(downward_df, ["转股价变动日期"])
    failure_date_col = _optional_existing_column(downward_df, ["下修失败/取消公告日期"])

    downward_df[code_col] = downward_df[code_col].astype(str).str.strip()
    cutoff = pd.Timestamp(as_of_date or pd.Timestamp.today()).normalize()
    board_sort_col = "__董事会发布日期排序"
    downward_df[board_sort_col] = pd.to_datetime(downward_df[board_col], errors="coerce")
    downward_df = downward_df[
        downward_df[board_sort_col].dt.normalize().le(cutoff)
    ].copy()
    latest_df = (
        downward_df.sort_values(
            [code_col, board_sort_col],
            ascending=[True, False],
            na_position="last",
            kind="mergesort",
        )
        .drop_duplicates(code_col, keep="first")
        .copy()
    )

    failed_or_cancelled = pd.Series(False, index=latest_df.index)
    failure_marker_cols = {
        col
        for col in [change_date_col, resolution_col, failure_date_col]
        if col is not None
    }
    failure_marker_cols.update(
        col
        for col in latest_df.columns
        if "失败" in str(col) or "取消" in str(col)
    )
    for marker_col in failure_marker_cols:
        failed_or_cancelled |= latest_df[marker_col].astype("string").str.contains(
            r"下修失败|取消",
            regex=True,
            na=False,
        )
    if failure_date_col:
        failed_or_cancelled |= _non_empty_mask(latest_df[failure_date_col])

    board_dates = pd.to_datetime(latest_df[board_col], errors="coerce").dt.normalize()
    if change_date_col:
        effective_dates = pd.to_datetime(
            latest_df[change_date_col],
            errors="coerce",
        ).dt.normalize()
    else:
        effective_dates = pd.Series(pd.NaT, index=latest_df.index, dtype="datetime64[ns]")
    effective_date_not_reached = effective_dates.isna() | effective_dates.gt(cutoff)
    group = latest_df[
        board_dates.notna()
        & effective_date_not_reached
        & ~failed_or_cancelled
    ].copy()
    group.drop(columns=[board_sort_col], inplace=True, errors="ignore")
    group.insert(0, "分组", "实施下修中")
    group = group.sort_values(board_col, ascending=False, na_position="last")
    group = _merge_with_base(base_data, group, code_col=code_col)

    _rename_column_if_present(group, board_col, "董事会提议公告时间")
    _rename_column_if_present(group, resolution_col, "股东大会决议时间")
    _rename_column_if_present(group, name_col, "转债名称")
    _rename_column_if_present(group, revised_price_col, "修正后转股价")

    if "转股价" in group.columns:
        if "修正前转股价" in group.columns:
            group["修正前转股价"] = group["转股价"]
        else:
            group.insert(group.columns.get_loc("转股价") + 1, "修正前转股价", group["转股价"])

    group = _add_downward_valuation_columns(group)
    group = _apply_announced_downward_floor(group, announced_floor_col)
    display_columns = [
        "转债代码",
        "转债名称",
        "时间区间",
        "计算天数",
        "董事会提议公告时间",
        "股东大会决议时间",
        "修正前转股价",
        "修正后转股价",
        "理论下修底价",
        "下修到底对应平价",
        "正股收盘价",
        "转债收盘价",
        "平价",
        "转债余额",
        "回售起始日期",
        "距离回售天数",
        "转股稀释率",
    ]
    group = _select_existing_columns(group, display_columns)
    return group


def _parse_parquet_col_to_ts(col: Any) -> pd.Timestamp:
    if isinstance(col, pd.Timestamp):
        return col.normalize()
    if isinstance(col, datetime):
        return pd.Timestamp(col).normalize()
    ts = pd.to_datetime(col, errors="coerce")
    if pd.notna(ts):
        return pd.Timestamp(ts).normalize()
    return pd.NaT


@lru_cache(maxsize=8)
def load_parquet_sheet_wide(
    sheet_name: str,
    parquet_root: Path = DEFAULT_PARQUET_ROOT,
) -> pd.DataFrame:
    """从按月 parquet 目录还原指定 sheet 的完整历史宽表。"""
    if not parquet_root.is_dir():
        raise FileNotFoundError(f"未找到 parquet 目录: {parquet_root}")

    merged = read_metric_wide(parquet_root, sheet_name).T
    merged.index = merged.index.astype(str).str.strip().str.upper()
    merged.index.name = "转债代码"
    return merged


def load_downward_cumulative_days_wide(parquet_root: Path = DEFAULT_PARQUET_ROOT) -> pd.DataFrame:
    """从 parquet 目录还原「下修累计天数」宽表。"""
    return load_parquet_sheet_wide(DOWNWARD_DAYS_SHEET, parquet_root)


def load_latest_downward_cumulative_days(parquet_root: Path = DEFAULT_PARQUET_ROOT) -> tuple[pd.DataFrame, pd.Timestamp]:
    """读取 parquet 中最新交易日的下修累计天数快照。"""
    wide = load_downward_cumulative_days_wide(parquet_root)
    latest_date = pd.Timestamp(wide.columns[-1]).normalize()
    days = pd.to_numeric(wide.iloc[:, -1], errors="coerce")
    cumulative_df = pd.DataFrame(
        {
            "转债代码": wide.index.astype(str).str.strip(),
            "数据日期": latest_date,
            "下修累计天数": days.values,
        }
    )
    cumulative_df = cumulative_df[cumulative_df["下修累计天数"].notna() & (cumulative_df["下修累计天数"] > 0)]
    return cumulative_df.reset_index(drop=True), latest_date


def _bond_code_base(code: Any) -> str:
    return str(code or "").strip().upper().split(".")[0]


def _downward_to_str_code(value: Any) -> str:
    if isinstance(value, pd.Series):
        non_na = value.dropna()
        value = non_na.iloc[0] if not non_na.empty else ""
    if value is None or pd.isna(value):
        return ""
    return str(value).strip().upper()


def _downward_is_failure_marker(value: Any) -> bool:
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return False
    return "下修失败" in str(value).strip()


def _downward_parse_maybe_date(value: Any) -> pd.Timestamp:
    if value is None:
        return pd.NaT
    if isinstance(value, pd.Timestamp):
        return value if pd.notna(value) else pd.NaT
    if isinstance(value, (int, float, np.integer, np.floating)) and not pd.isna(value):
        return pd.to_datetime(float(value), unit="D", origin="1899-12-30", errors="coerce")
    text = str(value).strip()
    if not text or text.lower() in {"nan", "nat", "none"}:
        return pd.NaT
    direct = pd.to_datetime(text, errors="coerce")
    if pd.notna(direct):
        return pd.Timestamp(direct)
    match = re.search(r"(20\d{2})[./-年](\d{1,2})[./-月](\d{1,2})", text)
    if not match:
        return pd.NaT
    year, month, day = match.groups()
    try:
        return pd.Timestamp(year=int(year), month=int(month), day=int(day))
    except ValueError:
        return pd.NaT


def _load_downward_backtest_announcements(
    database_path: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """读取累计天数状态机所需的下修与不下修公告。"""
    if not database_path.is_file():
        raise FileNotFoundError(f"未找到公告文件: {database_path}")
    downward_df = pd.read_excel(database_path, sheet_name="下修")
    non_downward_df = pd.read_excel(database_path, sheet_name="不下修")

    downward_required = ("转债代码", "董事会发布日期", "转股价变动日期")
    non_downward_required = ("转债代码", "公告时间", "承诺何日之前不行使")
    for column in downward_required:
        if column not in downward_df.columns:
            raise ValueError(f"下修 sheet 缺少必要字段: {column}")
    for column in non_downward_required:
        if column not in non_downward_df.columns:
            raise ValueError(f"不下修 sheet 缺少必要字段: {column}")

    if "向下修正发布日期" not in downward_df.columns:
        downward_df["向下修正发布日期"] = pd.NaT
    downward_df = downward_df[
        [*downward_required, "向下修正发布日期"]
    ].copy()
    non_downward_df = non_downward_df[list(non_downward_required)].copy()

    downward_df["转债代码"] = downward_df["转债代码"].map(_downward_to_str_code)
    non_downward_df["转债代码"] = non_downward_df["转债代码"].map(_downward_to_str_code)
    downward_df["转股价变动日期_原文"] = downward_df["转股价变动日期"]
    downward_df["董事会发布日期"] = pd.to_datetime(
        downward_df["董事会发布日期"], errors="coerce"
    )
    downward_df["向下修正发布日期"] = pd.to_datetime(
        downward_df["向下修正发布日期"], errors="coerce"
    )
    non_downward_df["公告时间"] = pd.to_datetime(
        non_downward_df["公告时间"], errors="coerce"
    )
    raw_promise = non_downward_df["承诺何日之前不行使"].copy()
    promise_end = raw_promise.map(_downward_parse_maybe_date)
    promise_text = raw_promise.astype("string").str.strip()
    no_promise_marker = (
        raw_promise.isna()
        | promise_text.fillna("").isin({"", "——", "—", "-"})
        | promise_text.fillna("").str.lower().isin({"nan", "nat", "none"})
    )
    is_silu = non_downward_df["转债代码"].map(_bond_code_base).eq("123138")
    silu_fallback = is_silu & promise_end.isna() & ~no_promise_marker
    promise_end.loc[silu_fallback] = (
        non_downward_df.loc[silu_fallback, "公告时间"]
        .dt.normalize()
        .map(lambda date: date + pd.DateOffset(months=6))
    )
    non_downward_df["承诺何日之前不行使"] = promise_end

    downward_df = downward_df.dropna(subset=["董事会发布日期"])
    non_downward_df = non_downward_df.dropna(subset=["公告时间"])
    return downward_df, non_downward_df


def _build_downward_backtest_event_maps(
    downward_df: pd.DataFrame,
    non_downward_df: pd.DataFrame,
) -> tuple[dict[str, list[tuple[pd.Timestamp, pd.Timestamp | None, bool]]], dict[str, list[tuple[pd.Timestamp, pd.Timestamp | None]]]]:
    downward_map: dict[str, list[tuple[pd.Timestamp, pd.Timestamp | None, bool]]] = {}
    non_downward_map: dict[str, list[tuple[pd.Timestamp, pd.Timestamp | None]]] = {}

    for _, row in downward_df.iterrows():
        code = _downward_to_str_code(row["转债代码"])
        if not code:
            continue
        start = pd.Timestamp(row["董事会发布日期"])
        raw_change = row.get("转股价变动日期_原文", row.get("转股价变动日期"))
        if _downward_is_failure_marker(raw_change):
            failure_date = row.get("向下修正发布日期", pd.NaT)
            end = pd.Timestamp(failure_date) if pd.notna(failure_date) else None
            event = (start, end, True)
        else:
            change_date = pd.to_datetime(raw_change, errors="coerce")
            end = pd.Timestamp(change_date) if pd.notna(change_date) else None
            event = (start, end, False)
        for key in {code, _bond_code_base(code)}:
            if key:
                downward_map.setdefault(key, []).append(event)

    for _, row in non_downward_df.iterrows():
        code = _downward_to_str_code(row["转债代码"])
        if not code:
            continue
        publish_date = pd.Timestamp(row["公告时间"])
        promise_end = row.get("承诺何日之前不行使", pd.NaT)
        end = pd.Timestamp(promise_end) if pd.notna(promise_end) else None
        event = (publish_date, end)
        for key in {code, _bond_code_base(code)}:
            if key:
                non_downward_map.setdefault(key, []).append(event)

    for event_map in (downward_map, non_downward_map):
        for key in event_map:
            event_map[key] = sorted(set(event_map[key]), key=lambda event: event[0])
    return downward_map, non_downward_map


def _merge_downward_backtest_events(event_map: dict[str, list], code: str) -> list:
    exact = _downward_to_str_code(code)
    base = _bond_code_base(exact)
    merged = [*event_map.get(exact, [])]
    if base != exact:
        merged.extend(event_map.get(base, []))
    return sorted(set(merged), key=lambda event: event[0]) if merged else []


def _next_trade_day(
    current_day: pd.Timestamp,
    trade_days: pd.DatetimeIndex,
) -> pd.Timestamp | None:
    position = int(trade_days.searchsorted(current_day, side="right"))
    return None if position >= len(trade_days) else pd.Timestamp(trade_days[position])


def _calculate_single_bond_downward_days(
    stock_price: pd.Series,
    conversion_price: pd.Series,
    stock_trade_status: pd.Series,
    issue_date: pd.Timestamp,
    listing_date: pd.Timestamp,
    last_trade_date: pd.Timestamp,
    trigger_ratio: float,
    trigger_days: int,
    window_days: int,
    downward_events: Iterable[tuple[pd.Timestamp, pd.Timestamp | None, bool]],
    non_downward_events: Iterable[tuple[pd.Timestamp, pd.Timestamp | None]],
) -> pd.Series:
    """按正股价格、有效转股价和正股交易状态计算单券下修累计进度。"""
    stock_price = pd.to_numeric(stock_price, errors="coerce")
    conversion_price = pd.to_numeric(conversion_price, errors="coerce")
    stock_trade_status = stock_trade_status.astype("string")
    trade_days = pd.DatetimeIndex(stock_price.index).sort_values()
    stock_price = stock_price.reindex(trade_days)
    conversion_price = conversion_price.reindex(trade_days)
    stock_trade_status = stock_trade_status.reindex(trade_days)
    result = pd.Series(index=trade_days, dtype="float64")
    if pd.isna(issue_date):
        return result
    issue_date = pd.Timestamp(issue_date).normalize()
    listing_date = (
        pd.Timestamp(listing_date).normalize()
        if pd.notna(listing_date)
        else pd.NaT
    )
    last_trade_date = (
        pd.Timestamp(last_trade_date).normalize()
        if pd.notna(last_trade_date)
        else pd.NaT
    )

    # 临时口径：发行后至上市前缺失的转股价，使用该券发行后首个可得
    # 转股价向前回填。上市后不跨期填补，避免掩盖转股价序列自身的缺口。
    first_available = conversion_price.loc[conversion_price.index >= issue_date].first_valid_index()
    if first_available is not None:
        first_price = conversion_price.loc[first_available]
        prelisting_mask = conversion_price.index >= issue_date
        if pd.notna(listing_date):
            prelisting_mask &= conversion_price.index < listing_date
        else:
            prelisting_mask &= conversion_price.index <= first_available
        fill_mask = prelisting_mask & conversion_price.isna()
        conversion_price.loc[fill_mask] = first_price

    def next_trade_day_normalized(anchor: pd.Timestamp) -> pd.Timestamp | None:
        next_day = _next_trade_day(pd.Timestamp(anchor).normalize(), trade_days)
        return None if next_day is None else next_day.normalize()

    placeholder_end = trade_days.max().normalize() if len(trade_days) else None
    far_future = pd.Timestamp("2099-12-31")
    downward_intervals: list[tuple[pd.Timestamp, pd.Timestamp]] = []
    downward_reset_days: set[pd.Timestamp] = set()
    failure_reset_next_days: set[pd.Timestamp] = set()
    for start_raw, end_raw, is_failure in downward_events:
        start = pd.Timestamp(start_raw).normalize()
        if is_failure:
            if end_raw is not None and pd.notna(end_raw):
                end = pd.Timestamp(end_raw).normalize()
                if end > start:
                    downward_intervals.append((start, end))
                    next_day = next_trade_day_normalized(end)
                else:
                    next_day = next_trade_day_normalized(start)
            else:
                next_day = next_trade_day_normalized(start)
            if next_day is not None:
                failure_reset_next_days.add(next_day)
            continue

        if end_raw is not None and pd.notna(end_raw):
            end = pd.Timestamp(end_raw).normalize()
            if end <= start:
                end = placeholder_end if placeholder_end is not None and placeholder_end > start else far_future
            else:
                downward_reset_days.add(end)
        else:
            end = placeholder_end if placeholder_end is not None and placeholder_end > start else far_future
        downward_intervals.append((start, end))
    downward_intervals.sort(key=lambda interval: interval[0])
    downward_announce_days = {start for start, _ in downward_intervals}

    non_downward_intervals: list[tuple[pd.Timestamp, pd.Timestamp]] = []
    non_downward_reset_next_days: set[pd.Timestamp] = set()
    for publish_raw, end_raw in non_downward_events:
        publish = pd.Timestamp(publish_raw).normalize()
        if end_raw is not None and pd.notna(end_raw):
            end = pd.Timestamp(end_raw).normalize()
            if end <= publish:
                end = publish
                next_day = next_trade_day_normalized(publish)
                if next_day is not None:
                    non_downward_reset_next_days.add(next_day)
        else:
            end = publish
            next_day = next_trade_day_normalized(publish)
            if next_day is not None:
                non_downward_reset_next_days.add(next_day)
        non_downward_intervals.append((publish, end))
    non_downward_intervals.sort(key=lambda interval: interval[0])
    non_downward_announce_days = {start for start, _ in non_downward_intervals}

    freeze_intervals = [
        *( (start, end, "downward") for start, end in downward_intervals ),
        *( (start, end, "non_downward") for start, end in non_downward_intervals ),
    ]
    freeze_intervals.sort(key=lambda interval: interval[0])

    trigger_days = max(int(trigger_days), 1)
    window_days = max(int(window_days), 1)
    window_hits: list[int] = []
    window_hit_sum = 0
    trigger_reached_previous_day = False

    for date in trade_days:
        normalized_date = date.normalize()
        if pd.notna(last_trade_date) and normalized_date > last_trade_date:
            result.loc[date] = np.nan
            continue
        if normalized_date < issue_date:
            result.loc[date] = np.nan
            continue

        frozen = False
        for start, end, source in freeze_intervals:
            in_interval = (
                start < normalized_date < end
                if source == "downward"
                else start < normalized_date <= end
            )
            if in_interval:
                frozen = True
                break
        if frozen:
            window_hits.clear()
            window_hit_sum = 0
            trigger_reached_previous_day = False
            result.loc[date] = 0.0
            continue

        if (
            normalized_date in non_downward_reset_next_days
            or normalized_date in downward_reset_days
            or normalized_date in failure_reset_next_days
        ):
            window_hits.clear()
            window_hit_sum = 0
            trigger_reached_previous_day = False

        # 仅正股发生交易的日期推进最近 B 个交易日窗口；盘中停牌当日已有
        # 实际交易，也视为有效交易日。全天/连续停牌或状态缺失时累计值
        # 保持不变。转债自身是否上市或交易不再参与判断。
        status_value = stock_trade_status.loc[date]
        is_stock_trading_day = (
            pd.notna(status_value)
            and (
                status_value in {"交易", "新股上市"}
                or "盘中停牌" in str(status_value)
            )
        )
        if not is_stock_trading_day:
            result.loc[date] = float(window_hit_sum)
            continue

        stock_value = stock_price.loc[date]
        conversion_value = conversion_price.loc[date]
        if pd.isna(stock_value) or pd.isna(conversion_value):
            result.loc[date] = float(window_hit_sum)
            continue

        trigger_price = float(conversion_value) * float(trigger_ratio) / 100.0
        hit_today = int(float(stock_value) < trigger_price)
        window_hits.append(hit_today)
        window_hit_sum += hit_today
        if len(window_hits) > window_days:
            window_hit_sum -= window_hits.pop(0)
        result.loc[date] = float(window_hit_sum)

        triggered_today = window_hit_sum >= trigger_days
        if triggered_today and not trigger_reached_previous_day:
            if normalized_date not in downward_announce_days and normalized_date not in non_downward_announce_days:
                window_hits.clear()
                window_hit_sum = 0
        trigger_reached_previous_day = triggered_today

    return result


def calculate_downward_cumulative_days(
    parquet_root: Path,
    database_path: Path,
) -> pd.DataFrame:
    """仅用 Parquet 底层字段与公告库全历史计算下修累计天数。"""
    total_path = parquet_root / "_special" / "总表.parquet"
    if not total_path.is_file():
        raise FileNotFoundError(f"未找到总表 Parquet: {total_path}")
    total = pd.read_parquet(total_path)
    required = {
        BOND_CODE,
        "发行日期",
        "上市日期",
        "最后交易日",
        "下修触发比例",
        "重设触发计算时间区间",
        "重设触发计算最大时间区间",
    }
    missing = sorted(required - set(total.columns))
    if missing:
        raise KeyError(f"总表缺少下修累计测算字段: {missing}")
    total[BOND_CODE] = total[BOND_CODE].map(_downward_to_str_code)
    total = total.drop_duplicates(BOND_CODE, keep="last").set_index(BOND_CODE)

    stock_price = load_parquet_sheet_wide("正股收盘价", parquet_root)
    conversion_price = load_parquet_sheet_wide("转股价", parquet_root)
    stock_trade_status = load_parquet_sheet_wide("正股交易状态", parquet_root)
    dates = pd.DatetimeIndex(stock_price.columns).sort_values()
    stock_price = stock_price.reindex(columns=dates)
    conversion_price = conversion_price.reindex(columns=dates)
    stock_trade_status = stock_trade_status.reindex(columns=dates)
    common_codes = (
        stock_price.index
        .intersection(total.index)
        .intersection(conversion_price.index)
        .intersection(stock_trade_status.index)
    )
    result = pd.DataFrame(np.nan, index=common_codes, columns=dates, dtype="float64")

    downward_df, non_downward_df = _load_downward_backtest_announcements(database_path)
    downward_map, non_downward_map = _build_downward_backtest_event_maps(
        downward_df,
        non_downward_df,
    )
    print(
        f"[downward] 全历史重算：{len(common_codes)}只转债 × {len(dates)}个交易日。",
        flush=True,
    )
    for position, code in enumerate(common_codes, start=1):
        trigger_ratio = pd.to_numeric(
            pd.Series([total.at[code, "下修触发比例"]]), errors="coerce"
        ).iloc[0]
        trigger_days = pd.to_numeric(
            pd.Series([total.at[code, "重设触发计算时间区间"]]), errors="coerce"
        ).iloc[0]
        window_days = pd.to_numeric(
            pd.Series([total.at[code, "重设触发计算最大时间区间"]]), errors="coerce"
        ).iloc[0]
        if pd.isna(trigger_ratio) or pd.isna(trigger_days) or pd.isna(window_days):
            continue
        one_result = _calculate_single_bond_downward_days(
            stock_price=stock_price.loc[code],
            conversion_price=conversion_price.loc[code],
            stock_trade_status=stock_trade_status.loc[code],
            issue_date=pd.to_datetime(total.at[code, "发行日期"], errors="coerce"),
            listing_date=pd.to_datetime(total.at[code, "上市日期"], errors="coerce"),
            last_trade_date=pd.to_datetime(total.at[code, "最后交易日"], errors="coerce"),
            trigger_ratio=float(trigger_ratio),
            trigger_days=int(trigger_days),
            window_days=int(window_days),
            downward_events=_merge_downward_backtest_events(downward_map, code),
            non_downward_events=_merge_downward_backtest_events(non_downward_map, code),
        )
        result.loc[code] = one_result.reindex(dates).to_numpy()
        if position % 100 == 0 or position == len(common_codes):
            print(f"[downward] 已完成 {position}/{len(common_codes)}只。", flush=True)

    result.index.name = BOND_CODE
    result.columns.name = "交易日期"
    return result


def recalculate_and_write_downward_cumulative_days(
    parquet_root: Path,
    database_path: Path,
) -> tuple[pd.DataFrame, int]:
    """全历史重算并将正累计值写回月度 Parquet。"""
    result = calculate_downward_cumulative_days(parquet_root, database_path)
    parquet_result = result.mask(result.eq(0))
    observations = replace_monthly_metric_from_wide(
        parquet_root,
        DOWNWARD_DAYS_SHEET,
        parquet_result,
    )
    load_parquet_sheet_wide.cache_clear()
    print(
        f"[downward] 已写回 Parquet：{observations}个非空累计天数观测。",
        flush=True,
    )
    return parquet_result, observations


def _load_announcement_date_maps(
    database_path: Path,
) -> tuple[dict[str, set[str]], dict[str, set[str]]]:
    """读取全部下修/不下修公告日，供最近30日平价单元格着色。"""
    downward_df = pd.read_excel(database_path, sheet_name="下修")
    non_downward_df = pd.read_excel(database_path, sheet_name="不下修")
    downward_code_col = _first_existing_column(downward_df, ["转债代码"])
    downward_date_col = _first_existing_column(
        downward_df,
        ["董事会发布日期", "董事会提议时间", "董事会提议日期"],
    )
    non_downward_code_col = _first_existing_column(non_downward_df, ["转债代码"])
    non_downward_date_col = _first_existing_column(
        non_downward_df,
        ["公告时间", "不下修公告日", "公告日期"],
    )

    def _build_map(df: pd.DataFrame, code_col: str, date_col: str) -> dict[str, set[str]]:
        result: dict[str, set[str]] = {}
        dates = pd.to_datetime(df[date_col], errors="coerce")
        for code_raw, date_value in zip(df[code_col], dates):
            if pd.isna(date_value):
                continue
            code = str(code_raw or "").strip().upper()
            date_label = pd.Timestamp(date_value).strftime("%Y-%m-%d")
            for key in {code, _bond_code_base(code)}:
                if key:
                    result.setdefault(key, set()).add(date_label)
        return result

    return (
        _build_map(downward_df, downward_code_col, downward_date_col),
        _build_map(non_downward_df, non_downward_code_col, non_downward_date_col),
    )


def _build_recent_parity_history_sheet(
    groups: dict[str, pd.DataFrame],
    base_data: pd.DataFrame,
    database_path: Path,
    parquet_root: Path,
) -> pd.DataFrame:
    """构造独立的近30日平价 sheet，并准备命中日/公告日配色元数据。"""
    parity_wide = load_parquet_sheet_wide(PARITY_SHEET, parquet_root)
    cumulative_wide = load_downward_cumulative_days_wide(parquet_root)
    recent_dates = pd.DatetimeIndex(parity_wide.columns[-RECENT_PARITY_DAYS:])
    history_labels = [pd.Timestamp(date).strftime("%Y-%m-%d") for date in recent_dates]
    cumulative_recent = cumulative_wide.reindex(columns=recent_dates)
    downward_map, non_downward_map = _load_announcement_date_maps(database_path)

    trigger_ratios: dict[str, float] = {}
    if "下修触发比例" in base_data.columns:
        for code_raw, ratio_raw in base_data["下修触发比例"].items():
            code = str(code_raw or "").strip().upper()
            ratio = pd.to_numeric(pd.Series([ratio_raw]), errors="coerce").iloc[0]
            for key in {code, _bond_code_base(code)}:
                if key:
                    trigger_ratios[key] = ratio

    rows: list[pd.DataFrame] = []
    hit_dates_by_code: dict[str, set[str]] = {}
    downward_dates_by_code: dict[str, set[str]] = {}
    non_downward_dates_by_code: dict[str, set[str]] = {}

    for sheet_name, source_df in groups.items():
        if sheet_name == RECENT_PARITY_SHEET or source_df.empty or "转债代码" not in source_df.columns:
            continue

        columns = ["转债代码"]
        if "分组" in source_df.columns:
            columns.insert(0, "分组")
        for optional_col in ["转债名称", "目前天数累计"]:
            if optional_col in source_df.columns:
                columns.append(optional_col)
        group = source_df[columns].copy()
        if "分组" not in group.columns:
            group.insert(0, "分组", sheet_name)
        else:
            group["分组"] = group["分组"].fillna(sheet_name)

        codes = group["转债代码"].astype(str).str.strip().str.upper()
        for date_value, date_label in zip(recent_dates, history_labels):
            parity_on_date = pd.to_numeric(parity_wide.get(date_value), errors="coerce")
            parity_lookup = parity_on_date.to_dict() if parity_on_date is not None else {}
            group[date_label] = codes.map(parity_lookup)

        for code in codes:
            base_code = _bond_code_base(code)
            ratio = trigger_ratios.get(code, trigger_ratios.get(base_code, float("nan")))
            hit_dates: set[str] = set()
            if pd.notna(ratio) and code in parity_wide.index:
                parity_row = pd.to_numeric(
                    parity_wide.loc[code].reindex(recent_dates),
                    errors="coerce",
                )
                if code in cumulative_recent.index:
                    cumulative_row = pd.to_numeric(
                        cumulative_recent.loc[code].reindex(recent_dates),
                        errors="coerce",
                    )
                    hit_mask = parity_row.lt(float(ratio)) & cumulative_row.gt(0)
                    hit_dates = {
                        pd.Timestamp(date).strftime("%Y-%m-%d")
                        for date in recent_dates[hit_mask.fillna(False).to_numpy()]
                    }
            hit_dates_by_code[code] = hit_dates
            downward_dates_by_code[code] = (
                downward_map.get(code, set()) | downward_map.get(base_code, set())
            )
            non_downward_dates_by_code[code] = (
                non_downward_map.get(code, set()) | non_downward_map.get(base_code, set())
            )

        rows.append(group)

    if rows:
        parity_sheet = pd.concat(rows, ignore_index=True)
    else:
        parity_sheet = pd.DataFrame(columns=["分组", "转债代码", "转债名称", "目前天数累计", *history_labels])

    # 部分分组（如“实施下修中”）没有“目前天数累计”，直接 concat 会把该列
    # 追加到所有日期列之后。显式重排，确保累计天数始终紧邻转债名称。
    leading_columns = ["分组", "转债代码", "转债名称", "目前天数累计"]
    parity_sheet = parity_sheet.reindex(columns=[*leading_columns, *history_labels])

    parity_sheet.attrs["history_date_columns"] = history_labels
    parity_sheet.attrs["hit_dates_by_code"] = hit_dates_by_code
    parity_sheet.attrs["downward_dates_by_code"] = downward_dates_by_code
    parity_sheet.attrs["non_downward_dates_by_code"] = non_downward_dates_by_code
    return parity_sheet


def _parse_non_downward_promise_end(
    raw_value: Any,
    *,
    code: str,
    announce_date: pd.Timestamp,
) -> pd.Timestamp:
    """解析不下修承诺期末；丝路转债的非空非日期文本按公告日起六个月兜底。"""
    if raw_value is None or (not isinstance(raw_value, str) and pd.isna(raw_value)):
        return pd.NaT
    if isinstance(raw_value, (pd.Timestamp, datetime)):
        return pd.Timestamp(raw_value).normalize()
    if isinstance(raw_value, (int, float)) and not pd.isna(raw_value):
        parsed_serial = pd.to_datetime(float(raw_value), unit="D", origin="1899-12-30", errors="coerce")
        return pd.Timestamp(parsed_serial).normalize() if pd.notna(parsed_serial) else pd.NaT

    text = str(raw_value).strip()
    if not text or text.lower() in {"nan", "nat", "none"} or text in {"——", "—", "-"}:
        return pd.NaT
    parsed = pd.to_datetime(text, errors="coerce")
    if pd.notna(parsed):
        return pd.Timestamp(parsed).normalize()

    match = re.search(r"(20\d{2})[./-年](\d{1,2})[./-月](\d{1,2})", text)
    if match:
        year, month, day = match.groups()
        try:
            return pd.Timestamp(year=int(year), month=int(month), day=int(day))
        except ValueError:
            return pd.NaT

    if str(code).strip().split(".")[0] == "123138" and pd.notna(announce_date):
        return pd.Timestamp(announce_date).normalize() + pd.DateOffset(months=6)
    return pd.NaT


def _load_latest_non_downward_records(
    database_path: Path,
    as_of_date: pd.Timestamp | None = None,
) -> pd.DataFrame:
    """读取截至基准日每只转债最新一条不下修公告，并解析承诺期末。"""
    non_downward_df = pd.read_excel(database_path, sheet_name="不下修")
    code_col = _first_existing_column(non_downward_df, ["转债代码"])
    announce_col = _first_existing_column(non_downward_df, ["公告时间", "不下修公告日", "公告日期"])
    promise_col = _first_existing_column(non_downward_df, ["承诺何日之前不行使"])
    non_downward_df = non_downward_df[[code_col, announce_col, promise_col]].copy()
    non_downward_df[code_col] = non_downward_df[code_col].astype(str).str.strip()
    non_downward_df[announce_col] = pd.to_datetime(non_downward_df[announce_col], errors="coerce")
    if as_of_date is not None:
        cutoff = pd.Timestamp(as_of_date).normalize()
        non_downward_df = non_downward_df[
            non_downward_df[announce_col].dt.normalize().le(cutoff)
        ].copy()
    non_downward_df = non_downward_df.sort_values(announce_col, ascending=False, na_position="last")
    non_downward_df = non_downward_df.drop_duplicates(code_col, keep="first")
    non_downward_df["__承诺截止日"] = non_downward_df.apply(
        lambda row: _parse_non_downward_promise_end(
            row[promise_col],
            code=row[code_col],
            announce_date=row[announce_col],
        ),
        axis=1,
    )
    non_downward_df.rename(
        columns={
            code_col: "转债代码",
            announce_col: "不下修公告日",
            promise_col: "承诺何日之前不行使",
        },
        inplace=True,
    )
    return non_downward_df


def _cumulative_style_display_columns() -> list[str]:
    """“当前具有累计天数”及同布局分组的统一列顺序。"""
    return [
        "转债代码",
        "转债名称",
        "时间区间",
        "计算天数",
        "不下修公告日",
        "承诺何日之前不行使",
        "目前天数累计",
        "转股价",
        "理论下修底价",
        "下修到底对应平价",
        "正股收盘价",
        "转债收盘价",
        "平价",
        "转债余额",
        "回售起始日期",
        "距离回售天数",
        "转股稀释率",
    ]


def build_current_cumulative_group(
    base_data: pd.DataFrame,
    parquet_root: Path | None = None,
    database_path: Path = DEFAULT_DATABASE,
    exclude_codes: set[str] | None = None,
) -> pd.DataFrame:
    """当前有下修累计天数的转债。"""
    root = parquet_root or DEFAULT_PARQUET_ROOT
    try:
        cumulative_df, latest_date = load_latest_downward_cumulative_days(root)
    except (FileNotFoundError, RuntimeError) as exc:
        print(f"⚠️ 未能从 parquet 读取下修累计天数: {exc}")
        return pd.DataFrame()

    code_col = "转债代码"
    days_col = "下修累计天数"
    non_downward_df = _load_latest_non_downward_records(database_path, as_of_date=latest_date)
    announcement_on_latest_date = pd.to_datetime(
        non_downward_df["不下修公告日"],
        errors="coerce",
    ).dt.normalize().eq(latest_date)
    announcement_day_codes = set(
        non_downward_df.loc[announcement_on_latest_date, code_col]
        .dropna()
        .astype(str)
        .str.strip()
    )

    # parquet 保留公告日实际累计值；展示层在不下修公告日将其视作已归零，
    # 再由后续分组按是否存在有效承诺期分别归入承诺期页或无累计页。
    group = cumulative_df[~cumulative_df[code_col].isin(announcement_day_codes)].copy()
    if exclude_codes:
        group = group[~group[code_col].isin(exclude_codes)]
    group.insert(0, "分组", "当前具有累计天数")
    group.insert(1, "累计天数来源", f"{root.name} @ {latest_date.strftime('%Y-%m-%d')}")
    group = group.sort_values(days_col, ascending=False, na_position="last")
    group = _merge_with_base(base_data, group, code_col=code_col)

    group = group.merge(non_downward_df, on=code_col, how="left")

    _rename_column_if_present(group, _optional_existing_column(group, ["转债名称", "转债简称"]), "转债名称")
    _rename_column_if_present(group, days_col, "目前天数累计")
    group = _add_downward_valuation_columns(group)
    return _select_existing_columns(group, _cumulative_style_display_columns())


def build_active_non_downward_commitment_group(
    base_data: pd.DataFrame,
    database_path: Path = DEFAULT_DATABASE,
    as_of_date: pd.Timestamp | None = None,
    exclude_codes: set[str] | None = None,
) -> pd.DataFrame:
    """当前仍处于不下修承诺期内的转债，列布局与累计天数组一致。"""
    code_col = "转债代码"
    days_col = "下修累计天数"
    cutoff = pd.Timestamp(as_of_date or pd.Timestamp.today()).normalize()
    group = _load_latest_non_downward_records(database_path, as_of_date=cutoff)

    current_codes = set(base_data.index.dropna().astype(str).str.strip())
    promise_end = pd.to_datetime(group["__承诺截止日"], errors="coerce").dt.normalize()
    announce_date = pd.to_datetime(group["不下修公告日"], errors="coerce").dt.normalize()
    active_mask = promise_end.ge(cutoff) & promise_end.gt(announce_date)
    group = group[group[code_col].isin(current_codes) & active_mask].copy()
    if exclude_codes:
        group = group[~group[code_col].isin(exclude_codes)]
    group.insert(0, "分组", "当前不下修承诺期内")
    group[days_col] = 0
    group = _merge_with_base(base_data, group, code_col=code_col)
    _rename_column_if_present(group, _optional_existing_column(group, ["转债名称", "转债简称"]), "转债名称")
    _rename_column_if_present(group, days_col, "目前天数累计")
    group = _add_downward_valuation_columns(group)
    group["__不下修公告日排序"] = pd.to_datetime(
        group["不下修公告日"], errors="coerce"
    ).dt.normalize()
    group = group.sort_values(
        ["__不下修公告日排序", code_col],
        ascending=[False, True],
        na_position="last",
        kind="mergesort",
    )
    group.drop(columns=["__不下修公告日排序"], inplace=True)
    return _select_existing_columns(group, _cumulative_style_display_columns())


def build_no_commitment_no_cumulative_group(
    base_data: pd.DataFrame,
    database_path: Path = DEFAULT_DATABASE,
    as_of_date: pd.Timestamp | None = None,
    exclude_codes: set[str] | None = None,
) -> pd.DataFrame:
    """不在不下修承诺期内、也没有累计天数的其余当前转债。"""
    code_col = "转债代码"
    days_col = "下修累计天数"
    cutoff = pd.Timestamp(as_of_date or pd.Timestamp.today()).normalize()
    excluded = {str(code).strip() for code in (exclude_codes or set())}
    current_codes = set(base_data.index.dropna().astype(str).str.strip())
    remaining_codes = sorted(current_codes - excluded)

    group = pd.DataFrame({code_col: remaining_codes})
    latest_non_downward = _load_latest_non_downward_records(database_path, as_of_date=cutoff)
    group = group.merge(latest_non_downward, on=code_col, how="left")
    group.insert(0, "分组", "非承诺期且无累计天数")
    group[days_col] = 0
    group = _merge_with_base(base_data, group, code_col=code_col)
    _rename_column_if_present(group, _optional_existing_column(group, ["转债名称", "转债简称"]), "转债名称")
    _rename_column_if_present(group, days_col, "目前天数累计")
    group = _add_downward_valuation_columns(group)
    group = group.sort_values(["转债名称", code_col], ascending=True, na_position="last")
    return _select_existing_columns(group, _cumulative_style_display_columns())


def build_display_groups(
    base_data: pd.DataFrame,
    database_path: Path = DEFAULT_DATABASE,
    parquet_root: Path | None = None,
) -> dict[str, pd.DataFrame]:
    try:
        _, latest_date = load_latest_downward_cumulative_days(parquet_root or DEFAULT_PARQUET_ROOT)
    except (FileNotFoundError, RuntimeError):
        latest_date = pd.Timestamp.today().normalize()

    implementing_df = build_implementing_downward_group(
        base_data,
        database_path=database_path,
        as_of_date=latest_date,
    )
    implementing_codes = set()
    if "转债代码" in implementing_df.columns:
        implementing_codes = set(implementing_df["转债代码"].dropna().astype(str).str.strip())
    cumulative_df = build_current_cumulative_group(
        base_data,
        parquet_root=parquet_root,
        database_path=database_path,
        exclude_codes=implementing_codes,
    )
    cumulative_codes = set()
    if "转债代码" in cumulative_df.columns:
        cumulative_codes = set(cumulative_df["转债代码"].dropna().astype(str).str.strip())

    active_non_downward_df = build_active_non_downward_commitment_group(
        base_data,
        database_path=database_path,
        as_of_date=latest_date,
        exclude_codes=implementing_codes | cumulative_codes,
    )
    active_non_downward_codes = set()
    if "转债代码" in active_non_downward_df.columns:
        active_non_downward_codes = set(
            active_non_downward_df["转债代码"].dropna().astype(str).str.strip()
        )

    no_commitment_no_cumulative_df = build_no_commitment_no_cumulative_group(
        base_data,
        database_path=database_path,
        as_of_date=latest_date,
        exclude_codes=implementing_codes | cumulative_codes | active_non_downward_codes,
    )
    groups = {
        "实施下修中": implementing_df,
        "当前具有累计天数": cumulative_df,
        "当前不下修承诺期内": active_non_downward_df,
        "非承诺期且无累计天数": no_commitment_no_cumulative_df,
    }
    try:
        groups[RECENT_PARITY_SHEET] = _build_recent_parity_history_sheet(
            groups,
            base_data=base_data,
            database_path=database_path,
            parquet_root=parquet_root or DEFAULT_PARQUET_ROOT,
        )
    except (FileNotFoundError, RuntimeError) as exc:
        print(f"⚠️ 未能追加最近{RECENT_PARITY_DAYS}日平价数据: {exc}")
    return groups


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="使用本地 Parquet 生成可转债下修进度跟踪表")
    parser.add_argument("--date", default=None, help="数据截止日期，格式 YYYY-MM-DD；默认使用 Parquet 最新日期")
    parser.add_argument("--out", default=str(DEFAULT_OUTPUT), help="输出 Excel 路径")
    parser.add_argument("--image-out", default=None, help="输出下修进度长图 PNG 路径；默认与 Excel 同目录同名")
    parser.add_argument("--header-image", default=str(DEFAULT_HEADER_IMAGE), help="长图表头模板图片路径")
    parser.add_argument("--image-title-date", default=None, help="长图标题日期；默认使用最近30日平价中的最后一个日期")
    parser.add_argument("--no-image", action="store_true", help="只输出 Excel，不生成下修进度长图")
    parser.add_argument("--database", default=str(DEFAULT_DATABASE), help="转股价修正信息跟踪数据库路径")
    parser.add_argument("--parquet-root", default=str(DEFAULT_PARQUET_ROOT), help="底稿 Parquet 根目录")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    parquet_root = Path(args.parquet_root)
    database_path = Path(args.database)
    recalculate_and_write_downward_cumulative_days(
        parquet_root=parquet_root,
        database_path=database_path,
    )
    merged_data = fetch_downward_base_data_local(
        args.date,
        parquet_root=parquet_root,
    )
    groups = build_display_groups(
        merged_data,
        database_path=database_path,
        parquet_root=parquet_root,
    )

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    write_excel_groups(out_path, groups)

    if not args.no_image:
        parity_sheet = groups.get(RECENT_PARITY_SHEET, pd.DataFrame())
        history_dates = list(parity_sheet.attrs.get("history_date_columns", []))
        image_title_date = args.image_title_date or (history_dates[-1] if history_dates else _time.strftime("%Y-%m-%d"))
        image_out_path = Path(args.image_out) if args.image_out else out_path.with_suffix(".png")
        write_downward_tracking_image(
            image_out_path,
            groups,
            header_template=Path(args.header_image),
            title_date=image_title_date,
        )
        print(f"已输出长图: {image_out_path}")

    print(f"已输出: {out_path}")
    if LOCAL_AS_OF_DATE is not None:
        print(f"Parquet 数据日期: {LOCAL_AS_OF_DATE.strftime('%Y-%m-%d')}")
    print(f"基础数据行数: {len(merged_data)}")
    print(f"实施下修中: {len(groups['实施下修中'])}")
    print(f"当前具有累计天数: {len(groups['当前具有累计天数'])}")
    print(f"当前不下修承诺期内: {len(groups['当前不下修承诺期内'])}")
    print(f"非承诺期且无累计天数: {len(groups['非承诺期且无累计天数'])}")


if __name__ == "__main__":
    main()
