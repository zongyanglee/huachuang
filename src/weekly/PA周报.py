"""
PA 周报一键生成脚本（纯 Python 单文件版）。

运行：
    py PA周报.py

本脚本不再调用外部 py 或 JS 文件。它会直接完成：
1. 读取「转债个券历史序列」parquet 数据；
2. 自动识别最新交易日；
3. 缺失时打开Excel执行工作簿自带的Wind EDB信用债数据抓取；
4. 通过 iFinD 自动补齐1.1的转债指数和全市场成交额；
5. 计算 1.1、1.3、1.4、1.5、1.6、1.7、1.11及2.1；
6. 用 openpyxl 直接生成 Excel 工作簿。

输出：
    PA周报YYYYMMDD/PA转债周度数据.xlsx
其中 YYYYMMDD 自动取 parquet 数据中的最新交易日。
"""

from __future__ import annotations

import json
import math
import os
import re
import sys
import tempfile
import time
import zipfile
from copy import deepcopy
from configparser import ConfigParser
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import pyarrow.parquet as pq
from lxml import etree as ET
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from pandas.tseries.offsets import BDay
from scipy.optimize import curve_fit

import sys

_COMMON_MODULE_DIR = Path(__file__).resolve().parents[1] / "common"
if str(_COMMON_MODULE_DIR) not in sys.path:
    sys.path.insert(0, str(_COMMON_MODULE_DIR))

from 转债Parquet标准读写模块 import BOND_CODE, INDEX_NAME, INDEX_VALUE, TRADE_DATE


ROOT = Path(__file__).resolve().parents[2]
PARQUET_ROOT = ROOT / "data/转债个券历史序列"
IFIND_CREDENTIAL_FILE = ROOT / "private/ifind账号.txt"
PA_WORKBOOK_NAME = "PA转债周度数据.xlsx"


def resolve_pa_workbook() -> Path:
    primary = ROOT / "runs" / "weekly" / PA_WORKBOOK_NAME
    if primary.exists():
        return primary
    previous_outputs = sorted(
        [
            *(ROOT / "runs" / "weekly").glob(f"PA周报20*/{PA_WORKBOOK_NAME}"),
            *(ROOT / "runs" / "weekly").glob(f"TK周报20*/{PA_WORKBOOK_NAME}"),
        ],
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return previous_outputs[0] if previous_outputs else primary


PA_WORKBOOK = resolve_pa_workbook()


def latest_parquet_trade_date() -> pd.Timestamp:
    files = sorted(PARQUET_ROOT.glob("20??/20????.parquet"))
    if not files:
        raise FileNotFoundError(f"未在 parquet 数据中找到交易日：{PARQUET_ROOT}")
    dates = pd.read_parquet(files[-1], columns=[TRADE_DATE])[TRADE_DATE]
    if dates.empty:
        raise FileNotFoundError(f"未在 parquet 数据中找到交易日：{PARQUET_ROOT}")
    return pd.Timestamp(dates.max())


LATEST_TRADE_DATE = latest_parquet_trade_date()
OUT = ROOT / "runs" / "weekly" / f"PA周报{LATEST_TRADE_DATE:%Y%m%d}"
OUT.mkdir(parents=True, exist_ok=True)

OUTPUT_XLSX = OUT / PA_WORKBOOK.name
META_JSON = OUT / "meta.json"

START_DATE = pd.Timestamp("2017-01-01")
DECOMP_START_DATE = pd.Timestamp("2018-01-01")
ROLLING_WINDOW = 20
CREDIT_REFRESH_TIMEOUT_SECONDS = 180
WEEKLY_MOVER_COUNT = 20

MARKET_COLUMNS = [
    "日期",
    "转债指数",
    "价格中位数",
    "价格均值",
    "价格余额加权",
    "百元拟合溢价率",
    "百元拟合溢价率2017年以来分位数",
    "隐含波动率均值",
    "隐含波动率自2017年以来分位数",
    "平价中位数",
    "转股溢价率中位数",
    "转股溢价率均值",
    "纯债价值中位数",
    "纯债溢价率中位数",
    "纯债溢价率均值",
    "YTM中位数",
    "YTM均值",
    "YTM大于0的比例",
    "YTM大于3年AA信用债比例(%)",
    "全市场成交额",
    "正股市场波动率",
    "转债市场隐含波动率中位数",
    "隐波差中位数",
]

STYLE_COLUMNS = [
    "日期",
    "转债等权指数",
    "转债加权指数",
    "正股等权指数",
    "正股加权指数",
    "转债大盘指数",
    "正股大盘指数",
    "转债中盘指数",
    "正股中盘指数",
    "转债小盘指数",
    "正股小盘指数",
    "转债AAA指数",
    "转债AA+指数",
    "转债AA指数",
    "转债AA-及以下指数",
    "正股AAA指数",
    "正股AA+指数",
    "正股AA指数",
    "正股AA-及以下指数",
]

HISTORY_VALUATION_COLUMNS = [
    "日期",
    "转债等权涨跌幅",
    "总成交额",
    "平均换手率",
    "上涨家数",
    "下跌家数",
    "前五成交额占总成交额比例",
    "转股溢价率中位数",
    "YTM中位数",
    "纯债溢价率中位数",
    "隐含波动率中位数",
    "剩余期限5.5-6年转债隐含波动率中位数",
    "70平价溢价率",
    "百元溢价率",
    "120平价溢价率",
    "收盘价中位数",
    "双低中位数",
    "YTM>0占比",
    "破面率",
    "跌破债底占比",
]

HISTORY_VALUATION_DAILY_COLUMNS = [
    "日期",
    "中证转债涨跌幅",
    *HISTORY_VALUATION_COLUMNS[1:],
]

DECOMP_COLUMNS = ["日期", "转债收益率", "债券贡献", "正股贡献", "估值贡献"]
DECOMP_GROUP_COLUMNS = ["分类", "转债收益率", "债券贡献", "正股贡献", "估值贡献"]

METRIC_NAMES = [
    "余额",
    "收盘价",
    "平价",
    "转股溢价率",
    "纯债价值",
    "纯债溢价率",
    "平价底价溢价率",
    "YTM",
    "换手率",
    "成交额",
    "涨跌幅",
    "剩余期限",
    "隐含波动率",
    "正股20日波动率",
    "正股收盘价",
    "债项评级",
]
TEXT_METRICS = {"债项评级"}

# 周报唯一有效的分平价口径；旧版七档分组不再使用。
PARITY_BUCKETS = [
    ("≤70", None, 70.0),
    ("70-80", 70.0, 80.0),
    ("80-90", 80.0, 90.0),
    ("90-100", 90.0, 100.0),
    ("100-110", 100.0, 110.0),
    ("110-120", 110.0, 120.0),
    ("120-130", 120.0, 130.0),
    ("130-140", 130.0, 140.0),
    ("140-150", 140.0, 150.0),
    ("150以上", 150.0, None),
]

TERM_BUCKETS = [
    ("[0, 0.5)", 0.0, 0.5),
    ("[0.5, 1)", 0.5, 1.0),
    ("[1, 1.5)", 1.0, 1.5),
    ("[1.5, 2)", 1.5, 2.0),
    ("[2, 2.5)", 2.0, 2.5),
    ("[2.5, 3)", 2.5, 3.0),
    ("[3, 3.5)", 3.0, 3.5),
    ("[3.5, 4)", 3.5, 4.0),
    ("[4, 4.5)", 4.0, 4.5),
    ("[4.5, 5)", 4.5, 5.0),
    ("[5, 5.5)", 5.0, 5.5),
    ("[5.5, 6)", 5.5, 6.0),
]

RATING_BUCKETS = ["AAA", "AA+", "AA", "AA-", "A+", "A", "A-"]

SIZE_BUCKETS = [
    ("0-5亿", 0.0, 5.0),
    ("5-10亿", 5.0, 10.0),
    ("10-20亿", 10.0, 20.0),
    ("20-50亿", 20.0, 50.0),
    ("50亿以上", 50.0, None),
]


def log(message: str) -> None:
    print(message, flush=True)


def format_elapsed(seconds: float) -> str:
    total_seconds = int(round(seconds))
    minutes, remaining_seconds = divmod(total_seconds, 60)
    if minutes:
        return f"{minutes}分{remaining_seconds}秒"
    return f"{remaining_seconds}秒"


def inverse_cubic(x: np.ndarray | float, a: float, b: float, c: float, d: float) -> np.ndarray | float:
    return a / np.power(x, 3) + b / np.power(x, 2) + c / x + d


def safe_mean(values: pd.Series) -> float | None:
    x = pd.to_numeric(values, errors="coerce").dropna()
    return float(x.mean()) if len(x) else None


def safe_median(values: pd.Series) -> float | None:
    x = pd.to_numeric(values, errors="coerce").dropna()
    return float(x.median()) if len(x) else None


def weighted_mean(values: pd.Series, weights: pd.Series) -> float | None:
    x = pd.to_numeric(values, errors="coerce")
    w = pd.to_numeric(weights, errors="coerce")
    valid = x.notna() & w.notna() & (w > 0)
    if not valid.any():
        return None
    denom = w[valid].sum()
    if pd.isna(denom) or denom == 0:
        return None
    return float((x[valid] * w[valid]).sum() / denom)


def weighted_return(ret: pd.Series, weight: pd.Series) -> float | None:
    r = pd.to_numeric(ret, errors="coerce")
    w = pd.to_numeric(weight, errors="coerce")
    valid = r.notna() & w.notna() & (w > 0)
    if not valid.any():
        return None
    denom = w[valid].sum()
    if pd.isna(denom) or denom == 0:
        return None
    return float((r[valid] * w[valid]).sum() / denom)


def equal_return(ret: pd.Series) -> float | None:
    r = pd.to_numeric(ret, errors="coerce").dropna()
    return float(r.mean()) if len(r) else None


def winsorize_premium_sample(sample: pd.DataFrame, low_q: float = 0.03, high_q: float = 0.97) -> pd.DataFrame:
    if sample.empty or "转股溢价率" not in sample.columns:
        return sample
    premium = pd.to_numeric(sample["转股溢价率"], errors="coerce")
    if premium.notna().sum() < 8:
        return sample.iloc[0:0]
    low = np.nanquantile(premium, low_q)
    high = np.nanquantile(premium, high_q)
    return sample[(premium > low) & (premium < high)].dropna(axis=0)


def fit_premium_at_x(
    plain: pd.Series,
    premium: pd.Series,
    turnover: pd.Series,
    target_x: float,
) -> float | None:
    """按分组多因子脚本的反三次口径计算指定平价处的拟合溢价率。"""
    sample = pd.DataFrame({"平价": plain, "转股溢价率": premium, "换手率": turnover})
    sample = sample.replace("", np.nan)
    sample["平价"] = pd.to_numeric(sample["平价"], errors="coerce").replace(0, np.nan)
    sample["转股溢价率"] = pd.to_numeric(sample["转股溢价率"], errors="coerce").replace(0, np.nan)
    sample["换手率"] = pd.to_numeric(sample["换手率"], errors="coerce")
    sample = sample.dropna(subset=["平价", "转股溢价率", "换手率"])
    sample = sample[
        (sample["换手率"] < 50)
        & (sample["平价"] > 70)
        & (sample["平价"] < 130)
    ]
    sample = winsorize_premium_sample(sample)
    if len(sample) < 8:
        return None
    try:
        popt, _ = curve_fit(
            inverse_cubic,
            sample["平价"].astype(float).values,
            sample["转股溢价率"].astype(float).values,
            maxfev=20000,
        )
        return float(inverse_cubic(float(target_x), *popt))
    except Exception:
        return None


def expanding_percentile(values: pd.Series) -> pd.Series:
    result: list[float | None] = []
    history: list[float] = []
    for value in values:
        if pd.isna(value):
            result.append(None)
            continue
        v = float(value)
        history.append(v)
        arr = np.asarray(history, dtype=float)
        result.append(float((np.sum(arr < v) + 0.5 * np.sum(arr == v)) / len(arr)))
    return pd.Series(result, index=values.index, dtype="float64")


def normalize_rating(value: object) -> str | None:
    if pd.isna(value):
        return None
    text = str(value).strip().upper().replace(" ", "")
    text = text.replace("STI", "")
    text = text.replace("（", "(").replace("）", ")")
    if not text or text in {"NAN", "NONE", "<NA>"}:
        return None
    return text


def combine_current_with_previous(current: pd.Series, previous: pd.Series) -> pd.Series:
    """以当期非空值优先合并历史值，并兼容 pandas 3.x 的 Series 名称处理。"""
    return current.rename(None).combine_first(previous.rename(None))


def load_3y_credit_yields(workbook_path: Path) -> pd.DataFrame:
    """读取PA工作簿中AAA至A-的3年企业债到期收益率百分点序列。"""
    if not workbook_path.exists():
        raise FileNotFoundError(f"未找到PA转债周度数据：{workbook_path}")
    with zipfile.ZipFile(workbook_path, "r") as archive:
        try:
            sheet_path = _sheet_xml_path(archive, "3年AA信用债")
            sheet_name = "3年AA信用债"
        except KeyError:
            sheet_path = _sheet_xml_path(archive, "Sheet1")
            sheet_name = "Sheet1"

        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            strings_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in strings_root.findall(_qname(_SHEET_NS, "si")):
                shared_strings.append("".join(item.itertext()))

        sheet_root = ET.fromstring(archive.read(sheet_path))
        rows = sheet_root.findall(f".//{_qname(_SHEET_NS, 'row')}")

        def read_cell(cell) -> object:
            value_node = cell.find(_qname(_SHEET_NS, "v"))
            if value_node is None or value_node.text is None:
                inline = cell.find(_qname(_SHEET_NS, "is"))
                return "".join(inline.itertext()) if inline is not None else None
            if cell.get("t") == "s":
                return shared_strings[int(value_node.text)]
            return value_node.text

        header_row = None
        rating_columns: dict[str, str] = {}
        for row in rows[:20]:
            for cell in row.findall(_qname(_SHEET_NS, "c")):
                value = read_cell(cell)
                text = "" if value is None else str(value).replace(" ", "")
                normalized_text = text.replace("（", "(").replace("）", ")")
                for rating in RATING_BUCKETS:
                    if f"到期收益率({rating}):3年" in normalized_text:
                        header_row = int(row.get("r", "0"))
                        rating_columns[rating] = _cell_column(cell.get("r", ""))
                        break
            if len(rating_columns) == len(RATING_BUCKETS):
                break
        missing_ratings = [rating for rating in RATING_BUCKETS if rating not in rating_columns]
        if header_row is None or missing_ratings:
            raise KeyError(
                f"{sheet_name} 中缺少3年信用债到期收益率列：{', '.join(missing_ratings)}"
            )

        records: list[dict[str, object]] = []
        for row in rows:
            row_number = int(row.get("r", "0"))
            if row_number <= header_row:
                continue
            cells = {
                _cell_column(cell.get("r", "")): cell
                for cell in row.findall(_qname(_SHEET_NS, "c"))
            }
            date_raw = pd.to_numeric(read_cell(cells.get("A")) if cells.get("A") is not None else None, errors="coerce")
            if pd.isna(date_raw):
                continue
            date = (
                pd.Timestamp("1899-12-30")
                + pd.to_timedelta(float(date_raw), unit="D")
            ).normalize()
            record: dict[str, object] = {"日期": date}
            for rating, column in rating_columns.items():
                yield_raw = pd.to_numeric(
                    read_cell(cells.get(column)) if cells.get(column) is not None else None,
                    errors="coerce",
                )
                record[rating] = float(yield_raw) if pd.notna(yield_raw) else np.nan
            records.append(record)
        if not records:
            raise ValueError(f"{sheet_name} 中没有可用的3年信用债收益率数据")
        result = pd.DataFrame(records).drop_duplicates("日期", keep="last").set_index("日期")
        return result[RATING_BUCKETS].sort_index().astype(float)


def load_3y_aa_credit_yields(workbook_path: Path) -> pd.Series:
    """读取3年AA企业债到期收益率，返回按日期索引的百分点序列。"""
    return load_3y_credit_yields(workbook_path)["AA"]


def _excel_date(value: object) -> pd.Timestamp | None:
    """将 Excel COM 返回的日期序列值或日期文本转换为 Timestamp。"""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return (
            pd.Timestamp("1899-12-30")
            + pd.to_timedelta(float(value), unit="D")
        ).normalize()
    parsed = pd.to_datetime(value, errors="coerce")
    return pd.Timestamp(parsed).normalize() if pd.notna(parsed) else None


def refresh_3y_credit_yields_via_excel(
    workbook_path: Path,
    expected_latest_date: pd.Timestamp,
    timeout_seconds: int = CREDIT_REFRESH_TIMEOUT_SECONDS,
) -> None:
    """打开 Excel 运行工作簿自带的 Wind EDB 抓取，并保存刷新结果。"""
    try:
        import pythoncom
        import win32com.client
    except ImportError as exc:
        raise RuntimeError("缺少pywin32，无法自动刷新3年AA信用债数据") from exc

    excel = None
    workbook = None
    pythoncom.CoInitialize()
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = True
        excel.ScreenUpdating = False
        try:
            workbook = excel.Workbooks.Open(
                str(workbook_path.resolve()),
                UpdateLinks=3,
                ReadOnly=False,
            )
        except Exception as exc:
            raise RuntimeError(
                f"无法用Excel打开PA工作簿，请确认文件未被占用：{workbook_path}"
            ) from exc
        if bool(workbook.ReadOnly):
            raise RuntimeError(
                f"PA工作簿以只读方式打开，无法保存信用债刷新结果；请先关闭Excel：{workbook_path}"
            )

        try:
            sheet = workbook.Worksheets("3年AA信用债")
        except Exception:
            try:
                sheet = workbook.Worksheets("Sheet1")
            except Exception as exc:
                raise KeyError("PA工作簿缺少3年AA信用债 Sheet") from exc

        # A1 的 =edb() 由 WindFunc.xla 提供；UpdateLinks、RefreshAll 和全量重算
        # 共同触发其自带数据抓取。A3:H3 是倒序序列的最新一期。
        workbook.RefreshAll()
        excel.CalculateFullRebuild()
        try:
            excel.CalculateUntilAsyncQueriesDone()
        except Exception:
            pass

        deadline = time.monotonic() + timeout_seconds
        latest_date: pd.Timestamp | None = None
        latest_values: tuple[object, ...] = ()
        while time.monotonic() < deadline:
            time.sleep(2)
            try:
                excel.CalculateUntilAsyncQueriesDone()
            except Exception:
                pass
            latest_date = _excel_date(sheet.Range("A3").Value2)
            raw_values = sheet.Range("B3:H3").Value2
            if isinstance(raw_values, tuple) and raw_values:
                first_row = raw_values[0] if isinstance(raw_values[0], tuple) else raw_values
                latest_values = tuple(first_row)
            else:
                latest_values = ()
            numeric_complete = len(latest_values) == len(RATING_BUCKETS) and all(
                pd.notna(pd.to_numeric(value, errors="coerce"))
                for value in latest_values
            )
            if (
                latest_date is not None
                and latest_date >= pd.Timestamp(expected_latest_date).normalize()
                and numeric_complete
                and int(excel.CalculationState) == 0
            ):
                workbook.Save()
                log(f"3年AA信用债数据已通过Excel/Wind刷新至：{latest_date:%Y-%m-%d}")
                return

        latest_text = latest_date.strftime("%Y-%m-%d") if latest_date is not None else "无有效日期"
        raise TimeoutError(
            "Excel/Wind刷新3年AA信用债数据超时："
            f"期望截至{pd.Timestamp(expected_latest_date):%Y-%m-%d}，实际截至{latest_text}。"
            "请确认Wind插件已登录且工作簿外部链接可用。"
        )
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def ensure_3y_credit_yields_current(
    workbook_path: Path,
    required_dates: pd.DatetimeIndex,
) -> pd.DataFrame:
    """信用债序列未覆盖所需交易日时，先调用Excel自带抓取再重新读取。"""
    required = pd.DatetimeIndex(required_dates).normalize().unique().sort_values()
    credit_yields = load_3y_credit_yields(workbook_path)
    complete_dates = credit_yields.dropna(subset=RATING_BUCKETS).index
    missing = required.difference(complete_dates)
    if missing.empty:
        return credit_yields

    log(
        "3年AA信用债数据未更新，启动Excel执行工作簿自带Wind EDB抓取："
        f"缺少{len(missing)}个交易日，最新缺失日{missing.max():%Y-%m-%d}"
    )
    refresh_3y_credit_yields_via_excel(workbook_path, missing.max())
    refreshed = load_3y_credit_yields(workbook_path)
    still_missing = required.difference(refreshed.dropna(subset=RATING_BUCKETS).index)
    if not still_missing.empty:
        missing_text = ", ".join(value.strftime("%Y-%m-%d") for value in still_missing[:10])
        raise RuntimeError(f"Excel/Wind刷新后3年AA信用债仍缺少：{missing_text}")
    return refreshed


def load_11_index_and_turnover_values(workbook_path: Path) -> pd.DataFrame:
    """读取1.1中iFinD公式已缓存的转债指数和全市场成交额数值。"""
    if not workbook_path.exists():
        raise FileNotFoundError(f"未找到PA转债周度数据：{workbook_path}")
    with zipfile.ZipFile(workbook_path, "r") as archive:
        sheet_path = _sheet_xml_path(archive, "1.1指标汇总")
        sheet_root = ET.fromstring(archive.read(sheet_path))
        records: list[dict[str, object]] = []
        for row in sheet_root.findall(f".//{_qname(_SHEET_NS, 'row')}"):
            if int(row.get("r", "0")) < 14:
                continue
            cells = {
                _cell_column(cell.get("r", "")): cell
                for cell in row.findall(_qname(_SHEET_NS, "c"))
            }

            def cached_number(column: str) -> float | None:
                cell = cells.get(column)
                value_node = cell.find(_qname(_SHEET_NS, "v")) if cell is not None else None
                value = pd.to_numeric(
                    value_node.text if value_node is not None else None,
                    errors="coerce",
                )
                return float(value) if pd.notna(value) else None

            date_serial = cached_number("A")
            if date_serial is None:
                continue
            records.append(
                {
                    "日期": (
                        pd.Timestamp("1899-12-30")
                        + pd.to_timedelta(date_serial, unit="D")
                    ).normalize(),
                    "转债指数": cached_number("B"),
                    "全市场成交额": cached_number("T"),
                }
            )
    values = pd.DataFrame(records)
    if values.empty:
        raise ValueError("1.1指标汇总中没有可用的转债指数和全市场成交额缓存值")
    return values.drop_duplicates("日期", keep="last").set_index("日期").sort_index()


def load_ifind_credentials() -> tuple[str, str]:
    """读取项目统一的 iFinD 登录账号，不在日志中输出凭据。"""
    if not IFIND_CREDENTIAL_FILE.is_file():
        raise FileNotFoundError(f"未找到iFinD账号文件：{IFIND_CREDENTIAL_FILE}")
    config = ConfigParser(interpolation=None)
    config.read(IFIND_CREDENTIAL_FILE, encoding="utf-8")
    username = config.get("ifind", "username", fallback="").strip()
    password = config.get("ifind", "password", fallback="").strip()
    if not username or not password:
        raise RuntimeError("ifind账号.txt中的[ifind] username或password为空")
    return username, password


def _ifind_error_message(module: object, code: int) -> str:
    try:
        detail = module.THS_GetErrorInfo(code)
        if isinstance(detail, dict):
            return str(detail.get("errmsg", detail))
        return str(detail)
    except Exception:
        return f"状态码 {code}"


def fetch_11_index_and_turnover_values(
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
) -> pd.DataFrame:
    """用 iFinD API 获取中证转债指数及全市场成交额，结果单位与1.1一致。"""
    try:
        import iFinDPy as ifind
    except ImportError as exc:
        raise RuntimeError("未安装iFinDPy，无法自动补齐1.1的iFinD数据") from exc

    username, password = load_ifind_credentials()
    login_code = int(ifind.THS_iFinDLogin(username, password))
    if login_code not in (0, -201):
        raise RuntimeError(
            "iFinD登录失败："
            f"{_ifind_error_message(ifind, login_code)}（状态码 {login_code}）"
        )
    owns_login = login_code == 0
    try:
        result = ifind.THS_DS(
            "000832.CSI",
            "ths_close_price_index;ths_amt_nd_index",
            ";",
            "block:history",
            f"{pd.Timestamp(start_date):%Y-%m-%d}",
            f"{pd.Timestamp(end_date):%Y-%m-%d}",
        )
        error_code = getattr(result, "errorcode", None)
        if error_code not in (None, 0):
            detail = getattr(result, "errmsg", "") or _ifind_error_message(ifind, int(error_code))
            raise RuntimeError(f"iFinD获取1.1指标失败：{detail}（状态码 {error_code}）")
        raw = getattr(result, "data", None)
        required_columns = {"time", "ths_close_price_index", "ths_amt_nd_index"}
        if raw is None or raw.empty:
            raise RuntimeError("iFinD未返回1.1的转债指数和全市场成交额")
        if not required_columns.issubset(raw.columns):
            raise RuntimeError(f"iFinD返回1.1字段异常：{raw.columns.tolist()}")

        values = raw.rename(
            columns={
                "time": "日期",
                "ths_close_price_index": "转债指数",
                "ths_amt_nd_index": "全市场成交额",
            }
        )[["日期", "转债指数", "全市场成交额"]].copy()
        values["日期"] = pd.to_datetime(values["日期"], errors="coerce").dt.normalize()
        values["转债指数"] = pd.to_numeric(values["转债指数"], errors="coerce")
        values["全市场成交额"] = (
            pd.to_numeric(values["全市场成交额"], errors="coerce") / 100_000_000.0
        )
        values = values.dropna(subset=["日期", "转债指数", "全市场成交额"])
        return values.drop_duplicates("日期", keep="last").set_index("日期").sort_index()
    finally:
        if owns_login:
            try:
                ifind.THS_iFinDLogout()
            except Exception:
                pass


def ensure_11_index_and_turnover_values(
    workbook_path: Path,
    required_dates: pd.DatetimeIndex,
) -> pd.DataFrame:
    """复用工作簿历史值，并通过 iFinD 自动补齐1.1缺失交易日。"""
    cached = load_11_index_and_turnover_values(workbook_path)
    required = pd.DatetimeIndex(required_dates).normalize().unique().sort_values()
    complete_cached = cached.dropna(subset=["转债指数", "全市场成交额"])
    missing = required.difference(complete_cached.index)
    if missing.empty:
        return cached

    log(
        "1.1指标汇总缺少iFinD数值，自动获取："
        f"{missing.min():%Y-%m-%d} 至 {missing.max():%Y-%m-%d}，"
        f"共{len(missing)}个交易日"
    )
    fetched = fetch_11_index_and_turnover_values(missing.min(), missing.max())
    fetched = fetched.loc[fetched.index.intersection(missing)]
    combined = pd.concat([cached, fetched]).sort_index()
    combined = combined[~combined.index.duplicated(keep="last")]
    still_missing = missing.difference(
        combined.dropna(subset=["转债指数", "全市场成交额"]).index
    )
    if not still_missing.empty:
        missing_text = ", ".join(value.strftime("%Y-%m-%d") for value in still_missing[:10])
        raise RuntimeError(f"iFinD自动获取后1.1指标仍缺少：{missing_text}")
    log(f"1.1指标汇总iFinD数值已补齐：{len(fetched)}个交易日")
    return combined


def load_master() -> tuple[pd.Series, pd.Series, pd.Series, pd.Series]:
    master_path = PARQUET_ROOT / "_special" / "总表.parquet"
    master = pd.read_parquet(master_path).set_index(BOND_CODE)
    listing = pd.to_datetime(master["上市日期"], errors="coerce")
    last_trade = pd.to_datetime(master["最后交易日"], errors="coerce")
    industry = master["申万行业"].replace("", np.nan)
    bond_names = master["转债名称"].replace("", np.nan)
    return listing, last_trade, industry, bond_names


def date_columns_from_file(file: Path) -> list[str]:
    dates = pd.read_parquet(file, columns=[TRADE_DATE])[TRADE_DATE]
    return sorted(pd.to_datetime(dates).drop_duplicates().tolist())


def build_trading_calendar(files: list[Path]) -> pd.DatetimeIndex:
    dates: list[str] = []
    for file in files:
        dates.extend(date_columns_from_file(file))
    return pd.DatetimeIndex(sorted(set(pd.to_datetime(dates))))


def build_inclusion_dates(listing: pd.Series, trading_calendar: pd.DatetimeIndex) -> pd.Series:
    """上市前 10 个交易日不纳入风格指数，第 11 个交易日起纳入。"""
    out = pd.Series(pd.Timestamp.min, index=listing.index, dtype="datetime64[ns]")
    valid_listing = listing.dropna()
    positions = trading_calendar.searchsorted(pd.DatetimeIndex(valid_listing), side="left")
    cutoff = []
    for pos in positions:
        cutoff.append(trading_calendar[pos + 10] if pos + 10 < len(trading_calendar) else pd.NaT)
    out.loc[valid_listing.index] = pd.to_datetime(cutoff)
    return out


def load_month_blocks(file: Path) -> tuple[dict[str, pd.DataFrame], pd.DataFrame, list[str]]:
    raw = pd.read_parquet(file)
    raw[TRADE_DATE] = pd.to_datetime(raw[TRADE_DATE])
    date_cols = sorted(raw[TRADE_DATE].drop_duplicates().tolist())
    blocks: dict[str, pd.DataFrame] = {}
    for name in METRIC_NAMES:
        block = raw.pivot(index=BOND_CODE, columns=TRADE_DATE, values=name).reindex(columns=date_cols)
        blocks[name] = block if name in TEXT_METRICS else block.apply(pd.to_numeric, errors="coerce")
    index_long = pd.read_parquet(PARQUET_ROOT / "_special" / "指数.parquet")
    index_long[TRADE_DATE] = pd.to_datetime(index_long[TRADE_DATE])
    index_long = index_long.loc[index_long[TRADE_DATE].isin(date_cols)]
    index_block = index_long.pivot(index=INDEX_NAME, columns=TRADE_DATE, values=INDEX_VALUE).reindex(columns=date_cols)
    return blocks, index_block, date_cols


def active_ids_for_date(
    all_ids: pd.Index,
    dt: pd.Timestamp,
    listing: pd.Series,
    last_trade: pd.Series,
    values: dict[str, pd.Series],
) -> pd.Index:
    list_s = listing.reindex(all_ids)
    last_s = last_trade.reindex(all_ids)
    start_ok = ((list_s - pd.Timedelta(days=4)) <= dt) | values["成交额"].notna()
    end_ok = (last_s + BDay(1)) >= dt
    stale_nontrading = values["成交额"].isna() & values["剩余期限"].notna()
    active = start_ok & end_ok & ~stale_nontrading & (values["余额"] > 0) & values["收盘价"].notna()
    return all_ids[active.fillna(False)]


def calculate_weekly_bond_movers(
    listing: pd.Series,
    last_trade: pd.Series,
    bond_names: pd.Series,
    trading_calendar: pd.DatetimeIndex,
) -> pd.DataFrame:
    """计算最新交易日所在自然周的个券复合涨跌幅前20名和后20名。"""
    calendar = pd.DatetimeIndex(trading_calendar).normalize().unique().sort_values()
    calendar = calendar[calendar <= LATEST_TRADE_DATE]
    if calendar.empty:
        raise RuntimeError("缺少交易日历，无法计算1.11转债周度涨跌幅个券")
    latest_date = pd.Timestamp(calendar.max()).normalize()
    week_start = latest_date - pd.Timedelta(days=latest_date.weekday())
    week_dates = calendar[(calendar >= week_start) & (calendar <= latest_date)]
    if week_dates.empty:
        raise RuntimeError("最新自然周没有交易日，无法计算1.11转债周度涨跌幅个券")

    month_files = sorted(
        {
            PARQUET_ROOT / f"{date:%Y}" / f"{date:%Y%m}.parquet"
            for date in week_dates
        }
    )
    missing_files = [str(path) for path in month_files if not path.is_file()]
    if missing_files:
        raise FileNotFoundError(f"计算1.11缺少月度parquet：{', '.join(missing_files)}")
    columns = [BOND_CODE, TRADE_DATE, "涨跌幅", "成交额", "剩余期限", "余额", "收盘价"]
    raw = pd.concat(
        [pd.read_parquet(path, columns=columns) for path in month_files],
        ignore_index=True,
    )
    raw[TRADE_DATE] = pd.to_datetime(raw[TRADE_DATE], errors="coerce").dt.normalize()
    raw = raw.loc[raw[TRADE_DATE].isin(week_dates)].copy()
    if raw.empty:
        raise RuntimeError("最新自然周parquet没有个券数据，无法计算1.11")

    returns = raw.pivot(index=BOND_CODE, columns=TRADE_DATE, values="涨跌幅")
    returns = returns.reindex(columns=week_dates).apply(pd.to_numeric, errors="coerce")
    complete = returns.notna().sum(axis=1) == len(week_dates)
    weekly_returns = (1.0 + returns / 100.0).prod(axis=1) - 1.0
    weekly_returns = weekly_returns.where(complete).replace([np.inf, -np.inf], np.nan)

    latest_rows = (
        raw.loc[raw[TRADE_DATE] == latest_date]
        .drop_duplicates(BOND_CODE, keep="last")
        .set_index(BOND_CODE)
    )
    all_ids = weekly_returns.index
    latest_values = {
        name: pd.to_numeric(latest_rows[name], errors="coerce").reindex(all_ids)
        for name in ("成交额", "剩余期限", "余额", "收盘价")
    }
    active_ids = active_ids_for_date(
        all_ids,
        latest_date,
        listing,
        last_trade,
        latest_values,
    )
    ranked = pd.DataFrame(
        {
            "转债代码": active_ids.astype(str),
            "转债名称": bond_names.reindex(active_ids).to_numpy(),
            "周涨跌幅": weekly_returns.reindex(active_ids).to_numpy(),
        }
    ).dropna(subset=["转债名称", "周涨跌幅"])
    if len(ranked) < WEEKLY_MOVER_COUNT * 2:
        raise RuntimeError(
            f"1.11周度涨跌幅有效个券仅{len(ranked)}只，不足前后各{WEEKLY_MOVER_COUNT}只"
        )

    top = (
        ranked.sort_values(
            ["周涨跌幅", "转债代码"],
            ascending=[False, True],
            kind="mergesort",
        )
        .head(WEEKLY_MOVER_COUNT)
        .assign(分组="前20名")
    )
    bottom = (
        ranked.sort_values(
            ["周涨跌幅", "转债代码"],
            ascending=[True, True],
            kind="mergesort",
        )
        .head(WEEKLY_MOVER_COUNT)
        .assign(分组="后20名")
    )
    return pd.concat([top, bottom], ignore_index=True)[
        ["分组", "转债代码", "转债名称", "周涨跌幅"]
    ]


def calculate_redemption_trigger_candidates(
    listing: pd.Series,
    last_trade: pd.Series,
    bond_names: pd.Series,
) -> pd.DataFrame:
    """筛选最新交易日赎回累计天数大于0的存续转债。"""
    latest_file = PARQUET_ROOT / f"{LATEST_TRADE_DATE:%Y}" / f"{LATEST_TRADE_DATE:%Y%m}.parquet"
    columns = [BOND_CODE, TRADE_DATE, "赎回累计天数", "余额", "收盘价"]
    raw = pd.read_parquet(latest_file, columns=columns)
    raw[TRADE_DATE] = pd.to_datetime(raw[TRADE_DATE], errors="coerce").dt.normalize()
    latest = (
        raw.loc[raw[TRADE_DATE] == LATEST_TRADE_DATE.normalize()]
        .drop_duplicates(BOND_CODE, keep="last")
        .set_index(BOND_CODE)
    )
    cumulative_days = pd.to_numeric(latest["赎回累计天数"], errors="coerce")
    balance = pd.to_numeric(latest["余额"], errors="coerce")
    close = pd.to_numeric(latest["收盘价"], errors="coerce")
    listed = listing.reindex(latest.index).le(LATEST_TRADE_DATE)
    not_delisted = (last_trade.reindex(latest.index) + BDay(1)).ge(LATEST_TRADE_DATE)
    mask = (
        cumulative_days.gt(0)
        & balance.gt(0)
        & close.notna()
        & listed.fillna(False)
        & not_delisted.fillna(False)
    )
    ids = latest.index[mask]
    candidates = pd.DataFrame(
        {
            "转债代码": ids.astype(str),
            "转债名称": bond_names.reindex(ids).to_numpy(),
            "累计天数": cumulative_days.reindex(ids).to_numpy(),
        }
    ).dropna(subset=["转债名称", "累计天数"])
    candidates["累计天数"] = candidates["累计天数"].astype(int)
    return candidates.sort_values(
        ["累计天数", "转债代码"],
        ascending=[False, True],
        kind="mergesort",
    ).reset_index(drop=True)


def history_valuation_row(day: str, x: pd.DataFrame) -> dict[str, object]:
    balance = x["余额"]
    price = x["收盘价"]
    plain = x["平价"]
    premium = x["转股溢价率"]
    ytm = x["YTM"]
    bond_premium = x["纯债溢价率"]
    implied_vol = x["隐含波动率"]
    maturity = x["剩余期限"]
    amount = x["成交额"]
    bond_ret = x["涨跌幅"]
    turnover = x["换手率"]
    total_amount = amount.sum(min_count=1)
    top5_amount = amount.dropna().nlargest(5).sum()
    ytm_valid = ytm.dropna()
    price_valid = price.dropna()
    below_bond_value_valid = price.notna() & x["纯债价值"].notna()
    long_maturity = (maturity >= 5.5) & (maturity <= 6.0)
    avg_turnover = safe_mean(turnover)

    return {
        "日期": day,
        "转债等权涨跌幅": safe_mean(bond_ret),
        "总成交额": float(total_amount) if pd.notna(total_amount) else None,
        "平均换手率": avg_turnover / 100.0 if pd.notna(avg_turnover) else None,
        "上涨家数": int((bond_ret > 0).sum()),
        "下跌家数": int((bond_ret < 0).sum()),
        "前五成交额占总成交额比例": (
            float(top5_amount / total_amount) if pd.notna(total_amount) and total_amount > 0 else None
        ),
        "转股溢价率中位数": safe_median(premium),
        "YTM中位数": safe_median(ytm),
        "纯债溢价率中位数": safe_median(bond_premium),
        "隐含波动率中位数": safe_median(implied_vol),
        "剩余期限5.5-6年转债隐含波动率中位数": safe_median(implied_vol[long_maturity]),
        "70平价溢价率": fit_premium_at_x(plain, premium, turnover, 70.0),
        "百元溢价率": fit_premium_at_x(plain, premium, turnover, 100.0),
        "120平价溢价率": fit_premium_at_x(plain, premium, turnover, 120.0),
        "收盘价中位数": safe_median(price),
        "双低中位数": safe_median(price + premium),
        "YTM>0占比": float((ytm_valid > 0).sum() / len(ytm_valid)) if len(ytm_valid) else None,
        "破面率": float((price_valid < 100).sum() / len(price_valid)) if len(price_valid) else None,
        "跌破债底占比": (
            float((price[below_bond_value_valid] < x.loc[below_bond_value_valid, "纯债价值"]).sum() / below_bond_value_valid.sum())
            if below_bond_value_valid.sum()
            else None
        ),
    }


def style_daily_returns(
    x: pd.DataFrame,
    style_ids: pd.Index,
    bond_ret: pd.Series,
    stock_ret: pd.Series,
) -> dict[str, float | None]:
    balance = x.loc[style_ids, "余额"]
    rating = x.loc[style_ids, "债项评级"].map(normalize_rating)

    groups = {
        "large": balance >= 50,
        "mid": (balance >= 10) & (balance < 50),
        "small": (balance >= 0.3) & (balance < 10),
        "AAA": rating.eq("AAA"),
        "AA+": rating.eq("AA+"),
        "AA": rating.eq("AA"),
        "AA-及以下": rating.notna() & ~rating.isin(["AAA", "AA+", "AA"]),
    }

    out: dict[str, float | None] = {}
    out["转债等权指数"] = equal_return(bond_ret.loc[style_ids])
    out["转债加权指数"] = weighted_return(bond_ret.loc[style_ids], balance)
    out["正股等权指数"] = equal_return(stock_ret.loc[style_ids])
    out["正股加权指数"] = weighted_return(stock_ret.loc[style_ids], balance)

    for group_key, bond_name, stock_name in [
        ("large", "转债大盘指数", "正股大盘指数"),
        ("mid", "转债中盘指数", "正股中盘指数"),
        ("small", "转债小盘指数", "正股小盘指数"),
        ("AAA", "转债AAA指数", "正股AAA指数"),
        ("AA+", "转债AA+指数", "正股AA+指数"),
        ("AA", "转债AA指数", "正股AA指数"),
        ("AA-及以下", "转债AA-及以下指数", "正股AA-及以下指数"),
    ]:
        ids = style_ids[groups[group_key].fillna(False)]
        out[bond_name] = weighted_return(bond_ret.loc[ids], balance.loc[ids])
        out[stock_name] = weighted_return(stock_ret.loc[ids], balance.loc[ids])
    return out


def weighted_decomp_group(sample: pd.DataFrame, label: object | None = None) -> dict[str, object] | None:
    valid = sample["期初余额"].notna() & (sample["期初余额"] > 0)
    if not valid.any():
        return None
    s = sample.loc[valid]
    weights = s["期初余额"].astype(float)
    row: dict[str, object] = {}
    if label is not None:
        row["分类"] = label
    for col in ["转债收益率", "债券贡献", "正股贡献", "估值贡献"]:
        values = pd.to_numeric(s[col], errors="coerce")
        mask = values.notna() & weights.notna() & (weights > 0)
        row[col] = float((values[mask] * weights[mask]).sum() / weights[mask].sum()) if mask.any() else None
    return row


def build_return_decomposition_outputs(observations: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if observations.empty:
        empty_ts = pd.DataFrame(columns=DECOMP_COLUMNS)
        empty_group = pd.DataFrame(columns=DECOMP_GROUP_COLUMNS)
        return empty_ts, empty_group, empty_group.copy(), empty_group.copy()

    observations = observations.sort_values(["转债代码", "日期"]).reset_index(drop=True)
    market_dates = pd.DatetimeIndex(sorted(pd.to_datetime(observations["日期"]).dropna().unique()))
    rows: list[dict[str, object]] = []
    factor_cols = ["纯债价值变动", "平价变动", "转股溢价率变动"]

    for code, group in observations.groupby("转债代码", sort=False):
        g = group.sort_values("日期").drop_duplicates("日期", keep="last").set_index("日期")
        g = g.reindex(market_dates)
        n = len(g)
        if n < ROLLING_WINDOW:
            continue

        x = g[factor_cols].astype(float).to_numpy()
        y = g["转债日收益率"].astype(float).to_numpy()
        price = g["收盘价"].astype(float).to_numpy()
        prev_price = g["前收盘价"].astype(float).to_numpy()
        row_valid = np.isfinite(y) & np.all(np.isfinite(x), axis=1)

        x_filled = np.where(np.isfinite(x), x, 0.0)
        y_filled = np.where(np.isfinite(y), y, 0.0)
        xtx_terms = np.einsum("ni,nj->nij", x_filled, x_filled)
        xty_terms = x_filled * y_filled[:, None]
        xsum_terms = x_filled
        c_xtx = np.concatenate([np.zeros((1, 3, 3)), np.cumsum(xtx_terms, axis=0)], axis=0)
        c_xty = np.concatenate([np.zeros((1, 3)), np.cumsum(xty_terms, axis=0)], axis=0)
        c_xsum = np.concatenate([np.zeros((1, 3)), np.cumsum(xsum_terms, axis=0)], axis=0)
        c_valid = np.concatenate([[0], np.cumsum(row_valid.astype(int))])

        for end_pos in range(ROLLING_WINDOW - 1, n):
            start_pos = end_pos - ROLLING_WINDOW + 1
            end_dt = pd.Timestamp(market_dates[end_pos])
            if end_dt < DECOMP_START_DATE:
                continue
            if c_valid[end_pos + 1] - c_valid[start_pos] < ROLLING_WINDOW:
                continue

            xtx = c_xtx[end_pos + 1] - c_xtx[start_pos]
            xty = c_xty[end_pos + 1] - c_xty[start_pos]
            xsum = c_xsum[end_pos + 1] - c_xsum[start_pos]
            try:
                beta = np.linalg.solve(xtx, xty)
            except np.linalg.LinAlgError:
                beta = np.linalg.lstsq(x[start_pos : end_pos + 1], y[start_pos : end_pos + 1], rcond=None)[0]

            contribution = beta * xsum
            start_price = prev_price[start_pos]
            end_price = price[end_pos]
            start_balance = g["期初余额"].iloc[start_pos]
            if (
                not np.isfinite(start_price)
                or start_price <= 0
                or not np.isfinite(end_price)
                or pd.isna(start_balance)
                or float(start_balance) <= 0
            ):
                continue

            rows.append(
                {
                    "日期": end_dt,
                    "转债代码": code,
                    "转债收益率": float(end_price / start_price - 1.0),
                    "债券贡献": float(contribution[0]),
                    "正股贡献": float(contribution[1]),
                    "估值贡献": float(contribution[2]),
                    "期初余额": float(start_balance),
                    "申万行业": g["申万行业"].iloc[start_pos],
                    "期初平价": g["期初平价"].iloc[start_pos],
                    "期初平价底价溢价率": g["期初平价底价溢价率"].iloc[start_pos],
                }
            )

    decomp = pd.DataFrame(rows)
    if decomp.empty:
        empty_ts = pd.DataFrame(columns=DECOMP_COLUMNS)
        empty_group = pd.DataFrame(columns=DECOMP_GROUP_COLUMNS)
        return empty_ts, empty_group, empty_group.copy(), empty_group.copy()

    ts_rows: list[dict[str, object]] = []
    for day, sample in decomp.groupby("日期", sort=True):
        row = weighted_decomp_group(sample)
        if row is None:
            continue
        row["日期"] = day
        ts_rows.append(row)
    timeseries = pd.DataFrame(ts_rows)[DECOMP_COLUMNS].sort_values("日期", ascending=False)

    latest_date = decomp["日期"].max()
    latest = decomp[decomp["日期"].eq(latest_date)].copy()

    industry_rows: list[dict[str, object]] = []
    for industry, sample in latest.dropna(subset=["申万行业"]).groupby("申万行业", sort=False):
        row = weighted_decomp_group(sample, industry)
        if row is not None:
            industry_rows.append(row)
    industry = pd.DataFrame(industry_rows, columns=DECOMP_GROUP_COLUMNS)
    if not industry.empty:
        industry = industry.sort_values("转债收益率", ascending=False)

    parity_bins = [
        ("100-110", (latest["期初平价"] > 100) & (latest["期初平价"] <= 110)),
        ("110-120", (latest["期初平价"] > 110) & (latest["期初平价"] <= 120)),
        ("120-130", (latest["期初平价"] > 120) & (latest["期初平价"] <= 130)),
        ("130以上", latest["期初平价"] > 130),
    ]
    parity_rows: list[dict[str, object]] = []
    for label, mask in parity_bins:
        row = weighted_decomp_group(latest[mask.fillna(False)], label)
        parity_rows.append(row if row is not None else {"分类": label, **{col: None for col in DECOMP_GROUP_COLUMNS[1:]}})
    parity = pd.DataFrame(parity_rows, columns=DECOMP_GROUP_COLUMNS)

    floor = latest["期初平价底价溢价率"]
    type_bins = [
        ("偏股型", floor > 20),
        ("平衡型", (floor > -20) & (floor < 20)),
        ("偏债型", floor < -20),
    ]
    type_rows: list[dict[str, object]] = []
    for label, mask in type_bins:
        row = weighted_decomp_group(latest[mask.fillna(False)], label)
        type_rows.append(row if row is not None else {"分类": label, **{col: None for col in DECOMP_GROUP_COLUMNS[1:]}})
    bond_type = pd.DataFrame(type_rows, columns=DECOMP_GROUP_COLUMNS)

    return timeseries, industry, parity, bond_type


def market_indicator_row(
    day: object,
    sample: pd.DataFrame,
    index_value: float | None,
    market_turnover: float | None,
    aa_credit_yield: float | None,
) -> dict[str, object]:
    """按1.1统一口径计算单日市场指标。"""
    balance = sample["余额"]
    price = sample["收盘价"]
    plain = sample["平价"]
    premium = sample["转股溢价率"]
    implied_vol = sample["隐含波动率"]
    stock_vol = sample["正股20日波动率"]
    ytm = sample["YTM"]
    ytm_valid = ytm.dropna()
    ytm_above_aa_ratio = (
        float((ytm_valid > float(aa_credit_yield)).sum() / len(ytm_valid))
        if len(ytm_valid) and pd.notna(aa_credit_yield)
        else None
    )
    return {
        "日期": day,
        "转债指数": index_value,
        "价格中位数": safe_median(price),
        "价格均值": safe_mean(price),
        "价格余额加权": weighted_mean(price, balance),
        "百元拟合溢价率": fit_premium_at_x(plain, premium, sample["换手率"], 100.0),
        "隐含波动率均值": safe_mean(implied_vol),
        "平价中位数": safe_median(plain),
        "转股溢价率中位数": safe_median(premium),
        "转股溢价率均值": safe_mean(premium),
        "纯债价值中位数": safe_median(sample["纯债价值"]),
        "纯债溢价率中位数": safe_median(sample["纯债溢价率"]),
        "纯债溢价率均值": safe_mean(sample["纯债溢价率"]),
        "YTM中位数": safe_median(ytm),
        "YTM均值": safe_mean(ytm),
        "YTM大于0的比例": (
            float((ytm_valid > 0).sum() / len(ytm_valid)) if len(ytm_valid) else None
        ),
        "YTM大于3年AA信用债比例(%)": ytm_above_aa_ratio,
        "全市场成交额": market_turnover,
        "正股市场波动率": safe_mean(stock_vol),
        "转债市场隐含波动率中位数": safe_median(implied_vol),
        "隐波差中位数": safe_median((implied_vol - stock_vol).dropna()),
    }


def parity_valuation_row(day: object, sample: pd.DataFrame) -> dict[str, object]:
    """计算各平价区间的转股溢价率及隐含波动率算术均值。"""
    plain = pd.to_numeric(sample["平价"], errors="coerce")
    premium = pd.to_numeric(sample["转股溢价率"], errors="coerce").dropna()
    implied_volatility = pd.to_numeric(sample["隐含波动率"], errors="coerce").dropna()
    row: dict[str, object] = {
        "日期": day,
        "整体转股溢价率均值": safe_mean(premium),
        "整体隐含波动率均值": safe_mean(implied_volatility),
        "整体转股溢价率有效个数": int(len(premium)),
        "整体隐含波动率有效个数": int(len(implied_volatility)),
    }
    for label, lower, upper in PARITY_BUCKETS:
        if lower is None:
            mask = plain <= float(upper)
        elif upper is None:
            mask = plain > float(lower)
        else:
            mask = (plain > float(lower)) & (plain <= float(upper))
        bucket = sample.loc[mask.fillna(False)]
        row[f"转股溢价率_{label}"] = safe_mean(bucket["转股溢价率"])
        row[f"隐含波动率_{label}"] = safe_mean(bucket["隐含波动率"])
    return row


def term_valuation_row(day: object, sample: pd.DataFrame) -> dict[str, object]:
    """计算各剩余期限区间的转股溢价率算术均值及个券数量。"""
    maturity = pd.to_numeric(sample["剩余期限"], errors="coerce")
    premium = pd.to_numeric(sample["转股溢价率"], errors="coerce")
    labels = [label for label, _, _ in TERM_BUCKETS]
    groups = pd.cut(
        maturity,
        bins=[bucket[1] for bucket in TERM_BUCKETS] + [TERM_BUCKETS[-1][2]],
        labels=labels,
        right=False,
        include_lowest=True,
    )
    means = premium.groupby(groups, observed=False).mean()
    counts = groups.value_counts(sort=False)
    row: dict[str, object] = {"日期": day}
    for label in labels:
        mean_value = means.get(label, np.nan)
        row[f"转股溢价率_{label}"] = float(mean_value) if pd.notna(mean_value) else None
        row[f"个券数量_{label}"] = int(counts.get(label, 0))
    return row


def rating_valuation_row(
    day: object,
    sample: pd.DataFrame,
    credit_yields: pd.Series | None,
) -> dict[str, object]:
    """计算各债项评级的估值均值及YTM相对同评级3年信用债利差。"""
    ratings = sample["债项评级"].map(normalize_rating)
    premium = pd.to_numeric(sample["转股溢价率"], errors="coerce")
    implied_volatility = pd.to_numeric(sample["隐含波动率"], errors="coerce")
    ytm = pd.to_numeric(sample["YTM"], errors="coerce")
    premium_means = premium.groupby(ratings).mean()
    volatility_means = implied_volatility.groupby(ratings).mean()
    ytm_means = ytm.groupby(ratings).mean()
    counts = ratings.value_counts()
    row: dict[str, object] = {"日期": day}
    for rating in RATING_BUCKETS:
        premium_value = premium_means.get(rating, np.nan)
        volatility_value = volatility_means.get(rating, np.nan)
        ytm_value = ytm_means.get(rating, np.nan)
        credit_value = (
            pd.to_numeric(credit_yields.get(rating), errors="coerce")
            if credit_yields is not None
            else np.nan
        )
        row[f"转股溢价率_{rating}"] = float(premium_value) if pd.notna(premium_value) else None
        row[f"隐含波动率_{rating}"] = float(volatility_value) if pd.notna(volatility_value) else None
        row[f"YTM_{rating}"] = float(ytm_value) if pd.notna(ytm_value) else None
        row[f"YTM信用利差_{rating}"] = (
            float(ytm_value - credit_value)
            if pd.notna(ytm_value) and pd.notna(credit_value)
            else None
        )
        row[f"个券数量_{rating}"] = int(counts.get(rating, 0))
    return row


def size_valuation_row(day: object, sample: pd.DataFrame) -> dict[str, object]:
    """计算各余额规模区间的转股溢价率及隐含波动率算术均值。"""
    balance = pd.to_numeric(sample["余额"], errors="coerce")
    labels = [label for label, _, _ in SIZE_BUCKETS]
    groups = pd.cut(
        balance,
        bins=[0.0, 5.0, 10.0, 20.0, 50.0, np.inf],
        labels=labels,
        right=False,
        include_lowest=True,
    )
    premium = pd.to_numeric(sample["转股溢价率"], errors="coerce")
    implied_volatility = pd.to_numeric(sample["隐含波动率"], errors="coerce")
    premium_means = premium.groupby(groups, observed=False).mean()
    volatility_means = implied_volatility.groupby(groups, observed=False).mean()
    counts = groups.value_counts(sort=False)
    row: dict[str, object] = {"日期": day}
    for label in labels:
        premium_value = premium_means.get(label, np.nan)
        volatility_value = volatility_means.get(label, np.nan)
        row[f"转股溢价率_{label}"] = float(premium_value) if pd.notna(premium_value) else None
        row[f"隐含波动率_{label}"] = float(volatility_value) if pd.notna(volatility_value) else None
        row[f"个券数量_{label}"] = int(counts.get(label, 0))
    return row


def seasoning_valuation_row(day: object, sample: pd.DataFrame) -> dict[str, object]:
    """计算全市场、次新券及老券估值；次新券定义为剩余期限[5.5, 6)。"""
    maturity = pd.to_numeric(sample["剩余期限"], errors="coerce")
    implied_volatility = pd.to_numeric(sample["隐含波动率"], errors="coerce")
    premium = pd.to_numeric(sample["转股溢价率"], errors="coerce")
    new_mask = (maturity >= 5.5) & (maturity < 6.0)
    old_mask = (maturity >= 0.0) & (maturity < 5.5)
    return {
        "日期": day,
        "全部转债隐含波动率均值": safe_mean(implied_volatility),
        "次新券隐含波动率均值": safe_mean(implied_volatility[new_mask.fillna(False)]),
        "次新券转股溢价率均值": safe_mean(premium[new_mask.fillna(False)]),
        "老券隐含波动率均值": safe_mean(implied_volatility[old_mask.fillna(False)]),
    }


def build_major_valuation_metrics(market: pd.DataFrame) -> pd.DataFrame:
    """生成主要估值指标Sheet所需的250交易日滚动统计。"""
    ordered = market.sort_values("日期").reset_index(drop=True)
    result = pd.DataFrame({"日期": ordered["日期"]})
    for source_column, output_prefix in (
        ("百元拟合溢价率", "百元拟合溢价率"),
        ("隐含波动率均值", "隐含波动率"),
        ("正股市场波动率", "正股波动率"),
        ("隐波差中位数", "隐波差"),
    ):
        values = pd.to_numeric(ordered[source_column], errors="coerce") / 100.0
        rolling = values.rolling(window=250, min_periods=1)
        rolling_mean = rolling.mean()
        rolling_std = rolling.std(ddof=1)
        result[output_prefix] = values
        result[f"{output_prefix}_250日均值"] = rolling_mean
        result[f"{output_prefix}_250标准差"] = rolling_std
        result[f"{output_prefix}_+1倍标准差"] = rolling_mean + rolling_std
        result[f"{output_prefix}_+2倍标准差"] = rolling_mean + 2.0 * rolling_std
        result[f"{output_prefix}_-1倍标准差"] = rolling_mean - rolling_std
        result[f"{output_prefix}_-2倍标准差"] = rolling_mean - 2.0 * rolling_std
    return result.sort_values("日期", ascending=False).reset_index(drop=True)


def calculate_all() -> dict[str, pd.DataFrame]:
    listing, last_trade, industry, bond_names = load_master()
    all_files = sorted(PARQUET_ROOT.glob("20??/20????.parquet"))
    trading_calendar = build_trading_calendar(all_files)
    required_11_dates = trading_calendar[
        (trading_calendar >= START_DATE) & (trading_calendar <= LATEST_TRADE_DATE)
    ]
    credit_yields = ensure_3y_credit_yields_current(
        PA_WORKBOOK,
        required_11_dates,
    )
    aa_credit_yields = credit_yields["AA"]
    index_and_turnover = ensure_11_index_and_turnover_values(
        PA_WORKBOOK,
        required_11_dates,
    )
    weekly_bond_movers = calculate_weekly_bond_movers(
        listing,
        last_trade,
        bond_names,
        trading_calendar,
    )
    redemption_trigger_candidates = calculate_redemption_trigger_candidates(
        listing,
        last_trade,
        bond_names,
    )
    inclusion_dates = build_inclusion_dates(listing, trading_calendar)
    files = [file for file in all_files if int(file.stem[:4]) >= START_DATE.year - 1]

    market_rows: list[dict[str, object]] = []
    market_trimmed_rows: list[dict[str, object]] = []
    style_rows: list[dict[str, object]] = []
    history_valuation_rows: list[dict[str, object]] = []
    parity_valuation_rows: list[dict[str, object]] = []
    term_valuation_rows: list[dict[str, object]] = []
    rating_valuation_rows: list[dict[str, object]] = []
    size_valuation_rows: list[dict[str, object]] = []
    seasoning_valuation_rows: list[dict[str, object]] = []
    decomp_observation_rows: list[dict[str, object]] = []

    style_levels = {col: 100.0 for col in STYLE_COLUMNS if col != "日期"}
    prev_stock_close = pd.Series(dtype=float)
    prev_decomp_values = {
        "收盘价": pd.Series(dtype=float),
        "纯债价值": pd.Series(dtype=float),
        "平价": pd.Series(dtype=float),
        "转股溢价率": pd.Series(dtype=float),
        "余额": pd.Series(dtype=float),
        "平价底价溢价率": pd.Series(dtype=float),
    }

    started = time.perf_counter()
    for file_no, file in enumerate(files, 1):
        blocks, index_block, date_cols = load_month_blocks(file)
        all_ids = blocks["收盘价"].index
        for day in date_cols:
            dt = pd.Timestamp(day)
            values = {name: blocks[name][day] for name in METRIC_NAMES}
            ids = active_ids_for_date(all_ids, dt, listing, last_trade, values)
            if len(ids) == 0:
                prev_stock_close = combine_current_with_previous(values["正股收盘价"], prev_stock_close)
                prev_stock_close.update(values["正股收盘价"].dropna())
                for name in prev_decomp_values:
                    prev_decomp_values[name] = combine_current_with_previous(values[name], prev_decomp_values[name])
                    prev_decomp_values[name].update(values[name].dropna())
                continue

            x = pd.DataFrame({name: values[name].reindex(ids) for name in METRIC_NAMES})
            balance = x["余额"]
            price = x["收盘价"]
            plain = x["平价"]
            premium = x["转股溢价率"]
            implied_vol = x["隐含波动率"]
            stock_vol = x["正股20日波动率"]
            ytm = x["YTM"]
            amount = x["成交额"]
            bond_ret = x["涨跌幅"] / 100.0
            stock_prev = prev_stock_close.reindex(ids)
            stock_ret = x["正股收盘价"] / stock_prev - 1.0

            prev_price = prev_decomp_values["收盘价"].reindex(ids)
            prev_bond_value = prev_decomp_values["纯债价值"].reindex(ids)
            prev_plain = prev_decomp_values["平价"].reindex(ids)
            prev_premium = prev_decomp_values["转股溢价率"].reindex(ids)
            prev_balance = prev_decomp_values["余额"].reindex(ids)
            prev_floor_premium = prev_decomp_values["平价底价溢价率"].reindex(ids)
            decomp_valid = (
                bond_ret.notna()
                & price.notna()
                & prev_price.notna()
                & (prev_price > 0)
                & x["纯债价值"].notna()
                & prev_bond_value.notna()
                & (prev_bond_value > 0)
                & plain.notna()
                & prev_plain.notna()
                & (prev_plain > 0)
                & premium.notna()
                & prev_premium.notna()
                & prev_balance.notna()
                & (prev_balance > 0)
            )
            if decomp_valid.any():
                dids = ids[decomp_valid.fillna(False)]
                obs = pd.DataFrame(index=dids)
                obs["日期"] = dt
                obs["转债代码"] = obs.index.astype(str)
                obs["转债日收益率"] = bond_ret.reindex(dids).astype(float)
                obs["纯债价值变动"] = x.loc[dids, "纯债价值"].astype(float) / prev_bond_value.reindex(dids).astype(float) - 1.0
                obs["平价变动"] = plain.reindex(dids).astype(float) / prev_plain.reindex(dids).astype(float) - 1.0
                obs["转股溢价率变动"] = (premium.reindex(dids).astype(float) - prev_premium.reindex(dids).astype(float)) / 100.0
                obs["收盘价"] = price.reindex(dids).astype(float)
                obs["前收盘价"] = prev_price.reindex(dids).astype(float)
                obs["期初余额"] = prev_balance.reindex(dids).astype(float)
                obs["申万行业"] = industry.reindex(dids)
                obs["期初平价"] = prev_plain.reindex(dids).astype(float)
                obs["期初平价底价溢价率"] = prev_floor_premium.reindex(dids)
                obs = obs.replace([np.inf, -np.inf], np.nan).dropna(
                    subset=["转债日收益率", "纯债价值变动", "平价变动", "转股溢价率变动", "收盘价", "前收盘价", "期初余额"]
                )
                if not obs.empty:
                    decomp_observation_rows.extend(obs.reset_index(drop=True).to_dict("records"))

            total_amount = amount.sum(min_count=1)
            aa_credit_yield = aa_credit_yields.get(dt.normalize(), np.nan)

            if dt >= START_DATE:
                index_value = (
                    float(index_block.at["转债指数", day])
                    if "转债指数" in index_block.index
                    and pd.notna(index_block.at["转债指数", day])
                    else None
                )
                market_turnover = float(total_amount) if pd.notna(total_amount) else None
                market_rows.append(
                    market_indicator_row(
                        day,
                        x,
                        index_value,
                        market_turnover,
                        aa_credit_yield,
                    )
                )
                trimmed_mask = plain.between(60, 140, inclusive="both")
                market_trimmed_rows.append(
                    market_indicator_row(
                        day,
                        x.loc[trimmed_mask.fillna(False)],
                        index_value,
                        market_turnover,
                        aa_credit_yield,
                    )
                )
                history_valuation_rows.append(history_valuation_row(day, x))
                parity_valuation_rows.append(parity_valuation_row(day, x))
                term_valuation_rows.append(term_valuation_row(day, x))
                credit_row = (
                    credit_yields.loc[dt.normalize()]
                    if dt.normalize() in credit_yields.index
                    else None
                )
                rating_valuation_rows.append(rating_valuation_row(day, x, credit_row))
                size_valuation_rows.append(size_valuation_row(day, x))
                seasoning_valuation_rows.append(seasoning_valuation_row(day, x))

                include_dates = inclusion_dates.reindex(ids)
                style_mask = (balance >= 0.3) & (include_dates.notna()) & (dt >= include_dates)
                style_ids = ids[style_mask.fillna(False)]
                daily_returns = style_daily_returns(x, style_ids, bond_ret, stock_ret)
                style_row: dict[str, object] = {"日期": day}
                for col in STYLE_COLUMNS[1:]:
                    ret = daily_returns.get(col)
                    if ret is not None and pd.notna(ret):
                        style_levels[col] *= 1.0 + float(ret)
                    style_row[col] = style_levels[col]
                style_rows.append(style_row)

            prev_stock_close = combine_current_with_previous(values["正股收盘价"], prev_stock_close)
            prev_stock_close.update(values["正股收盘价"].dropna())
            for name in prev_decomp_values:
                prev_decomp_values[name] = combine_current_with_previous(values[name], prev_decomp_values[name])
                prev_decomp_values[name].update(values[name].dropna())

        if file_no % 24 == 0:
            log(f"processed {file_no}/{len(files)} files, rows={len(market_rows)}")

    market = pd.DataFrame(market_rows)
    market["日期"] = pd.to_datetime(market["日期"])
    market = market.sort_values("日期").drop_duplicates("日期", keep="last")
    market_dates = market["日期"].dt.normalize()
    market["转债指数"] = market_dates.map(index_and_turnover["转债指数"])
    market["全市场成交额"] = market_dates.map(index_and_turnover["全市场成交额"])
    missing_market_values = market.loc[
        market[["转债指数", "全市场成交额"]].isna().any(axis=1), "日期"
    ]
    if not missing_market_values.empty:
        missing_text = ", ".join(
            pd.Timestamp(value).strftime("%Y-%m-%d")
            for value in missing_market_values.head(10)
        )
        raise ValueError(
            "PA转债周度数据的1.1指标汇总在自动iFinD取值后仍缺少："
            f"{missing_text}。请检查iFinD返回值及交易日口径。"
        )
    market["百元拟合溢价率2017年以来分位数"] = expanding_percentile(market["百元拟合溢价率"])
    market["隐含波动率自2017年以来分位数"] = expanding_percentile(market["隐含波动率均值"])
    market = market[MARKET_COLUMNS].sort_values("日期", ascending=False)

    market_trimmed = pd.DataFrame(market_trimmed_rows)
    market_trimmed["日期"] = pd.to_datetime(market_trimmed["日期"])
    market_trimmed = market_trimmed.sort_values("日期").drop_duplicates("日期", keep="last")
    trimmed_dates = market_trimmed["日期"].dt.normalize()
    market_trimmed["转债指数"] = trimmed_dates.map(index_and_turnover["转债指数"])
    market_trimmed["全市场成交额"] = trimmed_dates.map(index_and_turnover["全市场成交额"])
    market_trimmed["百元拟合溢价率2017年以来分位数"] = expanding_percentile(
        market_trimmed["百元拟合溢价率"]
    )
    market_trimmed["隐含波动率自2017年以来分位数"] = expanding_percentile(
        market_trimmed["隐含波动率均值"]
    )
    market_trimmed = market_trimmed[MARKET_COLUMNS].sort_values("日期", ascending=False)

    style = pd.DataFrame(style_rows)
    style["日期"] = pd.to_datetime(style["日期"])
    style = style.sort_values("日期").drop_duplicates("日期", keep="last")
    style = style[STYLE_COLUMNS].sort_values("日期", ascending=False)

    history_valuation_daily = pd.DataFrame(history_valuation_rows)
    history_valuation_daily["日期"] = pd.to_datetime(history_valuation_daily["日期"])
    history_valuation_daily = (
        history_valuation_daily.sort_values("日期")
        .drop_duplicates("日期", keep="last")
        .sort_values("日期", ascending=False)
        .reset_index(drop=True)
    )
    index_returns = market[["日期", "转债指数"]].copy().sort_values("日期")
    index_returns["中证转债涨跌幅"] = (
        pd.to_numeric(index_returns["转债指数"], errors="coerce")
        .pct_change(fill_method=None)
        .mul(100.0)
    )
    history_valuation_daily = history_valuation_daily.merge(
        index_returns[["日期", "中证转债涨跌幅"]],
        on="日期",
        how="left",
    )
    history_valuation_daily = history_valuation_daily[
        HISTORY_VALUATION_DAILY_COLUMNS
    ].sort_values("日期", ascending=False).reset_index(drop=True)

    history_valuation_latest = history_valuation_daily[
        [column for column in HISTORY_VALUATION_COLUMNS if column in history_valuation_daily.columns]
    ].head(1)
    history_valuation_long = history_valuation_latest.copy()
    history_valuation_long["日期"] = history_valuation_long["日期"].dt.strftime("%Y-%m-%d")
    history_valuation_long = history_valuation_long.T.reset_index()
    history_valuation_long.columns = ["指标", "数值"]

    parity_valuation = pd.DataFrame(parity_valuation_rows)
    parity_valuation["日期"] = pd.to_datetime(parity_valuation["日期"])
    parity_valuation = (
        parity_valuation.sort_values("日期")
        .drop_duplicates("日期", keep="last")
        .sort_values("日期", ascending=False)
        .reset_index(drop=True)
    )

    term_valuation = pd.DataFrame(term_valuation_rows)
    term_valuation["日期"] = pd.to_datetime(term_valuation["日期"])
    term_valuation = (
        term_valuation.sort_values("日期")
        .drop_duplicates("日期", keep="last")
        .sort_values("日期", ascending=False)
        .reset_index(drop=True)
    )

    rating_valuation = pd.DataFrame(rating_valuation_rows)
    rating_valuation["日期"] = pd.to_datetime(rating_valuation["日期"])
    rating_valuation = (
        rating_valuation.sort_values("日期")
        .drop_duplicates("日期", keep="last")
        .sort_values("日期", ascending=False)
        .reset_index(drop=True)
    )

    size_valuation = pd.DataFrame(size_valuation_rows)
    size_valuation["日期"] = pd.to_datetime(size_valuation["日期"])
    size_valuation = (
        size_valuation.sort_values("日期")
        .drop_duplicates("日期", keep="last")
        .sort_values("日期", ascending=False)
        .reset_index(drop=True)
    )

    seasoning_valuation = pd.DataFrame(seasoning_valuation_rows)
    seasoning_valuation["日期"] = pd.to_datetime(seasoning_valuation["日期"])
    seasoning_valuation = (
        seasoning_valuation.sort_values("日期")
        .drop_duplicates("日期", keep="last")
        .sort_values("日期", ascending=False)
        .reset_index(drop=True)
    )

    decomp_observations = pd.DataFrame(decomp_observation_rows)
    decomp_timeseries, decomp_industry, decomp_parity, decomp_type = build_return_decomposition_outputs(decomp_observations)
    major_valuation_metrics = build_major_valuation_metrics(market)

    elapsed = time.perf_counter() - started
    log(f"calculation completed, rows={len(market)}, elapsed={format_elapsed(elapsed)}")

    return {
        "market": market,
        "market_trimmed": market_trimmed,
        "major_valuation_metrics": major_valuation_metrics,
        "style": style,
        "history_valuation": history_valuation_long,
        "history_valuation_daily": history_valuation_daily,
        "parity_valuation": parity_valuation,
        "term_valuation": term_valuation,
        "rating_valuation": rating_valuation,
        "size_valuation": size_valuation,
        "seasoning_valuation": seasoning_valuation,
        "decomp_timeseries": decomp_timeseries,
        "decomp_industry": decomp_industry,
        "decomp_parity": decomp_parity,
        "decomp_type": decomp_type,
        "weekly_bond_movers": weekly_bond_movers,
        "redemption_trigger_candidates": redemption_trigger_candidates,
    }


_SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def _qname(namespace: str, local_name: str) -> str:
    return f"{{{namespace}}}{local_name}"


def _sheet_xml_path(archive: zipfile.ZipFile, sheet_name: str) -> str:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    sheet = next(
        (
            node
            for node in workbook_root.findall(f".//{_qname(_SHEET_NS, 'sheet')}")
            if node.get("name") == sheet_name
        ),
        None,
    )
    if sheet is None:
        raise KeyError(f"PA转债周度数据缺少 sheet：{sheet_name}")
    relationship_id = sheet.get(_qname(_OFFICE_REL_NS, "id"))
    relationships_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relationship = next(
        (
            node
            for node in relationships_root.findall(_qname(_PACKAGE_REL_NS, "Relationship"))
            if node.get("Id") == relationship_id
        ),
        None,
    )
    if relationship is None:
        raise KeyError(f"无法解析 sheet 关系：{sheet_name}")
    target = relationship.get("Target", "").replace("\\", "/")
    if target.startswith("/"):
        return target.lstrip("/")
    while target.startswith("../"):
        target = target[3:]
    return target if target.startswith("xl/") else f"xl/{target}"


def _sheet_id(archive: zipfile.ZipFile, sheet_name: str) -> str:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    sheet = next(
        (
            node
            for node in workbook_root.findall(f".//{_qname(_SHEET_NS, 'sheet')}")
            if node.get("name") == sheet_name
        ),
        None,
    )
    if sheet is None or not sheet.get("sheetId"):
        raise KeyError(f"PA转债周度数据缺少 sheet：{sheet_name}")
    return str(sheet.get("sheetId"))


def _cell_column(cell_reference: str) -> str:
    match = re.match(r"[A-Z]+", cell_reference or "")
    return match.group(0) if match else ""


def _set_xml_cell_value(cell, value: object = None, formula: str | None = None) -> None:
    for child in list(cell):
        cell.remove(child)
    cell.attrib.pop("t", None)
    if formula is not None:
        formula_node = ET.SubElement(cell, _qname(_SHEET_NS, "f"))
        formula_node.text = formula[1:] if formula.startswith("=") else formula
        return
    if value is None or pd.isna(value):
        return
    value_node = ET.SubElement(cell, _qname(_SHEET_NS, "v"))
    if isinstance(value, (np.integer, int)):
        value_node.text = str(int(value))
    else:
        value_node.text = repr(float(value))


def _set_xml_cell_text(cell, value: str | None) -> None:
    for child in list(cell):
        cell.remove(child)
    cell.attrib.pop("t", None)
    if value is None:
        return
    cell.set("t", "inlineStr")
    inline = ET.SubElement(cell, _qname(_SHEET_NS, "is"))
    text_node = ET.SubElement(inline, _qname(_SHEET_NS, "t"))
    text_node.text = str(value)


def _patch_summary_sheet_xml(sheet_xml: bytes, market: pd.DataFrame) -> bytes:
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(_qname(_SHEET_NS, "sheetData"))
    if sheet_data is None:
        raise RuntimeError("1.1指标汇总缺少 sheetData")
    template_row = next(
        (row for row in sheet_data.findall(_qname(_SHEET_NS, "row")) if row.get("r") == "14"),
        None,
    )
    if template_row is None:
        raise RuntimeError("1.1指标汇总缺少第14行格式模板")
    style_by_column = {
        _cell_column(cell.get("r", "")): cell.get("s")
        for cell in template_row.findall(_qname(_SHEET_NS, "c"))
    }
    if style_by_column.get("R"):
        style_by_column["S"] = style_by_column["R"]

    for row in list(sheet_data.findall(_qname(_SHEET_NS, "row"))):
        if int(row.get("r", "0")) >= 14:
            sheet_data.remove(row)

    value_columns = {
        "C": "价格中位数",
        "D": "价格均值",
        "E": "价格余额加权",
        "F": "百元拟合溢价率",
        "G": "百元拟合溢价率2017年以来分位数",
        "H": "隐含波动率均值",
        "I": "隐含波动率自2017年以来分位数",
        "J": "平价中位数",
        "K": "转股溢价率中位数",
        "L": "转股溢价率均值",
        "M": "纯债价值中位数",
        "N": "纯债溢价率中位数",
        "O": "纯债溢价率均值",
        "P": "YTM中位数",
        "Q": "YTM均值",
        "R": "YTM大于0的比例",
        "S": "YTM大于3年AA信用债比例(%)",
        "U": "正股市场波动率",
        "V": "转债市场隐含波动率中位数",
        "W": "隐波差中位数",
    }
    market = market.sort_values("日期", ascending=False).reset_index(drop=True)
    for offset, (_, record) in enumerate(market.iterrows()):
        row_number = 14 + offset
        row = deepcopy(template_row)
        row.set("r", str(row_number))
        if row.get("spans") is not None:
            row.set("spans", "1:23")
        cells = {
            _cell_column(cell.get("r", "")): cell
            for cell in row.findall(_qname(_SHEET_NS, "c"))
        }
        for column_number in range(1, 24):
            column = get_column_letter(column_number)
            cell = cells.get(column)
            if cell is None:
                cell = ET.SubElement(row, _qname(_SHEET_NS, "c"))
                cells[column] = cell
            cell.set("r", f"{column}{row_number}")
            style_id = style_by_column.get(column)
            if style_id is not None:
                cell.set("s", style_id)

        date = pd.Timestamp(record["日期"]).normalize()
        excel_date = int((date - pd.Timestamp("1899-12-30")).days)
        _set_xml_cell_value(cells["A"], excel_date)
        _set_xml_cell_value(cells["B"], record.get("转债指数"))
        for column, metric in value_columns.items():
            _set_xml_cell_value(cells[column], record.get(metric))
        _set_xml_cell_value(cells["T"], record.get("全市场成交额"))
        sheet_data.append(row)

    last_row = 13 + len(market)
    dimension = root.find(_qname(_SHEET_NS, "dimension"))
    if dimension is not None:
        dimension.set("ref", f"A1:W{last_row}")
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _patch_bond_indicator_sheet_xml(sheet_xml: bytes, market: pd.DataFrame) -> bytes:
    """将1.1统一口径的债性指标历史序列写入“债性指标”Sheet。"""
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(_qname(_SHEET_NS, "sheetData"))
    if sheet_data is None:
        raise RuntimeError("债性指标缺少sheetData")
    template_row = next(
        (row for row in sheet_data.findall(_qname(_SHEET_NS, "row")) if row.get("r") == "23"),
        None,
    )
    if template_row is None:
        raise RuntimeError("债性指标缺少第23行格式模板")

    style_by_column = {
        _cell_column(cell.get("r", "")): cell.get("s")
        for cell in template_row.findall(_qname(_SHEET_NS, "c"))
    }
    # G/H均为比例，统一使用模板中G列的百分比数字格式。
    if style_by_column.get("G"):
        style_by_column["H"] = style_by_column["G"]

    for row in list(sheet_data.findall(_qname(_SHEET_NS, "row"))):
        if int(row.get("r", "0")) >= 23:
            sheet_data.remove(row)

    column_metrics = {
        "B": "纯债价值中位数",
        "C": "纯债溢价率中位数",
        "D": "纯债溢价率均值",
        "E": "YTM中位数",
        "F": "YTM均值",
        "G": "YTM大于0的比例",
        "H": "YTM大于3年AA信用债比例(%)",
    }
    market = market.sort_values("日期", ascending=False).reset_index(drop=True)
    for offset, (_, record) in enumerate(market.iterrows()):
        row_number = 23 + offset
        row = deepcopy(template_row)
        row.set("r", str(row_number))
        if row.get("spans") is not None:
            row.set("spans", "1:8")
        cells = {
            _cell_column(cell.get("r", "")): cell
            for cell in row.findall(_qname(_SHEET_NS, "c"))
        }
        for column_number in range(1, 9):
            column = get_column_letter(column_number)
            cell = cells.get(column)
            if cell is None:
                cell = ET.SubElement(row, _qname(_SHEET_NS, "c"))
                cells[column] = cell
            cell.set("r", f"{column}{row_number}")
            style_id = style_by_column.get(column)
            if style_id is not None:
                cell.set("s", style_id)

        date = pd.Timestamp(record["日期"]).normalize()
        excel_date = int((date - pd.Timestamp("1899-12-30")).days)
        _set_xml_cell_value(cells["A"], excel_date)
        for column, metric in column_metrics.items():
            _set_xml_cell_value(cells[column], record.get(metric))
        sheet_data.append(row)

    last_row = 22 + len(market)
    dimension = root.find(_qname(_SHEET_NS, "dimension"))
    if dimension is not None:
        dimension.set("ref", f"A1:H{last_row}")
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _patch_parity_valuation_sheet_xml(
    sheet_xml: bytes,
    metrics: pd.DataFrame,
) -> bytes:
    """写入1.3分平价转股溢价率和隐含波动率历史序列。"""
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(_qname(_SHEET_NS, "sheetData"))
    if sheet_data is None:
        raise RuntimeError("1.3分平价估值缺少sheetData")
    template_row = next(
        (row for row in sheet_data.findall(_qname(_SHEET_NS, "row")) if row.get("r") == "32"),
        None,
    )
    if template_row is None:
        raise RuntimeError("1.3分平价估值缺少第32行格式模板")

    template_styles = {
        _cell_column(cell.get("r", "")): cell.get("s")
        for cell in template_row.findall(_qname(_SHEET_NS, "c"))
    }
    date_style = template_styles.get("A") or template_styles.get("O")
    numeric_style = template_styles.get("D") or template_styles.get("C")
    style_by_column = dict(template_styles)
    style_by_column["A"] = date_style
    style_by_column["O"] = date_style
    for column in [
        *[get_column_letter(value) for value in range(2, 13)],
        *[get_column_letter(value) for value in range(16, 27)],
    ]:
        style_by_column[column] = numeric_style

    metrics = metrics.sort_values("日期", ascending=False).reset_index(drop=True)
    summary_formula_rows = {
        21: "={column}32",
        22: "={column}37",
        23: "={column}21-{column}22",
        24: "=_xlfn.PERCENTRANK.INC({column}32:INDIRECT(\"{column}\"&'1.1指标汇总'!$C$1),{column}32)",
        25: "=_xlfn.PERCENTRANK.INC({column}32:INDIRECT(\"{column}\"&'1.1指标汇总'!$D$1),{column}32)",
        26: "=_xlfn.PERCENTRANK.INC({column}32:{column}274,{column}32)",
        27: "=_xlfn.PERCENTRANK.INC({column}32:{column}759,{column}32)",
        28: "=_xlfn.PERCENTRANK.INC({column}32:{column}1243,{column}32)",
    }
    summary_rows = {
        int(row.get("r", "0")): row
        for row in sheet_data.findall(_qname(_SHEET_NS, "row"))
        if 21 <= int(row.get("r", "0")) <= 29
    }
    summary_columns = {
        "L": ("K", "整体转股溢价率有效个数"),
        "Z": ("Y", "整体隐含波动率有效个数"),
    }
    latest_record = metrics.iloc[0] if not metrics.empty else None
    for column, (style_source_column, count_metric) in summary_columns.items():
        for row_number in range(21, 30):
            row = summary_rows.get(row_number)
            if row is None:
                raise RuntimeError(f"1.3分平价估值缺少第{row_number}行摘要模板")
            cells = {
                _cell_column(cell.get("r", "")): cell
                for cell in row.findall(_qname(_SHEET_NS, "c"))
            }
            cell = cells.get(column)
            if cell is None:
                cell = ET.SubElement(row, _qname(_SHEET_NS, "c"))
                cells[column] = cell
            cell.set("r", f"{column}{row_number}")
            style_source = cells.get(style_source_column)
            if style_source is not None and style_source.get("s") is not None:
                cell.set("s", style_source.get("s"))
            if row_number == 29:
                count_value = None if latest_record is None else latest_record.get(count_metric)
                _set_xml_cell_value(cell, count_value)
            else:
                _set_xml_cell_value(
                    cell,
                    formula=summary_formula_rows[row_number].format(column=column),
                )
            cell_nodes = row.findall(_qname(_SHEET_NS, "c"))
            for node in cell_nodes:
                row.remove(node)
            for node in sorted(
                cell_nodes,
                key=lambda item: column_index_from_string(_cell_column(item.get("r", ""))),
            ):
                row.append(node)

    for row in list(sheet_data.findall(_qname(_SHEET_NS, "row"))):
        if int(row.get("r", "0")) >= 32:
            sheet_data.remove(row)

    premium_columns = {
        get_column_letter(2 + offset): f"转股溢价率_{label}"
        for offset, (label, _, _) in enumerate(PARITY_BUCKETS)
    }
    premium_columns["L"] = "整体转股溢价率均值"
    volatility_columns = {
        get_column_letter(16 + offset): f"隐含波动率_{label}"
        for offset, (label, _, _) in enumerate(PARITY_BUCKETS)
    }
    volatility_columns["Z"] = "整体隐含波动率均值"
    for offset, (_, record) in enumerate(metrics.iterrows()):
        row_number = 32 + offset
        row = deepcopy(template_row)
        row.set("r", str(row_number))
        if row.get("spans") is not None:
            row.set("spans", "1:26")
        cells = {
            _cell_column(cell.get("r", "")): cell
            for cell in row.findall(_qname(_SHEET_NS, "c"))
        }
        for column_number in range(1, 27):
            column = get_column_letter(column_number)
            cell = cells.get(column)
            if cell is None:
                cell = ET.SubElement(row, _qname(_SHEET_NS, "c"))
                cells[column] = cell
            cell.set("r", f"{column}{row_number}")
            style_id = style_by_column.get(column)
            if style_id is not None:
                cell.set("s", style_id)
            _set_xml_cell_value(cell, None)

        date = pd.Timestamp(record["日期"]).normalize()
        excel_date = int((date - pd.Timestamp("1899-12-30")).days)
        _set_xml_cell_value(cells["A"], excel_date)
        _set_xml_cell_value(cells["O"], excel_date)
        for column, metric in premium_columns.items():
            _set_xml_cell_value(cells[column], record.get(metric))
        for column, metric in volatility_columns.items():
            _set_xml_cell_value(cells[column], record.get(metric))

        cell_nodes = row.findall(_qname(_SHEET_NS, "c"))
        for cell in cell_nodes:
            row.remove(cell)
        for cell in sorted(
            cell_nodes,
            key=lambda node: column_index_from_string(_cell_column(node.get("r", ""))),
        ):
            row.append(cell)
        sheet_data.append(row)

    last_row = 31 + len(metrics)
    dimension = root.find(_qname(_SHEET_NS, "dimension"))
    if dimension is not None:
        dimension.set("ref", f"A1:Z{last_row}")
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _patch_term_valuation_sheet_xml(
    sheet_xml: bytes,
    metrics: pd.DataFrame,
) -> bytes:
    """写入1.4分期限估值的转股溢价率历史序列及摘要。"""
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(_qname(_SHEET_NS, "sheetData"))
    if sheet_data is None:
        raise RuntimeError("1.4分期限估值缺少sheetData")
    template_row = next(
        (row for row in sheet_data.findall(_qname(_SHEET_NS, "row")) if row.get("r") == "32"),
        None,
    )
    if template_row is None:
        raise RuntimeError("1.4分期限估值缺少第32行格式模板")

    template_styles = {
        _cell_column(cell.get("r", "")): cell.get("s")
        for cell in template_row.findall(_qname(_SHEET_NS, "c"))
    }
    date_style = template_styles.get("A")
    numeric_style = template_styles.get("D") or template_styles.get("B")
    style_by_column = {"A": date_style}
    for column_number in range(2, 14):
        style_by_column[get_column_letter(column_number)] = numeric_style

    metrics = metrics.sort_values("日期", ascending=False).reset_index(drop=True)
    summary_formula_rows = {
        21: "={column}32",
        22: "={column}37",
        23: "={column}21-{column}22",
        24: "=_xlfn.PERCENTRANK.INC({column}32:INDIRECT(\"{column}\"&'1.1指标汇总'!$C$1),{column}32)",
        25: "=_xlfn.PERCENTRANK.INC({column}32:INDIRECT(\"{column}\"&'1.1指标汇总'!$D$1),{column}32)",
        26: "=_xlfn.PERCENTRANK.INC({column}32:{column}274,{column}32)",
        27: "=_xlfn.PERCENTRANK.INC({column}32:{column}759,{column}32)",
        28: "=_xlfn.PERCENTRANK.INC({column}32:{column}1243,{column}32)",
    }
    summary_rows = {
        int(row.get("r", "0")): row
        for row in sheet_data.findall(_qname(_SHEET_NS, "row"))
        if 21 <= int(row.get("r", "0")) <= 29
    }
    latest_record = metrics.iloc[0] if not metrics.empty else None
    for column_number, (label, _, _) in enumerate(TERM_BUCKETS, start=2):
        column = get_column_letter(column_number)
        for row_number in range(21, 30):
            row = summary_rows.get(row_number)
            if row is None:
                raise RuntimeError(f"1.4分期限估值缺少第{row_number}行摘要模板")
            cells = {
                _cell_column(cell.get("r", "")): cell
                for cell in row.findall(_qname(_SHEET_NS, "c"))
            }
            cell = cells.get(column)
            if cell is None:
                cell = ET.SubElement(row, _qname(_SHEET_NS, "c"))
                cells[column] = cell
            cell.set("r", f"{column}{row_number}")
            style_source = cells.get("B")
            if style_source is not None and style_source.get("s") is not None:
                cell.set("s", style_source.get("s"))
            if row_number == 29:
                count_value = None if latest_record is None else latest_record.get(f"个券数量_{label}")
                _set_xml_cell_value(cell, count_value)
            else:
                _set_xml_cell_value(
                    cell,
                    formula=summary_formula_rows[row_number].format(column=column),
                )
    for row in summary_rows.values():
        cell_nodes = row.findall(_qname(_SHEET_NS, "c"))
        for cell in cell_nodes:
            row.remove(cell)
        for cell in sorted(
            cell_nodes,
            key=lambda node: column_index_from_string(_cell_column(node.get("r", ""))),
        ):
            row.append(cell)

    for row in list(sheet_data.findall(_qname(_SHEET_NS, "row"))):
        if int(row.get("r", "0")) >= 32:
            sheet_data.remove(row)

    value_columns = {
        get_column_letter(2 + offset): f"转股溢价率_{label}"
        for offset, (label, _, _) in enumerate(TERM_BUCKETS)
    }
    for offset, (_, record) in enumerate(metrics.iterrows()):
        row_number = 32 + offset
        row = deepcopy(template_row)
        row.set("r", str(row_number))
        if row.get("spans") is not None:
            row.set("spans", "1:13")
        cells = {
            _cell_column(cell.get("r", "")): cell
            for cell in row.findall(_qname(_SHEET_NS, "c"))
        }
        for column_number in range(1, 14):
            column = get_column_letter(column_number)
            cell = cells.get(column)
            if cell is None:
                cell = ET.SubElement(row, _qname(_SHEET_NS, "c"))
                cells[column] = cell
            cell.set("r", f"{column}{row_number}")
            style_id = style_by_column.get(column)
            if style_id is not None:
                cell.set("s", style_id)
            _set_xml_cell_value(cell, None)

        date = pd.Timestamp(record["日期"]).normalize()
        excel_date = int((date - pd.Timestamp("1899-12-30")).days)
        _set_xml_cell_value(cells["A"], excel_date)
        for column, metric in value_columns.items():
            _set_xml_cell_value(cells[column], record.get(metric))

        cell_nodes = row.findall(_qname(_SHEET_NS, "c"))
        for cell in cell_nodes:
            row.remove(cell)
        for cell in sorted(
            cell_nodes,
            key=lambda node: column_index_from_string(_cell_column(node.get("r", ""))),
        ):
            row.append(cell)
        sheet_data.append(row)

    last_row = 31 + len(metrics)
    dimension = root.find(_qname(_SHEET_NS, "dimension"))
    if dimension is not None:
        dimension.set("ref", f"A1:M{last_row}")
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _patch_group_valuation_sheet_xml(
    sheet_xml: bytes,
    metrics: pd.DataFrame,
    style_reference_xml: bytes,
    categories: list[str],
    blocks: list[tuple[int, str, str]],
) -> bytes:
    """按1.3页样式建立分组估值摘要及2017年以来日度历史区。"""
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(_qname(_SHEET_NS, "sheetData"))
    if sheet_data is None:
        raise RuntimeError("分组估值Sheet缺少sheetData")
    reference_root = ET.fromstring(style_reference_xml)
    reference_sheet_data = reference_root.find(_qname(_SHEET_NS, "sheetData"))
    if reference_sheet_data is None:
        raise RuntimeError("分组估值样式模板缺少sheetData")
    reference_rows = {
        int(row.get("r", "0")): row
        for row in reference_sheet_data.findall(_qname(_SHEET_NS, "row"))
    }
    if any(row_number not in reference_rows for row_number in range(20, 33)):
        raise RuntimeError("分组估值样式模板缺少第20至32行")

    def style_id(row_number: int, source_column: str) -> str | None:
        row = reference_rows[row_number]
        for cell in row.findall(_qname(_SHEET_NS, "c")):
            if _cell_column(cell.get("r", "")) == source_column:
                return cell.get("s")
        return None

    def make_row(row_number: int, template_row_number: int) -> object:
        template = reference_rows[template_row_number]
        attributes = dict(template.attrib)
        attributes["r"] = str(row_number)
        max_column = max(start + len(categories) for start, _, _ in blocks)
        attributes["spans"] = f"1:{max_column}"
        return ET.Element(_qname(_SHEET_NS, "row"), attributes)

    def add_cell(
        row,
        row_number: int,
        column_number: int,
        source_column: str,
        source_row_number: int,
    ):
        column = get_column_letter(column_number)
        cell = ET.SubElement(
            row,
            _qname(_SHEET_NS, "c"),
            {"r": f"{column}{row_number}"},
        )
        source_style = style_id(source_row_number, source_column)
        if source_style is not None:
            cell.set("s", source_style)
        return cell

    for row in list(sheet_data.findall(_qname(_SHEET_NS, "row"))):
        if int(row.get("r", "0")) >= 20:
            sheet_data.remove(row)

    metrics = metrics.sort_values("日期", ascending=False).reset_index(drop=True)
    latest_record = metrics.iloc[0] if not metrics.empty else None
    summary_labels = {
        22: "上周",
        23: "较上周变动",
        24: "2017年以来",
        25: "2022年以来",
        26: "滚动1年分位数",
        27: "滚动3年分位数",
        28: "滚动5年分位数",
        29: "个券数量",
    }
    for row_number in range(20, 32):
        row = make_row(row_number, row_number)
        for start_column, title, metric_prefix in blocks:
            date_column = get_column_letter(start_column)
            if row_number == 20:
                _set_xml_cell_text(
                    add_cell(row, row_number, start_column, "A", 20),
                    "日期",
                )
                for offset, category in enumerate(categories, start=1):
                    _set_xml_cell_text(
                        add_cell(row, row_number, start_column + offset, "B", 20),
                        category,
                    )
            elif row_number == 21:
                _set_xml_cell_value(
                    add_cell(row, row_number, start_column, "A", 21),
                    formula=f"={date_column}32",
                )
                for offset in range(1, len(categories) + 1):
                    column = get_column_letter(start_column + offset)
                    _set_xml_cell_value(
                        add_cell(row, row_number, start_column + offset, "B", 21),
                        formula=f"={column}32",
                    )
            elif row_number in summary_labels:
                _set_xml_cell_text(
                    add_cell(row, row_number, start_column, "A", row_number),
                    summary_labels[row_number],
                )
                for offset, category in enumerate(categories, start=1):
                    column_number = start_column + offset
                    column = get_column_letter(column_number)
                    cell = add_cell(row, row_number, column_number, "B", row_number)
                    if row_number == 22:
                        formula = f"={column}37"
                        _set_xml_cell_value(cell, formula=formula)
                    elif row_number == 23:
                        formula = f"={column}21-{column}22"
                        _set_xml_cell_value(cell, formula=formula)
                    elif row_number == 24:
                        formula = (
                            f"=_xlfn.PERCENTRANK.INC({column}32:"
                            f"INDIRECT(\"{column}\"&'1.1指标汇总'!$C$1),{column}32)"
                        )
                        _set_xml_cell_value(cell, formula=formula)
                    elif row_number == 25:
                        formula = (
                            f"=_xlfn.PERCENTRANK.INC({column}32:"
                            f"INDIRECT(\"{column}\"&'1.1指标汇总'!$D$1),{column}32)"
                        )
                        _set_xml_cell_value(cell, formula=formula)
                    elif row_number == 26:
                        _set_xml_cell_value(
                            cell,
                            formula=f"=_xlfn.PERCENTRANK.INC({column}32:{column}274,{column}32)",
                        )
                    elif row_number == 27:
                        _set_xml_cell_value(
                            cell,
                            formula=f"=_xlfn.PERCENTRANK.INC({column}32:{column}759,{column}32)",
                        )
                    elif row_number == 28:
                        _set_xml_cell_value(
                            cell,
                            formula=f"=_xlfn.PERCENTRANK.INC({column}32:{column}1243,{column}32)",
                        )
                    else:
                        count_value = (
                            None
                            if latest_record is None
                            else latest_record.get(f"个券数量_{category}")
                        )
                        _set_xml_cell_value(cell, count_value)
            elif row_number == 30:
                _set_xml_cell_text(
                    add_cell(row, row_number, start_column + 1, "B", 30),
                    title,
                )
            elif row_number == 31:
                _set_xml_cell_text(
                    add_cell(row, row_number, start_column, "A", 31),
                    "日期",
                )
                for offset, category in enumerate(categories, start=1):
                    _set_xml_cell_text(
                        add_cell(row, row_number, start_column + offset, "B", 31),
                        category,
                    )
        sheet_data.append(row)

    for offset, (_, record) in enumerate(metrics.iterrows()):
        row_number = 32 + offset
        row = make_row(row_number, 32)
        date = pd.Timestamp(record["日期"]).normalize()
        excel_date = int((date - pd.Timestamp("1899-12-30")).days)
        for start_column, _, metric_prefix in blocks:
            _set_xml_cell_value(
                add_cell(row, row_number, start_column, "A", 32),
                excel_date,
            )
            for category_offset, category in enumerate(categories, start=1):
                _set_xml_cell_value(
                    add_cell(
                        row,
                        row_number,
                        start_column + category_offset,
                        "B",
                        32,
                    ),
                    record.get(f"{metric_prefix}{category}"),
                )
        sheet_data.append(row)

    max_column = max(start + len(categories) for start, _, _ in blocks)
    existing_cols = root.find(_qname(_SHEET_NS, "cols"))
    if existing_cols is not None:
        root.remove(existing_cols)
    cols = ET.Element(_qname(_SHEET_NS, "cols"))
    block_columns = {
        column_number
        for start_column, _, _ in blocks
        for column_number in range(start_column, start_column + len(categories) + 1)
    }
    date_columns = {start_column for start_column, _, _ in blocks}
    for column_number in range(1, max_column + 1):
        width = 12.0 if column_number in date_columns else 11.0
        if column_number not in block_columns:
            width = 2.5
        ET.SubElement(
            cols,
            _qname(_SHEET_NS, "col"),
            {
                "min": str(column_number),
                "max": str(column_number),
                "width": str(width),
                "customWidth": "1",
            },
        )
    sheet_data_index = list(root).index(sheet_data)
    root.insert(sheet_data_index, cols)

    last_row = 31 + len(metrics)
    dimension = root.find(_qname(_SHEET_NS, "dimension"))
    if dimension is None:
        dimension = ET.Element(_qname(_SHEET_NS, "dimension"))
        root.insert(0, dimension)
    dimension.set("ref", f"A1:{get_column_letter(max_column)}{last_row}")
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _patch_seasoning_valuation_sheet_xml(
    sheet_xml: bytes,
    metrics: pd.DataFrame,
) -> bytes:
    """写入1.7次新券估值2017年以来日度历史序列。"""
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(_qname(_SHEET_NS, "sheetData"))
    if sheet_data is None:
        raise RuntimeError("1.7次新券估值缺少sheetData")
    template_row = next(
        (row for row in sheet_data.findall(_qname(_SHEET_NS, "row")) if row.get("r") == "18"),
        None,
    )
    header_row = next(
        (row for row in sheet_data.findall(_qname(_SHEET_NS, "row")) if row.get("r") == "17"),
        None,
    )
    if template_row is None or header_row is None:
        raise RuntimeError("1.7次新券估值缺少第17或18行模板")

    headers = [
        "日期",
        "全部转债隐含波动率均值",
        "次新券隐含波动率均值",
        "次新券转股溢价率均值",
        "老券隐含波动率均值",
    ]
    header_cells = {
        _cell_column(cell.get("r", "")): cell
        for cell in header_row.findall(_qname(_SHEET_NS, "c"))
    }
    for column_number, header in enumerate(headers, start=1):
        column = get_column_letter(column_number)
        cell = header_cells.get(column)
        if cell is None:
            cell = ET.SubElement(header_row, _qname(_SHEET_NS, "c"))
        cell.set("r", f"{column}17")
        _set_xml_cell_text(cell, header)

    for row in list(sheet_data.findall(_qname(_SHEET_NS, "row"))):
        if int(row.get("r", "0")) >= 18:
            sheet_data.remove(row)

    template_styles = {
        _cell_column(cell.get("r", "")): cell.get("s")
        for cell in template_row.findall(_qname(_SHEET_NS, "c"))
    }
    date_style = template_styles.get("A")
    numeric_style = template_styles.get("B") or template_styles.get("C")
    column_metrics = {
        "B": "全部转债隐含波动率均值",
        "C": "次新券隐含波动率均值",
        "D": "次新券转股溢价率均值",
        "E": "老券隐含波动率均值",
    }
    metrics = metrics.sort_values("日期", ascending=False).reset_index(drop=True)
    for offset, (_, record) in enumerate(metrics.iterrows()):
        row_number = 18 + offset
        row = deepcopy(template_row)
        row.set("r", str(row_number))
        if row.get("spans") is not None:
            row.set("spans", "1:5")
        cells = {
            _cell_column(cell.get("r", "")): cell
            for cell in row.findall(_qname(_SHEET_NS, "c"))
        }
        for column_number in range(1, 6):
            column = get_column_letter(column_number)
            cell = cells.get(column)
            if cell is None:
                cell = ET.SubElement(row, _qname(_SHEET_NS, "c"))
                cells[column] = cell
            cell.set("r", f"{column}{row_number}")
            style_id = date_style if column == "A" else numeric_style
            if style_id is not None:
                cell.set("s", style_id)
            _set_xml_cell_value(cell, None)
        date = pd.Timestamp(record["日期"]).normalize()
        excel_date = int((date - pd.Timestamp("1899-12-30")).days)
        _set_xml_cell_value(cells["A"], excel_date)
        for column, metric in column_metrics.items():
            _set_xml_cell_value(cells[column], record.get(metric))
        cell_nodes = row.findall(_qname(_SHEET_NS, "c"))
        for cell in cell_nodes:
            row.remove(cell)
        for cell in sorted(
            cell_nodes,
            key=lambda node: column_index_from_string(_cell_column(node.get("r", ""))),
        ):
            row.append(cell)
        sheet_data.append(row)

    last_row = 17 + len(metrics)
    dimension = root.find(_qname(_SHEET_NS, "dimension"))
    if dimension is not None:
        dimension.set("ref", f"A1:E{last_row}")
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _patch_history_valuation_sheet_xml(
    sheet_xml: bytes,
    metrics: pd.DataFrame,
) -> bytes:
    """将1.8历史估值对比的2017年以来日度数据写入O:AI。"""
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(_qname(_SHEET_NS, "sheetData"))
    if sheet_data is None:
        raise RuntimeError("1.8历史估值对比缺少sheetData")
    rows = sheet_data.findall(_qname(_SHEET_NS, "row"))
    header_template = next((row for row in rows if row.get("r") == "1"), None)
    data_template = next((row for row in rows if row.get("r") == "2"), None)
    if header_template is None or data_template is None:
        raise RuntimeError("1.8历史估值对比缺少第1或第2行模板")

    header_styles = {
        _cell_column(cell.get("r", "")): cell.get("s")
        for cell in header_template.findall(_qname(_SHEET_NS, "c"))
    }
    data_styles = {
        _cell_column(cell.get("r", "")): cell.get("s")
        for cell in data_template.findall(_qname(_SHEET_NS, "c"))
    }
    row_by_number = {int(row.get("r", "0")): row for row in rows}

    for row in rows:
        for cell in list(row.findall(_qname(_SHEET_NS, "c"))):
            column = _cell_column(cell.get("r", ""))
            column_number = column_index_from_string(column) if column else 0
            if 15 <= column_number <= 35:
                row.remove(cell)

    def get_row(row_number: int) -> object:
        row = row_by_number.get(row_number)
        if row is None:
            row = deepcopy(data_template)
            for cell in list(row.findall(_qname(_SHEET_NS, "c"))):
                row.remove(cell)
            row.set("r", str(row_number))
            sheet_data.append(row)
            row_by_number[row_number] = row
        else:
            row.set("r", str(row_number))
        if row.get("spans") is not None:
            row.set("spans", "1:35")
        return row

    headers = [
        "日期",
        "中证转债涨跌幅%",
        "全转债等权涨跌幅%",
        "总成交额(亿元)",
        "平均换手率%",
        "上涨家数",
        "下跌家数",
        "前五成交额占比",
        "转股溢价率中位数",
        "YTM中位数",
        "纯债溢价率中位数",
        "隐含波动率中位数",
        "次新券隐含波动率中位数",
        "70平价溢价率",
        "百元溢价率",
        "120平价溢价率",
        "收盘价中位数",
        "双低中位数",
        "YTM>0占比",
        "破面率",
        "跌破债底比例",
    ]
    header_row = get_row(1)
    for column_number, header in enumerate(headers, start=15):
        column = get_column_letter(column_number)
        cell = ET.SubElement(
            header_row,
            _qname(_SHEET_NS, "c"),
            {"r": f"{column}1"},
        )
        style_id = header_styles.get(column) or header_styles.get("O")
        if style_id is not None:
            cell.set("s", style_id)
        _set_xml_cell_text(cell, header)

    value_metrics = [
        "中证转债涨跌幅",
        "转债等权涨跌幅",
        "总成交额",
        "平均换手率",
        "上涨家数",
        "下跌家数",
        "前五成交额占总成交额比例",
        "转股溢价率中位数",
        "YTM中位数",
        "纯债溢价率中位数",
        "隐含波动率中位数",
        "剩余期限5.5-6年转债隐含波动率中位数",
        "70平价溢价率",
        "百元溢价率",
        "120平价溢价率",
        "收盘价中位数",
        "双低中位数",
        "YTM>0占比",
        "破面率",
        "跌破债底占比",
    ]
    metrics = (
        metrics.sort_values("日期", ascending=False)
        .drop_duplicates("日期", keep="last")
        .reset_index(drop=True)
    )
    for offset, (_, record) in enumerate(metrics.iterrows(), start=2):
        row = get_row(offset)
        date = pd.Timestamp(record["日期"]).normalize()
        excel_date = int((date - pd.Timestamp("1899-12-30")).days)
        date_cell = ET.SubElement(
            row,
            _qname(_SHEET_NS, "c"),
            {"r": f"O{offset}"},
        )
        if data_styles.get("O") is not None:
            date_cell.set("s", data_styles["O"])
        _set_xml_cell_value(date_cell, excel_date)
        for column_number, metric in enumerate(value_metrics, start=16):
            column = get_column_letter(column_number)
            cell = ET.SubElement(
                row,
                _qname(_SHEET_NS, "c"),
                {"r": f"{column}{offset}"},
            )
            style_id = data_styles.get(column) or data_styles.get("P")
            if style_id is not None:
                cell.set("s", style_id)
            value = record.get(metric)
            if metric == "平均换手率" and value is not None and pd.notna(value):
                value = float(value) * 100.0
            _set_xml_cell_value(cell, value)

    last_row = 1 + len(metrics)
    for row_number, row in list(row_by_number.items()):
        if row_number <= last_row:
            continue
        if not row.findall(_qname(_SHEET_NS, "c")):
            sheet_data.remove(row)
            del row_by_number[row_number]

    row_nodes = sheet_data.findall(_qname(_SHEET_NS, "row"))
    for row in row_nodes:
        cell_nodes = row.findall(_qname(_SHEET_NS, "c"))
        for cell in cell_nodes:
            row.remove(cell)
        for cell in sorted(
            cell_nodes,
            key=lambda node: column_index_from_string(_cell_column(node.get("r", ""))),
        ):
            row.append(cell)
        sheet_data.remove(row)
    for row in sorted(row_nodes, key=lambda node: int(node.get("r", "0"))):
        sheet_data.append(row)

    dimension = root.find(_qname(_SHEET_NS, "dimension"))
    if dimension is not None:
        dimension.set("ref", f"A1:AI{last_row}")
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _patch_major_valuation_sheet_xml(
    sheet_xml: bytes,
    metrics: pd.DataFrame,
) -> bytes:
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(_qname(_SHEET_NS, "sheetData"))
    if sheet_data is None:
        raise RuntimeError("主要估值指标缺少sheetData")
    rows = sheet_data.findall(_qname(_SHEET_NS, "row"))
    template_row = next((row for row in rows if row.get("r") == "23"), None)
    if template_row is None:
        raise RuntimeError("主要估值指标缺少第23行格式模板")

    template_styles = {
        _cell_column(cell.get("r", "")): cell.get("s")
        for cell in template_row.findall(_qname(_SHEET_NS, "c"))
    }
    style_by_column = {
        "A": template_styles.get("A"),
        "B": template_styles.get("B"),
        "C": template_styles.get("C"),
        "D": template_styles.get("D"),
        "E": template_styles.get("E"),
        "F": template_styles.get("F"),
        "G": template_styles.get("G"),
        "H": template_styles.get("H"),
        "K": template_styles.get("K") or template_styles.get("A"),
        "L": template_styles.get("L") or template_styles.get("B"),
        "M": template_styles.get("M") or template_styles.get("C"),
        "N": template_styles.get("N") or template_styles.get("D"),
        "O": template_styles.get("O") or template_styles.get("E"),
        "P": template_styles.get("P") or template_styles.get("F"),
        "Q": template_styles.get("Q") or template_styles.get("G"),
        "R": template_styles.get("R") or template_styles.get("H"),
        "T": template_styles.get("T") or template_styles.get("A"),
        "U": template_styles.get("U") or template_styles.get("B"),
        "V": template_styles.get("V") or template_styles.get("C"),
        "W": template_styles.get("W") or template_styles.get("D"),
        "X": template_styles.get("X") or template_styles.get("E"),
        "Y": template_styles.get("Y") or template_styles.get("F"),
        "Z": template_styles.get("Z") or template_styles.get("G"),
        "AA": template_styles.get("AA") or template_styles.get("H"),
        "AB": template_styles.get("AB") or template_styles.get("B"),
        "AC": template_styles.get("AC") or template_styles.get("C"),
        "AD": template_styles.get("AD") or template_styles.get("D"),
        "AE": template_styles.get("AE") or template_styles.get("E"),
        "AF": template_styles.get("AF") or template_styles.get("F"),
        "AG": template_styles.get("AG") or template_styles.get("G"),
        "AH": template_styles.get("AH") or template_styles.get("H"),
    }
    target_columns = set(style_by_column)
    row_by_number = {int(row.get("r", "0")): row for row in rows}
    original_last_row = max(row_by_number, default=22)

    for row_number, row in list(row_by_number.items()):
        if row_number < 23:
            continue
        for cell in list(row.findall(_qname(_SHEET_NS, "c"))):
            if _cell_column(cell.get("r", "")) in target_columns:
                row.remove(cell)

    column_metrics = {
        "B": "百元拟合溢价率",
        "C": "百元拟合溢价率_250日均值",
        "D": "百元拟合溢价率_250标准差",
        "E": "百元拟合溢价率_+1倍标准差",
        "F": "百元拟合溢价率_+2倍标准差",
        "G": "百元拟合溢价率_-1倍标准差",
        "H": "百元拟合溢价率_-2倍标准差",
        "L": "隐含波动率",
        "M": "隐含波动率_250日均值",
        "N": "隐含波动率_250标准差",
        "O": "隐含波动率_+1倍标准差",
        "P": "隐含波动率_+2倍标准差",
        "Q": "隐含波动率_-1倍标准差",
        "R": "隐含波动率_-2倍标准差",
        "U": "正股波动率",
        "V": "正股波动率_250日均值",
        "W": "正股波动率_250标准差",
        "X": "正股波动率_+1倍标准差",
        "Y": "正股波动率_+2倍标准差",
        "Z": "正股波动率_-1倍标准差",
        "AA": "正股波动率_-2倍标准差",
        "AB": "隐波差",
        "AC": "隐波差_250日均值",
        "AD": "隐波差_250标准差",
        "AE": "隐波差_+1倍标准差",
        "AF": "隐波差_+2倍标准差",
        "AG": "隐波差_-1倍标准差",
        "AH": "隐波差_-2倍标准差",
    }
    metrics = metrics.sort_values("日期", ascending=False).reset_index(drop=True)
    for offset, (_, record) in enumerate(metrics.iterrows()):
        row_number = 23 + offset
        row = row_by_number.get(row_number)
        if row is None:
            row = ET.Element(_qname(_SHEET_NS, "row"), r=str(row_number), spans="1:34")
            sheet_data.append(row)
            row_by_number[row_number] = row
        else:
            row.set("r", str(row_number))
            if row.get("spans") is not None:
                row.set("spans", "1:34")

        cells = {
            _cell_column(cell.get("r", "")): cell
            for cell in row.findall(_qname(_SHEET_NS, "c"))
        }
        for column in sorted(target_columns, key=column_index_from_string):
            cell = cells.get(column)
            if cell is None:
                cell = ET.SubElement(row, _qname(_SHEET_NS, "c"))
                cells[column] = cell
            cell.set("r", f"{column}{row_number}")
            style_id = style_by_column.get(column)
            if style_id is not None:
                cell.set("s", style_id)

        date = pd.Timestamp(record["日期"]).normalize()
        excel_date = int((date - pd.Timestamp("1899-12-30")).days)
        _set_xml_cell_value(cells["A"], excel_date)
        _set_xml_cell_value(cells["K"], excel_date)
        _set_xml_cell_value(cells["T"], excel_date)
        for column, metric in column_metrics.items():
            _set_xml_cell_value(cells[column], record.get(metric))

        cell_nodes = row.findall(_qname(_SHEET_NS, "c"))
        for cell in cell_nodes:
            row.remove(cell)
        for cell in sorted(
            cell_nodes,
            key=lambda node: column_index_from_string(_cell_column(node.get("r", ""))),
        ):
            row.append(cell)

    row_nodes = sheet_data.findall(_qname(_SHEET_NS, "row"))
    for row in row_nodes:
        sheet_data.remove(row)
    for row in sorted(row_nodes, key=lambda node: int(node.get("r", "0"))):
        sheet_data.append(row)

    last_row = 22 + len(metrics)
    dimension = root.find(_qname(_SHEET_NS, "dimension"))
    if dimension is not None:
        dimension.set("ref", f"A1:AH{max(original_last_row, last_row)}")
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _patch_style_factor_sheet_xml(
    sheet_xml: bytes,
    metrics: pd.DataFrame,
) -> bytes:
    """将2017年以来风格指数写入1.9，并保留最新一周涨跌幅区。"""
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(_qname(_SHEET_NS, "sheetData"))
    if sheet_data is None:
        raise RuntimeError("1.9转债风格因子缺少sheetData")
    rows = sheet_data.findall(_qname(_SHEET_NS, "row"))
    header_template = next((row for row in rows if row.get("r") == "1"), None)
    data_template = next((row for row in rows if row.get("r") == "2"), None)
    if header_template is None or data_template is None:
        raise RuntimeError("1.9转债风格因子缺少第1或第2行模板")

    header_styles = {
        _cell_column(cell.get("r", "")): cell.get("s")
        for cell in header_template.findall(_qname(_SHEET_NS, "c"))
    }
    data_styles = {
        _cell_column(cell.get("r", "")): cell.get("s")
        for cell in data_template.findall(_qname(_SHEET_NS, "c"))
    }
    row_by_number = {
        int(row.get("r", "0")): row
        for row in rows
    }

    def get_row(row_number: int, template) -> object:
        row = row_by_number.get(row_number)
        if row is None:
            row = deepcopy(template)
            for cell in list(row.findall(_qname(_SHEET_NS, "c"))):
                row.remove(cell)
            row.set("r", str(row_number))
            sheet_data.append(row)
            row_by_number[row_number] = row
        else:
            row.set("r", str(row_number))
        if row.get("spans") is not None:
            row.set("spans", "1:39")
        return row

    for row in rows:
        for cell in list(row.findall(_qname(_SHEET_NS, "c"))):
            column = _cell_column(cell.get("r", ""))
            if column and 1 <= column_index_from_string(column) <= 19:
                row.remove(cell)

    header_row = get_row(1, header_template)
    for column_number, header in enumerate(STYLE_COLUMNS, start=1):
        column = get_column_letter(column_number)
        cell = ET.SubElement(header_row, _qname(_SHEET_NS, "c"), {"r": f"{column}1"})
        style_id = header_styles.get(column) or header_styles.get("A")
        if style_id is not None:
            cell.set("s", style_id)
        _set_xml_cell_text(cell, header)

    metrics = (
        metrics.sort_values("日期", ascending=False)
        .drop_duplicates("日期", keep="last")
        .reset_index(drop=True)
    )
    for offset, (_, record) in enumerate(metrics.iterrows(), start=2):
        row = get_row(offset, data_template)
        date = pd.Timestamp(record["日期"]).normalize()
        excel_date = int((date - pd.Timestamp("1899-12-30")).days)
        for column_number, metric in enumerate(STYLE_COLUMNS, start=1):
            column = get_column_letter(column_number)
            cell = ET.SubElement(
                row,
                _qname(_SHEET_NS, "c"),
                {"r": f"{column}{offset}"},
            )
            style_id = data_styles.get(column) or (
                data_styles.get("A") if column == "A" else data_styles.get("B")
            )
            if style_id is not None:
                cell.set("s", style_id)
            _set_xml_cell_value(
                cell,
                excel_date if metric == "日期" else record.get(metric),
            )

    last_row = 1 + len(metrics)
    for row_number, row in list(row_by_number.items()):
        if row_number <= last_row:
            continue
        if not row.findall(_qname(_SHEET_NS, "c")):
            sheet_data.remove(row)
            del row_by_number[row_number]

    weekly_row = get_row(2, data_template)
    weekly_cells = {
        _cell_column(cell.get("r", "")): cell
        for cell in weekly_row.findall(_qname(_SHEET_NS, "c"))
    }
    for source_column_number, target_column_number in zip(range(2, 20), range(22, 40)):
        source_column = get_column_letter(source_column_number)
        target_column = get_column_letter(target_column_number)
        cell = weekly_cells.get(target_column)
        if cell is None:
            cell = ET.SubElement(
                weekly_row,
                _qname(_SHEET_NS, "c"),
                {"r": f"{target_column}2"},
            )
        cell.set("r", f"{target_column}2")
        _set_xml_cell_value(cell, formula=f"={source_column}2/{source_column}7-1")

    row_nodes = sheet_data.findall(_qname(_SHEET_NS, "row"))
    for row in row_nodes:
        cell_nodes = row.findall(_qname(_SHEET_NS, "c"))
        for cell in cell_nodes:
            row.remove(cell)
        for cell in sorted(
            cell_nodes,
            key=lambda node: column_index_from_string(_cell_column(node.get("r", ""))),
        ):
            row.append(cell)
        sheet_data.remove(row)
    for row in sorted(row_nodes, key=lambda node: int(node.get("r", "0"))):
        sheet_data.append(row)

    dimension = root.find(_qname(_SHEET_NS, "dimension"))
    if dimension is not None:
        dimension.set("ref", f"A1:AM{last_row}")
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _market_decomp_row(timeseries: pd.DataFrame) -> pd.DataFrame:
    if timeseries.empty:
        return pd.DataFrame(columns=DECOMP_GROUP_COLUMNS)
    latest = timeseries.sort_values("日期", ascending=False).iloc[0]
    return pd.DataFrame(
        [
            {
                "分类": "转债等权",
                **{metric: latest[metric] for metric in DECOMP_COLUMNS[1:]},
            }
        ],
        columns=DECOMP_GROUP_COLUMNS,
    )


def _patch_return_decomposition_sheet_xml(
    sheet_xml: bytes,
    timeseries: pd.DataFrame,
    industry: pd.DataFrame,
    parity: pd.DataFrame,
    bond_type: pd.DataFrame,
) -> bytes:
    """写入1.10的滚动20日序列及最新行业、平价、类型拆解。"""
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(_qname(_SHEET_NS, "sheetData"))
    if sheet_data is None:
        raise RuntimeError("1.10回报拆解缺少sheetData")
    rows = sheet_data.findall(_qname(_SHEET_NS, "row"))
    row_by_number = {int(row.get("r", "0")): row for row in rows}
    required_rows = (1, 2, 34, 35, 41, 42)
    if any(row_number not in row_by_number for row_number in required_rows):
        raise RuntimeError("1.10回报拆解缺少布局模板行")

    style_maps: dict[int, dict[str, str | None]] = {}
    for row_number in required_rows:
        style_maps[row_number] = {
            _cell_column(cell.get("r", "")): cell.get("s")
            for cell in row_by_number[row_number].findall(_qname(_SHEET_NS, "c"))
        }

    for row in rows:
        for cell in list(row.findall(_qname(_SHEET_NS, "c"))):
            column = _cell_column(cell.get("r", ""))
            column_number = column_index_from_string(column) if column else 0
            if 1 <= column_number <= 5 or 9 <= column_number <= 13:
                row.remove(cell)

    def get_row(row_number: int, template_row_number: int) -> object:
        row = row_by_number.get(row_number)
        if row is None:
            row = deepcopy(row_by_number[template_row_number])
            for cell in list(row.findall(_qname(_SHEET_NS, "c"))):
                row.remove(cell)
            row.set("r", str(row_number))
            sheet_data.append(row)
            row_by_number[row_number] = row
        else:
            row.set("r", str(row_number))
        if row.get("spans") is not None:
            row.set("spans", "1:13")
        return row

    def add_cell(
        row,
        row_number: int,
        column_number: int,
        template_row_number: int,
        template_column: str,
    ):
        column = get_column_letter(column_number)
        cell = ET.SubElement(
            row,
            _qname(_SHEET_NS, "c"),
            {"r": f"{column}{row_number}"},
        )
        style_id = style_maps[template_row_number].get(template_column)
        if style_id is not None:
            cell.set("s", style_id)
        return cell

    market_row = _market_decomp_row(timeseries)
    blocks = [
        (1, pd.concat([market_row, industry], ignore_index=True)),
        (34, pd.concat([market_row, parity], ignore_index=True)),
        (41, pd.concat([market_row, bond_type], ignore_index=True)),
    ]
    if len(blocks[0][1]) + 1 >= 34:
        raise RuntimeError("1.10回报拆解行业数量超出模板区域")
    block_headers = ["20日拆分", "转债收益率", "债券贡献", "正股贡献", "估值贡献"]
    for start_row, frame in blocks:
        header_template_row = start_row
        data_template_row = start_row + 1
        header_row = get_row(start_row, header_template_row)
        for column_number, header in enumerate(block_headers, start=1):
            _set_xml_cell_text(
                add_cell(
                    header_row,
                    start_row,
                    column_number,
                    header_template_row,
                    get_column_letter(column_number),
                ),
                header,
            )
        for offset, (_, record) in enumerate(frame.iterrows(), start=1):
            row_number = start_row + offset
            row = get_row(row_number, data_template_row)
            _set_xml_cell_text(
                add_cell(row, row_number, 1, data_template_row, "A"),
                record.get("分类"),
            )
            for column_number, metric in enumerate(DECOMP_GROUP_COLUMNS[1:], start=2):
                _set_xml_cell_value(
                    add_cell(
                        row,
                        row_number,
                        column_number,
                        data_template_row,
                        get_column_letter(column_number),
                    ),
                    record.get(metric),
                )

    timeseries = (
        timeseries.sort_values("日期", ascending=False)
        .drop_duplicates("日期", keep="last")
        .reset_index(drop=True)
    )
    timeseries_header_row = get_row(1, 1)
    for offset, header in enumerate(DECOMP_COLUMNS, start=9):
        _set_xml_cell_text(
            add_cell(
                timeseries_header_row,
                1,
                offset,
                1,
                get_column_letter(offset),
            ),
            header,
        )
    for offset, (_, record) in enumerate(timeseries.iterrows(), start=2):
        row = get_row(offset, 2)
        date = pd.Timestamp(record["日期"]).normalize()
        excel_date = int((date - pd.Timestamp("1899-12-30")).days)
        _set_xml_cell_value(add_cell(row, offset, 9, 2, "I"), excel_date)
        for column_number, metric in enumerate(DECOMP_COLUMNS[1:], start=10):
            _set_xml_cell_value(
                add_cell(
                    row,
                    offset,
                    column_number,
                    2,
                    get_column_letter(column_number),
                ),
                record.get(metric),
            )

    last_row = max(45, 1 + len(timeseries))
    for row_number, row in list(row_by_number.items()):
        if row_number <= last_row:
            continue
        if not row.findall(_qname(_SHEET_NS, "c")):
            sheet_data.remove(row)
            del row_by_number[row_number]

    row_nodes = sheet_data.findall(_qname(_SHEET_NS, "row"))
    for row in row_nodes:
        cell_nodes = row.findall(_qname(_SHEET_NS, "c"))
        for cell in cell_nodes:
            row.remove(cell)
        for cell in sorted(
            cell_nodes,
            key=lambda node: column_index_from_string(_cell_column(node.get("r", ""))),
        ):
            row.append(cell)
        sheet_data.remove(row)
    for row in sorted(row_nodes, key=lambda node: int(node.get("r", "0"))):
        sheet_data.append(row)

    dimension = root.find(_qname(_SHEET_NS, "dimension"))
    if dimension is not None:
        dimension.set("ref", f"A1:M{last_row}")
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _patch_weekly_bond_movers_sheet_xml(
    sheet_xml: bytes,
    movers: pd.DataFrame,
) -> bytes:
    """仅更新1.11的A:B列，保留C列及其右侧原有公式和格式。"""
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(_qname(_SHEET_NS, "sheetData"))
    if sheet_data is None:
        raise RuntimeError("1.11转债周度涨跌幅个券缺少sheetData")
    row_by_number = {
        int(row.get("r", "0")): row
        for row in sheet_data.findall(_qname(_SHEET_NS, "row"))
    }
    blocks = (("前20名", 3), ("后20名", 27))
    for group, start_row in blocks:
        group_data = movers.loc[movers["分组"] == group].reset_index(drop=True)
        if len(group_data) != WEEKLY_MOVER_COUNT:
            raise RuntimeError(f"1.11的{group}数量不是{WEEKLY_MOVER_COUNT}只")
        for offset, (_, record) in enumerate(group_data.iterrows()):
            row_number = start_row + offset
            row = row_by_number.get(row_number)
            if row is None:
                raise RuntimeError(f"1.11缺少第{row_number}行格式模板")
            cells = {
                _cell_column(cell.get("r", "")): cell
                for cell in row.findall(_qname(_SHEET_NS, "c"))
            }
            for column in ("A", "B"):
                cell = cells.get(column)
                if cell is None:
                    cell = ET.SubElement(row, _qname(_SHEET_NS, "c"))
                    cell.set("r", f"{column}{row_number}")
                    cells[column] = cell
            _set_xml_cell_text(cells["A"], str(record["转债代码"]))
            _set_xml_cell_text(cells["B"], str(record["转债名称"]))

            cell_nodes = row.findall(_qname(_SHEET_NS, "c"))
            for cell in cell_nodes:
                row.remove(cell)
            for cell in sorted(
                cell_nodes,
                key=lambda node: column_index_from_string(_cell_column(node.get("r", ""))),
            ):
                row.append(cell)
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _translate_template_formula_row(
    formula: str | None,
    source_row: int,
    target_row: int,
) -> str | None:
    if not formula or source_row == target_row:
        return formula

    def replace(match: re.Match[str]) -> str:
        column, absolute_row, _ = match.groups()
        row_number = source_row if absolute_row else target_row
        return f"{column}{absolute_row}{row_number}"

    return re.sub(
        rf"(\$?[A-Z]{{1,3}})(\$?)({source_row})\b",
        replace,
        formula,
    )


def _patch_redemption_trigger_sheet_xml(
    sheet_xml: bytes,
    candidates: pd.DataFrame,
) -> bytes:
    """动态更新2.1的A、B、L列，其余列沿用Excel模板公式。"""
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(_qname(_SHEET_NS, "sheetData"))
    if sheet_data is None:
        raise RuntimeError("2.1即将触发赎回缺少sheetData")
    template_row = next(
        (
            row
            for row in sheet_data.findall(_qname(_SHEET_NS, "row"))
            if row.get("r") == "2"
        ),
        None,
    )
    if template_row is None:
        raise RuntimeError("2.1即将触发赎回缺少第2行公式模板")
    for row in list(sheet_data.findall(_qname(_SHEET_NS, "row"))):
        if int(row.get("r", "0")) >= 2:
            sheet_data.remove(row)

    for offset, (_, record) in enumerate(candidates.iterrows(), start=2):
        row = deepcopy(template_row)
        row.set("r", str(offset))
        cells: dict[str, object] = {}
        for cell in row.findall(_qname(_SHEET_NS, "c")):
            column = _cell_column(cell.get("r", ""))
            cell.set("r", f"{column}{offset}")
            cells[column] = cell
            formula_node = cell.find(_qname(_SHEET_NS, "f"))
            if formula_node is not None:
                formula_node.text = _translate_template_formula_row(
                    formula_node.text,
                    2,
                    offset,
                )
                formula_node.set("ca", "1")
                value_node = cell.find(_qname(_SHEET_NS, "v"))
                if value_node is not None:
                    cell.remove(value_node)
        for column in ("A", "B", "H", "L"):
            if column not in cells:
                cell = ET.SubElement(row, _qname(_SHEET_NS, "c"))
                cell.set("r", f"{column}{offset}")
                cells[column] = cell
        _set_xml_cell_text(cells["A"], str(record["转债代码"]))
        _set_xml_cell_text(cells["B"], str(record["转债名称"]))
        _set_xml_cell_value(cells["H"], None)
        _set_xml_cell_value(cells["L"], int(record["累计天数"]))

        cell_nodes = row.findall(_qname(_SHEET_NS, "c"))
        for cell in cell_nodes:
            row.remove(cell)
        for cell in sorted(
            cell_nodes,
            key=lambda node: column_index_from_string(_cell_column(node.get("r", ""))),
        ):
            row.append(cell)
        sheet_data.append(row)

    dimension = root.find(_qname(_SHEET_NS, "dimension"))
    if dimension is not None:
        dimension.set("ref", f"A1:L{max(1, 1 + len(candidates))}")
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _disable_forced_excel_recalculation(workbook_xml: bytes) -> bytes:
    root = ET.fromstring(workbook_xml)
    calc_properties = root.find(_qname(_SHEET_NS, "calcPr"))
    if calc_properties is None:
        return workbook_xml
    changed = False
    for attribute in ("fullCalcOnLoad", "forceFullCalc"):
        if attribute in calc_properties.attrib:
            del calc_properties.attrib[attribute]
            changed = True
    if not changed:
        return workbook_xml
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def _remove_value_cells_from_calc_chain(
    calc_chain_xml: bytes,
    summary_sheet_ids: set[str],
    parity_sheet_id: str | None = None,
    term_sheet_id: str | None = None,
    rating_sheet_id: str | None = None,
    size_sheet_id: str | None = None,
    seasoning_sheet_id: str | None = None,
    history_sheet_id: str | None = None,
    style_sheet_id: str | None = None,
    decomposition_sheet_id: str | None = None,
    redemption_trigger_sheet_id: str | None = None,
) -> bytes:
    """从计算链移除已转为纯数值的历史单元格。"""
    root = ET.fromstring(calc_chain_xml)
    current_sheet_id: str | None = None
    changed = False
    for cell in list(root):
        if cell.get("i") is not None:
            current_sheet_id = str(cell.get("i"))
        match = re.fullmatch(r"([A-Z]+)(\d+)", cell.get("r", ""))
        is_summary_value = (
            current_sheet_id in summary_sheet_ids
            and match is not None
            and match.group(1) in {"B", "T"}
            and int(match.group(2)) >= 14
        )
        is_parity_value = (
            current_sheet_id == parity_sheet_id
            and match is not None
            and 1 <= column_index_from_string(match.group(1)) <= 26
            and int(match.group(2)) >= 32
        )
        is_term_value = (
            current_sheet_id == term_sheet_id
            and match is not None
            and 1 <= column_index_from_string(match.group(1)) <= 13
            and int(match.group(2)) >= 32
        )
        is_rating_value = (
            current_sheet_id == rating_sheet_id
            and match is not None
            and 1 <= column_index_from_string(match.group(1)) <= 35
            and int(match.group(2)) >= 32
        )
        is_size_value = (
            current_sheet_id == size_sheet_id
            and match is not None
            and 1 <= column_index_from_string(match.group(1)) <= 13
            and int(match.group(2)) >= 32
        )
        is_seasoning_value = (
            current_sheet_id == seasoning_sheet_id
            and match is not None
            and 1 <= column_index_from_string(match.group(1)) <= 5
            and int(match.group(2)) >= 18
        )
        is_style_value = (
            current_sheet_id == style_sheet_id
            and match is not None
            and 1 <= column_index_from_string(match.group(1)) <= 19
            and int(match.group(2)) >= 2
        )
        is_history_value = (
            current_sheet_id == history_sheet_id
            and match is not None
            and 15 <= column_index_from_string(match.group(1)) <= 35
            and int(match.group(2)) >= 2
        )
        is_decomposition_value = (
            current_sheet_id == decomposition_sheet_id
            and match is not None
            and (
                1 <= column_index_from_string(match.group(1)) <= 5
                or 9 <= column_index_from_string(match.group(1)) <= 13
            )
            and int(match.group(2)) >= 2
        )
        is_redemption_trigger_formula = (
            current_sheet_id == redemption_trigger_sheet_id
            and match is not None
            and 1 <= column_index_from_string(match.group(1)) <= 12
            and int(match.group(2)) >= 2
        )
        if (
            is_summary_value
            or is_parity_value
            or is_term_value
            or is_rating_value
            or is_size_value
            or is_seasoning_value
            or is_history_value
            or is_style_value
            or is_decomposition_value
            or is_redemption_trigger_formula
        ):
            root.remove(cell)
            changed = True
    if not changed:
        return calc_chain_xml
    return ET.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def write_workbook(data: dict[str, pd.DataFrame]) -> None:
    if not PA_WORKBOOK.exists():
        raise FileNotFoundError(f"未找到PA转债周度数据：{PA_WORKBOOK}")
    OUT.mkdir(parents=True, exist_ok=True)
    handle, temporary_name = tempfile.mkstemp(prefix="PA转债周度数据_", suffix=".tmp.xlsx", dir=OUT)
    os.close(handle)
    Path(temporary_name).unlink(missing_ok=True)
    temporary_path = Path(temporary_name)
    try:
        with zipfile.ZipFile(PA_WORKBOOK, "r") as source:
            summary_sheets = {
                _sheet_xml_path(source, "1.1指标汇总"): data["market"],
                _sheet_xml_path(source, "1.1指标汇总_剔除极值"): data["market_trimmed"],
            }
            major_valuation_path = _sheet_xml_path(source, "主要估值指标")
            bond_indicator_path = _sheet_xml_path(source, "债性指标")
            parity_valuation_path = _sheet_xml_path(source, "1.3分平价估值")
            term_valuation_path = _sheet_xml_path(source, "1.4分期限估值")
            rating_valuation_path = _sheet_xml_path(source, "1.5分评级估值")
            size_valuation_path = _sheet_xml_path(source, "1.6分规模估值")
            seasoning_valuation_path = _sheet_xml_path(source, "1.7次新券估值")
            history_valuation_path = _sheet_xml_path(source, "1.8历史估值对比")
            style_factor_path = _sheet_xml_path(source, "1.9转债风格因子")
            return_decomposition_path = _sheet_xml_path(source, "1.10回报拆解")
            weekly_bond_movers_path = _sheet_xml_path(source, "1.11转债周度涨跌幅个券")
            redemption_trigger_path = _sheet_xml_path(source, "2.1即将触发赎回")
            style_reference_xml = source.read(parity_valuation_path)
            summary_sheet_ids = {
                _sheet_id(source, "1.1指标汇总"),
                _sheet_id(source, "1.1指标汇总_剔除极值"),
            }
            parity_sheet_id = _sheet_id(source, "1.3分平价估值")
            term_sheet_id = _sheet_id(source, "1.4分期限估值")
            rating_sheet_id = _sheet_id(source, "1.5分评级估值")
            size_sheet_id = _sheet_id(source, "1.6分规模估值")
            seasoning_sheet_id = _sheet_id(source, "1.7次新券估值")
            history_sheet_id = _sheet_id(source, "1.8历史估值对比")
            style_sheet_id = _sheet_id(source, "1.9转债风格因子")
            decomposition_sheet_id = _sheet_id(source, "1.10回报拆解")
            redemption_trigger_sheet_id = _sheet_id(source, "2.1即将触发赎回")
            with zipfile.ZipFile(temporary_path, "w") as target:
                for item in source.infolist():
                    payload = source.read(item.filename)
                    if item.filename in summary_sheets:
                        payload = _patch_summary_sheet_xml(
                            payload,
                            summary_sheets[item.filename],
                        )
                    elif item.filename == major_valuation_path:
                        payload = _patch_major_valuation_sheet_xml(
                            payload,
                            data["major_valuation_metrics"],
                        )
                    elif item.filename == bond_indicator_path:
                        payload = _patch_bond_indicator_sheet_xml(
                            payload,
                            data["market"],
                        )
                    elif item.filename == parity_valuation_path:
                        payload = _patch_parity_valuation_sheet_xml(
                            payload,
                            data["parity_valuation"],
                        )
                    elif item.filename == term_valuation_path:
                        payload = _patch_group_valuation_sheet_xml(
                            payload,
                            data["term_valuation"],
                            style_reference_xml,
                            [label for label, _, _ in TERM_BUCKETS],
                            [(1, "分剩余期限区间转股溢价率均值", "转股溢价率_")],
                        )
                    elif item.filename == rating_valuation_path:
                        payload = _patch_group_valuation_sheet_xml(
                            payload,
                            data["rating_valuation"],
                            style_reference_xml,
                            RATING_BUCKETS,
                            [
                                (1, "分评级转股溢价率均值", "转股溢价率_"),
                                (10, "分评级隐含波动率均值", "隐含波动率_"),
                                (19, "分评级YTM均值", "YTM_"),
                                (28, "分评级YTM较同评级3年期信用债利差均值", "YTM信用利差_"),
                            ],
                        )
                    elif item.filename == size_valuation_path:
                        payload = _patch_group_valuation_sheet_xml(
                            payload,
                            data["size_valuation"],
                            style_reference_xml,
                            [label for label, _, _ in SIZE_BUCKETS],
                            [
                                (1, "分规模转股溢价率均值", "转股溢价率_"),
                                (8, "分规模隐含波动率均值", "隐含波动率_"),
                            ],
                        )
                    elif item.filename == seasoning_valuation_path:
                        payload = _patch_seasoning_valuation_sheet_xml(
                            payload,
                            data["seasoning_valuation"],
                        )
                    elif item.filename == history_valuation_path:
                        payload = _patch_history_valuation_sheet_xml(
                            payload,
                            data["history_valuation_daily"],
                        )
                    elif item.filename == style_factor_path:
                        payload = _patch_style_factor_sheet_xml(
                            payload,
                            data["style"],
                        )
                    elif item.filename == return_decomposition_path:
                        payload = _patch_return_decomposition_sheet_xml(
                            payload,
                            data["decomp_timeseries"],
                            data["decomp_industry"],
                            data["decomp_parity"],
                            data["decomp_type"],
                        )
                    elif item.filename == weekly_bond_movers_path:
                        payload = _patch_weekly_bond_movers_sheet_xml(
                            payload,
                            data["weekly_bond_movers"],
                        )
                    elif item.filename == redemption_trigger_path:
                        payload = _patch_redemption_trigger_sheet_xml(
                            payload,
                            data["redemption_trigger_candidates"],
                        )
                    elif item.filename == "xl/workbook.xml":
                        payload = _disable_forced_excel_recalculation(payload)
                    elif item.filename == "xl/calcChain.xml":
                        payload = _remove_value_cells_from_calc_chain(
                            payload,
                            summary_sheet_ids,
                            parity_sheet_id,
                            term_sheet_id,
                            rating_sheet_id,
                            size_sheet_id,
                            seasoning_sheet_id,
                            history_sheet_id,
                            style_sheet_id,
                            decomposition_sheet_id,
                            redemption_trigger_sheet_id,
                        )
                    target.writestr(item, payload)
        temporary_path.replace(OUTPUT_XLSX)
    finally:
        temporary_path.unlink(missing_ok=True)


def _verify_group_valuation_sheet(
    worksheet,
    expected_metrics: pd.DataFrame,
    categories: list[str],
    blocks: list[tuple[int, str]],
) -> None:
    """校验分期限、分评级、分规模估值页的结构、数值与摘要区。"""
    expected = (
        expected_metrics.sort_values("日期", ascending=False)
        .drop_duplicates("日期", keep="last")
        .reset_index(drop=True)
    )
    max_column = max(start_column + len(categories) for start_column, _ in blocks)
    history_rows = list(
        worksheet.iter_rows(
            min_row=32,
            max_row=31 + len(expected),
            min_col=1,
            max_col=max_column,
        )
    )
    if len(history_rows) != len(expected):
        raise RuntimeError(f"{worksheet.title}历史数据行数不正确")
    for start_column, metric_prefix in blocks:
        if worksheet.cell(20, start_column).value != "日期":
            raise RuntimeError(f"{worksheet.title}摘要日期表头缺失")
        if worksheet.cell(31, start_column).value != "日期":
            raise RuntimeError(f"{worksheet.title}历史日期表头缺失")
        for offset, category in enumerate(categories, start=1):
            column_number = start_column + offset
            if worksheet.cell(20, column_number).value != category:
                raise RuntimeError(f"{worksheet.title}摘要分类表头不正确")
            if worksheet.cell(31, column_number).value != category:
                raise RuntimeError(f"{worksheet.title}历史分类表头不正确")
            for row_number in range(21, 29):
                cell = worksheet.cell(row_number, column_number)
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    raise RuntimeError(f"{worksheet.title}摘要公式缺失")
            count_cell = worksheet.cell(29, column_number)
            if not isinstance(count_cell.value, (int, float)):
                raise RuntimeError(f"{worksheet.title}个券数量缺失")

        for history_row, (_, record) in zip(history_rows, expected.iterrows()):
            date_cell = history_row[start_column - 1]
            if date_cell.data_type == "f" or pd.Timestamp(date_cell.value) != pd.Timestamp(record["日期"]):
                raise RuntimeError(f"{worksheet.title}日期与计算结果不一致")
            for category_offset, category in enumerate(categories, start=1):
                cell = history_row[start_column + category_offset - 1]
                if cell.data_type == "f":
                    raise RuntimeError(f"{worksheet.title}历史区域存在公式单元格")
                actual_value = cell.value
                expected_value = record[f"{metric_prefix}{category}"]
                if pd.isna(expected_value):
                    if actual_value is not None:
                        raise RuntimeError(f"{worksheet.title}空值与计算结果不一致")
                elif actual_value is None or not math.isclose(
                    float(actual_value),
                    float(expected_value),
                    rel_tol=1e-12,
                    abs_tol=1e-12,
                ):
                    raise RuntimeError(f"{worksheet.title}与计算结果不一致")

        latest = expected.iloc[0]
        for offset, category in enumerate(categories, start=1):
            actual_count = worksheet.cell(29, start_column + offset).value
            expected_count = latest[f"个券数量_{category}"]
            if int(actual_count) != int(expected_count):
                raise RuntimeError(f"{worksheet.title}个券数量与计算结果不一致")


def _verify_history_valuation_sheet(
    worksheet,
    expected_metrics: pd.DataFrame,
) -> None:
    expected = (
        expected_metrics.sort_values("日期", ascending=False)
        .drop_duplicates("日期", keep="last")
        .reset_index(drop=True)
    )
    expected_headers = [
        "日期",
        "中证转债涨跌幅%",
        "全转债等权涨跌幅%",
        "总成交额(亿元)",
        "平均换手率%",
        "上涨家数",
        "下跌家数",
        "前五成交额占比",
        "转股溢价率中位数",
        "YTM中位数",
        "纯债溢价率中位数",
        "隐含波动率中位数",
        "次新券隐含波动率中位数",
        "70平价溢价率",
        "百元溢价率",
        "120平价溢价率",
        "收盘价中位数",
        "双低中位数",
        "YTM>0占比",
        "破面率",
        "跌破债底比例",
    ]
    actual_headers = [worksheet.cell(1, column).value for column in range(15, 36)]
    if actual_headers != expected_headers:
        raise RuntimeError("1.8历史估值对比日度数据表头不正确")
    rows = list(
        worksheet.iter_rows(
            min_row=2,
            max_row=1 + len(expected),
            min_col=15,
            max_col=35,
        )
    )
    if len(rows) != len(expected):
        raise RuntimeError("1.8历史估值对比日度数据行数不正确")
    value_metrics = [
        "中证转债涨跌幅",
        "转债等权涨跌幅",
        "总成交额",
        "平均换手率",
        "上涨家数",
        "下跌家数",
        "前五成交额占总成交额比例",
        "转股溢价率中位数",
        "YTM中位数",
        "纯债溢价率中位数",
        "隐含波动率中位数",
        "剩余期限5.5-6年转债隐含波动率中位数",
        "70平价溢价率",
        "百元溢价率",
        "120平价溢价率",
        "收盘价中位数",
        "双低中位数",
        "YTM>0占比",
        "破面率",
        "跌破债底占比",
    ]
    for actual_row, (_, record) in zip(rows, expected.iterrows()):
        if actual_row[0].data_type == "f" or pd.Timestamp(actual_row[0].value) != pd.Timestamp(record["日期"]):
            raise RuntimeError("1.8历史估值对比日期与计算结果不一致")
        for offset, metric in enumerate(value_metrics, start=1):
            cell = actual_row[offset]
            if cell.data_type == "f":
                raise RuntimeError("1.8历史估值对比日度区域存在公式单元格")
            expected_value = record[metric]
            if metric == "平均换手率" and pd.notna(expected_value):
                expected_value = float(expected_value) * 100.0
            actual_value = cell.value
            if pd.isna(expected_value):
                if actual_value is not None:
                    raise RuntimeError("1.8历史估值对比空值与计算结果不一致")
            elif actual_value is None or not math.isclose(
                float(actual_value),
                float(expected_value),
                rel_tol=1e-12,
                abs_tol=1e-12,
            ):
                raise RuntimeError("1.8历史估值对比与计算结果不一致")
    if worksheet["O2"].number_format == "General":
        raise RuntimeError("1.8历史估值对比日期格式缺失")
    for column in ("V",):
        if "%" not in worksheet[f"{column}2"].number_format:
            raise RuntimeError("1.8历史估值对比比例格式不正确")


def _verify_style_factor_sheet(worksheet, expected_metrics: pd.DataFrame) -> None:
    expected = (
        expected_metrics.sort_values("日期", ascending=False)
        .drop_duplicates("日期", keep="last")
        .reset_index(drop=True)
    )
    headers = [worksheet.cell(1, column).value for column in range(1, 20)]
    if headers != STYLE_COLUMNS:
        raise RuntimeError("1.9转债风格因子表头不正确")
    rows = list(
        worksheet.iter_rows(
            min_row=2,
            max_row=1 + len(expected),
            min_col=1,
            max_col=19,
        )
    )
    if len(rows) != len(expected):
        raise RuntimeError("1.9转债风格因子历史数据行数不正确")
    for actual_row, (_, record) in zip(rows, expected.iterrows()):
        for column_number, metric in enumerate(STYLE_COLUMNS):
            cell = actual_row[column_number]
            if cell.data_type == "f":
                raise RuntimeError("1.9转债风格因子历史区域存在公式单元格")
            expected_value = record[metric]
            actual_value = cell.value
            if metric == "日期":
                if pd.Timestamp(actual_value) != pd.Timestamp(expected_value):
                    raise RuntimeError("1.9转债风格因子日期与计算结果不一致")
            elif pd.isna(expected_value):
                if actual_value is not None:
                    raise RuntimeError("1.9转债风格因子空值与计算结果不一致")
            elif actual_value is None or not math.isclose(
                float(actual_value),
                float(expected_value),
                rel_tol=1e-12,
                abs_tol=1e-12,
            ):
                raise RuntimeError("1.9转债风格因子与计算结果不一致")
    for source_column_number, target_column_number in zip(range(2, 20), range(22, 40)):
        source_column = get_column_letter(source_column_number)
        target_column = get_column_letter(target_column_number)
        cell = worksheet[f"{target_column}2"]
        expected_formula = f"={source_column}2/{source_column}7-1"
        if cell.data_type != "f" or cell.value != expected_formula:
            raise RuntimeError("1.9转债风格因子最新一周涨跌幅公式不正确")


def _verify_return_decomposition_sheet(
    worksheet,
    timeseries: pd.DataFrame,
    industry: pd.DataFrame,
    parity: pd.DataFrame,
    bond_type: pd.DataFrame,
) -> None:
    expected_timeseries = (
        timeseries.sort_values("日期", ascending=False)
        .drop_duplicates("日期", keep="last")
        .reset_index(drop=True)
    )
    timeseries_headers = [worksheet.cell(1, column).value for column in range(9, 14)]
    if timeseries_headers != DECOMP_COLUMNS:
        raise RuntimeError("1.10回报拆解时间序列表头不正确")
    rows = list(
        worksheet.iter_rows(
            min_row=2,
            max_row=1 + len(expected_timeseries),
            min_col=9,
            max_col=13,
        )
    )
    if len(rows) != len(expected_timeseries):
        raise RuntimeError("1.10回报拆解时间序列行数不正确")
    for actual_row, (_, record) in zip(rows, expected_timeseries.iterrows()):
        for offset, metric in enumerate(DECOMP_COLUMNS):
            cell = actual_row[offset]
            if cell.data_type == "f":
                raise RuntimeError("1.10回报拆解时间序列存在公式单元格")
            expected_value = record[metric]
            actual_value = cell.value
            if metric == "日期":
                if pd.Timestamp(actual_value) != pd.Timestamp(expected_value):
                    raise RuntimeError("1.10回报拆解日期与计算结果不一致")
            elif pd.isna(expected_value):
                if actual_value is not None:
                    raise RuntimeError("1.10回报拆解时间序列空值不一致")
            elif actual_value is None or not math.isclose(
                float(actual_value),
                float(expected_value),
                rel_tol=1e-12,
                abs_tol=1e-12,
            ):
                raise RuntimeError("1.10回报拆解时间序列与计算结果不一致")

    market_row = _market_decomp_row(expected_timeseries)
    blocks = [
        (1, pd.concat([market_row, industry], ignore_index=True)),
        (34, pd.concat([market_row, parity], ignore_index=True)),
        (41, pd.concat([market_row, bond_type], ignore_index=True)),
    ]
    expected_headers = ["20日拆分", "转债收益率", "债券贡献", "正股贡献", "估值贡献"]
    for start_row, frame in blocks:
        actual_headers = [worksheet.cell(start_row, column).value for column in range(1, 6)]
        if actual_headers != expected_headers:
            raise RuntimeError("1.10回报拆解分组表头不正确")
        actual_rows = list(
            worksheet.iter_rows(
                min_row=start_row + 1,
                max_row=start_row + len(frame),
                min_col=1,
                max_col=5,
            )
        )
        for actual_row, (_, record) in zip(actual_rows, frame.iterrows()):
            if actual_row[0].value != record["分类"]:
                raise RuntimeError("1.10回报拆解分组标签不一致")
            for offset, metric in enumerate(DECOMP_GROUP_COLUMNS[1:], start=1):
                cell = actual_row[offset]
                if cell.data_type == "f":
                    raise RuntimeError("1.10回报拆解分组区域存在公式单元格")
                expected_value = record[metric]
                actual_value = cell.value
                if pd.isna(expected_value):
                    if actual_value is not None:
                        raise RuntimeError("1.10回报拆解分组空值不一致")
                elif actual_value is None or not math.isclose(
                    float(actual_value),
                    float(expected_value),
                    rel_tol=1e-12,
                    abs_tol=1e-12,
                ):
                    raise RuntimeError("1.10回报拆解分组与计算结果不一致")


def _verify_weekly_bond_movers_sheet(
    worksheet,
    movers: pd.DataFrame,
) -> None:
    """校验1.11前后各20只个券的A:B代码与名称。"""
    for group, start_row in (("前20名", 3), ("后20名", 27)):
        expected = movers.loc[movers["分组"] == group].reset_index(drop=True)
        if len(expected) != WEEKLY_MOVER_COUNT:
            raise RuntimeError(f"1.11的{group}计算结果数量不正确")
        for offset, (_, record) in enumerate(expected.iterrows()):
            row_number = start_row + offset
            code_cell = worksheet.cell(row_number, 1)
            name_cell = worksheet.cell(row_number, 2)
            if code_cell.data_type == "f" or name_cell.data_type == "f":
                raise RuntimeError("1.11的A:B列不应包含公式")
            if str(code_cell.value) != str(record["转债代码"]):
                raise RuntimeError(f"1.11的{group}转债代码与计算结果不一致")
            if str(name_cell.value) != str(record["转债名称"]):
                raise RuntimeError(f"1.11的{group}转债名称与计算结果不一致")


def _verify_redemption_trigger_sheet(
    worksheet,
    candidates: pd.DataFrame,
) -> None:
    """校验2.1的A、B、L列及其余Excel公式模板。"""
    if worksheet.max_row != max(1, 1 + len(candidates)):
        raise RuntimeError("2.1即将触发赎回明细行数与计算结果不一致")
    for offset, (_, record) in enumerate(candidates.iterrows(), start=2):
        code_cell = worksheet.cell(offset, 1)
        name_cell = worksheet.cell(offset, 2)
        days_cell = worksheet.cell(offset, 12)
        if any(cell.data_type == "f" for cell in (code_cell, name_cell, days_cell)):
            raise RuntimeError("2.1的A、B、L列不应包含公式")
        if str(code_cell.value) != str(record["转债代码"]):
            raise RuntimeError("2.1转债代码与计算结果不一致")
        if str(name_cell.value) != str(record["转债名称"]):
            raise RuntimeError("2.1转债名称与计算结果不一致")
        if days_cell.value is None or int(days_cell.value) != int(record["累计天数"]):
            raise RuntimeError("2.1累计天数与计算结果不一致")
        for column_number in (3, 4, 5, 6, 7, 9):
            if worksheet.cell(offset, column_number).data_type != "f":
                raise RuntimeError("2.1由Excel计算的公式列缺失")


def verify_workbook(
    path: Path,
    expected_data: dict[str, pd.DataFrame] | None = None,
) -> None:
    wb = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        for sheet_name in ("1.1指标汇总", "1.1指标汇总_剔除极值"):
            if sheet_name not in wb.sheetnames:
                raise RuntimeError(f"工作簿缺少 sheet：{sheet_name}")
            ws = wb[sheet_name]
            if ws.max_row < 14:
                raise RuntimeError(f"{sheet_name}没有历史数据")
            rows = ws.iter_rows(min_row=14, min_col=1, max_col=20)
            first_row = next(rows, None)
            if first_row is None:
                raise RuntimeError(f"{sheet_name}没有历史数据")
            if first_row[1].data_type == "f" or not isinstance(first_row[1].value, (int, float)):
                raise RuntimeError(f"{sheet_name}的转债指数不是纯数值")
            if first_row[19].data_type == "f" or not isinstance(first_row[19].value, (int, float)):
                raise RuntimeError(f"{sheet_name}的全市场成交额不是纯数值")
            ratios = [first_row[18].value, *(row[18].value for row in rows)]
            invalid = [value for value in ratios if value is not None and not (0 <= float(value) <= 1)]
            if invalid:
                raise RuntimeError(f"{sheet_name}的YTM大于3年AA信用债比例存在越界值")

        valuation_sheet = wb["主要估值指标"]
        summary_sheet = wb["1.1指标汇总"]
        valuation_rows = list(
            valuation_sheet.iter_rows(
                min_row=23,
                max_row=22 + summary_sheet.max_row - 13,
                min_col=1,
                max_col=34,
            )
        )
        if not valuation_rows:
            raise RuntimeError("主要估值指标没有历史数据")
        first_valuation_row = valuation_rows[0]
        target_indexes = [
            *range(0, 8),
            *range(10, 18),
            *range(19, 34),
        ]
        for index in target_indexes:
            cell = first_valuation_row[index]
            if cell.data_type == "f" or not isinstance(cell.value, (int, float, datetime)):
                raise RuntimeError("主要估值指标存在非数值或公式单元格")

        premium_window = np.array(
            [
                float(summary_sheet.cell(row, 6).value) / 100.0
                for row in range(14, min(summary_sheet.max_row, 263) + 1)
                if summary_sheet.cell(row, 6).value is not None
            ],
            dtype=float,
        )
        volatility_window = np.array(
            [
                float(summary_sheet.cell(row, 8).value) / 100.0
                for row in range(14, min(summary_sheet.max_row, 263) + 1)
                if summary_sheet.cell(row, 8).value is not None
            ],
            dtype=float,
        )
        stock_volatility_window = np.array(
            [
                float(summary_sheet.cell(row, 21).value) / 100.0
                for row in range(14, min(summary_sheet.max_row, 263) + 1)
                if summary_sheet.cell(row, 21).value is not None
            ],
            dtype=float,
        )
        volatility_gap_window = np.array(
            [
                float(summary_sheet.cell(row, 23).value) / 100.0
                for row in range(14, min(summary_sheet.max_row, 263) + 1)
                if summary_sheet.cell(row, 23).value is not None
            ],
            dtype=float,
        )
        expected = {
            1: premium_window[0],
            2: premium_window.mean(),
            3: premium_window.std(ddof=1),
            11: volatility_window[0],
            12: volatility_window.mean(),
            13: volatility_window.std(ddof=1),
            20: stock_volatility_window[0],
            21: stock_volatility_window.mean(),
            22: stock_volatility_window.std(ddof=1),
            27: volatility_gap_window[0],
            28: volatility_gap_window.mean(),
            29: volatility_gap_window.std(ddof=1),
        }
        for index, expected_value in expected.items():
            actual_value = float(first_valuation_row[index].value)
            if not math.isclose(actual_value, float(expected_value), rel_tol=1e-12, abs_tol=1e-12):
                raise RuntimeError("主要估值指标的250日滚动统计校验失败")

        bond_sheet = wb["债性指标"]
        bond_rows = list(
            bond_sheet.iter_rows(
                min_row=23,
                max_row=22 + summary_sheet.max_row - 13,
                min_col=1,
                max_col=8,
            )
        )
        summary_bond_rows = list(
            summary_sheet.iter_rows(
                min_row=14,
                max_row=summary_sheet.max_row,
                min_col=1,
                max_col=19,
            )
        )
        if len(bond_rows) != len(summary_bond_rows):
            raise RuntimeError("债性指标历史数据行数不正确")
        summary_indexes = (0, 12, 13, 14, 15, 16, 17, 18)
        for bond_row, summary_row in zip(bond_rows, summary_bond_rows):
            for bond_cell, summary_index in zip(bond_row, summary_indexes):
                expected_value = summary_row[summary_index].value
                actual_value = bond_cell.value
                if bond_cell.data_type == "f":
                    raise RuntimeError("债性指标存在公式单元格")
                if expected_value is None and actual_value is None:
                    continue
                if expected_value is None or actual_value is None:
                    raise RuntimeError("债性指标与1.1指标汇总存在空值差异")
                if isinstance(expected_value, datetime):
                    if pd.Timestamp(actual_value) != pd.Timestamp(expected_value):
                        raise RuntimeError("债性指标日期与1.1指标汇总不一致")
                elif not math.isclose(
                    float(actual_value),
                    float(expected_value),
                    rel_tol=1e-12,
                    abs_tol=1e-12,
                ):
                    raise RuntimeError("债性指标与1.1指标汇总口径不一致")
        if bond_sheet["G23"].number_format != bond_sheet["H23"].number_format:
            raise RuntimeError("债性指标的比例列格式不一致")

        parity_sheet = wb["1.3分平价估值"]
        parity_rows = list(
            parity_sheet.iter_rows(
                min_row=32,
                max_row=31 + len(summary_bond_rows),
                min_col=1,
                max_col=26,
            )
        )
        if len(parity_rows) != len(summary_bond_rows):
            raise RuntimeError("1.3分平价估值历史数据行数不正确")
        parity_target_indexes = [0, *range(1, 12), 14, *range(15, 26)]
        for parity_row, summary_row in zip(parity_rows, summary_bond_rows):
            expected_date = pd.Timestamp(summary_row[0].value)
            for index in parity_target_indexes:
                cell = parity_row[index]
                if cell.data_type == "f":
                    raise RuntimeError("1.3分平价估值历史区域存在公式单元格")
                if index in {0, 14}:
                    if pd.Timestamp(cell.value) != expected_date:
                        raise RuntimeError("1.3分平价估值日期与1.1指标汇总不一致")
                elif cell.value is not None and not isinstance(cell.value, (int, float)):
                    raise RuntimeError("1.3分平价估值存在非数值单元格")
        parity_number_format = parity_sheet["D32"].number_format
        for column in [
            *[get_column_letter(value) for value in range(2, 13)],
            *[get_column_letter(value) for value in range(16, 27)],
        ]:
            if parity_sheet[f"{column}32"].number_format != parity_number_format:
                raise RuntimeError("1.3分平价估值数值格式不一致")
        for column, reference_column in (("L", "K"), ("Z", "Y")):
            for row_number in range(21, 29):
                cell = parity_sheet[f"{column}{row_number}"]
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    raise RuntimeError("1.3分平价估值整体均值摘要公式缺失")
                if cell.number_format != parity_sheet[f"{reference_column}{row_number}"].number_format:
                    raise RuntimeError("1.3分平价估值整体均值摘要格式不一致")
            if not isinstance(parity_sheet[f"{column}29"].value, (int, float)):
                raise RuntimeError("1.3分平价估值整体均值有效个数缺失")
        if expected_data is not None:
            expected_parity = (
                expected_data["parity_valuation"]
                .sort_values("日期", ascending=False)
                .reset_index(drop=True)
            )
            if len(expected_parity) != len(parity_rows):
                raise RuntimeError("1.3分平价估值与计算结果行数不一致")
            latest_parity = expected_parity.iloc[0]
            for column, metric in (
                ("L", "整体转股溢价率有效个数"),
                ("Z", "整体隐含波动率有效个数"),
            ):
                if int(parity_sheet[f"{column}29"].value) != int(latest_parity[metric]):
                    raise RuntimeError("1.3分平价估值整体均值有效个数不一致")
            premium_metrics = [
                f"转股溢价率_{label}" for label, _, _ in PARITY_BUCKETS
            ]
            premium_metrics.append("整体转股溢价率均值")
            volatility_metrics = [
                f"隐含波动率_{label}" for label, _, _ in PARITY_BUCKETS
            ]
            volatility_metrics.append("整体隐含波动率均值")
            for parity_row, (_, expected_row) in zip(
                parity_rows,
                expected_parity.iterrows(),
            ):
                expected_values = [
                    *[expected_row[metric] for metric in premium_metrics],
                    *[expected_row[metric] for metric in volatility_metrics],
                ]
                actual_cells = [
                    *[parity_row[index] for index in range(1, 12)],
                    *[parity_row[index] for index in range(15, 26)],
                ]
                for cell, expected_value in zip(actual_cells, expected_values):
                    actual_value = cell.value
                    if pd.isna(expected_value):
                        if actual_value is not None:
                            raise RuntimeError("1.3分平价估值空值与计算结果不一致")
                    elif actual_value is None or not math.isclose(
                        float(actual_value),
                        float(expected_value),
                        rel_tol=1e-12,
                        abs_tol=1e-12,
                    ):
                        raise RuntimeError("1.3分平价估值与计算结果不一致")

        term_sheet = wb["1.4分期限估值"]
        term_rows = list(
            term_sheet.iter_rows(
                min_row=32,
                max_row=31 + len(summary_bond_rows),
                min_col=1,
                max_col=13,
            )
        )
        if len(term_rows) != len(summary_bond_rows):
            raise RuntimeError("1.4分期限估值历史数据行数不正确")
        for term_row, summary_row in zip(term_rows, summary_bond_rows):
            if term_row[0].data_type == "f" or pd.Timestamp(term_row[0].value) != pd.Timestamp(summary_row[0].value):
                raise RuntimeError("1.4分期限估值日期与1.1指标汇总不一致")
            for cell in term_row[1:]:
                if cell.data_type == "f":
                    raise RuntimeError("1.4分期限估值历史区域存在公式单元格")
                if cell.value is not None and not isinstance(cell.value, (int, float)):
                    raise RuntimeError("1.4分期限估值存在非数值单元格")
        term_number_format = term_sheet["D32"].number_format
        for column_number in range(2, 14):
            column = get_column_letter(column_number)
            if term_sheet[f"{column}32"].number_format != term_number_format:
                raise RuntimeError("1.4分期限估值数值格式不一致")
            for row_number in range(21, 29):
                cell = term_sheet[f"{column}{row_number}"]
                if cell.data_type != "f" or not isinstance(cell.value, str):
                    raise RuntimeError("1.4分期限估值摘要公式缺失")
                if cell.number_format != term_sheet[f"B{row_number}"].number_format:
                    raise RuntimeError("1.4分期限估值摘要格式不一致")
            if not isinstance(term_sheet[f"{column}29"].value, (int, float)):
                raise RuntimeError("1.4分期限估值个券数量缺失")
            if term_sheet[f"{column}29"].number_format != term_sheet["B29"].number_format:
                raise RuntimeError("1.4分期限估值个券数量格式不一致")
        if expected_data is not None:
            expected_term = (
                expected_data["term_valuation"]
                .sort_values("日期", ascending=False)
                .reset_index(drop=True)
            )
            if len(expected_term) != len(term_rows):
                raise RuntimeError("1.4分期限估值与计算结果行数不一致")
            term_metrics = [
                f"转股溢价率_{label}" for label, _, _ in TERM_BUCKETS
            ]
            for term_row, (_, expected_row) in zip(term_rows, expected_term.iterrows()):
                for cell, metric in zip(term_row[1:], term_metrics):
                    actual_value = cell.value
                    expected_value = expected_row[metric]
                    if pd.isna(expected_value):
                        if actual_value is not None:
                            raise RuntimeError("1.4分期限估值空值与计算结果不一致")
                    elif actual_value is None or not math.isclose(
                        float(actual_value),
                        float(expected_value),
                        rel_tol=1e-12,
                        abs_tol=1e-12,
                    ):
                        raise RuntimeError("1.4分期限估值与计算结果不一致")
            latest_term = expected_term.iloc[0]
            for column_number, (label, _, _) in enumerate(TERM_BUCKETS, start=2):
                column = get_column_letter(column_number)
                if int(term_sheet[f"{column}29"].value) != int(latest_term[f"个券数量_{label}"]):
                    raise RuntimeError("1.4分期限估值个券数量与计算结果不一致")

        if expected_data is not None:
            _verify_group_valuation_sheet(
                term_sheet,
                expected_data["term_valuation"],
                [label for label, _, _ in TERM_BUCKETS],
                [(1, "转股溢价率_")],
            )
            _verify_group_valuation_sheet(
                wb["1.5分评级估值"],
                expected_data["rating_valuation"],
                RATING_BUCKETS,
                [
                    (1, "转股溢价率_"),
                    (10, "隐含波动率_"),
                    (19, "YTM_"),
                    (28, "YTM信用利差_"),
                ],
            )
            _verify_group_valuation_sheet(
                wb["1.6分规模估值"],
                expected_data["size_valuation"],
                [label for label, _, _ in SIZE_BUCKETS],
                [
                    (1, "转股溢价率_"),
                    (8, "隐含波动率_"),
                ],
            )

            seasoning_sheet = wb["1.7次新券估值"]
            expected_seasoning = (
                expected_data["seasoning_valuation"]
                .sort_values("日期", ascending=False)
                .drop_duplicates("日期", keep="last")
                .reset_index(drop=True)
            )
            expected_headers = [
                "日期",
                "全部转债隐含波动率均值",
                "次新券隐含波动率均值",
                "次新券转股溢价率均值",
                "老券隐含波动率均值",
            ]
            actual_headers = [seasoning_sheet.cell(17, column).value for column in range(1, 6)]
            if actual_headers != expected_headers:
                raise RuntimeError("1.7次新券估值表头不正确")
            seasoning_metrics = expected_headers[1:]
            seasoning_rows = list(
                seasoning_sheet.iter_rows(
                    min_row=18,
                    max_row=17 + len(expected_seasoning),
                    min_col=1,
                    max_col=5,
                )
            )
            if len(seasoning_rows) != len(expected_seasoning):
                raise RuntimeError("1.7次新券估值历史数据行数不正确")
            for seasoning_row, (_, record) in zip(
                seasoning_rows,
                expected_seasoning.iterrows(),
            ):
                date_cell = seasoning_row[0]
                if date_cell.data_type == "f" or pd.Timestamp(date_cell.value) != pd.Timestamp(record["日期"]):
                    raise RuntimeError("1.7次新券估值日期与计算结果不一致")
                for column_number, metric in enumerate(seasoning_metrics, start=2):
                    cell = seasoning_row[column_number - 1]
                    if cell.data_type == "f":
                        raise RuntimeError("1.7次新券估值历史区域存在公式单元格")
                    actual_value = cell.value
                    expected_value = record[metric]
                    if pd.isna(expected_value):
                        if actual_value is not None:
                            raise RuntimeError("1.7次新券估值空值与计算结果不一致")
                    elif actual_value is None or not math.isclose(
                        float(actual_value),
                        float(expected_value),
                        rel_tol=1e-12,
                        abs_tol=1e-12,
                    ):
                        raise RuntimeError("1.7次新券估值与计算结果不一致")

            _verify_history_valuation_sheet(
                wb["1.8历史估值对比"],
                expected_data["history_valuation_daily"],
            )
            _verify_style_factor_sheet(
                wb["1.9转债风格因子"],
                expected_data["style"],
            )
            _verify_return_decomposition_sheet(
                wb["1.10回报拆解"],
                expected_data["decomp_timeseries"],
                expected_data["decomp_industry"],
                expected_data["decomp_parity"],
                expected_data["decomp_type"],
            )
            _verify_weekly_bond_movers_sheet(
                wb["1.11转债周度涨跌幅个券"],
                expected_data["weekly_bond_movers"],
            )
            _verify_redemption_trigger_sheet(
                wb["2.1即将触发赎回"],
                expected_data["redemption_trigger_candidates"],
            )
    finally:
        wb.close()

    with zipfile.ZipFile(path, "r") as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        calc_properties = workbook_root.find(_qname(_SHEET_NS, "calcPr"))
        if calc_properties is not None:
            if calc_properties.get("fullCalcOnLoad") in {"1", "true", "True"}:
                raise RuntimeError("工作簿仍启用了fullCalcOnLoad")
            if calc_properties.get("forceFullCalc") in {"1", "true", "True"}:
                raise RuntimeError("工作簿仍启用了forceFullCalc")
        if "xl/calcChain.xml" in archive.namelist():
            summary_sheet_ids = {
                _sheet_id(archive, "1.1指标汇总"),
                _sheet_id(archive, "1.1指标汇总_剔除极值"),
            }
            parity_sheet_id = _sheet_id(archive, "1.3分平价估值")
            term_sheet_id = _sheet_id(archive, "1.4分期限估值")
            rating_sheet_id = _sheet_id(archive, "1.5分评级估值")
            size_sheet_id = _sheet_id(archive, "1.6分规模估值")
            seasoning_sheet_id = _sheet_id(archive, "1.7次新券估值")
            history_sheet_id = _sheet_id(archive, "1.8历史估值对比")
            style_sheet_id = _sheet_id(archive, "1.9转债风格因子")
            decomposition_sheet_id = _sheet_id(archive, "1.10回报拆解")
            redemption_trigger_sheet_id = _sheet_id(archive, "2.1即将触发赎回")
            calc_chain_root = ET.fromstring(archive.read("xl/calcChain.xml"))
            current_sheet_id: str | None = None
            for cell in calc_chain_root:
                if cell.get("i") is not None:
                    current_sheet_id = str(cell.get("i"))
                match = re.fullmatch(r"([A-Z]+)(\d+)", cell.get("r", ""))
                if (
                    current_sheet_id in summary_sheet_ids
                    and match is not None
                    and match.group(1) in {"B", "T"}
                    and int(match.group(2)) >= 14
                ):
                    raise RuntimeError("计算链仍包含1.1页面中已转为数值的B/T单元格")
                if (
                    current_sheet_id == parity_sheet_id
                    and match is not None
                    and 1 <= column_index_from_string(match.group(1)) <= 26
                    and int(match.group(2)) >= 32
                ):
                    raise RuntimeError("计算链仍包含1.3分平价估值中已转为数值的单元格")
                if (
                    current_sheet_id == term_sheet_id
                    and match is not None
                    and 1 <= column_index_from_string(match.group(1)) <= 13
                    and int(match.group(2)) >= 32
                ):
                    raise RuntimeError("计算链仍包含1.4分期限估值中已转为数值的单元格")
                if (
                    current_sheet_id == rating_sheet_id
                    and match is not None
                    and 1 <= column_index_from_string(match.group(1)) <= 35
                    and int(match.group(2)) >= 32
                ):
                    raise RuntimeError("计算链仍包含1.5分评级估值中已转为数值的单元格")
                if (
                    current_sheet_id == size_sheet_id
                    and match is not None
                    and 1 <= column_index_from_string(match.group(1)) <= 13
                    and int(match.group(2)) >= 32
                ):
                    raise RuntimeError("计算链仍包含1.6分规模估值中已转为数值的单元格")
                if (
                    current_sheet_id == seasoning_sheet_id
                    and match is not None
                    and 1 <= column_index_from_string(match.group(1)) <= 5
                    and int(match.group(2)) >= 18
                ):
                    raise RuntimeError("计算链仍包含1.7次新券估值中已转为数值的单元格")
                if (
                    current_sheet_id == style_sheet_id
                    and match is not None
                    and 1 <= column_index_from_string(match.group(1)) <= 19
                    and int(match.group(2)) >= 2
                ):
                    raise RuntimeError("计算链仍包含1.9转债风格因子中已转为数值的单元格")
                if (
                    current_sheet_id == history_sheet_id
                    and match is not None
                    and 15 <= column_index_from_string(match.group(1)) <= 35
                    and int(match.group(2)) >= 2
                ):
                    raise RuntimeError("计算链仍包含1.8历史估值对比中已转为数值的单元格")
                if (
                    current_sheet_id == decomposition_sheet_id
                    and match is not None
                    and (
                        1 <= column_index_from_string(match.group(1)) <= 5
                        or 9 <= column_index_from_string(match.group(1)) <= 13
                    )
                    and int(match.group(2)) >= 2
                ):
                    raise RuntimeError("计算链仍包含1.10回报拆解中已转为数值的单元格")
                if (
                    current_sheet_id == redemption_trigger_sheet_id
                    and match is not None
                    and 1 <= column_index_from_string(match.group(1)) <= 12
                    and int(match.group(2)) >= 2
                ):
                    raise RuntimeError("计算链仍包含2.1即将触发赎回的旧公式引用")
        parity_sheet_xml = ET.fromstring(
            archive.read(_sheet_xml_path(archive, "1.3分平价估值"))
        )
        parity_sheet_data = parity_sheet_xml.find(_qname(_SHEET_NS, "sheetData"))
        if parity_sheet_data is None:
            raise RuntimeError("1.3分平价估值缺少sheetData")
        for row in parity_sheet_data.findall(_qname(_SHEET_NS, "row")):
            if int(row.get("r", "0")) < 21:
                continue
            column_indexes = [
                column_index_from_string(_cell_column(cell.get("r", "")))
                for cell in row.findall(_qname(_SHEET_NS, "c"))
            ]
            if column_indexes != sorted(set(column_indexes)):
                raise RuntimeError("1.3分平价估值单元格顺序不符合Excel规范")
        for sheet_name, start_row in (
            ("1.4分期限估值", 20),
            ("1.5分评级估值", 20),
            ("1.6分规模估值", 20),
            ("1.7次新券估值", 17),
            ("1.8历史估值对比", 1),
            ("1.9转债风格因子", 1),
            ("1.10回报拆解", 1),
            ("2.1即将触发赎回", 1),
        ):
            sheet_xml = ET.fromstring(
                archive.read(_sheet_xml_path(archive, sheet_name))
            )
            sheet_data = sheet_xml.find(_qname(_SHEET_NS, "sheetData"))
            if sheet_data is None:
                raise RuntimeError(f"{sheet_name}缺少sheetData")
            row_numbers = [
                int(row.get("r", "0"))
                for row in sheet_data.findall(_qname(_SHEET_NS, "row"))
            ]
            if row_numbers != sorted(set(row_numbers)):
                raise RuntimeError(f"{sheet_name}行顺序不符合Excel规范")
            for row in sheet_data.findall(_qname(_SHEET_NS, "row")):
                if int(row.get("r", "0")) < start_row:
                    continue
                column_indexes = [
                    column_index_from_string(_cell_column(cell.get("r", "")))
                    for cell in row.findall(_qname(_SHEET_NS, "c"))
                ]
                if column_indexes != sorted(set(column_indexes)):
                    raise RuntimeError(f"{sheet_name}单元格顺序不符合Excel规范")


def save_meta(data: dict[str, pd.DataFrame]) -> None:
    market = data["market"]
    meta = {
        "start": market["日期"].min().strftime("%Y-%m-%d"),
        "end": market["日期"].max().strftime("%Y-%m-%d"),
        "rows": int(len(market)),
        "output": str(OUTPUT_XLSX),
        "sheets_updated": [
            "1.1指标汇总",
            "1.1指标汇总_剔除极值",
            "主要估值指标",
            "债性指标",
            "1.3分平价估值",
            "1.4分期限估值",
            "1.5分评级估值",
            "1.6分规模估值",
            "1.7次新券估值",
            "1.8历史估值对比",
            "1.9转债风格因子",
            "1.10回报拆解",
            "1.11转债周度涨跌幅个券",
            "2.1即将触发赎回",
        ],
        "latest_return_decomposition": data["decomp_timeseries"].head(1).replace({np.nan: None}).to_dict("records"),
    }
    META_JSON.write_text(json.dumps(meta, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def main() -> int:
    started = time.perf_counter()
    try:
        log("开始计算 PA 周报数据...")
        data = calculate_all()
        log("开始写入 Excel...")
        write_workbook(data)
        verify_workbook(OUTPUT_XLSX, data)
        save_meta(data)
    except Exception as exc:
        print(f"PA 周报生成失败：{exc}", file=sys.stderr)
        return 1

    elapsed = time.perf_counter() - started
    log(f"PA 周报生成完成，用时 {format_elapsed(elapsed)}：{OUTPUT_XLSX}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
