"""华创转债周报独立一键生成脚本（Parquet Schema v2.1）。

正常运行：
    py 华创转债周报自动更新.py

离线验证（不调用 Wind / iFinD）：
    py 华创转债周报自动更新.py --offline

默认自动选择最近一个完整周，报告日期取该周结束后的下一个周一。
也可显式指定：
    py 华创转债周报自动更新.py --week-end 2026-08-21 --report-date 2026-08-24

默认在 ``【华创】转债周报YYYYMMDD`` 文件夹生成4张PNG、完整周报TXT和
包含“图表25”“图表30”“数据来源”三个工作表的Excel底稿。脚本不导入、
不调用 ``人保周报自动更新.py`` 或其他项目脚本；iFinD不可用时自动使用
本地Parquet、公告统计簿和最近一期华创底稿降级运行。
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import subprocess
import tempfile
import time
import warnings
from configparser import ConfigParser
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlencode

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
from matplotlib.ticker import FormatStrFormatter, FuncFormatter, MultipleLocator
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from scipy.optimize import OptimizeWarning, curve_fit


ROOT = Path(__file__).resolve().parents[2]
IFIND_CREDENTIAL_FILE = ROOT / "private/ifind账号.txt"
PARQUET_ROOT = ROOT / "data/转债个券历史序列"
MASTER_PATH = PARQUET_ROOT / "_special" / "总表.parquet"
INDEX_PATH = PARQUET_ROOT / "_special" / "指数.parquet"
REDEMPTION_BOOK = ROOT / "data/clauses/【华创固收】赎回和不赎回公告统计.xlsx"
DOWN_REVISION_BOOK = ROOT / "data/clauses/【华创固收】下修和不下修公告统计.xlsx"

BOND_CODE = "转债代码"
TRADE_DATE = "交易日期"
INDEX_NAME = "指数名称"
INDEX_VALUE = "指数值"

RED = "#E6121B"
BLUE = "#0262BA"
GRAY = "#A6A6A6"
LIGHT_RED = "#E7B8B8"
START_CHART = pd.Timestamp("2023-01-01")


def load_ifind_credentials() -> tuple[str, str]:
    """从项目目录的 ifind账号.txt 读取统一登录账号。"""
    if not IFIND_CREDENTIAL_FILE.is_file():
        raise FileNotFoundError(f"未找到iFinD账号文件：{IFIND_CREDENTIAL_FILE}")
    config = ConfigParser(interpolation=None)
    config.read(IFIND_CREDENTIAL_FILE, encoding="utf-8")
    username = config.get("ifind", "username", fallback="").strip()
    password = config.get("ifind", "password", fallback="").strip()
    if not username or not password:
        raise RuntimeError("ifind账号.txt中的[ifind] username或password为空")
    return username, password


def print_ifind_usage(module: object) -> None:
    """显示iFinD各数据项的已用额度比例。"""
    try:
        result = module.THS_DataStatistics()
        tables = result.get("tables", {}) if isinstance(result, dict) else {}
        if not tables:
            detail = result.get("errmsg", "未返回额度数据") if isinstance(result, dict) else str(result)
            print(f"[警告] iFinD使用额度查询失败：{detail}")
            return
        print("iFinD使用额度：")
        for key, value in tables.items():
            ratio = value.get("ratio", "N/A") if isinstance(value, dict) else value
            print(f"{key} 已用：{ratio}")
    except Exception as exc:
        print(f"[警告] iFinD使用额度查询失败：{exc}")

FIGURE2_INDICES = ["沪深300", "中证500", "中证1000", "中证2000", "转债指数"]
FUND_INDICES = ["混合债券型一级基金", "混合债券型二级基金", "可转换债券型基金", "灵活配置型基金"]
LOCAL_METRICS = [
    "余额",
    "收盘价",
    "平价",
    "转股溢价率",
    "换手率",
    "成交额",
    "涨跌幅",
    "剩余期限",
    "赎回累计天数",
    "下修累计天数",
    "平价底价溢价率",
    "主体评级",
    "债项评级",
    "正股市值",
]

PROPOSAL_COLUMNS = [
    "最新公告日期",
    "公司代码",
    "公司名称",
    "方案进度",
    "发行方式",
    "发行规模(亿元)",
    "发行期限(年)",
    "预案公告日",
    "股东大会公告日",
    "发审委审核公告日",
    "证监会批准公告日",
    "申万行业",
]


def log(message: str) -> None:
    print(message, flush=True)


def chinese_join(values: Iterable[str]) -> str:
    return "、".join(str(value) for value in values if str(value).strip())


def month_day(ts: pd.Timestamp) -> str:
    return f"{ts.month}月{ts.day}日"


def date_range_text(start: pd.Timestamp, end: pd.Timestamp) -> str:
    return f"{month_day(start)}至{month_day(end)}"


def monthly_files() -> list[Path]:
    return sorted(
        path
        for year_dir in PARQUET_ROOT.iterdir()
        if year_dir.is_dir() and year_dir.name.isdigit()
        for path in year_dir.glob("*.parquet")
        if re.fullmatch(r"\d{6}", path.stem)
    )


def load_master() -> pd.DataFrame:
    master = pd.read_parquet(MASTER_PATH)
    if BOND_CODE not in master:
        raise ValueError(f"{MASTER_PATH}缺少字段：{BOND_CODE}")
    master[BOND_CODE] = master[BOND_CODE].astype("string")
    for column in ["上市日期", "最后交易日", "摘牌日期", "到期日期", "发行日期", "赎回公告日", "转股期起始日"]:
        if column in master:
            master[column] = pd.to_datetime(master[column], errors="coerce")
    for column in ["发行规模", "赎回触发比例", "下修触发比例"]:
        if column in master:
            master[column] = pd.to_numeric(master[column], errors="coerce")
    return master.set_index(BOND_CODE, drop=False)


def load_panel(start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    parts: list[pd.DataFrame] = []
    columns = [BOND_CODE, TRADE_DATE, *LOCAL_METRICS]
    for path in monthly_files():
        year_month = pd.Timestamp(f"{path.stem[:4]}-{path.stem[4:]}-01")
        if year_month > end.to_period("M").to_timestamp() or year_month.to_period("M") < start.to_period("M"):
            continue
        frame = pd.read_parquet(path, columns=columns)
        frame[TRADE_DATE] = pd.to_datetime(frame[TRADE_DATE])
        frame = frame[frame[TRADE_DATE].between(start, end)]
        if not frame.empty:
            parts.append(frame)
    if not parts:
        raise FileNotFoundError(f"{start:%Y-%m-%d}至{end:%Y-%m-%d}没有月度parquet数据")
    panel = pd.concat(parts, ignore_index=True)
    panel[BOND_CODE] = panel[BOND_CODE].astype("string")
    panel = panel.sort_values([TRADE_DATE, BOND_CODE], kind="stable")
    if panel.duplicated([BOND_CODE, TRADE_DATE]).any():
        raise ValueError("月度parquet存在重复的转债代码＋交易日期")
    return panel.reset_index(drop=True)


def load_indices(start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    frame = pd.read_parquet(INDEX_PATH)
    required = {INDEX_NAME, TRADE_DATE, INDEX_VALUE}
    if not required.issubset(frame.columns):
        raise ValueError(f"{INDEX_PATH}不是指数长表Schema")
    frame[TRADE_DATE] = pd.to_datetime(frame[TRADE_DATE])
    frame[INDEX_VALUE] = pd.to_numeric(frame[INDEX_VALUE], errors="coerce")
    frame = frame[frame[TRADE_DATE].between(start, end)].dropna(subset=[INDEX_VALUE])
    return frame.pivot(index=TRADE_DATE, columns=INDEX_NAME, values=INDEX_VALUE).sort_index()


def all_panel_dates() -> pd.DatetimeIndex:
    paths = monthly_files()
    if not paths:
        raise FileNotFoundError(f"未找到月度parquet：{PARQUET_ROOT}")
    dates = pd.read_parquet(paths[-1], columns=[TRADE_DATE])[TRADE_DATE]
    return pd.DatetimeIndex(pd.to_datetime(dates).drop_duplicates().sort_values())


def choose_period(explicit_week_end: str | None, explicit_report_date: str | None) -> tuple[pd.Timestamp, pd.Timestamp, pd.Timestamp]:
    dates = all_panel_dates()
    latest = pd.Timestamp(dates.max()).normalize()
    today = pd.Timestamp.today().normalize()
    if explicit_week_end:
        week_end = pd.Timestamp(explicit_week_end).normalize()
        if week_end not in dates:
            candidates = dates[dates <= week_end]
            if len(candidates) == 0:
                raise ValueError(f"指定周末{week_end:%Y-%m-%d}之前没有交易日")
            week_end = pd.Timestamp(candidates.max()).normalize()
    else:
        latest_week_start = latest - pd.Timedelta(days=latest.weekday())
        current_week_complete = latest.weekday() >= 4 or today >= latest_week_start + pd.Timedelta(days=5)
        if current_week_complete:
            week_start = latest_week_start
        else:
            week_start = latest_week_start - pd.Timedelta(days=7)
        candidates = dates[(dates >= week_start) & (dates < week_start + pd.Timedelta(days=7))]
        if len(candidates) == 0:
            raise ValueError("无法识别最近完整周")
        week_end = pd.Timestamp(candidates.max()).normalize()
    week_start = week_end - pd.Timedelta(days=week_end.weekday())
    report_date = (
        pd.Timestamp(explicit_report_date).normalize()
        if explicit_report_date
        else week_start + pd.Timedelta(days=7)
    )
    return week_start, week_end, report_date


def prior_base_date(index: pd.DatetimeIndex, week_start: pd.Timestamp) -> pd.Timestamp:
    candidates = index[index < week_start]
    if len(candidates) == 0:
        raise ValueError("没有找到本周开始前的基准交易日")
    return pd.Timestamp(candidates.max()).normalize()


def inverse_cubic(x, a, b, c, d):
    return a / np.power(x, 3) + b / np.power(x, 2) + c / x + d


def fit_premium_at_100(sample: pd.DataFrame) -> float:
    data = sample[["平价", "转股溢价率", "换手率"]].apply(pd.to_numeric, errors="coerce")
    data = data.replace(0, np.nan).dropna()
    data = data[data["平价"].between(70, 130, inclusive="both") & (data["换手率"] <= 50)]
    if len(data) < 8:
        return float("nan")
    low, high = data["转股溢价率"].quantile([0.03, 0.97])
    data = data[(data["转股溢价率"] > low) & (data["转股溢价率"] < high)]
    if len(data) < 8:
        return float("nan")
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", OptimizeWarning)
            params, _ = curve_fit(
                inverse_cubic,
                data["平价"].to_numpy(float),
                data["转股溢价率"].to_numpy(float),
                maxfev=20000,
            )
        return float(inverse_cubic(100.0, *params))
    except Exception:
        return float("nan")


def valid_listing_mask(frame: pd.DataFrame, master: pd.DataFrame, day: pd.Timestamp) -> pd.Series:
    codes = frame[BOND_CODE]
    listing = master["上市日期"].reindex(codes).reset_index(drop=True)
    last_trade = master["最后交易日"].reindex(codes).reset_index(drop=True)
    return (
        (listing.isna() | (listing <= day))
        & (last_trade.isna() | (last_trade >= day))
        & frame["收盘价"].notna().reset_index(drop=True)
    )


def build_daily_valuation(panel: pd.DataFrame, master: pd.DataFrame, end: pd.Timestamp) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for day, raw in panel[panel[TRADE_DATE] >= START_CHART].groupby(TRADE_DATE, sort=True):
        day = pd.Timestamp(day)
        sample = raw.reset_index(drop=True)
        valid = valid_listing_mask(sample, master, day)
        sample = sample.loc[valid.to_numpy()].copy()
        rows.append(
            {
                "日期": day,
                "转债价格中位数": pd.to_numeric(sample["收盘价"], errors="coerce").median(),
                "百元平价拟合溢价率": fit_premium_at_100(sample),
            }
        )
    return pd.DataFrame(rows).sort_values("日期").drop_duplicates("日期", keep="last")


def expanded_limits(values: pd.Series, base_min: float, base_max: float, step: float) -> tuple[float, float]:
    minimum = float(pd.to_numeric(values, errors="coerce").min())
    maximum = float(pd.to_numeric(values, errors="coerce").max())
    low = base_min if minimum >= base_min else math.floor(minimum / step) * step - step
    high = base_max if maximum <= base_max else math.ceil(maximum / step) * step + step
    return low, high


def setup_fonts() -> tuple[FontProperties, FontProperties]:
    chinese = Path(r"C:\Windows\Fonts\simsun.ttc")
    latin = Path(r"C:\Windows\Fonts\times.ttf")
    if not chinese.exists() or not latin.exists():
        raise FileNotFoundError("未找到宋体或Times New Roman字体")
    return FontProperties(fname=str(chinese), size=13), FontProperties(fname=str(latin), size=12.5)


def figure1(indices: pd.DataFrame, output: Path) -> None:
    data = indices.loc[:, ["转债指数", "万得全A"]].dropna()
    chinese_font, latin_font = setup_fonts()
    fig, left = plt.subplots(figsize=(8, 4.75), dpi=250)
    right = left.twinx()
    fig.patch.set_facecolor("white")
    line_cb, = left.plot(data.index, data["转债指数"], color=RED, linewidth=2.2, label="中证转债")
    line_wind, = right.plot(data.index, data["万得全A"], color=BLUE, linewidth=2.2, label="万得全A")
    left.set_xlim(data.index.min(), data.index.max())
    left.set_ylim(*expanded_limits(data["转债指数"], 200, 600, 50))
    right.set_ylim(*expanded_limits(data["万得全A"], 3000, 7500, 500))
    left.yaxis.set_major_locator(MultipleLocator(50))
    right.yaxis.set_major_locator(MultipleLocator(500))
    right.yaxis.set_major_formatter(FuncFormatter(lambda value, _: f"{value:,.0f}"))
    ticks = [data.index.min().replace(year=year) for year in range(data.index.min().year, data.index.max().year + 1)]
    left.set_xticks(ticks)
    left.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
    for axis in (left, right):
        axis.grid(False)
        axis.spines["top"].set_visible(False)
        axis.tick_params(axis="both", labelsize=12.5, length=4, width=0.8)
        for label in axis.get_xticklabels() + axis.get_yticklabels():
            label.set_fontproperties(latin_font)
    left.spines["right"].set_visible(False)
    right.spines["left"].set_visible(False)
    legend = left.legend(
        [line_cb, line_wind], ["中证转债", "万得全A"], loc="upper center",
        bbox_to_anchor=(0.5, -0.17), ncol=2, frameon=False, handlelength=3.2,
        columnspacing=2.2, prop=chinese_font,
    )
    for handle in legend.legend_handles:
        handle.set_linewidth(2.5)
    fig.subplots_adjust(left=0.105, right=0.895, top=0.96, bottom=0.25)
    fig.savefig(output, dpi=250, facecolor="white")
    plt.close(fig)


def weekly_returns(indices: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> tuple[pd.Timestamp, pd.Series]:
    data = indices.loc[indices.index <= week_end]
    base = prior_base_date(data.index, week_start)
    end_candidates = data.index[data.index <= week_end]
    end = pd.Timestamp(end_candidates.max())
    missing = [name for name in FIGURE2_INDICES if name not in data.columns]
    if missing:
        raise ValueError(f"指数parquet缺少：{missing}")
    result = (data.loc[end, FIGURE2_INDICES] / data.loc[base, FIGURE2_INDICES] - 1) * 100
    result.index = ["沪深300", "中证500", "中证1000", "中证2000", "中证转债"]
    return base, result


def figure2(returns: pd.Series, output: Path) -> None:
    chinese_font, latin_font = setup_fonts()
    fig, axis = plt.subplots(figsize=(8, 4.75), dpi=250)
    x = np.arange(len(returns))
    bars = axis.bar(x, returns.to_numpy(), width=0.52, color=RED, edgecolor="none")
    axis.axhline(0, color="black", linewidth=0.8)
    axis.set_xticks(x, returns.index.tolist())
    for label in axis.get_xticklabels():
        label.set_fontproperties(chinese_font)
    for label in axis.get_yticklabels():
        label.set_fontproperties(latin_font)
    minimum, maximum = min(0.0, float(returns.min())), max(0.0, float(returns.max()))
    span = max(maximum - minimum, 1.0)
    lower = 0.0 if minimum >= 0 else math.floor((minimum - span * 0.12) / 2) * 2
    upper = 0.0 if maximum <= 0 else max(2.0, math.ceil((maximum + span * 0.18) / 2) * 2)
    axis.set_ylim(lower, upper)
    axis.yaxis.set_major_locator(MultipleLocator(2))
    axis.tick_params(axis="both", labelsize=12.5, length=4, width=0.8)
    offset = span * 0.035
    for bar, value in zip(bars, returns.to_numpy()):
        axis.text(
            bar.get_x() + bar.get_width() / 2,
            value + offset if value >= 0 else value - offset,
            f"{value:.2f}", ha="center", va="bottom" if value >= 0 else "top",
            fontproperties=latin_font, color="black",
        )
    axis.grid(False)
    axis.spines["top"].set_visible(False)
    axis.spines["right"].set_visible(False)
    axis.margins(x=0.08)
    fig.subplots_adjust(left=0.105, right=0.97, top=0.95, bottom=0.17)
    fig.savefig(output, dpi=250, facecolor="white")
    plt.close(fig)


def two_month_ticks(start: pd.Timestamp, end: pd.Timestamp) -> list[pd.Timestamp]:
    tick = pd.Timestamp(start.year, start.month, 3)
    if tick < start:
        tick += pd.DateOffset(months=2)
    ticks: list[pd.Timestamp] = []
    while tick <= end:
        ticks.append(tick)
        tick += pd.DateOffset(months=2)
    return ticks


def nice_ylim(values: pd.Series, step: float, zero_floor: bool) -> tuple[float, float]:
    finite = pd.to_numeric(values, errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
    low, high = float(finite.min()), float(finite.max())
    pad = max(step * 0.25, (high - low) * 0.05)
    lower = math.floor((low - pad) / step) * step
    upper = math.ceil((high + pad) / step) * step
    if zero_floor:
        lower = max(0.0, lower)
    return lower, upper if upper > lower else lower + step


def valuation_figure(data: pd.DataFrame, column: str, legend_name: str, output: Path, y_step: float, zero_floor: bool) -> dict[str, float]:
    series = data.set_index("日期")[column].dropna().sort_index()
    quantiles = series.quantile([0.25, 0.50, 0.75])
    plt.rcParams.update({"font.family": ["SimSun", "Times New Roman"], "axes.unicode_minus": False, "font.size": 20})
    fig, ax = plt.subplots(figsize=(14.25, 8.535), dpi=200)
    ax.plot(series.index, series.values, color=RED, linewidth=3.8, label=legend_name, zorder=4)
    for q, color, label in [(0.25, BLUE, "25%"), (0.50, GRAY, "50%"), (0.75, LIGHT_RED, "75%")]:
        ax.axhline(float(quantiles.loc[q]), color=color, linewidth=3.5, linestyle=(0, (4.5, 4.5)), dash_capstyle="round", label=label)
    ax.set_xlim(series.index.min(), series.index.max())
    ax.set_xticks(two_month_ticks(series.index.min(), series.index.max()))
    ax.xaxis.set_major_formatter(FuncFormatter(lambda value, _: f"{mdates.num2date(value).year}/{mdates.num2date(value).month}/{mdates.num2date(value).day}"))
    plt.setp(ax.get_xticklabels(), rotation=48, ha="right", rotation_mode="anchor")
    all_y = pd.concat([series, pd.Series(quantiles.values)])
    ax.set_ylim(*nice_ylim(all_y, y_step, zero_floor))
    ax.yaxis.set_major_locator(MultipleLocator(y_step))
    ax.yaxis.set_major_formatter(FormatStrFormatter("%.2f"))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(axis="both", labelsize=20, width=1.2, length=6, pad=8)
    ax.grid(False)
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.25), ncol=4, frameon=False, fontsize=21, handlelength=4.5, columnspacing=2.0)
    fig.subplots_adjust(left=0.10, right=0.985, top=0.965, bottom=0.31)
    fig.savefig(output, dpi=200, facecolor="white")
    plt.close(fig)
    return {
        "最新值": float(series.iloc[-1]),
        "25%分位数": float(quantiles.loc[0.25]),
        "50%分位数": float(quantiles.loc[0.50]),
        "75%分位数": float(quantiles.loc[0.75]),
    }


@dataclass
class IFindSession:
    active: bool = False
    module: object | None = None

    def open(self, auto: bool, strict: bool) -> bool:
        if not auto:
            return False
        username, password = load_ifind_credentials()
        try:
            import iFinDPy as module
            result = module.THS_iFinDLogin(username, password)
            code = int(result) if isinstance(result, (int, np.integer)) else int(getattr(result, "errorcode", -1))
            if code not in {0, -201}:
                raise RuntimeError(f"iFinD登录失败，错误码{code}")
            self.active, self.module = True, module
            print_ifind_usage(module)
            return True
        except Exception as exc:
            if strict:
                raise
            log(f"[警告] iFinD实时接口不可用，改用本地数据：{exc}")
            return False

    def close(self) -> None:
        if self.active and self.module is not None:
            try:
                self.module.THS_iFinDLogout()
            except Exception:
                pass
        self.active = False


def normalize_ifind_frame(
    result,
    api: str,
    *,
    allow_no_data: bool = False,
    empty_columns: Iterable[str] = (),
) -> pd.DataFrame:
    code = int(getattr(result, "errorcode", -1))
    if allow_no_data and code == -4001:
        return pd.DataFrame(columns=list(empty_columns))
    if code != 0:
        raise RuntimeError(f"{api}失败：{code} {getattr(result, 'errmsg', '')}")
    data = getattr(result, "data", None)
    if not isinstance(data, pd.DataFrame):
        raise RuntimeError(f"{api}未返回DataFrame")
    return data.copy()


def fetch_proposals(session: IFindSession, start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    fields = (
        "p03153_f029:Y,p03153_f001:Y,p03153_f002:Y,p03153_f003:Y,p03153_f004:Y,"
        "p03153_f005:Y,p03153_f006:Y,p03153_f012:Y,p03153_f013:Y,p03153_f023:Y,"
        "p03153_f024:Y,p03153_f026:Y"
    )
    params = f"STARTTRADEDATE={start:%Y%m%d};ENDTRADEDATE={end:%Y%m%d};PROJECTKIND=全部;fxqk=否"
    raw = normalize_ifind_frame(session.module.THS_DR("p03153", params, fields, "format:dataframe"), "THS_DR(p03153)")
    mapping = dict(zip(
        ["p03153_f029", "p03153_f001", "p03153_f002", "p03153_f003", "p03153_f004", "p03153_f005", "p03153_f006", "p03153_f012", "p03153_f013", "p03153_f023", "p03153_f024", "p03153_f026"],
        PROPOSAL_COLUMNS,
    ))
    missing = set(mapping) - set(raw.columns)
    if missing:
        raise RuntimeError(f"p03153缺少字段：{sorted(missing)}")
    return standardize_proposals(raw[list(mapping)].rename(columns=mapping))


def standardize_proposals(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    for column in PROPOSAL_COLUMNS:
        if column not in result:
            result[column] = pd.NA
    result = result[PROPOSAL_COLUMNS]
    result = result[~result["发行方式"].astype("string").str.contains("定向", na=False)]
    for column in ["最新公告日期", "预案公告日", "股东大会公告日", "发审委审核公告日", "证监会批准公告日"]:
        result[column] = pd.to_datetime(result[column], format="mixed", errors="coerce")
    for column in ["发行规模(亿元)", "发行期限(年)"]:
        result[column] = pd.to_numeric(result[column], errors="coerce")
    result = result.sort_values(["最新公告日期", "公司代码"], ascending=[False, True], na_position="last")
    # p03153扩大查询窗口后可能返回同一公司的历史状态；仅保留最新公告对应记录。
    result = result.drop_duplicates("公司代码", keep="first")
    return result.reset_index(drop=True)


def load_proposal_fallback(report_date: pd.Timestamp, output_dir: Path) -> pd.DataFrame:
    candidates = [
        output_dir / f"转债周报图表底稿{report_date:%Y%m%d}.xlsx",
        ROOT / "runs" / "weekly" / f"发行预案{report_date:%Y%m%d}.xlsx",
    ]
    candidates.extend(sorted((ROOT / "runs" / "weekly").glob("【华创】转债周报*/转债周报图表底稿*.xlsx"), reverse=True))
    candidates.extend(sorted(ROOT.glob("outputs/**/转债发行预案_剔除定向_*.xlsx"), reverse=True))
    for path in candidates:
        if not path.exists():
            continue
        try:
            sheets = pd.ExcelFile(path).sheet_names
            sheet = "发行预案" if "发行预案" in sheets else ("原始发行预案" if "原始发行预案" in sheets else sheets[0])
            frame = pd.read_excel(path, sheet_name=sheet)
            if {"公司名称", "方案进度", "发行规模(亿元)"}.issubset(frame.columns):
                log(f"[本地] 使用发行预案：{path.name}")
                return standardize_proposals(frame)
        except Exception:
            continue
    return pd.DataFrame(columns=PROPOSAL_COLUMNS)


def fetch_market_overview(session: IFindSession, as_of: pd.Timestamp) -> tuple[pd.DataFrame, pd.DataFrame]:
    dr = normalize_ifind_frame(
        session.module.THS_DR(
            "p00570",
            f"jyzt=未到期;sfdb=全部;jysc=全部;edate={as_of:%Y%m%d}",
            "jydm:Y,jydm_mc:Y,p00570_f001:Y,p00570_f019:Y",
            "format:dataframe",
        ),
        "THS_DR(p00570)",
    )
    codes = ",".join(dr["jydm"].astype(str))
    raw = normalize_ifind_frame(
        session.module.THS_BD(
            codes,
            "ths_convertible_debt_short_name_cbond;ths_stock_code_cbond;ths_stock_short_name_cbond;"
            "ths_issue_method_cbond;ths_trading_status_bond;ths_bond_balance_cbond;ths_listed_date_cbond",
            f";;;;;{as_of:%Y-%m-%d};",
        ),
        "THS_BD(转债基础信息)",
    )
    rename = {
        "thscode": BOND_CODE,
        "ths_convertible_debt_short_name_cbond": "转债名称",
        "ths_stock_code_cbond": "正股代码",
        "ths_stock_short_name_cbond": "公司名称",
        "ths_issue_method_cbond": "发行方式",
        "ths_trading_status_bond": "交易状态",
        "ths_bond_balance_cbond": "余额",
        "ths_listed_date_cbond": "上市日期",
    }
    basic = raw.rename(columns=rename)
    basic = basic[~basic["发行方式"].astype("string").str.contains("定向", na=False)]
    basic = basic[~basic[BOND_CODE].astype("string").str.contains("NQ", na=False)]
    basic = basic[~basic["交易状态"].astype("string").str.contains("终止上市", na=False)]
    basic["余额"] = pd.to_numeric(basic["余额"], errors="coerce")
    basic["上市日期"] = pd.to_datetime(basic["上市日期"], errors="coerce")
    upcoming = normalize_ifind_frame(
        session.module.THS_DR("p00600", "zqlx=640007", "p00600_f001:Y,p00600_f004:Y,p00600_f044:Y", "format:dataframe"),
        "THS_DR(p00600)",
        allow_no_data=True,
        empty_columns=["p00600_f001", "p00600_f004", "p00600_f044"],
    )
    return basic, upcoming


def local_market_overview(panel: pd.DataFrame, master: pd.DataFrame, as_of: pd.Timestamp) -> tuple[pd.DataFrame, pd.DataFrame]:
    snap = panel[panel[TRADE_DATE].eq(as_of)].set_index(BOND_CODE)
    result = master.copy()
    result["余额"] = pd.to_numeric(snap["余额"].reindex(result.index), errors="coerce")
    issued = result["发行日期"].notna() & (result["发行日期"] <= as_of)
    unlisted = issued & (result["上市日期"].isna() | (result["上市日期"] > as_of))
    # 存续数量按“已发行且尚未摘牌”统计；最后交易日至摘牌日期之间仍属于未到期存续券。
    delisted = result["摘牌日期"].notna() & (result["摘牌日期"] <= as_of)
    keep = issued & ~delisted
    basic = result.loc[keep, [BOND_CODE, "转债名称", "上市日期", "余额"]].copy()
    basic["公司名称"] = basic["转债名称"]
    basic["交易状态"] = np.where(basic["上市日期"].isna() | (basic["上市日期"] > as_of), "已发行未上市", "交易")
    basic["发行方式"] = ""
    return basic.reset_index(drop=True), pd.DataFrame()


def fetch_delist_dates_wind(codes: list[str], timeout_seconds: int = 300) -> pd.Series:
    import win32com.client
    with tempfile.TemporaryDirectory(prefix="renbao_wind_") as temp_dir:
        workbook_path = Path(temp_dir) / "delist.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.append(["wind_code", "s_info_delistdate"])
        for row, code in enumerate(codes, 2):
            sheet.cell(row=row, column=1, value=code)
            sheet.cell(row=row, column=2, value=f'=WSS(A{row},"s_info_delistdate")')
        book.save(workbook_path)
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = None
        try:
            wb = excel.Workbooks.Open(str(workbook_path))
            excel.CalculateFullRebuild()
            try:
                excel.CalculateUntilAsyncQueriesDone()
            except Exception:
                pass
            ws = wb.Worksheets(1)
            deadline = time.monotonic() + timeout_seconds
            values = None
            while time.monotonic() < deadline:
                time.sleep(2)
                values = ws.Range(f"B2:B{len(codes) + 1}").Value2
                valid = sum(row[0] not in (None, "", "数据获取中") for row in values)
                if valid and excel.CalculationState == 0:
                    break
            if values is None:
                raise RuntimeError("Wind Excel没有返回摘牌日期")
            parsed = []
            for row in values:
                value = row[0]
                if isinstance(value, (int, float)):
                    parsed.append(pd.Timestamp("1899-12-30") + pd.to_timedelta(float(value), unit="D"))
                else:
                    parsed.append(pd.to_datetime(value, errors="coerce"))
            return pd.Series(parsed, index=codes, name="摘牌日期")
        finally:
            if wb is not None:
                wb.Close(False)
            excel.Quit()


def current_balances(panel: pd.DataFrame, day: pd.Timestamp) -> pd.Series:
    snap = panel[panel[TRADE_DATE].eq(day)].set_index(BOND_CODE)
    return pd.to_numeric(snap["余额"], errors="coerce")


def delist_and_redeeming(master: pd.DataFrame, balances: pd.Series, as_of: pd.Timestamp, delist_dates: pd.Series) -> tuple[pd.DataFrame, pd.DataFrame]:
    data = master.copy()
    data["摘牌日期"] = data.index.to_series().map(delist_dates)
    data["余额"] = balances.reindex(data.index)
    window_end = as_of + pd.DateOffset(months=1)
    selected = data[data["摘牌日期"].gt(as_of) & data["摘牌日期"].le(window_end)].copy()
    selected = selected[selected["余额"].fillna(0).gt(0)].sort_values("摘牌日期")
    not_delisted = data["摘牌日期"].gt(as_of) | (data["摘牌日期"].isna() & data["最后交易日"].gt(as_of))
    redeeming = data[data["赎回公告日"].notna() & data["赎回公告日"].le(as_of) & not_delisted].copy()
    redeeming = redeeming[redeeming["余额"].fillna(0).gt(0)].sort_values("赎回公告日", ascending=False)
    return selected, redeeming


def read_redemption_book() -> tuple[pd.DataFrame, pd.DataFrame]:
    if not REDEMPTION_BOOK.exists():
        return pd.DataFrame(), pd.DataFrame()
    redemption = pd.read_excel(REDEMPTION_BOOK, sheet_name="赎回")
    redemption = redemption.iloc[:, :4]
    redemption.columns = [BOND_CODE, "转债名称", "公告时间", "公告标题"]
    redemption["公告时间"] = pd.to_datetime(redemption["公告时间"], errors="coerce")
    no_redemption = pd.read_excel(REDEMPTION_BOOK, sheet_name="不赎回")
    no_redemption = no_redemption.iloc[:, :5]
    no_redemption.columns = [BOND_CODE, "转债名称", "公告时间", "承诺截止日", "公告标题"]
    no_redemption["公告时间"] = pd.to_datetime(no_redemption["公告时间"], errors="coerce")
    no_redemption["承诺截止日"] = pd.to_datetime(no_redemption["承诺截止日"], errors="coerce")
    return redemption.dropna(subset=[BOND_CODE]), no_redemption.dropna(subset=[BOND_CODE])


def strong_redemption_text(panel: pd.DataFrame, master: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> str:
    redemption, no_redemption = read_redemption_book()
    end_exclusive = week_end + pd.Timedelta(days=1)
    announced = redemption[redemption["公告时间"].ge(week_start) & redemption["公告时间"].lt(end_exclusive)] if not redemption.empty else pd.DataFrame()
    announced_names = announced["转债名称"].dropna().astype(str).drop_duplicates().tolist()

    snap = panel[panel[TRADE_DATE].eq(week_end)].set_index(BOND_CODE)
    trigger = pd.to_numeric(snap["赎回累计天数"], errors="coerce").dropna()
    strong_codes = set(redemption[BOND_CODE].astype(str)) if not redemption.empty else set()
    recently_declined = set()
    if not no_redemption.empty:
        active = no_redemption["承诺截止日"].ge(week_end) | no_redemption["公告时间"].ge(week_start - pd.Timedelta(days=7))
        recently_declined = set(no_redemption.loc[active, BOND_CODE].astype(str))
    candidates = trigger[~trigger.index.astype(str).isin(strong_codes | recently_declined)]
    candidates = candidates[candidates > 0].sort_values(ascending=False)
    if not candidates.empty:
        cutoff = candidates.iloc[min(2, len(candidates) - 1)]
        candidates = candidates[candidates >= cutoff].head(4)
    near_names = master["转债名称"].reindex(candidates.index).dropna().astype(str).tolist()
    announced_part = chinese_join(announced_names) if announced_names else "暂无新增转债"
    near_part = chinese_join(near_names) if near_names else "暂无"
    return f"强赎方面，公告强赎的有{announced_part}，强赎触发较近的有{near_part}。"


def individual_weekly_returns(panel: pd.DataFrame, master: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> pd.DataFrame:
    sample = panel[panel[TRADE_DATE].between(week_start, week_end)].copy()
    sample = sample.join(master[["上市日期", "最后交易日"]], on=BOND_CODE)
    valid = (
        (sample["上市日期"].isna() | (sample[TRADE_DATE] >= sample["上市日期"]))
        & (sample["最后交易日"].isna() | (sample[TRADE_DATE] <= sample["最后交易日"]))
        & sample["涨跌幅"].notna()
    )
    sample = sample.loc[valid]
    ret = sample.groupby(BOND_CODE)["涨跌幅"].apply(
        lambda values: ((1 + pd.to_numeric(values, errors="coerce").dropna() / 100).prod() - 1) * 100
    )
    result = master[["转债名称", "申万行业", "申万三级行业", "上市日期", "最后交易日"]].copy()
    result["周涨跌幅"] = ret.reindex(result.index)
    return result[result["周涨跌幅"].notna()].sort_values("周涨跌幅", ascending=False)


def industry_text(individual: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> str:
    grouped = individual.dropna(subset=["申万行业"]).groupby("申万行业")["周涨跌幅"].mean().sort_values(ascending=False)
    positive = grouped[grouped > 0].head(4)
    negative = grouped[grouped < 0].sort_values().head(3)
    up = "、".join(f"{name}（{value:+.2f}%）" for name, value in positive.items()) or "无行业"
    down = "、".join(f"{name}（{value:+.2f}%）" for name, value in negative.items()) or "无行业"
    return f"转债行业方面，{date_range_text(week_start, week_end)}，转债市场各行业表现分化，{up}等上涨；{down}下跌居前。"


def individual_text(individual: pd.DataFrame) -> str:
    top = individual.head(5)
    bottom = individual.tail(5).sort_values("周涨跌幅")
    def items(frame: pd.DataFrame) -> str:
        return chinese_join(
            f"{row['转债名称']}（{row['申万三级行业'] if pd.notna(row['申万三级行业']) else '行业未分类'}）"
            for _, row in frame.iterrows()
        )
    return f"个券方面，{items(top)}转债涨幅靠前；{items(bottom)}跌幅靠前。"


def fund_text(indices: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> str:
    base = prior_base_date(indices.index, week_start)
    values = (indices.loc[week_end, FUND_INDICES] / indices.loc[base, FUND_INDICES] - 1) * 100
    verb = "整体表现积极" if (values > 0).sum() >= 3 else "表现有所分化"
    def pair(first: float, second: float) -> str:
        if first >= 0 and second >= 0:
            return f"上涨{first:.2f}%、{second:.2f}%"
        if first < 0 and second < 0:
            return f"下跌{abs(first):.2f}%、{abs(second):.2f}%"
        return (
            f"{'上涨' if first >= 0 else '下跌'}{abs(first):.2f}%、"
            f"{'上涨' if second >= 0 else '下跌'}{abs(second):.2f}%"
        )
    return (
        f"基金表现方面，“固收+”基金{verb}，一级债基及二级债基指数分别"
        f"{pair(float(values.iloc[0]), float(values.iloc[1]))}，"
        f"可转债基金及灵活配置型基金指数则分别"
        f"{pair(float(values.iloc[2]), float(values.iloc[3]))}。"
    )


def turnover_text(panel: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> str:
    daily = panel.groupby(TRADE_DATE)["成交额"].sum(min_count=1).sort_index()
    current = daily[(daily.index >= week_start) & (daily.index <= week_end)].dropna()
    previous_start = week_start - pd.Timedelta(days=7)
    previous_end = week_start - pd.Timedelta(days=1)
    previous = daily[(daily.index >= previous_start) & (daily.index <= previous_end)].dropna()
    total, avg, previous_avg = float(current.sum()), float(current.mean()), float(previous.mean())
    change = avg / previous_avg - 1
    action = "放量" if change >= 0 else "缩量"
    return (
        f"成交额方面，{date_range_text(week_start, week_end)}，转债市场总成交额{total:.2f}亿元，"
        f"日均成交额{avg:.2f}亿元，较{date_range_text(pd.Timestamp(previous.index.min()), pd.Timestamp(previous.index.max()))}均值"
        f"{previous_avg:.2f}亿元{action}{abs(change):.2%}。"
    )


def market_description(returns: pd.Series) -> str:
    equities = returns.iloc[:4]
    if (equities > 0).all():
        direction = "普遍上涨"
    elif (equities < 0).all():
        direction = "普遍下跌"
    else:
        direction = "表现分化"
    large = float(equities.iloc[0])
    small = float(equities.iloc[1:].mean())
    size = "中小盘表现占优" if small > large else "大盘表现相对占优"
    cb = float(returns.iloc[4])
    equity_avg = float(equities.mean())
    if cb >= 0 and equity_avg >= 0:
        relative = "转债市场同步上涨" + ("但涨幅相对有限" if cb < equity_avg else "且表现相对占优")
    elif cb < 0 and equity_avg < 0:
        relative = "转债市场同步下跌" + ("但相对抗跌" if cb > equity_avg else "且跌幅相对较大")
    else:
        relative = "转债市场与权益市场表现有所背离"
    return f"主要股指{direction}，{size}，{relative}。"


def overview_text(basic: pd.DataFrame, upcoming: pd.DataFrame, parquet_balance: float) -> str:
    count = len(basic)
    # 周报存续余额统一使用报告周五 Parquet 截面的“余额”汇总，
    # iFinD 基础信息仅用于存续数量、未上市名单和未来发行日历。
    balance = float(parquet_balance)
    unlisted = basic[basic["交易状态"].astype("string").eq("已发行未上市") | basic["上市日期"].isna()]
    if unlisted.empty:
        listed_part = "已发行转债均已上市进行交易，"
    else:
        listed_part = f"已发行转债中，{chinese_join(unlisted['转债名称'].astype(str))}尚未上市进行交易，"
    upcoming_names = []
    if not upcoming.empty and "p00600_f044" in upcoming:
        upcoming_names = upcoming["p00600_f044"].dropna().astype(str).tolist()
    upcoming_part = f"其中{chinese_join(upcoming_names)}即将网上发行。" if upcoming_names else "目前尚无将发行转债。"
    return f"现已发行未到期可转债有{count}支，余额规模{balance:.2f}亿元，{listed_part}{upcoming_part}"


def exit_text(delist: pd.DataFrame, redeeming: pd.DataFrame) -> str:
    delist_names = chinese_join(delist["转债名称"].astype(str))
    delist_balance = pd.to_numeric(delist["余额"], errors="coerce").sum(min_count=1)
    redeem_balance = pd.to_numeric(redeeming["余额"], errors="coerce").sum(min_count=1)
    return (
        f"退出方面，未来一个月内将摘牌转债目前有{len(delist)}只，合计余额{delist_balance:.2f}亿元，"
        f"包括{delist_names}，正在赎回中的转债共{len(redeeming)}只，合计余额{redeem_balance:.2f}亿元。"
    )


def proposal_commentary(proposals: pd.DataFrame, master: pd.DataFrame, basic: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> str:
    issued = master[master["发行日期"].between(week_start, week_end, inclusive="both")]
    listed = master[master["上市日期"].between(week_start, week_end, inclusive="both")]
    issued_names = chinese_join(issued["转债名称"].astype(str))
    # 新券上市统一使用转债名称，不使用发行主体或正股简称。
    listed_names = chinese_join(listed["转债名称"].astype(str))
    parts = [
        f"一级市场方面，{date_range_text(week_start, week_end)}，{len(issued)}只转债发行，包括{issued_names}，合计规模{issued['发行规模'].sum():.2f}亿元；"
        f"{listed_names}新券上市，规模合计{listed['发行规模'].sum():.2f}亿元。"
    ]
    events = [
        ("预案公告日", "新增董事会预案"),
        ("股东大会公告日", "新增股东大会通过"),
        ("发审委审核公告日", "新增发审委审批通过"),
        ("证监会批准公告日", "新增证监会核准"),
    ]
    for column, label in events:
        if column not in proposals:
            continue
        names = proposals.loc[proposals[column].between(week_start, week_end, inclusive="both"), "公司名称"].dropna().astype(str).drop_duplicates().tolist()
        if names:
            parts.append(f"{chinese_join(names)}{label}。")
    active = proposals[~proposals["方案进度"].astype("string").str.contains("停止实施|终止", na=False)]
    registered = active[active["方案进度"].astype("string").str.contains("证监会批准|同意注册", na=False)]
    committee = active[active["方案进度"].astype("string").str.contains("上市委通过|发审委通过", na=False)]
    parts.append(
        f"截至目前，待发可转债共计{len(active)}只，合计规模{active['发行规模(亿元)'].sum():,.2f}亿元，"
        f"其中已被同意注册的{len(registered)}只，规模合计{registered['发行规模(亿元)'].sum():.2f}亿元；"
        f"已获上市委通过的{len(committee)}只，规模合计{committee['发行规模(亿元)'].sum():.2f}亿元。"
    )
    return "".join(parts)


# ---------------------------------------------------------------------------
# 华创周报正文与图表底稿
# ---------------------------------------------------------------------------

def action(value: float, *, down_word: str = "压缩") -> str:
    if pd.isna(value):
        return "变动数据暂缺"
    return f"{'抬升' if value >= 0 else down_word}{abs(value):.2f}"


def return_phrase(value: float) -> str:
    if pd.isna(value):
        return "数据暂缺"
    return f"{'上涨' if value >= 0 else '下跌'}{abs(value):.2f}%"


def named_subject(names: Iterable[str], empty: str = "无转债") -> str:
    clean = list(dict.fromkeys(str(name).strip() for name in names if str(name).strip() and str(name) != "nan"))
    if not clean:
        return empty
    if len(clean) <= 2:
        return chinese_join(clean)
    return f"{len(clean)}只转债"


def active_snapshot(panel: pd.DataFrame, master: pd.DataFrame, day: pd.Timestamp) -> pd.DataFrame:
    sample = panel[panel[TRADE_DATE].eq(day)].reset_index(drop=True).copy()
    if sample.empty:
        raise ValueError(f"{day:%Y-%m-%d}没有转债截面数据")
    valid = valid_listing_mask(sample, master, day)
    sample = sample.loc[valid.to_numpy()].copy()
    sample["发行规模"] = pd.to_numeric(master["发行规模"].reindex(sample[BOND_CODE]).to_numpy(), errors="coerce")
    for column in ["余额", "收盘价", "平价", "转股溢价率", "平价底价溢价率", "涨跌幅", "正股市值"]:
        if column in sample:
            sample[column] = pd.to_numeric(sample[column], errors="coerce")
    sample = sample[sample["余额"].fillna(0).gt(0)]
    return sample


def weighted_average(frame: pd.DataFrame, value: str, weight: str = "余额") -> float:
    data = frame[[value, weight]].apply(pd.to_numeric, errors="coerce").dropna()
    data = data[data[weight] > 0]
    if data.empty:
        return float("nan")
    return float(np.average(data[value], weights=data[weight]))


def category_price_stats(panel: pd.DataFrame, master: pd.DataFrame, day: pd.Timestamp) -> dict[str, float]:
    snap = active_snapshot(panel, master, day)
    score = snap["平价底价溢价率"]
    return {
        "偏股": weighted_average(snap[score > 20], "收盘价"),
        "偏债": weighted_average(snap[score < -20], "收盘价"),
        "平衡": weighted_average(snap[score.between(-20, 20, inclusive="both")], "收盘价"),
        "中位数": float(snap["收盘价"].median()),
    }


def _mean_premium(frame: pd.DataFrame) -> float:
    values = pd.to_numeric(frame["转股溢价率"], errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
    if len(values) < 3:
        return float("nan")
    low, high = values.quantile([0.01, 0.99])
    return float(values.clip(low, high).mean())


def premium_group_changes(panel: pd.DataFrame, master: pd.DataFrame, base: pd.Timestamp, end: pd.Timestamp) -> dict[str, dict[str, float]]:
    before, after = active_snapshot(panel, master, base), active_snapshot(panel, master, end)

    def rating_group(frame: pd.DataFrame, label: str) -> pd.Series:
        rating = frame["债项评级"].fillna(frame["主体评级"]).astype("string").str.upper().str.strip()
        definitions = {
            "AAA/AA+": rating.isin(["AAA", "AA+"]),
            "AA/AA-": rating.isin(["AA", "AA-"]),
            "A+/A": rating.isin(["A+", "A", "A-"]),
        }
        return definitions[label]

    rating_changes: dict[str, float] = {}
    for label in ["AAA/AA+", "AA/AA-", "A+/A"]:
        rating_changes[label] = _mean_premium(after[rating_group(after, label)]) - _mean_premium(before[rating_group(before, label)])

    scale_defs = {
        "0—3亿元": lambda x: (x > 0) & (x < 3),
        "3—10亿元": lambda x: (x >= 3) & (x < 10),
        "10—20亿元": lambda x: (x >= 10) & (x < 20),
        "20—50亿元": lambda x: (x >= 20) & (x < 50),
        "50亿元以上": lambda x: x >= 50,
    }
    scale_changes = {
        label: _mean_premium(after[selector(after["发行规模"])]) - _mean_premium(before[selector(before["发行规模"])])
        for label, selector in scale_defs.items()
    }
    parity_defs = {
        "70—90元": lambda x: x.between(70, 90, inclusive="left"),
        "90—110元": lambda x: x.between(90, 110, inclusive="left"),
        "110—130元": lambda x: x.between(110, 130, inclusive="both"),
    }
    parity_changes = {
        label: _mean_premium(after[selector(after["平价"])]) - _mean_premium(before[selector(before["平价"])])
        for label, selector in parity_defs.items()
    }
    return {"rating": rating_changes, "scale": scale_changes, "parity": parity_changes}


def grouped_weekly_returns(panel: pd.DataFrame, master: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> tuple[dict[str, float], dict[str, float]]:
    sample = panel[panel[TRADE_DATE].between(week_start, week_end)].copy()
    sample["发行规模"] = pd.to_numeric(master["发行规模"].reindex(sample[BOND_CODE]).to_numpy(), errors="coerce")
    sample["评级"] = sample["债项评级"].fillna(sample["主体评级"]).astype("string").str.upper().str.strip()
    sample["余额"] = pd.to_numeric(sample["余额"], errors="coerce")
    sample["涨跌幅"] = pd.to_numeric(sample["涨跌幅"], errors="coerce")
    sample = sample.dropna(subset=["余额", "涨跌幅"])
    sample = sample[sample["余额"] > 0]

    def calculate(mask: pd.Series) -> float:
        chosen = sample.loc[mask].copy()
        daily = chosen.groupby(TRADE_DATE).apply(lambda x: np.average(x["涨跌幅"], weights=x["余额"]), include_groups=False)
        return float(((1 + daily / 100).prod() - 1) * 100) if not daily.empty else float("nan")

    scale = {
        "大盘": calculate(sample["发行规模"] >= 50),
        "中盘": calculate(sample["发行规模"].between(10, 50, inclusive="left")),
        "小盘": calculate(sample["发行规模"].between(0.3, 10, inclusive="left")),
    }
    rating = {
        "AAA": calculate(sample["评级"].eq("AAA")),
        "AA+": calculate(sample["评级"].eq("AA+")),
        "AA": calculate(sample["评级"].eq("AA")),
        "AA-及以下": calculate(~sample["评级"].isin(["AAA", "AA+", "AA"])),
    }
    return scale, rating


def first_market_section(index_returns: pd.Series, valuation: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> str:
    equity = index_returns.iloc[:4]
    if (equity < 0).all():
        equity_desc = "普遍回调"
    elif (equity > 0).all():
        equity_desc = "普遍上涨"
    else:
        equity_desc = "分化表现"
    cb = float(index_returns.iloc[4])
    eq_mean = float(equity.mean())
    relative = "表现相对占优" if cb > eq_mean else "表现相对偏弱"
    size = "大盘指数相对抗跌" if float(equity.iloc[0]) >= float(equity.iloc[1:].max()) else "中小盘指数相对占优"
    series = valuation.set_index("日期")["百元平价拟合溢价率"].dropna().sort_index()
    base = prior_base_date(series.index, week_start)
    latest = float(series.loc[:week_end].iloc[-1])
    change = latest - float(series.loc[base])
    market_title = "转债周度上涨" if cb >= 0 else "转债周度下跌"
    valuation_title = "估值小幅抬升" if change >= 0 else "估值有所压缩"
    cb_desc = "上涨" if cb >= 0 else "下跌"
    return (
        f"一、市场复盘：{market_title}，{valuation_title}\n\n"
        f"（一）周度市场行情：转债市场{cb_desc}，权益热点{'集中' if equity.std() < 0.4 else '分散'}\n\n"
        f"上周主要股指{equity_desc}，转债{relative}。"
        f"沪深300指数{return_phrase(float(index_returns.iloc[0]))}，中证500指数{return_phrase(float(index_returns.iloc[1]))}，"
        f"中证1000指数{return_phrase(float(index_returns.iloc[2]))}，中证2000指数{return_phrase(float(index_returns.iloc[3]))}，"
        f"中证转债指数{return_phrase(cb)}。本周权益市场{equity_desc}，{size}，转债市场"
        f"{'上涨' if cb >= 0 else '回调'}，估值{action(change)}pct至{latest:.2f}%。"
    )


def second_valuation_section(panel: pd.DataFrame, master: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp, base: pd.Timestamp) -> str:
    before, after = category_price_stats(panel, master, base), category_price_stats(panel, master, week_end)
    changes = premium_group_changes(panel, master, base, week_end)
    scale_ret, rating_ret = grouped_weekly_returns(panel, master, week_start, week_end)
    price_change = {key: (after[key] / before[key] - 1) * 100 for key in before}
    rating_values = list(changes["rating"].values())
    rating_summary = "整体抬升" if all(v >= 0 for v in rating_values if pd.notna(v)) else ("整体压缩" if all(v <= 0 for v in rating_values if pd.notna(v)) else "表现分化")
    rating_best = max(changes["rating"], key=lambda k: changes["rating"][k] if pd.notna(changes["rating"][k]) else -1e9)
    scale_sorted = sorted(changes["scale"].items(), key=lambda item: item[1] if pd.notna(item[1]) else 1e9)
    parity_sorted = sorted(changes["parity"].items(), key=lambda item: item[1] if pd.notna(item[1]) else 1e9)
    scale_words = "，".join(f"{name}{action(value)}pct" for name, value in scale_sorted)
    parity_words = "，".join(f"{name}{action(value)}pct" for name, value in parity_sorted)
    scale_best = max(scale_ret, key=lambda k: scale_ret[k] if pd.notna(scale_ret[k]) else -1e9)
    rating_best_ret = max(rating_ret, key=lambda k: rating_ret[k] if pd.notna(rating_ret[k]) else -1e9)
    return (
        f"（二）估值表现：各评级转债拟合溢价率均值{rating_summary}\n\n"
        f"从价格角度来看，偏股型转债的收盘价为{after['偏股']:.2f}元，较上周五{action(price_change['偏股'])}%；"
        f"偏债型转债的收盘价为{after['偏债']:.2f}元，较上周五{action(price_change['偏债'])}%；"
        f"平衡型转债的收盘价为{after['平衡']:.2f}元，较上周五{action(price_change['平衡'])}%。"
        f"价格中位数为{after['中位数']:.2f}元，较上周五{action(price_change['中位数'], down_word='下降')}%。"
        f"从评级和规模来看，各评级转债拟合溢价率均值{rating_summary}，{rating_best}评级表现相对占优，其中"
        + "，".join(f"{name}评级溢价率{action(value)}pct" for name, value in changes["rating"].items())
        + f"；分规模看，{scale_words}。从平价区间来看，{parity_words}。"
        f"从构建价格指数来看，转债{scale_best}指数表现相对占优；转债大盘指数上周{return_phrase(scale_ret['大盘'])}，"
        f"中盘{return_phrase(scale_ret['中盘'])}，小盘{return_phrase(scale_ret['小盘'])}。评级方面，转债{rating_best_ret}指数相对占优，"
        f"AAA指数{return_phrase(rating_ret['AAA'])}，AA+指数{return_phrase(rating_ret['AA+'])}，"
        f"AA指数{return_phrase(rating_ret['AA'])}，AA-及以下指数{return_phrase(rating_ret['AA-及以下'])}。\n\n"
        f"结构上，估值变化并非简单的单边扩张或收缩，而是在评级、规模与平价区间之间重新分配。"
        f"短期应优先观察正股趋势能否持续消化溢价率，并控制高估值、低流动性品种的回撤风险。"
    )


def latest_strategy_workbook(week_end: pd.Timestamp) -> Path | None:
    candidates: list[tuple[pd.Timestamp, Path]] = []
    for path in (ROOT / "runs" / "research").glob("策略回测*/回测合集.xlsx"):
        match = re.search(r"(20\d{6})", path.parent.name)
        if match:
            day = pd.to_datetime(match.group(1), format="%Y%m%d", errors="coerce")
            if pd.notna(day) and day <= week_end:
                candidates.append((day, path))
    return max(candidates, default=(pd.NaT, None), key=lambda item: item[0])[1]


def strategy_section(week_end: pd.Timestamp, cb_return: float) -> str:
    path = latest_strategy_workbook(week_end)
    if path is None:
        return "（三）策略回顾：策略收益随市场结构分化\n\n本周策略回测底稿暂缺，建议结合正股趋势、转债Delta与溢价率消化能力进行组合复核。"
    summary = pd.read_excel(path, sheet_name="总结表格")
    returns = summary.set_index("策略")["本周收益率"].dropna().astype(float) * 100
    strategy_names = [name for name in returns.index if name not in ["万得全A", "转债指数"]]
    ranked = returns.reindex(strategy_names).sort_values(ascending=False)
    top, bottom = ranked.head(3), ranked.tail(3).sort_values()
    comment = pd.read_excel(path, sheet_name="本周策略点评", header=None, names=["项目", "内容"])
    metrics = dict(zip(comment["项目"].astype(str), comment["内容"]))
    def parse_percent(value: Any) -> float:
        text = str(value).strip()
        has_sign = "%" in text
        number = pd.to_numeric(pd.Series([text.replace("%", "")]), errors="coerce").iloc[0]
        if pd.notna(number) and not has_sign and abs(number) <= 1:
            number *= 100
        return float(number)
    stock = parse_percent(metrics.get("近20日正股贡献"))
    valuation = parse_percent(metrics.get("近20日估值贡献"))
    top_words = "、".join(f"{name}{value:.2f}%" for name, value in top.items())
    bottom_words = "、".join(f"{name}{value:.2f}%" for name, value in bottom.items())
    driver = "正股收益能够覆盖估值损耗" if stock + valuation >= 0 else "估值损耗超过正股收益"
    focus = "适度保留权益弹性" if stock > 0 and stock + valuation >= 0 else "提高组合防御性"
    return (
        f"（三）策略回顾：{top.index[0]}策略相对占优\n\n"
        f"上周策略收益分化，{top_words}表现居前，{bottom_words}相对靠后。"
        f"占优策略较中证转债指数跑赢{top.iloc[0] - cb_return:.2f}个百分点，反映当周收益更多来自组合对股性、价格与溢价率约束的差异化暴露。"
        f"近20日正股贡献为{stock:.2f}%，估值贡献为{valuation:.2f}%，说明市场呈现“正股支撑、转债估值压缩”的结构，{driver}。"
        f"短期配置应更重视溢价率消化能力与转债Delta，{focus}，并避免单纯追逐高弹性。"
    )


def clause_data(panel: pd.DataFrame, master: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> dict[str, list[str] | list[tuple[str, str]]]:
    redemption, no_redemption = read_redemption_book()
    window = lambda series: pd.to_datetime(series, errors="coerce").dt.normalize().between(week_start, week_end, inclusive="both")
    announced = redemption.loc[window(redemption["公告时间"]), "转债名称"].dropna().astype(str).drop_duplicates().tolist() if not redemption.empty else []
    declined = no_redemption.loc[window(no_redemption["公告时间"]), "转债名称"].dropna().astype(str).drop_duplicates().tolist() if not no_redemption.empty else []
    snap = active_snapshot(panel, master, week_end).set_index(BOND_CODE)
    trigger = pd.to_numeric(snap["赎回累计天数"], errors="coerce")
    base = prior_base_date(pd.DatetimeIndex(panel[TRADE_DATE].drop_duplicates().sort_values()), week_start)
    base_snap = active_snapshot(panel, master, base).set_index(BOND_CODE)
    base_trigger = pd.to_numeric(base_snap["赎回累计天数"], errors="coerce").reindex(trigger.index).fillna(0)
    excluded = set(redemption[BOND_CODE].astype(str)) | set(no_redemption.loc[no_redemption["承诺截止日"].ge(week_end), BOND_CODE].astype(str))
    # “预计满足强赎”仅列本周首次跨过10/15阈值的品种，避免连续数周重复列示。
    expected_codes = trigger[(trigger >= 10) & (base_trigger < 10) & ~trigger.index.astype(str).isin(excluded)].sort_values(ascending=False).index
    expected = master["转债名称"].reindex(expected_codes).dropna().astype(str).tolist()

    proposed: list[str] = []
    results: list[tuple[str, str]] = []
    no_down: list[str] = []
    expected_down: list[str] = []
    if DOWN_REVISION_BOOK.exists():
        down = pd.read_excel(DOWN_REVISION_BOOK, sheet_name="下修")
        proposed = down.loc[window(down["董事会发布日期"]), "转债名称"].dropna().astype(str).drop_duplicates().tolist()
        result_rows = down.loc[window(down["向下修正发布日期"])].copy()
        result_rows = result_rows[
            ~result_rows["是否下修到底\n（带函数，<5%）"].astype("string").str.contains("失败", na=False)
        ]
        for _, row in result_rows.iterrows():
            raw = row.get("是否下修到底\n（带函数，<5%）", pd.NA)
            value = str(raw).strip()
            label = "到底" if value in {"是", "1", "True", "TRUE"} else "未到底"
            results.append((str(row["转债名称"]), label))
        no = pd.read_excel(DOWN_REVISION_BOOK, sheet_name="不下修")
        no_down = no.loc[window(no["公告时间"]), "转债简称"].dropna().astype(str).drop_duplicates().tolist()
        no_down = [name for name in no_down if name not in {item[0] for item in results}]
        exp = pd.read_excel(DOWN_REVISION_BOOK, sheet_name="预计下修")
        expected_down = exp.loc[window(exp["公告时间"]), "转债简称"].dropna().astype(str).drop_duplicates().tolist()
    return {
        "strong": announced, "no_strong": declined, "expected_strong": expected,
        "proposed": proposed, "results": results, "no_down": no_down, "expected_down": expected_down,
    }


def clauses_text(data: dict[str, Any], week_end: pd.Timestamp) -> tuple[str, str]:
    strong_subject = named_subject(data["strong"])
    proposed_subject = named_subject(data["proposed"])
    strong_title = f"{strong_subject}公告强赎" if data["strong"] else "无转债公告强赎"
    proposed_title = f"{proposed_subject}提议下修" if data["proposed"] else "无转债提议下修"
    no_strong = chinese_join(data["no_strong"]) if data["no_strong"] else "无转债"
    expected = chinese_join(data["expected_strong"]) if data["expected_strong"] else "无转债"
    result_text = "、".join(f"{name}（{label}）" for name, label in data["results"]) if data["results"] else "无转债"
    no_down = chinese_join(data["no_down"]) if data["no_down"] else "无转债"
    expected_down = chinese_join(data["expected_down"])
    expected_down_text = (
        f"{named_subject(data['expected_down'])}公告预计触发下修，包括{expected_down}。"
        if len(data["expected_down"]) > 2 else
        f"{expected_down or '无转债'}公告预计触发下修。"
    )
    body = (
        f"截至{month_day(week_end)}，{chinese_join(data['strong']) + '公告新增提前赎回' if data['strong'] else '无公告新增提前赎回'}；"
        f"{no_strong}公告不提前赎回；{expected}公告预计满足强赎条件。\n"
        f"截至{month_day(week_end)}，上周{chinese_join(data['proposed']) if data['proposed'] else '无转债'}发布董事会提议向下修正议案的公告，"
        f"{result_text}公告下修结果；{no_down}公告不下修，{expected_down_text}"
    )
    return f"（一）条款：上周{strong_title}，{proposed_title}", body


def issuance_pipeline_text(proposals: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> tuple[str, dict[str, float]]:
    active = proposals[~proposals["方案进度"].astype("string").str.contains("停止实施|终止", na=False)].copy()
    registered = active[active["方案进度"].astype("string").str.contains("证监会批准|同意注册", na=False)]
    new_registered = active[active["证监会批准公告日"].between(week_start, week_end, inclusive="both")]
    new_committee = active[active["发审委审核公告日"].between(week_start, week_end, inclusive="both")]
    new_board = active[active["预案公告日"].between(week_start, week_end, inclusive="both")]

    def event_sentence(frame: pd.DataFrame, label: str) -> str:
        names = frame["公司名称"].dropna().astype(str).drop_duplicates().tolist()
        amount = frame.drop_duplicates("公司代码")["发行规模(亿元)"].sum(min_count=1)
        if not names:
            return f"上周无公司新增{label}"
        return f"上周{chinese_join(names)}新增{label}" if label == "转债发行批文" else f"{chinese_join(names)}新增{label}，规模{float(amount):.2f}亿元"

    first = event_sentence(new_registered, "转债发行批文")
    second = event_sentence(new_committee, "通过发审委")
    third = event_sentence(new_board, "董事会预案")
    if "规模" not in third:
        third += "，合计规模0.00亿元"
    else:
        third = third.replace("，规模", "，合计规模")
    text = (
        f"截至{month_day(week_end)}，{first}，尚在证监会同意注册阶段的合计{len(registered)}家，"
        f"总计拟发行规模{registered['发行规模(亿元)'].sum():.2f}亿元。{second}，{third}。"
    )
    stats = {
        "active_count": len(active), "active_amount": float(active["发行规模(亿元)"].sum()),
        "registered_count": len(registered), "registered_amount": float(registered["发行规模(亿元)"].sum()),
    }
    return text, stats


def clean_scalar(value: Any) -> Any:
    if value is None or (not isinstance(value, (date, datetime)) and pd.isna(value)):
        return None
    text = str(value).strip()
    return None if text in {"", "--", "None", "nan", "NaT"} else value


def parse_ifind_date(value: Any) -> date | None:
    parsed = pd.to_datetime(value, errors="coerce")
    return None if pd.isna(parsed) else parsed.date()


def ensure_market_suffix(code: Any) -> str:
    text = str(code).strip().upper()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    if text.endswith((".SH", ".SZ")):
        return text
    return f"{text}.SH" if text.startswith(("10", "11")) else f"{text}.SZ"


def ifind_report(session: IFindSession, report_id: str, fields: str, label: str) -> pd.DataFrame:
    result = normalize_ifind_frame(
        session.module.THS_DR(report_id, "zqlx=640007;gnfl=0", fields, "format:dataframe"),
        f"{label} {report_id}", allow_no_data=True, empty_columns=[],
    )
    if result.empty:
        return pd.DataFrame(columns=[BOND_CODE])
    field = f"{report_id}_f001"
    if field not in result:
        raise RuntimeError(f"{report_id}缺少字段{field}")
    result[BOND_CODE] = result[field].map(ensure_market_suffix)
    return result.drop_duplicates(BOND_CODE)


def fetch_unlisted_report(session: IFindSession, as_of: pd.Timestamp) -> pd.DataFrame:
    fields = "jydm:Y,jydm_mc:Y,p05479_f001:Y,p05479_f002:Y,p05479_f018:Y,p05479_f019:Y,p05479_f044:Y"
    raw = normalize_ifind_frame(
        session.module.THS_DR(
            "p05479", f"jyzt=2;sfdb=1;jysc=1;sszt=213006;edate={as_of:%Y%m%d};gnfl=0",
            fields, "format:dataframe",
        ), "p05479", allow_no_data=True, empty_columns=["jydm"],
    )
    if raw.empty:
        return pd.DataFrame(columns=[BOND_CODE])
    raw = raw[~raw["jydm"].astype(str).str.contains("NQ", case=False, na=False)]
    raw = raw[~raw["jydm_mc"].astype(str).str.contains("定转", na=False)]
    return raw.rename(columns={
        "jydm": BOND_CODE, "jydm_mc": "转债简称", "p05479_f001": "发行规模",
        "p05479_f002": "期限", "p05479_f018": "发行公告日", "p05479_f019": "上市日期",
        "p05479_f044": "信用等级",
    })


def fetch_issue_details(session: IFindSession, codes: Iterable[str], as_of: pd.Timestamp) -> pd.DataFrame:
    codes = list(dict.fromkeys(ensure_market_suffix(code) for code in codes if str(code).strip()))
    if not codes:
        return pd.DataFrame(columns=[BOND_CODE])
    event = normalize_ifind_frame(
        session.module.THS_BD(",".join(codes), "ths_convertible_debt_short_name_cbond;ths_online_issue_date_cbond;ths_listed_date_cbond", ";;"),
        "THS_BD(发行与上市日期)",
    )
    event = event.rename(columns={"thscode": BOND_CODE})
    event.columns = [BOND_CODE, "转债简称", "网上发行日期", "上市日期"]
    display = normalize_ifind_frame(
        session.module.THS_BD(
            ",".join(codes),
            "ths_object_the_sw_bond;ths_issue_total_amt_bond;ths_debt_rating_primary_rating_agency_bond;ths_stock_code_cbond;ths_stock_short_name_cbond",
            f"100,{as_of:%Y-%m-%d};;;;",
        ), "THS_BD(发行展示字段)",
    ).rename(columns={"thscode": BOND_CODE})
    display.columns = [BOND_CODE, "申万行业", "发行规模", "信用等级", "正股代码", "公司名称"]
    result = event.merge(display, on=BOND_CODE, how="outer")
    for column in ["网上发行日期", "上市日期"]:
        result[column] = pd.to_datetime(result[column], errors="coerce")
    result["发行规模"] = pd.to_numeric(result["发行规模"], errors="coerce").map(
        lambda x: x / 100_000_000 if pd.notna(x) and abs(x) >= 1_000_000 else x
    )
    return result


def cninfo_query(start: pd.Timestamp, end: pd.Timestamp, keyword: str = "发行公告") -> list[dict[str, Any]]:
    endpoint = "https://www.cninfo.com.cn/new/hisAnnouncement/query"
    result: list[dict[str, Any]] = []
    for column, plate in [("szse", ""), ("sse", "sh")]:
        for page in range(1, 12):
            body = urlencode({
                "pageNum": str(page), "pageSize": "50", "column": column, "tabName": "fulltext",
                "plate": plate, "stock": "", "searchkey": keyword, "secid": "", "category": "", "trade": "",
                "seDate": f"{start:%Y-%m-%d}~{end:%Y-%m-%d}", "sortName": "", "sortType": "", "isHLtitle": "true",
            }, safe=";")
            command = [
                "curl.exe", "-sS", "-L", "--compressed", "--connect-timeout", "30", endpoint,
                "-H", "User-Agent: Mozilla/5.0", "-H", "Referer: https://www.cninfo.com.cn/",
                "-H", "X-Requested-With: XMLHttpRequest", "--data", body,
            ]
            proc = subprocess.run(command, capture_output=True, text=True, encoding="utf-8", timeout=90)
            if proc.returncode != 0:
                raise RuntimeError(f"巨潮公告查询失败：{proc.stderr.strip()}")
            items = (json.loads(proc.stdout).get("announcements") or [])
            result.extend(items)
            if len(items) < 50:
                break
    return list({str(item.get("announcementId") or item.get("adjunctUrl")): item for item in result}.values())


def announcement_date(items: list[dict[str, Any]], stock_code: str, issue_date: pd.Timestamp) -> tuple[pd.Timestamp | pd.NaT, str | None]:
    match = re.search(r"\d{6}", str(stock_code))
    if not match or pd.isna(issue_date):
        return pd.NaT, None
    candidates: list[tuple[pd.Timestamp, str]] = []
    for item in items:
        if str(item.get("secCode") or "").zfill(6) != match.group(0):
            continue
        title = re.sub(r"<[^>]+>", "", str(item.get("announcementTitle") or "")).strip()
        compact = re.sub(r"\s+", "", title)
        if "可转换公司债券发行公告" not in compact or any(word in compact for word in ["发行结果", "发行提示", "上市公告", "中签率", "中签号码"]):
            continue
        stamp = item.get("announcementTime")
        if stamp is None:
            continue
        day = pd.Timestamp(datetime.fromtimestamp(int(stamp) / 1000, tz=timezone.utc).astimezone(timezone(timedelta(hours=8))).date())
        if issue_date - pd.Timedelta(days=14) <= day <= issue_date:
            candidates.append((day, "https://static.cninfo.com.cn/" + str(item.get("adjunctUrl") or "")))
    return max(candidates, default=(pd.NaT, None), key=lambda item: item[0])


def sw1_industry(value: Any) -> str | None:
    text = str(value).strip()
    if not text or text in {"nan", "None"}:
        return None
    mapping = {
        "其他汽车零部件": "汽车", "高速公路": "交通运输", "面板": "电子", "分立器件": "电子", "印制电路板": "电子",
        "其他专业工程": "建筑装饰", "预加工食品": "食品饮料", "化学制剂": "医药生物", "制冷空调设备": "机械设备",
        "住宅开发": "房地产", "其他化学制品": "基础化工", "医疗耗材": "医药生物", "环保设备": "环保",
    }
    return mapping.get(text, text)


def table25_online(session: IFindSession, master: pd.DataFrame, week_start: pd.Timestamp, report_date: pd.Timestamp) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    window_end = report_date + pd.Timedelta(days=4)
    parquet_events = master[
        master["发行日期"].between(week_start, window_end, inclusive="both")
        | master["上市日期"].between(week_start, window_end, inclusive="both")
    ]
    unlisted = fetch_unlisted_report(session, report_date)
    issuing = ifind_report(session, "p04647", "p04647_f001:Y,p04647_f002:Y,p04647_f004:Y,p04647_f009:Y,p04647_f026:Y,p04647_f042:Y,p04647_f043:Y", "正在发行")
    pending = ifind_report(session, "p04649", "p04649_f001:Y,p04649_f002:Y,p04649_f004:Y,p04649_f009:Y,p04649_f026:Y,p04649_f043:Y,p04649_f044:Y", "待发行")
    pool = list(dict.fromkeys(
        parquet_events[BOND_CODE].astype(str).tolist()
        + unlisted.get(BOND_CODE, pd.Series(dtype=str)).astype(str).tolist()
        + issuing.get(BOND_CODE, pd.Series(dtype=str)).astype(str).tolist()
        + pending.get(BOND_CODE, pd.Series(dtype=str)).astype(str).tolist()
    ))
    details = fetch_issue_details(session, pool, report_date)
    details = details[
        details["网上发行日期"].between(week_start, window_end, inclusive="both")
        | details["上市日期"].between(week_start, window_end, inclusive="both")
    ].copy()
    unlisted_map = unlisted.set_index(BOND_CODE).to_dict("index") if not unlisted.empty else {}
    notices = cninfo_query(report_date - pd.Timedelta(days=60), report_date)
    rows: list[dict[str, Any]] = []
    audit: list[dict[str, Any]] = []
    for row in details.itertuples(index=False):
        code = str(getattr(row, BOND_CODE))
        issue_date = pd.Timestamp(row.网上发行日期) if pd.notna(row.网上发行日期) else pd.NaT
        notice = pd.to_datetime(unlisted_map.get(code, {}).get("发行公告日"), errors="coerce")
        url = None
        if pd.isna(notice):
            notice, url = announcement_date(notices, row.正股代码, issue_date)
        master_row = master.loc[code] if code in master.index else pd.Series(dtype=object)
        maturity = pd.to_datetime(master_row.get("到期日期"), errors="coerce")
        term = round((maturity - issue_date).days / 365.25) if pd.notna(maturity) and pd.notna(issue_date) else unlisted_map.get(code, {}).get("期限")
        amount = pd.to_numeric(pd.Series([row.发行规模]), errors="coerce").iloc[0]
        if pd.isna(amount):
            amount = pd.to_numeric(master_row.get("发行规模"), errors="coerce")
        rows.append({
            "发行公告日": notice, "转债代码": code, "转债简称": row.转债简称,
            "网上发行日期": issue_date, "上市日期": row.上市日期,
            "发行规模(亿元)": amount, "期限(年)": term, "信用等级": row.信用等级,
            "申万行业": sw1_industry(row.申万行业),
        })
        audit.append({"数据项": f"图表25-{code}", "来源": "iFinD/巨潮资讯", "链接或文件": url or "iFinD终端", "备注": "发行、上市及公告信息"})
    frame = pd.DataFrame(rows)
    if not frame.empty:
        frame = frame.sort_values(["发行公告日", "转债代码"], ascending=[False, True]).reset_index(drop=True)
    return frame, audit


def load_table25_fallback(week_end: pd.Timestamp, master: pd.DataFrame, week_start: pd.Timestamp, report_date: pd.Timestamp) -> pd.DataFrame:
    candidates = sorted((ROOT / "runs" / "weekly").glob("【华创】转债周报*/转债周报图表底稿*.xlsx"), reverse=True)
    candidates += sorted(ROOT.glob("outputs/**/转债周报图表底稿*.xlsx"), reverse=True)
    for path in candidates:
        try:
            frame = pd.read_excel(path, sheet_name="图表25", header=1)
            required = {"发行公告日", "转债代码", "转债简称", "网上发行日期", "上市日期", "发行规模(亿元)", "期限(年)", "信用等级", "申万行业"}
            if required.issubset(frame.columns):
                for column in ["发行公告日", "网上发行日期", "上市日期"]:
                    frame[column] = pd.to_datetime(frame[column], errors="coerce")
                window_end = report_date + pd.Timedelta(days=4)
                selected = frame[
                    frame["网上发行日期"].between(week_start, window_end, inclusive="both")
                    | frame["上市日期"].between(week_start, window_end, inclusive="both")
                ]
                if not selected.empty:
                    return selected[list(required)].sort_values("发行公告日", ascending=False)
        except Exception:
            pass
    events = master[
        master["发行日期"].between(week_start, report_date + pd.Timedelta(days=4), inclusive="both")
        | master["上市日期"].between(week_start, report_date + pd.Timedelta(days=4), inclusive="both")
    ].copy()
    return pd.DataFrame({
        "发行公告日": pd.NaT, "转债代码": events[BOND_CODE], "转债简称": events["转债名称"],
        "网上发行日期": events["发行日期"], "上市日期": events["上市日期"], "发行规模(亿元)": events["发行规模"],
        "期限(年)": np.where(events["到期日期"].notna() & events["发行日期"].notna(), ((events["到期日期"] - events["发行日期"]).dt.days / 365.25).round(), np.nan),
        "信用等级": pd.NA, "申万行业": events["申万行业"],
    })


def enrich_table25_from_local(frame: pd.DataFrame, master: pd.DataFrame, week_start: pd.Timestamp, report_date: pd.Timestamp) -> pd.DataFrame:
    """用最近一期华创底稿补齐在线接口没有返回的历史发行公告日等静态字段。"""
    fallback = load_table25_fallback(report_date - pd.Timedelta(days=3), master, week_start, report_date)
    if fallback.empty or frame.empty or BOND_CODE not in fallback:
        return frame
    fallback = fallback.drop_duplicates(BOND_CODE).set_index(BOND_CODE)
    result = frame.copy()
    for column in ["发行公告日", "期限(年)", "信用等级", "申万行业"]:
        if column not in result or column not in fallback:
            continue
        mapped = result[BOND_CODE].map(fallback[column])
        result[column] = result[column].where(result[column].notna(), mapped)
    return result


def table30_from_proposals(proposals: pd.DataFrame) -> pd.DataFrame:
    active = proposals[~proposals["方案进度"].astype("string").str.contains("停止实施|终止", na=False)]
    registered = active[active["方案进度"].astype("string").str.contains("证监会批准|同意注册", na=False)].copy()
    result = pd.DataFrame({
        "公告日期": registered["最新公告日期"], "公司代码": registered["公司代码"], "公司名称": registered["公司名称"],
        "同意注册日": registered["证监会批准公告日"], "发行规模(亿元)": registered["发行规模(亿元)"],
        "发行期限(年)": registered["发行期限(年)"], "申万行业": registered["申万行业"].map(sw1_industry),
    })
    return result.sort_values(["公告日期", "公司代码"], ascending=[False, True]).reset_index(drop=True)


def primary_market_sections(
    master: pd.DataFrame,
    basic: pd.DataFrame,
    upcoming: pd.DataFrame,
    proposals: pd.DataFrame,
    balances: pd.Series,
    delist: pd.DataFrame,
    redeeming: pd.DataFrame,
    week_start: pd.Timestamp,
    week_end: pd.Timestamp,
) -> tuple[str, str, str, dict[str, float]]:
    issued = master[master["发行日期"].between(week_start, week_end, inclusive="both")]
    listed = master[master["上市日期"].between(week_start, week_end, inclusive="both")]
    pipeline, stats = issuance_pipeline_text(proposals, week_start, week_end)
    issue_title = f"上周{len(issued)}只转债新券发行" if len(issued) else "上周无转债新券发行"
    list_title = f"{len(listed)}只转债上市" if len(listed) else "无转债上市"
    issued_names = chinese_join(issued["转债名称"].astype(str))
    listed_names = chinese_join(listed["转债名称"].astype(str))
    issued_sentence = (
        f"上周{len(issued)}只转债发行，包括{issued_names}，合计规模{issued['发行规模'].sum():.2f}亿元；"
        if len(issued) else "上周无转债发行；"
    )
    listed_sentence = (
        f"{listed_names}新券上市，规模合计{listed['发行规模'].sum():.2f}亿元；"
        if len(listed) else "无新券上市；"
    )
    overview = overview_text(basic, upcoming, float(balances.sum(min_count=1)))
    exits = exit_text(delist, redeeming)
    text1 = f"1、{issue_title}，{list_title}\n\n{issued_sentence}{listed_sentence}{overview}{exits}"

    active = proposals[~proposals["方案进度"].astype("string").str.contains("停止实施|终止", na=False)]
    new_board = active[active["预案公告日"].between(week_start, week_end, inclusive="both")]
    new_names = chinese_join(new_board["公司名称"].dropna().astype(str).drop_duplicates())
    board_title = (
        f"{len(new_board)}家公司新增董事会预案" if len(new_board) > 1
        else (f"{new_names}新增董事会预案" if len(new_board) == 1 else "无公司新增董事会预案")
    )
    text2 = (
        f"2、待发规模合计{stats['active_amount']:.2f}亿元，{board_title}\n\n"
        f"截至{month_day(week_end)}，待发可转债共计{int(stats['active_count'])}只，合计规模{stats['active_amount']:.2f}亿元。{pipeline}"
    )
    section_title = f"（二）一级市场：{issue_title}，总代发规模{stats['active_amount']:.2f}亿元"
    return section_title, text1, text2, stats


def _write_report_sheet(workbook: Workbook, name: str, title: str, frame: pd.DataFrame, source: str) -> None:
    sheet = workbook.create_sheet(name)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A3"
    columns = frame.columns.tolist()
    end_col = get_column_letter(max(1, len(columns)))
    sheet.merge_cells(f"A1:{end_col}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(name="宋体", size=14, bold=True, color="C00000")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 26
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for column, value in enumerate(columns, 1):
        cell = sheet.cell(2, column, value)
        cell.font = Font(name="宋体", size=10, bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r, row in enumerate(frame.itertuples(index=False, name=None), 3):
        for c, value in enumerate(row, 1):
            if value is pd.NA or (not isinstance(value, (date, datetime, pd.Timestamp)) and pd.isna(value)):
                value = None
            cell = sheet.cell(r, c, value)
            cell.font = Font(name="宋体", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if isinstance(value, (pd.Timestamp, datetime, date)):
                cell.number_format = "yyyy-mm-dd"
            elif columns[c - 1] in ["发行规模(亿元)"]:
                cell.number_format = "0.00"
            elif columns[c - 1] in ["期限(年)", "发行期限(年)"]:
                cell.number_format = "0"
        sheet.row_dimensions[r].height = 20
    source_row = 3 + len(frame)
    sheet.merge_cells(start_row=source_row, start_column=1, end_row=source_row, end_column=max(1, len(columns)))
    sheet.cell(source_row, 1, source)
    sheet.cell(source_row, 1).font = Font(name="宋体", size=9)
    sheet.cell(source_row, 1).alignment = Alignment(horizontal="left", vertical="center")
    default_widths = {
        "发行公告日": 13, "公告日期": 13, "转债代码": 14, "公司代码": 14, "转债简称": 14, "公司名称": 16,
        "网上发行日期": 14, "上市日期": 13, "同意注册日": 13, "发行规模(亿元)": 15,
        "期限(年)": 10, "发行期限(年)": 13, "信用等级": 11, "申万行业": 14,
    }
    for c, column in enumerate(columns, 1):
        sheet.column_dimensions[get_column_letter(c)].width = default_widths.get(column, 16)
    sheet.auto_filter.ref = f"A2:{end_col}{max(2, 2 + len(frame))}"
    sheet.print_title_rows = "1:2"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def write_combined_workbook(
    output: Path,
    table25: pd.DataFrame,
    table30: pd.DataFrame,
    sources: list[dict[str, Any]],
    week_end: pd.Timestamp,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    table25_columns = ["发行公告日", "转债代码", "转债简称", "网上发行日期", "上市日期", "发行规模(亿元)", "期限(年)", "信用等级", "申万行业"]
    table30_columns = ["公告日期", "公司代码", "公司名称", "同意注册日", "发行规模(亿元)", "发行期限(年)", "申万行业"]
    for column in table25_columns:
        if column not in table25:
            table25[column] = pd.NA
    for column in table30_columns:
        if column not in table30:
            table30[column] = pd.NA
    _write_report_sheet(workbook, "图表25", "图表25  上周及本周发行及上市转债", table25[table25_columns], "资料来源：Wind，华创证券")
    _write_report_sheet(workbook, "图表30", f"图表30  {len(table30)}家上市公司公开发行A股可转债获证监会核准", table30[table30_columns], "资料来源：Wind，华创证券")
    source_frame = pd.DataFrame(sources)
    source = workbook.create_sheet("数据来源")
    source.sheet_view.showGridLines = False
    source["A1"] = "数据来源与口径说明"
    source["A1"].font = Font(name="宋体", size=14, bold=True, color="C00000")
    source.merge_cells("A1:D1")
    columns = ["数据项", "来源", "链接或文件", "备注"]
    for c, value in enumerate(columns, 1):
        cell = source.cell(2, c, value)
        cell.font = Font(name="宋体", size=10, bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r, item in enumerate(source_frame.reindex(columns=columns).itertuples(index=False, name=None), 3):
        for c, value in enumerate(item, 1):
            source.cell(r, c, None if pd.isna(value) else value)
            source.cell(r, c).font = Font(name="宋体", size=10)
            source.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)
    source.column_dimensions["A"].width = 24
    source.column_dimensions["B"].width = 28
    source.column_dimensions["C"].width = 52
    source.column_dimensions["D"].width = 52
    source.freeze_panes = "A3"
    workbook.properties.title = f"华创转债周报图表底稿{week_end:%Y%m%d}"
    workbook.properties.subject = "发行上市与同意注册转债明细"
    workbook.save(output)


def write_full_report(output: Path, sections: list[str]) -> None:
    output.write_text("\n\n".join(section.strip() for section in sections if section.strip()) + "\n", encoding="utf-8")


def write_proposal_workbook(proposals: pd.DataFrame, output: Path, source_note: str) -> None:
    display = proposals.copy()
    date_columns = ["最新公告日期", "预案公告日", "股东大会公告日", "发审委审核公告日", "证监会批准公告日"]
    for column in date_columns:
        display[column] = pd.to_datetime(display[column], errors="coerce")
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        display.to_excel(writer, sheet_name="发行预案", index=False)
        pd.DataFrame(
            {
                "项目": ["数据来源", "筛选规则", "记录数"],
                "说明": [source_note, "剔除发行方式包含‘定向’的记录", len(display)],
            }
        ).to_excel(writer, sheet_name="说明", index=False)
    workbook = openpyxl.load_workbook(output)
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.row_dimensions[1].height = 24
        for cell in sheet[1]:
            cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="C00000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="微软雅黑", size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column in range(1, sheet.max_column + 1):
            width = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, min(sheet.max_row, 200) + 1))
            sheet.column_dimensions[get_column_letter(column)].width = min(max(width * 1.5 + 2, 12), 28)
        sheet.auto_filter.ref = sheet.dimensions
    proposal_sheet = workbook["发行预案"]
    date_column_numbers = [PROPOSAL_COLUMNS.index(column) + 1 for column in date_columns]
    for row in range(2, proposal_sheet.max_row + 1):
        for column in date_column_numbers:
            proposal_sheet.cell(row, column).number_format = "yyyy-mm-dd"
            proposal_sheet.cell(row, column).alignment = Alignment(horizontal="center", vertical="center")
        for column in [6, 7]:
            proposal_sheet.cell(row, column).number_format = "0.00"
            proposal_sheet.cell(row, column).alignment = Alignment(horizontal="right", vertical="center")
    workbook.save(output)


def write_text(output: Path, paragraphs: list[str], week_start: pd.Timestamp, week_end: pd.Timestamp) -> None:
    content = [f"数据周期：{week_start.year}年{month_day(week_start)}至{week_end.year}年{month_day(week_end)}", ""]
    for paragraph in paragraphs:
        content.extend([paragraph, ""])
    output.write_text("\n".join(content).rstrip() + "\n", encoding="utf-8")


def validate_outputs(output_dir: Path, week_end: pd.Timestamp) -> None:
    expected = [
        f"转债周报文本{week_end:%Y%m%d}.txt",
        "图1_中证转债与万得全A时序图.png",
        "图2_主要指数上周涨跌幅.png",
        "图3_转债价格中位数及历史分位数.png",
        "图4_转债百元拟合溢价率及历史分位数.png",
        f"转债周报图表底稿{week_end:%Y%m%d}.xlsx",
    ]
    missing = [name for name in expected if not (output_dir / name).exists() or (output_dir / name).stat().st_size == 0]
    if missing:
        raise RuntimeError(f"输出校验失败：{missing}")
    log("输出校验通过：" + "、".join(expected))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="基于Parquet v2.1、iFinD和公告统计簿一键生成华创转债周报")
    parser.add_argument("--week-end", help="报告周截止交易日，YYYY-MM-DD")
    parser.add_argument("--report-date", help="报告运行日，YYYY-MM-DD；默认取周末后的周一")
    parser.add_argument("--output-dir", type=Path, help="自定义输出目录，默认【华创】转债周报YYYYMMDD（周五日期）")
    parser.add_argument("--offline", action="store_true", help="不调用iFinD/巨潮，使用Parquet及本地最近一期底稿")
    parser.add_argument("--strict-external", action="store_true", help="外部接口失败时直接终止")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    week_start, week_end, report_date = choose_period(args.week_end, args.report_date)
    output_dir = (args.output_dir or (ROOT / "runs" / "weekly" / f"【华创】转债周报{week_end:%Y%m%d}")).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    log(f"报告周期：{week_start:%Y-%m-%d}至{week_end:%Y-%m-%d}；运行日：{report_date:%Y-%m-%d}")

    master = load_master()
    panel_start = min(START_CHART, week_start - pd.Timedelta(days=14))
    panel = load_panel(panel_start, week_end)
    indices = load_indices(START_CHART, week_end)
    valuation = build_daily_valuation(panel, master, week_end)

    figure1(indices, output_dir / "图1_中证转债与万得全A时序图.png")
    _, index_returns = weekly_returns(indices, week_start, week_end)
    figure2(index_returns, output_dir / "图2_主要指数上周涨跌幅.png")
    fig3_stats = valuation_figure(
        valuation, "转债价格中位数", "转债价格中位数",
        output_dir / "图3_转债价格中位数及历史分位数.png", 10.0, False,
    )
    fig4_stats = valuation_figure(
        valuation, "百元平价拟合溢价率", "百元平价拟合溢价率",
        output_dir / "图4_转债百元拟合溢价率及历史分位数.png", 5.0, True,
    )

    session = IFindSession()
    session.open(auto=not args.offline, strict=args.strict_external)
    proposal_source = "本地最近一次发行预案"
    table25_audit: list[dict[str, Any]] = []
    try:
        if session.active:
            proposals = fetch_proposals(session, report_date - pd.DateOffset(years=1), report_date)
            # p03153偶尔不返回仍在推进、但近期没有新公告的项目；与最近一期本地全量表合并，
            # 同一公司仍以在线最新公告记录覆盖，以兼顾完整性和状态更新。
            historical = load_proposal_fallback(report_date, output_dir)
            if not historical.empty:
                proposals = standardize_proposals(pd.concat([proposals, historical], ignore_index=True))
            proposal_source = f"iFinD THS_DR(p03153)+本地最近一期全量表，截至{report_date:%Y-%m-%d}"
            basic, upcoming = fetch_market_overview(session, week_end)
            try:
                table25, table25_audit = table25_online(session, master, week_start, report_date)
                table25 = enrich_table25_from_local(table25, master, week_start, report_date)
            except Exception as exc:
                if args.strict_external:
                    raise
                log(f"[警告] 图表25在线取数失败，改用本地底稿：{exc}")
                table25 = load_table25_fallback(week_end, master, week_start, report_date)
        else:
            proposals = load_proposal_fallback(report_date, output_dir)
            basic, upcoming = local_market_overview(panel, master, week_end)
            table25 = load_table25_fallback(week_end, master, week_start, report_date)
    finally:
        session.close()

    balances = current_balances(panel, week_end)
    parquet_balance = balances.sum(min_count=1)
    # “未来一个月将摘牌”按最后交易日统计，避免提前赎回品种的摘牌日更新滞后。
    delist_dates = master["最后交易日"].copy()
    delist, redeeming = delist_and_redeeming(master, balances, week_end, delist_dates)

    base = prior_base_date(pd.DatetimeIndex(panel[TRADE_DATE].drop_duplicates().sort_values()), week_start)
    first = first_market_section(index_returns, valuation, week_start, week_end)
    second = second_valuation_section(panel, master, week_start, week_end, base)
    strategy = strategy_section(week_end, float(index_returns.iloc[4]))
    clauses = clause_data(panel, master, week_start, week_end)
    clause_title, clause_body = clauses_text(clauses, week_end)
    primary_title, primary1, primary2, primary_stats = primary_market_sections(
        master, basic, upcoming, proposals, balances, delist, redeeming, week_start, week_end,
    )
    strong_subject = named_subject(clauses["strong"])
    chapter_strong = f"{strong_subject}公告强赎" if clauses["strong"] else "无转债公告强赎"
    chapter3 = f"三、条款及供给：{chapter_strong}，总代发规模{primary_stats['active_amount']:.2f}亿元"
    sections = [first, second, strategy, chapter3, clause_title, clause_body, primary_title, primary1, primary2]
    text_path = output_dir / f"转债周报文本{week_end:%Y%m%d}.txt"
    write_full_report(text_path, sections)

    table30 = table30_from_proposals(proposals)
    sources = [
        {"数据项": "图1—图4及正文市场数据", "来源": "本地Parquet Schema v2.1", "链接或文件": str(PARQUET_ROOT), "备注": f"截至{week_end:%Y-%m-%d}"},
        {"数据项": "发行预案与待发规模", "来源": proposal_source, "链接或文件": "iFinD p03153/本地最近一期底稿", "备注": "剔除定向发行及终止项目"},
        {"数据项": "存续余额", "来源": "本地Parquet周五截面", "链接或文件": str(PARQUET_ROOT), "备注": f"余额合计{parquet_balance:.2f}亿元"},
        {"数据项": "未来一个月摘牌", "来源": "总表最后交易日", "链接或文件": str(MASTER_PATH), "备注": "按最后交易日而非公告摘牌日统计"},
        {"数据项": "条款公告", "来源": "华创赎回/下修公告统计簿", "链接或文件": f"{REDEMPTION_BOOK.name}；{DOWN_REVISION_BOOK.name}", "备注": "按公告日在报告周内筛选"},
    ] + table25_audit
    workbook_path = output_dir / f"转债周报图表底稿{week_end:%Y%m%d}.xlsx"
    write_combined_workbook(workbook_path, table25, table30, sources, week_end)
    validate_outputs(output_dir, week_end)

    log(
        f"图3：最新{fig3_stats['最新值']:.2f}，25%/50%/75%="
        f"{fig3_stats['25%分位数']:.2f}/{fig3_stats['50%分位数']:.2f}/{fig3_stats['75%分位数']:.2f}"
    )
    log(
        f"图4：最新{fig4_stats['最新值']:.2f}%，25%/50%/75%="
        f"{fig4_stats['25%分位数']:.2f}%/{fig4_stats['50%分位数']:.2f}%/{fig4_stats['75%分位数']:.2f}%"
    )
    log(f"完成：{output_dir}")


if __name__ == "__main__":
    main()
