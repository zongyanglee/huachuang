"""人保转债市场周报一键生成脚本（Parquet Schema v2.1）。

正常运行：
    py 人保周报自动更新.py

离线验证（不调用 Wind / iFinD）：
    py 人保周报自动更新.py --offline

默认自动选择最近一个完整周，报告日期取该周结束后的下一个周一。
也可显式指定：
    py 人保周报自动更新.py --week-end 2026-08-07 --report-date 2026-08-10

最终只在 ``人保周报YYYYMMDD`` 文件夹保留：4张PNG、1个TXT、1个发行预案XLSX。
"""

from __future__ import annotations

import argparse
import math
import os
import re
import tempfile
import time
import warnings
from configparser import ConfigParser
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
from matplotlib.ticker import FormatStrFormatter, FuncFormatter, MultipleLocator
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
from scipy.optimize import OptimizeWarning, curve_fit


ROOT = Path(__file__).resolve().parents[2]
IFIND_CREDENTIAL_FILE = ROOT / "private/ifind账号.txt"
PARQUET_ROOT = ROOT / "data/转债个券历史序列"
MASTER_PATH = PARQUET_ROOT / "_special" / "总表.parquet"
INDEX_PATH = PARQUET_ROOT / "_special" / "指数.parquet"
REDEMPTION_BOOK = ROOT / "data/clauses/【华创固收】赎回和不赎回公告统计.xlsx"

BOND_CODE = "转债代码"
TRADE_DATE = "交易日期"
INDEX_NAME = "指数名称"
INDEX_VALUE = "指数值"

RED = "#E6121B"
BLUE = "#0262BA"
GRAY = "#A6A6A6"
LIGHT_RED = "#E7B8B8"
START_CHART = pd.Timestamp("2023-01-01")


def load_ifind_credentials() -> tuple[str, str]:
    """从项目目录的 ifind账号.txt 读取统一登录账号。"""
    if not IFIND_CREDENTIAL_FILE.is_file():
        raise FileNotFoundError(f"未找到iFinD账号文件：{IFIND_CREDENTIAL_FILE}")
    config = ConfigParser(interpolation=None)
    config.read(IFIND_CREDENTIAL_FILE, encoding="utf-8")
    username = config.get("ifind", "username", fallback="").strip()
    password = config.get("ifind", "password", fallback="").strip()
    if not username or not password:
        raise RuntimeError("ifind账号.txt中的[ifind] username或password为空")
    return username, password


def print_ifind_usage(module: object) -> None:
    """显示iFinD各数据项的已用额度比例。"""
    try:
        result = module.THS_DataStatistics()
        tables = result.get("tables", {}) if isinstance(result, dict) else {}
        if not tables:
            detail = result.get("errmsg", "未返回额度数据") if isinstance(result, dict) else str(result)
            print(f"[警告] iFinD使用额度查询失败：{detail}")
            return
        print("iFinD使用额度：")
        for key, value in tables.items():
            ratio = value.get("ratio", "N/A") if isinstance(value, dict) else value
            print(f"{key} 已用：{ratio}")
    except Exception as exc:
        print(f"[警告] iFinD使用额度查询失败：{exc}")

FIGURE2_INDICES = ["沪深300", "中证500", "中证1000", "中证2000", "转债指数"]
FUND_INDICES = ["混合债券型一级基金", "混合债券型二级基金", "可转换债券型基金", "灵活配置型基金"]
LOCAL_METRICS = [
    "余额",
    "收盘价",
    "平价",
    "转股溢价率",
    "换手率",
    "成交额",
    "涨跌幅",
    "剩余期限",
    "赎回累计天数",
]

PROPOSAL_COLUMNS = [
    "最新公告日期",
    "公司代码",
    "公司名称",
    "方案进度",
    "发行方式",
    "发行规模(亿元)",
    "发行期限(年)",
    "预案公告日",
    "股东大会公告日",
    "发审委审核公告日",
    "证监会批准公告日",
    "申万行业",
]


def log(message: str) -> None:
    print(message, flush=True)


def chinese_join(values: Iterable[str]) -> str:
    return "、".join(str(value) for value in values if str(value).strip())


def month_day(ts: pd.Timestamp) -> str:
    return f"{ts.month}月{ts.day}日"


def date_range_text(start: pd.Timestamp, end: pd.Timestamp) -> str:
    return f"{month_day(start)}至{month_day(end)}"


def monthly_files() -> list[Path]:
    return sorted(
        path
        for year_dir in PARQUET_ROOT.iterdir()
        if year_dir.is_dir() and year_dir.name.isdigit()
        for path in year_dir.glob("*.parquet")
        if re.fullmatch(r"\d{6}", path.stem)
    )


def load_master() -> pd.DataFrame:
    master = pd.read_parquet(MASTER_PATH)
    if BOND_CODE not in master:
        raise ValueError(f"{MASTER_PATH}缺少字段：{BOND_CODE}")
    master[BOND_CODE] = master[BOND_CODE].astype("string")
    for column in ["上市日期", "最后交易日", "发行日期", "赎回公告日", "转股期起始日"]:
        if column in master:
            master[column] = pd.to_datetime(master[column], errors="coerce")
    for column in ["发行规模", "赎回触发比例", "下修触发比例"]:
        if column in master:
            master[column] = pd.to_numeric(master[column], errors="coerce")
    return master.set_index(BOND_CODE, drop=False)


def load_panel(start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    parts: list[pd.DataFrame] = []
    columns = [BOND_CODE, TRADE_DATE, *LOCAL_METRICS]
    for path in monthly_files():
        year_month = pd.Timestamp(f"{path.stem[:4]}-{path.stem[4:]}-01")
        if year_month > end.to_period("M").to_timestamp() or year_month.to_period("M") < start.to_period("M"):
            continue
        frame = pd.read_parquet(path, columns=columns)
        frame[TRADE_DATE] = pd.to_datetime(frame[TRADE_DATE])
        frame = frame[frame[TRADE_DATE].between(start, end)]
        if not frame.empty:
            parts.append(frame)
    if not parts:
        raise FileNotFoundError(f"{start:%Y-%m-%d}至{end:%Y-%m-%d}没有月度parquet数据")
    panel = pd.concat(parts, ignore_index=True)
    panel[BOND_CODE] = panel[BOND_CODE].astype("string")
    panel = panel.sort_values([TRADE_DATE, BOND_CODE], kind="stable")
    if panel.duplicated([BOND_CODE, TRADE_DATE]).any():
        raise ValueError("月度parquet存在重复的转债代码＋交易日期")
    return panel.reset_index(drop=True)


def load_indices(start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    frame = pd.read_parquet(INDEX_PATH)
    required = {INDEX_NAME, TRADE_DATE, INDEX_VALUE}
    if not required.issubset(frame.columns):
        raise ValueError(f"{INDEX_PATH}不是指数长表Schema")
    frame[TRADE_DATE] = pd.to_datetime(frame[TRADE_DATE])
    frame[INDEX_VALUE] = pd.to_numeric(frame[INDEX_VALUE], errors="coerce")
    frame = frame[frame[TRADE_DATE].between(start, end)].dropna(subset=[INDEX_VALUE])
    return frame.pivot(index=TRADE_DATE, columns=INDEX_NAME, values=INDEX_VALUE).sort_index()


def all_panel_dates() -> pd.DatetimeIndex:
    paths = monthly_files()
    if not paths:
        raise FileNotFoundError(f"未找到月度parquet：{PARQUET_ROOT}")
    dates = pd.read_parquet(paths[-1], columns=[TRADE_DATE])[TRADE_DATE]
    return pd.DatetimeIndex(pd.to_datetime(dates).drop_duplicates().sort_values())


def choose_period(explicit_week_end: str | None, explicit_report_date: str | None) -> tuple[pd.Timestamp, pd.Timestamp, pd.Timestamp]:
    dates = all_panel_dates()
    latest = pd.Timestamp(dates.max()).normalize()
    today = pd.Timestamp.today().normalize()
    if explicit_week_end:
        week_end = pd.Timestamp(explicit_week_end).normalize()
        if week_end not in dates:
            candidates = dates[dates <= week_end]
            if len(candidates) == 0:
                raise ValueError(f"指定周末{week_end:%Y-%m-%d}之前没有交易日")
            week_end = pd.Timestamp(candidates.max()).normalize()
    else:
        latest_week_start = latest - pd.Timedelta(days=latest.weekday())
        current_week_complete = latest.weekday() >= 4 or today >= latest_week_start + pd.Timedelta(days=5)
        if current_week_complete:
            week_start = latest_week_start
        else:
            week_start = latest_week_start - pd.Timedelta(days=7)
        candidates = dates[(dates >= week_start) & (dates < week_start + pd.Timedelta(days=7))]
        if len(candidates) == 0:
            raise ValueError("无法识别最近完整周")
        week_end = pd.Timestamp(candidates.max()).normalize()
    week_start = week_end - pd.Timedelta(days=week_end.weekday())
    report_date = (
        pd.Timestamp(explicit_report_date).normalize()
        if explicit_report_date
        else week_start + pd.Timedelta(days=7)
    )
    return week_start, week_end, report_date


def prior_base_date(index: pd.DatetimeIndex, week_start: pd.Timestamp) -> pd.Timestamp:
    candidates = index[index < week_start]
    if len(candidates) == 0:
        raise ValueError("没有找到本周开始前的基准交易日")
    return pd.Timestamp(candidates.max()).normalize()


def inverse_cubic(x, a, b, c, d):
    return a / np.power(x, 3) + b / np.power(x, 2) + c / x + d


def fit_premium_at_100(sample: pd.DataFrame) -> float:
    data = sample[["平价", "转股溢价率", "换手率"]].apply(pd.to_numeric, errors="coerce")
    data = data.replace(0, np.nan).dropna()
    data = data[data["平价"].between(70, 130, inclusive="both") & (data["换手率"] <= 50)]
    if len(data) < 8:
        return float("nan")
    low, high = data["转股溢价率"].quantile([0.03, 0.97])
    data = data[(data["转股溢价率"] > low) & (data["转股溢价率"] < high)]
    if len(data) < 8:
        return float("nan")
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", OptimizeWarning)
            params, _ = curve_fit(
                inverse_cubic,
                data["平价"].to_numpy(float),
                data["转股溢价率"].to_numpy(float),
                maxfev=20000,
            )
        return float(inverse_cubic(100.0, *params))
    except Exception:
        return float("nan")


def valid_listing_mask(frame: pd.DataFrame, master: pd.DataFrame, day: pd.Timestamp) -> pd.Series:
    codes = frame[BOND_CODE]
    listing = master["上市日期"].reindex(codes).reset_index(drop=True)
    last_trade = master["最后交易日"].reindex(codes).reset_index(drop=True)
    return (
        (listing.isna() | (listing <= day))
        & (last_trade.isna() | (last_trade >= day))
        & frame["收盘价"].notna().reset_index(drop=True)
    )


def build_daily_valuation(panel: pd.DataFrame, master: pd.DataFrame, end: pd.Timestamp) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for day, raw in panel[panel[TRADE_DATE] >= START_CHART].groupby(TRADE_DATE, sort=True):
        day = pd.Timestamp(day)
        sample = raw.reset_index(drop=True)
        valid = valid_listing_mask(sample, master, day)
        sample = sample.loc[valid.to_numpy()].copy()
        rows.append(
            {
                "日期": day,
                "转债价格中位数": pd.to_numeric(sample["收盘价"], errors="coerce").median(),
                "百元平价拟合溢价率": fit_premium_at_100(sample),
            }
        )
    return pd.DataFrame(rows).sort_values("日期").drop_duplicates("日期", keep="last")


def expanded_limits(values: pd.Series, base_min: float, base_max: float, step: float) -> tuple[float, float]:
    minimum = float(pd.to_numeric(values, errors="coerce").min())
    maximum = float(pd.to_numeric(values, errors="coerce").max())
    low = base_min if minimum >= base_min else math.floor(minimum / step) * step - step
    high = base_max if maximum <= base_max else math.ceil(maximum / step) * step + step
    return low, high


def setup_fonts() -> tuple[FontProperties, FontProperties]:
    chinese = Path(r"C:\Windows\Fonts\simsun.ttc")
    latin = Path(r"C:\Windows\Fonts\times.ttf")
    if not chinese.exists() or not latin.exists():
        raise FileNotFoundError("未找到宋体或Times New Roman字体")
    return FontProperties(fname=str(chinese), size=13), FontProperties(fname=str(latin), size=12.5)


def figure1(indices: pd.DataFrame, output: Path) -> None:
    data = indices.loc[:, ["转债指数", "万得全A"]].dropna()
    chinese_font, latin_font = setup_fonts()
    fig, left = plt.subplots(figsize=(8, 4.75), dpi=250)
    right = left.twinx()
    fig.patch.set_facecolor("white")
    line_cb, = left.plot(data.index, data["转债指数"], color=RED, linewidth=2.2, label="中证转债")
    line_wind, = right.plot(data.index, data["万得全A"], color=BLUE, linewidth=2.2, label="万得全A")
    left.set_xlim(data.index.min(), data.index.max())
    left.set_ylim(*expanded_limits(data["转债指数"], 200, 600, 50))
    right.set_ylim(*expanded_limits(data["万得全A"], 3000, 7500, 500))
    left.yaxis.set_major_locator(MultipleLocator(50))
    right.yaxis.set_major_locator(MultipleLocator(500))
    right.yaxis.set_major_formatter(FuncFormatter(lambda value, _: f"{value:,.0f}"))
    ticks = [data.index.min().replace(year=year) for year in range(data.index.min().year, data.index.max().year + 1)]
    left.set_xticks(ticks)
    left.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
    for axis in (left, right):
        axis.grid(False)
        axis.spines["top"].set_visible(False)
        axis.tick_params(axis="both", labelsize=12.5, length=4, width=0.8)
        for label in axis.get_xticklabels() + axis.get_yticklabels():
            label.set_fontproperties(latin_font)
    left.spines["right"].set_visible(False)
    right.spines["left"].set_visible(False)
    legend = left.legend(
        [line_cb, line_wind], ["中证转债", "万得全A"], loc="upper center",
        bbox_to_anchor=(0.5, -0.17), ncol=2, frameon=False, handlelength=3.2,
        columnspacing=2.2, prop=chinese_font,
    )
    for handle in legend.legend_handles:
        handle.set_linewidth(2.5)
    fig.subplots_adjust(left=0.105, right=0.895, top=0.96, bottom=0.25)
    fig.savefig(output, dpi=250, facecolor="white")
    plt.close(fig)


def weekly_returns(indices: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> tuple[pd.Timestamp, pd.Series]:
    data = indices.loc[indices.index <= week_end]
    base = prior_base_date(data.index, week_start)
    end_candidates = data.index[data.index <= week_end]
    end = pd.Timestamp(end_candidates.max())
    missing = [name for name in FIGURE2_INDICES if name not in data.columns]
    if missing:
        raise ValueError(f"指数parquet缺少：{missing}")
    result = (data.loc[end, FIGURE2_INDICES] / data.loc[base, FIGURE2_INDICES] - 1) * 100
    result.index = ["沪深300", "中证500", "中证1000", "中证2000", "中证转债"]
    return base, result


def figure2(returns: pd.Series, output: Path) -> None:
    chinese_font, latin_font = setup_fonts()
    fig, axis = plt.subplots(figsize=(8, 4.75), dpi=250)
    x = np.arange(len(returns))
    bars = axis.bar(x, returns.to_numpy(), width=0.52, color=RED, edgecolor="none")
    axis.axhline(0, color="black", linewidth=0.8)
    axis.set_xticks(x, returns.index.tolist())
    for label in axis.get_xticklabels():
        label.set_fontproperties(chinese_font)
    for label in axis.get_yticklabels():
        label.set_fontproperties(latin_font)
    minimum, maximum = min(0.0, float(returns.min())), max(0.0, float(returns.max()))
    span = max(maximum - minimum, 1.0)
    lower = 0.0 if minimum >= 0 else math.floor((minimum - span * 0.12) / 2) * 2
    upper = 0.0 if maximum <= 0 else max(2.0, math.ceil((maximum + span * 0.18) / 2) * 2)
    axis.set_ylim(lower, upper)
    axis.yaxis.set_major_locator(MultipleLocator(2))
    axis.tick_params(axis="both", labelsize=12.5, length=4, width=0.8)
    offset = span * 0.035
    for bar, value in zip(bars, returns.to_numpy()):
        axis.text(
            bar.get_x() + bar.get_width() / 2,
            value + offset if value >= 0 else value - offset,
            f"{value:.2f}", ha="center", va="bottom" if value >= 0 else "top",
            fontproperties=latin_font, color="black",
        )
    axis.grid(False)
    axis.spines["top"].set_visible(False)
    axis.spines["right"].set_visible(False)
    axis.margins(x=0.08)
    fig.subplots_adjust(left=0.105, right=0.97, top=0.95, bottom=0.17)
    fig.savefig(output, dpi=250, facecolor="white")
    plt.close(fig)


def two_month_ticks(start: pd.Timestamp, end: pd.Timestamp) -> list[pd.Timestamp]:
    tick = pd.Timestamp(start.year, start.month, 3)
    if tick < start:
        tick += pd.DateOffset(months=2)
    ticks: list[pd.Timestamp] = []
    while tick <= end:
        ticks.append(tick)
        tick += pd.DateOffset(months=2)
    return ticks


def nice_ylim(values: pd.Series, step: float, zero_floor: bool) -> tuple[float, float]:
    finite = pd.to_numeric(values, errors="coerce").replace([np.inf, -np.inf], np.nan).dropna()
    low, high = float(finite.min()), float(finite.max())
    pad = max(step * 0.25, (high - low) * 0.05)
    lower = math.floor((low - pad) / step) * step
    upper = math.ceil((high + pad) / step) * step
    if zero_floor:
        lower = max(0.0, lower)
    return lower, upper if upper > lower else lower + step


def valuation_figure(data: pd.DataFrame, column: str, legend_name: str, output: Path, y_step: float, zero_floor: bool) -> dict[str, float]:
    series = data.set_index("日期")[column].dropna().sort_index()
    quantiles = series.quantile([0.25, 0.50, 0.75])
    plt.rcParams.update({"font.family": ["SimSun", "Times New Roman"], "axes.unicode_minus": False, "font.size": 20})
    fig, ax = plt.subplots(figsize=(14.25, 8.535), dpi=200)
    ax.plot(series.index, series.values, color=RED, linewidth=3.8, label=legend_name, zorder=4)
    for q, color, label in [(0.25, BLUE, "25%"), (0.50, GRAY, "50%"), (0.75, LIGHT_RED, "75%")]:
        ax.axhline(float(quantiles.loc[q]), color=color, linewidth=3.5, linestyle=(0, (4.5, 4.5)), dash_capstyle="round", label=label)
    ax.set_xlim(series.index.min(), series.index.max())
    ax.set_xticks(two_month_ticks(series.index.min(), series.index.max()))
    ax.xaxis.set_major_formatter(FuncFormatter(lambda value, _: f"{mdates.num2date(value).year}/{mdates.num2date(value).month}/{mdates.num2date(value).day}"))
    plt.setp(ax.get_xticklabels(), rotation=48, ha="right", rotation_mode="anchor")
    all_y = pd.concat([series, pd.Series(quantiles.values)])
    ax.set_ylim(*nice_ylim(all_y, y_step, zero_floor))
    ax.yaxis.set_major_locator(MultipleLocator(y_step))
    ax.yaxis.set_major_formatter(FormatStrFormatter("%.2f"))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(axis="both", labelsize=20, width=1.2, length=6, pad=8)
    ax.grid(False)
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, -0.25), ncol=4, frameon=False, fontsize=21, handlelength=4.5, columnspacing=2.0)
    fig.subplots_adjust(left=0.10, right=0.985, top=0.965, bottom=0.31)
    fig.savefig(output, dpi=200, facecolor="white")
    plt.close(fig)
    return {
        "最新值": float(series.iloc[-1]),
        "25%分位数": float(quantiles.loc[0.25]),
        "50%分位数": float(quantiles.loc[0.50]),
        "75%分位数": float(quantiles.loc[0.75]),
    }


@dataclass
class IFindSession:
    active: bool = False
    module: object | None = None

    def open(self, auto: bool, strict: bool) -> bool:
        if not auto:
            return False
        username, password = load_ifind_credentials()
        try:
            import iFinDPy as module
            result = module.THS_iFinDLogin(username, password)
            code = int(result) if isinstance(result, (int, np.integer)) else int(getattr(result, "errorcode", -1))
            if code not in {0, -201}:
                raise RuntimeError(f"iFinD登录失败，错误码{code}")
            self.active, self.module = True, module
            print_ifind_usage(module)
            return True
        except Exception as exc:
            if strict:
                raise
            log(f"[警告] iFinD实时接口不可用，改用本地数据：{exc}")
            return False

    def close(self) -> None:
        if self.active and self.module is not None:
            try:
                self.module.THS_iFinDLogout()
            except Exception:
                pass
        self.active = False


def normalize_ifind_frame(
    result,
    api: str,
    *,
    allow_no_data: bool = False,
    empty_columns: Iterable[str] = (),
) -> pd.DataFrame:
    code = int(getattr(result, "errorcode", -1))
    if allow_no_data and code == -4001:
        return pd.DataFrame(columns=list(empty_columns))
    if code != 0:
        raise RuntimeError(f"{api}失败：{code} {getattr(result, 'errmsg', '')}")
    data = getattr(result, "data", None)
    if not isinstance(data, pd.DataFrame):
        raise RuntimeError(f"{api}未返回DataFrame")
    return data.copy()


def fetch_proposals(session: IFindSession, start: pd.Timestamp, end: pd.Timestamp) -> pd.DataFrame:
    fields = (
        "p03153_f029:Y,p03153_f001:Y,p03153_f002:Y,p03153_f003:Y,p03153_f004:Y,"
        "p03153_f005:Y,p03153_f006:Y,p03153_f012:Y,p03153_f013:Y,p03153_f023:Y,"
        "p03153_f024:Y,p03153_f026:Y"
    )
    params = f"STARTTRADEDATE={start:%Y%m%d};ENDTRADEDATE={end:%Y%m%d};PROJECTKIND=全部;fxqk=否"
    raw = normalize_ifind_frame(session.module.THS_DR("p03153", params, fields, "format:dataframe"), "THS_DR(p03153)")
    mapping = dict(zip(
        ["p03153_f029", "p03153_f001", "p03153_f002", "p03153_f003", "p03153_f004", "p03153_f005", "p03153_f006", "p03153_f012", "p03153_f013", "p03153_f023", "p03153_f024", "p03153_f026"],
        PROPOSAL_COLUMNS,
    ))
    missing = set(mapping) - set(raw.columns)
    if missing:
        raise RuntimeError(f"p03153缺少字段：{sorted(missing)}")
    return standardize_proposals(raw[list(mapping)].rename(columns=mapping))


def standardize_proposals(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    for column in PROPOSAL_COLUMNS:
        if column not in result:
            result[column] = pd.NA
    result = result[PROPOSAL_COLUMNS]
    result = result[~result["发行方式"].astype("string").str.contains("定向", na=False)]
    for column in ["最新公告日期", "预案公告日", "股东大会公告日", "发审委审核公告日", "证监会批准公告日"]:
        result[column] = pd.to_datetime(result[column], format="mixed", errors="coerce")
    for column in ["发行规模(亿元)", "发行期限(年)"]:
        result[column] = pd.to_numeric(result[column], errors="coerce")
    return result.sort_values(["最新公告日期", "公司代码"], ascending=[False, True], na_position="last").reset_index(drop=True)


def load_proposal_fallback(report_date: pd.Timestamp, output_dir: Path) -> pd.DataFrame:
    candidates = [
        ROOT / "runs" / "weekly" / f"人保周报{report_date:%Y%m%d}" / f"转债发行预案_剔除定向_{report_date:%Y%m%d}.xlsx",
        output_dir / f"转债发行预案_剔除定向_{report_date:%Y%m%d}.xlsx",
        ROOT / "runs" / "weekly" / f"发行预案{report_date:%Y%m%d}.xlsx",
    ]
    candidates.extend(sorted((ROOT / "runs" / "weekly").glob("人保周报*/转债发行预案_剔除定向_*.xlsx"), reverse=True))
    for path in candidates:
        if not path.exists():
            continue
        try:
            frame = pd.read_excel(path, sheet_name=0)
            if {"公司名称", "方案进度", "发行规模(亿元)"}.issubset(frame.columns):
                log(f"[本地] 使用发行预案：{path.name}")
                return standardize_proposals(frame)
        except Exception:
            continue
    return pd.DataFrame(columns=PROPOSAL_COLUMNS)


def fetch_market_overview(session: IFindSession, as_of: pd.Timestamp) -> tuple[pd.DataFrame, pd.DataFrame]:
    dr = normalize_ifind_frame(
        session.module.THS_DR(
            "p00570",
            f"jyzt=未到期;sfdb=全部;jysc=全部;edate={as_of:%Y%m%d}",
            "jydm:Y,jydm_mc:Y,p00570_f001:Y,p00570_f019:Y",
            "format:dataframe",
        ),
        "THS_DR(p00570)",
    )
    codes = ",".join(dr["jydm"].astype(str))
    raw = normalize_ifind_frame(
        session.module.THS_BD(
            codes,
            "ths_convertible_debt_short_name_cbond;ths_stock_code_cbond;ths_stock_short_name_cbond;"
            "ths_issue_method_cbond;ths_trading_status_bond;ths_bond_balance_cbond;ths_listed_date_cbond",
            f";;;;;{as_of:%Y-%m-%d};",
        ),
        "THS_BD(转债基础信息)",
    )
    rename = {
        "thscode": BOND_CODE,
        "ths_convertible_debt_short_name_cbond": "转债名称",
        "ths_stock_code_cbond": "正股代码",
        "ths_stock_short_name_cbond": "公司名称",
        "ths_issue_method_cbond": "发行方式",
        "ths_trading_status_bond": "交易状态",
        "ths_bond_balance_cbond": "余额",
        "ths_listed_date_cbond": "上市日期",
    }
    basic = raw.rename(columns=rename)
    basic = basic[~basic["发行方式"].astype("string").str.contains("定向", na=False)]
    basic = basic[~basic[BOND_CODE].astype("string").str.contains("NQ", na=False)]
    basic = basic[~basic["交易状态"].astype("string").str.contains("终止上市", na=False)]
    basic["余额"] = pd.to_numeric(basic["余额"], errors="coerce")
    basic["上市日期"] = pd.to_datetime(basic["上市日期"], errors="coerce")
    upcoming = normalize_ifind_frame(
        session.module.THS_DR("p00600", "zqlx=640007", "p00600_f001:Y,p00600_f004:Y,p00600_f044:Y", "format:dataframe"),
        "THS_DR(p00600)",
        allow_no_data=True,
        empty_columns=["p00600_f001", "p00600_f004", "p00600_f044"],
    )
    return basic, upcoming


def local_market_overview(panel: pd.DataFrame, master: pd.DataFrame, as_of: pd.Timestamp) -> tuple[pd.DataFrame, pd.DataFrame]:
    snap = panel[panel[TRADE_DATE].eq(as_of)].set_index(BOND_CODE)
    result = master.copy()
    result["余额"] = pd.to_numeric(snap["余额"].reindex(result.index), errors="coerce")
    issued = result["发行日期"].notna() & (result["发行日期"] <= as_of)
    positive = result["余额"].fillna(0).gt(0)
    unlisted = issued & (result["上市日期"].isna() | (result["上市日期"] > as_of))
    keep = positive | unlisted
    basic = result.loc[keep, [BOND_CODE, "转债名称", "上市日期", "余额"]].copy()
    basic["公司名称"] = basic["转债名称"]
    basic["交易状态"] = np.where(basic["上市日期"].isna() | (basic["上市日期"] > as_of), "已发行未上市", "交易")
    basic["发行方式"] = ""
    return basic.reset_index(drop=True), pd.DataFrame()


def fetch_delist_dates_wind(codes: list[str], timeout_seconds: int = 300) -> pd.Series:
    import win32com.client
    with tempfile.TemporaryDirectory(prefix="renbao_wind_") as temp_dir:
        workbook_path = Path(temp_dir) / "delist.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.append(["wind_code", "s_info_delistdate"])
        for row, code in enumerate(codes, 2):
            sheet.cell(row=row, column=1, value=code)
            sheet.cell(row=row, column=2, value=f'=WSS(A{row},"s_info_delistdate")')
        book.save(workbook_path)
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = None
        try:
            wb = excel.Workbooks.Open(str(workbook_path))
            excel.CalculateFullRebuild()
            try:
                excel.CalculateUntilAsyncQueriesDone()
            except Exception:
                pass
            ws = wb.Worksheets(1)
            deadline = time.monotonic() + timeout_seconds
            values = None
            while time.monotonic() < deadline:
                time.sleep(2)
                values = ws.Range(f"B2:B{len(codes) + 1}").Value2
                valid = sum(row[0] not in (None, "", "数据获取中") for row in values)
                if valid and excel.CalculationState == 0:
                    break
            if values is None:
                raise RuntimeError("Wind Excel没有返回摘牌日期")
            parsed = []
            for row in values:
                value = row[0]
                if isinstance(value, (int, float)):
                    parsed.append(pd.Timestamp("1899-12-30") + pd.to_timedelta(float(value), unit="D"))
                else:
                    parsed.append(pd.to_datetime(value, errors="coerce"))
            return pd.Series(parsed, index=codes, name="摘牌日期")
        finally:
            if wb is not None:
                wb.Close(False)
            excel.Quit()


def current_balances(panel: pd.DataFrame, day: pd.Timestamp) -> pd.Series:
    snap = panel[panel[TRADE_DATE].eq(day)].set_index(BOND_CODE)
    return pd.to_numeric(snap["余额"], errors="coerce")


def delist_and_redeeming(master: pd.DataFrame, balances: pd.Series, as_of: pd.Timestamp, delist_dates: pd.Series) -> tuple[pd.DataFrame, pd.DataFrame]:
    data = master.copy()
    data["摘牌日期"] = data.index.to_series().map(delist_dates)
    data["余额"] = balances.reindex(data.index)
    window_end = as_of + pd.DateOffset(months=1)
    selected = data[data["摘牌日期"].gt(as_of) & data["摘牌日期"].le(window_end)].copy()
    selected = selected[selected["余额"].fillna(0).gt(0)].sort_values("摘牌日期")
    not_delisted = data["摘牌日期"].gt(as_of) | (data["摘牌日期"].isna() & data["最后交易日"].gt(as_of))
    redeeming = data[data["赎回公告日"].notna() & data["赎回公告日"].le(as_of) & not_delisted].copy()
    redeeming = redeeming[redeeming["余额"].fillna(0).gt(0)].sort_values("赎回公告日", ascending=False)
    return selected, redeeming


def read_redemption_book() -> tuple[pd.DataFrame, pd.DataFrame]:
    if not REDEMPTION_BOOK.exists():
        return pd.DataFrame(), pd.DataFrame()
    redemption = pd.read_excel(REDEMPTION_BOOK, sheet_name="赎回")
    redemption = redemption.iloc[:, :4]
    redemption.columns = [BOND_CODE, "转债名称", "公告时间", "公告标题"]
    redemption["公告时间"] = pd.to_datetime(redemption["公告时间"], errors="coerce")
    no_redemption = pd.read_excel(REDEMPTION_BOOK, sheet_name="不赎回")
    no_redemption = no_redemption.iloc[:, :5]
    no_redemption.columns = [BOND_CODE, "转债名称", "公告时间", "承诺截止日", "公告标题"]
    no_redemption["公告时间"] = pd.to_datetime(no_redemption["公告时间"], errors="coerce")
    no_redemption["承诺截止日"] = pd.to_datetime(no_redemption["承诺截止日"], errors="coerce")
    return redemption.dropna(subset=[BOND_CODE]), no_redemption.dropna(subset=[BOND_CODE])


def strong_redemption_text(panel: pd.DataFrame, master: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> str:
    redemption, no_redemption = read_redemption_book()
    end_exclusive = week_end + pd.Timedelta(days=1)
    announced = redemption[redemption["公告时间"].ge(week_start) & redemption["公告时间"].lt(end_exclusive)] if not redemption.empty else pd.DataFrame()
    announced_names = announced["转债名称"].dropna().astype(str).drop_duplicates().tolist()

    snap = panel[panel[TRADE_DATE].eq(week_end)].set_index(BOND_CODE)
    trigger = pd.to_numeric(snap["赎回累计天数"], errors="coerce").dropna()
    strong_codes = set(redemption[BOND_CODE].astype(str)) if not redemption.empty else set()
    recently_declined = set()
    if not no_redemption.empty:
        active = no_redemption["承诺截止日"].ge(week_end) | no_redemption["公告时间"].ge(week_start - pd.Timedelta(days=7))
        recently_declined = set(no_redemption.loc[active, BOND_CODE].astype(str))
    candidates = trigger[~trigger.index.astype(str).isin(strong_codes | recently_declined)]
    candidates = candidates[candidates > 0].sort_values(ascending=False)
    if not candidates.empty:
        cutoff = candidates.iloc[min(2, len(candidates) - 1)]
        candidates = candidates[candidates >= cutoff].head(4)
    near_names = master["转债名称"].reindex(candidates.index).dropna().astype(str).tolist()
    announced_part = chinese_join(announced_names) if announced_names else "暂无新增转债"
    near_part = chinese_join(near_names) if near_names else "暂无"
    return f"强赎方面，公告强赎的有{announced_part}，强赎触发较近的有{near_part}。"


def individual_weekly_returns(panel: pd.DataFrame, master: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> pd.DataFrame:
    sample = panel[panel[TRADE_DATE].between(week_start, week_end)].copy()
    sample = sample.join(master[["上市日期", "最后交易日"]], on=BOND_CODE)
    valid = (
        (sample["上市日期"].isna() | (sample[TRADE_DATE] >= sample["上市日期"]))
        & (sample["最后交易日"].isna() | (sample[TRADE_DATE] <= sample["最后交易日"]))
        & sample["涨跌幅"].notna()
    )
    sample = sample.loc[valid]
    ret = sample.groupby(BOND_CODE)["涨跌幅"].apply(
        lambda values: ((1 + pd.to_numeric(values, errors="coerce").dropna() / 100).prod() - 1) * 100
    )
    result = master[["转债名称", "申万行业", "申万三级行业", "上市日期", "最后交易日"]].copy()
    result["周涨跌幅"] = ret.reindex(result.index)
    return result[result["周涨跌幅"].notna()].sort_values("周涨跌幅", ascending=False)


def industry_text(individual: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> str:
    grouped = individual.dropna(subset=["申万行业"]).groupby("申万行业")["周涨跌幅"].mean().sort_values(ascending=False)
    positive = grouped[grouped > 0].head(4)
    negative = grouped[grouped < 0].sort_values().head(3)
    up = "、".join(f"{name}（{value:+.2f}%）" for name, value in positive.items()) or "无行业"
    down = "、".join(f"{name}（{value:+.2f}%）" for name, value in negative.items()) or "无行业"
    return f"转债行业方面，{date_range_text(week_start, week_end)}，转债市场各行业表现分化，{up}等上涨；{down}下跌居前。"


def individual_text(individual: pd.DataFrame) -> str:
    top = individual.head(5)
    bottom = individual.tail(5).sort_values("周涨跌幅")
    def items(frame: pd.DataFrame) -> str:
        return chinese_join(
            f"{row['转债名称']}（{row['申万三级行业'] if pd.notna(row['申万三级行业']) else '行业未分类'}）"
            for _, row in frame.iterrows()
        )
    return f"个券方面，{items(top)}转债涨幅靠前；{items(bottom)}跌幅靠前。"


def fund_text(indices: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> str:
    base = prior_base_date(indices.index, week_start)
    values = (indices.loc[week_end, FUND_INDICES] / indices.loc[base, FUND_INDICES] - 1) * 100
    verb = "整体表现积极" if (values > 0).sum() >= 3 else "表现有所分化"
    def pair(first: float, second: float) -> str:
        if first >= 0 and second >= 0:
            return f"上涨{first:.2f}%、{second:.2f}%"
        if first < 0 and second < 0:
            return f"下跌{abs(first):.2f}%、{abs(second):.2f}%"
        return (
            f"{'上涨' if first >= 0 else '下跌'}{abs(first):.2f}%、"
            f"{'上涨' if second >= 0 else '下跌'}{abs(second):.2f}%"
        )
    return (
        f"基金表现方面，“固收+”基金{verb}，一级债基及二级债基指数分别"
        f"{pair(float(values.iloc[0]), float(values.iloc[1]))}，"
        f"可转债基金及灵活配置型基金指数则分别"
        f"{pair(float(values.iloc[2]), float(values.iloc[3]))}。"
    )


def turnover_text(panel: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> str:
    daily = panel.groupby(TRADE_DATE)["成交额"].sum(min_count=1).sort_index()
    current = daily[(daily.index >= week_start) & (daily.index <= week_end)].dropna()
    previous_start = week_start - pd.Timedelta(days=7)
    previous_end = week_start - pd.Timedelta(days=1)
    previous = daily[(daily.index >= previous_start) & (daily.index <= previous_end)].dropna()
    total, avg, previous_avg = float(current.sum()), float(current.mean()), float(previous.mean())
    change = avg / previous_avg - 1
    action = "放量" if change >= 0 else "缩量"
    return (
        f"成交额方面，{date_range_text(week_start, week_end)}，转债市场总成交额{total:.2f}亿元，"
        f"日均成交额{avg:.2f}亿元，较{date_range_text(pd.Timestamp(previous.index.min()), pd.Timestamp(previous.index.max()))}均值"
        f"{previous_avg:.2f}亿元{action}{abs(change):.2%}。"
    )


def market_description(returns: pd.Series) -> str:
    equities = returns.iloc[:4]
    if (equities > 0).all():
        direction = "普遍上涨"
    elif (equities < 0).all():
        direction = "普遍下跌"
    else:
        direction = "表现分化"
    large = float(equities.iloc[0])
    small = float(equities.iloc[1:].mean())
    size = "中小盘表现占优" if small > large else "大盘表现相对占优"
    cb = float(returns.iloc[4])
    equity_avg = float(equities.mean())
    if cb >= 0 and equity_avg >= 0:
        relative = "转债市场同步上涨" + ("但涨幅相对有限" if cb < equity_avg else "且表现相对占优")
    elif cb < 0 and equity_avg < 0:
        relative = "转债市场同步下跌" + ("但相对抗跌" if cb > equity_avg else "且跌幅相对较大")
    else:
        relative = "转债市场与权益市场表现有所背离"
    return f"主要股指{direction}，{size}，{relative}。"


def overview_text(basic: pd.DataFrame, upcoming: pd.DataFrame, parquet_balance: float) -> str:
    count = len(basic)
    # 周报存续余额统一使用报告周五 Parquet 截面的“余额”汇总，
    # iFinD 基础信息仅用于存续数量、未上市名单和未来发行日历。
    balance = float(parquet_balance)
    unlisted = basic[basic["交易状态"].astype("string").eq("已发行未上市") | basic["上市日期"].isna()]
    if unlisted.empty:
        listed_part = "已发行转债均已上市进行交易，"
    else:
        listed_part = f"已发行转债中，{chinese_join(unlisted['转债名称'].astype(str))}尚未上市进行交易，"
    upcoming_names = []
    if not upcoming.empty and "p00600_f044" in upcoming:
        upcoming_names = upcoming["p00600_f044"].dropna().astype(str).tolist()
    upcoming_part = f"其中{chinese_join(upcoming_names)}即将网上发行。" if upcoming_names else "目前尚无将发行转债。"
    return f"现已发行未到期可转债有{count}支，余额规模{balance:.2f}亿元，{listed_part}{upcoming_part}"


def exit_text(delist: pd.DataFrame, redeeming: pd.DataFrame) -> str:
    delist_names = chinese_join(delist["转债名称"].astype(str))
    delist_balance = pd.to_numeric(delist["余额"], errors="coerce").sum(min_count=1)
    redeem_balance = pd.to_numeric(redeeming["余额"], errors="coerce").sum(min_count=1)
    return (
        f"退出方面，未来一个月内将摘牌转债目前有{len(delist)}只，合计余额{delist_balance:.2f}亿元，"
        f"包括{delist_names}，正在赎回中的转债共{len(redeeming)}只，合计余额{redeem_balance:.2f}亿元。"
    )


def proposal_commentary(proposals: pd.DataFrame, master: pd.DataFrame, basic: pd.DataFrame, week_start: pd.Timestamp, week_end: pd.Timestamp) -> str:
    issued = master[master["发行日期"].between(week_start, week_end, inclusive="both")]
    listed = master[master["上市日期"].between(week_start, week_end, inclusive="both")]
    issued_names = chinese_join(issued["转债名称"].astype(str))
    # 新券上市统一使用转债名称，不使用发行主体或正股简称。
    listed_names = chinese_join(listed["转债名称"].astype(str))
    parts = [
        f"一级市场方面，{date_range_text(week_start, week_end)}，{len(issued)}只转债发行，包括{issued_names}，合计规模{issued['发行规模'].sum():.2f}亿元；"
        f"{listed_names}新券上市，规模合计{listed['发行规模'].sum():.2f}亿元。"
    ]
    events = [
        ("预案公告日", "新增董事会预案"),
        ("股东大会公告日", "新增股东大会通过"),
        ("发审委审核公告日", "新增发审委审批通过"),
        ("证监会批准公告日", "新增证监会核准"),
    ]
    for column, label in events:
        if column not in proposals:
            continue
        names = proposals.loc[proposals[column].between(week_start, week_end, inclusive="both"), "公司名称"].dropna().astype(str).drop_duplicates().tolist()
        if names:
            parts.append(f"{chinese_join(names)}{label}。")
    active = proposals[~proposals["方案进度"].astype("string").str.contains("停止实施|终止", na=False)]
    registered = active[active["方案进度"].astype("string").str.contains("证监会批准|同意注册", na=False)]
    committee = active[active["方案进度"].astype("string").str.contains("上市委通过|发审委通过", na=False)]
    parts.append(
        f"截至目前，待发可转债共计{len(active)}只，合计规模{active['发行规模(亿元)'].sum():,.2f}亿元，"
        f"其中已被同意注册的{len(registered)}只，规模合计{registered['发行规模(亿元)'].sum():.2f}亿元；"
        f"已获上市委通过的{len(committee)}只，规模合计{committee['发行规模(亿元)'].sum():.2f}亿元。"
    )
    return "".join(parts)


def write_proposal_workbook(proposals: pd.DataFrame, output: Path, source_note: str) -> None:
    display = proposals.copy()
    date_columns = ["最新公告日期", "预案公告日", "股东大会公告日", "发审委审核公告日", "证监会批准公告日"]
    for column in date_columns:
        display[column] = pd.to_datetime(display[column], errors="coerce")
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        display.to_excel(writer, sheet_name="发行预案", index=False)
        pd.DataFrame(
            {
                "项目": ["数据来源", "筛选规则", "记录数"],
                "说明": [source_note, "剔除发行方式包含‘定向’的记录", len(display)],
            }
        ).to_excel(writer, sheet_name="说明", index=False)
    workbook = openpyxl.load_workbook(output)
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.row_dimensions[1].height = 24
        for cell in sheet[1]:
            cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="C00000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="微软雅黑", size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for column in range(1, sheet.max_column + 1):
            width = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, min(sheet.max_row, 200) + 1))
            sheet.column_dimensions[get_column_letter(column)].width = min(max(width * 1.5 + 2, 12), 28)
        sheet.auto_filter.ref = sheet.dimensions
    proposal_sheet = workbook["发行预案"]
    date_column_numbers = [PROPOSAL_COLUMNS.index(column) + 1 for column in date_columns]
    for row in range(2, proposal_sheet.max_row + 1):
        for column in date_column_numbers:
            proposal_sheet.cell(row, column).number_format = "yyyy-mm-dd"
            proposal_sheet.cell(row, column).alignment = Alignment(horizontal="center", vertical="center")
        for column in [6, 7]:
            proposal_sheet.cell(row, column).number_format = "0.00"
            proposal_sheet.cell(row, column).alignment = Alignment(horizontal="right", vertical="center")
    workbook.save(output)


def write_text(output: Path, paragraphs: list[str], week_start: pd.Timestamp, week_end: pd.Timestamp) -> None:
    content = [f"数据周期：{week_start.year}年{month_day(week_start)}至{week_end.year}年{month_day(week_end)}", ""]
    for paragraph in paragraphs:
        content.extend([paragraph, ""])
    output.write_text("\n".join(content).rstrip() + "\n", encoding="utf-8")


def validate_outputs(output_dir: Path, report_date: pd.Timestamp) -> None:
    expected = [
        f"人保周报{report_date:%Y%m%d}.txt",
        "图1_中证转债与万得全A时序图.png",
        "图2_主要指数上周涨跌幅.png",
        "图3_转债价格中位数及历史分位数.png",
        "图4_转债百元拟合溢价率及历史分位数.png",
        f"转债发行预案_剔除定向_{report_date:%Y%m%d}.xlsx",
    ]
    missing = [name for name in expected if not (output_dir / name).exists() or (output_dir / name).stat().st_size == 0]
    if missing:
        raise RuntimeError(f"输出校验失败：{missing}")
    obsolete = output_dir / "图3_图4_数据_2023年以来.csv"
    if obsolete.exists():
        obsolete.unlink()
    log("输出校验通过：" + "、".join(expected))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="基于Parquet v2.1一键生成人保转债市场周报")
    parser.add_argument("--week-end", help="报告周截止交易日，YYYY-MM-DD")
    parser.add_argument("--report-date", help="文件命名日期，YYYY-MM-DD")
    parser.add_argument("--output-dir", type=Path, help="自定义输出目录，默认人保周报YYYYMMDD")
    parser.add_argument("--offline", action="store_true", help="不调用Wind和iFinD，仅用于本地验证")
    parser.add_argument("--strict-external", action="store_true", help="外部接口失败时直接终止")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    week_start, week_end, report_date = choose_period(args.week_end, args.report_date)
    output_dir = (args.output_dir or (ROOT / "runs" / "weekly" / f"人保周报{report_date:%Y%m%d}")).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    log(f"报告周期：{week_start:%Y-%m-%d}至{week_end:%Y-%m-%d}；命名日期：{report_date:%Y-%m-%d}")

    master = load_master()
    panel_start = min(START_CHART, week_start - pd.Timedelta(days=14))
    panel = load_panel(panel_start, week_end)
    indices = load_indices(START_CHART, week_end)
    valuation = build_daily_valuation(panel, master, week_end)

    figure1(indices, output_dir / "图1_中证转债与万得全A时序图.png")
    _, index_returns = weekly_returns(indices, week_start, week_end)
    figure2(index_returns, output_dir / "图2_主要指数上周涨跌幅.png")
    fig3_stats = valuation_figure(
        valuation, "转债价格中位数", "转债价格中位数",
        output_dir / "图3_转债价格中位数及历史分位数.png", 10.0, False,
    )
    fig4_stats = valuation_figure(
        valuation, "百元平价拟合溢价率", "百元平价拟合溢价率",
        output_dir / "图4_转债百元拟合溢价率及历史分位数.png", 5.0, True,
    )

    session = IFindSession()
    session.open(auto=not args.offline, strict=args.strict_external)
    proposal_source = "本地最近一次发行预案"
    try:
        if session.active:
            proposals = fetch_proposals(session, report_date - pd.DateOffset(years=1), report_date)
            proposal_source = f"iFinD THS_DR(p03153)，截至{report_date:%Y-%m-%d}"
            basic, upcoming = fetch_market_overview(session, week_end)
        else:
            proposals = load_proposal_fallback(report_date, output_dir)
            basic, upcoming = local_market_overview(panel, master, week_end)
    finally:
        session.close()

    balances = current_balances(panel, week_end)
    parquet_balance = balances.sum(min_count=1)
    # “未来一个月将摘牌”按最后交易日统计，避免提前赎回品种的摘牌日更新滞后。
    delist_dates = master["最后交易日"].copy()
    delist, redeeming = delist_and_redeeming(master, balances, week_end, delist_dates)

    individual = individual_weekly_returns(panel, master, week_start, week_end)
    first_paragraph = (
        f"{date_range_text(week_start, week_end)}，{market_description(index_returns)}"
        f"{overview_text(basic, upcoming, parquet_balance)}{exit_text(delist, redeeming)}"
    )
    primary = proposal_commentary(proposals, master, basic, week_start, week_end)
    paragraphs = [
        first_paragraph,
        primary,
        individual_text(individual),
        industry_text(individual, week_start, week_end),
        fund_text(indices, week_start, week_end),
        turnover_text(panel, week_start, week_end),
        strong_redemption_text(panel, master, week_start, week_end),
        (
            f"图1数据说明：中证转债使用左轴及红色折线（{RED}），万得全A使用右轴及蓝色折线（{BLUE}）；"
            f"数据来自parquet，区间为{indices.index.min():%Y年%m月%d日}至{indices.index.max():%Y年%m月%d日}，共{len(indices[['转债指数','万得全A']].dropna())}个共同交易日。"
        ),
    ]
    write_text(output_dir / f"人保周报{report_date:%Y%m%d}.txt", paragraphs, week_start, week_end)
    write_proposal_workbook(proposals, output_dir / f"转债发行预案_剔除定向_{report_date:%Y%m%d}.xlsx", proposal_source)
    validate_outputs(output_dir, report_date)

    log(
        f"图3：最新{fig3_stats['最新值']:.2f}，25%/50%/75%="
        f"{fig3_stats['25%分位数']:.2f}/{fig3_stats['50%分位数']:.2f}/{fig3_stats['75%分位数']:.2f}"
    )
    log(
        f"图4：最新{fig4_stats['最新值']:.2f}%，25%/50%/75%="
        f"{fig4_stats['25%分位数']:.2f}%/{fig4_stats['50%分位数']:.2f}%/{fig4_stats['75%分位数']:.2f}%"
    )
    log(f"完成：{output_dir}")


if __name__ == "__main__":
    main()
