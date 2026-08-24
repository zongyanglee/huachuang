#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl
from datetime import datetime, timedelta
from pathlib import Path
import time
import warnings

# 仅以只读方式提取已保存的单元格值，不会回写底稿；忽略不影响取数的图形/格式兼容提示。
warnings.filterwarnings('ignore', message='wmf image format is not supported.*')
warnings.filterwarnings('ignore', message='Conditional Formatting extension is not supported.*')
warnings.filterwarnings('ignore', message='Unknown extension is not supported.*')

MMDD_today = str(time.strftime("%m%d", time.localtime()))
YYYYMMDD_today = str(time.strftime("%Y%m%d", time.localtime()))
folder_name = MMDD_today + '数据更新'

# 日报所有行情描述统一使用非剔妖统计口径。
filename = Path(f'【华创固收】转债日报底稿-{YYYYMMDD_today}-快照1.xlsx').resolve()
market_stats_path = (Path(folder_name) / f'{MMDD_today}数据更新（清理后）统计.xlsx').resolve()

for input_path in (filename, market_stats_path):
    if not input_path.exists():
        raise FileNotFoundError(f'未找到日报输入文件：{input_path}')


outputname = '【华创固收】转债日报' + MMDD_today + '.txt'
# 读取底稿中已保存的公式计算结果，无需启动 Excel。
wb2 = openpyxl.load_workbook(filename, data_only=True)
sht = wb2['展示页']

# 计算日期字符串
date_str = sht['D13'].value.strftime('%Y年%m月%d日')

# 动态生成日报文案中的日期标签
运行日期 = datetime.now().date()
运行日标签 = f"{运行日期.month}月{运行日期.day}日"

# 从底稿的实际交易日期中取运行日前最近一个交易日，可正确处理周末及休市日
交易日期原始值 = [单元格.value for 行 in wb2['2、资金表现']['A9:A60'] for 单元格 in 行]
交易日期列表 = [值.date() if isinstance(值, datetime) else 值 for 值 in 交易日期原始值 if 值 is not None]
历史交易日期 = [日期 for 日期 in 交易日期列表 if 日期 < 运行日期]
if not 历史交易日期:
    raise ValueError('底稿中未找到运行日前的交易日期，无法生成上一交易日标签')
上一交易日 = max(历史交易日期)
上一交易日标签 = f"{上一交易日.month}月{上一交易日.day}日"


# 读取 G17、G21 单元格的值
g17_value = sht['G17'].value
g21_value = sht['G21'].value

# 计算结果字符串
# result = date_str + '可转债市场主要指数变化情况：\n\n'
result = "市场概况："
if g17_value > 0 and g21_value > 0:
    if g17_value > g21_value:
        result += f'{运行日标签}转债跟随正股上涨，'
    elif g17_value < g21_value:
        result += f'{运行日标签}转债跟随正股上涨，'
    else:
        result += f'{运行日标签}转债跟随正股上涨，'
elif g17_value < 0 and g21_value < 0:
    if g17_value < g21_value:
        result += f'{运行日标签}转债跟随正股下跌，'
    elif g17_value > g21_value:
        result += f'{运行日标签}转债跟随正股下跌，'
    else:
        result += f'{运行日标签}转债跟随正股下跌，'
elif g17_value > 0 and g21_value < 0:
    result += f'{运行日标签}转债表现强于正股，转债市场逆势上涨{g17_value / 100:.2%}'
elif g17_value < 0 and g21_value > 0:
    result += f'{运行日标签}转债表现弱于正股，转债市场下降{g17_value / 100:.2%}'

result += "\n指数表现："
# 输出结果到文件
with open(outputname, 'w', encoding='utf-8') as f:
    f.write(result)


# In[2]:


ws = sht

# 获取各单元格位置的值
E21 = ws['E21'].value
G21 = ws['G21'].value
E22 = ws['E22'].value
G22 = ws['G22'].value
E23 = ws['E23'].value
G23 = ws['G23'].value
E24 = ws['E24'].value
G24 = ws['G24'].value
E25 = ws['E25'].value
G25 = ws['G25'].value
E17 = ws['E17'].value
G17 = ws['G17'].value
# 定义函数，输入参数为指数和变化率，输出结果为字符串
def format_result(index, change):
    if change > 0:
        return f"{index}指数环比上涨{abs(change/100):.2%}"
    elif change < 0:
        return f"{index}指数环比下降{abs(change/100):.2%}"
    else:
        return f"{index}指数保持不变"
    
# 定义指数和变化率的列表
indexes = [E17, E24, E25]
changes = [G17, G24, G25]

# 使用列表推导式将所有指数和变化率的结果生成为字符串
results = [format_result(index, change) for index, change in zip(indexes, changes)]

def format_result_zs(index_zs, change_zs):
    if change_zs > 0:
        return f"{index_zs}环比上涨{abs(change_zs/100):.2%}"
    elif change_zs < 0:
        return f"{index_zs}环比下降{abs(change_zs/100):.2%}"
    else:
        return f"{index_zs}保持不变"
    
# 定义指数和变化率的列表
indexes_zs = [E21, E22, E23]
changes_zs= [G21, G22, G23]

# 使用列表推导式将所有指数和变化率的结果生成为字符串
results_zs = [format_result_zs(index_zs, change_zs) for index_zs, change_zs in zip(indexes_zs, changes_zs)]

position = 1
results[position:position] = results_zs
# 将所有结果连接起来，并用逗号隔开
result_str = '、'.join(results)
result_str +=  '。'

# 追加写入文件
with open(outputname, 'a', encoding='utf-8') as f:
    f.write(result_str)


# In[3]:


ws = wb2['展示页']

# 定义市场风格指数名称列表
index_names = ['大盘成长', '大盘价值', '中盘成长', '中盘价值', '小盘成长', '小盘价值']
data = []
# 获取市场风格指数最大值的行号
# data = [cell.value for cell in ws['N20:N25']]
data = ws['N20:N25']
max_value = -float('inf')
max_row_num = None
for row_num, row in enumerate(data, 20):
    for cell in row:
        if cell.value and cell.value > max_value:
            max_value = cell.value
            max_row_num = row_num
# print(max_row_num)

# # 获取市场风格指数相对占优的结果
# is_index_up = '上涨' if ws.cell(row=max_row_num, column=14).value > 0 else '下降'

# # 获取市场风格指数相对占优的百分比
# up_pct = float(ws.cell(max_row_num, 14).value) / 100

def get_updown_info(row_num, ws):
    is_index_up = '环比上涨' if ws.cell(row=row_num, column=14).value > 0 else '环比下降'
    up_pct = float(ws.cell(row_num, 14).value) / 100
    return is_index_up, up_pct

is_index_up, up_pct = get_updown_info(max_row_num, ws)

# 获取市场风格指数最小值的行号
data = ws['N20:N25']
min_value = float('inf')
min_row_num = None
for row_num, row in enumerate(data, 20):
    for cell in row:
        if cell.value and cell.value < min_value:
            min_value = cell.value
            min_row_num = row_num
# print(min_row_num)

# # 获取市场风格指数相对不占优的结果
# is_index_down = '上涨' if ws.cell(row=min_row_num, column=14).value > 0 else '下降'

# # 获取市场风格指数相对不占优的百分比
# down_pct = float(ws.cell(min_row_num, 14).value) / 100

is_index_down, down_pct = get_updown_info(min_row_num, ws)

# 构造结果字符串
# result_str = f"\n\n市场风格指数方面：{index_names[max_row_num-20]}相对占优，{is_index_up}{abs(up_pct):.2%}；{index_names[min_row_num-20]}相对不占优，{is_index_down}{abs(down_pct):.2%}。此外，"
# result_str = f"\n市场风格：{index_names[max_row_num-20]}相对占优。具体来看，{index_names[max_row_num-20]}{is_index_up}{abs(up_pct):.2%}，"
result_str = f"\n市场风格：{index_names[max_row_num-20]}相对占优。"
# 获取其他四个未被输出指数的行数据和涨跌幅数据
# other_indices = [index_names[i] for i in range(len(index_names)) if i != max_row_num-20]
other_indices = [index_names[i] for i in range(len(index_names))]

def get_row_num_by_index_name(index_name):
    index_names_to_row_nums = {
        '大盘成长': 20,
        '大盘价值': 21,
        '中盘成长': 22,
        '中盘价值': 23,
        '小盘成长': 24,
        '小盘价值': 25,
    }
    return index_names_to_row_nums.get(index_name, None)

other_str = ""
for index_name in other_indices:
    other_row_num = get_row_num_by_index_name(index_name)
    is_index_up, up_pct = get_updown_info(other_row_num, ws)
    other_str += f"{index_names[other_row_num-20]}{is_index_up}{abs(up_pct):.2%}、"
other_str = other_str[:-1].replace('、', '、', 1)
result_str += other_str

# 追加写入文件
with open(outputname, 'a', encoding='utf-8') as f:
    f.write(result_str)
    f.write('。')


# In[4]:


# 


# In[5]:


sheet = wb2['2、资金表现']

# 获取单元格值
# d13_value = sheet['D13'].value
# tdays_offset_value = sheet['D13'].offset(row=1, column=0).value
bond_turnover_value = sheet['B9'].value
bond_turnover_value_lastday = sheet['B10'].value
ashare_turnover_value = sheet['C9'].value
ashare_turnover_value_lastday = sheet['C10'].value
rate_value = sheet['G8'].value
rate_yesterday_value = sheet['G9'].value
main_net_value = sheet['L13'].value

# 格式化字符串
result = (
    f"\n资金表现：转债市场成交情绪{'减弱' if bond_turnover_value - bond_turnover_value_lastday < 0 else '升温' if bond_turnover_value - bond_turnover_value_lastday > 0 else '维持'}"
    f"。可转债市场成交额为{bond_turnover_value:.2f}亿元，环比"
    f"{'减少' if bond_turnover_value - bond_turnover_value_lastday <= 0 else '增长'}"
    f"{abs((bond_turnover_value - bond_turnover_value_lastday) / bond_turnover_value_lastday) * 100:.2f}%"
    f"；万得全A总成交额为{ashare_turnover_value:.2f}亿元，环比"
    f"{'减少' if ashare_turnover_value - ashare_turnover_value_lastday <= 0 else '增长'}"
    f"{abs((ashare_turnover_value - ashare_turnover_value_lastday) / ashare_turnover_value_lastday) * 100:.2f}%"
    f"；沪深两市主力净{'流入' if main_net_value > 0 else '流出'}"
    f"{abs(main_net_value):.2f}{'亿元' if abs(main_net_value) > 1 else '万元'}。"
#     f"十年国债收益率环比"
#     f"{'降低' if rate_value - rate_yesterday_value <= 0 else '上升'}"
#     f"{abs(rate_value - rate_yesterday_value) * 100:.2f}bp至"
#     f"{rate_value / 100:.2%}。\n"
)

# 追加写入文件
with open(outputname, 'a', encoding='utf-8') as f:
    f.write(result)


# In[6]:


import openpyxl

# 打开 Excel 文件
market_stats_wb = openpyxl.load_workbook(market_stats_path)

# 选择名为“全样本余额”的工作表
worksheet = market_stats_wb["全样本余额"]

# 获取“转债收盘价加权平均值”这一列的数据
column_c = worksheet["C"]

# 获取“转债收盘价加权平均值”这一列的最后一个单元格和倒数第二个单元格
last_cell = column_c[-1]
previous_cell = column_c[-2]

# 获取昨日的“转债收盘价加权平均值”
previous_value = previous_cell.value

# 获取今天的“转债收盘价加权平均值”
current_value = last_cell.value

# 计算变化率
if previous_value is not None and current_value is not None:
    change_rate = (current_value - previous_value) / previous_value
else:
    change_rate = None

# 将变化率转换为上升或下降的描述
if change_rate is not None:
    if change_rate > 0:
        change_desc = f"上升{abs(change_rate):.2%}"
    elif change_rate < 0:
        change_desc = f"下降{abs(change_rate):.2%}"
    else:
        change_desc = "维持同一水平"
else:
    change_desc = "未知，请检查"
# 临时，仅表述方向，将变化率转换为上升或下降的描述    
if change_rate is not None:
    if change_rate > 0:
        change_desc_dir = f"提升"
    elif change_rate < 0:
        change_desc_dir = f"下降"
    else:
        change_desc_dir = "维持同一水平"
else:
    change_desc_dir = "未知，请检查"
# 将结果输出到txt文件中
# with open(outputname, 'a', encoding='utf-8') as f:
    
#     f.write(f"\n转债整体收盘价加权平均值为{current_value:.2f}，")
#     if change_rate is not None:
#         f.write(f"环比昨日{change_desc}。")
#     else:
#         f.write(change_desc)


# In[7]:


import openpyxl

# 打开 Excel 文件
# 选择名为“平底分类余额加权收盘价”的工作表
worksheet = market_stats_wb["平底分类余额加权收盘价"]

# 获取“偏股型转债的收盘价”、“偏债型转债的收盘价”、“平衡型转债的收盘价”这三列的数据
column_b = worksheet["B"]
column_c = worksheet["C"]
column_d = worksheet["D"]

# 获取每列的最后一个单元格和倒数第二个单元格的值
last_b = column_b[-1].value
last_c = column_c[-1].value
last_d = column_d[-1].value
previous_b = column_b[-2].value
previous_c = column_c[-2].value
previous_d = column_d[-2].value

# 计算变化率
change_rate_b = (last_b - previous_b) / previous_b if previous_b is not None and last_b is not None else None
change_rate_c = (last_c - previous_c) / previous_c if previous_c is not None and last_c is not None else None
change_rate_d = (last_d - previous_d) / previous_d if previous_d is not None and last_d is not None else None

# 将变化率转换为上升或下降的描述
change_desc_b = f"上升{abs(change_rate_b):.2%}" if change_rate_b is not None and change_rate_b > 0 else f"下降{abs(change_rate_b):.2%}" if change_rate_b is not None and change_rate_b < 0 else "维持同一水平" if change_rate_b == 0 else "未知"
change_desc_c = f"上升{abs(change_rate_c):.2%}" if change_rate_c is not None and change_rate_c > 0 else f"下降{abs(change_rate_c):.2%}" if change_rate_c is not None and change_rate_c < 0 else "维持同一水平" if change_rate_c == 0 else "未知"
change_desc_d = f"上升{abs(change_rate_d):.2%}" if change_rate_d is not None and change_rate_d > 0 else f"下降{abs(change_rate_d):.2%}" if change_rate_d is not None and change_rate_d < 0 else "维持同一水平" if change_rate_d == 0 else "未知"

# 将结果输出到txt文件中
# with open(outputname, 'a', encoding='utf-8') as f:
#     f.write(f"偏股型转债的收盘价{last_b:.2f}，")
#     f.write(f"环比{change_desc_b}；")
#     f.write(f"偏债型转债的收盘价{last_c:.2f}，")
#     f.write(f"环比{change_desc_c}；")
#     f.write(f"平衡型转债的收盘价{last_d:.2f}，")
#     f.write(f"环比{change_desc_d}。")


# In[8]:


import openpyxl

# 读取Excel文件
# 选择“收盘价区间数量比例”sheet
sheet = market_stats_wb["收盘价区间数量比例"]

# 获取最后一行和倒数第二行的单元格对象
today_row = sheet.max_row
yesterday_row = today_row - 1
today_cells = sheet[today_row][1:]
yesterday_cells = sheet[yesterday_row][1:]

# 统计高价券个数占比并计算变化率
today_high_ratio = sum([cell.value for cell in today_cells[6:8]]) / sum([cell.value for cell in today_cells[:]])
yesterday_high_ratio = sum([cell.value for cell in yesterday_cells[6:8]]) / sum([cell.value for cell in yesterday_cells[:]])
change_ratio = (today_high_ratio - yesterday_high_ratio) * 100

# 输出上升或下降百分点数值
def pct_change_ratio(change_ratio):
    if change_ratio > 0:
        result = "上升" + str(round(change_ratio, 2)) + "pct"
    elif change_ratio < 0:
        result = "下降" + str(round(abs(change_ratio), 2)) + "pct"
    else:
        result = "维持同一水平"
    return result
        
# 计算每列占比变化率并输出变化最大的列名和变化率
max_col = None
max_change_ratio = 0
for col in ["E", "F", "G"]:
    today_ratio = sheet[col][-1].value / sum([cell.value for cell in today_cells[:]])
    yesterday_ratio = sheet[col][-2].value / sum([cell.value for cell in yesterday_cells[:]])
    col_change_ratio = (today_ratio - yesterday_ratio) * 100
    if abs(col_change_ratio) > abs(max_change_ratio):
        max_ratio = today_ratio
        max_col = sheet[col + "1"].value
        max_change_ratio = col_change_ratio
        
# formatted_ratio = f"{today_high_ratio:.2%}"  
百元至130元区间数量 = sheet["B"][-1].value + sheet["C"][-1].value + sheet["D"][-1].value
     
高价券结论 = "高价券占比" + ('提升。' if change_ratio > 0 else '下降。' if change_ratio < 0 else '维持相同水平。')

# 输出结果到txt文件
with open(outputname, 'a', encoding='utf-8') as f:
    f.write(f"转债价格：转债中枢{change_desc_dir}，{高价券结论}")
    f.write(f"转债整体收盘价加权平均值为{current_value:.2f}元，")
    if change_rate is not None:
        f.write(f"环比{上一交易日标签}{change_desc}。")
    else:
        f.write(change_desc)
    f.write(f"其中偏股型转债的收盘价为{last_b:.2f}元，")
    f.write(f"环比{change_desc_b}；")
    f.write(f"偏债型转债的收盘价为{last_c:.2f}元，")
    f.write(f"环比{change_desc_c}；")
    f.write(f"平衡型转债的收盘价为{last_d:.2f}元，")
    f.write(f"环比{change_desc_d}。")
    f.write(f"从转债收盘价分布情况看，")
    f.write("130元以上高价券个数占比" + f"{today_high_ratio:.2%}" + f"，较{上一交易日标签}环比" + pct_change_ratio(change_ratio) + "；")
    f.write("占比变化最大的区间为" + str(max_col) + "，占比" + f"{max_ratio:.2%}，")
    f.write(f"较{上一交易日标签}" + pct_change_ratio(max_change_ratio) + "；")
    f.write("收盘价在100元以下的个券有" + str(百元至130元区间数量) + "只。")


# In[ ]:





# In[9]:


worksheet = market_stats_wb['收盘价分位数统计']

# 获取 D 列最后一个有数字的单元格的行数
last_row = None
for cell in reversed(list(worksheet['D'])):
    if cell.value is not None:
        last_row = cell.row
        break

# 获取 D 列最后一个有数字的单元格的值
最新中位数 = worksheet.cell(row=last_row, column=4).value

# 获取 D 列倒数第二个有数字的单元格的行数
second_last_row = None
for cell in reversed(list(worksheet['D'])):
    if cell.row != last_row and cell.value is not None:
        second_last_row = cell.row
        break

# 获取 Q 列倒数第二个有数字的单元格的值
昨日中位数 = worksheet.cell(row=second_last_row, column=4).value

中位数变化率 = (最新中位数-昨日中位数)/昨日中位数

with open(outputname, 'a', encoding='utf-8') as f:
    f.write(f"价格中位数为{最新中位数:.2f}元，环比{上一交易日标签}{'上升' if 中位数变化率 > 0 else '下降' if 中位数变化率 < 0 else '维持相同水平'}{'' if 中位数变化率 == 0 else abs(中位数变化率):.2%}。\n")


# In[10]:


worksheet = market_stats_wb['收盘价分位数统计']

# 列名及对应列号
column_names = {
    "B": 2,
    "C": 3,
    "E": 5,
    "F": 6,
    "G": 7
}
result = ""
# 循环计算每列的变化率
for col, col_num in column_names.items():
    # 获取最后一个有数字的单元格的行数
    last_row = None
    for cell in reversed(list(worksheet[col])):
        if cell.value is not None:
            last_row = cell.row
            break

    # 获取最新值
    latest_value = worksheet.cell(row=last_row, column=col_num).value

    # 获取倒数第二个有数字的单元格的行数
    second_last_row = None
    for cell in reversed(list(worksheet[col])):
        if cell.row != last_row and cell.value is not None:
            second_last_row = cell.row
            break

    # 获取昨日值
    yesterday_value = worksheet.cell(row=second_last_row, column=col_num).value

    # 计算变化率
    change_ratio = (latest_value - yesterday_value) / yesterday_value
    # 输出结果
    result += ('+' if change_ratio > 0 else ('-' if change_ratio < 0 else '+')) + str(round(change_ratio * 100, 2)) + "%、"


result = f"此外，5%、25%、75%、80%及90%分位数分别较{上一交易日标签}环比：" + result.rstrip("、") + "。"
# with open(outputname, 'a', encoding='utf-8') as f:
#     f.write(result)
    


# In[11]:


import openpyxl
worksheet = wb2['3、转债估值 (19年)']
# 获取 G 列最后一个有数字的单元格的行数
last_row = None
for cell in reversed(list(worksheet['AV'])):
    if cell.value is not None:
        last_row = cell.row
        break

# 获取 G 列最后一个有数字的单元格的值
最新百元溢价率 = worksheet.cell(row=last_row, column=48).value

# 获取 G 列倒数第二个有数字的单元格的行数
second_last_row = None
for cell in reversed(list(worksheet['AV'])):
    if cell.row != last_row and cell.value is not None:
        second_last_row = cell.row
        break

# 获取 G 列倒数第二个有数字的单元格的值
昨日百元溢价率 = worksheet.cell(row=second_last_row, column=48).value

百元溢价率变化 = 最新百元溢价率 - 昨日百元溢价率


# In[12]:


# 获取 L 列最后一个有数字的单元格的行数
last_row = None
for cell in reversed(list(worksheet['L'])):
    if cell.value is not None:
        last_row = cell.row
        break

# 获取 L 列最后一个有数字的单元格的值
最新平价 = worksheet.cell(row=last_row, column=12).value

# 获取 L 列倒数第二个有数字的单元格的行数
second_last_row = None
for cell in reversed(list(worksheet['L'])):
    if cell.row != last_row and cell.value is not None:
        second_last_row = cell.row
        break

# 获取 L 列倒数第二个有数字的单元格的值
昨日平价 = worksheet.cell(row=second_last_row, column=12).value

平价变化率 = (最新平价-昨日平价)/昨日平价


# In[13]:


# 获取 X 列最后一个有数字的单元格的行数
last_row = None
for cell in reversed(list(worksheet['X'])):
    if cell.value is not None:
        last_row = cell.row
        break

# 获取 X 列最后一个有数字的单元格的值
最新偏股溢价率 = worksheet.cell(row=last_row, column=24).value

# 获取 X 列倒数第二个有数字的单元格的行数
second_last_row = None
for cell in reversed(list(worksheet['X'])):
    if cell.row != last_row and cell.value is not None:
        second_last_row = cell.row
        break

# 获取 X 列倒数第二个有数字的单元格的值
昨日偏股溢价率 = worksheet.cell(row=second_last_row, column=24).value

偏股溢价率变动 = 最新偏股溢价率 - 昨日偏股溢价率


# In[14]:


# 获取 Y 列最后一个有数字的单元格的行数
last_row = None
for cell in reversed(list(worksheet['Y'])):
    if cell.value is not None:
        last_row = cell.row
        break

# 获取 Y 列最后一个有数字的单元格的值
最新偏债溢价率 = worksheet.cell(row=last_row, column=25).value

# 获取 Y 列倒数第二个有数字的单元格的行数
second_last_row = None
for cell in reversed(list(worksheet['Y'])):
    if cell.row != last_row and cell.value is not None:
        second_last_row = cell.row
        break

# 获取 Y 列倒数第二个有数字的单元格的值
昨日偏债溢价率 = worksheet.cell(row=second_last_row, column=25).value

偏债溢价率变动 = 最新偏债溢价率 - 昨日偏债溢价率


# In[15]:


# 获取 Z 列最后一个有数字的单元格的行数
last_row = None
for cell in reversed(list(worksheet['Z'])):
    if cell.value is not None:
        last_row = cell.row
        break

# 获取 Z 列最后一个有数字的单元格的值
最新平衡溢价率 = worksheet.cell(row=last_row, column=26).value

# 获取 Z 列倒数第二个有数字的单元格的行数
second_last_row = None
for cell in reversed(list(worksheet['Z'])):
    if cell.row != last_row and cell.value is not None:
        second_last_row = cell.row
        break

# 获取 Z 列倒数第二个有数字的单元格的值
昨日平衡溢价率 = worksheet.cell(row=second_last_row, column=26).value

平衡溢价率变动 = 最新平衡溢价率 - 昨日平衡溢价率


# In[16]:


def calc_premium_change(stock_premium, bond_premium, balance_premium):
    change_list = [stock_premium, bond_premium, balance_premium]
    change_max = max(change_list)
    change_min = min(change_list)
    if all(x > 0 for x in change_list):
        result = "偏股偏债型转债溢价率均向上，" + max(zip(change_list, ['偏股', '偏债', '平衡']))[1] + "型转债市场偏好度相对较高。"
    elif all(x < 0 for x in change_list):
        result = "偏股偏债型转债溢价率均向下，" + min(zip(change_list, ['偏股', '偏债', '平衡']))[1] + "型转债跌幅最大" + max(zip(change_list, ['偏股', '偏债', '平衡']))[1] + "型转债相对抗跌。"
    else:
        result = "转债估值股债型分化，市场相对看好" + max(zip(change_list, ['偏股', '偏债', '平衡']))[1] + "型转债，" + min(zip(change_list, ['偏股', '偏债', '平衡']))[1] + "型转债跌幅最大。"
    return result


# In[17]:


# 构造字符串
result2 = f"转债估值：估值{'抬升' if 百元溢价率变化 > 0 else '压缩' if 百元溢价率变化 < 0 else '维持此前水平'}"\
          f"。百元平价拟合转股溢价率为{最新百元溢价率/100:.2%}，环比{上一交易日标签}{'上升' if 百元溢价率变化 > 0 else '下降' if 百元溢价率变化 < 0 else '不变'}{'' if 百元溢价率变化 == 0 else abs(百元溢价率变化):.2f}{'pct' if 百元溢价率变化 != 0 else ''}；"\
          f"整体加权平价为{最新平价:.2f}元，环比{上一交易日标签}{'上升' if 平价变化率 > 0 else '下降' if 平价变化率 < 0 else '维持相同水平'}{'' if 平价变化率 == 0 else abs(平价变化率):.2%}。" 
#           f"价格中位数为{最新中位数:.2f}，环比昨日{'上升' if 中位数变化率 > 0 else '下降' if 中位数变化率 < 0 else '维持相同水平'}{'' if 中位数变化率 == 0 else abs(中位数变化率):.2%}。"


股债表现概况 = calc_premium_change(偏股溢价率变动, 偏债溢价率变动, 平衡溢价率变动)

result3 = f"偏股型转债溢价率为{最新偏股溢价率/100:.2%}{'，环比上升' if 偏股溢价率变动 > 0 else '，环比下降' if 偏股溢价率变动 < 0 else '，环比维持相同水平'}{'' if 偏股溢价率变动 == 0 else f'{abs(偏股溢价率变动):.2f}pct；'}"
result3 += f"偏债型转债溢价率为{最新偏债溢价率/100:.2%}{'，环比上升' if 偏债溢价率变动 > 0 else '，环比下降' if 偏债溢价率变动 < 0 else '，环比维持相同水平'}{'' if 偏债溢价率变动 == 0 else f'{abs(偏债溢价率变动):.2f}pct；'}" 
result3 += f"平衡型转债溢价率为{最新平衡溢价率/100:.2%}{'，环比上升' if 平衡溢价率变动 > 0 else '，环比下降' if 平衡溢价率变动 < 0 else '，环比维持相同水平'}{'' if 平衡溢价率变动 == 0 else f'{abs(平衡溢价率变动):.2f}pct。'}"

# result = result2 + 股债表现概况 + result3
# 追加写入文件
with open(outputname, 'a', encoding='utf-8') as f:
    f.write(result2)
#     f.write(股债表现概况)
    f.write(result3)


# In[18]:


from openpyxl.utils.cell import range_boundaries
worksheet = wb2['5、行业轮动情况']

# range_boundaries('C6:C36') 应当返回元组 (3, 2, 35, 2)
# start_row, start_col, end_row, end_col = range_boundaries('C8:C36')
start_row = 6
start_col = 3
end_row = 36
end_col = 3
# 获取涨跌幅数据
changes = []
for row in range(start_row, end_row):
    for col in range(start_col, end_col+1):
        cell = worksheet.cell(row=row, column=col)
        changes.append(cell.value)

# print(changes)
# 获取涨跌幅为正的个数
positive_count = sum(change > 0 for change in changes)

# 获取涨跌幅为正的指数名称及其涨跌幅
positive_changes = [(worksheet.cell(row=row, column=start_col-1).value, worksheet.cell(row=row, column=start_col).value) for row in range(start_row, end_row) if worksheet.cell(row=row, column=start_col).value > 0]
positive_changes.sort(key=lambda x: x[1], reverse=True)


# 获取涨跌幅为负的个数
negative_count = sum(change < 0 for change in changes)

# 获取涨跌幅为负的指数名称及其涨跌幅
negative_changes = [(worksheet.cell(row=row, column=start_col-1).value, worksheet.cell(row=row, column=start_col).value) for row in range(start_row, end_row) if worksheet.cell(row=row, column=start_col).value < 0]
negative_changes.sort(key=lambda x: x[1])


# 输出结果到txt文件
with open(outputname, 'a', encoding='utf-8') as f:
    f.write('\n行业表现：')
    if positive_count == len(changes):
        f.write(f'{运行日标签}各行业均不同程度上涨。')
        f.write('A股市场中，涨幅前三位行业为')
        name = [item[0] for item in positive_changes]
        change = [item[1] for item in positive_changes]
        f.write(f'{name[0]}（+{change[0]/100:.2%}）、{name[1]}（+{change[1]/100:.2%}）、{name[2]}（+{change[2]/100:.2%}）。')
#         for name, change in positive_changes[:3]:
#             f.write(f'{name}（+{change/100:.2%}）、')
    elif positive_count >= 26 and positive_count != len(changes):
        f.write(f'{运行日标签}正股行业普涨，共计')
        f.write(str(positive_count) + "个行业上涨。")
        f.write('A股市场中，涨幅前三位行业为')
        name = [item[0] for item in positive_changes]
        change = [item[1] for item in positive_changes]
        f.write(f'{name[0]}（+{change[0]/100:.2%}）、{name[1]}（+{change[1]/100:.2%}）、{name[2]}（+{change[2]/100:.2%}）；')
#         for name, change in positive_changes[:3]:
#             f.write(f'{name}（+{change/100:.2%}）、') 
        if len(negative_changes) == 1:
            f.write('唯一下跌的行业为')
            for name, change in negative_changes[:1]:
                f.write(f'{name}（{change/100:.2%}）。')
        elif len(negative_changes) == 2:
            f.write('仅两个行业下跌，分别为')
            name = [item[0] for item in negative_changes]
            change = [item[1] for item in negative_changes]
            f.write(f'{name[0]}（{change[0]/100:.2%}）、{name[1]}（{change[1]/100:.2%}）。')
#             for name, change in negative_changes[:2]:
#                 f.write(f'{name}（{change/100:.2%}）、')              
        elif len(negative_changes) >= 3:
            f.write('跌幅前三位行业为')
            name = [item[0] for item in negative_changes]
            change = [item[1] for item in negative_changes]
            f.write(f'{name[0]}（{change[0]/100:.2%}）、{name[1]}（{change[1]/100:.2%}）、{name[2]}（{change[2]/100:.2%}）。')
#             for name, change in negative_changes[:3]:
#                 f.write(f'{name}（{change/100:.2%}）、')
    elif positive_count / len(changes) >= 0.5 and positive_count < 26:
        f.write(f'{运行日标签}正股行业指数上涨占比过半，共计')
        f.write(str(positive_count) + "个行业上涨。")
        f.write('A股市场中，涨幅前三位行业为')
        name = [item[0] for item in positive_changes]
        change = [item[1] for item in positive_changes]
        f.write(f'{name[0]}（+{change[0]/100:.2%}）、{name[1]}（+{change[1]/100:.2%}）、{name[2]}（+{change[2]/100:.2%}）；')
#         for name, change in positive_changes[:3]:
#             f.write(f'{name}（+{change/100:.2%}）、')   
        if len(negative_changes) == 1:
            f.write('唯一下跌的行业为')
            for name, change in negative_changes[:1]:
                f.write(f'{name}（{change/100:.2%}）。')
        elif len(negative_changes) == 2:
            f.write('仅两个行业下跌，分别为')
            name = [item[0] for item in negative_changes]
            change = [item[1] for item in negative_changes]
            f.write(f'{name[0]}（{change[0]/100:.2%}）、{name[1]}（{change[1]/100:.2%}）。')
#             for name, change in negative_changes[:2]:
#                 f.write(f'{name}（{change/100:.2%}）、')               
        elif len(negative_changes) >= 3:
            f.write('跌幅前三位行业为')
            name = [item[0] for item in negative_changes]
            change = [item[1] for item in negative_changes]
            f.write(f'{name[0]}（{change[0]/100:.2%}）、{name[1]}（{change[1]/100:.2%}）、{name[2]}（{change[2]/100:.2%}）。')
#             for name, change in negative_changes[:3]:
#                 f.write(f'{name}（{change/100:.2%}）、') 
    elif negative_count / len(changes) > 0.5 and negative_count < 26:
        f.write(f'{运行日标签}正股行业指数下降占比过半，共计')
        f.write(str(negative_count) + "个行业下跌。")
        f.write('A股市场中，跌幅前三位行业为')
        name = [item[0] for item in negative_changes]
        change = [item[1] for item in negative_changes]
        f.write(f'{name[0]}（{change[0]/100:.2%}）、{name[1]}（{change[1]/100:.2%}）、{name[2]}（{change[2]/100:.2%}）；')
#         for name, change in negative_changes[:3]:
#             f.write(f'{name}（{change/100:.2%}）、') 
        if len(positive_changes) == 1:
            f.write('唯一逆势上涨的行业为')
            for name, change in positive_changes[:1]:
                f.write(f'{name}（+{change/100:.2%}）。')
        elif len(positive_changes) == 2:
            f.write('仅两个行业逆势上涨，分别为')
            name = [item[0] for item in positive_changes]
            change = [item[1] for item in positive_changes]
            f.write(f'{name[0]}（+{change[0]/100:.2%}）、{name[1]}（+{change[1]/100:.2%}）。')
#             for name, change in positive_changes[:2]:
#                 f.write(f'{name}（+{change/100:.2%}）、')                
        elif len(positive_changes) >= 3:
            f.write('涨幅前三位行业为')
            name = [item[0] for item in positive_changes]
            change = [item[1] for item in positive_changes]
            f.write(f'{name[0]}（+{change[0]/100:.2%}）、{name[1]}（+{change[1]/100:.2%}）、{name[2]}（+{change[2]/100:.2%}）。')
#             for name, change in positive_changes[:3]:
#                 f.write(f'{name}（+{change/100:.2%}）、')        
    elif negative_count >= 26 and negative_count != len(changes):
        f.write(f'{运行日标签}正股行业普跌，共计')
        f.write(str(negative_count) + "个行业下跌。")
        f.write('A股市场中，跌幅前三位行业为')
        name = [item[0] for item in negative_changes]
        change = [item[1] for item in negative_changes]
        f.write(f'{name[0]}（{change[0]/100:.2%}）、{name[1]}（{change[1]/100:.2%}）、{name[2]}（{change[2]/100:.2%}）；')
#         for name, change in negative_changes[:3]:
#             f.write(f'{name}（ {change/100:.2%}）、')  
        if len(positive_changes) == 1:
            f.write('唯一逆势上涨的行业为')
            for name, change in positive_changes[:1]:
                f.write(f'{name}（+{change/100:.2%}）。')
        elif len(positive_changes) == 2:
            f.write('仅两个行业逆势上涨，分别为')
            name = [item[0] for item in positive_changes]
            change = [item[1] for item in positive_changes]
            f.write(f'{name[0]}（+{change[0]/100:.2%}）、{name[1]}（+{change[1]/100:.2%}）。')
#             for name, change in positive_changes[:2]:
#                 f.write(f'{name}（+{change/100:.2%}）、')                
        elif len(positive_changes) >= 3:
            f.write('涨幅前三位行业为')
            name = [item[0] for item in positive_changes]
            change = [item[1] for item in positive_changes]
            f.write(f'{name[0]}（+{change[0]/100:.2%}）、{name[1]}（+{change[1]/100:.2%}）、{name[2]}（+{change[2]/100:.2%}）。')
#             for name, change in positive_changes[:3]:
#                 f.write(f'{name}（+{change/100:.2%}）、')
    elif negative_count == len(changes):
        f.write(f'{运行日标签}各行业均不同程度下跌。')
        f.write('A股市场中，跌幅前三位行业为')
        name = [item[0] for item in negative_changes]
        change = [item[1] for item in negative_changes]
        f.write(f'{name[0]}（{change[0]/100:.2%}）、{name[1]}（{change[1]/100:.2%}）、{name[2]}（{change[2]/100:.2%}）。')    
#         for name, change in negative_changes[:3]:
#             f.write(f'{name}（{change/100:.2%}）、')


# In[19]:


# 转债市场涨跌幅
from openpyxl.utils.cell import range_boundaries
worksheet = wb2['5、行业轮动情况']

# range_boundaries('C6:C36') 应当返回元组 (3, 2, 35, 2)
# start_row, start_col, end_row, end_col = range_boundaries('C8:C36')
start_row = 6
start_col = 7
end_row = 36
end_col = 7
# 获取涨跌幅数据
changes = []
for row in range(start_row, end_row):
    for col in range(start_col, end_col+1):
        cell = worksheet.cell(row=row, column=col)
        changes.append(cell.value)

changes = [x for x in changes if x and str(x) != 'nan']    
        
# print(changes)



# 获取涨跌幅为正的指数名称及其涨跌幅
positive_changes = [(worksheet.cell(row=row, column=start_col-5).value, worksheet.cell(row=row, column=start_col).value)
                    for row in range(start_row, end_row)
                    if worksheet.cell(row=row, column=start_col).value is not None and worksheet.cell(row=row, column=start_col).value > 0]
positive_changes.sort(key=lambda x: x[1], reverse=True)
# 获取涨跌幅为正的个数
positive_count = len(positive_changes)



# 获取涨跌幅为负的指数名称及其涨跌幅
negative_changes = [(worksheet.cell(row=row, column=start_col-5).value, worksheet.cell(row=row, column=start_col).value)
                    for row in range(start_row, end_row)
                    if worksheet.cell(row=row, column=start_col).value is not None and worksheet.cell(row=row, column=start_col).value < 0]
negative_changes.sort(key=lambda x: x[1])
# 获取涨跌幅为负的个数
negative_count = len(negative_changes)


# 输出结果到txt文件
with open(outputname, 'a', encoding='utf-8') as f:
    f.write('\n')
    if positive_count == len(changes):
        f.write('各行业转债均上涨，')
        f.write('涨幅前三位行业为')
        name = [item[0] for item in positive_changes]
        change = [item[1] for item in positive_changes]
        f.write(f'{name[0]}（+{change[0]:.2%}）、{name[1]}（+{change[1]:.2%}）、{name[2]}（+{change[2]:.2%}）。')   
#         for name, change in positive_changes[:3]:
#             f.write(f'{name}（+{change:.2%}）、')
    elif positive_count >= 26 and positive_count != len(changes):
        f.write('转债市场共计')
        f.write(str(positive_count) + "个行业上涨，")
        f.write('涨幅前三位行业为')
        name = [item[0] for item in positive_changes]
        change = [item[1] for item in positive_changes]
        f.write(f'{name[0]}（+{change[0]:.2%}）、{name[1]}（+{change[1]:.2%}）、{name[2]}（+{change[2]:.2%}）；')  
#         for name, change in positive_changes[:3]:
#             f.write(f'{name}（+{change:.2%}）、') 
        if len(negative_changes) == 1:
            f.write('唯一下跌的行业为')
            for name, change in negative_changes[:1]:
                f.write(f'{name}（{change:.2%}）。')
        elif len(negative_changes) == 2:
            f.write('仅两个行业下跌，分别为')
            name = [item[0] for item in negative_changes]
            change = [item[1] for item in negative_changes]
            f.write(f'{name[0]}（{change[0]:.2%}）、{name[1]}（{change[1]:.2%}）。')  
#             for name, change in negative_changes[:2]:
#                 f.write(f'{name}（{change:.2%}）、')              
        elif len(negative_changes) >= 3:
            f.write('跌幅前三位行业为')
            name = [item[0] for item in negative_changes]
            change = [item[1] for item in negative_changes]
            f.write(f'{name[0]}（{change[0]:.2%}）、{name[1]}（{change[1]:.2%}）、{name[2]}（{change[2]:.2%}）。')  
#             for name, change in negative_changes[:3]:
#                 f.write(f'{name}（{change:.2%}）、')
    elif positive_count / len(changes) >= 0.5 and positive_count < 26:
        f.write('转债市场共计')
        f.write(str(positive_count) + "个行业上涨，")
        f.write('涨幅前三位行业为')
        name = [item[0] for item in positive_changes]
        change = [item[1] for item in positive_changes]
        f.write(f'{name[0]}（+{change[0]:.2%}）、{name[1]}（+{change[1]:.2%}）、{name[2]}（+{change[2]:.2%}）；')  
#         for name, change in positive_changes[:3]:
#             f.write(f'{name}（+{change:.2%}）、')   
        if len(negative_changes) == 1:
            f.write('唯一下跌的行业为')
            for name, change in negative_changes[:1]:
                f.write(f'{name}（{change:.2%}）。')
        elif len(negative_changes) == 2:
            f.write('仅两个行业下跌，分别为')
            name = [item[0] for item in negative_changes]
            change = [item[1] for item in negative_changes]
            f.write(f'{name[0]}（{change[0]:.2%}）、{name[1]}（{change[1]:.2%}）。')  
#             for name, change in negative_changes[:2]:
#                 f.write(f'{name}（{change:.2%}）、')               
        elif len(negative_changes) >= 3:
            f.write('跌幅前三位行业为')
            name = [item[0] for item in negative_changes]
            change = [item[1] for item in negative_changes]
            f.write(f'{name[0]}（{change[0]:.2%}）、{name[1]}（{change[1]:.2%}）、{name[2]}（{change[2]:.2%}）。')  
#             for name, change in negative_changes[:3]:
#                 f.write(f'{name}（{change:.2%}）、') 
    elif negative_count / len(changes) > 0.5 and negative_count < 26:
        f.write('转债市场共计')
        f.write(str(negative_count) + "个行业下跌，")
        f.write('跌幅前三位行业为')
        name = [item[0] for item in negative_changes]
        change = [item[1] for item in negative_changes]
        f.write(f'{name[0]}（{change[0]:.2%}）、{name[1]}（{change[1]:.2%}）、{name[2]}（{change[2]:.2%}）；')  
#         for name, change in negative_changes[:3]:
#             f.write(f'{name}（{change:.2%}）、') 
        if len(positive_changes) == 1:
            f.write('唯一逆势上涨的行业为')
            for name, change in positive_changes[:1]:
                f.write(f'{name}（+{change:.2%}）。')
        elif len(positive_changes) == 2:
            f.write('仅两个行业逆势上涨，分别为')
            name = [item[0] for item in positive_changes]
            change = [item[1] for item in positive_changes]
            f.write(f'{name[0]}（+{change[0]:.2%}）、{name[1]}（+{change[1]:.2%}）。')  
#             for name, change in positive_changes[:2]:
#                 f.write(f'{name}（+{change:.2%}）、')                
        elif len(positive_changes) >= 3:
            f.write('涨幅前三位行业为')
            name = [item[0] for item in positive_changes]
            change = [item[1] for item in positive_changes]
            f.write(f'{name[0]}（+{change[0]:.2%}）、{name[1]}（+{change[1]:.2%}）、{name[2]}（+{change[2]:.2%}）。')  
#             for name, change in positive_changes[:3]:
#                 f.write(f'{name}（+{change:.2%}）、')        
    elif negative_count >= 26 and negative_count != len(changes):
        f.write('转债市场共计')
        f.write(str(negative_count) + "个行业下跌，")
        f.write('跌幅前三位行业为')
        name = [item[0] for item in negative_changes]
        change = [item[1] for item in negative_changes]
        f.write(f'{name[0]}（{change[0]:.2%}）、{name[1]}（{change[1]:.2%}）、{name[2]}（{change[2]:.2%}）；') 
        for name, change in negative_changes[:3]:
            f.write(f'{name}（ {change:.2%}）、')  
        if len(positive_changes) == 1:
            f.write('唯一逆势上涨的行业为')
            for name, change in positive_changes[:1]:
                f.write(f'{name}（+{change:.2%}）。')
        elif len(positive_changes) == 2:
            f.write('仅两个行业逆势上涨，分别为')
            name = [item[0] for item in positive_changes]
            change = [item[1] for item in positive_changes]
            f.write(f'{name[0]}（+{change[0]:.2%}）、{name[1]}（+{change[1]:.2%}）。')  
#             for name, change in positive_changes[:2]:
#                 f.write(f'{name}（+{change:.2%}）、')                
        elif len(positive_changes) >= 3:
            f.write('涨幅前三位行业为')
            for name, change in positive_changes[:3]:
                f.write(f'{name}（+{change:.2%}）、')
    elif negative_count == len(changes):
        f.write('转债市场均下跌，')
        f.write('跌幅前三位行业为')
        name = [item[0] for item in negative_changes]
        change = [item[1] for item in negative_changes]
        f.write(f'{name[0]}（{change[0]:.2%}）、{name[1]}（{change[1]:.2%}）、{name[2]}（{change[2]:.2%}）。')  
        
    else:
         f.write('转债市场行业涨跌幅表述有误，')
#         for name, change in negative_changes[:3]:
#             f.write(f'{name}（{change:.2%}）、')


# In[20]:


sheet = wb2['4、行业表现（核心指标）']

# 初始化变量，用于存储每个行业的变化率
changes = {}

# 计算每个行业的变化率
for i in range(1, 6):
    # 获取今日数据和昨日数据
    today_data = sheet.cell(row=i+4, column=sheet.max_column).value
    yesterday_data = sheet.cell(row=i+4, column=sheet.max_column-1).value
    
    # 计算变化率
    change = (today_data - yesterday_data) / yesterday_data
    
    # 将行业和变化率存储到字典中
    industry = sheet.cell(row=i+4, column=2).value
    changes[industry] = change
    
content_str = ""
for industry, change in changes.items():
    if change > 0:
        symbol = '+'
    else:
        symbol = ''
    content_str +=  f"{industry}环比{symbol}{change:.2%}、"
content_str = content_str[:-1].replace('、', '、', 1)
# 将变化率输出到txt文件
with open(outputname, 'a', encoding='utf-8') as f:
    f.write('\n(1) 收盘价：')
    f.write(content_str)
    f.write('。')


# In[21]:


sheet = wb2['4、行业表现（核心指标）']

# 初始化变量，用于存储每个行业的变化率
changes = {}

# 计算每个行业的变化率
for i in range(9, 14):
    # 获取今日数据和昨日数据
    today_data = sheet.cell(row=i+4, column=sheet.max_column).value
    yesterday_data = sheet.cell(row=i+4, column=sheet.max_column-1).value
    
    # 计算变化率
    change = (today_data - yesterday_data)
    
    # 将行业和变化率存储到字典中
    industry = sheet.cell(row=i+4, column=2).value
    changes[industry] = change

content_str = ""
for industry, change in changes.items():
    if change > 0:
        symbol = '+'
    else:
        symbol = ''
    content_str +=  f"{industry}环比{symbol}{change:.2}pct、"
content_str = content_str[:-1].replace('、', '、', 1)
# 将变化率输出到txt文件
with open(outputname, 'a', encoding='utf-8') as f:
    f.write('\n(2) 转股溢价率：')
    f.write(content_str)
    f.write('。')


# In[22]:


sheet = wb2['4、行业表现（核心指标）']

# 初始化变量，用于存储每个行业的变化率
changes = {}

# 计算每个行业的变化率
for i in range(17, 22):
    # 获取今日数据和昨日数据
    today_data = sheet.cell(row=i+4, column=sheet.max_column).value
    yesterday_data = sheet.cell(row=i+4, column=sheet.max_column-1).value
    
    # 计算变化率
    change = (today_data - yesterday_data) / yesterday_data
    
    # 将行业和变化率存储到字典中
    industry = sheet.cell(row=i+4, column=2).value
    changes[industry] = change

content_str = ""
for industry, change in changes.items():
    if change > 0:
        symbol = '+'
    else:
        symbol = ''
    content_str +=  f"{industry}环比{symbol}{change:.2%}、"
content_str = content_str[:-1].replace('、', '、', 1)
# 将变化率输出到txt文件
with open(outputname, 'a', encoding='utf-8') as f:
    f.write('\n(3) 转换价值：')
    f.write(content_str)
    f.write('。')


# In[23]:


sheet = wb2['4、行业表现（核心指标）']

# 初始化变量，用于存储每个行业的变化率
changes = {}

# 计算每个行业的变化率
for i in range(25, 30):
    # 获取今日数据和昨日数据
    today_data = sheet.cell(row=i+4, column=sheet.max_column).value
    yesterday_data = sheet.cell(row=i+4, column=sheet.max_column-1).value
    
    # 计算变化率
    change = (today_data - yesterday_data) 
    
    # 将行业和变化率存储到字典中
    industry = sheet.cell(row=i+4, column=2).value
    changes[industry] = change

content_str = ""
for industry, change in changes.items():
    if change > 0:
        symbol = '+'
    else:
        symbol = ''
    content_str +=  f"{industry}环比{symbol}{change:.2}pct、"
content_str = content_str[:-1].replace('、', '、', 1)
# 将变化率输出到txt文件
with open(outputname, 'a', encoding='utf-8') as f:
    f.write('\n(4) 纯债溢价率：')
    f.write(content_str)
    f.write('。')


# In[24]:


# import pandas as pd

# # 读取Excel文件
# excel_file = pd.ExcelFile(folder_name + '/预期变动.xlsx')

# # 定义要处理的sheet名称列表
# sheet_names = ['净利润预期', '营收预期', 'EPS预期']

# # 创建一个字典用于存储每个公司的涨跌幅数据
# company_data = {}

# # 遍历每个sheet
# for sheet_name in sheet_names:
#     # 读取当前sheet的数据
#     df = excel_file.parse(sheet_name)
    
#     # 遍历每行数据
#     for index, row in df.iterrows():
#         company_name = row['名称']
#         institution_code = row['机构代码']
#         change_percent = row['差异']
        
#         # 如果公司名称不存在于字典中，则创建一个新的字典项
#         if company_name not in company_data:
#             company_data[company_name] = {'名称': company_name}
        
#         # 添加或更新涨跌幅数据
#         if sheet_name not in company_data[company_name]:
#             company_data[company_name][sheet_name] = change_percent
#         elif company_data[company_name][sheet_name] != change_percent:
#             company_data[company_name][sheet_name] += ', ' + change_percent

# # 输出结果

# with open(outputname, 'a', encoding='utf-8') as f:
#     f.write("\n")
#     # 遍历每个公司的数据
#     for company_name, data in company_data.items():
#         f.write(company_name + "\n")
#         for sheet_name in sheet_names:
#             if sheet_name in data:
#                 change_percent = data[sheet_name]
#                 change_percent_list = change_percent.split(", ")  # 分割涨跌幅值
#                 processed_change_percent = []
#                 for percent in change_percent_list:
#                     percent = percent.strip('%')  # 去除百分号
#                     if float(percent) > 0:  # 判断涨跌幅是否为正值
#                         processed_change_percent.append("+" + percent + "%")
#                     else:
#                         processed_change_percent.append(percent + "%")
#                 formatted_change_percent = ", ".join(processed_change_percent)  # 拼接处理后的涨跌幅值
#                 f.write(sheet_name + ": " + formatted_change_percent + "\n")
#         f.write("\n")




# In[25]:


# import openpyxl
# from datetime import datetime

# def concatenate_non_blank(range):
#     result = ""
#     for cell in range:
#         if cell.value is not None and cell.value != 0 and str(cell.value).strip() != "":
#             if isinstance(cell.value, datetime):
#                 date_str = cell.value.strftime('%Y-%m-%d')
#                 result += f"{date_str}、"
#             elif isinstance(cell.value, float):
#                 date_num = int(cell.value)
#                 date_obj = datetime.fromordinal(date_num + 693594)  # 将数字转换为datetime对象
#                 date_str = date_obj.strftime('%Y%m%d')  # 将日期格式化为YYYYMMDD
#                 result += f"{date_str}、"
#             else:
#                 result += f"{cell.value}、"
#     if result != "":
#         result = result[:-1]  # 去除最后一个逗号
#     return result

# # wb2 = openpyxl.load_workbook(filename)

# sheet = wb2['附6、可转债列表']

# count = 0
# for cell in sheet["C"]:
#     if cell.value not in ["私募", "定向", "0", 0, None]:
#         count += 1

# balance = round(sum([cell.value for cell in sheet["D:D"] if isinstance(cell.value, (int, float))]),2)

# # content = f"\n发行信息：\n现已发行未到期可转债{count}只，余额规模约{balance}亿元。{concatenate_non_blank(sheet['L:L'])}尚未发布上市公告，此外{concatenate_non_blank(sheet['M:M'])}即将上市交易，上市日期"
# content = f"{concatenate_non_blank(sheet['M:M'])}即将上市交易，上市日期"
# if "、" in concatenate_non_blank(sheet['M:M']):
#     content += "分别为："
# else:
#     content += "为："

# # 追加写入文件
# with open(outputname, 'a', encoding='utf-8') as f:
#     f.write(content)


# In[26]:


# sheet = wb2['附6、可转债列表']

# from datetime import datetime
# # 创建一个空列表用于存储数据
# bond_data = []

# # 遍历K列和M列，将债券名称和对应的日期添加到列表中
# for row in range(1, sheet.max_row+1):
#     if sheet.cell(row=row, column=11).value is None or sheet.cell(row=row, column=11).value == 0:
#         continue  # 如果K列单元格为空，则跳过该行
#     try:
#         date_value = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + sheet.cell(row=row, column=11).value - 2)
#     except TypeError:
#         continue  # 如果K列单元格不是日期，则跳过该行
#     bond_name = sheet.cell(row=row, column=13).value
#     bond_data.append((bond_name, date_value))

# # 按照日期先后排序
# bond_data = sorted(bond_data, key=lambda x: x[1])

# # 将结果输出到txt文件中
# with open(outputname, 'a', encoding='utf-8') as f:
#     for i in range(len(bond_data)):
#         if i > 0 and bond_data[i][1] == bond_data[i-1][1]:
#             # 如果日期相同，将名称添加到同一行
#             f.write(f'、{bond_data[i][0]}')
#         else:
#             # 如果日期不同，新起一行输出
#             f.write('、')
#             f.write(f'{bond_data[i][0]}')
#         f.write(f': {bond_data[i][1].month}月{bond_data[i][1].day}日')
#     f.write("。")


# In[27]:


wb2.close()
del wb2
market_stats_wb.close()


